"""KURO xlsx `expected_mutations` sheet adapter (Blocker B resolved)."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from mame.models import ExpectedMutation

_EXPECTED_SHEET = "expected_mutations"
_EXPECTED_HEADER = [
    "mutant_id",
    "position",
    "wt_aa",
    "mt_aa",
    "wt_codon",
    "mt_codon",
    "group_id",
    "primer_set_ref",
    "notation_type",
    "status",
]


def read_expected_mutations(path: Path) -> list[ExpectedMutation]:
    """Read the `expected_mutations` sheet from a KURO xlsx export.

    Only rows with status == "DESIGNED" are returned; FAILED rows are Phase 2.
    Raises ValueError if the expected sheet is missing (old KURO version).
    """

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if _EXPECTED_SHEET not in wb.sheetnames:
            raise ValueError(
                f"KURO xlsx '{path}' is missing the required '{_EXPECTED_SHEET}' sheet. "
                "Re-export from a KURO build >= commit 8c47037."
            )
        ws = wb[_EXPECTED_SHEET]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"'{_EXPECTED_SHEET}' sheet in '{path}' is empty.")
        header_list = [str(c) if c is not None else "" for c in header]
        expected = [h.lower() for h in _EXPECTED_HEADER]
        got = [h.strip().lower() for h in header_list]
        if got[: len(expected)] != expected:
            raise ValueError(
                f"'{_EXPECTED_SHEET}' header mismatch. Expected {expected}, got {got}."
            )

        results: list[ExpectedMutation] = []
        for raw in rows_iter:
            if raw is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
                continue
            cells = list(raw) + [None] * (len(_EXPECTED_HEADER) - len(raw))
            status = _s(cells[9])
            if status.upper() != "DESIGNED":
                continue
            results.append(
                ExpectedMutation(
                    mutant_id=_s(cells[0]),
                    position=_int(cells[1]),
                    wt_aa=_s(cells[2]),
                    mt_aa=_s(cells[3]),
                    wt_codon=_s(cells[4]),
                    mt_codon=_s(cells[5]),
                    group_id=_s(cells[6]),
                    primer_set_ref=_s(cells[7]),
                    notation_type=_s(cells[8]),
                    status=status,
                )
            )
        return results
    finally:
        wb.close()


def expected_to_labels(expected: list[ExpectedMutation]) -> list[str]:
    """Produce the human-readable mutation label list consumed by compare.verdict."""

    return [f"{m.wt_aa}{m.position}{m.mt_aa}" for m in expected]


def _s(cell: object) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


def _int(cell: object) -> int:
    if cell is None:
        return 0
    try:
        return int(str(cell).strip())
    except (TypeError, ValueError):
        return 0
