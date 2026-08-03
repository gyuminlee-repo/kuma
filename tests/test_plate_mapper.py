"""Tests for plate mapping, Excel export, and order CSV export."""

from __future__ import annotations

import csv
from pathlib import Path

import pytest

from kuma_core.kuro.plate_mapper import (
    PlateMapping,
    _split_echo_volume,
    deduplicate_reverse,
    export_echo_mapping_csv,
    export_idt_csv,
    export_janus_mapping_csv,
    export_plate_excel,
    export_twist_csv,
    generate_plate_map,
    _to_384_well_fwd,
    _to_384_well_rev,
)
from kuma_core.kuro.sdm_engine import SdmPrimerResult, design_sdm_primers
from tests.conftest import FIXTURES_DIR, TARGET_START


@pytest.fixture(scope="module")
def sdm_results(genbank_path, mutations_csv) -> list[SdmPrimerResult]:
    results, _, _f = design_sdm_primers(
        fasta_path=genbank_path,
        target_start=TARGET_START,
        mutations_csv=mutations_csv,
        polymerase="Q5",
        overlap_len=18,
    )
    return results


class TestDeduplicateReverse:
    def test_no_duplicates(self, sdm_results):
        rev_map = deduplicate_reverse(sdm_results)
        total_mutations = sum(len(v) for v in rev_map.values())
        assert total_mutations == len(sdm_results)


class TestGeneratePlateMap:
    def test_plate_map_count(self, sdm_results):
        fwd_map, rev_map = generate_plate_map(sdm_results, deduplicate_rev=True)
        assert len(fwd_map) == len(sdm_results)
        assert len(rev_map) <= len(sdm_results)

    def test_well_names_valid(self, sdm_results):
        fwd_map, rev_map = generate_plate_map(sdm_results)
        for m in fwd_map + rev_map:
            assert m.well[0] in "ABCDEFGH"
            col = int(m.well[1:])
            assert 1 <= col <= 12

    def test_column_order(self, sdm_results):
        fwd_map, _ = generate_plate_map(sdm_results, well_order="column")
        if len(fwd_map) >= 8:
            first_8 = [m.well for m in fwd_map[:8]]
            expected = ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]
            assert first_8 == expected

    def test_no_dedup(self, sdm_results):
        fwd_map, rev_map = generate_plate_map(sdm_results, deduplicate_rev=False)
        assert len(fwd_map) == len(sdm_results)
        assert len(rev_map) == len(sdm_results)


class TestExportExcel:
    def test_export_creates_file(self, sdm_results, tmp_path):
        fwd_map, rev_map = generate_plate_map(sdm_results)
        xlsx_path = tmp_path / "test_plate.xlsx"
        export_plate_excel(fwd_map + rev_map, xlsx_path)
        assert xlsx_path.exists()
        assert xlsx_path.stat().st_size > 0

    def test_export_four_sheets(self, sdm_results, tmp_path):
        from openpyxl import load_workbook
        fwd_map, rev_map = generate_plate_map(sdm_results)
        xlsx_path = tmp_path / "test_plate.xlsx"
        export_plate_excel(fwd_map + rev_map, xlsx_path)

        wb = load_workbook(xlsx_path)
        assert "Fwd List" in wb.sheetnames
        assert "Fwd Plate" in wb.sheetnames
        assert "Rev List" in wb.sheetnames
        assert "Rev Plate" in wb.sheetnames


class TestExportIdtCsv:
    def test_idt_csv_creates_file(self, sdm_results, tmp_path):
        csv_path = tmp_path / "idt_order.csv"
        export_idt_csv(sdm_results, csv_path)
        assert csv_path.exists()
        assert csv_path.stat().st_size > 0

    def test_idt_csv_header(self, sdm_results, tmp_path):
        csv_path = tmp_path / "idt_order.csv"
        export_idt_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert header == ["Name", "Sequence", "Scale", "Purification"]

    def test_idt_csv_row_count(self, sdm_results, tmp_path):
        csv_path = tmp_path / "idt_order.csv"
        export_idt_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        # header + 2 rows per result (fwd + rev)
        assert len(rows) == 1 + len(sdm_results) * 2

    def test_idt_csv_naming_convention(self, sdm_results, tmp_path):
        csv_path = tmp_path / "idt_order.csv"
        export_idt_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for i, row in enumerate(reader):
                name = row[0]
                if i % 2 == 0:
                    assert name.endswith("_F")
                else:
                    assert name.endswith("_R")

    def test_idt_csv_default_scale_and_purification(self, sdm_results, tmp_path):
        csv_path = tmp_path / "idt_order.csv"
        export_idt_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                assert row[2] == "25nm"
                assert row[3] == "STD"

    def test_idt_csv_custom_scale(self, sdm_results, tmp_path):
        csv_path = tmp_path / "idt_order.csv"
        export_idt_csv(sdm_results, csv_path, scale="100nm", purification="PAGE")
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                assert row[2] == "100nm"
                assert row[3] == "PAGE"


class TestExportTwistCsv:
    def test_twist_csv_creates_file(self, sdm_results, tmp_path):
        csv_path = tmp_path / "twist_order.csv"
        export_twist_csv(sdm_results, csv_path)
        assert csv_path.exists()
        assert csv_path.stat().st_size > 0

    def test_twist_csv_header(self, sdm_results, tmp_path):
        csv_path = tmp_path / "twist_order.csv"
        export_twist_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert header == ["Name", "Sequence", "Notes"]

    def test_twist_csv_row_count(self, sdm_results, tmp_path):
        csv_path = tmp_path / "twist_order.csv"
        export_twist_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) == 1 + len(sdm_results) * 2

    def test_twist_csv_notes_contain_mutation(self, sdm_results, tmp_path):
        csv_path = tmp_path / "twist_order.csv"
        export_twist_csv(sdm_results, csv_path)
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                # Notes column should contain the mutation name
                assert len(row[2]) > 0


class Test384WellConversion:
    def test_fwd_row_mapping(self):
        assert _to_384_well_fwd("A1") == "A1"
        assert _to_384_well_fwd("B1") == "C1"
        assert _to_384_well_fwd("C1") == "E1"
        assert _to_384_well_fwd("H1") == "O1"
        assert _to_384_well_fwd("A12") == "A12"

    def test_rev_row_mapping(self):
        assert _to_384_well_rev("A1") == "B1"
        assert _to_384_well_rev("B1") == "D1"
        assert _to_384_well_rev("C1") == "F1"
        assert _to_384_well_rev("H1") == "P1"

    def test_fwd_rev_no_overlap(self):
        wells_96 = [f"{r}{c}" for c in range(1, 13) for r in "ABCDEFGH"]
        fwd_384 = {_to_384_well_fwd(w) for w in wells_96}
        rev_384 = {_to_384_well_rev(w) for w in wells_96}
        assert fwd_384.isdisjoint(rev_384)


class TestEchoMappingExport:
    def test_split_echo_volume_above_single_transfer_limit(self):
        assert _split_echo_volume(1000) == [500, 500]
        assert _split_echo_volume(600) == [500, 100]
        assert _split_echo_volume(300) == [300]

    def test_row_count(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        rev_groups = deduplicate_reverse(sdm_results)
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path, rev_groups=rev_groups)

        with open(csv_path, encoding="utf-8") as f:
            rows = list(csv.reader(f))

        # Header + one row per fwd + one row per (rev→dest) pair
        # With dedup, rev rows = len(fwd) (one dest per mutation)
        assert len(rows) == 1 + len(fwd) + len(fwd)

    def test_header(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            header = next(csv.reader(f))

        assert header == [
            "Source Plate Name", "Source Well Name", "Source Well",
            "Dest Plate Name", "Dest Well Name", "Dest Well", "Transfer Vol",
        ]

    def test_source_plate_constant(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                assert row[0] == "Source [1]"
                assert row[3] == "Destination [1]"

    def test_transfer_vol_default(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                assert row[6] == "100"

    def test_transfer_vol_split_rows_above_echo_limit(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        rev_groups = deduplicate_reverse(sdm_results)
        csv_path = tmp_path / "echo_split.csv"
        export_echo_mapping_csv(fwd, rev, csv_path, transfer_vol=600, rev_groups=rev_groups)

        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            rows = list(reader)

        assert len(rows) == 1 + (len(fwd) * 4) - 1
        assert {row[6] for row in rows} == {"500", "100"}

    def test_fwd_source_well_odd_rows(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path)

        odd_384_rows = set("ACEGIKMO")
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            fwd_rows = [row for row in reader if row[1].endswith("_F")]
        for row in fwd_rows:
            assert row[2][0] in odd_384_rows, f"Fwd source well {row[2]} not in odd rows"

    def test_rev_source_well_even_rows(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path)

        even_384_rows = set("BDFHJLNP")
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            rev_rows = [row for row in reader if row[1].endswith("_R")]
        for row in rev_rows:
            assert row[2][0] in even_384_rows, f"Rev source well {row[2]} not in even rows"


class TestJanusMappingExport:
    def test_row_count(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        rev_groups = deduplicate_reverse(sdm_results)
        csv_path = tmp_path / "janus.csv"
        export_janus_mapping_csv(fwd, rev, csv_path, rev_groups=rev_groups)

        with open(csv_path, encoding="utf-8") as f:
            rows = list(csv.reader(f))

        assert len(rows) == 1 + len(fwd) + len(fwd)

    def test_header(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "janus.csv"
        export_janus_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            header = next(csv.reader(f))

        assert header == [
            "name", "type", "Dsp. Rack", "no",
            "Asp. Rack", "Asp. Posi", "Dsp. Rack", "Dsp. Posi", "volume",
        ]

    def test_asp_rack_separation(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "janus.csv"
        export_janus_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            rows = list(reader)

        fw_rows = [r for r in rows if r[0].endswith("-F")]
        rv_rows = [r for r in rows if r[0].endswith("-R")]
        assert all(r[4] == "1" for r in fw_rows), "fw rows must use Asp. Rack 1"
        assert all(r[4] == "2" for r in rv_rows), "rv rows must use Asp. Rack 2"
        assert all(r[6] == "3" for r in fw_rows + rv_rows), "all rows must use Dsp. Rack 3"

    def test_transfer_vol_default(self, sdm_results, tmp_path):
        fwd, rev = generate_plate_map(sdm_results, deduplicate_rev=True)
        csv_path = tmp_path / "janus.csv"
        export_janus_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                assert float(row[8]) == 2.0


def _shared_rev_fixture() -> tuple[list[PlateMapping], list[PlateMapping], dict[str, list[str]]]:
    """Two mutations sharing one reverse primer, plus one unshared mutation.

    Built directly from PlateMapping so sharing is guaranteed, unlike the
    ``sdm_results`` fixture whose reverse primers happen to be all distinct.
    """
    shared_seq = "CCCGGGAAATTTCCCGGG"
    solo_seq = "TTTGGGCCCAAATTTGGG"
    fwd = [
        PlateMapping("A1", "M1_F", "AAATTTCCCGGGAAATTT", "forward", "M1"),
        PlateMapping("B1", "M2_F", "AAATTTCCCGGGAAAGGG", "forward", "M2"),
        PlateMapping("C1", "M3_F", "AAATTTCCCGGGAAACCC", "forward", "M3"),
    ]
    rev = [
        PlateMapping("A1", "M1_R", shared_seq, "reverse", "M1"),
        PlateMapping("B1", "M3_R", solo_seq, "reverse", "M3"),
    ]
    rev_groups = {shared_seq: ["M1", "M2"], solo_seq: ["M3"]}
    return fwd, rev, rev_groups


class TestSharedReverseWithoutRevGroups:
    """Regression: dedup groups missing must not silently drop reverse rows."""

    def test_echo_csv_raises_when_rev_groups_missing(self, tmp_path):
        fwd, rev, _ = _shared_rev_fixture()
        with pytest.raises(ValueError) as exc:
            export_echo_mapping_csv(fwd, rev, tmp_path / "echo.csv")
        assert "M2" in str(exc.value)

    def test_janus_csv_raises_when_rev_groups_missing(self, tmp_path):
        fwd, rev, _ = _shared_rev_fixture()
        with pytest.raises(ValueError) as exc:
            export_janus_mapping_csv(fwd, rev, tmp_path / "janus.csv")
        assert "M2" in str(exc.value)

    def test_rev_groups_present_keeps_every_mutation(self, tmp_path):
        fwd, rev, rev_groups = _shared_rev_fixture()
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path, rev_groups=rev_groups)

        with open(csv_path, encoding="utf-8") as f:
            rows = list(csv.reader(f))[1:]

        rev_dest = [r[4] for r in rows if r[1].endswith("_R")]
        assert sorted(rev_dest) == ["M1", "M2", "M3"]
        assert len(rows) == len(fwd) * 2

    def test_undeduplicated_rev_needs_no_rev_groups(self, tmp_path):
        """One reverse mapping per mutation carries its own link, so no raise."""
        fwd, _, _ = _shared_rev_fixture()
        rev = [
            PlateMapping("A1", "M1_R", "CCCGGGAAATTTCCCGGG", "reverse", "M1"),
            PlateMapping("B1", "M2_R", "CCCGGGAAATTTCCCAAA", "reverse", "M2"),
            PlateMapping("C1", "M3_R", "TTTGGGCCCAAATTTGGG", "reverse", "M3"),
        ]
        csv_path = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, csv_path)

        with open(csv_path, encoding="utf-8") as f:
            rows = list(csv.reader(f))[1:]
        rev_dest = [r[4] for r in rows if r[1].endswith("_R")]
        assert sorted(rev_dest) == ["M1", "M2", "M3"]


class TestRevUsageTable:
    """layout sheet reports shared count and total transfer volume."""

    def _read_usage(self, ws):
        header_row = None
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=False):
            if row[0].value == "Reverse primer usage":
                header_row = row[0].row
                break
        assert header_row is not None, "usage table header missing"
        rows = []
        r = header_row + 2
        while ws.cell(row=r, column=1).value and ws.cell(row=r, column=3).value is not None:
            rows.append([ws.cell(row=r, column=c).value for c in range(1, 6)])
            r += 1
        note = ws.cell(row=r, column=1).value
        return [c.value for c in ws[header_row + 1][:5]], rows, note

    def test_echo_layout_usage_table(self, tmp_path):
        from openpyxl import load_workbook
        from kuma_core.kuro.plate_mapper import export_echo_mapping_xlsx

        fwd, rev, rev_groups = _shared_rev_fixture()
        path = tmp_path / "echo.xlsx"
        export_echo_mapping_xlsx(
            fwd, rev, path, transfer_vol=600, rev_groups=rev_groups,
        )

        ws = load_workbook(path)["layout"]
        headers, rows, note = self._read_usage(ws)
        assert headers == [
            "Source Well", "Primer Name", "Shared By (reactions)",
            "Volume Per Reaction (nL)", "Total Transfer Volume (nL)",
        ]
        # Shared M1/M2 well is aspirated twice; M3 once. Volume splitting
        # (600 nL -> 500 + 100) must not change the totals.
        assert rows == [
            ["A1", "M1_R", 2, 600, 1200],
            ["B1", "M3_R", 1, 600, 600],
        ]
        assert "dead volume" in note

    def test_janus_layout_usage_table(self, tmp_path):
        from openpyxl import load_workbook
        from kuma_core.kuro.plate_mapper import export_janus_mapping_xlsx

        fwd, rev, rev_groups = _shared_rev_fixture()
        path = tmp_path / "janus.xlsx"
        export_janus_mapping_xlsx(
            fwd, rev, path, transfer_vol=2.0, rev_groups=rev_groups,
        )

        ws = load_workbook(path)["layout"]
        headers, rows, note = self._read_usage(ws)
        assert headers[3] == "Volume Per Reaction (µL)"
        assert headers[4] == "Total Transfer Volume (µL)"
        assert rows == [
            ["A1", "M1_R", 2, 2.0, 4.0],
            ["B1", "M3_R", 1, 2.0, 2.0],
        ]
        assert "dead volume" in note


class TestExpectedMutationsSheet:
    """Tests for _write_expected_mutations_sheet and integration into export_plate_excel."""

    def _make_sdm_result(self, raw: str, wt_aa: str, position: int, mt_aa: str,
                          wt_codon: str, mt_codon: str, group_id=None):
        """Build a minimal SdmPrimerResult for testing."""
        from kuma_core.kuro.mutation import Mutation
        from kuma_core.kuro.overlap import OverlapWindow
        from kuma_core.kuro.sdm_engine import SdmPrimerResult

        mut = Mutation(
            raw=raw,
            wt_aa=wt_aa,
            position=position,
            mt_aa=mt_aa,
            codon_start=0,
            wt_codon=wt_codon,
            mt_codon=mt_codon,
            group_id=group_id,
        )
        window = OverlapWindow(sequence="AAATTTCCC", start=0, end=9, codon_offset=3)
        return SdmPrimerResult(
            mutation=mut,
            forward_seq="AAATTTCCCGGG",
            reverse_seq="CCCGGGAAATTT",
            forward_binding="AAATTT",
            reverse_binding="CCCGGG",
            overlap_window=window,
            tm_fwd=60.0,
            tm_rev=58.0,
            tm_overlap=55.0,
            tm_condition_met=True,
        )

    def _make_mock_results(self):
        """
        Construct 4 SdmPrimerResult objects:
          - V5F   (single, group_id=None)
          - K53N  (single, group_id=None)
          - A40P  (multi, group_id="A40P/E61Y")
          - E61Y  (multi, group_id="A40P/E61Y")

        Note: spec §7 example stated 3 results but yielded max_row==4 inconsistency.
        A40P/E61Y multi-mutation produces two separate SdmPrimerResult objects
        (one per sub-Mutation), so the correct fixture is 4 results → max_row == 5.
        """
        return [
            self._make_sdm_result("V5F", "V", 5, "F", "GTG", "TTT"),
            self._make_sdm_result("K53N", "K", 53, "N", "AAG", "AAC"),
            self._make_sdm_result("A40P", "A", 40, "P", "GCG", "CCG", group_id="A40P/E61Y"),
            self._make_sdm_result("E61Y", "E", 61, "Y", "GAA", "TAT", group_id="A40P/E61Y"),
        ]

    def test_write_expected_mutations_sheet_basic(self):
        """4 DESIGNED results (incl. 1 multi-mutation group) → expected_mutations 시트 검증."""
        from openpyxl import Workbook
        from kuma_core.kuro.plate_mapper import _write_expected_mutations_sheet

        mock_results = self._make_mock_results()
        wb = Workbook()
        _write_expected_mutations_sheet(wb, mock_results)

        # 시트 존재 확인
        assert "expected_mutations" in wb.sheetnames

        ws = wb["expected_mutations"]

        # 헤더 행 확인
        headers = [c.value for c in ws[1]]
        assert headers == [
            "mutant_id", "position", "wt_aa", "mt_aa",
            "wt_codon", "mt_codon", "group_id", "primer_set_ref",
            "notation_type", "status", "rescue_type", "rescue_stage", "rescued_from",
        ]

        # 데이터 행 수 확인: header(1) + 4 data rows = 5
        assert ws.max_row == 5

        # 모든 행 notation_type, status 상수 확인
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[8] == "substitution"
            assert row[9] == "DESIGNED"
            assert row[10] == ""
            assert row[11] == ""
            assert row[12] == ""

        # 단일 변이(V5F)의 group_id는 빈 문자열
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        v5f_row = next(r for r in data_rows if r[0] == "V5F")
        assert v5f_row[6] == ""

        # multi-mutation group_id 확인
        a40p_row = next(r for r in data_rows if r[0] == "A40P")
        e61y_row = next(r for r in data_rows if r[0] == "E61Y")
        assert a40p_row[6] == "A40P/E61Y"
        assert e61y_row[6] == "A40P/E61Y"

        # primer_set_ref == mutant_id
        for row in data_rows:
            assert row[0] == row[7], f"primer_set_ref mismatch for {row[0]}"

    def test_write_expected_mutations_sheet_preserves_status_and_adds_rescue_metadata(self):
        from openpyxl import Workbook
        from kuma_core.kuro.plate_mapper import _write_expected_mutations_sheet

        mock_results = self._make_mock_results()
        wb = Workbook()
        _write_expected_mutations_sheet(
            wb,
            mock_results,
            rescued_info=[
                {
                    "original": "V5F",
                    "rescued_by": "K53N",
                    "type": "auto_suggestion_l2",
                    "stage": 2,
                }
            ],
        )

        rows = list(wb["expected_mutations"].iter_rows(min_row=2, values_only=True))
        k53n_row = next(r for r in rows if r[0] == "K53N")
        assert k53n_row[9] == "DESIGNED"
        assert k53n_row[10] == "auto_suggestion_l2"
        assert k53n_row[11] == 2
        assert k53n_row[12] == "V5F"

    def test_export_plate_excel_without_results_no_extra_sheet(self, sdm_results, tmp_path):
        """results=None のとき expected_mutations シートが生成されない (하위 호환 확인)."""
        from openpyxl import load_workbook

        fwd_map, rev_map = generate_plate_map(sdm_results)
        xlsx_path = tmp_path / "no_extra.xlsx"
        export_plate_excel(fwd_map + rev_map, xlsx_path)  # results 생략

        wb = load_workbook(xlsx_path)
        assert "expected_mutations" not in wb.sheetnames
        # 기존 4개 시트는 그대로
        assert "Fwd List" in wb.sheetnames
        assert "Fwd Plate" in wb.sheetnames
        assert "Rev List" in wb.sheetnames
        assert "Rev Plate" in wb.sheetnames

    def test_cmd_design_xlsx_has_expected_mutations_sheet(self, sdm_results, tmp_path):
        """export_plate_excel(results=...) 직접 호출 → 생성된 xlsx에 5번째 시트 확인."""
        from openpyxl import load_workbook
        from kuma_core.kuro.plate_mapper import deduplicate_reverse

        rev_groups = deduplicate_reverse(sdm_results)
        fwd_map, rev_map = generate_plate_map(sdm_results, deduplicate_rev=True)
        xlsx_path = tmp_path / "with_expected.xlsx"
        export_plate_excel(fwd_map + rev_map, xlsx_path, rev_groups=rev_groups, results=sdm_results)

        wb = load_workbook(xlsx_path)
        assert "expected_mutations" in wb.sheetnames

        ws = wb["expected_mutations"]
        assert ws.max_row >= 2  # 최소 1건 이상

        # 헤더 정확성 확인
        headers = [c.value for c in ws[1]]
        assert headers == [
            "mutant_id", "position", "wt_aa", "mt_aa",
            "wt_codon", "mt_codon", "group_id", "primer_set_ref",
            "notation_type", "status", "rescue_type", "rescue_stage", "rescued_from",
        ]

        # 모든 데이터 행의 상수 컬럼 검증
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[8] == "substitution"
            assert row[9] == "DESIGNED"

        # expected_mutations 시트가 마지막 시트
        assert wb.sheetnames[-1] == "expected_mutations"


class TestExpectedMutationsFollowsPlateOrder:
    """MAME reads row i of expected_mutations as well i, so that order is a coordinate."""

    def _result(self, raw, wt, pos, mt):
        from kuma_core.kuro.mutation import Mutation
        from kuma_core.kuro.overlap import OverlapWindow
        from kuma_core.kuro.sdm_engine import SdmPrimerResult

        mut = Mutation(
            raw=raw, wt_aa=wt, position=pos, mt_aa=mt,
            codon_start=0, wt_codon="GTG", mt_codon="TTT", group_id=None,
        )
        return SdmPrimerResult(
            mutation=mut,
            forward_seq="AAATTTCCCGGG",
            reverse_seq="CCCGGGAAATTT",
            forward_binding="AAATTT",
            reverse_binding="CCCGGG",
            overlap_window=OverlapWindow(sequence="AAATTTCCC", start=0, end=9, codon_offset=3),
            tm_fwd=60.0, tm_rev=58.0, tm_overlap=55.0, tm_condition_met=True,
        )

    def _mapping(self, well, mutation, primer_type):
        from kuma_core.kuro.plate_mapper import PlateMapping

        suffix = "F" if primer_type == "forward" else "R"
        return PlateMapping(
            well=well,
            primer_name=f"{mutation}_{suffix}",
            sequence="AAATTTCCC",
            primer_type=primer_type,
            mutation=mutation,
        )

    def test_rows_are_reordered_to_match_the_forward_plate(self, tmp_path):
        from openpyxl import load_workbook

        from kuma_core.kuro.plate_mapper import export_plate_excel

        # 설계 순서(점수 순 등)와 플레이트 순서가 어긋난 상태를 재현한다.
        results = [self._result(*a) for a in (
            ("I92D", "I", 92, "D"), ("S11I", "S", 11, "I"), ("K53I", "K", 53, "I"),
        )]
        mappings = [
            self._mapping("A1", "S11I", "forward"),
            self._mapping("B1", "K53I", "forward"),
            self._mapping("C1", "I92D", "forward"),
            self._mapping("A2", "S11I", "reverse"),
        ]
        out = tmp_path / "platemap.xlsx"

        export_plate_excel(mappings, out, results=results)

        ws = load_workbook(out)["expected_mutations"]
        ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
        # 260722 R2-1 export 가 낸 증상: 같은 94종이 두 시트에서 다른 순서였다.
        assert ids == ["S11I", "K53I", "I92D"]

    def test_a_result_without_a_forward_mapping_is_kept_at_the_end(self, tmp_path):
        from openpyxl import load_workbook

        from kuma_core.kuro.plate_mapper import export_plate_excel

        results = [self._result(*a) for a in (
            ("I92D", "I", 92, "D"), ("S11I", "S", 11, "I"),
        )]
        mappings = [self._mapping("A1", "S11I", "forward")]
        out = tmp_path / "platemap.xlsx"

        export_plate_excel(mappings, out, results=results)

        ws = load_workbook(out)["expected_mutations"]
        ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
        # 매핑이 없다고 설계된 변이를 시트에서 잃지는 않는다.
        assert ids == ["S11I", "I92D"]
