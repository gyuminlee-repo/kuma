"""Phase 1 (core) tests for MAME well-layout: build_draft_layout + WT-aware scoping.

Covers:
- build_draft_layout column-major order, WT placement, and 96-well clamping.
- well_layout injection into run_analyze: mutant wells scoped by ground truth,
  WT well carries an empty expected scope (clean -> PASS, variant -> fail), and a
  contaminated WT well is NOT mis-grouped into a real mutant's replicate group.

Self-contained fixtures; no minimap2 dependency (barcode-mode FASTA ingest).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.ingest import IngestMode
from kuma_core.mame.layout import (
    apply_well_selection,
    build_draft_layout,
    normalise_selected_wells,
)
from kuma_core.mame.models import ExpectedMutation, VerdictClass
from kuma_core.mame.pipeline import _norm_well, run_analyze

# Reference ATG GGG TTT -> M G F (9 bp, table 11).
_REFERENCE_NT = "ATGGGGTTT"
# Well A02 (custom_barcode "1_2"): G2A -> M A F
_G2A_NT = "ATGGCGTTT"
# Well B01 (custom_barcode "2_1"): F3W -> M G W
_F3W_NT = "ATGGGGTGG"
# Clean WT consensus == reference -> M G F
_WT_NT = "ATGGGGTTT"
_PAD = "\n" * (52 * 1024)


def _em(mutant_id: str, position: int, wt_aa: str, mt_aa: str) -> ExpectedMutation:
    return ExpectedMutation(
        mutant_id=mutant_id, position=position, wt_aa=wt_aa, mt_aa=mt_aa,
        wt_codon="", mt_codon="", group_id="", primer_set_ref="",
        notation_type="substitution", status="DESIGNED",
    )


# ---------------------------------------------------------------------------
# build_draft_layout
# ---------------------------------------------------------------------------

def test_build_draft_layout_column_major_order_and_wt_position() -> None:
    """well 1..N column-major -> mutant_id; well N+1 -> WT."""
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D"), _em("M3", 3, "E", "F")]
    draft = build_draft_layout(expected)
    # seq 1->A1, 2->B1, 3->C1 (column-major), WT at seq 4 -> D1
    assert draft.layout["A1"] == "M1"
    assert draft.layout["B1"] == "M2"
    assert draft.layout["C1"] == "M3"
    assert draft.layout["D1"] == "WT"
    assert len(draft.layout) == 4
    assert draft.is_complete
    assert draft.dropped_mutant_ids == []


def test_build_draft_layout_places_wt_at_the_ordinal_the_source_stated() -> None:
    """An explicit WT row takes its own well and moves the rest one down.

    The regression this pins: the reader used to discard the WT row, so the
    mutants after it each moved one well UP and the result read as a correct
    plate. M2 belongs in C1 here because WT sits in B1, not in B1 itself.
    """
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D"), _em("M3", 3, "E", "F")]
    draft = build_draft_layout(expected, wt_ordinal=2)
    assert list(draft.layout.items()) == [
        ("A1", "M1"),
        ("B1", "WT"),
        ("C1", "M2"),
        ("D1", "M3"),
    ]
    assert draft.is_complete


def test_build_draft_layout_wt_first_is_the_same_rule() -> None:
    """WT at ordinal 1 is not a special case, just the first occupant."""
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D")]
    draft = build_draft_layout(expected, wt_ordinal=1)
    assert list(draft.layout.items()) == [("A1", "WT"), ("B1", "M1"), ("C1", "M2")]


def test_build_draft_layout_refuses_96_mutants_because_wt_needs_a_well() -> None:
    """N == 96 does not fit: the WT control is a 97th occupant.

    Judging capacity on N alone let this through, and the placement loop then
    asked ``seq_to_well`` for well 97 and raised mid-run. Nothing partial comes
    back, because a 96-row layout missing its last mutant reads as a full plate.
    """
    expected = [_em(f"M{i}", i, "A", "B") for i in range(1, 97)]
    draft = build_draft_layout(expected)
    assert draft.layout == {}
    assert draft.dropped_mutant_ids == ["M96"]
    assert not draft.is_complete


def test_build_draft_layout_95_mutants_plus_wt_is_exactly_full() -> None:
    """The largest set that fits, and it fills the plate to H12."""
    expected = [_em(f"M{i}", i, "A", "B") for i in range(1, 96)]
    draft = build_draft_layout(expected)
    assert len(draft.layout) == 96
    assert draft.layout["H12"] == "WT"
    assert draft.is_complete


def test_build_draft_layout_names_every_mutant_that_does_not_fit() -> None:
    """N > 95: nothing is placed, and the overflow is named rather than counted."""
    expected = [_em(f"M{i}", i, "A", "B") for i in range(1, 110)]
    draft = build_draft_layout(expected)
    assert draft.layout == {}
    # A truncated draft reads as a correct full plate, so the mutants that do not
    # fit are reported by id rather than only by a count a caller may ignore.
    assert draft.dropped_mutant_ids == [f"M{i}" for i in range(96, 110)]
    assert not draft.is_complete


# ---------------------------------------------------------------------------
# Declaring which wells the campaign occupies
# ---------------------------------------------------------------------------

def test_selecting_the_leading_wells_reproduces_the_draft_exactly() -> None:
    """The default has to be a no-op, or every existing run changes.

    ``selected_wells`` defaults to absent, and absent means the leading N+1
    wells. Selecting those explicitly must therefore produce the same mapping,
    byte for byte, that the draft already had.
    """
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D")]
    draft = build_draft_layout(expected)

    reseated = apply_well_selection(draft, ["A1", "B1", "C1"])

    assert reseated.layout == draft.layout


def test_occupants_keep_the_wells_the_draft_gave_them() -> None:
    """The selection narrows the draft, it does not re-seat it.

    Occupant *i* used to take the *i*th declared well, so leaving one out slid
    every later variant up and the plate the operator was looking at rearranged
    itself under a click meant to describe it.
    """
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D")]
    draft = build_draft_layout(expected)

    narrowed = apply_well_selection(draft, ["A1", "C1", "E1"])

    # The draft is A1=M1, B1=M2, C1=WT. B1 was not declared, so M2 is off the
    # plate and WT stays in C1 rather than moving up into it.
    assert narrowed.layout == {"A1": "M1", "C1": "WT"}
    assert narrowed.excluded_occupants == {"B1": "M2"}
    assert narrowed.unused_wells == ["E1"]


def test_the_selection_order_the_caller_sent_does_not_matter() -> None:
    """Plate order is imposed here, so a click order cannot re-place the plate."""
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D")]
    draft = build_draft_layout(expected)

    scrambled = apply_well_selection(draft, ["E1", "A1", "C1"])
    ordered = apply_well_selection(draft, ["A1", "C1", "E1"])

    assert scrambled.layout == ordered.layout


def test_a_selection_smaller_than_the_campaign_names_what_it_leaves_out() -> None:
    """A partly filled plate is the ordinary case, not a refusal.

    It was refused while the rule was re-seating, because a short list left an
    occupant with nowhere to go. Nothing moves now, so the wells past the
    declaration simply hold nothing, and what the draft put there is named:
    those variants have no verdict anywhere on the run.
    """
    expected = [_em("M1", 1, "A", "B"), _em("M2", 2, "C", "D")]
    draft = build_draft_layout(expected)

    narrowed = apply_well_selection(draft, ["A1", "B1"])

    assert narrowed.layout == {"A1": "M1", "B1": "M2"}
    assert narrowed.excluded_occupants == {"C1": "WT"}
    assert narrowed.unused_wells == []


def test_normalise_selected_wells_sorts_dedupes_and_bounds() -> None:
    assert normalise_selected_wells(["B1", "A1", "A1", " a2 "]) == ["A1", "B1", "A2"]
    # Off-plate labels are dropped rather than raising: a caller that cares
    # compares the length it sent with the length it got back.
    assert normalise_selected_wells(["A1", "I1", "A13", "nonsense"]) == ["A1"]


# ---------------------------------------------------------------------------
# well_layout injection into run_analyze
# ---------------------------------------------------------------------------

def _write_fasta(path: Path, header: str, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    # Carry a depth=N header well above the recommended min_read_count=30 so the
    # real read-depth gate clears these wells; these tests exercise well-layout
    # scoping, not the LOWDEPTH gate. custom_barcode is header.split()[0], so the
    # trailing metadata does not disturb the barcode token.
    path.write_text(f">{header} depth=100\n{body}\n{_PAD}", encoding="utf-8")


def _make_kuro_xlsx(dest: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fwd List"
    ws.append(["Well", "Primer Name", "Sequence", "Length", "Tm", "Tm_Overlap",
               "WT_Codon", "MT_Codon", "Mutation"])
    ws2 = wb.create_sheet("expected_mutations")
    ws2.append(["mutant_id", "position", "wt_aa", "mt_aa", "wt_codon", "mt_codon",
                "group_id", "primer_set_ref", "notation_type", "status"])
    ws2.append(["G2A", 2, "G", "A", "GGG", "GCG", "", "G2A", "substitution", "DESIGNED"])
    ws2.append(["F3W", 3, "F", "W", "TTT", "TGG", "", "F3W", "substitution", "DESIGNED"])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def _make_reference_fasta(tmp_path: Path) -> Path:
    ref = tmp_path / "reference.fasta"
    ref.write_text(f">ref\n{_REFERENCE_NT}\n", encoding="utf-8")
    return ref


def test_well_layout_scopes_mutant_wells_and_clean_wt_passes(tmp_path: Path) -> None:
    """well_layout injection: mutant wells scoped (PASS), clean WT well PASSes.

    Layout: A02 -> G2A, B01 -> F3W, C01 -> WT.
      custom_barcode "1_2" -> seq 9 -> A2;  "2_1" -> seq 2 -> B1;  "3_1" -> seq 3 -> C1
    """
    ingest = tmp_path / "consensus"
    _write_fasta(ingest / "NB01" / "1_2.fasta", "1_2", _G2A_NT)   # A02 G2A
    _write_fasta(ingest / "NB01" / "2_1.fasta", "2_1", _F3W_NT)   # B01 F3W
    _write_fasta(ingest / "NB01" / "3_1.fasta", "3_1", _WT_NT)    # C01 WT (clean)
    reference = _make_reference_fasta(tmp_path)
    kuro = tmp_path / "kuro.xlsx"
    _make_kuro_xlsx(kuro)

    well_layout = {"A2": "G2A", "B1": "F3W", "C1": "WT"}

    verdicts, _ = run_analyze(
        input_dir=ingest, reference_path=reference, expected_path=kuro,
        output_path=tmp_path / "out.xlsx", cds_start=0, cds_end=9,
        min_file_size_kb=0.0, ingest_mode=IngestMode.BARCODE, well_layout=well_layout,
    )
    by_custom = {v.translated.barcode.custom_barcode: v for v in verdicts}
    assert by_custom["1_2"].verdict is VerdictClass.PASS, by_custom["1_2"].verdict_notes
    assert by_custom["2_1"].verdict is VerdictClass.PASS, by_custom["2_1"].verdict_notes
    # WT well scoped to [] (empty expected); clean consensus == reference -> PASS.
    wt = by_custom["3_1"]
    assert wt.expected_mutations == [], wt.expected_mutations
    assert wt.verdict is VerdictClass.PASS, wt.verdict_notes


def test_well_layout_wt_well_with_variant_fails_and_not_mis_grouped(tmp_path: Path) -> None:
    """A WT well observing a variant at a real mutant's position must fail AND must
    not be grouped under that mutant (ground-truth attribution beats the heuristic)."""
    ingest = tmp_path / "consensus"
    # WT well C01 is contaminated with G2A (matches mutant G2A's position 2).
    _write_fasta(ingest / "NB01" / "3_1.fasta", "3_1", _G2A_NT)   # C01 declared WT, observes G2A
    reference = _make_reference_fasta(tmp_path)
    kuro = tmp_path / "kuro.xlsx"
    _make_kuro_xlsx(kuro)

    well_layout = {"C1": "WT"}

    verdicts, replicates = run_analyze(
        input_dir=ingest, reference_path=reference, expected_path=kuro,
        output_path=tmp_path / "out.xlsx", cds_start=0, cds_end=9,
        min_file_size_kb=0.0, ingest_mode=IngestMode.BARCODE, well_layout=well_layout,
    )
    wt = verdicts[0]
    assert wt.expected_mutations == []
    # Empty expected + observed variant -> not a PASS (WRONG_AA: unexpected extra).
    assert wt.verdict is not VerdictClass.PASS, wt.verdict_notes
    # Must NOT be grouped under the real mutant G2A; ground truth pins it to "WT".
    mutant_ids = {r.mutant_id for r in replicates}
    assert "G2A" not in mutant_ids, (
        f"contaminated WT well mis-grouped into G2A; groups={mutant_ids}"
    )
    assert "WT" in mutant_ids, f"WT well not attributed to WT group; groups={mutant_ids}"


# ---------------------------------------------------------------------------
# IPC: mame.build_well_layout RPC handler + dispatcher registration
# ---------------------------------------------------------------------------

def test_build_well_layout_handler_returns_ordered_draft(tmp_path: Path) -> None:
    """Handler reads the KURO xlsx and returns an ordered draft + count."""
    from sidecar_mame.handlers.build_well_layout import handle_build_well_layout

    kuro = tmp_path / "kuro.xlsx"
    _make_kuro_xlsx(kuro)  # G2A (seq1->A1), F3W (seq2->B1); WT at seq3 -> C1

    result = handle_build_well_layout({"expected_mutations_xlsx": str(kuro)})

    assert result["count"] == 3
    draft = result["draft"]
    assert draft == [
        {"well": "A1", "sample": "G2A"},
        {"well": "B1", "sample": "F3W"},
        {"well": "C1", "sample": "WT"},
    ], draft


def test_build_well_layout_params_rejects_missing_file(tmp_path: Path) -> None:
    """A non-existent xlsx path fails validation (fail-fast, no silent fallback)."""
    from sidecar_mame.models import BuildWellLayoutParams

    with pytest.raises(ValueError):
        BuildWellLayoutParams.model_validate(
            {"expected_mutations_xlsx": str(tmp_path / "missing.xlsx")}
        )


def test_build_well_layout_params_rejects_path_traversal() -> None:
    """Path-traversal segments are rejected."""
    from sidecar_mame.models import BuildWellLayoutParams

    with pytest.raises(ValueError):
        BuildWellLayoutParams.model_validate(
            {"expected_mutations_xlsx": "../../etc/passwd.xlsx"}
        )


def test_build_well_layout_registered_synchronous() -> None:
    """build_well_layout is in _METHODS and is NOT an async method (stat-only)."""
    from sidecar_mame.dispatcher import _ASYNC_METHODS, _METHODS

    assert "mame.build_well_layout" in _METHODS
    assert "mame.build_well_layout" not in _ASYNC_METHODS


# ---------------------------------------------------------------------------
# Plain variant lists on the analysis and layout paths
#
# The generic adapter used to be wired to the sample map template only, so a
# workbook with a single ``variant`` column reached ``mame.build_well_layout``
# and ``analyze`` as a KURO export that was missing its sheet. These pin the two
# shapes side by side: a plain list is read, and a KURO export is unchanged.
# ---------------------------------------------------------------------------


def _make_variant_list_xlsx(dest: Path, labels: list[str], header: str = "variant") -> Path:
    """A workbook with one variant column and no ``expected_mutations`` sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Variants"
    ws.append([header])
    for label in labels:
        ws.append([label])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def test_build_well_layout_reads_a_plain_variant_list(tmp_path: Path) -> None:
    from sidecar_mame.handlers.build_well_layout import handle_build_well_layout

    path = _make_variant_list_xlsx(tmp_path / "variants.xlsx", ["G2A", "F3W"])

    result = handle_build_well_layout({"expected_mutations_xlsx": str(path)})

    assert result["draft"] == [
        {"well": "A1", "sample": "G2A"},
        {"well": "B1", "sample": "F3W"},
        {"well": "C1", "sample": "WT"},
    ]
    assert result["count"] == 3


def test_build_well_layout_does_not_append_a_second_wt(tmp_path: Path) -> None:
    """A source carrying its own WT row gets exactly one control well: that one.

    Not zero, which is what dropping the row produced. The row states where the
    control sits, so it takes the well its position names and the mutants after
    it stay where they were.
    """
    from sidecar_mame.handlers.build_well_layout import handle_build_well_layout

    path = _make_variant_list_xlsx(tmp_path / "with_wt.xlsx", ["G2A", "WT", "F3W"])

    result = handle_build_well_layout({"expected_mutations_xlsx": str(path)})

    assert result["draft"] == [
        {"well": "A1", "sample": "G2A"},
        {"well": "B1", "sample": "WT"},
        {"well": "C1", "sample": "F3W"},
    ]
    assert [row["sample"] for row in result["draft"]].count("WT") == 1


def test_build_well_layout_honours_an_explicit_sheet_and_column(tmp_path: Path) -> None:
    """Two candidate columns are ambiguous until the caller names one."""
    from sidecar_mame.handlers.build_well_layout import handle_build_well_layout

    path = tmp_path / "two_columns.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Round 2"
    ws.append(["alpha", "beta"])
    ws.append(["G2A", "F3W"])
    wb.save(path)

    result = handle_build_well_layout(
        {
            "expected_mutations_xlsx": str(path),
            "variant_sheet": "Round 2",
            "variant_column": "beta",
        }
    )

    assert result["draft"][0] == {"well": "A1", "sample": "F3W"}


def test_build_well_layout_kuro_export_is_unchanged(tmp_path: Path) -> None:
    """The KURO shape keeps taking the strict reader, WT control included."""
    from sidecar_mame.handlers.build_well_layout import handle_build_well_layout

    kuro = tmp_path / "kuro.xlsx"
    _make_kuro_xlsx(kuro)

    assert handle_build_well_layout({"expected_mutations_xlsx": str(kuro)}) == {
        "draft": [
            {"well": "A1", "sample": "G2A"},
            {"well": "B1", "sample": "F3W"},
            {"well": "C1", "sample": "WT"},
        ],
        "count": 3,
        "dropped_mutant_ids": [],
    }


def test_run_analyze_reads_a_plain_variant_list(tmp_path: Path) -> None:
    """The analysis path scores wells from a workbook with no KURO sheet."""
    ingest = tmp_path / "consensus"
    _write_fasta(ingest / "NB01" / "1_2.fasta", "1_2", _G2A_NT)   # A02 G2A
    _write_fasta(ingest / "NB01" / "2_1.fasta", "2_1", _F3W_NT)   # B01 F3W
    reference = _make_reference_fasta(tmp_path)
    variants = _make_variant_list_xlsx(tmp_path / "variants.xlsx", ["G2A", "F3W"])

    verdicts, _ = run_analyze(
        input_dir=ingest, reference_path=reference, expected_path=variants,
        output_path=tmp_path / "out.xlsx", cds_start=0, cds_end=9,
        min_file_size_kb=0.0, ingest_mode=IngestMode.BARCODE,
        well_layout={"A2": "G2A", "B1": "F3W"},
    )

    by_custom = {v.translated.barcode.custom_barcode: v for v in verdicts}
    assert by_custom["1_2"].verdict is VerdictClass.PASS, by_custom["1_2"].verdict_notes
    assert by_custom["2_1"].verdict is VerdictClass.PASS, by_custom["2_1"].verdict_notes


def test_validate_inputs_accepts_a_plain_variant_list(tmp_path: Path) -> None:
    """Validation used to demand the KURO sheet and rejected a readable list."""
    from sidecar_mame.handlers.analyze import handle_validate_inputs

    ingest = tmp_path / "consensus"
    _write_fasta(ingest / "NB01" / "1_2.fasta", "1_2", _G2A_NT)
    reference = _make_reference_fasta(tmp_path)
    variants = _make_variant_list_xlsx(tmp_path / "variants.xlsx", ["G2A", "F3W"])

    result = handle_validate_inputs(
        {
            "input_dir": str(ingest),
            "reference": str(reference),
            "expected": str(variants),
            "cds_end": 9,
        }
    )

    assert result == {"valid": True, "errors": []}


def test_validate_inputs_reports_why_a_file_cannot_be_read(tmp_path: Path) -> None:
    """The error names what is needed instead of pointing at a KURO commit."""
    from sidecar_mame.handlers.analyze import handle_validate_inputs

    ingest = tmp_path / "consensus"
    _write_fasta(ingest / "NB01" / "1_2.fasta", "1_2", _G2A_NT)
    reference = _make_reference_fasta(tmp_path)
    path = tmp_path / "ambiguous.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["alpha", "beta"])
    ws.append(["G2A", "F3W"])
    wb.save(path)

    result = handle_validate_inputs(
        {
            "input_dir": str(ingest),
            "reference": str(reference),
            "expected": str(path),
            "cds_end": 9,
        }
    )

    assert result["valid"] is False
    (message,) = result["errors"]
    assert "which column holds the variants" in message
    assert "8c47037" not in message


# ---------------------------------------------------------------------------
# The inferred layout follows the workbook's own row order
# ---------------------------------------------------------------------------

def _expected_xlsx_in_order(dest: Path, mutant_ids: list[str]) -> Path:
    """A KURO export whose ``expected_mutations`` rows appear in *mutant_ids* order."""
    rows = {
        "G2A": ["G2A", 2, "G", "A", "GGG", "GCG", "", "G2A", "substitution", "DESIGNED"],
        "F3W": ["F3W", 3, "F", "W", "TTT", "TGG", "", "F3W", "substitution", "DESIGNED"],
    }
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "expected_mutations"
    ws.append(["mutant_id", "position", "wt_aa", "mt_aa", "wt_codon", "mt_codon",
               "group_id", "primer_set_ref", "notation_type", "status"])
    for mutant_id in mutant_ids:
        ws.append(rows[mutant_id])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def test_inferred_layout_follows_the_workbook_row_order_and_nothing_else(
    tmp_path: Path,
) -> None:
    """Reorder the workbook's rows and the wells move with them.

    ``test_handle_analyze_auto_scopes_from_expected_when_layout_omitted``
    (tests/sidecar_mame/test_analyze_raw_run.py) already pins that a run given no
    ``well_layout`` places wells from the chosen workbook. It reads one workbook, whose rows happen to be in residue-position
    order, so it cannot say WHICH property of that file the placement followed.

    This varies only the row order, holding the reads and the plate fixed. The
    designed rows are written F3W (position 3) before G2A (position 2), so sheet
    order and position order now disagree and the wells must follow the sheet:
    A01 declared F3W while its reads carry G2A, B01 the reverse, both failing.
    Sorting by position instead -- ``canonical_plate_order`` is one import away
    and its docstring offers "the plate the bench actually fills" -- would leave
    every other test in the suite green while silently re-placing every well of
    any workbook not already position-ordered.

    That is the 2026-08-06 shape exactly: a 95-well layout in design-ranking
    order scored a plate filled in position order, agreed with it at 0 of 95
    wells, and reported 0 PASS from sequencing that was in fact 262/285 correct.
    """
    from sidecar_mame.handlers.analyze import handle_analyze

    ingest = tmp_path / "consensus"
    _write_fasta(ingest / "NB01" / "1_1.fasta", "1_1", _G2A_NT)   # A01 observes G2A
    _write_fasta(ingest / "NB01" / "2_1.fasta", "2_1", _F3W_NT)   # B01 observes F3W

    result = handle_analyze({
        "input_dir": str(ingest),
        "reference": str(_make_reference_fasta(tmp_path)),
        "expected": str(_expected_xlsx_in_order(tmp_path / "expected.xlsx", ["F3W", "G2A"])),
        "output": str(tmp_path / "out.xlsx"),
        "cds_start": 0,
        "cds_end": 9,
        "min_file_size_kb": 0.0,
        "min_read_count": 0,
        "ingest_mode": "barcode",
    })

    # The run has to name this branch, or a layout it derived itself could pass
    # downstream as one the operator supplied.
    assert result["layout_provenance"]["source"] == "inferred_draft_layout"
    assert result["layout_provenance"]["selected_wells"] is None

    by_custom = {v["custom_barcode"]: v for v in result["verdicts"]}
    assert by_custom["1_1"]["expected_mutations"] == ["F3W"], by_custom["1_1"]
    assert by_custom["2_1"]["expected_mutations"] == ["G2A"], by_custom["2_1"]
    assert by_custom["1_1"]["verdict"] != "PASS", by_custom["1_1"]
    assert by_custom["2_1"]["verdict"] != "PASS", by_custom["2_1"]
