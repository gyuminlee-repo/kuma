"""KURO xlsx reader tests (Blocker B acceptance)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.io.kuro_reader import (
    expected_to_labels,
    read_expected_mutations,
    read_expected_mutations_with_rows,
)


def test_read_expected_mutations_returns_two_rows(kuro_xlsx_path: Path) -> None:
    result = read_expected_mutations(kuro_xlsx_path)
    assert len(result) == 2
    labels = expected_to_labels(result)
    assert labels == ["V5F", "K53N"]


def test_missing_expected_sheet_raises(tmp_path: Path) -> None:
    bad = tmp_path / "KURO_legacy.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fwd List"
    ws.append(["Well", "Primer Name"])
    wb.save(bad)

    with pytest.raises(ValueError):
        read_expected_mutations(bad)


def test_read_expected_mutations_accepts_rescue_status_from_interim_exports(tmp_path: Path) -> None:
    path = tmp_path / "KURO_rescue_status.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "expected_mutations"
    ws.append([
        "mutant_id", "position", "wt_aa", "mt_aa",
        "wt_codon", "mt_codon", "group_id", "primer_set_ref",
        "notation_type", "status", "rescue_type", "rescue_stage", "rescued_from",
    ])
    ws.append(["K53N", 53, "K", "N", "AAG", "AAC", "", "K53N", "substitution", "auto_suggestion_l2", "", "", ""])
    ws.append(["E61Y", 61, "E", "Y", "GAA", "TAT", "", "E61Y", "substitution", "FAILED", "", "", ""])
    wb.save(path)

    result = read_expected_mutations(path)

    assert expected_to_labels(result) == ["K53N"]


def _write_export(path: Path, rows: list[tuple[str, str]]) -> Path:
    """`(mutant_id, status)` rows on a minimal `expected_mutations` sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "expected_mutations"
    ws.append([
        "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
        "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
    ])
    for mutant_id, status in rows:
        ws.append([mutant_id, 0, "-", "-", "-", "-", "G0", "-", "wt", status])
    wb.save(path)
    return path


def test_wild_type_row_survives_a_status_outside_the_designed_set(tmp_path: Path) -> None:
    """KURO writes its control row as `status=control`, which is no design.

    Dropping it is not a neutral omission: the caller reads row order as plate
    order, so the row it never sees moves every later mutant one well up and the
    whole workbook is refused over it. Membership is decided by the label here.
    """
    path = _write_export(
        tmp_path / "kuro.xlsx",
        [("M001", "designed"), ("WT", "control"), ("M002", "designed")],
    )

    read = read_expected_mutations_with_rows(path)

    assert [m.mutant_id for m in read.expected] == ["M001", "WT", "M002"]
    assert read.row_numbers == [2, 3, 4]
    assert read.dropped_rows == []


def test_the_control_row_keeps_the_values_the_sheet_wrote(tmp_path: Path) -> None:
    path = _write_export(tmp_path / "kuro.xlsx", [("WT", "control")])

    (wt,) = read_expected_mutations_with_rows(path).expected

    assert (wt.position, wt.wt_aa, wt.mt_aa, wt.status) == (0, "-", "-", "control")


def test_a_mutant_row_is_still_dropped_on_status(tmp_path: Path) -> None:
    """판정은 라벨로만 넓어졌다. status 자체를 넓히면 진짜 실패 행이 통과한다."""
    path = _write_export(
        tmp_path / "kuro.xlsx",
        [("M001", "designed"), ("M002", "control"), ("M003", "FAILED")],
    )

    read = read_expected_mutations_with_rows(path)

    assert [m.mutant_id for m in read.expected] == ["M001"]
    assert read.dropped_rows == [(3, "control"), (4, "FAILED")]


def test_the_wild_type_exemption_ignores_status_entirely(tmp_path: Path) -> None:
    """면제는 라벨로만 판정한다. status 는 WT 행에 대해 아무 힘이 없다.

    조작자가 대조군 행을 손으로 FAILED 로 표시해도 그 행은 웰을 차지한다.
    행을 빼면 뒤 변이가 한 칸씩 밀리는데, 대조군의 실패는 그 자체로 판정
    대상이지 배치에서 지울 근거가 아니다. 이 선택을 여기에 고정해 둔다.
    """
    path = _write_export(
        tmp_path / "kuro.xlsx",
        [("M001", "designed"), ("WT", "FAILED"), ("M002", "designed")],
    )

    read = read_expected_mutations_with_rows(path)

    assert [m.mutant_id for m in read.expected] == ["M001", "WT", "M002"]
    assert read.dropped_rows == []


@pytest.mark.parametrize("label", ["WT", "wt", "Wild-Type", "control"])
def test_every_wild_type_spelling_is_recognised(tmp_path: Path, label: str) -> None:
    path = _write_export(tmp_path / "kuro.xlsx", [(label, "control")])

    assert [m.mutant_id for m in read_expected_mutations(path)] == [label]
