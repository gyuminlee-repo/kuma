"""The issued template is what the reader expects to get back.

Two things are worth failing over. The wells must be in the plate order MAME
fills (down a column, not across a row), and the control must land where the
bench puts it. The well assertions are written as literals rather than derived
from ``seq_to_well``, because deriving them would only prove the template calls
the same function twice: row 2 is ``B1`` under column fill and ``A2`` under row
fill, and that is the pair that tells the two apart.
"""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from kuma_core.mame.io.variant_template import (
    CONTROL_LABEL,
    TEMPLATE_HEADERS,
    TEMPLATE_SHEET,
    default_control_well,
    write_variant_template,
)
from kuma_core.mame.models import WT_LABELS


def _read(path):
    """Sheet name and the rows below the header, as ``(well, variant)``."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = [
            (row[0], row[1] if len(row) > 1 else None)
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()
    return workbook.sheetnames[0], rows[0], rows[1:]


def test_template_has_one_row_per_well_in_column_fill_order(tmp_path):
    _, header, rows = _read(write_variant_template(tmp_path / "t.xlsx"))

    assert header == TEMPLATE_HEADERS
    assert len(rows) == 96
    wells = [row[0] for row in rows]
    # Literal anchors, not derived: A2 at row 2 would mean row-major fill.
    assert wells[0] == "A1"
    assert wells[1] == "B1"
    assert wells[7] == "H1"
    assert wells[8] == "A2"
    assert wells[95] == "H12"
    assert len(set(wells)) == 96


def test_control_defaults_to_the_last_well(tmp_path):
    _, _, rows = _read(write_variant_template(tmp_path / "t.xlsx"))

    filled = [(well, variant) for well, variant in rows if variant is not None]
    assert filled == [("H12", CONTROL_LABEL)]
    assert default_control_well() == "H12"


def test_control_label_is_one_the_reader_treats_as_wild_type():
    # The template is only useful if the label it writes is the label the
    # reader recognises, so the two are asserted against each other.
    assert CONTROL_LABEL.lower() in WT_LABELS


def test_sheet_is_not_named_like_a_kuro_export(tmp_path):
    # `expected_mutations` routes the file to the strict ten-column reader.
    name, _, _ = _read(write_variant_template(tmp_path / "t.xlsx"))
    assert name == TEMPLATE_SHEET
    assert name != "expected_mutations"


def test_control_well_is_an_argument(tmp_path):
    _, _, rows = _read(
        write_variant_template(tmp_path / "t.xlsx", control_well="A1")
    )

    filled = [(well, variant) for well, variant in rows if variant is not None]
    assert filled == [("A1", CONTROL_LABEL)]


def test_control_can_be_left_out_entirely(tmp_path):
    _, _, rows = _read(
        write_variant_template(tmp_path / "t.xlsx", include_control=False)
    )

    assert len(rows) == 96
    assert all(variant is None for _, variant in rows)


def test_control_well_off_the_plate_is_refused(tmp_path):
    with pytest.raises(ValueError):
        write_variant_template(tmp_path / "t.xlsx", control_well="A13")
    with pytest.raises(ValueError):
        write_variant_template(tmp_path / "t.xlsx", control_well="I1")
