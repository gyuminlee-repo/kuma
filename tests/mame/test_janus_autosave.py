"""The pick list is written when the analysis finishes, not on demand.

A run produces two artefacts: the result workbook and the pick list beside it,
recording which variant was selected, where it sits, and where it should be
collected. Until v0.15.6 that automatic file was written in the instrument
schema (then named ``device9``), which made it a robot worklist carrying a
volume and a deck. Those values describe the deck standing in the room at that
moment, so a file stating them was produced by every exploratory re-run and none
of them could be trusted later. The automatic file is now ``legacy5``: the
conclusion of the run and nothing about the instrument.

Properties pinned here, the later ones mattering more than the first: no
instrument setting may block the automatic file, an empty pick list must not be
written (an empty file reads like a finished plate), a plate keeps the name the
run gave it, and a file that cannot be built must not cost the analysis, which by
then has already run to completion. The instrument (``device``) sheet is a
different file: it is never written by ``handle_analyze`` (v0.15.14 removed that
autosave, since a robot worklist states a deck that describes the room at the
moment it is written, not at analyze time), only by a manual
``export_janus_mapping`` call, and that contrast is tested here too.

``device9`` is the former name of the instrument schema, kept alive at the wire
edge so a saved project asking for it still resolves. It appears below only in
payloads standing in for such a client.

Fixtures are self-contained barcode-mode consensus FASTA, shared byte-for-byte
with ``test_analyze_liveness``: no minimap2 needed.
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

# Reference ATG GGG TTT -> M G F (9 bp, table 11).
_REFERENCE_NT = "ATGGGGTTT"
_G2A_NT = "ATGGCGTTT"  # well A02, custom_barcode "1_2"
_F3W_NT = "ATGGGGTGG"  # well B01, custom_barcode "2_1"
_PAD = "\n" * (52 * 1024)

# The five columns of the automatic file, as declared by the core module.
_LEGACY5_HEADER = ["name", "source_plate", "source_well", "dest_well", "priority_score"]

# The liquid class has no default: it describes how the operator pipetted this
# run. The eight column instrument sheet has no column for it, so it is recorded
# on the settings and written to neither file. It appears below only to prove it
# reaches no cell.
_LIQUID_CLASS = "cell_stock_100"


def _write_fasta(path: Path, header: str, body: str, depth: int = 100) -> None:
    """One consensus file. *depth* becomes ``read_count``, and so priority_score.

    It is a parameter because the pick list still carries that number after it
    stopped ordering the file, and telling the two apart needs a run whose depth
    ranking and plate order disagree.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(f">{header} depth={depth}\n{body}\n{_PAD}", encoding="utf-8")


def _make_kuro_xlsx(dest: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "expected_mutations"
    ws.append(["mutant_id", "position", "wt_aa", "mt_aa", "wt_codon", "mt_codon",
               "group_id", "primer_set_ref", "notation_type", "status"])
    ws.append(["G2A", 2, "G", "A", "GGG", "GCG", "", "G2A", "substitution", "DESIGNED"])
    ws.append(["F3W", 3, "F", "W", "TTT", "TGG", "", "F3W", "substitution", "DESIGNED"])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def _make_reference_fasta(tmp_path: Path) -> Path:
    ref = tmp_path / "reference.fasta"
    ref.write_text(f">ref\n{_REFERENCE_NT}\n", encoding="utf-8")
    return ref


def _run(
    tmp_path: Path,
    bodies: dict[str, str],
    plate: str = "NB01",
    depths: dict[str, int] | None = None,
    **extra,
) -> dict:
    """Run ``handle_analyze`` over one well per entry of *bodies*.

    *depths* sets ``depth=`` per well; wells it does not name keep the default.
    """
    ingest = tmp_path / "consensus"
    for barcode, body in bodies.items():
        _write_fasta(
            ingest / plate / f"{barcode}.fasta",
            barcode,
            body,
            depth=(depths or {}).get(barcode, 100),
        )
    from sidecar_mame.handlers.analyze import handle_analyze

    return handle_analyze(
        {
            "input_dir": str(ingest),
            "reference": str(_make_reference_fasta(tmp_path)),
            "expected": str(_make_kuro_xlsx(tmp_path / "kuro.xlsx")),
            "output": str(tmp_path / "260804_ref_MAME.xlsx"),
            "cds_start": 0,
            "cds_end": 9,
            "min_file_size_kb": 0.0,
            "ingest_mode": "barcode",
            **extra,
        }
    )


def _read_csv(path: str) -> list[list[str]]:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        return [row for row in csv.reader(handle) if row and any(c.strip() for c in row)]


def test_pick_list_is_written_next_to_the_result_workbook(tmp_path: Path) -> None:
    """Same folder, same stem as the workbook, so one run reads as one run."""
    result = _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})

    autosave = result["janus_autosave"]
    assert autosave["status"] == "saved", autosave
    written = Path(autosave["output_path"])
    assert written == tmp_path / "260804_ref_MAME_picks.csv"
    assert written.exists()
    assert autosave["format"] == "csv"


def test_the_automatic_file_carries_the_selection_not_instrument_columns(
    tmp_path: Path,
) -> None:
    """legacy5, so no volume and no plate name is stated."""
    result = _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})

    body = _read_csv(result["janus_autosave"]["output_path"])
    assert body[0] == _LEGACY5_HEADER


def test_the_pick_list_reads_in_plate_order_and_still_carries_the_depth(
    tmp_path: Path,
) -> None:
    """The file a run writes by itself is the one the operator picks against.

    It is read beside the step 2.2 plate map, so it runs in the same direction
    as the plate: A1 first, then B1. It sorted by ``priority_score`` DESC until
    this change, which would have put ``F3W`` first here.

    The depths are set apart on purpose. With the fixture default both wells
    carry 100 reads, the two rules produce one list, and the case would prove
    nothing. ``priority_score`` is still written for every row: the number that
    says how deeply a clone was sequenced stays in the operator's hands, it just
    stopped deciding where anything goes.
    """
    result = _run(
        tmp_path,
        {"1_1": _G2A_NT, "2_1": _F3W_NT},
        depths={"1_1": 40, "2_1": 900},
    )

    body = _read_csv(result["janus_autosave"]["output_path"])
    header, rows = body[0], body[1:]
    assert header == _LEGACY5_HEADER
    column = {name: header.index(name) for name in _LEGACY5_HEADER}
    assert [r[column["name"]] for r in rows] == ["G2A", "F3W"]
    assert [r[column["source_well"]] for r in rows] == ["A1", "B1"]
    assert [r[column["dest_well"]] for r in rows] == ["A1", "B1"]
    assert [r[column["priority_score"]] for r in rows] == ["40.0", "900.0"]


def test_no_liquid_class_no_longer_blocks_the_automatic_file(tmp_path: Path) -> None:
    """The point of the change: an unset instrument value is not a reason to
    withhold the conclusion of the run."""
    result = _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})

    autosave = result["janus_autosave"]
    assert autosave["status"] == "saved", autosave
    assert autosave["errors"] == []


@pytest.mark.parametrize("stored_schema", ["device", "device9"])
def test_an_instrument_choice_in_the_dialog_does_not_reach_the_automatic_file(
    tmp_path: Path, stored_schema: str
) -> None:
    """The two files answer different questions, so the schema is not shared.

    An operator who set up the instrument sheet in the dialog still gets the
    selection beside the workbook, not a second worklist.

    The step 3 panel no longer offers an output-column choice and always sends
    the instrument schema, which is exactly the payload below, so this case is
    the only thing standing between the automatic pick list and that schema.
    ``SCHEMA_LEGACY5`` in ``kuma_core`` is therefore live code, not a dead path
    left over from the removed control.

    Both names are run because the former name is folded into the current one at
    the wire edge. A fold written loosely enough to catch ``legacy5`` too would
    turn the pick list into a worklist, and the ``device9`` case is what would
    catch that.
    """
    result = _run(
        tmp_path,
        {"1_2": _G2A_NT, "2_1": _F3W_NT},
        janus_settings={
            "output_schema": stored_schema,
            "liquid_class": _LIQUID_CLASS,
        },
    )

    autosave = result["janus_autosave"]
    assert autosave["status"] == "saved", autosave
    assert _read_csv(autosave["output_path"])[0] == _LEGACY5_HEADER


def test_selection_settings_are_still_honoured(tmp_path: Path) -> None:
    """``dest_layout`` describes how the picks are gathered, not the deck."""
    result = _run(
        tmp_path,
        {"1_2": _G2A_NT, "2_1": _F3W_NT},
        janus_settings={"dest_layout": "source"},
    )

    body = _read_csv(result["janus_autosave"]["output_path"])
    header, rows = body[0], body[1:]
    source_idx = header.index("source_well")
    dest_idx = header.index("dest_well")
    # "source" mirrors the source position; "compact" would have filled from A1.
    assert [r[dest_idx] for r in rows] == [r[source_idx] for r in rows]
    assert [r[source_idx] for r in rows] == ["B1"]


def test_a_plate_is_labelled_the_way_every_other_export_labels_it(
    tmp_path: Path,
) -> None:
    """Native barcode folders are named per run, and the label must survive it.

    Observed on a real v0.15.6 run: the plates were ``sort_barcode07`` and up,
    and the Janus export was the one export that did not run them through
    ``nb_label``. It carried a fixed NB01->P1 dictionary instead, which such a
    name never matched, so the raw folder name was written while the result
    workbook said ``NB07`` for the same plate.
    """
    result = _run(tmp_path, {"2_1": _F3W_NT}, plate="sort_barcode07")

    autosave = result["janus_autosave"]
    assert autosave["status"] == "saved", autosave
    body = _read_csv(autosave["output_path"])
    plate_idx = body[0].index("source_plate")
    assert {row[plate_idx] for row in body[1:]} == {"NB07"}


def test_the_workbook_and_the_pick_list_name_one_plate_one_way(
    tmp_path: Path,
) -> None:
    """The two files of a run are read side by side, so a label split is the bug.

    Both artefacts of the same run are inspected here rather than each against a
    literal: a future change that moves one label moves it in both, and only
    comparing them catches a change that moves just one.
    """
    result = _run(tmp_path, {"2_1": _F3W_NT}, plate="sort_barcode07")

    plate_idx = _read_csv(result["janus_autosave"]["output_path"])[0].index(
        "source_plate"
    )
    picks_labels = {
        row[plate_idx] for row in _read_csv(result["janus_autosave"]["output_path"])[1:]
    }

    workbook = openpyxl.load_workbook(result["output_path"])
    sheet = workbook["NGS Results"]
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    selected_idx = header.index("selected_NB")
    workbook_labels = {
        str(row[selected_idx])
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[selected_idx]
    }

    assert picks_labels == workbook_labels
    assert picks_labels == {"NB07"}


def test_no_pass_writes_no_file_and_says_so(tmp_path: Path) -> None:
    """An empty pick list reads like a finished plate, so none is written."""
    # A well whose consensus is the reference carries no designed mutation, so
    # nothing reaches the pick list.
    result = _run(tmp_path, {"1_2": _REFERENCE_NT})

    autosave = result["janus_autosave"]
    assert autosave["status"] == "skipped", autosave
    assert autosave["row_count"] == 0
    assert autosave["output_path"] is None
    assert not (tmp_path / "260804_ref_MAME_picks.csv").exists()
    # The analysis itself is untouched.
    assert result["output_path"] == str(tmp_path / "260804_ref_MAME.xlsx")


def test_autosaved_file_carries_only_the_selected_replicate(tmp_path: Path) -> None:
    """Three copies of one mutant contribute the one pick, not three rows.

    Well ``2_1`` is B1, the layout position the draft assigns to F3W, so all
    three plates carry the same declared mutant and the selector has a real
    choice to make.
    """
    ingest = tmp_path / "consensus"
    for plate in ("NB01", "NB02", "NB03"):
        _write_fasta(ingest / plate / "2_1.fasta", "2_1", _F3W_NT)
    from sidecar_mame.handlers.analyze import handle_analyze

    result = handle_analyze(
        {
            "input_dir": str(ingest),
            "reference": str(_make_reference_fasta(tmp_path)),
            "expected": str(_make_kuro_xlsx(tmp_path / "kuro.xlsx")),
            "output": str(tmp_path / "260804_ref_MAME.xlsx"),
            "cds_start": 0,
            "cds_end": 9,
            "min_file_size_kb": 0.0,
            "ingest_mode": "barcode",
        }
    )

    autosave = result["janus_autosave"]
    assert autosave["status"] == "saved", autosave
    assert autosave["row_count"] == 1
    # Header plus exactly one pick, whatever the replicate count behind it.
    assert len(_read_csv(autosave["output_path"])) == 2


def test_the_dialog_writes_the_instrument_sheet_and_never_the_liquid_class(
    tmp_path: Path,
) -> None:
    """The instrument schema by default, and the liquid class reaches no cell.

    Same session as the run above, so this is the very state the dialog exports
    from. The sheet once refused to be written without a liquid class, then
    shipped a blank column and warned about it; now the column is gone, so the
    value is recorded on the settings and written nowhere. Exporting with and
    without one has to produce the same eight columns, and neither file may
    contain the string.
    """
    _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})
    from sidecar_mame.handlers.export import handle_export_janus_mapping

    blank = tmp_path / "worklist_blank.csv"
    written_blank = handle_export_janus_mapping({"output": str(blank)})
    assert Path(written_blank["output_path"]).exists()
    assert len(_read_csv(written_blank["output_path"])[0]) == 8
    # Nothing is said about a value that reaches no file.
    assert [w["code"] for w in written_blank["warnings"]] == ["derived_source_rack"]

    target = tmp_path / "worklist.csv"
    written = handle_export_janus_mapping(
        {"output": str(target), "liquid_class": _LIQUID_CLASS}
    )
    assert Path(written["output_path"]).exists()
    # Eight instrument columns, the point of that file.
    assert len(_read_csv(written["output_path"])[0]) == 8
    # Scanning every cell, rather than the column it used to sit in, is what
    # catches the liquid class reappearing somewhere else.
    rows = _read_csv(written["output_path"])
    assert not [r for r in rows if _LIQUID_CLASS in r]


# Column-major well order over the expected-mutation sheet: seq 1 ("1_1"),
# seq 2 ("2_1"), seq 3 ("3_1").
_M1L_NT = "CTGGGGTTT"  # well A01, custom_barcode "1_1"


def _make_three_mutant_xlsx(dest: Path) -> Path:
    """Expected-mutation sheet for the three-plate run below."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "expected_mutations"
    ws.append(["mutant_id", "position", "wt_aa", "mt_aa", "wt_codon", "mt_codon",
               "group_id", "primer_set_ref", "notation_type", "status"])
    ws.append(["M1L", 1, "M", "L", "ATG", "CTG", "", "M1L", "substitution", "DESIGNED"])
    ws.append(["G2A", 2, "G", "A", "GGG", "GCG", "", "G2A", "substitution", "DESIGNED"])
    ws.append(["F3W", 3, "F", "W", "TTT", "TGG", "", "F3W", "substitution", "DESIGNED"])
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    return dest


def test_a_three_plate_native_barcode_run_writes_the_instrument_sheet_on_export(
    tmp_path: Path,
) -> None:
    """``sort_barcode07/08/09``, the run that produced no mapping file at all.

    Its plates matched nothing in the fixed NB01/NB02/NB03 rack map, so v0.15.6
    refused to write and v0.15.7 stopped writing an instrument sheet
    automatically at all; v0.15.14 stopped analyze from writing one altogether
    (only a manual ``export_janus_mapping`` call does now, tested here directly
    the way the dialog path already is above). Nothing is configured here: no
    liquid class, no plate names. The file has to come out anyway, with the
    deck taken from the plates of the run.

    Three plates also separates the naming rule from the plate number: the run
    is sort_barcode07/08/09, so the plates are the first, second and third stock
    plate rather than the seventh, eighth and ninth.
    """
    ingest = tmp_path / "consensus"
    _write_fasta(ingest / "sort_barcode07" / "1_1.fasta", "1_1", _M1L_NT)
    _write_fasta(ingest / "sort_barcode08" / "2_1.fasta", "2_1", _G2A_NT)
    _write_fasta(ingest / "sort_barcode09" / "3_1.fasta", "3_1", _F3W_NT)

    from sidecar_mame.handlers.analyze import handle_analyze
    from sidecar_mame.handlers.export import handle_export_janus_mapping

    result = handle_analyze(
        {
            "input_dir": str(ingest),
            "reference": str(_make_reference_fasta(tmp_path)),
            "expected": str(_make_three_mutant_xlsx(tmp_path / "kuro.xlsx")),
            "output": str(tmp_path / "260804_ref_MAME.xlsx"),
            "cds_start": 0,
            "cds_end": 9,
            "min_file_size_kb": 0.0,
            "ingest_mode": "barcode",
        }
    )
    # Analyze wrote no instrument sheet at all: only the pick list.
    assert "janus_mapping_autosave" not in result
    assert not (tmp_path / "260804_ref_MAME_janus.csv").exists()

    target = tmp_path / "worklist.csv"
    mapping = handle_export_janus_mapping({"output": str(target)})
    written = Path(mapping["output_path"])
    assert written == target
    assert written.exists()

    body = _read_csv(str(written))
    assert len(body[0]) == 8
    rows = body[1:]
    assert len(rows) == 3
    # Asp. Rack (index 3) by plate, Dsp. Rack (index 5) the one culture plate.
    by_name = {row[0]: row for row in rows}
    assert by_name["M1L"][3] == "Stock plate1"
    assert by_name["G2A"][3] == "Stock plate2"
    assert by_name["F3W"][3] == "Stock plate3"
    assert {row[5] for row in rows} == {"final culture plate"}
    # Every row is a cell stock, the word the seeding workbook uses.
    assert {row[1] for row in rows} == {"cell stock"}
    assert sorted(w["code"] for w in mapping["warnings"]) == ["derived_source_rack"]

    # The pick list was written beside the workbook by analyze itself, no
    # export needed: two files, two writers, two questions.
    picks = result["janus_autosave"]
    assert picks["status"] == "saved", picks
    assert Path(picks["output_path"]) == tmp_path / "260804_ref_MAME_picks.csv"
    assert len(_read_csv(picks["output_path"])[0]) == 5


def test_operator_plate_names_reach_the_exported_instrument_sheet(
    tmp_path: Path,
) -> None:
    """A plate name supplied for the export wins over the generated one."""
    _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})
    from sidecar_mame.handlers.export import handle_export_janus_mapping

    target = tmp_path / "worklist.csv"
    mapping = handle_export_janus_mapping(
        {
            "output": str(target),
            "liquid_class": _LIQUID_CLASS,
            "source_racks": {"NB01": "bench stock plate"},
            "dest_rack": "bench culture plate",
            "volume": 40.0,
        }
    )

    rows = _read_csv(mapping["output_path"])[1:]
    assert {row[3] for row in rows} == {"bench stock plate"}
    assert {row[5] for row in rows} == {"bench culture plate"}
    assert {row[7] for row in rows} == {"40.0"}
    assert not [r for r in rows if _LIQUID_CLASS in r]
    # Every name was supplied, so nothing is reported as derived.
    assert mapping["warnings"] == []


def test_a_rack_number_left_over_from_the_old_deck_is_dropped_not_written(
    tmp_path: Path,
) -> None:
    """The one way this migration could silently undo itself.

    A client that stored its deck while the columns carried numbers sends
    integers. Writing ``str(5)`` would put a bare "5" where the robot expects a
    labware name, and every gate would stay green because a written cell is a
    written cell. The stale value is dropped instead, which restores the
    generated name, and the export still produces a file rather than failing an
    operator for what an older build saved.
    """
    _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})
    from sidecar_mame.handlers.export import handle_export_janus_mapping

    target = tmp_path / "worklist.csv"
    mapping = handle_export_janus_mapping(
        {"output": str(target), "source_racks": {"NB01": 5}, "dest_rack": 8}
    )

    rows = _read_csv(mapping["output_path"])[1:]
    assert rows
    assert {row[3] for row in rows} == {"Stock plate1"}
    assert {row[5] for row in rows} == {"final culture plate"}
    assert not [r for r in rows if "5" in r or "8" in r]


def test_an_export_failure_does_not_cost_the_analysis(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The run is already complete by then; losing it to a file write is worse."""
    from kuma_core.mame.export import janus_mapping as janus_mod

    def _boom(*_args, **_kwargs):
        raise OSError("disk went away")

    monkeypatch.setattr(janus_mod, "export_mame_janus_csv", _boom)
    monkeypatch.setattr(
        "kuma_core.mame.export.export_mame_janus_csv", _boom, raising=False
    )

    result = _run(tmp_path, {"1_2": _G2A_NT, "2_1": _F3W_NT})

    autosave = result["janus_autosave"]
    assert autosave["status"] == "failed", autosave
    assert autosave["errors"][0]["code"] == "autosave_failed"
    assert "disk went away" in autosave["errors"][0]["message"]
    assert len(result["verdicts"]) == 2
    assert result["summary"]["total"] == 2


def test_autosave_path_derives_from_the_workbook_name() -> None:
    """The workbook name is the rule; the pick list only appends its own token."""
    from sidecar_mame.handlers.analyze import picks_autosave_path

    workbook = Path("/runs/260804_pTSN-PtIspS-idi_KanR_MAME_95verdicts.xlsx")

    assert picks_autosave_path(workbook) == Path(
        "/runs/260804_pTSN-PtIspS-idi_KanR_MAME_95verdicts_picks.csv"
    )
