from __future__ import annotations

import gzip
from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.ingest.amplicon_reference import (
    AmpliconReferenceError,
    check_coverage_reachable,
    resolve_amplicon_reference,
)

_F_TAIL = "cacaggaggttaaacc"
_R_TAIL = "tgcgttgcgctctag"
_COMPLEMENT = str.maketrans("ACGTacgt", "TGCAtgca")


def _reverse_complement(sequence: str) -> str:
    return sequence.translate(_COMPLEMENT)[::-1]


def _write_barcodes(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["target_f_1", "AATCCCACTAC" + _F_TAIL])
    sheet.append(["target_f_2", "TGAACTGAGCG" + _F_TAIL])
    sheet.append(["target_r_1", "CCCTATGACA" + _R_TAIL])
    sheet.append(["target_r_2", "TAATGGCAAG" + _R_TAIL])
    workbook.save(path)


def test_whole_plasmid_is_reduced_to_primer_bounded_amplicon(tmp_path: Path) -> None:
    coding = "ATG" + "GCT" * 19 + "TAA"
    amplicon = _F_TAIL + coding + _reverse_complement(_R_TAIL)
    plasmid = "G" * 80 + amplicon + "C" * 70
    reference = tmp_path / "plasmid.fa"
    reference.write_text(f">plasmid\n{plasmid}\n", encoding="utf-8")
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)

    resolution = resolve_amplicon_reference(reference, barcodes, tmp_path / "out")

    assert resolution.extracted is True
    assert resolution.span is not None
    assert resolution.span.start == 80
    assert resolution.span.end == 80 + len(amplicon)
    written = "".join(
        line.strip()
        for line in resolution.reference_fasta.read_text(encoding="utf-8").splitlines()
        if not line.startswith(">")
    )
    assert written == amplicon.upper()
    assert resolution.cds_start == len(_F_TAIL)
    assert resolution.cds_end == len(_F_TAIL) + len(coding)


def test_existing_amplicon_reference_is_left_unchanged(tmp_path: Path) -> None:
    reference = tmp_path / "amplicon.fa"
    reference.write_text(">amplicon\nATGGCTGCTTAA\n", encoding="utf-8")
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)

    resolution = resolve_amplicon_reference(reference, barcodes, tmp_path / "out")

    assert resolution.extracted is False
    assert resolution.reference_fasta == reference
    assert resolution.cds_start == 0
    assert resolution.cds_end == 0


# ---------------------------------------------------------------------------
# Undetected ORF (silent (0, 0) CDS)
# ---------------------------------------------------------------------------


def test_amplicon_without_an_orf_says_so_instead_of_reporting_zero_bounds(
    tmp_path: Path,
) -> None:
    """An extracted amplicon holding no forward ORF used to report ``(0, 0)``.

    Nothing said the search had failed, and ``(0, 0)`` is indistinguishable
    from the bounds the skipped-extraction branches report, so the one case
    where the CDS is genuinely unknown looked exactly like the ordinary ones.
    """
    # No ATG anywhere: not in the tails, not in the filler, not across a joint.
    payload = "GCT" * 20
    amplicon = _F_TAIL.upper() + payload + _reverse_complement(_R_TAIL).upper()
    assert "ATG" not in amplicon
    plasmid = "GGG" * 25 + amplicon + "CCC" * 25
    reference = tmp_path / "plasmid.fa"
    reference.write_text(f">plasmid\n{plasmid}\n", encoding="utf-8")
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)

    resolution = resolve_amplicon_reference(reference, barcodes, tmp_path / "out")

    assert resolution.extracted is True
    assert resolution.coding_bounds_found is False
    assert resolution.cds_start == 0
    assert resolution.cds_end == 0
    assert "no forward reading frame" in resolution.note


def test_an_extracted_amplicon_with_an_orf_reports_bounds_as_found(
    tmp_path: Path,
) -> None:
    coding = "ATG" + "GCT" * 19 + "TAA"
    amplicon = _F_TAIL + coding + _reverse_complement(_R_TAIL)
    plasmid = "G" * 80 + amplicon + "C" * 70
    reference = tmp_path / "plasmid.fa"
    reference.write_text(f">plasmid\n{plasmid}\n", encoding="utf-8")
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)

    resolution = resolve_amplicon_reference(reference, barcodes, tmp_path / "out")

    assert resolution.coding_bounds_found is True


def test_a_skipped_extraction_does_not_claim_coding_bounds(tmp_path: Path) -> None:
    reference = tmp_path / "amplicon.fa"
    reference.write_text(">amplicon\nATGGCTGCTTAA\n", encoding="utf-8")
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)

    resolution = resolve_amplicon_reference(reference, barcodes, tmp_path / "out")

    assert resolution.extracted is False
    assert resolution.coding_bounds_found is False


# ---------------------------------------------------------------------------
# Multi-record reference guard (silent concatenation)
# ---------------------------------------------------------------------------


def test_multi_record_reference_is_refused_instead_of_concatenated(
    tmp_path: Path,
) -> None:
    """Two records in one file used to be glued into one nonexistent sequence.

    The backbone and the target were joined base to base, and the joint is a
    junction no molecule has. Everything downstream then ran on that chimera
    without a single warning, so the refusal below is the whole point: the
    length assertion states what the old reader would have produced.
    """
    backbone = "GGGG" * 25
    target = "ATG" + "GCT" * 19 + "TAA"
    reference = tmp_path / "two_records.fa"
    reference.write_text(
        f">backbone\n{backbone}\n>target_gene\n{target}\n", encoding="utf-8"
    )
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)

    with pytest.raises(AmpliconReferenceError) as excinfo:
        resolve_amplicon_reference(reference, barcodes, tmp_path / "out")

    message = str(excinfo.value)
    # The count and the names both have to be there: the operator has to pick
    # one record, and cannot without knowing which ones the file offers.
    assert "2 sequence records" in message
    assert "backbone" in message
    assert "target_gene" in message
    # No output was written: the refusal happens before anything downstream can
    # read a chimera off disk.
    assert not (tmp_path / "out").exists()


def test_multi_record_reference_is_refused_by_the_coverage_precheck(
    tmp_path: Path,
) -> None:
    """The reachability pre-flight reads the same file and must refuse it too.

    It measures reads against the reference LENGTH, and a concatenated length
    makes the gate look unreachable for a reason the operator cannot act on.
    """
    reference = tmp_path / "two_records.fa"
    reference.write_text(">a\nACGT\n>b\nACGT\n", encoding="utf-8")
    run_dir = _write_run(tmp_path / "run", read_length=100)

    with pytest.raises(AmpliconReferenceError):
        check_coverage_reachable(reference, run_dir, _COVERAGE_FRACTION)


def test_headerless_and_single_record_references_still_load(tmp_path: Path) -> None:
    """The two shapes that were always valid stay valid.

    A bare sequence file (no header at all) is the pre-existing behaviour this
    guard must not break, and one record is the ordinary case.
    """
    headerless = tmp_path / "bare.txt"
    headerless.write_text("ACGTACGTAC\nGTACGTACGT\n", encoding="utf-8")
    single = tmp_path / "one.fa"
    single.write_text(">only\nACGTACGTACGTACGTACGT\n", encoding="utf-8")
    run_dir = _write_run(tmp_path / "run", read_length=100)

    assert check_coverage_reachable(
        headerless, run_dir, _COVERAGE_FRACTION
    ).reference_length == 20
    assert check_coverage_reachable(
        single, run_dir, _COVERAGE_FRACTION
    ).reference_length == 20


# ---------------------------------------------------------------------------
# Coverage-reachability guard (silent whole-plasmid fallback)
# ---------------------------------------------------------------------------

_COVERAGE_FRACTION = 0.98


def _write_platemap_barcodes(path: Path) -> None:
    """A workbook whose rows do NOT follow the ``*_f_<n>`` / ``*_r_<n>`` rule."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["A1", "variant_001"])
    sheet.append(["A2", "variant_002"])
    workbook.save(path)


def _write_run(run_dir: Path, read_length: int, n_reads: int = 4) -> Path:
    barcode_dir = run_dir / "fastq_pass" / "barcode01"
    barcode_dir.mkdir(parents=True, exist_ok=True)
    fastq = barcode_dir / "reads.fastq.gz"
    with gzip.open(fastq, "wt") as handle:
        for index in range(n_reads):
            sequence = ("ACGT" * read_length)[:read_length]
            handle.write(f"@read{index}\n{sequence}\n+\n{'I' * read_length}\n")
    return run_dir


def test_whole_plasmid_reference_cannot_reach_the_coverage_gate(tmp_path: Path) -> None:
    reference = tmp_path / "plasmid.fa"
    reference.write_text(">plasmid\n" + "ACGT" * 225 + "\n", encoding="utf-8")
    run_dir = _write_run(tmp_path / "run", read_length=200)

    reachability = check_coverage_reachable(reference, run_dir, _COVERAGE_FRACTION)

    assert reachability.reference_length == 900
    assert reachability.required_span == 882
    assert reachability.sampled_reads == 4
    assert reachability.longest_read_length == 200
    assert reachability.reachable is False


def test_short_amplicon_reference_still_reaches_the_coverage_gate(tmp_path: Path) -> None:
    reference = tmp_path / "amplicon.fa"
    reference.write_text(">amplicon\n" + "ACGT" * 50 + "\n", encoding="utf-8")
    run_dir = _write_run(tmp_path / "run", read_length=200)

    reachability = check_coverage_reachable(reference, run_dir, _COVERAGE_FRACTION)

    assert reachability.reference_length == 200
    assert reachability.required_span == 196
    assert reachability.reachable is True


def test_reachability_does_not_block_when_no_reads_are_found(tmp_path: Path) -> None:
    reference = tmp_path / "plasmid.fa"
    reference.write_text(">plasmid\n" + "ACGT" * 225 + "\n", encoding="utf-8")
    empty_run = tmp_path / "empty_run"
    (empty_run / "fastq_pass").mkdir(parents=True)

    reachability = check_coverage_reachable(reference, empty_run, _COVERAGE_FRACTION)

    assert reachability.sampled_reads == 0
    assert reachability.reachable is True


def test_extracted_amplicon_reference_reaches_the_gate(tmp_path: Path) -> None:
    """A proper barcodes workbook extracts the amplicon, so the run proceeds."""
    coding = "ATG" + "GCT" * 55 + "TAA"
    amplicon = _F_TAIL + coding + _reverse_complement(_R_TAIL)
    plasmid = "G" * 400 + amplicon + "C" * 400
    reference = tmp_path / "plasmid.fa"
    reference.write_text(f">plasmid\n{plasmid}\n", encoding="utf-8")
    barcodes = tmp_path / "barcodes.xlsx"
    _write_barcodes(barcodes)
    run_dir = _write_run(tmp_path / "run", read_length=len(amplicon))

    resolution = resolve_amplicon_reference(reference, barcodes, tmp_path / "out")
    assert resolution.extracted is True

    whole = check_coverage_reachable(reference, run_dir, _COVERAGE_FRACTION)
    extracted = check_coverage_reachable(
        resolution.reference_fasta, run_dir, _COVERAGE_FRACTION
    )

    assert whole.reachable is False
    assert extracted.reachable is True


def test_analyze_refuses_a_whole_plasmid_reference_with_platemap_barcodes(
    tmp_path: Path,
) -> None:
    from sidecar_mame.handlers.analyze import handle_analyze

    reference = tmp_path / "plasmid.fa"
    reference.write_text(">plasmid\n" + "ACGT" * 1625 + "\n", encoding="utf-8")
    barcodes = tmp_path / "platemap.xlsx"
    _write_platemap_barcodes(barcodes)
    # A readable variant list, because this test is about the reference and the
    # barcodes. An empty workbook stands in for nothing: the plate-capacity gate
    # reads the expected list before the raw-run block and refuses a file it
    # cannot place, so the fixture would be answered with a sentence about the
    # wrong input.
    expected = tmp_path / "expected.xlsx"
    variants = openpyxl.Workbook()
    variant_sheet = variants.active
    assert variant_sheet is not None
    variant_sheet.append(["variant"])
    variant_sheet.append(["A2G"])
    variants.save(expected)
    run_dir = _write_run(tmp_path / "run", read_length=1700)

    with pytest.raises(ValueError) as excinfo:
        handle_analyze(
            {
                "input_dir": str(run_dir),
                "reference": str(reference),
                "expected": str(expected),
                "output": str(tmp_path / "result.xlsx"),
                "custom_barcodes_xlsx": str(barcodes),
            }
        )

    message = str(excinfo.value)
    assert "shared primer tails could not be derived" in message
    assert "_f_<n>" in message
    assert "6500 bp" in message
    assert "0.98" in message
    assert "6370 bp" in message
    assert "1700 bp" in message
