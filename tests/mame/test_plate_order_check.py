"""Does an exported workbook describe one plate, or two? (io/plate_order_check.py)"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from kuma_core.mame.io.plate_order_check import check_plate_order

REAL_EXPORT = Path(
    "/mnt/d/_workspace/020.admin/projects/070.KUMA_elements/260730 MAME test"
    "/260722_Ep_R2-1_platemap.xlsx"
)


#: The ten columns a KURO export writes, in order. `read_expected_mutations_with_rows`
#: checks them exactly, so a fixture that abbreviates them is not an export and is
#: refused before any comparison runs.
_STRICT_HEADER = [
    "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
    "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
]


def _expected_row(mutant_id, status="DESIGNED"):
    """One `expected_mutations` row. WT labels carry the control shape KURO writes."""
    if not mutant_id[1:-1].isdigit():
        return [mutant_id, 0, "-", "-", "-", "-", "G0", "-", "wt", status]
    return [
        mutant_id, int(mutant_id[1:-1]), mutant_id[0], mutant_id[-1],
        "GTG", "TTT", "", "", "single", status,
    ]


def _write_expected_sheet(workbook, rows):
    strict = workbook.create_sheet("expected_mutations")
    strict.append(_STRICT_HEADER)
    for row in rows:
        strict.append(row if isinstance(row, list) else _expected_row(row))
    return strict


def _workbook(path, plate_rows, expected_ids, plate_title="Fwd List"):
    wb = Workbook()
    plate = wb.worksheets[0]
    plate.title = plate_title
    plate.append(["Well", "Primer Name", "Mutation"])
    for well, mutation in plate_rows:
        plate.append([well, f"{mutation}_F", mutation])
    _write_expected_sheet(wb, list(expected_ids))
    wb.save(path)
    return path


class TestAgreement:
    def test_matching_order_is_reported_clean(self, tmp_path):
        path = _workbook(
            tmp_path / "ok.xlsx",
            [("A1", "S11I"), ("B1", "S22T"), ("C1", "N28S")],
            ["S11I", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.ok is True

    def test_the_well_column_decides_not_the_row_order(self, tmp_path):
        """플레이트 시트의 행이 섞여 있어도 Well 열이 좌표다."""
        path = _workbook(
            tmp_path / "shuffled.xlsx",
            [("C1", "N28S"), ("A1", "S11I"), ("B1", "S22T")],
            ["S11I", "S22T", "N28S"],
        )

        assert check_plate_order(path).ok is True


class TestDisagreement:
    def test_a_reordered_expected_sheet_is_reported_with_the_well(self, tmp_path):
        path = _workbook(
            tmp_path / "reordered.xlsx",
            [("A1", "S11I"), ("B1", "S22T"), ("C1", "N28S")],
            ["N28S", "S11I", "S22T"],
        )

        report = check_plate_order(path)

        assert report.mismatched is True
        assert report.examples[0] == ("A1", "S11I", "N28S")

    def test_a_well_missing_from_the_expected_sheet_is_named(self, tmp_path):
        """V263I 사례. 뒤쪽 well 이 전부 한 칸 밀린다."""
        path = _workbook(
            tmp_path / "missing.xlsx",
            [("A1", "R262N"), ("B1", "V263I"), ("C1", "I277V")],
            ["R262N", "I277V"],
        )

        report = check_plate_order(path)

        assert report.missing_from_expected == ["V263I"]
        assert report.ok is False
        assert ("B1", "V263I", "I277V") in report.examples


class TestNotComparable:
    def test_a_file_without_the_expected_sheet_is_not_called_consistent(self, tmp_path):
        wb = Workbook()
        wb.worksheets[0].title = "Fwd List"
        wb.worksheets[0].append(["Well", "Mutation"])
        wb.worksheets[0].append(["A1", "S11I"])
        path = tmp_path / "plate-only.xlsx"
        wb.save(path)

        report = check_plate_order(path)

        # 비교 불가와 일치는 다르다. 침묵을 합격으로 읽지 못하게 한다.
        assert report.comparable is False

    def test_a_missing_file_is_not_comparable(self, tmp_path):
        assert check_plate_order(tmp_path / "nope.xlsx").comparable is False


class TestGridSheet:
    def test_a_plate_grid_can_supply_the_order(self, tmp_path):
        """그리드의 행 라벨·열 번호가 곧 well 이다.

        예전 픽스처는 A1, A2, B1 세 칸을 채우고 `ok is True` 를 기대했다.
        그건 틀린 기대였다. 그리드는 K53I 가 A2 에 있다고 말하는데 expected
        시트는 그것을 3번째 행, 즉 C1 에 둔다. 옛 판독기가 채워진 칸만 모아
        조밀한 목록으로 눌러 담아 그 어긋남을 지웠다. 여기서는 세 칸을 한
        열에 두어 두 시트가 실제로 같은 말을 하게 한다.
        """
        wb = Workbook()
        grid = wb.worksheets[0]
        grid.title = "Fwd Plate"
        grid.append(["", "1", "2"])
        grid.append(["A", "S11I_F", ""])
        grid.append(["B", "S22T_F", ""])
        grid.append(["C", "K53I_F", ""])
        _write_expected_sheet(wb, ["S11I", "S22T", "K53I"])
        path = tmp_path / "grid.xlsx"
        wb.save(path)

        report = check_plate_order(path)

        assert report.plate_sheet == "Fwd Plate"
        assert report.ok is True

    def test_a_hole_in_the_grid_does_not_renumber_the_wells(self, tmp_path):
        """빈 칸이 뒤 칸을 끌어올리지 않는다.

        옛 판독기는 채워진 칸만 모아 순번을 다시 매겼으므로 A2 의 K53I 가
        2번 자리로 올라가 B1 의 것과 비교됐다. 보고되는 well 라벨도 그
        순번에서 나와 실제 well 과 달랐다.
        """
        wb = Workbook()
        grid = wb.worksheets[0]
        grid.title = "Fwd Plate"
        grid.append(["", "1", "2"])
        grid.append(["A", "S11I_F", "K53I_F"])
        grid.append(["B", "", ""])
        _write_expected_sheet(wb, ["S11I", "S22T", "K53I"])
        path = tmp_path / "hole.xlsx"
        wb.save(path)

        report = check_plate_order(path)

        assert report.mismatched is True
        # expected 시트는 S22T 를 B1, K53I 를 C1 에 두는데 그리드는 B1 이
        # 비었고 K53I 가 A2 다. 라벨이 실제 well 이어야 한다.
        assert ("B1", "", "S22T") in report.examples
        assert ("C1", "", "K53I") in report.examples
        assert ("A2", "K53I", "") in report.examples


class TestWildTypeRow:
    """WT 행은 웰을 차지하고 프라이머가 없다. 그것은 불일치가 아니다.

    예전에는 `_expected_order` 가 WT 행을 그냥 세어 넣어 두 목록의 길이가
    1 어긋났고, 자기 WT 행을 가진 KURO export 는 전부 `mismatched=True` 에
    `absent_from_plate=['WT']` 로 보고됐다.
    """

    def test_a_trailing_wt_row_does_not_make_the_file_mismatch(self, tmp_path):
        path = _workbook(
            tmp_path / "wt-last.xlsx",
            [("A1", "S11I"), ("B1", "S22T"), ("C1", "N28S")],
            ["S11I", "S22T", "N28S", "WT"],
        )

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.ok is True
        assert report.absent_from_plate == []

    def test_a_control_status_wt_row_is_read_not_dropped(self, tmp_path):
        """KURO 가 실제로 쓰는 모양. status 가 `control` 이다."""
        wb = Workbook()
        plate = wb.worksheets[0]
        plate.title = "Fwd List"
        plate.append(["Well", "Primer Name", "Mutation"])
        for well, mutation in (("A1", "S11I"), ("B1", "S22T")):
            plate.append([well, f"{mutation}_F", mutation])
        _write_expected_sheet(
            wb, ["S11I", "S22T", _expected_row("WT", status="control")]
        )
        path = tmp_path / "wt-control.xlsx"
        wb.save(path)

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.ok is True

    def test_a_wt_row_in_the_middle_moves_the_wells_after_it(self, tmp_path):
        """WT 가 2번째면 그 뒤 mutant 는 한 칸 내려간 well 에서 비교된다."""
        path = _workbook(
            tmp_path / "wt-middle.xlsx",
            [("A1", "S11I"), ("C1", "S22T"), ("D1", "N28S")],
            ["S11I", "WT", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.ok is True

    def test_the_middle_wt_well_is_left_out_of_the_comparison(self, tmp_path):
        """B1 은 대조군 자리다. 플레이트에 없다고 어긋난 것이 아니다."""
        path = _workbook(
            tmp_path / "wt-middle.xlsx",
            [("A1", "S11I"), ("C1", "S22T"), ("D1", "N28S")],
            ["S11I", "WT", "S22T", "N28S"],
        )

        assert [w for w, _, _ in check_plate_order(path).examples] == []

    def test_the_old_ordinal_reading_would_have_reported_a_shift(self, tmp_path):
        """WT 를 세지 않던 옛 규칙이라면 S22T 가 B1 에서 어긋났을 자리다."""
        path = _workbook(
            tmp_path / "wt-middle.xlsx",
            [("A1", "S11I"), ("C1", "S22T"), ("D1", "N28S")],
            ["S11I", "WT", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.missing_from_expected == []
        assert report.absent_from_plate == []


class TestStatusFilteredRows:
    def test_a_status_dropped_row_lowers_the_file_to_not_comparable(self, tmp_path):
        """리더가 거부하는 파일은 여기서도 판정하지 않는다.

        옛 코드는 status 를 아예 보지 않고 그 행을 세어 넣어, 리더가 거절할
        파일에 대해 well 배치를 자신 있게 보고했다. 이제는 리더에게 물어보고,
        리더가 거절하면 비교 불가로 낮춘다. 예외를 다시 던지지는 않는다.
        """
        wb = Workbook()
        plate = wb.worksheets[0]
        plate.title = "Fwd List"
        plate.append(["Well", "Primer Name", "Mutation"])
        for well, mutation in (("A1", "S11I"), ("B1", "S22T")):
            plate.append([well, f"{mutation}_F", mutation])
        _write_expected_sheet(
            wb,
            [
                "S11I",
                _expected_row("Z99Z", status="FAILED"),
                "S22T",
            ],
        )
        path = tmp_path / "failed-row.xlsx"
        wb.save(path)

        report = check_plate_order(path)

        assert report.comparable is False
        assert report.mismatched is False

    def test_an_unreadable_expected_sheet_does_not_raise(self, tmp_path):
        """헤더가 KURO 모양이 아니면 리더가 ValueError 를 던진다. 새면 안 된다."""
        wb = Workbook()
        plate = wb.worksheets[0]
        plate.title = "Fwd List"
        plate.append(["Well", "Mutation"])
        plate.append(["A1", "S11I"])
        strict = wb.create_sheet("expected_mutations")
        strict.append(["mutant_id", "status"])
        strict.append(["S11I", "DESIGNED"])
        path = tmp_path / "loose-header.xlsx"
        wb.save(path)

        assert check_plate_order(path).comparable is False


@pytest.mark.skipif(not REAL_EXPORT.exists(), reason="260730 test data not on this machine")
def test_the_260722_export_is_reported_as_mismatched():
    report = check_plate_order(REAL_EXPORT)

    assert report.comparable is True
    assert report.mismatched is True
    assert report.missing_from_expected == ["V263I"]
    # 첫 well 부터 어긋난다. 프라이머 목록은 S11I 로 시작하고 expected 시트는 V233I 다.
    assert report.examples[0] == ("A1", "S11I", "V233I")
    assert report.plate_sheet == "Fwd List"
