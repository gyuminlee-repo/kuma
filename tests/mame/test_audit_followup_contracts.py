"""Three silent-wrong-answer paths the audit found, and the inputs that stay legal.

Each of the three fixes replaces a value the code invented when it could not read
one: position 0 for an unreadable cell, last-write-wins for a well declared
twice, and ``True`` for the string ``"false"``. An invented value is worse than a
refusal because the run continues and reports a number, so every test here comes
in a pair: one input that must now be refused, and one that must still be read.
The second half is the load-bearing one. A rule that refuses everything also
refuses the shipped templates, which is how the previous attempt at the position
fix broke them.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from kuma_core.mame.io.kuro_reader import (
    expected_to_labels,
    read_expected_mutations,
)
from kuma_core.mame.io.plate_order_check import check_plate_order

#: The ten columns a KURO export writes, in order.
_HEADER = [
    "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
    "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
]

#: The workbooks MAME ships. Both end with the wild-type control row
#: ``('WT', 0, '-', '-', '-', '-', 'G0', '-', 'wt', 'control')``, which is why
#: `position <= 0` cannot be the test for an unreadable position.
_REPO_ROOT = Path(__file__).resolve().parents[2]
SHIPPED_EXPECTED_WORKBOOKS = [
    _REPO_ROOT / "templates" / "03_mame_expected_mutations.xlsx",
    _REPO_ROOT / "src-tauri" / "samples" / "mame" / "03_mame_expected_mutations.xlsx",
]


def _expected_workbook(path: Path, rows: list[list[object]]) -> Path:
    workbook = Workbook()
    sheet = workbook.worksheets[0]
    sheet.title = "expected_mutations"
    sheet.append(_HEADER)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return path


def _designed(mutant_id: str, position: object) -> list[object]:
    return [
        mutant_id, position, mutant_id[0], mutant_id[-1],
        "GTG", "TTT", "G1", "PS_001", "single", "designed",
    ]


_WT_ROW: list[object] = ["WT", 0, "-", "-", "-", "-", "G0", "-", "wt", "control"]


class TestPositionCannotBeInvented:
    """`kuro_reader._position`: an unreadable cell is refused, not read as 0."""

    def test_an_unreadable_position_on_a_designed_row_is_refused(
        self, tmp_path: Path
    ) -> None:
        path = _expected_workbook(
            tmp_path / "unreadable.xlsx",
            [_designed("V5F", "n/a"), _WT_ROW],
        )

        with pytest.raises(ValueError) as raised:
            read_expected_mutations(path)

        message = str(raised.value)
        assert "position" in message
        # The operator has to find the cell, so the row number is part of the
        # refusal rather than a count of bad rows.
        assert "row 2" in message

    def test_a_readable_designed_row_is_still_read(self, tmp_path: Path) -> None:
        """The control half of the pair: the same shape, with a readable cell."""
        path = _expected_workbook(
            tmp_path / "readable.xlsx",
            [_designed("V5F", 5), _designed("K53N", 53), _WT_ROW],
        )

        result = read_expected_mutations(path)

        assert expected_to_labels(result) == ["V5F", "K53N", "-0-"]

    @pytest.mark.parametrize("cell", ["232", 232, 232.0, "232.0"])
    def test_a_whole_number_in_any_spelling_is_a_position(
        self, tmp_path: Path, cell: object
    ) -> None:
        """Excel hands the same number over as int, float, or text."""
        path = _expected_workbook(tmp_path / "spellings.xlsx", [_designed("Q232A", cell)])

        assert read_expected_mutations(path)[0].position == 232

    def test_a_blank_position_is_accepted(self, tmp_path: Path) -> None:
        """Blank is a column the exporter did not fill, not an unreadable cell."""
        path = _expected_workbook(tmp_path / "blank.xlsx", [_designed("V5F", None)])

        assert read_expected_mutations(path)[0].position == 0

    def test_the_control_row_keeps_position_zero(self, tmp_path: Path) -> None:
        """0 is what the control is. Refusing it would refuse every export."""
        path = _expected_workbook(tmp_path / "control.xlsx", [_designed("V5F", 5), _WT_ROW])

        result = read_expected_mutations(path)

        assert [m.mutant_id for m in result] == ["V5F", "WT"]
        assert result[1].position == 0

    def test_an_unreadable_position_on_the_control_row_is_tolerated(
        self, tmp_path: Path
    ) -> None:
        """The control row carries no residue, so its position is never compared.

        Refusing here would refuse a whole export over a cell nothing reads,
        which is the failure mode this rule was written to avoid.
        """
        path = _expected_workbook(
            tmp_path / "control_junk.xlsx",
            [
                _designed("V5F", 5),
                ["WT", "n/a", "-", "-", "-", "-", "G0", "-", "wt", "control"],
            ],
        )

        result = read_expected_mutations(path)

        assert [(m.mutant_id, m.position) for m in result] == [("V5F", 5), ("WT", 0)]

    @pytest.mark.parametrize(
        "path", SHIPPED_EXPECTED_WORKBOOKS, ids=lambda p: p.parent.name
    )
    def test_the_shipped_workbooks_are_still_read(self, path: Path) -> None:
        """The two files MAME ships, read as files rather than as a fixture.

        A synthetic fixture cannot say whether the rule refuses the templates,
        and the templates are what an operator starts from.
        """
        assert path.exists(), path

        result = read_expected_mutations(path)

        assert result, path
        control = [m for m in result if m.mutant_id.upper() == "WT"]
        assert len(control) == 1, path
        assert control[0].position == 0
        designed = [m for m in result if m.mutant_id.upper() != "WT"]
        assert designed, path
        assert all(m.position > 0 for m in designed), path


def _plate_workbook(
    path: Path,
    plate_rows: list[tuple[str, str]],
    expected_ids: list[str],
) -> Path:
    workbook = Workbook()
    plate = workbook.worksheets[0]
    plate.title = "Fwd List"
    plate.append(["Well", "Primer Name", "Mutation"])
    for well, mutation in plate_rows:
        plate.append([well, f"{mutation}_F", mutation])
    sheet = workbook.create_sheet("expected_mutations")
    sheet.append(_HEADER)
    for mutant_id in expected_ids:
        sheet.append(
            [
                mutant_id, int(mutant_id[1:-1]), mutant_id[0], mutant_id[-1],
                "GTG", "TTT", "", "", "single", "DESIGNED",
            ]
        )
    workbook.save(path)
    return path


def _grid_workbook(path: Path, grid_rows: list[list[object]], expected_ids: list[str]) -> Path:
    workbook = Workbook()
    plate = workbook.worksheets[0]
    plate.title = "Fwd Plate"
    plate.append(["", 1, 2, 3])
    for row in grid_rows:
        plate.append(row)
    sheet = workbook.create_sheet("expected_mutations")
    sheet.append(_HEADER)
    for mutant_id in expected_ids:
        sheet.append(
            [
                mutant_id, int(mutant_id[1:-1]), mutant_id[0], mutant_id[-1],
                "GTG", "TTT", "", "", "single", "DESIGNED",
            ]
        )
    workbook.save(path)
    return path


class TestAWellDeclaredTwiceIsReported:
    """`plate_order_check`: last-write-wins picked one of two plates in silence."""

    def test_a_well_with_two_different_occupants_is_reported(self, tmp_path: Path) -> None:
        path = _plate_workbook(
            tmp_path / "duplicate.xlsx",
            [("A1", "S11I"), ("A1", "S22T"), ("C1", "N28S")],
            ["S11I", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.duplicate_wells == ["A1"]
        assert report.ok is False

    def test_a_repeated_well_is_the_only_thing_wrong_and_still_reported(
        self, tmp_path: Path
    ) -> None:
        """A file whose every other check passes, failed by the repeat alone.

        The fixture above disagrees for a second reason (the plate carries a
        mutant the expected sheet places elsewhere), so it would be reported
        with or without this fix. Here the two sheets agree on every well that
        is named once, ``examples`` is empty, and nothing is missing or absent:
        the repeat is the whole finding. This is the shape the old code passed
        in silence, and it is what pins the repeat into ``mismatched`` rather
        than into the new field alone. ``mismatched`` is the flag the wire
        payload carries (``sidecar_mame.handlers.barcode_package``), so a repeat
        that reached only the field would still reach the operator as a clean
        file.
        """
        path = _plate_workbook(
            tmp_path / "duplicate_only.xlsx",
            [("A1", "S11I"), ("A1", "S22T"), ("B1", "S22T")],
            ["S11I", "S22T"],
        )

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.examples == []
        assert report.missing_from_expected == []
        assert report.absent_from_plate == []
        assert report.duplicate_wells == ["A1"]
        assert report.mismatched is True
        assert report.ok is False

    def test_a_plate_naming_each_well_once_is_clean(self, tmp_path: Path) -> None:
        """The control: the same sheet with the second A1 row moved to B1."""
        path = _plate_workbook(
            tmp_path / "no_duplicate.xlsx",
            [("A1", "S11I"), ("B1", "S22T"), ("C1", "N28S")],
            ["S11I", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.comparable is True
        assert report.duplicate_wells == []
        assert report.ok is True

    def test_a_well_repeated_with_the_same_occupant_is_not_a_disagreement(
        self, tmp_path: Path
    ) -> None:
        """Both readings name one plate, so there is nothing to choose between.

        This is the boundary of the rule, kept as a test because widening it
        would refuse a sheet that lists two primers for one well.
        """
        path = _plate_workbook(
            tmp_path / "same_twice.xlsx",
            [("A1", "S11I"), ("A1", "S11I"), ("B1", "S22T"), ("C1", "N28S")],
            ["S11I", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.duplicate_wells == []
        assert report.ok is True

    def test_a_grid_repeating_a_row_label_is_reported(self, tmp_path: Path) -> None:
        path = _grid_workbook(
            tmp_path / "grid_duplicate.xlsx",
            # Column-major down column 1, the order MAME assigns, plus a second
            # row labelled A: two occupants for well A1.
            [
                ["A", "S11I_F", None, None],
                ["B", "S22T_F", None, None],
                ["C", "N28S_F", None, None],
                ["A", "Q99R_F", None, None],
            ],
            ["S11I", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.duplicate_wells == ["A1"]
        assert report.ok is False

    def test_a_grid_naming_each_well_once_is_clean(self, tmp_path: Path) -> None:
        """The control for the grid reader."""
        path = _grid_workbook(
            tmp_path / "grid_ok.xlsx",
            [
                ["A", "S11I_F", None, None],
                ["B", "S22T_F", None, None],
                ["C", "N28S_F", None, None],
            ],
            ["S11I", "S22T", "N28S"],
        )

        report = check_plate_order(path)

        assert report.duplicate_wells == []
        assert report.ok is True


def _verdict_payload(**overrides: object) -> dict:
    payload: dict = {
        "native_barcode": "NB01",
        "custom_barcode": "CB01",
        "verdict": "PASS",
        "mutant_id": "V5F",
    }
    payload.update(overrides)
    return payload


class TestASerializedBooleanIsReadAsOne:
    """`analyze._as_bool`: ``bool("false")`` is ``True``, which inverted a gate."""

    def test_the_word_false_restores_as_false(self) -> None:
        from sidecar_mame.handlers.analyze import _deserialize_verdict

        record = _deserialize_verdict(
            _verdict_payload(consensus_n_fraction_evaluable="false")
        )

        assert record.translated.barcode.consensus_n_fraction_evaluable is False

    @pytest.mark.parametrize("stored", [True, "true", "TRUE", 1])
    def test_a_true_in_any_spelling_still_restores_as_true(self, stored: object) -> None:
        """The control: the fix must not flip the flag the other way."""
        from sidecar_mame.handlers.analyze import _deserialize_verdict

        record = _deserialize_verdict(
            _verdict_payload(consensus_n_fraction_evaluable=stored)
        )

        assert record.translated.barcode.consensus_n_fraction_evaluable is True

    def test_a_payload_without_the_key_keeps_the_field_default(self) -> None:
        """A payload written before the field existed restores as it does today."""
        from sidecar_mame.handlers.analyze import _deserialize_verdict

        record = _deserialize_verdict(_verdict_payload())

        assert record.translated.barcode.consensus_n_fraction_evaluable is True

    def test_a_real_boolean_false_is_unchanged(self) -> None:
        from sidecar_mame.handlers.analyze import _deserialize_verdict

        record = _deserialize_verdict(
            _verdict_payload(consensus_n_fraction_evaluable=False)
        )

        assert record.translated.barcode.consensus_n_fraction_evaluable is False

    def test_the_replicate_flags_read_the_word_false_too(self) -> None:
        """`failed` and `is_fallback` sit on the same restore path."""
        from sidecar_mame.handlers.analyze import _deserialize_replicate

        restored = _deserialize_replicate(
            {"mutant_id": "V5F", "failed": "false", "is_fallback": "false"}
        )

        assert restored.failed is False
        assert restored.is_fallback is False

    def test_the_replicate_flags_still_read_a_true(self) -> None:
        """The control for the replicate pair."""
        from sidecar_mame.handlers.analyze import _deserialize_replicate

        restored = _deserialize_replicate(
            {"mutant_id": "V5F", "failed": True, "is_fallback": "true"}
        )

        assert restored.failed is True
        assert restored.is_fallback is True
