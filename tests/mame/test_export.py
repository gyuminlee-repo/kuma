"""Export tests: well mapping, color rules, failed-well behavior, reference sheets, Janus."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

from kuma_core.mame.export import (
    WellMapper,
    export_mame_janus_csv,
    export_mame_janus_xlsx,
    seq_to_well,
    write_excel,
)
from kuma_core.mame.export.excel_writer import (
    _SHEET1_HEADER,
    _FINAL_HEADER,
    _matrix_header,
    _ngs_header,
    _run_native_barcodes,
    FAILED_FILL,
    VERDICT_FILL,
)
from kuma_core.mame.export.nb_label import nb_label
from kuma_core.mame.export.janus_mapping import _JANUS_HEADER
from kuma_core.mame.report.builder import build_run_report_data
from kuma_core.mame.report.html_renderer import render_html
from kuma_core.mame.models import (
    BarcodeRecord,
    ReplicateResult,
    TranslatedRecord,
    VerdictClass,
    VerdictRecord,
)


def test_seq_to_well_column_major() -> None:
    assert seq_to_well(1) == "A1"
    assert seq_to_well(2) == "B1"
    assert seq_to_well(8) == "H1"
    assert seq_to_well(9) == "A2"
    assert seq_to_well(96) == "H12"


def test_well_mapper_roundtrip() -> None:
    mapper = WellMapper()
    for seq in (1, 2, 8, 9, 17, 96):
        well = mapper.seq_to_well(seq)
        assert mapper.well_to_seq(well) == seq


def _make_verdict(
    nb: str,
    custom: str,
    verdict: VerdictClass,
    size_kb: float = 60.0,
    read_count: int | None = None,
    observed_aa: list[str] | None = None,
) -> VerdictRecord:
    barcode = BarcodeRecord(
        native_barcode=nb,
        custom_barcode=custom,
        consensus_seq="",
        file_size_kb=size_kb,
        source_path=Path("/tmp/mock.fasta"),
        read_count=read_count,
    )
    translated = TranslatedRecord(
        barcode=barcode,
        aa_sequence="",
        observed_nt_changes=[],
        observed_aa_changes=observed_aa or [],
    )
    return VerdictRecord(
        translated=translated,
        expected_mutations=[],
        verdict=verdict,
        verdict_notes="",
    )


def test_excel_sheet_colors(tmp_path: Path) -> None:
    verdicts = [
        _make_verdict("NB01", "1_1", VerdictClass.PASS),
        _make_verdict("NB01", "1_2", VerdictClass.AMBIGUOUS),
        _make_verdict("NB01", "1_3", VerdictClass.FRAMESHIFT),
        _make_verdict("NB01", "1_4", VerdictClass.MANY),
        _make_verdict("NB02", "1_1", VerdictClass.LOWDEPTH, size_kb=5.0),
    ]
    out = tmp_path / "out.xlsx"
    write_excel(
        verdict_records=verdicts,
        replicate_results=[],
        output_path=out,
    )
    wb = openpyxl.load_workbook(out)
    assert set(_SHEET1_HEADER).issubset({c.value for c in wb["NB01"][1]})

    # Look up each verdict row in NB01 / NB02 and confirm fill color.
    expected_fills = {
        "1_1": VERDICT_FILL[VerdictClass.PASS],
        "1_2": VERDICT_FILL[VerdictClass.AMBIGUOUS],
        "1_3": VERDICT_FILL[VerdictClass.FRAMESHIFT],
        "1_4": VERDICT_FILL[VerdictClass.MANY],
    }
    ws = wb["NB01"]
    header = [c.value for c in ws[1]]
    custom_col = header.index("custom_barcode") + 1
    for row in ws.iter_rows(min_row=2):
        label = row[custom_col - 1].value
        if label in expected_fills:
            fg = row[0].fill.fgColor.rgb or ""
            assert fg.endswith(expected_fills[label])


def test_excel_failed_well(tmp_path: Path) -> None:
    # One replicate result that failed -> appears with FAILED and red fill.
    plate_verdicts = {
        "NB01": _make_verdict("NB01", "1_1", VerdictClass.WRONG_AA),
    }
    rr = ReplicateResult(
        mutant_id="N63F",
        plate_verdicts=plate_verdicts,
        selected_plate=None,
        selection_reason="all fail",
        failed=True,
    )
    out = tmp_path / "failed.xlsx"
    write_excel(
        verdict_records=list(plate_verdicts.values()),
        replicate_results=[rr],
        output_path=out,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["Final"]

    found_failed = False
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        if "FAILED" in values:
            found_failed = True
            fg = row[1].fill.fgColor.rgb or ""
            assert fg.endswith(FAILED_FILL)
            break
    assert found_failed, "no FAILED row found in Final sheet"


# ---------------------------------------------------------------------------
# Reference-format sheets (G7 + G2)
# ---------------------------------------------------------------------------


def _make_replicate(
    mutant_id: str,
    nb: str,
    custom: str,
    verdict: VerdictClass = VerdictClass.PASS,
    size_kb: float = 80.0,
    read_count: int | None = None,
) -> ReplicateResult:
    vr = _make_verdict(nb, custom, verdict, size_kb=size_kb, read_count=read_count)
    return ReplicateResult(
        mutant_id=mutant_id,
        plate_verdicts={nb: vr},
        selected_plate=nb,
        selection_reason="pass",
        failed=False,
    )


def test_reference_sheets_present(tmp_path: Path) -> None:
    """write_excel must include 'NGS Results' and 'Final (matrix)' sheets."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "ref.xlsx"
    write_excel(
        verdict_records=[vr],
        replicate_results=[rr],
        output_path=out,
    )
    wb = openpyxl.load_workbook(out)
    assert "NGS Results" in wb.sheetnames, "NGS Results sheet missing"
    assert "Final (matrix)" in wb.sheetnames, "Final (matrix) sheet missing"
    # Legacy sheets preserved.
    assert "NB01" in wb.sheetnames
    assert "Final" in wb.sheetnames


def test_ngs_result_sheet_header(tmp_path: Path) -> None:
    """NGS Results row-1 matches the dynamic _ngs_header for the run's NBs."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "ngs.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["NGS Results"]
    actual_header = [c.value for c in ws[1]]
    nbs = _run_native_barcodes([vr], [rr])
    assert actual_header == _ngs_header(nbs)
    # selected_NB sits immediately after well, before custom_barcode.
    assert actual_header[2:5] == ["well", "selected_NB", "custom_barcode"]


def test_final_matrix_sheet_header(tmp_path: Path) -> None:
    """Final (matrix) row-1 matches the dynamic _matrix_header for the run's NBs."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "matrix.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Final (matrix)"]
    actual_header = [c.value for c in ws[1]]
    nbs = _run_native_barcodes([vr], [rr])
    assert actual_header == _matrix_header(nbs)
    # selected_NB sits immediately after well.
    assert actual_header[2:4] == ["well", "selected_NB"]


def test_final_matrix_pass_blank_and_selected_nb(tmp_path: Path) -> None:
    """(a)+(c): PASS plate => 'O', non-PASS => blank; selected_NB = nb_label(selected)."""
    vr = _make_verdict("NB02", "2_3", VerdictClass.PASS, size_kb=90.0)
    rr = ReplicateResult(
        mutant_id="K7R",
        plate_verdicts={
            "NB01": _make_verdict("NB01", "2_3", VerdictClass.AMBIGUOUS),
            "NB02": vr,
        },
        selected_plate="NB02",
        selection_reason="pass beats ambiguous",
        failed=False,
    )
    out = tmp_path / "matrix2.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Final (matrix)"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    sel_idx = header.index("selected_NB")
    nb01_idx = header.index(nb_label("NB01"))
    nb02_idx = header.index(nb_label("NB02"))
    assert row[sel_idx] == nb_label("NB02")  # selected_NB value, well's direct neighbor
    assert row[nb02_idx] == "O", "PASS plate should be 'O'"
    assert row[nb01_idx] in ("", None), "AMBIGUOUS plate should be blank"


def test_final_matrix_bold_only_pass_selection(tmp_path: Path) -> None:
    """(b): bold 'O' marks the single PASS final selection; at most one per mutant."""
    vr_sel = _make_verdict("NB02", "2_3", VerdictClass.PASS)
    rr = ReplicateResult(
        mutant_id="K7R",
        plate_verdicts={
            "NB01": _make_verdict("NB01", "2_3", VerdictClass.PASS),
            "NB02": vr_sel,
        },
        selected_plate="NB02",
        selection_reason="best pass",
        failed=False,
    )
    out = tmp_path / "matrix_bold.xlsx"
    write_excel(verdict_records=[vr_sel], replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Final (matrix)"]
    header = [c.value for c in ws[1]]
    nb01_col = header.index(nb_label("NB01")) + 1
    nb02_col = header.index(nb_label("NB02")) + 1
    # Both PASS plates carry "O"; only the selected one is bold.
    assert ws.cell(row=2, column=nb01_col).value == "O"
    assert ws.cell(row=2, column=nb02_col).value == "O"
    bold_count = sum(
        1
        for col in (nb01_col, nb02_col)
        if ws.cell(row=2, column=col).font.bold
    )
    assert bold_count == 1, "exactly one bold 'O' per mutant"
    assert ws.cell(row=2, column=nb02_col).font.bold is True
    assert ws.cell(row=2, column=nb01_col).font.bold in (False, None)


def test_ngs_reads_uses_read_count_not_filesize(tmp_path: Path) -> None:
    """(d): NGS reads carries read_count verbatim; None => blank; differs from file_size_kb."""
    vr_with = _make_verdict(
        "NB01", "1_1", VerdictClass.PASS, size_kb=123.0, read_count=4567
    )
    rr_with = ReplicateResult(
        mutant_id="V5F",
        plate_verdicts={"NB01": vr_with},
        selected_plate="NB01",
        selection_reason="pass",
        failed=False,
    )
    vr_none = _make_verdict(
        "NB01", "2_1", VerdictClass.PASS, size_kb=99.0, read_count=None
    )
    rr_none = ReplicateResult(
        mutant_id="K7R",
        plate_verdicts={"NB01": vr_none},
        selected_plate="NB01",
        selection_reason="pass",
        failed=False,
    )
    out = tmp_path / "ngs_reads.xlsx"
    write_excel(
        verdict_records=[vr_with, vr_none],
        replicate_results=[rr_with, rr_none],
        output_path=out,
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["NGS Results"]
    header = [c.value for c in ws[1]]
    reads_idx = header.index(f"{nb_label('NB01')}_reads")
    # Row 2 = rr_with: read_count value, never the file_size_kb proxy.
    assert ws[2][reads_idx].value == 4567
    assert ws[2][reads_idx].value != 123.0
    # Row 3 = rr_none: blank (read_count None, no file_size_kb fallback).
    assert ws[3][reads_idx].value in ("", None)


def test_dynamic_nb_columns_variable_names(tmp_path: Path) -> None:
    """(e): 4 sort_barcode NBs yield natural-ordered NB02/05/07/11 columns dynamically."""
    names = ["sort_barcode02", "sort_barcode05", "sort_barcode07", "sort_barcode11"]
    verdicts = [_make_verdict(nb, "1_1", VerdictClass.PASS) for nb in names]
    rr = ReplicateResult(
        mutant_id="V5F",
        plate_verdicts=dict(zip(names, verdicts)),
        selected_plate="sort_barcode05",
        selection_reason="pass",
        failed=False,
    )
    out = tmp_path / "dynamic.xlsx"
    write_excel(verdict_records=verdicts, replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    matrix = wb["Final (matrix)"]
    header = [c.value for c in matrix[1]]
    # Natural order by nb_order_key: 02, 05, 07, 11.
    assert header == _matrix_header(names)
    assert "NB05" in header and "NB07" in header and "NB11" in header
    row = [c.value for c in matrix[2]]
    sel_idx = header.index("selected_NB")
    assert row[sel_idx] == "NB05"
    # NGS sheet also has the dynamic per-NB triplets.
    ngs_header = [c.value for c in wb["NGS Results"][1]]
    assert "NB07_reads" in ngs_header and "NB11_quality" in ngs_header


def test_consensus_sheet_name_and_columns_preserved(tmp_path: Path) -> None:
    """(f): consensus native_barcode keeps its raw tab name and Sheet1 columns."""
    vr = _make_verdict("consensus", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "consensus", "1_1")
    out = tmp_path / "consensus.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    # nb_label("consensus") == "consensus" (no digits) -> tab name unchanged.
    assert "consensus" in wb.sheetnames
    header = [c.value for c in wb["consensus"][1]]
    assert header == _SHEET1_HEADER


def test_fallback_ambiguous_selection_no_matrix_o(tmp_path: Path) -> None:
    """(g) MEDIUM-3: AMBIGUOUS fallback selection => zero matrix 'O', selected_NB filled."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.AMBIGUOUS)
    rr = ReplicateResult(
        mutant_id="K7R",
        plate_verdicts={"NB01": vr},
        selected_plate="NB01",
        selection_reason="fallback to ambiguous",
        failed=False,
        is_fallback=True,
    )
    out = tmp_path / "fallback.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Final (matrix)"]
    header = [c.value for c in ws[1]]
    row = [c.value for c in ws[2]]
    sel_idx = header.index("selected_NB")
    nb01_idx = header.index(nb_label("NB01"))
    assert row[sel_idx] == nb_label("NB01"), "selected_NB must surface the fallback NB"
    # No "O" in any NB matrix cell (verdict is AMBIGUOUS, not PASS).
    nb_cells = [row[i] for i in range(sel_idx + 1, len(row))]
    assert "O" not in nb_cells, "AMBIGUOUS fallback must not produce a matrix 'O'"
    assert row[nb01_idx] in ("", None)


# ---------------------------------------------------------------------------
# Janus mapping export (K4)
# ---------------------------------------------------------------------------


def _make_janus_replicates() -> list[ReplicateResult]:
    """Two confirmed replicates with different sizes (for sort-order test)."""
    rr_high = ReplicateResult(
        mutant_id="V5F",
        plate_verdicts={"NB01": _make_verdict("NB01", "1_1", VerdictClass.PASS, size_kb=200.0)},
        selected_plate="NB01",
        selection_reason="pass",
        failed=False,
    )
    rr_low = ReplicateResult(
        mutant_id="K7R",
        plate_verdicts={"NB02": _make_verdict("NB02", "1_2", VerdictClass.PASS, size_kb=50.0)},
        selected_plate="NB02",
        selection_reason="pass",
        failed=False,
    )
    return [rr_low, rr_high]  # intentionally unsorted to test sort order


def test_janus_csv_header(tmp_path: Path) -> None:
    """CSV output must have the exact Janus header."""
    out = tmp_path / "janus.csv"
    export_mame_janus_csv(_make_janus_replicates(), out)
    with out.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        assert reader.fieldnames == _JANUS_HEADER


def test_janus_csv_sorted_desc(tmp_path: Path) -> None:
    """Rows must be sorted by priority_score DESC (V5F=200 before K7R=50)."""
    out = tmp_path / "janus_sorted.csv"
    export_mame_janus_csv(_make_janus_replicates(), out)
    with out.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert rows[0]["name"] == "V5F", "highest priority_score should come first"
    assert rows[1]["name"] == "K7R"


def test_janus_csv_plate_label(tmp_path: Path) -> None:
    """NB01->P1, NB02->P2 mapping must be applied in source_plate column."""
    out = tmp_path / "janus_plate.csv"
    export_mame_janus_csv(_make_janus_replicates(), out)
    with out.open(encoding="utf-8") as fh:
        rows = {r["name"]: r for r in csv.DictReader(fh)}
    assert rows["V5F"]["source_plate"] == "P1"
    assert rows["K7R"]["source_plate"] == "P2"


def test_janus_xlsx_sheet_name(tmp_path: Path) -> None:
    """XLSX output must have 'Janus Mapping' sheet with correct header."""
    out = tmp_path / "janus.xlsx"
    export_mame_janus_xlsx(_make_janus_replicates(), out)
    wb = openpyxl.load_workbook(out)
    assert "Janus Mapping" in wb.sheetnames
    ws = wb["Janus Mapping"]
    actual_header = [c.value for c in ws[1]]
    assert actual_header == _JANUS_HEADER


def test_janus_excludes_failed(tmp_path: Path) -> None:
    """Failed replicates must not appear in Janus output."""
    failed_rr = ReplicateResult(
        mutant_id="BAD",
        plate_verdicts={"NB01": _make_verdict("NB01", "1_3", VerdictClass.FRAMESHIFT)},
        selected_plate=None,
        selection_reason="all fail",
        failed=True,
    )
    ok_rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "janus_no_failed.csv"
    export_mame_janus_csv([failed_rr, ok_rr], out)
    with out.open(encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    names = [r["name"] for r in rows]
    assert "BAD" not in names, "failed replicate must be excluded"
    assert "V5F" in names

# ---------------------------------------------------------------------------
# Detected / recovery (재현율) — AC12-14
# ---------------------------------------------------------------------------


def _ngs_summary_pairs(ws) -> dict:
    """Collect key/value pairs from the NGS Results summary area (col A/B)."""
    pairs: dict = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] is not None and len(row) >= 2:
            pairs[row[0]] = row[1]
    return pairs


def test_ngs_recovered_column_present(tmp_path: Path) -> None:
    """AC14: NGS Results header gains a distinct `recovered` column (not NB0X_detected)."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    nbs = _run_native_barcodes([vr], [rr])
    out = tmp_path / "recovered.xlsx"
    write_excel(
        verdict_records=[vr],
        replicate_results=[rr],
        output_path=out,
        designed_mutant_ids=frozenset({"V5F"}),
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["NGS Results"]
    header = [c.value for c in ws[1]]
    assert "recovered" in header, "recovered column missing from NGS Results header"
    # The recovered column is distinct from the observed-AA NB0X_detected columns.
    assert "recovered" in _ngs_header(nbs)
    assert _ngs_header(nbs).count("recovered") == 1
    # Data row: V5F is PASS on its only plate -> recovered = Y.
    rec_idx = header.index("recovered")
    assert ws[2][rec_idx].value == "Y"


def test_ngs_recovery_summary(tmp_path: Path) -> None:
    """AC14: recovery summary reflects compute_recovery over designed_mutant_ids."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "recovery_summary.xlsx"
    write_excel(
        verdict_records=[vr],
        replicate_results=[rr],
        output_path=out,
        designed_mutant_ids=frozenset({"V5F"}),
    )
    ws = openpyxl.load_workbook(out)["NGS Results"]
    pairs = _ngs_summary_pairs(ws)
    assert pairs.get("recovered_mutants") == 1
    assert pairs.get("total_mutants") == 1
    assert pairs.get("Recovery (재현율)") == "100.0%"


def test_ngs_recovery_summary_na_when_unavailable(tmp_path: Path) -> None:
    """AC14: with no designed_mutant_ids the recovery summary renders n/a (never 0%)."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "recovery_na.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    ws = openpyxl.load_workbook(out)["NGS Results"]
    pairs = _ngs_summary_pairs(ws)
    assert pairs.get("Recovery (재현율)") == "n/a"
    assert "recovered_mutants" not in pairs


def test_report_detected_chip_and_plate_dt() -> None:
    """AC12-13: HTML report shows Detected/재현율 chip and per-plate 검출 D/T."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    data = build_run_report_data(
        [vr], [rr], designed_mutant_ids=frozenset({"V5F"})
    )
    html = render_html(data)
    assert "Detected / 재현율" in html
    assert "100% (1/1)" in html  # recovery chip value
    assert "검출 1/1" in html  # per-plate detected D/T


def test_report_detected_chip_na_when_unavailable() -> None:
    """AC12: chip renders n/a (not 0%) when recovery is unavailable."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    data = build_run_report_data([vr], [rr])
    html = render_html(data)
    assert "Detected / 재현율" in html
    # chip value cell renders n/a, never a fabricated 0%.
    assert ">n/a<" in html

# ---------------------------------------------------------------------------
# Selection / fallback marker columns + unified well sorting (AC-2.1/2.3/2.4)
# ---------------------------------------------------------------------------


def _fallback_replicate(
    mutant_id: str,
    nb: str,
    custom: str,
    verdict: VerdictClass = VerdictClass.AMBIGUOUS,
    fallback_reason: str = "no PASS; best AMBIGUOUS",
    verdict_notes: str = "",
) -> ReplicateResult:
    """A non-PASS fallback selection (is_fallback=True)."""
    vr = _make_verdict(nb, custom, verdict)
    vr.verdict_notes = verdict_notes
    return ReplicateResult(
        mutant_id=mutant_id,
        plate_verdicts={nb: vr},
        selected_plate=nb,
        selection_reason="fallback",
        failed=False,
        is_fallback=True,
        fallback_reason=fallback_reason,
    )


def _col(ws, name: str) -> int:
    """0-based column index of *name* in row 1."""
    return [c.value for c in ws[1]].index(name)


def test_sheet1_selection_marker_columns(tmp_path: Path) -> None:
    """AC-2.3: Sheet1 gains selected/is_fallback/fallback_reason; chosen well => 'Y'."""
    vr_sel = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    vr_other = _make_verdict("NB01", "1_2", VerdictClass.WRONG_AA)
    rr = _make_replicate("V5F", "NB01", "1_1")  # selects NB01 well 1_1
    out = tmp_path / "sheet1_sel.xlsx"
    write_excel(
        verdict_records=[vr_sel, vr_other],
        replicate_results=[rr],
        output_path=out,
    )
    ws = openpyxl.load_workbook(out)["NB01"]
    header = [c.value for c in ws[1]]
    for col in ("selected", "is_fallback", "fallback_reason"):
        assert col in header, f"{col} column missing from Sheet1"

    cb_col = _col(ws, "custom_barcode")
    sel_col = _col(ws, "selected")
    by_cb = {row[cb_col].value: row for row in ws.iter_rows(min_row=2)}
    assert by_cb["1_1"][sel_col].value == "Y"
    assert by_cb["1_2"][sel_col].value in ("", None)


def test_sheet1_fallback_marker(tmp_path: Path) -> None:
    """AC-2.4: Sheet1 fallback selection flags is_fallback='Y' + reason on chosen well."""
    vr_fb = _make_verdict("NB01", "1_1", VerdictClass.AMBIGUOUS)
    rr_fb = _fallback_replicate("V5F", "NB01", "1_1", fallback_reason="best AMBIGUOUS")
    out = tmp_path / "sheet1_fb.xlsx"
    write_excel(
        verdict_records=[vr_fb],
        replicate_results=[rr_fb],
        output_path=out,
    )
    ws = openpyxl.load_workbook(out)["NB01"]
    cb_col = _col(ws, "custom_barcode")
    sel_col = _col(ws, "selected")
    fb_col = _col(ws, "is_fallback")
    reason_col = _col(ws, "fallback_reason")
    row = next(r for r in ws.iter_rows(min_row=2) if r[cb_col].value == "1_1")
    assert row[sel_col].value == "Y"
    assert row[fb_col].value == "Y"
    assert row[reason_col].value == "best AMBIGUOUS"


def test_final_fallback_and_notes_columns(tmp_path: Path) -> None:
    """AC-2.4: Final gains is_fallback/fallback_reason/notes with values on selected row."""
    vr_fb = _make_verdict("NB01", "1_1", VerdictClass.AMBIGUOUS)
    rr_fb = _fallback_replicate(
        "V5F",
        "NB01",
        "1_1",
        fallback_reason="best AMBIGUOUS",
        verdict_notes="ambiguous extra mutations",
    )
    # Reuse the verdict record carrying the notes so the Final row reflects it.
    rr_fb.plate_verdicts["NB01"] = vr_fb
    vr_fb.verdict_notes = "ambiguous extra mutations"
    out = tmp_path / "final_fb.xlsx"
    write_excel(
        verdict_records=[vr_fb],
        replicate_results=[rr_fb],
        output_path=out,
    )
    ws = openpyxl.load_workbook(out)["Final"]
    header = [c.value for c in ws[1]]
    for col in ("is_fallback", "fallback_reason", "notes"):
        assert col in header, f"{col} column missing from Final"

    well_col = _col(ws, "well_id")
    fb_col = _col(ws, "is_fallback")
    reason_col = _col(ws, "fallback_reason")
    notes_col = _col(ws, "notes")
    row = next(r for r in ws.iter_rows(min_row=2) if r[well_col].value == "A1")
    assert row[fb_col].value == "Y"
    assert row[reason_col].value == "best AMBIGUOUS"
    assert row[notes_col].value == "ambiguous extra mutations"


def test_final_failed_row_blank_markers(tmp_path: Path) -> None:
    """AC-2.4: FAILED rows leave the new marker columns blank."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.WRONG_AA)
    rr = ReplicateResult(
        mutant_id="N63F",
        plate_verdicts={"NB01": vr},
        selected_plate=None,
        selection_reason="all fail",
        failed=True,
    )
    out = tmp_path / "final_failed.xlsx"
    write_excel(verdict_records=[vr], replicate_results=[rr], output_path=out)
    ws = openpyxl.load_workbook(out)["Final"]
    fb_col = _col(ws, "is_fallback")
    reason_col = _col(ws, "fallback_reason")
    notes_col = _col(ws, "notes")
    row = next(
        r for r in ws.iter_rows(min_row=2) if "FAILED" in [c.value for c in r]
    )
    assert row[fb_col].value in ("", None)
    assert row[reason_col].value in ("", None)
    assert row[notes_col].value in ("", None)


def _data_rows(ws, key_col: int):
    """Rows until the first blank-key row (stops before summary blocks)."""
    out = []
    for row in ws.iter_rows(min_row=2):
        if row[key_col].value in ("", None):
            break
        out.append(row)
    return out


def test_all_sheets_natural_well_sort(tmp_path: Path) -> None:
    """AC-2.1: NGS Results, Final (matrix), Sheet1, and Final use natural well order."""
    # Intentionally unsorted; wells 1_1, 1_2, 1_10 -> natural (not lexicographic).
    rr10 = _make_replicate("M10", "NB01", "1_10")
    rr2 = _make_replicate("M2", "NB01", "1_2")
    rr1 = _make_replicate("M1", "NB01", "1_1")
    rrs = [rr10, rr2, rr1]
    vrs = [rr.plate_verdicts["NB01"] for rr in rrs]
    out = tmp_path / "sort.xlsx"
    write_excel(verdict_records=vrs, replicate_results=rrs, output_path=out)
    wb = openpyxl.load_workbook(out)

    # (1) NGS Results: custom_barcode column natural-ordered, index re-assigned 1..N.
    ngs = wb["NGS Results"]
    cb_col = _col(ngs, "custom_barcode")
    idx_col = _col(ngs, "index")
    rows = _data_rows(ngs, idx_col)
    assert [r[cb_col].value for r in rows] == ["1_1", "1_2", "1_10"]
    assert [r[idx_col].value for r in rows] == [1, 2, 3]

    # (2) Final (matrix): well column natural-ordered (A1, A2, A10), index 1..N.
    matrix = wb["Final (matrix)"]
    well_col = _col(matrix, "well")
    midx_col = _col(matrix, "index")
    mrows = _data_rows(matrix, midx_col)
    assert [r[well_col].value for r in mrows] == ["A1", "A2", "A10"]
    assert [r[midx_col].value for r in mrows] == [1, 2, 3]

    # (3) per-NB Sheet1: custom_barcode natural-ordered.
    nb01 = wb["NB01"]
    s1_cb = _col(nb01, "custom_barcode")
    assert [r[s1_cb].value for r in nb01.iter_rows(min_row=2)] == [
        "1_1",
        "1_2",
        "1_10",
    ]

    # (4) Final (legacy): 1..96 grid placement keeps natural order (A10 holds 1_10).
    final = wb["Final"]
    f_well = _col(final, "well_id")
    f_cb = _col(final, "custom_barcode")
    placed = {
        r[f_well].value: r[f_cb].value
        for r in final.iter_rows(min_row=2)
        if r[f_cb].value not in ("", None)
    }
    assert placed["A1"] == "1_1"
    assert placed["A2"] == "1_2"
    assert placed["A10"] == "1_10"


def test_sheet1_header_includes_marker_columns() -> None:
    """Regression: _SHEET1_HEADER carries the new selection marker columns."""
    assert _SHEET1_HEADER[-3:] == ["selected", "is_fallback", "fallback_reason"]


def test_final_header_includes_marker_columns() -> None:
    """Regression: _FINAL_HEADER carries the new fallback/notes columns."""
    assert _FINAL_HEADER[-3:] == ["is_fallback", "fallback_reason", "notes"]
