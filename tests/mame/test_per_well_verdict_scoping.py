"""Per-well verdict scoping via sample_map (pipeline.py fix verification).

The discriminating test proves that scoping does the work:
- With sample_map: a well whose observation matches its assigned mutant -> PASS
- Without sample_map (full expected list): same well is missing the OTHER expected
  mutant -> WRONG_AA

Fixtures are self-contained (no dependency on create_fixtures.py or minimap2).
"""

from __future__ import annotations

import textwrap
from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.pipeline import _norm_well, run_analyze
from kuma_core.mame.models import VerdictClass


# ---------------------------------------------------------------------------
# Reference and per-well sequences
# ---------------------------------------------------------------------------
# Reference: ATG GGG TTT  ->  M G F  (9 bp, table 11)
_REFERENCE_NT = "ATGGGGTTT"

# Well A02 (custom_barcode "1_2"): G2A substitution only  -> M A F
# _custom_barcode_to_seq("1_2") = (2-1)*8+1 = 9
# seq_to_well(9) = "A2"  ->  _norm_well -> "A02"
_G2A_NT = "ATGGCGTTT"

# Well B01 (custom_barcode "2_1"): F3W substitution only  -> M G W
# _custom_barcode_to_seq("2_1") = (1-1)*8+2 = 2
# seq_to_well(2) = "B1"  ->  _norm_well -> "B01"
_F3W_NT = "ATGGGGTGG"

# Padding to clear the default 50 KB file-size threshold.
_PAD = "\n" * (52 * 1024 // 1)  # ~52 KB of newlines


def _write_fasta(path: Path, header: str, body: str, pad: bool = True) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = f">{header}\n{body}\n"
    if pad:
        text += _PAD
    path.write_text(text, encoding="utf-8")


def _make_kuro_xlsx(dest: Path) -> None:
    """Minimal KURO xlsx with two expected mutants: G2A (pos 2) and F3W (pos 3)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Fwd List"
    ws.append(["Well", "Primer Name", "Sequence", "Length", "Tm", "Tm_Overlap",
               "WT_Codon", "MT_Codon", "Mutation"])
    ws.append(["A1", "G2A_F", "ATGNNNNNNNN", 11, 60.0, 40.0, "GGG", "GCG", "G2A"])
    ws.append(["B1", "F3W_F", "ATGNNNNNNNN", 11, 60.0, 40.0, "TTT", "TGG", "F3W"])

    ws2 = wb.create_sheet("expected_mutations")
    ws2.append(["mutant_id", "position", "wt_aa", "mt_aa", "wt_codon", "mt_codon",
                "group_id", "primer_set_ref", "notation_type", "status"])
    # G2A: position 2, G->A
    ws2.append(["G2A", 2, "G", "A", "GGG", "GCG", "", "G2A", "substitution", "DESIGNED"])
    # F3W: position 3, F->W  (NOT in same column as G2A — transposition-proof)
    ws2.append(["F3W", 3, "F", "W", "TTT", "TGG", "", "F3W", "substitution", "DESIGNED"])

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def _make_sample_map_xlsx(dest: Path) -> None:
    """Sample map assigning well A2->G2A and well B1->F3W (non-zero-padded input).

    parse_sample_map normalises A2->A02, B1->B01; this exercises the normalisation.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    # Column A: sample name, Column B: well position
    ws.append(["G2A", "A2"])   # well A02 (non-zero-padded to exercise normalisation)
    ws.append(["F3W", "B1"])   # well B01 (non-zero-padded to exercise normalisation)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def _make_input_dir(tmp_path: Path) -> Path:
    """Create a barcode-mode ingest dir: NB01/1_2.fasta (G2A) and NB01/2_1.fasta (F3W)."""
    ingest_dir = tmp_path / "consensus"
    # Well A02: custom_barcode "1_2" -> r=1, f=2 -> seq=9 -> A2
    _write_fasta(ingest_dir / "NB01" / "1_2.fasta", header="1_2", body=_G2A_NT)
    # Well B01: custom_barcode "2_1" -> r=2, f=1 -> seq=2 -> B1
    _write_fasta(ingest_dir / "NB01" / "2_1.fasta", header="2_1", body=_F3W_NT)
    return ingest_dir


def _make_reference_fasta(tmp_path: Path) -> Path:
    ref = tmp_path / "reference.fasta"
    ref.write_text(f">ref\n{_REFERENCE_NT}\n", encoding="utf-8")
    return ref


# ---------------------------------------------------------------------------
# Primary discriminating test
# ---------------------------------------------------------------------------

def test_per_well_scoping_pass_vs_wrong_aa(tmp_path: pytest.FixtureDef) -> None:
    """Core discriminating test: PASS with sample_map, WRONG_AA without.

    - Well A02 (custom_barcode "1_2") observes G2A only.
    - Well B01 (custom_barcode "2_1") observes F3W only.
    - Full expected list = ["G2A", "F3W"]; each well misses one -> WRONG_AA.
    - Scoped list (via sample_map): each well scoped to its own mutant -> PASS.
    The wells are off-diagonal (A02 != B01) so a row/col transposition would
    scope the wrong mutant and flip PASS -> WRONG_AA, catching the bug.
    """
    ingest_dir = _make_input_dir(tmp_path)
    reference = _make_reference_fasta(tmp_path)
    kuro_xlsx = tmp_path / "kuro.xlsx"
    sample_map = tmp_path / "sample_map.xlsx"
    output_no_map = tmp_path / "out_no_map.xlsx"
    output_with_map = tmp_path / "out_with_map.xlsx"

    _make_kuro_xlsx(kuro_xlsx)
    _make_sample_map_xlsx(sample_map)

    from kuma_core.mame.ingest import IngestMode

    # ── Without sample_map: full list -> both wells WRONG_AA ──────────────
    verdicts_no_map, _ = run_analyze(
        input_dir=ingest_dir,
        reference_path=reference,
        expected_path=kuro_xlsx,
        output_path=output_no_map,
        cds_start=0,
        cds_end=9,
        min_file_size_kb=0.0,
        ingest_mode=IngestMode.BARCODE,
        sample_map_path=None,
    )
    by_custom = {v.translated.barcode.custom_barcode: v for v in verdicts_no_map}
    assert by_custom["1_2"].verdict is VerdictClass.WRONG_AA, (
        "Without sample_map, A02 (G2A-only) vs full list [G2A, F3W] must be WRONG_AA "
        f"(missing F3W); got {by_custom['1_2'].verdict}"
    )
    assert by_custom["2_1"].verdict is VerdictClass.WRONG_AA, (
        "Without sample_map, B01 (F3W-only) vs full list [G2A, F3W] must be WRONG_AA "
        f"(missing G2A); got {by_custom['2_1'].verdict}"
    )

    # ── With sample_map: scoped -> both wells PASS ─────────────────────────
    verdicts_with_map, _ = run_analyze(
        input_dir=ingest_dir,
        reference_path=reference,
        expected_path=kuro_xlsx,
        output_path=output_with_map,
        cds_start=0,
        cds_end=9,
        min_file_size_kb=0.0,
        ingest_mode=IngestMode.BARCODE,
        sample_map_path=sample_map,
    )
    by_custom_map = {v.translated.barcode.custom_barcode: v for v in verdicts_with_map}
    assert by_custom_map["1_2"].verdict is VerdictClass.PASS, (
        "With sample_map, A02 scoped to [G2A] must PASS; "
        f"got {by_custom_map['1_2'].verdict}: {by_custom_map['1_2'].verdict_notes}"
    )
    assert by_custom_map["2_1"].verdict is VerdictClass.PASS, (
        "With sample_map, B01 scoped to [F3W] must PASS; "
        f"got {by_custom_map['2_1'].verdict}: {by_custom_map['2_1'].verdict_notes}"
    )


# ---------------------------------------------------------------------------
# Unmapped well -> falls back to full list -> WRONG_AA
# ---------------------------------------------------------------------------

def test_unmapped_well_falls_back_to_full_list(tmp_path: pytest.FixtureDef) -> None:
    """A well not in the sample_map falls back to the full expected_labels list.

    Uses a sample_map that only covers A2->G2A (omits B1). Well B01 (F3W-only)
    is therefore unmapped: it receives the full list [G2A, F3W] -> WRONG_AA
    (missing G2A).  Well A02 (G2A-only) is mapped -> PASS.
    """
    ingest_dir = _make_input_dir(tmp_path)
    reference = _make_reference_fasta(tmp_path)
    kuro_xlsx = tmp_path / "kuro.xlsx"
    sample_map_partial = tmp_path / "sample_map_partial.xlsx"
    output = tmp_path / "out_partial.xlsx"

    _make_kuro_xlsx(kuro_xlsx)

    # Partial sample_map: only A2 entry
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["G2A", "A2"])
    wb.save(sample_map_partial)

    from kuma_core.mame.ingest import IngestMode

    verdicts, _ = run_analyze(
        input_dir=ingest_dir,
        reference_path=reference,
        expected_path=kuro_xlsx,
        output_path=output,
        cds_start=0,
        cds_end=9,
        min_file_size_kb=0.0,
        ingest_mode=IngestMode.BARCODE,
        sample_map_path=sample_map_partial,
    )
    by_custom = {v.translated.barcode.custom_barcode: v for v in verdicts}
    assert by_custom["1_2"].verdict is VerdictClass.PASS, (
        "Mapped A02 must PASS; "
        f"got {by_custom['1_2'].verdict}: {by_custom['1_2'].verdict_notes}"
    )
    assert by_custom["2_1"].verdict is VerdictClass.WRONG_AA, (
        "Unmapped B01 must fall back to full list -> WRONG_AA; "
        f"got {by_custom['2_1'].verdict}: {by_custom['2_1'].verdict_notes}"
    )


# ---------------------------------------------------------------------------
# _norm_well normalisation unit tests
# ---------------------------------------------------------------------------

def test_norm_well_zero_padding() -> None:
    """_norm_well normalises unpadded and already-padded labels identically."""
    assert _norm_well("A1") == "A01"
    assert _norm_well("A01") == "A01"
    assert _norm_well("H12") == "H12"
    assert _norm_well("b3") == "B03"   # lowercase normalised


def test_norm_well_lookup_matches_parse_sample_map_keys(tmp_path: Path) -> None:
    """Well IDs from parse_sample_map (zero-padded) match _norm_well output.

    parse_sample_map zero-pads entries (A2->A02, B1->B01).
    seq_to_well produces non-padded labels (A2, B1, H12).
    _norm_well bridges the two so lookup succeeds.
    """
    from kuma_core.mame.export.well_mapper import seq_to_well
    from kuma_core.mame.ingest.sort_barcode import parse_sample_map

    sample_map_xlsx = tmp_path / "sm.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["G2A", "A2"])   # non-padded input to parse_sample_map
    ws.append(["F3W", "B1"])
    wb.save(sample_map_xlsx)

    well_to_sample = parse_sample_map(sample_map_xlsx)
    # parse_sample_map produces zero-padded keys
    assert "A02" in well_to_sample
    assert "B01" in well_to_sample

    # seq_to_well produces non-padded labels; _norm_well must bridge them
    # custom_barcode "1_2": seq=9 -> "A2" -> _norm_well -> "A02"
    # custom_barcode "2_1": seq=2 -> "B1" -> _norm_well -> "B01"
    assert _norm_well(seq_to_well(9)) == "A02"   # 1_2 maps to A02
    assert _norm_well(seq_to_well(2)) == "B01"   # 2_1 maps to B01

    # Confirm these normalised keys exist in the sample_map
    assert well_to_sample[_norm_well(seq_to_well(9))] == "G2A"
    assert well_to_sample[_norm_well(seq_to_well(2))] == "F3W"
