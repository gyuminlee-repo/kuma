"""Tests for build_evolvepro_input_from_reports (reports mode).

Reports mode assembles an EVOLVEpro input xlsx from two raw Agilent FID1B
standard reports plus a plate layout, with no rank file:
  - round-1 report (well-named samples) + layout -> one relative replicate
    per mutant (fallback).
  - re-measure report (variant-labeled samples) -> n relative replicates per
    variant (authoritative).
Authoritative mean replaces fallback where both define a variant.

Synthetic fixtures are written in-process via openpyxl so the tests are
portable (no machine-specific data paths).
"""

from __future__ import annotations

import pytest

from kuma_core.mame.activity.build_evolvepro_input import (
    build_evolvepro_input_from_reports,
    _agilent_wt_mean,
    _build_authoritative_from_variant_report,
    _build_fallback_from_raw_report,
    _build_fallback_from_prev_evolvepro,
)
from kuma_core.mame.activity.evolvepro_xlsx import (
    parse_agilent_standard,
    read_evolvepro_rows,
    write_evolvepro_xlsx,
)


def _write_prev_evolvepro(path, rows):
    """Write a previous-round EVOLVEpro xlsx. rows: [(variant, activity)]."""
    write_evolvepro_xlsx(rows, path)


def _write_fid1b(path, pairs):
    """Write a FID1B 5-row-block standard report. pairs: [(sample_name, area)]."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Page 1"
    for name, area in pairs:
        ws.append(["Signal:", None, "FID1B", None, None])
        ws.append([None, "Area", None, "Sample Name", None])
        ws.append([None, area, None, name, None])
        ws.append(["Sum", area, None, None, None])
        ws.append([None, None, None, None, None])
    wb.save(path)


def _write_layout(path, rows):
    """Write a plate layout xlsx. rows: [(mutant, well)]."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Mutant", "Well Pos."])
    for mut, well in rows:
        ws.append([mut, well])
    wb.save(path)


# ---------------------------------------------------------------------------
# T1: round-1 fallback
# ---------------------------------------------------------------------------

def test_t1_round1_fallback(tmp_path):
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])

    round1 = tmp_path / "round1.xlsx"
    _write_fid1b(
        round1,
        [
            ("A1", 0.80),
            ("B1", 0.40),
            ("WT_1", 0.50),
            ("WT_2", 0.50),
            ("WT_3", 0.50),
            ("0", 72.5),  # calibration row, skipped by parser
        ],
    )

    fallback, well_by_variant, warnings = _build_fallback_from_raw_report(
        round1, layout
    )

    # wt_mean = 0.5; rel = area / 0.5
    assert warnings == []
    assert fallback["5F"] == pytest.approx([1.6])
    assert fallback["10L"] == pytest.approx([0.8])
    # calibration '0' skipped; no V*-mapped keys leaked through.
    assert "0" not in fallback
    assert all(not k.startswith("V") for k in fallback)
    assert set(fallback) == {"5F", "10L"}
    assert well_by_variant["5F"] == "A01"
    assert well_by_variant["10L"] == "B01"


# ---------------------------------------------------------------------------
# T2: authoritative variant-labeled
# ---------------------------------------------------------------------------

def test_t2_authoritative_variant_labeled(tmp_path):
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [
            ("5F", 0.60),
            ("5F", 0.66),
            ("5F", 0.54),
            ("WT_1", 0.60),
            ("WT_2", 0.60),
            ("WT_3", 0.60),
        ],
    )

    authoritative, warnings = _build_authoritative_from_variant_report(remeasure)

    # wt_mean = 0.6; rel = area / 0.6 -> [1.0, 1.1, 0.9]
    assert authoritative["5F"] == pytest.approx([1.0, 1.1, 0.9])
    assert set(authoritative) == {"5F"}
    assert warnings == []


# ---------------------------------------------------------------------------
# T3: full merge (authoritative replaces fallback, others kept)
# ---------------------------------------------------------------------------

def test_t3_full_merge(tmp_path):
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])

    round1 = tmp_path / "round1.xlsx"
    _write_fid1b(
        round1,
        [
            ("A1", 0.80),  # 5F -> 1.6 fallback
            ("B1", 0.40),  # 10L -> 0.8 fallback
            ("WT_1", 0.50),
            ("WT_2", 0.50),
            ("WT_3", 0.50),
        ],
    )

    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [
            ("5F", 0.60),  # -> 1.0
            ("5F", 0.66),  # -> 1.1
            ("5F", 0.54),  # -> 0.9 ; mean 1.0 != round1 1.6
            ("WT_1", 0.60),
            ("WT_2", 0.60),
            ("WT_3", 0.60),
        ],
    )

    out = tmp_path / "evolvepro_input.xlsx"
    result = build_evolvepro_input_from_reports(layout, round1, remeasure, out)

    assert result.output_path == out
    assert out.exists()
    assert result.n_variants == 2
    assert result.n_authoritative == 1
    assert result.n_fallback_only == 1

    rows = read_evolvepro_rows(out)
    by_variant = {v: a for v, a in rows}
    assert set(by_variant) == {"5F", "10L"}
    # 5F replaced by authoritative mean (1.0+1.1+0.9)/3 == 1.0
    assert by_variant["5F"] == pytest.approx(1.0)
    # 10L kept from round-1 fallback (0.8)
    assert by_variant["10L"] == pytest.approx(0.8)


# ---------------------------------------------------------------------------
# T4: internal-notation re-measure label normalises and merges
# ---------------------------------------------------------------------------

def test_t4_internal_notation_label(tmp_path):
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])

    round1 = tmp_path / "round1.xlsx"
    _write_fid1b(
        round1,
        [
            ("A1", 0.80),
            ("B1", 0.40),
            ("WT_1", 0.50),
            ("WT_2", 0.50),
            ("WT_3", 0.50),
        ],
    )

    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [
            ("V5F", 0.60),  # internal notation -> short '5F'
            ("WT_1", 0.60),
            ("WT_2", 0.60),
            ("WT_3", 0.60),
        ],
    )

    authoritative, warnings = _build_authoritative_from_variant_report(remeasure)
    assert warnings == []
    assert set(authoritative) == {"5F"}
    assert authoritative["5F"] == pytest.approx([1.0])

    out = tmp_path / "out.xlsx"
    result = build_evolvepro_input_from_reports(layout, round1, remeasure, out)
    rows = {v: a for v, a in read_evolvepro_rows(out)}
    # 5F came from authoritative (V5F normalised); 10L from fallback.
    assert rows["5F"] == pytest.approx(1.0)
    assert rows["10L"] == pytest.approx(0.8)
    assert result.n_authoritative == 1


# ---------------------------------------------------------------------------
# T5: non-variant re-measure label skipped with a warning
# ---------------------------------------------------------------------------

def test_t5_non_variant_label_skipped(tmp_path):
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [
            ("5F", 0.60),
            ("XYZ", 0.99),  # junk, not a variant label
            ("WT_1", 0.60),
            ("WT_2", 0.60),
            ("WT_3", 0.60),
        ],
    )

    authoritative, warnings = _build_authoritative_from_variant_report(remeasure)
    assert "XYZ" not in authoritative
    assert set(authoritative) == {"5F"}
    assert any("XYZ" in w for w in warnings)


# ---------------------------------------------------------------------------
# T6: report with no WT blocks -> _agilent_wt_mean raises
# ---------------------------------------------------------------------------

def test_t6_no_wt_raises(tmp_path):
    report = tmp_path / "no_wt.xlsx"
    _write_fid1b(report, [("5F", 0.60), ("10L", 0.40)])

    records = parse_agilent_standard(report)
    with pytest.raises(ValueError):
        _agilent_wt_mean(records)


# ---------------------------------------------------------------------------
# T7: prev-EVOLVEpro round-1 mode (round-1 baseline is a prior EVOLVEpro file)
# ---------------------------------------------------------------------------

def test_t7_prev_evolvepro_fallback_helper(tmp_path):
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.0), ("10L", 0.9), ("WT", 1.0)])
    fallback, warnings = _build_fallback_from_prev_evolvepro(prev)
    # WT row skipped; variant rows kept as one replicate each.
    assert set(fallback) == {"5F", "10L"}
    assert fallback["5F"] == [1.0]
    assert warnings == []


def test_t8_prev_evolvepro_merge(tmp_path):
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.0), ("10L", 0.9), ("99Z", 0.5)])
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [
            ("V5F", 0.60), ("V5F", 0.66), ("V5F", 0.54),  # mean 0.60 (rel, WT=1.0)
            ("WT_1", 1.0), ("WT_2", 1.0), ("WT_3", 1.0),
        ],
    )
    out = tmp_path / "out.xlsx"
    res = build_evolvepro_input_from_reports(
        None, None, remeasure, out, prev_evolvepro_xlsx=prev
    )
    rows = {v: a for v, a in read_evolvepro_rows(out)}
    assert res.n_authoritative == 1
    assert res.n_fallback_only == 2
    assert res.n_variants == 3
    assert rows["5F"] == pytest.approx(0.60)   # replaced by re-measure mean
    assert rows["10L"] == pytest.approx(0.9)   # kept from prev EVOLVEpro
    assert rows["99Z"] == pytest.approx(0.5)   # kept from prev EVOLVEpro
    assert res.n_ngs_excluded == 0


def test_t9_prev_evolvepro_verdict_without_layout_skips_gating(tmp_path):
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.0), ("10L", 0.9)])
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [("V5F", 0.6), ("V5F", 0.6), ("WT_1", 1.0), ("WT_2", 1.0)],
    )
    verdict = tmp_path / "verdict.xlsx"
    _write_verdict(verdict, [("A01", "V5F", "WRONG_AA"), ("B01", "V10L", "PASS")])
    out = tmp_path / "out.xlsx"
    res = build_evolvepro_input_from_reports(
        None, None, remeasure, out, prev_evolvepro_xlsx=prev, verdict_xlsx=verdict
    )
    # No layout -> no variant->well map -> gating skipped gracefully, all kept.
    assert res.n_ngs_excluded == 0
    assert res.n_variants == 2
    assert any("gating skipped" in w for w in res.warnings)


def test_t10_prev_evolvepro_with_layout_gates(tmp_path):
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.0), ("10L", 0.9)])
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [("V5F", 0.6), ("V5F", 0.6), ("V10L", 0.8), ("V10L", 0.8),
         ("WT_1", 1.0), ("WT_2", 1.0)],
    )
    verdict = tmp_path / "verdict.xlsx"
    _write_verdict(verdict, [("A01", "V5F", "PASS"), ("B01", "V10L", "WRONG_AA")])
    out = tmp_path / "out.xlsx"
    res = build_evolvepro_input_from_reports(
        layout, None, remeasure, out,
        prev_evolvepro_xlsx=prev, verdict_xlsx=verdict,
    )
    # 10L well B01 is WRONG_AA -> excluded; 5F (A01 PASS) survives.
    assert res.n_ngs_excluded == 1
    assert res.ngs_excluded == ["10L"]
    assert res.n_variants == 1


def test_t11_raw_mode_missing_inputs_raises(tmp_path):
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(remeasure, [("V5F", 0.6), ("WT_1", 1.0)])
    out = tmp_path / "out.xlsx"
    # Neither raw round-1 (layout + round1_report) nor prev_evolvepro_xlsx given.
    with pytest.raises(ValueError, match="raw round-1 mode requires"):
        build_evolvepro_input_from_reports(None, None, remeasure, out)


def test_t12_prev_evolvepro_layout_with_control_rows_does_not_crash(tmp_path):
    # prev-mode + verdict + a layout containing non-variant rows (control, WT
    # replicate label) that to_evolvepro would reject. Build must succeed and
    # skip those rows rather than crashing in _well_by_variant_from_layout.
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.0), ("10L", 0.9)])
    layout = tmp_path / "layout.xlsx"
    _write_layout(
        layout,
        [("V5F", "A1"), ("V10L", "B1"), ("BLANK", "C1"), ("WT_1", "D1")],
    )
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [("V5F", 0.6), ("V5F", 0.6), ("WT_1", 1.0), ("WT_2", 1.0)],
    )
    verdict = tmp_path / "verdict.xlsx"
    _write_verdict(verdict, [("A01", "V5F", "PASS"), ("B01", "V10L", "PASS")])
    out = tmp_path / "out.xlsx"
    res = build_evolvepro_input_from_reports(
        layout, None, remeasure, out,
        prev_evolvepro_xlsx=prev, verdict_xlsx=verdict,
    )
    # control/WT rows skipped from the variant->well map; both PASS variants kept.
    assert res.n_variants == 2
    assert res.n_ngs_excluded == 0


def _write_verdict(path, rows):
    """Write an Analyze verdict xlsx. rows: [(well_id, mutant_id, verdict)]."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Final"
    ws.append(
        ["well_id", "selected_plate", "custom_barcode", "mutant_id",
         "verdict", "is_fallback", "fallback_reason", "notes"]
    )
    for w, m, v in rows:
        ws.append([w, "P1", "", m, v, "", "", ""])
    wb.save(path)


# ---------------------------------------------------------------------------
# GC export: optional well-level relative activity artifact (raw round-1 only)
# ---------------------------------------------------------------------------

def _gc_export_case(tmp_path):
    """Shared fixture set for the gc_export_xlsx cases."""
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])

    round1 = tmp_path / "round1.xlsx"
    _write_fid1b(
        round1,
        [
            ("A1", 0.80),
            ("B1", 0.40),
            ("C1", 0.20),  # well absent from the layout: kept by the export
            ("WT_1", 0.50),
            ("WT_2", 0.50),
            ("0", 72.5),  # calibration row, dropped by the parser
        ],
    )

    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [("5F", 0.60), ("WT_1", 0.60), ("WT_2", 0.60)],
    )
    return layout, round1, remeasure


def test_gc_export_written_with_expected_header_and_row_count(tmp_path):
    """Raw round-1 + gc_export_xlsx writes a 'Sample Name' / 'Area' sheet."""
    import openpyxl

    from kuma_core.mame.activity.evolvepro_xlsx import RELATIVE_ACTIVITY_COLUMNS

    layout, round1, remeasure = _gc_export_case(tmp_path)
    gc_export = tmp_path / "gc_export.xlsx"

    res = build_evolvepro_input_from_reports(
        layout,
        round1,
        remeasure,
        tmp_path / "out.xlsx",
        gc_export_xlsx=gc_export,
    )

    assert res.gc_export_path == gc_export
    assert gc_export.exists()

    ws = openpyxl.load_workbook(gc_export).active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == RELATIVE_ACTIVITY_COLUMNS

    n_non_wt = sum(1 for r in parse_agilent_standard(round1) if not r.is_wt)
    assert len(rows) - 1 == n_non_wt == 3


def test_gc_export_values_are_area_over_wt_mean(tmp_path):
    """Every exported value equals raw area / mean WT block area."""
    from kuma_core.mame.activity.evolvepro_xlsx import parse_relative_only

    layout, round1, remeasure = _gc_export_case(tmp_path)
    gc_export = tmp_path / "gc_export.xlsx"

    build_evolvepro_input_from_reports(
        layout,
        round1,
        remeasure,
        tmp_path / "out.xlsx",
        gc_export_xlsx=gc_export,
    )

    records = parse_agilent_standard(round1)
    wt_mean = _agilent_wt_mean(records)
    expected = [
        (r.sample_name, r.area / wt_mean) for r in records if not r.is_wt
    ]

    # Round-trips through the pre-normalised reader, so the file really carries
    # the GC data shape rather than only the right header text.
    exported = parse_relative_only(gc_export)
    assert len(exported) == len(expected)
    for got, (name, value) in zip(exported, expected):
        assert got.sample_name == name
        assert got.area == pytest.approx(value, abs=1e-9)


def test_gc_export_warns_and_writes_nothing_on_prev_evolvepro_round1(tmp_path):
    """The prev-EVOLVEpro round-1 path has no raw report to project."""
    layout, _round1, remeasure = _gc_export_case(tmp_path)
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.6), ("10L", 0.8)])
    gc_export = tmp_path / "gc_export.xlsx"

    res = build_evolvepro_input_from_reports(
        layout,
        None,
        remeasure,
        tmp_path / "out.xlsx",
        prev_evolvepro_xlsx=prev,
        gc_export_xlsx=gc_export,
    )

    assert res.gc_export_path is None
    assert not gc_export.exists()
    ignored = [w for w in res.warnings if "gc_export_xlsx ignored" in w]
    assert len(ignored) == 1
