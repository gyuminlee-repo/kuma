"""Labelled corpus for the step 4.1 measurement-source detector.

The corpus is the specification: every bundled template carries the answer it
must produce, and the synthetic negatives carry the answer ``[]``.  Written
before the detector's rules were fixed, so a rule that happens to sort the six
positives is not enough on its own.

Every assertion is on the exact list, order included, because the whole point
of the ambiguous case is that it names two things and no third.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from kuma_core.mame.activity.detect_measurement_source import (
    CONFIRMATION_VARIANT_LABELS,
    GC_SHEET,
    LONG_FORMAT,
    NUMERIC_REPORT,
    RAW_REPORT,
    detect_measurement_source,
)

TEMPLATES = Path(__file__).resolve().parents[3] / "templates"

# (template file, expected candidates in order).
CORPUS: list[tuple[str, list[str]]] = [
    ("07_mame_activity_long.csv", [LONG_FORMAT]),
    ("07_mame_activity_long.xlsx", [LONG_FORMAT]),
    # The pre-normalised GC sheet is also a valid long-format file, because
    # 'sample name' is a label column and 'area' is a value column.  The two
    # readings disagree about the wild-type rows, so both are reported.
    ("10_mame_gc_prenormalised.xlsx", [GC_SHEET, LONG_FORMAT]),
    ("11_mame_gc_fid_round1_raw.xlsx", [RAW_REPORT]),
    ("12_mame_agilent_numeric_index.xlsx", [NUMERIC_REPORT]),
    ("09_mame_agilent_rep_batch.xlsx", [CONFIRMATION_VARIANT_LABELS]),
]


def _write_block_workbook(path: Path, sample_names: list[str]) -> Path:
    """Write a FID1B block workbook with one data row per name."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Agilent"
    for name in sample_names:
        ws.append(["Signal:", "FID1B"])
        ws.append(["Area", "Sample Name"])
        ws.append([1000.0, name])
        ws.append(["Sum", 1000.0])
        ws.append(["", ""])
    wb.save(path)
    return path


@pytest.mark.parametrize(("filename", "expected"), CORPUS, ids=[case[0] for case in CORPUS])
def test_bundled_templates_detect_as_labelled(filename: str, expected: list[str]) -> None:
    result = detect_measurement_source(TEMPLATES / filename)
    assert result.candidates == expected
    assert result.ambiguous is (len(expected) > 1)
    assert result.reason == ""


def test_gc_sheet_is_exactly_two_candidates() -> None:
    """The ambiguous case names those two and admits nothing else."""
    result = detect_measurement_source(TEMPLATES / "10_mame_gc_prenormalised.xlsx")
    assert result.candidates == [GC_SHEET, LONG_FORMAT]
    assert set(result.candidates) == {GC_SHEET, LONG_FORMAT}
    assert len(result.candidates) == 2
    assert result.ambiguous is True
    assert result.evidence["header"] == ["sample name", "area"]
    assert result.evidence["fid1b_signature"] is False


def test_block_evidence_names_the_namespace_and_the_wild_type_rows() -> None:
    result = detect_measurement_source(TEMPLATES / "12_mame_agilent_numeric_index.xlsx")
    assert result.evidence["fid1b_signature"] is True
    assert result.evidence["n_wt_rows"] == 3
    assert result.evidence["sample_name_namespaces"]["numericId"] == 18
    assert result.evidence["sample_name_namespaces"]["well"] == 0
    # '1-2' and '1-3' are replicates of numeric id 1, and a pure-integer test
    # would put them outside the namespace.
    assert "1-2" in result.evidence["sample_name_samples"]["numericId"]


def test_mixed_namespace_block_file_is_refused(tmp_path: Path) -> None:
    """A block file whose sample names are half wells and half numeric ids."""
    path = _write_block_workbook(
        tmp_path / "mixed.xlsx",
        ["WT1", "WT2", "WT3", "A4", "A5", "1", "2-2"],
    )
    result = detect_measurement_source(path)
    assert result.candidates == []
    assert result.ambiguous is False
    assert "well" in result.reason
    assert "numeric id" in result.reason
    assert "A4" in result.reason
    assert "1" in result.reason
    assert result.evidence["sample_name_namespaces"] == {
        "well": 2,
        "numericId": 2,
        "variantLabel": 0,
        "unclassified": 0,
    }
    assert result.evidence["sample_name_samples"]["well"] == ["A4", "A5"]
    assert result.evidence["sample_name_samples"]["numericId"] == ["1", "2-2"]


def test_bare_wild_type_row_is_not_a_sample(tmp_path: Path) -> None:
    """`parse_agilent_block_rep_batch` accepts a bare `WT`, so this must too.

    `WT_PATTERN` requires a replicate number, and the numeric-ID parser widens
    it (`evolvepro_xlsx.py:654`).  Reading the narrow test here would refuse a
    file that path parses.
    """
    path = _write_block_workbook(
        tmp_path / "bare_wt.xlsx",
        ["WT", "WT_2", "1", "1-2", "2"],
    )
    result = detect_measurement_source(path)
    assert result.candidates == [NUMERIC_REPORT]
    assert result.evidence["n_wt_rows"] == 2
    assert result.evidence["sample_name_namespaces"]["unclassified"] == 0


def test_block_file_with_only_wild_type_rows_is_refused(tmp_path: Path) -> None:
    path = _write_block_workbook(tmp_path / "wt_only.xlsx", ["WT1", "WT2", "WT3"])
    result = detect_measurement_source(path)
    assert result.candidates == []
    assert result.evidence["n_wt_rows"] == 3
    assert "wild-type" in result.reason


def test_empty_workbook_is_refused(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    Workbook().save(path)
    result = detect_measurement_source(path)
    assert result.candidates == []
    assert result.ambiguous is False
    assert result.reason
    assert result.evidence["fid1b_signature"] is False


def test_unrecognised_header_is_echoed(tmp_path: Path) -> None:
    """A refusal states what was seen, so an operator can tell why."""
    path = tmp_path / "layout_like.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Mutant", "Well Pos."])
    ws.append(["F89W", "A1"])
    wb.save(path)
    result = detect_measurement_source(path)
    assert result.candidates == []
    assert result.evidence["header"] == ["mutant", "well pos."]
    assert "mutant" in result.reason


def test_csv_without_a_value_column_is_refused(tmp_path: Path) -> None:
    path = tmp_path / "no_value.csv"
    path.write_text("plate_id,well_id,replicate_idx\nplate01,A1,1\n", encoding="utf-8")
    result = detect_measurement_source(path)
    assert result.candidates == []
    assert result.evidence["value_columns"] == []
    assert "well_id" in str(result.evidence["header"])


def test_missing_file_raises() -> None:
    with pytest.raises(FileNotFoundError):
        detect_measurement_source(TEMPLATES / "does_not_exist.xlsx")


def test_unsupported_extension_is_refused(tmp_path: Path) -> None:
    path = tmp_path / "measurements.json"
    path.write_text("{}", encoding="utf-8")
    result = detect_measurement_source(path)
    assert result.candidates == []
    assert ".json" in result.reason
