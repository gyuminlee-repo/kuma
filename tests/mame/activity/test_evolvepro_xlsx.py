"""Tests for kuma_core.mame.activity.evolvepro_xlsx.

Covers:
  - detect_format for all 4 formats
  - parse_agilent_standard (FID1B block)
  - parse_agilent_rep_batch (numeric ID)
  - parse_relative_only
  - read_evolvepro_xlsx / write_evolvepro_xlsx
"""

from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.activity.evolvepro_xlsx import (
    XlsxFormat,
    detect_format,
    parse_agilent_rep_batch,
    parse_agilent_standard,
    parse_relative_only,
    read_evolvepro_xlsx,
    write_evolvepro_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _save(wb: openpyxl.Workbook, path: Path) -> Path:
    wb.save(str(path))
    return path


def _make_relative_only(tmp_path: Path, filename: str = "gc.xlsx") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Sample Name", "Area"])
    ws.append(["F89W_A", 0.85])
    ws.append(["WT_1", 1.0])
    ws.append(["F89W_B", 0.90])
    return _save(wb, tmp_path / filename)


def _make_evolvepro(tmp_path: Path, filename: str = "ep.xlsx") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Variant", "activity"])
    ws.append(["89W", 0.85])
    ws.append(["10A", 1.10])
    ws.append(["WT", 1.0])
    return _save(wb, tmp_path / filename)


def _make_rep_batch(tmp_path: Path, filename: str = "rep.xlsx") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Sample Name", "Area"])
    for i in range(1, 5):
        for rep in range(1, 4):
            ws.append([f"{i}_rep{rep}", float(i * 100 + rep)])
    ws.append(["WT_1", 500.0])
    ws.append(["WT_2", 510.0])
    return _save(wb, tmp_path / filename)


def _make_agilent_standard(tmp_path: Path, filename: str = "std.xlsx") -> Path:
    """Minimal FID1B block xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    # Block 1
    ws.append(["Signal: FID1B,CD"])
    ws.append(["Sample Name", "Area"])
    ws.append(["F89W", 12345.6])
    ws.append(["WT_1", 15000.0])
    ws.append(["0", 9999.0])   # calibration row → should be skipped
    ws.append(["Sum", ""])
    # Block 2
    ws.append(["Signal: FID1B,CD"])
    ws.append(["Sample Name", "Area"])
    ws.append(["G10A", 11000.0])
    ws.append(["Sum", ""])
    return _save(wb, tmp_path / filename)


# ---------------------------------------------------------------------------
# detect_format
# ---------------------------------------------------------------------------

def test_detect_agilent_standard(tmp_path: Path):
    path = _make_agilent_standard(tmp_path)
    assert detect_format(path) == XlsxFormat.AGILENT_STANDARD


def test_detect_evolvepro(tmp_path: Path):
    path = _make_evolvepro(tmp_path)
    assert detect_format(path) == XlsxFormat.EVOLVEPRO


def test_detect_rep_batch(tmp_path: Path):
    path = _make_rep_batch(tmp_path)
    assert detect_format(path) == XlsxFormat.AGILENT_REP_BATCH


def test_detect_relative_only(tmp_path: Path):
    path = _make_relative_only(tmp_path)
    assert detect_format(path) == XlsxFormat.RELATIVE_ONLY


def test_detect_unknown_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Column1", "Column2"])
    ws.append([1, 2])
    path = tmp_path / "unknown.xlsx"
    _save(wb, path)
    with pytest.raises(ValueError, match="cannot determine format"):
        detect_format(path)


# ---------------------------------------------------------------------------
# parse_agilent_standard
# ---------------------------------------------------------------------------

def test_parse_agilent_standard_basic(tmp_path: Path):
    path = _make_agilent_standard(tmp_path)
    records = parse_agilent_standard(path)
    # calibration row '0' skipped; F89W + WT_1 + G10A = 3 records
    assert len(records) == 3
    names = [r.sample_name for r in records]
    assert "F89W" in names
    assert "WT_1" in names
    assert "G10A" in names
    assert "0" not in names


def test_parse_agilent_standard_wt_detected(tmp_path: Path):
    path = _make_agilent_standard(tmp_path)
    records = parse_agilent_standard(path)
    wt_records = [r for r in records if r.is_wt]
    assert len(wt_records) == 1
    assert wt_records[0].replicate_n == 1


def test_parse_agilent_standard_calibration_skipped(tmp_path: Path):
    path = _make_agilent_standard(tmp_path)
    records = parse_agilent_standard(path)
    assert all(r.sample_name != "0" for r in records)


def test_parse_agilent_standard_non_numeric_area_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Signal: FID1B"])
    ws.append(["Sample Name", "Area"])
    ws.append(["F89W", "NOT_A_NUMBER"])
    ws.append(["Sum", ""])
    path = tmp_path / "bad_area.xlsx"
    _save(wb, path)
    with pytest.raises(ValueError, match="Cannot convert"):
        parse_agilent_standard(path)


# ---------------------------------------------------------------------------
# parse_agilent_rep_batch
# ---------------------------------------------------------------------------

def test_parse_agilent_rep_batch_basic(tmp_path: Path):
    path = _make_rep_batch(tmp_path)
    records = parse_agilent_rep_batch(path)
    # 4 mutants × 3 reps + 2 WT = 14
    assert len(records) == 14


def test_parse_agilent_rep_batch_auto_mutant_count(tmp_path: Path):
    path = _make_rep_batch(tmp_path)
    # Pass mutant_count=None to trigger auto-estimation.
    records = parse_agilent_rep_batch(path, mutant_count=None)
    assert len(records) == 14


def test_parse_agilent_rep_batch_wt_detected(tmp_path: Path):
    path = _make_rep_batch(tmp_path)
    records = parse_agilent_rep_batch(path)
    wt_records = [r for r in records if r.is_wt]
    assert len(wt_records) == 2


def test_parse_agilent_rep_batch_missing_header_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["ID", "Value"])
    ws.append([1, 100.0])
    path = tmp_path / "bad_header.xlsx"
    _save(wb, path)
    with pytest.raises(ValueError, match="columns not found"):
        parse_agilent_rep_batch(path)


# ---------------------------------------------------------------------------
# parse_relative_only
# ---------------------------------------------------------------------------

def test_parse_relative_only_basic(tmp_path: Path):
    path = _make_relative_only(tmp_path)
    records = parse_relative_only(path)
    assert len(records) == 3
    assert all(r.is_relative is True for r in records)


def test_parse_relative_only_area_values(tmp_path: Path):
    path = _make_relative_only(tmp_path)
    records = parse_relative_only(path)
    area_map = {r.sample_name: r.area for r in records}
    assert area_map["F89W_A"] == pytest.approx(0.85)
    assert area_map["WT_1"] == pytest.approx(1.0)


def test_parse_relative_only_missing_header_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Name", "Value"])
    ws.append(["F89W", 0.9])
    path = tmp_path / "bad.xlsx"
    _save(wb, path)
    with pytest.raises(ValueError, match="Sample Name"):
        parse_relative_only(path)


def test_parse_relative_only_non_numeric_area_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Sample Name", "Area"])
    ws.append(["F89W", "n/a"])
    path = tmp_path / "bad_area.xlsx"
    _save(wb, path)
    with pytest.raises(ValueError, match="Cannot convert"):
        parse_relative_only(path)


# ---------------------------------------------------------------------------
# read_evolvepro_xlsx / write_evolvepro_xlsx
# ---------------------------------------------------------------------------

def test_read_evolvepro_xlsx(tmp_path: Path):
    path = _make_evolvepro(tmp_path)
    result = read_evolvepro_xlsx(path)
    assert "89W" in result
    assert result["89W"] == pytest.approx(0.85)
    assert "WT" in result


def test_write_then_read_evolvepro_xlsx(tmp_path: Path):
    data = [("89W", 0.85), ("10A", 1.10), ("200N", 0.50)]
    out_path = tmp_path / "out.xlsx"
    n = write_evolvepro_xlsx(data, out_path)
    assert n == 3
    assert out_path.exists()

    result = read_evolvepro_xlsx(out_path)
    assert len(result) == 3
    assert result["89W"] == pytest.approx(0.85)


def test_write_evolvepro_xlsx_missing_dir_raises(tmp_path: Path):
    out_path = tmp_path / "nonexistent_dir" / "out.xlsx"
    with pytest.raises(FileNotFoundError):
        write_evolvepro_xlsx([("89W", 0.85)], out_path)


def test_read_evolvepro_xlsx_missing_header_raises(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Mutation", "Score"])
    ws.append(["89W", 0.85])
    path = tmp_path / "bad.xlsx"
    _save(wb, path)
    with pytest.raises(ValueError, match="Variant"):
        read_evolvepro_xlsx(path)


def test_write_evolvepro_xlsx_header(tmp_path: Path):
    out_path = tmp_path / "ep_out.xlsx"
    write_evolvepro_xlsx([("89W", 1.0)], out_path)
    import python_calamine
    wb2 = python_calamine.CalamineWorkbook.from_path(str(out_path))
    rows = list(wb2.get_sheet_by_index(0).to_python())
    header = [str(c).strip() for c in rows[0]]
    assert header == ["Variant", "activity"]
