"""Handler-level contract test for reports-mode build_evolvepro_input.

Exercises handle_build_evolvepro_input end-to-end (model validation + dispatch +
response shape), which the core unit tests bypass by calling
build_evolvepro_input_from_reports directly. Pins the 13-key response contract
that the TS BuildEvolveproInputResult mirrors, including the reports-mode
hardcoded values (mapping_audit_path="", swap_warnings=[], prev_descending=True)
and the PR3 NGS gating fields.
"""

from __future__ import annotations

import openpyxl

from sidecar_mame.handlers.activity import handle_build_evolvepro_input


def _write_fid1b(path, pairs):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Page 1"
    for name, area in pairs:
        ws.append(["Signal:", None, "FID1B", None, None])
        ws.append([None, "Area", None, "Sample Name", None])
        ws.append([None, area, None, name, None])
        ws.append(["Sum", area, None, None, None])
        ws.append([None, None, None, None, None])
    wb.save(path)


def _write_layout(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Mutant", "Well Pos."])
    for mut, well in rows:
        ws.append([mut, well])
    wb.save(path)


def _write_verdict(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Final"
    ws.append(
        ["well_id", "selected_plate", "custom_barcode", "mutant_id",
         "verdict", "is_fallback", "fallback_reason", "notes"]
    )
    for w, m, v in rows:
        ws.append([w, "P1", "", m, v, "", "", ""])
    wb.save(path)


def _write_prev_evolvepro(path, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Variant", "activity"])
    for variant, activity in rows:
        ws.append([variant, activity])
    wb.save(path)


# The 13-key response contract mirrored by TS BuildEvolveproInputResult.
_RESULT_KEYS = {
    "output_path", "n_variants", "n_authoritative", "n_fallback_only",
    "mapping_audit", "mapping_audit_path", "prev_descending", "warnings",
    "swap_warnings", "mismatched", "mode", "n_ngs_excluded", "ngs_excluded",
}


def _inputs(tmp_path):
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])
    round1 = tmp_path / "round1.xlsx"
    _write_fid1b(
        round1,
        [("A1", 0.80), ("B1", 0.40), ("WT_1", 0.50), ("WT_2", 0.50), ("WT_3", 0.50)],
    )
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [("5F", 0.60), ("5F", 0.66), ("5F", 0.54),
         ("10L", 0.50), ("10L", 0.50), ("10L", 0.50),
         ("WT_1", 0.60), ("WT_2", 0.60), ("WT_3", 0.60)],
    )
    out = tmp_path / "out.xlsx"
    return layout, round1, remeasure, out


def test_handler_reports_mode_contract(tmp_path):
    layout, round1, remeasure, out = _inputs(tmp_path)
    resp = handle_build_evolvepro_input(
        {
            "layout_xlsx": str(layout),
            "round1_report_xlsx": str(round1),
            "remeasure_report_xlsx": str(remeasure),
            "output_xlsx": str(out),
        }
    )
    assert set(resp) == _RESULT_KEYS
    assert resp["mode"] == "reports"
    assert isinstance(resp["n_variants"], int) and resp["n_variants"] == 2
    assert isinstance(resp["n_authoritative"], int)
    assert isinstance(resp["n_fallback_only"], int)
    assert isinstance(resp["mapping_audit"], list)
    assert resp["mapping_audit_path"] == ""
    assert resp["prev_descending"] is True
    assert isinstance(resp["warnings"], list)
    assert resp["swap_warnings"] == []
    assert isinstance(resp["mismatched"], list)
    assert resp["n_ngs_excluded"] == 0
    assert resp["ngs_excluded"] == []
    assert out.exists()


def test_handler_reports_mode_with_ngs_gating(tmp_path):
    layout, round1, remeasure, out = _inputs(tmp_path)
    verdict = tmp_path / "verdict.xlsx"
    _write_verdict(verdict, [("A01", "V5F", "PASS"), ("B01", "V10L", "WRONG_AA")])
    resp = handle_build_evolvepro_input(
        {
            "layout_xlsx": str(layout),
            "round1_report_xlsx": str(round1),
            "remeasure_report_xlsx": str(remeasure),
            "verdict_xlsx": str(verdict),
            "output_xlsx": str(out),
        }
    )
    assert set(resp) == _RESULT_KEYS
    assert resp["mode"] == "reports"
    assert resp["n_ngs_excluded"] == 1
    assert resp["ngs_excluded"] == ["10L"]
    # only 5F (PASS well A01) survives gating; 10L (B01 WRONG_AA) excluded.
    assert resp["n_variants"] == 1


def test_handler_reports_mode_prev_evolvepro(tmp_path):
    # Round-1 baseline as a prior EVOLVEpro file (no layout / round1 report).
    prev = tmp_path / "prev.xlsx"
    _write_prev_evolvepro(prev, [("5F", 1.0), ("10L", 0.9), ("99Z", 0.5)])
    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [("V5F", 0.60), ("V5F", 0.66), ("V5F", 0.54),
         ("WT_1", 1.0), ("WT_2", 1.0), ("WT_3", 1.0)],
    )
    out = tmp_path / "out.xlsx"
    resp = handle_build_evolvepro_input(
        {
            "round1_evolvepro_xlsx": str(prev),
            "remeasure_report_xlsx": str(remeasure),
            "output_xlsx": str(out),
        }
    )
    assert set(resp) == _RESULT_KEYS
    assert resp["mode"] == "reports"
    assert resp["n_authoritative"] == 1   # 5F re-measured
    assert resp["n_fallback_only"] == 2   # 10L, 99Z kept from prev EVOLVEpro
    assert resp["n_variants"] == 3
    assert resp["n_ngs_excluded"] == 0
    assert out.exists()
