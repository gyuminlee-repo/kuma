"""Tests for optional NGS verdict gating in the reports-mode EVOLVEpro build.

When a verdict xlsx is supplied to build_evolvepro_input_from_reports, variants
whose layout well carries an explicit non-PASS verdict (NGS-failed designs) are
excluded from the assembled input. When the verdict file is absent, behaviour is
unchanged (graceful, layout-trust). A well present in the layout but absent from
the verdict file is treated as not-assessed and kept.

The verdict source mirrors the Analyze Excel report's Final sheet (header
includes well_id, mutant_id, verdict). Synthetic fixtures are written in-process
via openpyxl so the tests carry no machine-specific data paths.
"""

from __future__ import annotations

import pytest

from kuma_core.mame.activity.build_evolvepro_input import (
    build_evolvepro_input_from_reports,
)
from kuma_core.mame.activity.evolvepro_xlsx import read_evolvepro_rows
from kuma_core.mame.activity.verdict_ngs import parse_verdict_wells


def _write_fid1b(path, pairs):
    """Write a FID1B 5-row-block standard report. pairs: [(sample_name, area)]."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Page 1"
    for name, area in pairs:
        ws.append(["Signal:", None, "FID1B", None, None])
        ws.append([None, "Area", None, "Sample Name", None])
        ws.append([None, area, None, name, None])
        ws.append(["Sum", area, None, None, None])
        ws.append([None, None, None, None, None])
    wb.save(path)


def _write_layout(path, rows):
    """Write a plate layout xlsx. rows: [(mutant, well)]."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Mutant", "Well Pos."])
    for mut, well in rows:
        ws.append([mut, well])
    wb.save(path)


def _write_verdict(path, rows):
    """Write a verdict xlsx Final sheet. rows: [(well_id, mutant_id, verdict)]."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Final"
    ws.append(
        [
            "well_id",
            "selected_plate",
            "custom_barcode",
            "mutant_id",
            "verdict",
            "is_fallback",
            "fallback_reason",
            "notes",
        ]
    )
    for w, m, v in rows:
        ws.append([w, "P1", "", m, v, "", "", ""])
    wb.save(path)


def _standard_inputs(tmp_path):
    """Build the shared layout + round-1 + re-measure inputs used by G2..G5.

    Layout: 5F -> A1, 10L -> B1. Round-1 (wt_mean 0.5): A1 0.8 -> 1.6,
    B1 0.4 -> 0.8. Re-measure (wt_mean 0.5): 5F x3 -> [1.6,1.6,1.6],
    10L x3 -> [0.8,0.8,0.8]. Authoritative means match the round-1 fallbacks
    so no variant lands in replicate_stats.mismatched.
    """
    layout = tmp_path / "layout.xlsx"
    _write_layout(layout, [("V5F", "A1"), ("V10L", "B1")])

    round1 = tmp_path / "round1.xlsx"
    _write_fid1b(
        round1,
        [
            ("A1", 0.80),
            ("B1", 0.40),
            ("WT_1", 0.50),
            ("WT_2", 0.50),
            ("WT_3", 0.50),
        ],
    )

    remeasure = tmp_path / "remeasure.xlsx"
    _write_fid1b(
        remeasure,
        [
            ("5F", 0.80),
            ("5F", 0.80),
            ("5F", 0.80),
            ("10L", 0.40),
            ("10L", 0.40),
            ("10L", 0.40),
            ("WT_1", 0.50),
            ("WT_2", 0.50),
            ("WT_3", 0.50),
        ],
    )
    return layout, round1, remeasure


# ---------------------------------------------------------------------------
# G1: parse_verdict_wells basics
# ---------------------------------------------------------------------------

def test_g1_parse_verdict_wells(tmp_path):
    vfile = tmp_path / "verdict.xlsx"
    _write_verdict(
        vfile,
        [
            ("A01", "V5F", "PASS"),
            ("B01", "V10L", "WRONG_AA"),
            ("", "", ""),  # fully blank row -> skipped
        ],
    )

    result = parse_verdict_wells(vfile)
    assert result == {"A01": "PASS", "B01": "WRONG_AA"}


def test_g1_well_normalises(tmp_path):
    # 'A1'-style well input is normalised to zero-padded 'A01'.
    vfile = tmp_path / "verdict.xlsx"
    _write_verdict(vfile, [("a1", "V5F", "pass"), ("B1", "V10L", "wrong_aa")])

    result = parse_verdict_wells(vfile)
    assert result == {"A01": "PASS", "B01": "WRONG_AA"}


# ---------------------------------------------------------------------------
# G1b: PASS-priority dedupe
# ---------------------------------------------------------------------------

def test_g1b_pass_priority_dedupe(tmp_path):
    vfile = tmp_path / "verdict.xlsx"
    _write_verdict(
        vfile,
        [
            ("A01", "V5F", "AMBIGUOUS"),
            ("A01", "V5F", "PASS"),  # PASS wins over the earlier AMBIGUOUS row
        ],
    )

    result = parse_verdict_wells(vfile)
    assert result["A01"] == "PASS"


# ---------------------------------------------------------------------------
# G1c: no qualifying sheet -> ValueError
# ---------------------------------------------------------------------------

def test_g1c_no_qualifying_sheet_raises(tmp_path):
    import openpyxl

    vfile = tmp_path / "no_verdict.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Other"
    ws.append(["well_id", "mutant_id", "something_else"])  # no 'verdict' column
    ws.append(["A01", "V5F", "x"])
    wb.save(vfile)

    with pytest.raises(ValueError):
        parse_verdict_wells(vfile)


# ---------------------------------------------------------------------------
# G2: gating excludes a non-PASS variant
# ---------------------------------------------------------------------------

def test_g2_gating_excludes_non_pass(tmp_path):
    layout, round1, remeasure = _standard_inputs(tmp_path)

    vfile = tmp_path / "verdict.xlsx"
    _write_verdict(vfile, [("A01", "V5F", "PASS"), ("B01", "V10L", "WRONG_AA")])

    out = tmp_path / "out.xlsx"
    result = build_evolvepro_input_from_reports(
        layout, round1, remeasure, out, verdict_xlsx=vfile
    )

    assert result.n_ngs_excluded == 1
    assert "10L" in result.ngs_excluded
    by_variant = {v: a for v, a in read_evolvepro_rows(out)}
    assert set(by_variant) == {"5F"}
    assert any("excluded" in w.lower() and "10L" in w for w in result.warnings)
    # spec 2d: n_authoritative/n_fallback_only describe pre-gating source counts;
    # both 5F and 10L are in the re-measure report, so n_authoritative stays 2
    # even though 10L is gated out of the written rows.
    assert result.n_authoritative == 2


# ---------------------------------------------------------------------------
# G3: graceful absent (verdict_xlsx=None) -> unchanged behaviour
# ---------------------------------------------------------------------------

def test_g3_graceful_absent(tmp_path):
    layout, round1, remeasure = _standard_inputs(tmp_path)

    out = tmp_path / "out.xlsx"
    result = build_evolvepro_input_from_reports(
        layout, round1, remeasure, out, verdict_xlsx=None
    )

    assert result.n_ngs_excluded == 0
    assert result.ngs_excluded == []
    by_variant = {v: a for v, a in read_evolvepro_rows(out)}
    assert set(by_variant) == {"5F", "10L"}


# ---------------------------------------------------------------------------
# G4: PASS kept + unknown well kept
# ---------------------------------------------------------------------------

def test_g4_pass_and_unknown_well_kept(tmp_path):
    layout, round1, remeasure = _standard_inputs(tmp_path)

    # Verdict only assesses A01 (5F) as PASS; B01 (10L) is absent (not assessed).
    vfile = tmp_path / "verdict.xlsx"
    _write_verdict(vfile, [("A01", "V5F", "PASS")])

    out = tmp_path / "out.xlsx"
    result = build_evolvepro_input_from_reports(
        layout, round1, remeasure, out, verdict_xlsx=vfile
    )

    assert result.n_ngs_excluded == 0
    by_variant = {v: a for v, a in read_evolvepro_rows(out)}
    assert set(by_variant) == {"5F", "10L"}


# ---------------------------------------------------------------------------
# G5: all variants excluded -> ValueError
# ---------------------------------------------------------------------------

def test_g5_all_excluded_raises(tmp_path):
    layout, round1, remeasure = _standard_inputs(tmp_path)

    vfile = tmp_path / "verdict.xlsx"
    _write_verdict(
        vfile,
        [("A01", "V5F", "WRONG_AA"), ("B01", "V10L", "WRONG_AA")],
    )

    out = tmp_path / "out.xlsx"
    with pytest.raises(ValueError):
        build_evolvepro_input_from_reports(
            layout, round1, remeasure, out, verdict_xlsx=vfile
        )
