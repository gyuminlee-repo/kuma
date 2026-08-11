"""Tests for kuma_core.mame.ingest.run_meta and __kuma_meta__ sheet export.

Fixture layout mirrors a typical MinKNOW run:

    tmp_path/
      run_xyz/                           <- MinKNOW run dir (has final_summary)
        final_summary_PAX12345_abc.txt
        sample_sheet_PAX12345.csv
        sort_barcode06/                  <- input_dir supplied to discover
          NB01/
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.export.excel_writer import _write_kuma_meta_sheet, write_excel
from kuma_core.mame.export.janus_mapping import (
    JanusSettings,
    export_mame_janus_csv,
    export_mame_janus_xlsx,
)
from kuma_core.mame.ingest.run_meta import NgsRunMeta, discover_run_meta
from kuma_core.mame.models import (
    BarcodeRecord,
    ReplicateResult,
    TranslatedRecord,
    VerdictClass,
    VerdictRecord,
)


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------

_FINAL_SUMMARY_CONTENT = """\
instrument = PAX12345
position = X3
flow_cell_id = FAW12345
sample_id = my_sample
kit = SQK-LSK109
started = 2024-03-15T10:00:00Z
basecalling_enabled = true
"""

_SAMPLE_SHEET_CONTENT = """\
flow_cell_product_code,FLO-MIN106D
kit,SQK-LSK109
"""


# These cases assert the run-meta comment and the __kuma_meta__ sheet on the
# kuma-internal 5-column output, which is no longer the default schema.
_LEGACY5 = JanusSettings(output_schema="legacy5", dest_layout="source")


def _make_run_dir(parent: Path, dirname: str = "run_xyz") -> Path:
    """Create a mock MinKNOW run directory under *parent*."""
    run_dir = parent / dirname
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "final_summary_PAX12345_abc.txt").write_text(
        _FINAL_SUMMARY_CONTENT, encoding="utf-8"
    )
    (run_dir / "sample_sheet_PAX12345.csv").write_text(
        _SAMPLE_SHEET_CONTENT, encoding="utf-8"
    )
    return run_dir


def _make_verdict(
    nb: str,
    custom: str,
    verdict: VerdictClass = VerdictClass.PASS,
    size_kb: float = 80.0,
) -> VerdictRecord:
    barcode = BarcodeRecord(
        native_barcode=nb,
        custom_barcode=custom,
        consensus_seq="",
        file_size_kb=size_kb,
        source_path=Path("/tmp/mock.fasta"),
    )
    translated = TranslatedRecord(
        barcode=barcode,
        aa_sequence="",
        observed_nt_changes=[],
        observed_aa_changes=[],
    )
    return VerdictRecord(
        translated=translated,
        expected_mutations=[],
        verdict=verdict,
        verdict_notes="",
    )


def _make_replicate(
    mutant_id: str,
    nb: str,
    custom: str,
    verdict: VerdictClass = VerdictClass.PASS,
    size_kb: float = 80.0,
) -> ReplicateResult:
    vr = _make_verdict(nb, custom, verdict, size_kb=size_kb)
    return ReplicateResult(
        mutant_id=mutant_id,
        plate_verdicts={nb: vr},
        selected_plate=nb,
        selection_reason="pass",
        failed=False,
    )


# ---------------------------------------------------------------------------
# discover_run_meta tests
# ---------------------------------------------------------------------------


def test_discover_run_meta_from_input_dir(tmp_path: Path) -> None:
    """input_dir is a subdirectory of a MinKNOW run dir — meta is discovered."""
    run_dir = _make_run_dir(tmp_path)
    input_dir = run_dir / "sort_barcode06"
    input_dir.mkdir()

    meta = discover_run_meta(input_dir)

    assert meta is not None
    assert meta.instrument == "PAX12345"
    assert meta.position == "X3"
    assert meta.flow_cell_id == "FAW12345"
    assert meta.sample_id == "my_sample"
    assert meta.kit == "SQK-LSK109"
    assert meta.started == "2024-03-15T10:00:00Z"
    assert meta.basecalling_enabled is True
    assert meta.raw_run_dir is not None
    assert "run_xyz" in meta.raw_run_dir


def test_discover_run_meta_input_dir_is_run_dir(tmp_path: Path) -> None:
    """input_dir itself is the MinKNOW run dir."""
    run_dir = _make_run_dir(tmp_path)
    meta = discover_run_meta(run_dir)
    assert meta is not None
    assert meta.flow_cell_id == "FAW12345"


def test_discover_run_meta_sibling(tmp_path: Path) -> None:
    """Run dir and input_dir are siblings under the same parent."""
    run_dir = _make_run_dir(tmp_path)
    input_dir = tmp_path / "sort_barcode06"
    input_dir.mkdir()

    meta = discover_run_meta(input_dir)

    assert meta is not None
    assert meta.flow_cell_id == "FAW12345"


def test_discover_run_meta_no_run_dir_returns_none(tmp_path: Path) -> None:
    """No MinKNOW artefacts in the search tree — returns None."""
    input_dir = tmp_path / "plain_dir" / "barcode"
    input_dir.mkdir(parents=True)

    meta = discover_run_meta(input_dir)

    assert meta is None


def test_discover_run_meta_nonexistent_dir_returns_none(tmp_path: Path) -> None:
    """Nonexistent input_dir — returns None without raising."""
    input_dir = tmp_path / "does_not_exist"
    meta = discover_run_meta(input_dir)
    assert meta is None


def test_discover_run_meta_basecalling_false(tmp_path: Path) -> None:
    """basecalling_enabled = false is parsed correctly."""
    run_dir = tmp_path / "run2"
    run_dir.mkdir()
    (run_dir / "final_summary_X.txt").write_text(
        "basecalling_enabled = false\nflow_cell_id = ZZZ\n", encoding="utf-8"
    )
    input_dir = run_dir / "sub"
    input_dir.mkdir()
    meta = discover_run_meta(input_dir)
    assert meta is not None
    assert meta.basecalling_enabled is False
    assert meta.flow_cell_id == "ZZZ"


def test_discover_run_meta_kit_from_sample_sheet(tmp_path: Path) -> None:
    """Kit absent from final_summary but present in sample_sheet."""
    run_dir = tmp_path / "run3"
    run_dir.mkdir()
    # final_summary with no kit
    (run_dir / "final_summary_Y.txt").write_text(
        "flow_cell_id = ABC\n", encoding="utf-8"
    )
    (run_dir / "sample_sheet_Y.csv").write_text(
        "kit,SQK-RBK004\n", encoding="utf-8"
    )
    input_dir = run_dir / "barcode01"
    input_dir.mkdir()
    meta = discover_run_meta(input_dir)
    assert meta is not None
    assert meta.kit == "SQK-RBK004"


def test_discover_run_meta_single_sibling_run_dir(tmp_path: Path) -> None:
    """One sibling run folder next to the sorted output is still resolved."""
    base = tmp_path / "campaign"
    (base / "runA").mkdir(parents=True)
    (base / "sorted_output" / "barcode01").mkdir(parents=True)
    (base / "runA" / "final_summary_A.txt").write_text(
        "flow_cell_id=AAA\n", encoding="utf-8"
    )
    meta = discover_run_meta(base / "sorted_output")
    assert meta is not None
    assert meta.flow_cell_id == "AAA"


def test_discover_run_meta_ambiguous_sibling_run_dirs(tmp_path: Path) -> None:
    """Two sibling run folders is a guess with no answer, so nothing is claimed.

    The previous code returned whichever ``iterdir`` yielded first, stamping one
    run flow cell onto another run results with no warning.
    """
    base = tmp_path / "campaign"
    (base / "runA").mkdir(parents=True)
    (base / "runB").mkdir(parents=True)
    (base / "sorted_output" / "barcode01").mkdir(parents=True)
    (base / "runA" / "final_summary_A.txt").write_text(
        "flow_cell_id=AAA\n", encoding="utf-8"
    )
    (base / "runB" / "final_summary_B.txt").write_text(
        "flow_cell_id=BBB\n", encoding="utf-8"
    )
    assert discover_run_meta(base / "sorted_output") is None


def test_discover_run_meta_ancestor_beats_ambiguous_siblings(tmp_path: Path) -> None:
    """A run directory on the direct path is unambiguous and still wins."""
    run_dir = tmp_path / "the_run"
    (run_dir / "fastq_pass").mkdir(parents=True)
    (run_dir / "final_summary_R.txt").write_text(
        "flow_cell_id=REAL\n", encoding="utf-8"
    )
    for name in ("otherA", "otherB"):
        (tmp_path / name).mkdir()
        (tmp_path / name / f"final_summary_{name}.txt").write_text(
            "flow_cell_id=NOPE\n", encoding="utf-8"
        )
    meta = discover_run_meta(run_dir / "fastq_pass")
    assert meta is not None
    assert meta.flow_cell_id == "REAL"


def test_discover_run_meta_kit_from_sample_sheet_column(tmp_path: Path) -> None:
    """Kit as a CSV column, the layout MinKNOW actually writes.

    Verbatim header and row shape taken from a GridION run
    (``sample_sheet_FBF10847_20260212_2212_e7145f8e.csv``). The header starts
    with ``protocol_run_id``, so a line-prefix scan for ``kit,`` never fires.
    """
    run_dir = tmp_path / "run_col"
    run_dir.mkdir()
    (run_dir / "final_summary_Z.txt").write_text(
        "flow_cell_id=FBF10847\n", encoding="utf-8"
    )
    (run_dir / "sample_sheet_Z.csv").write_text(
        "protocol_run_id,position_id,flow_cell_id,sample_id,experiment_id,"
        "flow_cell_product_code,kit\n"
        "e7145f8e-9fba-4941-bba8-3056a32c8469,X4,FBF10847,260212_KHM,"
        "260212_KHM,FLO-MIN114,SQK-NBD114-24\n",
        encoding="utf-8",
    )
    input_dir = run_dir / "fastq_pass"
    input_dir.mkdir()
    meta = discover_run_meta(input_dir)
    assert meta is not None
    assert meta.kit == "SQK-NBD114-24"


def test_sample_sheet_kit_column_absent_returns_none(tmp_path: Path) -> None:
    """A sheet without a kit column yields None rather than a stray cell."""
    run_dir = tmp_path / "run_nokit"
    run_dir.mkdir()
    (run_dir / "final_summary_W.txt").write_text(
        "flow_cell_id=AAA\n", encoding="utf-8"
    )
    (run_dir / "sample_sheet_W.csv").write_text(
        "protocol_run_id,position_id,flow_cell_id\nabc,X1,AAA\n",
        encoding="utf-8",
    )
    input_dir = run_dir / "fastq_pass"
    input_dir.mkdir()
    meta = discover_run_meta(input_dir)
    assert meta is not None
    assert meta.kit is None


def test_sample_sheet_kit_column_blank_value_returns_none(tmp_path: Path) -> None:
    """An empty kit cell is not a kit."""
    run_dir = tmp_path / "run_blankkit"
    run_dir.mkdir()
    (run_dir / "final_summary_V.txt").write_text(
        "flow_cell_id=BBB\n", encoding="utf-8"
    )
    (run_dir / "sample_sheet_V.csv").write_text(
        "protocol_run_id,flow_cell_id,kit\nabc,BBB,\n", encoding="utf-8"
    )
    input_dir = run_dir / "fastq_pass"
    input_dir.mkdir()
    meta = discover_run_meta(input_dir)
    assert meta is not None
    assert meta.kit is None


# ---------------------------------------------------------------------------
# Excel __kuma_meta__ sheet tests
# ---------------------------------------------------------------------------


def test_excel_kuma_meta_sheet_present(tmp_path: Path) -> None:
    """write_excel always includes __kuma_meta__ sheet."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "with_meta.xlsx"
    run_dir = _make_run_dir(tmp_path)
    meta = discover_run_meta(run_dir)

    write_excel(
        verdict_records=[vr],
        replicate_results=[rr],
        output_path=out,
        ngs_run_meta=meta,
        kuma_version="1.2.3",
    )

    wb = openpyxl.load_workbook(out)
    assert "__kuma_meta__" in wb.sheetnames, "__kuma_meta__ sheet missing"


def test_excel_kuma_meta_sheet_values(tmp_path: Path) -> None:
    """__kuma_meta__ sheet contains expected flow_cell_id and kit rows."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "meta_values.xlsx"
    run_dir = _make_run_dir(tmp_path)
    meta = discover_run_meta(run_dir)
    assert meta is not None

    write_excel(
        verdict_records=[vr],
        replicate_results=[rr],
        output_path=out,
        ngs_run_meta=meta,
        kuma_version="test",
    )

    wb = openpyxl.load_workbook(out)
    ws = wb["__kuma_meta__"]
    kv = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert kv.get("flow_cell_id") == "FAW12345"
    assert kv.get("kit") == "SQK-LSK109"
    assert kv.get("instrument") == "PAX12345"
    assert kv.get("kuma_version") == "test"


def test_excel_kuma_meta_sheet_none_meta(tmp_path: Path) -> None:
    """write_excel with ngs_run_meta=None writes placeholder row."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    out = tmp_path / "no_meta.xlsx"
    write_excel(
        verdict_records=[vr],
        replicate_results=[],
        output_path=out,
        ngs_run_meta=None,
    )
    wb = openpyxl.load_workbook(out)
    assert "__kuma_meta__" in wb.sheetnames
    ws = wb["__kuma_meta__"]
    all_values = [row for row in ws.iter_rows(values_only=True) if any(v for v in row)]
    # Should have header + at least kuma_version + generated_at + placeholder rows
    assert len(all_values) >= 3


def test_excel_kuma_meta_carries_barcode_prefix_rule(tmp_path: Path) -> None:
    """How the barcode seeds were cut reaches the workbook the operator keeps.

    The sidecar log says it too, but nobody opens that file, and a seed cut at a
    guessed length names the wrong plate row while every other cell of the result
    looks ordinary. The row is written next to the run metadata because it is the
    same kind of fact: what this file was produced from.
    """
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    out = tmp_path / "prefix_rule.xlsx"
    note = "Barcode seeds on axis R were cut at a fixed length because ..."

    write_excel(
        verdict_records=[vr],
        replicate_results=[],
        output_path=out,
        ngs_run_meta=None,
        barcode_prefix_note=note,
    )

    ws = openpyxl.load_workbook(out)["__kuma_meta__"]
    kv = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert kv.get("barcode_prefix_rule") == note


def test_excel_kuma_meta_omits_barcode_prefix_rule_without_a_barcode_file(
    tmp_path: Path,
) -> None:
    """Consensus-dir mode reads no barcode workbook, so it claims nothing."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    out = tmp_path / "no_prefix_rule.xlsx"

    write_excel(
        verdict_records=[vr],
        replicate_results=[],
        output_path=out,
        ngs_run_meta=None,
    )

    ws = openpyxl.load_workbook(out)["__kuma_meta__"]
    keys = {row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert "barcode_prefix_rule" not in keys


def test_excel_legacy_sheets_preserved(tmp_path: Path) -> None:
    """Existing legacy sheets (NB01, Final, NGS Results, Final (matrix)) are unaffected."""
    vr = _make_verdict("NB01", "1_1", VerdictClass.PASS)
    rr = _make_replicate("V5F", "NB01", "1_1")
    out = tmp_path / "legacy.xlsx"
    write_excel(
        verdict_records=[vr],
        replicate_results=[rr],
        output_path=out,
        ngs_run_meta=None,
    )
    wb = openpyxl.load_workbook(out)
    for name in ("NB01", "Final", "NGS Results", "Final (matrix)", "__kuma_meta__"):
        assert name in wb.sheetnames, f"Sheet '{name}' missing"


# ---------------------------------------------------------------------------
# Janus mapping meta embedding (G3)
# ---------------------------------------------------------------------------


def test_janus_csv_header_is_line_one_with_meta(tmp_path: Path) -> None:
    """A ``ngs_run_meta`` still exports, but writes no comment line above the header.

    v0.16.6 dropped the ``# kuma_run_meta: ...`` line: it pushed the header (and
    every data row after it) down to line 2, which broke a plain
    ``csv.DictReader``/spreadsheet import, the way the lab actually opens this
    file.
    """
    rr = _make_replicate("V5F", "NB01", "1_1", size_kb=100.0)
    out = tmp_path / "janus_meta.csv"
    meta = NgsRunMeta(
        instrument="PAX12345",
        position="X3",
        flow_cell_id="FAW12345",
        sample_id="sample",
        kit="SQK-LSK109",
        started="2024-03-15T10:00:00Z",
        basecalling_enabled=True,
        raw_run_dir="/data/run_xyz",
    )
    export_mame_janus_csv([rr], out, ngs_run_meta=meta, settings=_LEGACY5)

    lines = out.read_text(encoding="utf-8").splitlines()
    assert lines[0].split(",")[0] == "name", f"Expected header row first, got: {lines[0]}"
    assert not lines[0].startswith("#")


def test_janus_csv_no_comment_when_meta_none(tmp_path: Path) -> None:
    """When ngs_run_meta=None, CSV starts with header row (no comment line)."""
    rr = _make_replicate("V5F", "NB01", "1_1", size_kb=100.0)
    out = tmp_path / "janus_no_comment.csv"
    export_mame_janus_csv([rr], out, ngs_run_meta=None, settings=_LEGACY5)

    with out.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        assert reader.fieldnames is not None
        assert reader.fieldnames[0] == "name", (
            f"Expected header row first, got: {reader.fieldnames}"
        )


def test_janus_csv_dict_reader_reads_rows_directly_no_skip_needed(tmp_path: Path) -> None:
    """No comment line to skip: DictReader reads the data row on the first pass."""
    rr = _make_replicate("V5F", "NB01", "1_1", size_kb=100.0)
    out = tmp_path / "janus_skip.csv"
    meta = NgsRunMeta(
        instrument=None, position=None, flow_cell_id="FC1",
        sample_id=None, kit=None, started=None,
        basecalling_enabled=None, raw_run_dir=None,
    )
    export_mame_janus_csv([rr], out, ngs_run_meta=meta, settings=_LEGACY5)

    with out.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = list(reader)
    assert len(rows) == 1
    assert rows[0]["name"] == "V5F"


def test_janus_xlsx_kuma_meta_sheet_present(tmp_path: Path) -> None:
    """Janus XLSX always contains __kuma_meta__ sheet."""
    rr = _make_replicate("V5F", "NB01", "1_1", size_kb=100.0)
    out = tmp_path / "janus.xlsx"
    export_mame_janus_xlsx([rr], out, ngs_run_meta=None, settings=_LEGACY5)
    wb = openpyxl.load_workbook(out)
    assert "__kuma_meta__" in wb.sheetnames, "__kuma_meta__ sheet missing from Janus XLSX"


def test_janus_xlsx_kuma_meta_values(tmp_path: Path) -> None:
    """Janus XLSX __kuma_meta__ sheet has flow_cell_id and kit rows."""
    rr = _make_replicate("V5F", "NB01", "1_1", size_kb=100.0)
    meta = NgsRunMeta(
        instrument="P2",
        position="Y1",
        flow_cell_id="FBW99",
        sample_id="samp",
        kit="SQK-RBK004",
        started="2024-06-01T08:00:00Z",
        basecalling_enabled=False,
        raw_run_dir="/data/run_abc",
    )
    out = tmp_path / "janus_kv.xlsx"
    export_mame_janus_xlsx([rr], out, ngs_run_meta=meta, kuma_version="2.0.0", settings=_LEGACY5)
    wb = openpyxl.load_workbook(out)
    ws = wb["__kuma_meta__"]
    kv = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    assert kv.get("flow_cell_id") == "FBW99"
    assert kv.get("kit") == "SQK-RBK004"
    assert kv.get("kuma_version") == "2.0.0"
    assert kv.get("basecalling_enabled") == "false"


def test_janus_xlsx_legacy_sheet_intact(tmp_path: Path) -> None:
    """Janus Mapping sheet (data) is still present after meta sheet added."""
    rr = _make_replicate("V5F", "NB01", "1_1", size_kb=100.0)
    out = tmp_path / "janus_legacy.xlsx"
    export_mame_janus_xlsx([rr], out, ngs_run_meta=None, settings=_LEGACY5)
    wb = openpyxl.load_workbook(out)
    assert "Janus Mapping" in wb.sheetnames
    assert "__kuma_meta__" in wb.sheetnames
