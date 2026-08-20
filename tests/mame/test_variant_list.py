"""Generic variant-list input for MAME (io/variant_list.py)."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from kuma_core.mame.io.variant_list import (
    inspect_variant_source,
    read_variant_source,
)
from kuma_core.mame.layout import WtPlacement, build_draft_layout
from kuma_core.mame.plate_geometry import seq_to_well


def _write_sheet(path, rows, sheet_title="Sheet1"):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = sheet_title
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def _write_kuro_export(path, mutants):
    """A KURO results xlsx: 10 fixed columns on an `expected_mutations` sheet."""
    wb = Workbook()
    wb.worksheets[0].title = "primers"
    ws = wb.create_sheet("expected_mutations")
    ws.append([
        "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
        "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
    ])
    for mutant_id, position, wt, mt, status in mutants:
        ws.append([mutant_id, position, wt, mt, "GTG", "TTT", "", "", "single", status])
    wb.save(path)
    return path


class TestPlainList:
    def test_rows_are_read_in_file_order(self, tmp_path):
        path = _write_sheet(
            tmp_path / "variants.xlsx",
            [["variant"], ["V5F"], ["K53N"], ["T10A"]],
        )

        result = read_variant_source(path)

        # 파일에 적힌 순서가 곧 plate 순서다. 정렬하지 않는다.
        assert [m.mutant_id for m in result.expected] == ["V5F", "K53N", "T10A"]

    def test_label_is_split_into_the_fields_verdict_comparison_uses(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["variant"], ["V5F"]])

        (mutation,) = read_variant_source(path).expected

        assert (mutation.wt_aa, mutation.position, mutation.mt_aa) == ("V", 5, "F")
        # MAME 가 읽지 않는 필드는 지어내지 않는다.
        assert mutation.wt_codon == ""
        assert mutation.group_id == ""

    def test_order_survives_into_the_draft_layout(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["K53N"], ["T10A"]],
        )

        result = read_variant_source(path)
        draft = build_draft_layout(result.expected, wt_ordinal=result.wt_ordinal)

        assert list(draft.layout.values())[:3] == ["V5F", "K53N", "T10A"]
        assert draft.is_complete

    @pytest.mark.parametrize("header", ["variant", "Mutation", "MUTANT_ID", "variants"])
    def test_common_header_names_are_recognised(self, tmp_path, header):
        path = _write_sheet(tmp_path / "v.xlsx", [[header], ["V5F"]])

        assert read_variant_source(path).variant_column == header

    def test_single_unnamed_column_is_used_without_asking(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["anything"], ["V5F"], ["K53N"]])

        assert len(read_variant_source(path).expected) == 2

    def test_ambiguous_columns_ask_instead_of_guessing(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["alpha", "beta"], ["V5F", "x"]],
        )

        with pytest.raises(ValueError, match="cannot tell which column"):
            read_variant_source(path)

    def test_named_column_wins(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["note", "target"], ["ignore", "V5F"], ["ignore", "K53N"]],
        )

        result = read_variant_source(path, variant_column="target")

        assert [m.mutant_id for m in result.expected] == ["V5F", "K53N"]

    def test_named_sheet_wins(self, tmp_path):
        wb = Workbook()
        first = wb.worksheets[0]
        first.title = "notes"
        first.append(["variant"])
        first.append(["A1G"])
        ws = wb.create_sheet("round2")
        ws.append(["variant"])
        ws.append(["V5F"])
        path = tmp_path / "v.xlsx"
        wb.save(path)

        result = read_variant_source(path, sheet="round2")

        assert [m.mutant_id for m in result.expected] == ["V5F"]
        assert result.sheet == "round2"

    def test_internal_blank_rows_are_refused_rather_than_skipped(self, tmp_path):
        """예전에는 조용히 흡수했다. 그게 K53N 을 한 칸 앞으로 당겼다."""
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], [None], [""], ["K53N"]],
        )

        with pytest.raises(ValueError, match="row 3.*row 4"):
            read_variant_source(path)

    def test_csv_is_accepted(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("variant\nV5F\nK53N\n", encoding="utf-8")

        result = read_variant_source(path)

        assert [m.mutant_id for m in result.expected] == ["V5F", "K53N"]
        assert result.sheet is None


class TestWildTypeRow:
    def test_explicit_wt_row_keeps_its_place_in_the_order(self, tmp_path):
        """WT 행은 버려지지 않고 배치 서수를 갖는다.

        예전에는 이 행을 ``continue`` 로 버려서, WT 뒤의 변이가 전부 한 칸씩
        앞으로 당겨졌다. 결과는 꽉 찬 정상 플레이트처럼 보였고 어디에도
        밀렸다는 표시가 없었다.
        """
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["WT"], ["K53N"]],
        )

        result = read_variant_source(path)

        assert result.wt_ordinal == 2
        assert result.has_explicit_wt is True
        assert [m.mutant_id for m in result.expected] == ["V5F", "K53N"]

    def test_the_wells_after_an_explicit_wt_do_not_move_up(self, tmp_path):
        """The regression, stated as wells rather than as an ordinal.

        The rule under test is the one that reads a wild-type ROW ORDINAL as a
        well, so it is pinned to ``AFTER_LAST_VARIANT`` rather than re-stated
        against the 2026-08-18 default. The default no longer answers this
        question at all: it puts the control in H12 whatever row it sat on.
        """
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["WT"], ["K53N"]],
        )

        result = read_variant_source(path)
        draft = build_draft_layout(
            result.expected,
            wt_ordinal=result.wt_ordinal,
            wt_placement=WtPlacement.AFTER_LAST_VARIANT,
        )

        assert list(draft.layout.items()) == [
            ("A1", "V5F"),
            ("B1", "WT"),
            ("C1", "K53N"),
        ]

    def test_plate_gets_one_wt_well_not_two(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["variant"], ["V5F"], ["WT"]])

        result = read_variant_source(path)
        draft = build_draft_layout(result.expected, wt_ordinal=result.wt_ordinal)

        assert list(draft.layout.values()).count("WT") == 1
        assert list(draft.layout.values()) == ["V5F", "WT"]

    def test_wt_is_appended_when_the_list_omits_it(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["variant"], ["V5F"]])

        result = read_variant_source(path)
        draft = build_draft_layout(result.expected, wt_ordinal=result.wt_ordinal)

        assert result.wt_ordinal is None
        assert result.has_explicit_wt is False
        assert list(draft.layout.values()) == ["V5F", "WT"]

    def test_two_wt_rows_name_both_of_them(self, tmp_path):
        """One plate carries one WT well, so the file has to say which row it is."""
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["WT"], ["K53N"], ["wt"]],
        )

        with pytest.raises(ValueError, match="rows 3 and 5"):
            read_variant_source(path)


class TestRowsThatCannotBePlaced:
    """행 순서가 곧 플레이트 순서이므로, 읽고 배치하지 못한 행은 거절된다."""

    def test_internal_blank_row_is_named_and_refused(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], [None], ["K53N"]],
        )

        with pytest.raises(ValueError, match="row 3"):
            read_variant_source(path)

    def test_trailing_blank_rows_are_openpyxl_phantoms_and_ignored(self, tmp_path):
        """마지막 값 뒤의 빈 행은 아무것도 밀지 않으므로 보고하지 않는다."""
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["K53N"], [None], [None]],
        )

        result = read_variant_source(path)

        assert [m.mutant_id for m in result.expected] == ["V5F", "K53N"]

    def test_kuro_status_filter_drops_are_named_and_refused(self, tmp_path):
        """두 판독기가 서로 다른 행 집합을 보던 자리.

        플레이트 순서 검사가 status 를 보지 않고 같은 시트를 따로 읽던 시절,
        조용히 걸러진 행 하나가 두 판독기의 웰 번호를 통째로 어긋나게 했다.
        그 검사는 이제 이 리더가 정한 배치를 그대로 쓴다.
        """
        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [
                ("V5F", 5, "V", "F", "DESIGNED"),
                ("Z9Z", 9, "Z", "Z", "FAILED"),
                ("K53N", 53, "K", "N", "DESIGNED"),
            ],
        )

        with pytest.raises(ValueError, match="row 3.*FAILED"):
            read_variant_source(path)


class TestRefusals:
    def test_unreadable_notation_names_the_row(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["not-a-mutation"]],
        )

        with pytest.raises(ValueError, match="row 3"):
            read_variant_source(path)

    def test_duplicate_variant_is_refused_with_both_rows(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["variant"], ["V5F"], ["K53N"], ["V5F"]],
        )

        with pytest.raises(ValueError, match="duplicate variant 'V5F'.*rows 2 and 4"):
            read_variant_source(path)

    def test_a_list_with_no_variants_is_refused(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["variant"]])

        with pytest.raises(ValueError, match="no variants found"):
            read_variant_source(path)

    def test_missing_named_column_lists_what_is_available(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["alpha", "beta"], ["V5F", "x"]])

        with pytest.raises(ValueError, match="alpha, beta"):
            read_variant_source(path, variant_column="gamma")

    def test_missing_file(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            read_variant_source(tmp_path / "absent.xlsx")


class TestKuroExportStillWorks:
    def test_kuro_export_is_routed_to_the_strict_reader(self, tmp_path):
        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [("V5F", 5, "V", "F", "DESIGNED"), ("K53N", 53, "K", "N", "DESIGNED")],
        )

        result = read_variant_source(path)

        assert [m.mutant_id for m in result.expected] == ["V5F", "K53N"]
        assert result.sheet == "expected_mutations"
        # 강한 판독 경로를 그대로 타므로 codon 이 살아 있다.
        assert result.expected[0].wt_codon == "GTG"

    def test_kuro_status_filter_itself_is_unchanged(self, tmp_path):
        """필터 규칙은 그대로다. 달라진 것은 조용하지 않다는 것뿐이다."""
        from kuma_core.mame.io.kuro_reader import read_expected_mutations_with_rows

        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [
                ("V5F", 5, "V", "F", "DESIGNED"),
                ("Z9Z", 9, "Z", "Z", "FAILED"),
                ("K53N", 53, "K", "N", "DESIGNED"),
            ],
        )

        read = read_expected_mutations_with_rows(path)

        assert [m.mutant_id for m in read.expected] == ["V5F", "K53N"]
        assert read.row_numbers == [2, 4]
        assert read.dropped_rows == [(3, "FAILED")]

    def test_kuro_export_wt_row_follows_the_same_rule_as_a_plain_list(self, tmp_path):
        """두 분기의 WT 규칙이 통일됐다.

        KURO 분기는 ``has_explicit_wt=False`` 를 하드코딩하고 있었다. 그래서
        자기 WT 행을 가진 KURO 시트에 대조군 웰이 하나 더 붙었고, 그 행 뒤의
        웰은 전부 한 칸 밀렸다.
        """
        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [
                ("V5F", 5, "V", "F", "DESIGNED"),
                ("WT", 0, "A", "A", "DESIGNED"),
                ("K53N", 53, "K", "N", "DESIGNED"),
            ],
        )

        result = read_variant_source(path)
        # 서수 규칙 자체를 못박는 테스트다. 기본값이 바뀌어도 이 규칙은
        # ``AFTER_LAST_VARIANT`` 로 살아 있으므로 선택지를 명시한다.
        draft = build_draft_layout(
            result.expected,
            wt_ordinal=result.wt_ordinal,
            wt_placement=WtPlacement.AFTER_LAST_VARIANT,
        )

        assert result.wt_ordinal == 2
        assert list(draft.layout.items()) == [
            ("A1", "V5F"),
            ("B1", "WT"),
            ("C1", "K53N"),
        ]

    def test_kuro_export_without_a_wt_row_reports_none(self, tmp_path):
        path = _write_kuro_export(tmp_path / "kuro.xlsx", [("V5F", 5, "V", "F", "DESIGNED")])

        assert read_variant_source(path).wt_ordinal is None


class TestInspect:
    def test_reports_sheets_and_headers_for_a_plain_workbook(self, tmp_path):
        wb = Workbook()
        first = wb.worksheets[0]
        first.title = "first"
        first.append(["variant", "note"])
        wb.create_sheet("second").append(["x"])
        path = tmp_path / "v.xlsx"
        wb.save(path)

        info = inspect_variant_source(path)

        assert info.is_kuro_export is False
        assert info.sheets == ["first", "second"]
        assert info.headers["first"] == ["variant", "note"]
        assert info.suggested_column == "variant"

    def test_flags_a_kuro_export_so_no_mapping_is_asked_for(self, tmp_path):
        path = _write_kuro_export(tmp_path / "kuro.xlsx", [("V5F", 5, "V", "F", "DESIGNED")])

        info = inspect_variant_source(path)

        assert info.is_kuro_export is True
        assert info.suggested_column is None

    def test_reports_no_suggestion_when_columns_are_ambiguous(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["alpha", "beta"], ["V5F", "x"]])

        assert inspect_variant_source(path).suggested_column is None


class TestExplicitSheetBeatsRecognition:
    """One workbook can hold the strict sheet next to the sheet built on the bench."""

    def _mixed_workbook(self, path):
        """`expected_mutations` and a plate sheet carrying the same set, reordered."""
        wb = Workbook()
        plate = wb.worksheets[0]
        plate.title = "Fwd List"
        plate.append(["Well", "Primer Name", "Mutation"])
        for well, mutant in (("A1", "S11I"), ("B1", "S22T"), ("C1", "K53I")):
            plate.append([well, f"{mutant}_F", mutant])
        strict = wb.create_sheet("expected_mutations")
        strict.append([
            "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
            "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
        ])
        # Same three mutants, design order rather than plate order.
        for mutant_id, position, wt, mt in (
            ("K53I", 53, "K", "I"), ("S11I", 11, "S", "I"), ("S22T", 22, "S", "T"),
        ):
            strict.append(
                [mutant_id, position, wt, mt, "GTG", "TTT", "", "", "single", "DESIGNED"]
            )
        wb.save(path)
        return path

    def test_naming_another_sheet_reads_that_sheet(self, tmp_path):
        path = self._mixed_workbook(tmp_path / "platemap.xlsx")

        result = read_variant_source(path, sheet="Fwd List", variant_column="Mutation")

        # 플레이트 시트를 고르면 그 순서가 곧 well 순서다.
        assert [m.mutant_id for m in result.expected] == ["S11I", "S22T", "K53I"]
        # column-major: seq 1..3 -> A1,B1,C1. 대조군은 기본값대로 마지막 웰이다.
        # 이 테스트가 지키는 것은 시트 선택이지 대조군 자리가 아니므로,
        # 선택지를 명시하지 않고 새 기본값의 결과를 그대로 확인한다.
        layout = build_draft_layout(result.expected).layout
        assert [layout[w] for w in ("A1", "B1", "C1")] == ["S11I", "S22T", "K53I"]
        assert layout["H12"] == "WT"
        assert "D1" not in layout

    def test_naming_no_sheet_keeps_the_strict_reader(self, tmp_path):
        path = self._mixed_workbook(tmp_path / "platemap.xlsx")

        result = read_variant_source(path)

        assert result.sheet == "expected_mutations"
        assert [m.mutant_id for m in result.expected] == ["K53I", "S11I", "S22T"]

    def test_naming_the_strict_sheet_keeps_the_strict_reader(self, tmp_path):
        path = self._mixed_workbook(tmp_path / "platemap.xlsx")

        result = read_variant_source(path, sheet="expected_mutations")

        assert result.variant_column == "mutant_id"
        assert [m.mutant_id for m in result.expected] == ["K53I", "S11I", "S22T"]

    def test_headers_are_offered_for_every_sheet_of_a_kuro_export(self, tmp_path):
        path = self._mixed_workbook(tmp_path / "platemap.xlsx")

        info = inspect_variant_source(path)

        # 강판독이 기본이라는 사실은 유지하되, 다른 시트를 고를 재료는 준다.
        assert info.is_kuro_export is True
        assert "Fwd List" in info.sheets
        assert info.headers["Fwd List"] == ["Well", "Primer Name", "Mutation"]


class TestControlStatusWildType:
    """KURO writes its WT row as `status=control`, which is not a design.

    The status filter removed it, the reader reported that removal as a row it
    could not place, and the whole workbook was refused. Every KURO export
    carrying its own control row was unreadable, the two shipped ones included.
    """

    def test_a_control_status_wt_row_does_not_refuse_the_file(self, tmp_path):
        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [
                ("M001", 5, "V", "F", "designed"),
                ("WT", 0, "-", "-", "control"),
                ("M002", 53, "K", "N", "designed"),
            ],
        )

        result = read_variant_source(path)

        assert [m.mutant_id for m in result.expected] == ["M001", "M002"]

    def test_the_control_row_occupies_the_well_its_row_number_names(self, tmp_path):
        """대조군 행은 자기 순서의 웰을 차지한다. 버리면 뒤가 한 칸씩 당겨진다."""
        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [
                ("M001", 5, "V", "F", "designed"),
                ("WT", 0, "-", "-", "control"),
                ("M002", 53, "K", "N", "designed"),
            ],
        )

        result = read_variant_source(path)
        # 이름 그대로 "행 번호가 가리키는 웰" 이 대상이므로 서수 선택지를
        # 명시한다.
        draft = build_draft_layout(
            result.expected,
            wt_ordinal=result.wt_ordinal,
            wt_placement=WtPlacement.AFTER_LAST_VARIANT,
        )

        assert result.wt_ordinal == 2
        assert list(draft.layout.items()) == [
            ("A1", "M001"),
            ("B1", "WT"),
            ("C1", "M002"),
        ]

    def test_a_mutant_row_is_still_judged_by_status_alone(self, tmp_path):
        """판정 기준은 라벨이지 status 가 아니다. 진짜 mutant 는 그대로 걸린다."""
        path = _write_kuro_export(
            tmp_path / "kuro.xlsx",
            [
                ("M001", 5, "V", "F", "designed"),
                ("M002", 9, "Z", "Z", "control"),
                ("M003", 53, "K", "N", "designed"),
            ],
        )

        with pytest.raises(ValueError, match="row 3.*control"):
            read_variant_source(path)

    # 배포·데모 워크북 자체를 실제 소비자에 태우는 검사는
    # ``test_shipped_plate_assets.py`` 로 옮겼다. 그쪽은 대상을 glob 으로 모으므로
    # 새 자산이 추가돼도 자동으로 걸리고, 읽기뿐 아니라 배치와 플레이트 시트
    # 대조까지 같이 본다. 여기 있던 점유자 수 단언은 그 모듈에 그대로 남아 있다.


class TestHeaderlessList:
    """헤더 없이 변이만 적은 파일에서 첫 변이가 컬럼명으로 먹히던 자리.

    예외도 경고도 없이 첫 변이가 사라졌고, 남은 변이가 전부 한 칸씩 앞으로
    당겨졌다. 이 모듈이 막으려는 바로 그 밀림이 이 경로만 무증상이었다.
    """

    def test_the_first_variant_is_not_eaten_as_a_column_name(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("S65T\nY66H\n", encoding="utf-8")

        result = read_variant_source(path)
        draft = build_draft_layout(result.expected, wt_ordinal=result.wt_ordinal)

        assert [m.mutant_id for m in result.expected] == ["S65T", "Y66H"]
        assert draft.layout["A1"] == "S65T"
        assert draft.layout["B1"] == "Y66H"

    def test_the_absence_of_a_header_is_reported_rather_than_a_variant(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("S65T\nY66H\n", encoding="utf-8")

        assert read_variant_source(path).variant_column == "(no header)"

    def test_row_numbers_start_at_one_because_no_row_is_a_header(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("S65T\nnot-a-mutation\n", encoding="utf-8")

        with pytest.raises(ValueError, match="row 2"):
            read_variant_source(path)

    def test_a_named_column_does_not_reopen_the_hole(self, tmp_path):
        """UI 는 inspect 의 제안을 그대로 넘긴다. 인자가 있어도 헤더리스다."""
        path = tmp_path / "v.csv"
        path.write_text("S65T\nY66H\n", encoding="utf-8")

        result = read_variant_source(path, variant_column="S65T")

        assert [m.mutant_id for m in result.expected] == ["S65T", "Y66H"]
        assert result.variant_column == "(no header)"

    def test_a_wt_first_row_is_recognised_too(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("WT\nS65T\n", encoding="utf-8")

        result = read_variant_source(path)

        assert result.wt_ordinal == 1
        assert [m.mutant_id for m in result.expected] == ["S65T"]

    def test_an_xlsx_without_a_header_reads_the_same_way(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["S65T"], ["Y66H"]])

        result = read_variant_source(path)

        assert [m.mutant_id for m in result.expected] == ["S65T", "Y66H"]

    def test_a_lower_case_first_variant_is_not_eaten_either(self, tmp_path):
        """소문자 표기도 데이터다. 컬럼명으로 먹히면 조용히 사라진다.

        ``parse_mutation_notation`` 은 대문자만 받으므로, 이 행은 헤더리스로
        인식된 뒤 파서에 넘어가 행 번호와 함께 큰 소리로 거절된다. 그것이
        아무 말 없이 한 칸 밀린 플레이트보다 낫다.
        """
        path = tmp_path / "v.csv"
        path.write_text("s65t\nY66H\n", encoding="utf-8")

        with pytest.raises(ValueError, match="row 1"):
            read_variant_source(path)

    def test_a_lower_case_first_variant_is_not_offered_as_a_column(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("s65t\nY66H\n", encoding="utf-8")

        assert inspect_variant_source(path).suggested_column is None

    def test_a_named_header_still_behaves_exactly_as_before(self, tmp_path):
        """회귀 방지. 헤더가 있으면 헤더고, 데이터 행은 2 부터다."""
        path = tmp_path / "v.csv"
        path.write_text("variant\nS65T\nnot-a-mutation\n", encoding="utf-8")

        with pytest.raises(ValueError, match="row 3"):
            read_variant_source(path)

    def test_a_header_that_is_a_variant_is_not_offered_as_a_column(self, tmp_path):
        """제안했다면 UI 가 그 값을 되넘겨 첫 변이를 잃게 했을 것이다."""
        path = tmp_path / "v.csv"
        path.write_text("S65T\nY66H\n", encoding="utf-8")

        assert inspect_variant_source(path).suggested_column is None

    def test_a_lone_ordinary_header_is_still_suggested(self, tmp_path):
        path = _write_sheet(tmp_path / "v.xlsx", [["anything"], ["V5F"]])

        assert inspect_variant_source(path).suggested_column == "anything"


class TestMultiColumnFirstRowIsRefused:
    """다열 파일에서는 첫 행이 헤더인지 데이터인지 파일이 말해주지 않는다.

    단열은 비어있지 않은 셀이 하나뿐이라 추론이 결정적이지만, 다열에서는
    나머지 열이 헤더일 수도 값일 수도 있다. 헤더로 읽으면 첫 변이를 잃고
    데이터로 읽으면 나머지 열의 이름을 지어내는 셈이라 어느 쪽도 고를 수
    없다. 그래서 추측하지 않고 거절한다.
    """

    def test_a_named_column_whose_header_is_a_variant_is_refused(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("S65T,plate1\nY66H,plate1\n", encoding="utf-8")

        with pytest.raises(ValueError, match="'S65T' in v.csv is itself a variant"):
            read_variant_source(path, variant_column="S65T")

    def test_the_refusal_says_what_to_do_about_it(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("S65T,plate1\nY66H,plate1\n", encoding="utf-8")

        with pytest.raises(ValueError, match="Add a header row"):
            read_variant_source(path, variant_column="S65T")

    def test_a_wt_label_header_is_refused_the_same_way(self, tmp_path):
        path = tmp_path / "v.csv"
        path.write_text("WT,plate1\nY66H,plate1\n", encoding="utf-8")

        with pytest.raises(ValueError, match="'WT' in v.csv is itself a variant"):
            read_variant_source(path, variant_column="WT")

    def test_the_variant_is_not_quietly_dropped_instead(self, tmp_path):
        """거절 전에는 이 입력이 `['Y66H']` 로 조용히 통과했다."""
        path = tmp_path / "v.csv"
        path.write_text("S65T,plate1\nY66H,plate1\n", encoding="utf-8")

        with pytest.raises(ValueError):
            read_variant_source(path, variant_column="S65T")

    def test_an_xlsx_is_refused_the_same_way(self, tmp_path):
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [["S65T", "plate1"], ["Y66H", "plate1"]],
        )

        with pytest.raises(ValueError, match="itself a variant"):
            read_variant_source(path, variant_column="S65T")

    def test_a_single_column_headerless_file_still_reads(self, tmp_path):
        """경계. 단열은 추론이 결정적이므로 거절 대상이 아니다."""
        path = tmp_path / "v.csv"
        path.write_text("S65T\nY66H\n", encoding="utf-8")

        result = read_variant_source(path, variant_column="S65T")

        assert [m.mutant_id for m in result.expected] == ["S65T", "Y66H"]

    @pytest.mark.parametrize("header", ["variant", "mutation", "mutant_id"])
    def test_ordinary_header_names_are_untouched(self, tmp_path, header):
        """정상 헤더명은 변이로 파싱되지 않으므로 이 규칙에 걸리지 않는다."""
        path = _write_sheet(
            tmp_path / "v.xlsx",
            [[header, "note"], ["S65T", "x"], ["Y66H", "x"]],
        )

        result = read_variant_source(path, variant_column=header)

        assert [m.mutant_id for m in result.expected] == ["S65T", "Y66H"]

    def test_naming_no_column_keeps_the_old_message(self, tmp_path):
        """컬럼을 안 넘긴 다열 파일의 거절 문구는 그대로다."""
        path = tmp_path / "v.csv"
        path.write_text("S65T,plate1\nY66H,plate1\n", encoding="utf-8")

        with pytest.raises(ValueError, match="cannot tell which column"):
            read_variant_source(path)


def _variant(index: int) -> str:
    """A distinct, parseable substitution label for filler rows."""
    return f"G{index + 2}A"


class TestStatedWells:
    """A ``Well`` column is the placement, and the reader computes nothing.

    Every test here asks the same question in a different shape: does the file
    decide, or does the reader. The row-order path is a separate class because
    it answers that question the other way and has to keep doing so.
    """

    def test_the_wells_the_file_names_are_the_wells_the_plate_uses(self, tmp_path):
        """Including a plate that starts nowhere near A1."""
        path = _write_sheet(
            tmp_path / "wells.xlsx",
            [["Well", "Variant"], ["C3", "V5F"], ["A1", "K53N"], ["H12", "T10A"]],
        )

        result = read_variant_source(path)
        draft = build_draft_layout(
            result.expected, wells=result.wells, wt_well=result.wt_well
        )

        # Row order says nothing here: the placement is read off the Well column
        # and then put into plate order, so A1 leads whatever line it sat on.
        assert draft.layout == {"A1": "K53N", "C3": "V5F", "H12": "T10A"}
        assert list(draft.layout) == ["A1", "C3", "H12"]

    def test_a_well_with_no_variant_is_a_well_this_campaign_did_not_use(
        self, tmp_path
    ):
        """The rule that makes a partial plate expressible.

        On the row-order path an empty variant cell is refused, because the row
        is read and every mutant after it moves one well up. Nothing moves here,
        so the row states that the well is empty and the layout leaves it out.
        """
        path = _write_sheet(
            tmp_path / "partial.xlsx",
            [["Well", "Variant"], ["A1", "V5F"], ["B1", None], ["C1", "K53N"]],
        )

        result = read_variant_source(path)
        draft = build_draft_layout(
            result.expected, wells=result.wells, wt_well=result.wt_well
        )

        assert result.dropped_rows == []
        assert draft.layout == {"A1": "V5F", "C1": "K53N"}
        assert "B1" not in draft.layout

    def test_ninety_six_variants_and_no_control_fill_the_plate(self, tmp_path):
        """Capacity is 96 on this path, because a control is not assumed.

        The 95 ceiling exists to keep a well free for a control the row-order
        path appends whether the file mentions one or not. A file that states
        its wells states whether there is a control, so a plate of 96 mutants is
        a plate the file is entitled to describe.
        """
        wells = [seq_to_well(seq) for seq in range(1, 97)]
        rows = [["Well", "Variant"]]
        rows += [[well, _variant(i)] for i, well in enumerate(wells)]
        path = _write_sheet(tmp_path / "full.xlsx", rows)

        result = read_variant_source(path)
        draft = build_draft_layout(
            result.expected, wells=result.wells, wt_well=result.wt_well
        )

        assert result.wt_well is None
        assert result.has_explicit_wt is False
        assert len(draft.layout) == 96
        assert draft.dropped_mutant_ids == []
        assert draft.has_wt_well is False
        assert draft.wt_well is None

    def test_the_same_well_twice_is_refused_by_row_and_value(self, tmp_path):
        path = _write_sheet(
            tmp_path / "dup.xlsx",
            [["Well", "Variant"], ["A1", "V5F"], ["B1", "K53N"], ["A1", "T10A"]],
        )

        with pytest.raises(ValueError) as excinfo:
            read_variant_source(path)

        message = str(excinfo.value)
        assert "A1" in message
        # Both rows, so the operator can see which two disagree.
        assert "2" in message and "4" in message

    def test_a_coordinate_off_the_plate_is_refused_by_row_and_value(self, tmp_path):
        for raw in ("I13", "A0", "Z9"):
            path = _write_sheet(
                tmp_path / f"off-{raw}.xlsx",
                [["Well", "Variant"], ["A1", "V5F"], [raw, "K53N"]],
            )

            with pytest.raises(ValueError) as excinfo:
                read_variant_source(path)

            message = str(excinfo.value)
            assert raw in message, message
            assert "row 3" in message, message

    def test_a_variant_with_an_empty_well_cell_is_refused(self, tmp_path):
        """The mirror of the empty-variant rule, and it is not symmetric.

        An empty variant says the well is unused. An empty well says nothing at
        all about where the variant goes, and there is no row order to fall back
        on once the file has claimed the addresses.
        """
        path = _write_sheet(
            tmp_path / "no-well.xlsx",
            [["Well", "Variant"], ["A1", "V5F"], [None, "K53N"]],
        )

        with pytest.raises(ValueError) as excinfo:
            read_variant_source(path)

        message = str(excinfo.value)
        assert "row 3" in message
        assert "K53N" in message

    def test_a01_and_a1_are_the_same_well(self, tmp_path):
        path = _write_sheet(
            tmp_path / "padded.xlsx",
            [["Well", "Variant"], ["A01", "V5F"], ["B01", "K53N"]],
        )

        result = read_variant_source(path)
        draft = build_draft_layout(
            result.expected, wells=result.wells, wt_well=result.wt_well
        )

        assert result.wells == ["A1", "B1"]
        assert draft.layout == {"A1": "V5F", "B1": "K53N"}

    def test_a01_collides_with_a1_rather_than_naming_a_second_well(self, tmp_path):
        path = _write_sheet(
            tmp_path / "padded-dup.xlsx",
            [["Well", "Variant"], ["A1", "V5F"], ["A01", "K53N"]],
        )

        with pytest.raises(ValueError, match="A1"):
            read_variant_source(path)

    def test_the_control_sits_in_the_well_the_file_gave_it(self, tmp_path):
        path = _write_sheet(
            tmp_path / "wt.xlsx",
            [["Well", "Variant"], ["A1", "V5F"], ["D7", "WT"], ["B1", "K53N"]],
        )

        result = read_variant_source(path)
        draft = build_draft_layout(
            result.expected, wells=result.wells, wt_well=result.wt_well
        )

        assert result.wt_well == "D7"
        assert result.has_explicit_wt is True
        # No ordinal is recorded: the file answered the question an ordinal is a
        # proxy for, so recording both would leave two statements about one well.
        assert result.wt_ordinal is None
        assert draft.layout == {"A1": "V5F", "B1": "K53N", "D7": "WT"}
        assert draft.wt_well == "D7"


class TestControlWellPlacementWithoutAWellColumn:
    """The row-order path, where where the control goes is a decision."""

    def _forty(self, tmp_path, wt: bool = True):
        rows = [["variant"]] + [[_variant(i)] for i in range(40)]
        if wt:
            rows.append(["WT"])
        return _write_sheet(tmp_path / "forty.xlsx", rows)

    def test_the_default_puts_the_control_in_the_last_well(self, tmp_path):
        """The regression this whole change exists for.

        Forty variants with a wild-type row on line 41 used to put the control
        in A6, because the row ordinal was read as a well. The bench pipettes it
        into H12. MAME scored A6 as the control and did not score H12 at all.
        """
        result = read_variant_source(self._forty(tmp_path))
        draft = build_draft_layout(result.expected, wt_ordinal=result.wt_ordinal)

        assert result.wt_ordinal == 41
        assert draft.wt_well == "H12"
        assert draft.layout["H12"] == "WT"
        # The variants are untouched: they still fill from A1 in file order.
        assert draft.layout["A1"] == _variant(0)
        assert draft.layout["H5"] == _variant(39)
        assert "A6" not in draft.layout

    def test_naming_the_ordinal_placement_keeps_the_old_answer(self, tmp_path):
        """The old rule is a choice now, not a default, and it still works."""
        result = read_variant_source(self._forty(tmp_path))
        draft = build_draft_layout(
            result.expected,
            wt_ordinal=result.wt_ordinal,
            wt_placement=WtPlacement.AFTER_LAST_VARIANT,
        )

        assert draft.wt_well == "A6"
        assert draft.layout["A6"] == "WT"

    def test_none_leaves_the_plate_without_a_control(self, tmp_path):
        """A campaign that ran no control is a plate, not an error."""
        result = read_variant_source(self._forty(tmp_path))
        draft = build_draft_layout(
            result.expected,
            wt_ordinal=result.wt_ordinal,
            wt_placement=WtPlacement.NONE,
        )

        assert draft.wt_well is None
        assert draft.has_wt_well is False
        assert "WT" not in draft.layout.values()
        assert len(draft.layout) == 40
        # The variants do not move: the control policy decides one well, not the
        # order of the other ninety-five.
        assert draft.layout["A1"] == _variant(0)
        assert draft.layout["H5"] == _variant(39)

    def test_ninety_six_variants_are_still_refused_without_a_well_column(
        self, tmp_path
    ):
        """The ceiling does not depend on where the control sits.

        It holds under NONE too. Whether a list fits must not change with a
        setting, or the same file would be a plate under one policy and not
        under another.
        """
        rows = [["variant"]] + [[_variant(i)] for i in range(96)]
        path = _write_sheet(tmp_path / "ninetysix.xlsx", rows)

        result = read_variant_source(path)

        for placement in WtPlacement:
            draft = build_draft_layout(result.expected, wt_placement=placement)
            assert draft.layout == {}, placement
            assert draft.dropped_mutant_ids == [_variant(95)], placement
            assert draft.is_complete is False, placement
