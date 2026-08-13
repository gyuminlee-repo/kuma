"""Cross-check the three KURO Echo writers against each other.

The CSV export, the XLSX "Echo mapping file" sheet and the sidecar preview
describe the same physical dispense, and each used to derive its own source
wells. Only the CSV read both placement parameters: the XLSX worklist sheet
read neither and the preview only ``mapping_range``. So one ``export_all`` wrote
a csv and an xlsx naming different wells for the same primer, and the preview
rendered above the quadrant selector agreed with neither, which is what the
operator checks before loading the csv onto the robot.

These tests deliberately compare the three outputs *to each other* rather than
to a literal table. Pinning literals only catches a writer drifting away from
the literal; a change that updates the literal and two writers still ships a
third writer out of step. Format pinning against the lab workbook lives in
tests/test_plate_mapper_reference_format.py and is not repeated here.
"""

from __future__ import annotations

import csv

import openpyxl
import pytest
from sidecar_kuro.handlers.export import _build_echo_preview_rows

from kuma_core.kuro.plate_mapper import (
    ECHO_DEVICE_HEADER,
    PlateMapping,
    echo_row_values,
    export_echo_mapping_csv,
    export_echo_mapping_xlsx,
)

TRANSFER_VOL = 100

# The three placements a caller can ask for. The no-parameter case alone proves
# nothing: the writers agreed on it before this was fixed, because the default
# is what each of them hard-coded. The other two are where they split.
PLACEMENTS = [
    pytest.param(None, None, id="no-placement-parameters"),
    pytest.param("B2", None, id="quadrant"),
    pytest.param(None, ("A", "P"), id="mapping-range"),
]


@pytest.fixture
def shared_rev_mappings():
    """Four mutations over two reverse primers, so shared-primer expansion runs."""
    fwd = [
        PlateMapping("A1", "A1G_F", "AAAATTTTCCCCGGGGAAAA", "forward", "A1G"),
        PlateMapping("B1", "C2T_F", "AAAATTTTCCCCGGGGCCCC", "forward", "C2T"),
        PlateMapping("C1", "G3A_F", "AAAATTTTCCCCGGGGGGGG", "forward", "G3A"),
        PlateMapping("H12", "T4C_F", "AAAATTTTCCCCGGGGTTTT", "forward", "T4C"),
    ]
    rev = [
        PlateMapping("A1", "A1G_R", "SHAREDTTTTCCCCGGGGAA", "reverse", "A1G"),
        PlateMapping("B1", "G3A_R", "REV2TTTTCCCCGGGGAAAA", "reverse", "G3A"),
    ]
    groups = {
        "SHAREDTTTTCCCCGGGGAA": ["A1G", "C2T"],
        "REV2TTTTCCCCGGGGAAAA": ["G3A", "T4C"],
    }
    return fwd, rev, groups


def _csv_table(path):
    with open(path, encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    return rows[0], rows[1:]


def _xlsx_table(path):
    ws = openpyxl.load_workbook(path)["Echo mapping file"]
    rows = [[c.value for c in r] for r in ws.iter_rows()]
    return rows[0], rows[1:]


def _layout_names(path):
    """Every primer name drawn on the xlsx layout sheet, keyed by 384 well."""
    ws = openpyxl.load_workbook(path)["layout"]
    found = {}
    for ri, row_letter in enumerate("ABCDEFGHIJKLMNOP"):
        for c in range(1, 25):
            value = ws.cell(row=5 + ri, column=c + 2).value
            if value:
                found[f"{row_letter}{c}"] = value
    return found


def _as_text(values):
    """Normalise one row so the writers are comparable.

    CSV carries every cell as text while XLSX round-trips numbers through Excel.
    Comparing numbers as numbers and everything else as text keeps the assertion
    about the values the writers chose, not about how each container stores them.
    """
    out = []
    for v in values:
        if v is None:
            out.append("")
            continue
        try:
            out.append(float(v))
        except (TypeError, ValueError):
            out.append(str(v))
    return out


def _write_both(fwd, rev, groups, tmp_path, quadrant, mapping_range):
    csv_path = tmp_path / "echo.csv"
    xlsx_path = tmp_path / "echo.xlsx"
    export_echo_mapping_csv(
        fwd, rev, csv_path, transfer_vol=TRANSFER_VOL, rev_groups=groups,
        mapping_range=mapping_range, quadrant=quadrant,
    )
    export_echo_mapping_xlsx(
        fwd, rev, xlsx_path, transfer_vol=TRANSFER_VOL, rev_groups=groups,
        mapping_range=mapping_range, quadrant=quadrant,
    )
    preview = _build_echo_preview_rows(
        fwd, rev, TRANSFER_VOL, groups,
        mapping_range=mapping_range, quadrant=quadrant,
    )
    return csv_path, xlsx_path, preview


class TestWritersAgree:
    @pytest.mark.parametrize(("quadrant", "mapping_range"), PLACEMENTS)
    def test_csv_xlsx_and_preview_write_the_same_rows(
        self, shared_rev_mappings, tmp_path, quadrant, mapping_range
    ):
        fwd, rev, groups = shared_rev_mappings
        csv_path, xlsx_path, preview = _write_both(
            fwd, rev, groups, tmp_path, quadrant, mapping_range
        )

        csv_header, csv_raw = _csv_table(csv_path)
        xlsx_header, xlsx_raw = _xlsx_table(xlsx_path)

        assert csv_header == ECHO_DEVICE_HEADER
        assert _as_text(xlsx_header) == _as_text(ECHO_DEVICE_HEADER)

        csv_rows = [_as_text(r) for r in csv_raw]
        assert csv_rows, "the fixture must produce rows for the comparison to mean anything"
        assert [_as_text(r) for r in xlsx_raw] == csv_rows
        assert [_as_text(echo_row_values(r)) for r in preview] == csv_rows

    @pytest.mark.parametrize(("quadrant", "mapping_range"), PLACEMENTS)
    def test_every_writer_names_the_same_source_wells(
        self, shared_rev_mappings, tmp_path, quadrant, mapping_range
    ):
        """Stated apart from the row comparison because this is the column that
        decides what the robot aspirates. A row list matching on six columns and
        differing on this one would still ruin the plate.
        """
        fwd, rev, groups = shared_rev_mappings
        csv_path, xlsx_path, preview = _write_both(
            fwd, rev, groups, tmp_path, quadrant, mapping_range
        )

        name_col = ECHO_DEVICE_HEADER.index("Source Well Name")
        well_col = ECHO_DEVICE_HEADER.index("Source Well")
        _, csv_raw = _csv_table(csv_path)
        _, xlsx_raw = _xlsx_table(xlsx_path)

        from_csv = {(r[name_col], r[well_col]) for r in csv_raw}
        from_xlsx = {(r[name_col], r[well_col]) for r in xlsx_raw}
        from_preview = {(r["source_well_name"], r["source_well"]) for r in preview}

        assert from_xlsx == from_csv
        assert from_preview == from_csv

    def test_shared_reverse_primer_expands_identically_everywhere(
        self, shared_rev_mappings, tmp_path
    ):
        """Every forward mutation gets a reverse row, in all three writers."""
        fwd, rev, groups = shared_rev_mappings
        csv_path, xlsx_path, preview = _write_both(
            fwd, rev, groups, tmp_path, "B2", None
        )

        expected = 2 * len(fwd)
        assert len(_csv_table(csv_path)[1]) == expected
        assert len(_xlsx_table(xlsx_path)[1]) == expected
        assert len(preview) == expected

    def test_preview_carries_every_instrument_column_plus_its_own_fields(
        self, shared_rev_mappings
    ):
        fwd, rev, groups = shared_rev_mappings
        preview = _build_echo_preview_rows(fwd, rev, TRANSFER_VOL, groups)

        for row in preview:
            # echo_row_values raises KeyError if an instrument column is absent.
            assert len(echo_row_values(row)) == len(ECHO_DEVICE_HEADER)
            assert row["mutation"]


class TestSpentQuadrantRefusal:
    def test_the_xlsx_export_refuses_before_writing(
        self, shared_rev_mappings, tmp_path
    ):
        fwd, rev, groups = shared_rev_mappings
        xlsx_path = tmp_path / "echo.xlsx"

        with pytest.raises(ValueError, match="already used"):
            export_echo_mapping_xlsx(
                fwd, rev, xlsx_path, transfer_vol=TRANSFER_VOL, rev_groups=groups,
                quadrant="A1", used_quadrants=["A1"],
            )

        assert not xlsx_path.exists()

    def test_the_preview_refuses_too(self, shared_rev_mappings):
        """So the operator is never shown a plate that cannot be dispensed onto
        and then told about it only at export time.
        """
        fwd, rev, groups = shared_rev_mappings

        with pytest.raises(ValueError, match="already used"):
            _build_echo_preview_rows(
                fwd, rev, TRANSFER_VOL, groups,
                quadrant="A1", used_quadrants=["A1"],
            )


class TestLayoutSheetIsUntouched:
    def test_the_layout_sheet_keeps_the_row_doubled_view(
        self, shared_rev_mappings, tmp_path
    ):
        """The layout sheet draws the default plate, not this transfer list, and
        sidecar_kuro/models.py says so. A quadrant must not move it.
        """
        fwd, rev, groups = shared_rev_mappings
        plain = tmp_path / "plain.xlsx"
        shifted = tmp_path / "shifted.xlsx"
        export_echo_mapping_xlsx(
            fwd, rev, plain, transfer_vol=TRANSFER_VOL, rev_groups=groups,
        )
        export_echo_mapping_xlsx(
            fwd, rev, shifted, transfer_vol=TRANSFER_VOL, rev_groups=groups,
            quadrant="B2",
        )

        assert _layout_names(shifted) == _layout_names(plain)
