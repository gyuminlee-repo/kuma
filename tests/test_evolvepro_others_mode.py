"""Tests for EVOLVEpro Others mode: xlsx, column override, asc sort."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest
import xlwt


def _write_csv(rows: list[list[str]], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def _write_xlsx(sheets: dict[str, list[list[str]]], path: Path) -> None:
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))


def _write_xls(sheets: dict[str, list[list[str]]], path: Path) -> None:
    wb = xlwt.Workbook()
    for sheet_name, rows in sheets.items():
        ws = wb.add_sheet(sheet_name)
        for ridx, row in enumerate(rows):
            for cidx, value in enumerate(row):
                ws.write(ridx, cidx, value)
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Task 4 tests: _load_evolvepro_rows
# ---------------------------------------------------------------------------

def test_load_rows_desc_baseline(tmp_path):
    """desc mode with alias columns matches original CSV behavior."""
    from kuma_core.kuro.evolvepro import _load_evolvepro_rows

    csv_file = tmp_path / "ep.csv"
    _write_csv([
        ["variant", "y_pred"],
        ["A1V", "0.9"],
        ["B2C", "0.7"],
        ["D3E", "0.8"],
    ], csv_file)

    rows = _load_evolvepro_rows(str(csv_file))

    if len(rows) != 3:
        pytest.fail(f"Expected 3 rows, got {len(rows)}")
    if len(rows[0]) != 3:
        pytest.fail(f"Expected 3-tuple, got {len(rows[0])}-tuple")

    variants = [r[0] for r in rows]
    sort_scores = [r[1] for r in rows]
    raw_scores = [r[2] for r in rows]

    if "A1V" not in variants:
        pytest.fail(f"Expected A1V in results, got {variants}")
    idx_a1v = variants.index("A1V")
    if abs(raw_scores[idx_a1v] - 0.9) > 1e-9:
        pytest.fail(f"Expected raw=0.9 for A1V, got {raw_scores[idx_a1v]}")
    if abs(sort_scores[idx_a1v] - 0.9) > 1e-9:
        pytest.fail(f"Expected sort_score=raw for desc, got {sort_scores[idx_a1v]}")


def test_load_rows_asc_rank(tmp_path):
    """asc mode: rank 1,2,3 input gives sort_score -1,-2,-3 so rank 1 wins."""
    from kuma_core.kuro.evolvepro import _load_evolvepro_rows

    csv_file = tmp_path / "rank.csv"
    _write_csv([
        ["variant", "rank"],
        ["A1V", "1"],
        ["B2C", "3"],
        ["D3E", "2"],
    ], csv_file)

    rows = _load_evolvepro_rows(
        str(csv_file),
        score_column="rank",
        score_order="asc",
    )

    if len(rows) != 3:
        pytest.fail(f"Expected 3 rows, got {len(rows)}")

    variants = [r[0] for r in rows]
    sort_scores = [r[1] for r in rows]
    raw_scores = [r[2] for r in rows]

    idx_a1v = variants.index("A1V")
    if abs(raw_scores[idx_a1v] - 1.0) > 1e-9:
        pytest.fail(f"Expected raw=1.0 for A1V, got {raw_scores[idx_a1v]}")
    if abs(sort_scores[idx_a1v] - (-1.0)) > 1e-9:
        pytest.fail(f"Expected sort_score=-1.0 for asc rank=1, got {sort_scores[idx_a1v]}")

    idx_b2c = variants.index("B2C")
    if abs(sort_scores[idx_b2c] - (-3.0)) > 1e-9:
        pytest.fail(f"Expected sort_score=-3.0 for asc rank=3, got {sort_scores[idx_b2c]}")


def test_load_rows_column_override(tmp_path):
    """Non-alias column name works when explicit variant_column/score_column given."""
    from kuma_core.kuro.evolvepro import _load_evolvepro_rows

    csv_file = tmp_path / "custom.csv"
    _write_csv([
        ["mutation_id", "ranking_score", "note"],
        ["X5Y", "0.85", "kept"],
        ["Z9W", "0.60", "kept"],
    ], csv_file)

    rows = _load_evolvepro_rows(
        str(csv_file),
        variant_column="mutation_id",
        score_column="ranking_score",
    )

    if len(rows) != 2:
        pytest.fail(f"Expected 2 rows, got {len(rows)}")
    variants = [r[0] for r in rows]
    if "X5Y" not in variants:
        pytest.fail(f"Expected X5Y in results, got {variants}")


def test_load_rows_xlsx_single_sheet(tmp_path):
    """XLSX with one sheet auto-selects it when sheet_name is None."""
    from kuma_core.kuro.evolvepro import _load_evolvepro_rows

    xlsx_file = tmp_path / "data.xlsx"
    _write_xlsx({"Sheet1": [
        ["variant", "y_pred"],
        ["P10Q", "0.75"],
        ["R11S", "0.65"],
    ]}, xlsx_file)

    rows = _load_evolvepro_rows(str(xlsx_file))

    if len(rows) != 2:
        pytest.fail(f"Expected 2 rows from xlsx, got {len(rows)}")
    variants = [r[0] for r in rows]
    if "P10Q" not in variants:
        pytest.fail(f"Expected P10Q, got {variants}")


def test_load_rows_xlsx_multi_sheet(tmp_path):
    """XLSX with multiple sheets reads only the specified sheet."""
    from kuma_core.kuro.evolvepro import _load_evolvepro_rows

    xlsx_file = tmp_path / "multi.xlsx"
    _write_xlsx({
        "Predictions": [
            ["variant", "y_pred"],
            ["M20L", "0.88"],
        ],
        "Other": [
            ["variant", "y_pred"],
            ["IGNORE1", "0.99"],
            ["IGNORE2", "0.98"],
        ],
    }, xlsx_file)

    rows = _load_evolvepro_rows(str(xlsx_file), sheet_name="Predictions")

    if len(rows) != 1:
        pytest.fail(f"Expected 1 row from Predictions sheet, got {len(rows)}")
    if rows[0][0] != "M20L":
        pytest.fail(f"Expected M20L, got {rows[0][0]}")


def test_load_rows_xls_multi_sheet(tmp_path):
    """Legacy XLS with multiple sheets reads the specified sheet."""
    from kuma_core.kuro.evolvepro import _load_evolvepro_rows

    xls_file = tmp_path / "multi.xls"
    _write_xls({
        "Predictions": [
            ["mutation_id", "ranking_score"],
            ["M20L", 0.88],
        ],
        "Other": [
            ["mutation_id", "ranking_score"],
            ["IGNORE1", 0.99],
        ],
    }, xls_file)

    rows = _load_evolvepro_rows(
        str(xls_file),
        sheet_name="Predictions",
        variant_column="mutation_id",
        score_column="ranking_score",
    )

    if len(rows) != 1:
        pytest.fail(f"Expected 1 row from XLS Predictions sheet, got {len(rows)}")
    if rows[0][0] != "M20L":
        pytest.fail(f"Expected M20L, got {rows[0][0]}")


# ---------------------------------------------------------------------------
# Task 5 tests: load_evolvepro_csv with asc mode + Top-N
# ---------------------------------------------------------------------------

def test_load_evolvepro_csv_asc_top_n(tmp_path):
    """asc mode: Top-N=3 picks rank 1, 2, 3 and yPredMap shows raw ranks."""
    from kuma_core.kuro.evolvepro import load_evolvepro_csv

    csv_file = tmp_path / "ranks.csv"
    _write_csv([
        ["variant", "rank"],
        ["A1V", "1"],
        ["B2C", "2"],
        ["D3E", "3"],
        ["E4F", "4"],
        ["G5H", "5"],
    ], csv_file)

    result = load_evolvepro_csv(
        str(csv_file),
        top_n=3,
        score_column="rank",
        score_order="asc",
    )

    selected = result["variants"]
    y_preds = result["y_preds"]

    if len(selected) != 3:
        pytest.fail(f"Expected 3 selected variants, got {len(selected)}: {selected}")

    if "A1V" not in selected:
        pytest.fail(f"Expected A1V (rank 1) in top 3, got {selected}")
    if "B2C" not in selected:
        pytest.fail(f"Expected B2C (rank 2) in top 3, got {selected}")
    if "D3E" not in selected:
        pytest.fail(f"Expected D3E (rank 3) in top 3, got {selected}")

    idx_a1v = selected.index("A1V")
    if abs(y_preds[idx_a1v] - 1.0) > 1e-6:
        pytest.fail(f"Expected yPredMap to show raw rank=1.0, got {y_preds[idx_a1v]}")
