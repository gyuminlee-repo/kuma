"""Tests for preview_evolvepro_source RPC handler."""

from __future__ import annotations

import csv
import tempfile
from pathlib import Path

import openpyxl
import pytest
import xlwt


def _make_csv(rows: list[list[str]], tmp_dir: str) -> str:
    """Write a temporary CSV file and return its path."""
    p = Path(tmp_dir) / "test_preview.csv"
    with open(p, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return str(p)


def _make_xlsx(sheets: dict[str, list[list[str]]], tmp_dir: str) -> str:
    """Write a temporary XLSX file with given sheets and return its path."""
    p = Path(tmp_dir) / "test_preview.xlsx"
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(p))
    return str(p)


def _make_xls(sheets: dict[str, list[list[str]]], tmp_dir: str) -> str:
    """Write a temporary legacy XLS file with given sheets and return its path."""
    p = Path(tmp_dir) / "test_preview.xls"
    wb = xlwt.Workbook()
    for sheet_name, rows in sheets.items():
        ws = wb.add_sheet(sheet_name)
        for ridx, row in enumerate(rows):
            for cidx, value in enumerate(row):
                ws.write(ridx, cidx, value)
    wb.save(str(p))
    return str(p)


def test_preview_csv_eight_rows():
    """CSV preview returns headers + up to 8 data rows, sheets=[]."""
    from sidecar_kuro.handlers.misc import handle_preview_evolvepro_source

    header = ["variant", "y_pred", "note"]
    data = [[f"M{i}A", str(0.9 - i * 0.05), "ok"] for i in range(12)]

    with tempfile.TemporaryDirectory() as tmp_dir:
        csv_path = _make_csv([header] + data, tmp_dir)
        result = handle_preview_evolvepro_source({"filepath": csv_path})

    if result["sheets"] != []:
        pytest.fail(f"Expected sheets=[] for CSV, got {result['sheets']}")
    if result["headers"] != header:
        pytest.fail(f"Expected headers={header}, got {result['headers']}")
    if len(result["rows"]) != 8:
        pytest.fail(f"Expected 8 rows (default max_rows), got {len(result['rows'])}")
    if result["rows"][0] != ["M0A", "0.9", "ok"]:
        pytest.fail(f"Unexpected first row: {result['rows'][0]}")


def test_preview_xlsx_multi_sheet():
    """XLSX preview lists all sheet names and reads specified sheet rows."""
    from sidecar_kuro.handlers.misc import handle_preview_evolvepro_source

    header = ["variant", "rank"]
    sheet1_data = [[f"V{i}A", str(i)] for i in range(1, 5)]
    sheet2_data = [[f"L{i}G", str(i * 2)] for i in range(1, 4)]

    sheets_data = {
        "Predictions": [header] + sheet1_data,
        "Summary": [header] + sheet2_data,
    }

    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = _make_xlsx(sheets_data, tmp_dir)

        result_sheet1 = handle_preview_evolvepro_source({
            "filepath": xlsx_path,
            "sheet_name": "Predictions",
        })
        result_sheet2 = handle_preview_evolvepro_source({
            "filepath": xlsx_path,
            "sheet_name": "Summary",
        })

    if set(result_sheet1["sheets"]) != {"Predictions", "Summary"}:
        pytest.fail(f"Expected both sheet names, got {result_sheet1['sheets']}")
    if result_sheet1["headers"] != header:
        pytest.fail(f"Unexpected headers: {result_sheet1['headers']}")
    if len(result_sheet1["rows"]) != 4:
        pytest.fail(f"Expected 4 rows from Predictions, got {len(result_sheet1['rows'])}")
    if len(result_sheet2["rows"]) != 3:
        pytest.fail(f"Expected 3 rows from Summary, got {len(result_sheet2['rows'])}")


def test_preview_xls_multi_sheet():
    """Legacy XLS preview lists sheets and reads selected sheet rows."""
    from sidecar_kuro.handlers.misc import handle_preview_evolvepro_source

    sheets_data = {
        "Predictions": [["mut", "rank"], ["A1V", 1], ["L2P", 2]],
        "Other": [["mut", "rank"], ["K3R", 3]],
    }

    with tempfile.TemporaryDirectory() as tmp_dir:
        xls_path = _make_xls(sheets_data, tmp_dir)
        result = handle_preview_evolvepro_source({
            "filepath": xls_path,
            "sheet_name": "Other",
        })

    if set(result["sheets"]) != {"Predictions", "Other"}:
        pytest.fail(f"Expected XLS sheet names, got {result['sheets']}")
    if result["headers"] != ["mut", "rank"]:
        pytest.fail(f"Unexpected XLS headers: {result['headers']}")
    if result["rows"] != [["K3R", "3.0"]]:
        pytest.fail(f"Unexpected XLS rows: {result['rows']}")


def test_preview_tsv_tab_delimited():
    """TSV preview parses tab-delimited rows the same as CSV."""
    from sidecar_kuro.handlers.misc import handle_preview_evolvepro_source

    with tempfile.TemporaryDirectory() as tmp_dir:
        p = Path(tmp_dir) / "test_preview.tsv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            f.write("variant\ty_pred\n")
            f.write("M1A\t0.85\n")
        result = handle_preview_evolvepro_source({"filepath": str(p)})

    if result["sheets"] != []:
        pytest.fail(f"Expected sheets=[] for TSV, got {result['sheets']}")
    if result["headers"] != ["variant", "y_pred"]:
        pytest.fail(f"Expected tab-split headers, got {result['headers']}")
    if result["rows"] != [["M1A", "0.85"]]:
        pytest.fail(f"Expected tab-split row, got {result['rows']}")


def test_preview_xlsx_invalid_sheet_name():
    """Requesting a non-existent sheet raises ValueError."""
    from sidecar_kuro.handlers.misc import handle_preview_evolvepro_source

    sheets_data = {"Sheet1": [["variant", "score"], ["A1V", "0.9"]]}

    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = _make_xlsx(sheets_data, tmp_dir)
        with pytest.raises(ValueError, match="sheet"):
            handle_preview_evolvepro_source({
                "filepath": xlsx_path,
                "sheet_name": "NonExistent",
            })
