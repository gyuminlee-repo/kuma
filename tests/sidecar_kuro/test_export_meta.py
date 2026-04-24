"""Tests for __kuma_meta__ sheet write in Kuro export."""
from __future__ import annotations

import openpyxl
import pytest

from sidecar_kuro.core import _state, _state_lock
from sidecar_kuro.handlers.export import handle_export_excel


@pytest.fixture
def minimal_state():
    """Populate sidecar_kuro state with a single mapping so export can run."""
    with _state_lock:
        saved_results = list(_state.results)
        saved_mappings = list(_state.plate_mappings)
        saved_dedup = dict(_state.dedup_info) if _state.dedup_info else {}

        _state.results = []
        _state.plate_mappings = []
        _state.dedup_info = {}
    yield
    with _state_lock:
        _state.results = saved_results
        _state.plate_mappings = saved_mappings
        _state.dedup_info = saved_dedup


def _make_mapping_item():
    return {
        "well": "A1",
        "primer_name": "M1_F",
        "sequence": "ACGTACGTACGT",
        "primer_type": "forward",
        "mutation": "M1",
    }


def test_xlsx_contains_kuma_meta_when_project_id_given(tmp_path, minimal_state):
    out = tmp_path / "out.xlsx"
    handle_export_excel(
        {
            "filepath": str(out),
            "mappings": [_make_mapping_item()],
            "dedup_info": {},
            "project_id": "abc-123",
            "kuma_version": "9.9.9",
        }
    )
    wb = openpyxl.load_workbook(out)
    assert "__kuma_meta__" in wb.sheetnames
    sheet = wb["__kuma_meta__"]
    assert sheet.sheet_state == "hidden"
    kv = {
        row[0].value: row[1].value
        for row in sheet.iter_rows(min_row=1, max_row=4, max_col=2)
    }
    assert kv["project_id"] == "abc-123"
    assert kv["kuma_version"] == "9.9.9"
    assert "kuro_module_version" in kv
    assert "exported_at" in kv


def test_xlsx_has_no_kuma_meta_in_scratch_mode(tmp_path, minimal_state):
    out = tmp_path / "out.xlsx"
    handle_export_excel(
        {
            "filepath": str(out),
            "mappings": [_make_mapping_item()],
            "dedup_info": {},
        }
    )
    wb = openpyxl.load_workbook(out)
    assert "__kuma_meta__" not in wb.sheetnames
