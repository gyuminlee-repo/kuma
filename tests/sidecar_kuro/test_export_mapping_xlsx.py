"""Regression tests for xlsx layout sheet in mapping exports.

Guards that ``handle_export_mapping`` with .xlsx output produces both a
``layout`` sheet and a format-specific data sheet (Echo / JANUS).
"""

from __future__ import annotations

from openpyxl import load_workbook

from sidecar_kuro.handlers.export import handle_export_mapping


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _make_mappings():
    """Minimal pair of forward + reverse mappings for plate export."""
    return [
        {
            "well": "A1",
            "primer_name": "M1_F",
            "sequence": "ACGTACGTACGT",
            "primer_type": "forward",
            "mutation": "M1",
        },
        {
            "well": "A2",
            "primer_name": "M1_R",
            "sequence": "TTTTACGTACGT",
            "primer_type": "reverse",
            "mutation": "M1",
        },
    ]


def test_export_mapping_xlsx_has_layout_sheet(tmp_path):
    """Echo xlsx export must contain both 'layout' and 'Echo mapping file' sheets."""
    path = tmp_path / "echo.xlsx"
    handle_export_mapping({
        "filepath": str(path),
        "format": "echo",
        "mappings": _make_mappings(),
        "dedup_info": {},
    })
    wb = load_workbook(path)
    sheets = wb.sheetnames
    _require("layout" in sheets, f"missing 'layout' sheet: {sheets}")
    _require("Echo mapping file" in sheets, f"missing 'Echo mapping file' sheet: {sheets}")


def test_export_janus_mapping_xlsx_has_layout_sheet(tmp_path):
    """JANUS xlsx export must contain both 'layout' and 'primer_mapping file' sheets."""
    path = tmp_path / "janus.xlsx"
    handle_export_mapping({
        "filepath": str(path),
        "format": "janus",
        "mappings": _make_mappings(),
        "dedup_info": {},
    })
    wb = load_workbook(path)
    sheets = wb.sheetnames
    _require("layout" in sheets, f"missing 'layout' sheet: {sheets}")
    _require("primer_mapping file" in sheets, f"missing 'primer_mapping file' sheet: {sheets}")
