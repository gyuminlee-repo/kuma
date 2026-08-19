"""Tests for the ``strategy.classify_round`` JSON-RPC handler -- Fork D (v0.4).

Contract change: params are now ``{round_files, c_next}`` (xlsx paths).
The old ``{round_id}`` param and sidecar _rounds store are no longer used.

Test structure:
  TestValidation     -- missing/invalid params raise ValueError
  TestXlsxParsing    -- column validation and anti-fallback (bad xlsx)
  TestMultiRound     -- 2-3 round fixtures produce non-deferred Decision
  TestLog2Fc         -- current_round_activities == log2(activity) (AC2)
  TestMissingColumns -- missing Variant/activity columns raise ValueError
  TestZeroActivity   -- activity <= 0 raises ValueError (anti-fallback)

Fixture design (AC3 rationale):
  sigma_assay=None (no WT) -> T2=NA, T_model=NA.
  Decision engine uses T1 and T3 only.
  For non-deferred result without WT:
    - n_rounds >= N_min=3 to pass calibration gate
    - hit_rate rising -> T3=False -> no saturation -> continue_walking
  switch_combinatorial/stop require wt_values (bootstrap gate); unreachable
  without WT import -- this is correct per spec (sigma deferred).  The handler
  reports that case as advisory="not_assessable" rather than as a deferred
  decision, so a question never asked is not counted as a judgement withheld.

anti-fallback: missing columns, bad Variant, activity<=0 all raise;
  no fabricated defaults.
"""

from __future__ import annotations

import importlib
import math

import openpyxl
import pytest

from sidecar_mame.handlers.classify_round import (
    _BOOTSTRAP_GATED_LABELS,
    _compute_delta_best_ema,
    _load_xlsx,
    _round_metrics,
    _wt_values,
    handle_classify_round,
)


# ---------------------------------------------------------------------------
# Helpers to build synthetic xlsx fixtures
# ---------------------------------------------------------------------------

def _make_xlsx(path, rows):
    """Write a minimal xlsx with Variant + activity columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["Variant", "activity"])
    for variant, activity in rows:
        ws.append([variant, activity])
    wb.save(path)


# ---------------------------------------------------------------------------
# TestValidation
# ---------------------------------------------------------------------------

class TestValidation:
    def test_missing_round_files_raises_value_error(self):
        with pytest.raises(ValueError, match="round_files"):
            handle_classify_round({})

    def test_empty_round_files_raises_value_error(self):
        with pytest.raises(ValueError, match="round_files"):
            handle_classify_round({"round_files": []})

    def test_bad_c_next_raises_value_error(self):
        with pytest.raises(ValueError, match="c_next"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": "/tmp/x.xlsx"}], "c_next": "bad"}
            )

    def test_missing_path_in_round_file_raises_value_error(self):
        with pytest.raises(ValueError, match="missing"):
            handle_classify_round({"round_files": [{"n": 1}]})


# ---------------------------------------------------------------------------
# TestXlsxParsing
# ---------------------------------------------------------------------------

class TestXlsxParsing:
    def test_file_not_found_raises_runtime_error(self):
        with pytest.raises(RuntimeError, match="not found"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": "/nonexistent/path/round.xlsx"}]}
            )

    def test_missing_variant_column_raises_value_error(self, tmp_path):
        bad_xlsx = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["SomeCol", "activity"])
        ws.append(["A100", 1.2])
        wb.save(str(bad_xlsx))
        with pytest.raises(ValueError, match="Variant"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": str(bad_xlsx)}]}
            )

    def test_missing_activity_column_raises_value_error(self, tmp_path):
        bad_xlsx = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Variant", "Score"])
        ws.append(["100A", 1.2])
        wb.save(str(bad_xlsx))
        with pytest.raises(ValueError, match="activity"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": str(bad_xlsx)}]}
            )

    def test_variant_without_leading_integer_raises_value_error(self, tmp_path):
        bad_xlsx = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Variant", "activity"])
        ws.append(["NoPosition", 1.5])
        wb.save(str(bad_xlsx))
        with pytest.raises(ValueError, match="no leading integer"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": str(bad_xlsx)}]}
            )

    def test_activity_zero_raises_value_error(self, tmp_path):
        bad_xlsx = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Variant", "activity"])
        ws.append(["100A", 0.0])
        wb.save(str(bad_xlsx))
        with pytest.raises(ValueError, match="<= 0"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": str(bad_xlsx)}]}
            )

    def test_activity_negative_raises_value_error(self, tmp_path):
        bad_xlsx = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Variant", "activity"])
        ws.append(["100A", -0.5])
        wb.save(str(bad_xlsx))
        with pytest.raises(ValueError, match="<= 0"):
            handle_classify_round(
                {"round_files": [{"n": 1, "path": str(bad_xlsx)}]}
            )


# ---------------------------------------------------------------------------
# TestLog2Fc -- AC2: current_round_activities == log2(activity)
# ---------------------------------------------------------------------------

class TestLog2Fc:
    def test_round_metrics_log2_activities(self, tmp_path):
        activities = [1.0, 2.0, 0.5, 4.0]
        rows = [(f"{100+i}A", a) for i, a in enumerate(activities)]
        xlsx = tmp_path / "r1.xlsx"
        _make_xlsx(str(xlsx), rows)
        records = _load_xlsx(str(xlsx))
        metrics = _round_metrics(records)
        expected_log2 = [math.log2(a) for a in activities]
        assert metrics["log2_activities"] == pytest.approx(expected_log2)

    def test_beneficial_uses_activity_gt_1_strictly(self, tmp_path):
        """activity=1.0 is NOT beneficial; must be strictly > 1.0."""
        activities = [0.5, 1.0, 1.5, 2.0]
        rows = [(f"{100+i}A", a) for i, a in enumerate(activities)]
        xlsx = tmp_path / "r1.xlsx"
        _make_xlsx(str(xlsx), rows)
        records = _load_xlsx(str(xlsx))
        metrics = _round_metrics(records)
        assert metrics["beneficial_count"] == 2

    def test_hit_rate_matches_beneficial_fraction(self, tmp_path):
        activities = [0.5, 0.8, 1.2, 1.5]
        rows = [(f"{100+i}A", a) for i, a in enumerate(activities)]
        xlsx = tmp_path / "r1.xlsx"
        _make_xlsx(str(xlsx), rows)
        records = _load_xlsx(str(xlsx))
        metrics = _round_metrics(records)
        assert metrics["hit_rate"] == pytest.approx(0.5)


# ---------------------------------------------------------------------------
# TestDeltaScale -- delta_best_ema and current_round_activities share a scale
# ---------------------------------------------------------------------------

class TestDeltaScale:
    """classify.py forms delta* = delta_best_ema + (best_n* - max(activities)).

    The bracket is a difference of log2 activities, so delta_best_ema has to be
    log2 too or that addition mixes two scales. It did: the EMA was fed
    max(activity) on the linear scale. The mixture was inert while sigma_assay
    stayed None, because delta_best_ema reaches nothing but compute_T2 and
    compute_T2 answers None without a sigma.

    The tests below fail on the linear spelling. TestEmaHelper does not: it
    calls the helper with bare numbers, so it holds on either scale and
    witnesses nothing about which one the handler supplies.
    """

    def test_round_best_log2_is_log2_of_round_best(self, tmp_path):
        activities = [0.5, 1.0, 2.0, 3.0]
        rows = [(f"{100+i}A", a) for i, a in enumerate(activities)]
        xlsx = tmp_path / "r1.xlsx"
        _make_xlsx(str(xlsx), rows)
        metrics = _round_metrics(_load_xlsx(str(xlsx)))
        assert metrics["round_best"] == pytest.approx(3.0)
        assert metrics["round_best_log2"] == pytest.approx(math.log2(3.0))

    def test_round_best_log2_is_the_max_of_the_activities_handed_over(self, tmp_path):
        """The two must not drift: one is what the EMA reads, the other is what
        the bootstrap resamples, and delta* subtracts the second from the first.
        """
        activities = [0.25, 1.6, 2.4, 0.9]
        rows = [(f"{100+i}A", a) for i, a in enumerate(activities)]
        xlsx = tmp_path / "r1.xlsx"
        _make_xlsx(str(xlsx), rows)
        metrics = _round_metrics(_load_xlsx(str(xlsx)))
        assert metrics["round_best_log2"] == pytest.approx(max(metrics["log2_activities"]))

    def test_the_handler_feeds_the_ema_log2_bests(self, tmp_path, monkeypatch):
        """Intercept the argument rather than recompute it.

        Reading round_best_log2 and asserting it is log2 proves nothing about
        which of the two keys the handler passes on, and that choice is the
        defect. The maxima below are 2.0 and 8.0, so the log2 list is [1.0, 3.0]
        and the linear list is [2.0, 8.0]: no coincidence makes them agree.
        """
        import sidecar_mame.handlers.classify_round as module

        seen: list[list[float]] = []
        original = module._compute_delta_best_ema

        def recording(round_bests):
            seen.append(list(round_bests))
            return original(round_bests)

        monkeypatch.setattr(module, "_compute_delta_best_ema", recording)

        r1 = tmp_path / "r1.xlsx"
        _make_xlsx(str(r1), [("101A", 2.0), ("102A", 0.7), ("103A", 0.6)])
        r2 = tmp_path / "r2.xlsx"
        _make_xlsx(str(r2), [("201A", 8.0), ("202A", 0.7), ("203A", 0.6)])
        r3 = tmp_path / "r3.xlsx"
        _make_xlsx(str(r3), [("301A", 8.0), ("302A", 1.5), ("303A", 1.4)])

        module.handle_classify_round(
            {
                "round_files": [
                    {"n": 1, "path": str(r1)},
                    {"n": 2, "path": str(r2)},
                    {"n": 3, "path": str(r3)},
                ]
            }
        )

        # The handler computes the final EMA first, then walks the interim
        # rounds, so the calls arrive full list first and growing prefixes
        # after. Both feed sites are separate lines and either could have been
        # left linear, so both are pinned.
        assert seen == [
            pytest.approx([1.0, 3.0, 3.0]),
            pytest.approx([1.0]),
            pytest.approx([1.0, 3.0]),
        ]
        # The linear spelling of the same rounds, which none of these may be.
        assert seen[0] != pytest.approx([2.0, 8.0, 8.0])


# ---------------------------------------------------------------------------
# TestEmaHelper
# ---------------------------------------------------------------------------

class TestEmaHelper:
    def test_single_round_returns_zero(self):
        assert _compute_delta_best_ema([2.0]) == pytest.approx(0.0)

    def test_two_rounds_ema_is_first_delta(self):
        assert _compute_delta_best_ema([1.0, 1.5]) == pytest.approx(0.5)

    def test_three_rounds_ema_value(self):
        # delta_0=0.5, delta_1=0.3; EMA_1 = 2/3*0.3 + 1/3*0.5
        expected = 2 / 3 * 0.3 + 1 / 3 * 0.5
        assert _compute_delta_best_ema([1.0, 1.5, 1.8]) == pytest.approx(expected)


# ---------------------------------------------------------------------------
# TestMultiRound -- AC3: non-deferred Decision
# ---------------------------------------------------------------------------

_VALID_LABELS = frozenset({"continue_walking", "switch_combinatorial", "stop", "deferred"})
_NON_DEFERRED = frozenset({"continue_walking", "switch_combinatorial", "stop"})


class TestMultiRound:
    """Fixture design rationale:
    sigma=None -> T2=NA, T_model=NA.
    n_rounds=3 -> clears N_min=3 calibration gate (n >= N_min required).
    hit_rates rising across 3 rounds -> T3=False -> no saturation.
    Decision: continue_walking("no_saturation_signal") -- non-deferred.

    cumulative_beneficial varies; K_throughput=14 (c_next=96 default).
    T1=False (not enough beneficials) -- throughput not met.
    T3=False (rising trend) -> no saturation -> continue_walking.
    """

    def _make_3round_files(self, tmp_path):
        r1 = tmp_path / "r1.xlsx"
        _make_xlsx(
            str(r1),
            [(f"{100+i}A", 1.5 if i < 2 else 0.7) for i in range(10)],
        )
        r2 = tmp_path / "r2.xlsx"
        _make_xlsx(
            str(r2),
            [(f"{200+i}A", 1.5 if i < 4 else 0.7) for i in range(10)],
        )
        r3 = tmp_path / "r3.xlsx"
        _make_xlsx(
            str(r3),
            [(f"{300+i}A", 1.5 if i < 6 else 0.7) for i in range(10)],
        )
        return [
            {"n": 1, "path": str(r1)},
            {"n": 2, "path": str(r2)},
            {"n": 3, "path": str(r3)},
        ]

    def test_3round_returns_advisory_decision(self, tmp_path):
        result = handle_classify_round({"round_files": self._make_3round_files(tmp_path)})
        assert result["advisory"] == "decision"

    def test_3round_declares_missing_inputs(self, tmp_path):
        """An answered decision still reports which inputs were unavailable.

        continue_walking is reached without the bootstrap gate, so the absent
        WT replicates never blocked anything here.  They still shaped the
        answer (T2 and T_model were NA), and the caller is told so.
        """
        result = handle_classify_round({"round_files": self._make_3round_files(tmp_path)})
        assert result["missing_inputs"] == ["wt_replicates"]

    def test_3round_label_is_valid(self, tmp_path):
        result = handle_classify_round({"round_files": self._make_3round_files(tmp_path)})
        assert result["label"] in _VALID_LABELS

    def test_3round_non_deferred(self, tmp_path):
        """Rising hit_rate -> T3=False -> continue_walking (non-deferred).

        T3=False: slope > 0 (0.2->0.4->0.6 = positive slope).
        all_na(T2=NA, T3=False, T_model=NA) = False (T3 evaluated).
        sat_now = any_true(T2=NA, T3=False, T_model=NA) = False.
        => continue_walking(no_saturation_signal). Not deferred.
        """
        result = handle_classify_round({"round_files": self._make_3round_files(tmp_path)})
        assert result["label"] in _NON_DEFERRED, (
            f"Got {result['label']!r} reason={result.get('reason')!r}"
        )

    def test_3round_reason_non_empty(self, tmp_path):
        result = handle_classify_round({"round_files": self._make_3round_files(tmp_path)})
        assert isinstance(result["reason"], str) and result["reason"]

    def test_3round_confidence_none_or_float(self, tmp_path):
        result = handle_classify_round({"round_files": self._make_3round_files(tmp_path)})
        conf = result["confidence"]
        if conf is not None:
            assert isinstance(conf, float)
            assert 0.0 <= conf <= 1.0

    def test_2round_returns_advisory_decision(self, tmp_path):
        r1 = tmp_path / "r1.xlsx"
        r2 = tmp_path / "r2.xlsx"
        _make_xlsx(str(r1), [(f"{100+i}A", 1.3 if i < 3 else 0.6) for i in range(10)])
        _make_xlsx(str(r2), [(f"{200+i}A", 1.3 if i < 5 else 0.6) for i in range(10)])
        result = handle_classify_round(
            {"round_files": [{"n": 1, "path": str(r1)}, {"n": 2, "path": str(r2)}]}
        )
        assert result["advisory"] == "decision"

    def test_c_next_custom(self, tmp_path):
        """c_next=10 -> K=5; small throughput bar."""
        r1 = tmp_path / "r1.xlsx"
        r2 = tmp_path / "r2.xlsx"
        r3 = tmp_path / "r3.xlsx"
        _make_xlsx(str(r1), [(f"{100+i}A", 1.5 if i < 3 else 0.7) for i in range(10)])
        _make_xlsx(str(r2), [(f"{200+i}A", 1.5 if i < 5 else 0.7) for i in range(10)])
        _make_xlsx(str(r3), [(f"{300+i}A", 1.5 if i < 7 else 0.7) for i in range(10)])
        result = handle_classify_round(
            {
                "round_files": [
                    {"n": 1, "path": str(r1)},
                    {"n": 2, "path": str(r2)},
                    {"n": 3, "path": str(r3)},
                ],
                "c_next": 10,
            }
        )
        assert result["advisory"] == "decision"
        assert result["label"] in _VALID_LABELS

    def test_round_files_out_of_order_accepted(self, tmp_path):
        """round_files provided in wrong n-order are sorted correctly."""
        r1 = tmp_path / "r1.xlsx"
        r2 = tmp_path / "r2.xlsx"
        r3 = tmp_path / "r3.xlsx"
        _make_xlsx(str(r1), [(f"{100+i}A", 1.2 if i < 2 else 0.8) for i in range(8)])
        _make_xlsx(str(r2), [(f"{200+i}A", 1.2 if i < 4 else 0.8) for i in range(8)])
        _make_xlsx(str(r3), [(f"{300+i}A", 1.2 if i < 6 else 0.8) for i in range(8)])
        result = handle_classify_round(
            {
                "round_files": [
                    {"n": 3, "path": str(r3)},
                    {"n": 1, "path": str(r1)},
                    {"n": 2, "path": str(r2)},
                ]
            }
        )
        assert result["advisory"] == "decision"


# ---------------------------------------------------------------------------
# TestSingleRound -- calibration period degenerate case
# ---------------------------------------------------------------------------

class TestSingleRound:
    def test_1round_returns_calibration_period(self, tmp_path):
        """Single round (n=1 < N_min=3) returns continue_walking(calibration_period)."""
        r1 = tmp_path / "r1.xlsx"
        _make_xlsx(str(r1), [(f"{100+i}A", 1.5 if i < 3 else 0.7) for i in range(10)])
        result = handle_classify_round(
            {"round_files": [{"n": 1, "path": str(r1)}]}
        )
        assert result["advisory"] == "decision"
        assert result["label"] == "continue_walking"
        assert result["reason"] == "calibration_period"


# ---------------------------------------------------------------------------
# TestDecliningSaturation -- T3=True path -> deferred(bootstrap_inputs_missing)
# ---------------------------------------------------------------------------

class TestDecliningSaturation:
    """Fixture: 3 rounds with declining hit_rates [0.6, 0.4, 0.2].

    T3 uses a 2-round sliding window of hit_rates.  Negative slope
    (0.4->0.2 in window 2-3) signals saturation.  classify() enters
    switch/stop evaluation.  wt_values=None triggers the bootstrap gate:
      Decision(label="deferred", reason="bootstrap_inputs_missing").
    The handler translates that one case into advisory="not_assessable".

    This test proves:
      (a) previous_signals chaining fires correctly (T3 reads prior signals).
      (b) saturated-looking data yields an explicit "cannot be asked" state
          rather than a fabricated label or a withheld-judgement label.
    """

    def _make_declining_3round(self, tmp_path):
        # Round 1: hit_rate=6/10=0.6, best=2.0
        r1 = tmp_path / "r1.xlsx"
        _make_xlsx(str(r1), [(f"{100+i}A", 2.0 if i < 6 else 0.5) for i in range(10)])
        # Round 2: hit_rate=4/10=0.4, best=1.8
        r2 = tmp_path / "r2.xlsx"
        _make_xlsx(str(r2), [(f"{200+i}A", 1.8 if i < 4 else 0.5) for i in range(10)])
        # Round 3: hit_rate=2/10=0.2, best=1.5
        r3 = tmp_path / "r3.xlsx"
        _make_xlsx(str(r3), [(f"{300+i}A", 1.5 if i < 2 else 0.5) for i in range(10)])
        return [
            {"n": 1, "path": str(r1)},
            {"n": 2, "path": str(r2)},
            {"n": 3, "path": str(r3)},
        ]

    def test_declining_hit_rate_is_not_assessable(self, tmp_path):
        """T3 saturation signal -> bootstrap gate -> not_assessable."""
        result = handle_classify_round(
            {"round_files": self._make_declining_3round(tmp_path)}
        )
        assert result["advisory"] == "not_assessable", (
            f"Expected the missing-input state; got {result['advisory']!r} "
            f"reason={result.get('reason')!r}"
        )
        assert result["reason"] == "wt_replicates_missing", (
            f"Expected wt_replicates_missing; got {result['reason']!r}"
        )

    def test_declining_names_the_missing_input(self, tmp_path):
        """The state says what is absent and what that costs."""
        result = handle_classify_round(
            {"round_files": self._make_declining_3round(tmp_path)}
        )
        assert result["missing_inputs"] == ["wt_replicates"]
        assert result["blocked_decisions"] == ["switch_combinatorial", "stop"]

    def test_declining_carries_no_decision_fields(self, tmp_path):
        """A question never asked has no label, reason code, or confidence."""
        result = handle_classify_round(
            {"round_files": self._make_declining_3round(tmp_path)}
        )
        assert "label" not in result
        assert "confidence" not in result

    def test_declining_reports_zero_replicates_on_record(self, tmp_path):
        """No WT recorded and too few WT recorded are different facts.

        Both leave the bootstrap gate shut, so both answer not_assessable.  The
        counts are what tells them apart, and this is the "none at all" side.
        """
        result = handle_classify_round(
            {"round_files": self._make_declining_3round(tmp_path)}
        )
        assert result["wt_replicate_count"] == 0
        assert result["wt_replicate_min"] == 4


# ---------------------------------------------------------------------------
# TestWtReplicatesForwarded -- step 4.1 replicates reaching the bootstrap gate
# ---------------------------------------------------------------------------

# Replicates of a wild-type well, on the scale of the activity column: each
# measurement over the mean of its own cohort, which is where they sit around
# 1.0.  Step 4.1 records these on the round it built and the caller forwards
# them on the matching round_files entry.
_WT_FOUR = [1.02, 0.97, 1.04, 0.99]
_WT_THREE = _WT_FOUR[:3]


class TestWtReplicatesForwarded:
    """The saturating fixture above, run with WT replicates beside the file.

    The same three rounds reach the bootstrap gate every time; what changes is
    whether the gate has anything to run on.  Below wt_replicate_min the
    replicates are withheld on purpose: compute_sigma_assay returns None under
    that count, so T2 and T_model would stay NA in every resample and the
    confirmation would fall back on the same lone T3 that proposed the branch.
    A T3 stable under resampling scores that agreement as high confidence, so
    forwarding three would print a single-signal switch as a near-certainty.
    """

    def _files(self, tmp_path, wt=None):
        files = TestDecliningSaturation()._make_declining_3round(tmp_path)
        if wt is not None:
            files[-1]["wt_values"] = list(wt)
        return files

    def test_four_replicates_reach_the_classifier(self, tmp_path):
        """With enough replicates the gate runs and a real verdict comes back."""
        result = handle_classify_round(
            {"round_files": self._files(tmp_path, _WT_FOUR)}
        )
        assert result["advisory"] == "decision", (
            f"Expected the classifier to answer; got {result!r}"
        )
        assert result["label"] in _BOOTSTRAP_GATED_LABELS, (
            f"The gate is only reached for switch/stop; got {result['label']!r}"
        )
        assert isinstance(result["confidence"], float)

    def test_answered_decision_still_reports_the_missing_input(self, tmp_path):
        """Supplied replicates do not reach the point signals, only the bootstrap.

        The handler passes sigma_assay=None either way, so the verdict itself
        was still reached with T2 and T_model NA and saturation resting on the
        hit-rate trend alone. That is what the caller note says, so it has to
        keep being reported here.
        """
        result = handle_classify_round(
            {"round_files": self._files(tmp_path, _WT_FOUR)}
        )
        assert result["missing_inputs"] == ["wt_replicates"]

    def _captured_wt_values(self, tmp_path, monkeypatch, wt):
        """Run the handler and return the wt_values classify() actually saw.

        The handler imports classify() inside the function body, so replacing
        the module attribute intercepts the real call.

        Asserted directly rather than through the answer because on this
        fixture the answer cannot tell replicate lists apart: T3 is True in
        every draw, sat_now is any_true over T2/T3/T_model, and a True T3
        settles it whatever sigma the draw derived. Both a tight and a
        scattered WT block return confidence 1.0 here (measured). Every other
        assertion in this class would therefore still pass if the handler
        forwarded a list of the right length holding the wrong numbers.
        """
        # import_module, not `import ... as`: the package re-exports a function
        # named classify, which shadows the submodule attribute.
        classify_module = importlib.import_module("kuma_core.strategy.classify")

        seen = {}
        real_classify = classify_module.classify

        def capture(round_state, registered):
            seen["wt_values"] = round_state.wt_values
            return real_classify(round_state, registered)

        monkeypatch.setattr(classify_module, "classify", capture)
        handle_classify_round({"round_files": self._files(tmp_path, wt)})
        return seen["wt_values"]

    def test_the_replicate_values_reach_the_classifier_unchanged(
        self, tmp_path, monkeypatch
    ):
        values = [0.4013, 1.9007, 0.5501, 1.7002]
        assert self._captured_wt_values(tmp_path, monkeypatch, values) == values

    def test_short_replicate_lists_reach_the_classifier_as_none(
        self, tmp_path, monkeypatch
    ):
        assert self._captured_wt_values(tmp_path, monkeypatch, _WT_THREE) is None

    def test_three_replicates_do_not_reach_the_classifier(self, tmp_path):
        """One short of the minimum is still not assessable."""
        result = handle_classify_round(
            {"round_files": self._files(tmp_path, _WT_THREE)}
        )
        assert result["advisory"] == "not_assessable"
        assert result["reason"] == "wt_replicates_insufficient"
        assert "label" not in result

    def test_three_replicates_are_counted_in_the_response(self, tmp_path):
        """The screen has to be able to say "3 on record, 4 needed"."""
        result = handle_classify_round(
            {"round_files": self._files(tmp_path, _WT_THREE)}
        )
        assert result["wt_replicate_count"] == 3
        assert result["wt_replicate_min"] == 4

    def test_replicates_on_earlier_rounds_are_not_read(self, tmp_path):
        """The bootstrap resamples the current round, so only its entry counts."""
        files = self._files(tmp_path)
        files[0]["wt_values"] = list(_WT_FOUR)
        result = handle_classify_round({"round_files": files})
        assert result["advisory"] == "not_assessable"
        assert result["reason"] == "wt_replicates_missing"
        assert result["wt_replicate_count"] == 0

    def test_empty_replicate_list_reads_as_none_recorded(self, tmp_path):
        result = handle_classify_round({"round_files": self._files(tmp_path, [])})
        assert result["reason"] == "wt_replicates_missing"
        assert result["wt_replicate_count"] == 0

    def test_unreadable_replicates_raise(self, tmp_path):
        """anti-fallback: a malformed value must not read as "none recorded"."""
        with pytest.raises(ValueError):
            handle_classify_round({"round_files": self._files(tmp_path, ["n/a", 1.0, 1.0, 1.0])})
        with pytest.raises(ValueError):
            handle_classify_round(
                {"round_files": self._files(tmp_path, [float("nan"), 1.0, 1.0, 1.0])}
            )

    def test_replicates_must_be_a_list(self, tmp_path):
        files = self._files(tmp_path)
        files[-1]["wt_values"] = 1.0
        with pytest.raises(ValueError):
            handle_classify_round({"round_files": files})

    def test_helper_reads_absent_and_present_values(self):
        assert _wt_values({"n": 1, "path": "x"}) == []
        assert _wt_values({"wt_values": None}) == []
        assert _wt_values({"wt_values": [1, "1.5"]}) == [1.0, 1.5]

