"""Pin the Echo and JANUS export format to the workbooks the lab imports.

The header strings in ``plate_mapper`` had no traceable source in the
repository: the commit that introduced them cites none, and a docstring
claiming a "lab reference format" left no artifact behind. Vendor
documentation spells several columns differently, which invites a correction
that would break an import the lab depends on.

``tests/fixtures/liquid_handler/reference_format.json`` carries the sheet
names, mapping-sheet headers, plate names, and layout anchors read out of the
real workbooks. See the README beside it for provenance and for why only the
format, and not the workbooks, is committed.

The JANUS half now pins "Project3_seeding mapping file (JANUS).xlsx", the eight
column seeding sheet with named racks. Its predecessor had nine columns, a
liquid class in the third, and integer racks; the fixture records both files
because only the older one carries the layout sheet that names the KURO plates.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
import pytest

from kuma_core.kuro.plate_mapper import (
    PlateMapping,
    export_echo_mapping_csv,
    export_echo_mapping_xlsx,
    export_janus_mapping_csv,
    export_janus_mapping_xlsx,
)
from kuma_core.shared.janus_deck import JANUS_DEVICE_HEADER

_REFERENCE = (
    Path(__file__).resolve().parent
    / "fixtures"
    / "liquid_handler"
    / "reference_format.json"
)


@pytest.fixture(scope="module")
def reference() -> dict:
    return json.loads(_REFERENCE.read_text(encoding="utf-8"))


@pytest.fixture
def mappings() -> tuple[list[PlateMapping], list[PlateMapping]]:
    """Two mutations, each with its own reverse primer."""
    fwd = [
        PlateMapping(
            well="A1",
            primer_name="V5F_F",
            sequence="AAAA",
            primer_type="forward",
            mutation="V5F",
        ),
        PlateMapping(
            well="B1",
            primer_name="V10L_F",
            sequence="CCCC",
            primer_type="forward",
            mutation="V10L",
        ),
    ]
    rev = [
        PlateMapping(
            well="A1",
            primer_name="V5F_R",
            sequence="TTTT",
            primer_type="reverse",
            mutation="V5F",
        ),
        PlateMapping(
            well="B1",
            primer_name="V10L_R",
            sequence="GGGG",
            primer_type="reverse",
            mutation="V10L",
        ),
    ]
    return fwd, rev


def _rev_groups(rev: list[PlateMapping]) -> dict[str, list[str]]:
    return {m.sequence: [m.mutation] for m in rev}


class TestEchoReferenceFormat:
    def test_csv_header_matches_the_workbook(self, mappings, reference, tmp_path):
        fwd, rev = mappings
        out = tmp_path / "echo.csv"
        export_echo_mapping_csv(fwd, rev, out, rev_groups=_rev_groups(rev))

        with open(out, encoding="utf-8") as f:
            header = next(csv.reader(f))

        assert header == reference["echo"]["mapping_header"]

    def test_xlsx_sheets_and_header_match_the_workbook(
        self, mappings, reference, tmp_path
    ):
        fwd, rev = mappings
        out = tmp_path / "echo.xlsx"
        export_echo_mapping_xlsx(fwd, rev, out, rev_groups=_rev_groups(rev))

        wb = openpyxl.load_workbook(out)
        expected = reference["echo"]
        assert expected["mapping_sheet"] in wb.sheetnames
        assert "layout" in wb.sheetnames

        ws = wb[expected["mapping_sheet"]]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert header[: len(expected["mapping_header"])] == expected["mapping_header"]

    def test_layout_names_the_same_labware(self, mappings, reference, tmp_path):
        fwd, rev = mappings
        out = tmp_path / "echo.xlsx"
        export_echo_mapping_xlsx(fwd, rev, out, rev_groups=_rev_groups(rev))

        lay = openpyxl.load_workbook(out)["layout"]
        cells = {
            str(c.value)
            for row in lay.iter_rows(max_row=10)
            for c in row
            if c.value is not None
        }
        anchors = reference["echo"]["layout_anchors"]
        assert anchors["labware_label"] in cells
        assert anchors["labware_value"] in cells


class TestJanusReferenceFormat:
    def test_csv_header_matches_the_workbook(self, mappings, reference, tmp_path):
        fwd, rev = mappings
        out = tmp_path / "janus.csv"
        export_janus_mapping_csv(fwd, rev, out, rev_groups=_rev_groups(rev))

        with open(out, encoding="utf-8") as f:
            header = next(csv.reader(f))

        assert header == reference["janus"]["mapping_header"]

    def test_the_header_is_exactly_these_eight_columns(self, reference):
        """Pin the column list itself, in order, against the shipped constant.

        The other tests here compare a writer to the fixture, so an edit that
        changed the fixture and the writer together would pass them all. This
        one restates the eight names as a literal, so the seeding workbook shape
        cannot drift without someone typing the new shape out.
        """
        expected = [
            "name",
            "type",
            "no",
            "Asp. Rack",
            "Asp. Posi",
            "Dsp. Rack",
            "Dsp. Posi",
            "volume",
        ]
        assert reference["janus"]["mapping_header"] == expected
        # The constant production writes from, and the transcription of the
        # workbook, are the same list. Nothing else in the chain has to agree
        # once these two do.
        assert JANUS_DEVICE_HEADER == expected

    def test_no_column_name_is_repeated_and_none_holds_a_liquid_class(
        self, reference
    ):
        """The predecessor sheet is a regression here, not a repair.

        The older workbook named ``Dsp. Rack`` twice and carried a liquid class
        in its third column, and a test once pinned that repetition as
        deliberate. Both are gone, so this states the opposite: the eight names
        are distinct, and no column exists for a liquid class to be written to.
        """
        header = reference["janus"]["mapping_header"]
        assert len(header) == len(set(header))
        assert not [c for c in header if "class" in c.lower()]

    def test_xlsx_sheets_and_header_match_the_workbook(
        self, mappings, reference, tmp_path
    ):
        fwd, rev = mappings
        out = tmp_path / "janus.xlsx"
        export_janus_mapping_xlsx(fwd, rev, out, rev_groups=_rev_groups(rev))

        wb = openpyxl.load_workbook(out)
        expected = reference["janus"]
        assert expected["mapping_sheet"] in wb.sheetnames
        # "layout" is KURO's own sheet, not something the seeding workbook has:
        # that workbook carries the mapping sheet alone. KURO keeps writing a
        # layout sheet because it is where the three plate names are spelled
        # out for whoever sets up the deck, which is why the fixture records the
        # older workbook as their source.
        assert "layout" in wb.sheetnames

        ws = wb[expected["mapping_sheet"]]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        assert header[: len(expected["mapping_header"])] == expected["mapping_header"]

    def test_the_layout_sheet_names_the_plates_the_mapping_sheet_aspirates_from(
        self, mappings, reference, tmp_path
    ):
        """The two sheets must call each plate by one name.

        The mapping sheet used to address plates by number while the layout
        sheet named them, so the same plate had two identities in one workbook.
        Now both carry the name, and this pins that they carry the *same* name.
        """
        fwd, rev = mappings
        out = tmp_path / "janus.xlsx"
        export_janus_mapping_xlsx(fwd, rev, out, rev_groups=_rev_groups(rev))

        wb = openpyxl.load_workbook(out)
        names = reference["janus"]["rack_names"]
        layout_cells = {
            str(c.value)
            for row in wb["layout"].iter_rows()
            for c in row
            if c.value is not None
        }
        for key in ("kuro_fwd", "kuro_rev", "kuro_dest"):
            assert names[key] in layout_cells

        ws = wb[reference["janus"]["mapping_sheet"]]
        header = [c.value for c in next(ws.iter_rows(max_row=1))]
        asp = header.index("Asp. Rack")
        dsp = header.index("Dsp. Rack")
        body = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
        assert {r[asp] for r in body} == {names["kuro_fwd"], names["kuro_rev"]}
        assert {r[dsp] for r in body} == {names["kuro_dest"]}


class TestReferenceFixtureItself:
    def test_records_provenance(self, reference):
        """A format claim without a source is what this fixture exists to prevent."""
        for key in ("echo", "janus"):
            entry = reference[key]
            assert entry["source_file"]
            assert len(entry["sha256"]) == 64

    def test_the_kuro_plate_names_carry_their_own_source(self, reference):
        """The names and the header come from two different workbooks.

        The seeding workbook has no layout sheet, so it cannot be the source of
        the KURO plate names. Recording the header against one file while the
        names silently kept the other file's authority is exactly the untraceable
        claim this fixture exists to stop, so the second source is written down
        and pinned here.
        """
        layout_source = reference["janus"]["kuro_layout_source"]
        assert layout_source["source_file"] != reference["janus"]["source_file"]
        assert len(layout_source["sha256"]) == 64
        assert layout_source["sheet"] not in reference["janus"]["sheet_names"]
