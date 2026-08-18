"""The analyze plate-capacity refusal, and what it deliberately does NOT refuse.

Fixtures that put the boundary where the plate actually is. A plate holds 96
wells and the WT control is a sequencing target that takes one of them, so a
designed list of N occupies N + 1 wells:

* 95 designed + no WT row  -> 96 occupants, exactly full, runs
* 96 designed + no WT row  -> 97 occupants, refused before the demux
* 97 designed              -> refused, and every mutant past the plate is named

The middle case is the whole point, and it used to be the bug. Capacity was
judged on ``len(expected)`` against 96, so 96 designed rows passed the gate and
the placement loop then asked ``seq_to_well`` for well 97 and raised in the
middle of the run. Judging on total occupancy BEFORE anything is placed is what
turns that into a refusal an operator can act on.

The other condition is orthogonal to all three and outranks them: the refusal is
only asked of a run that has to DRAFT its layout out of ``expected``. A run given
``well_layout`` was told which wells it scores, so a longer designed list is a
lookup table with spare rows rather than an overflowing plate, and refusing it
would break a configuration that ran correctly before this check existed.

Also covers the axis counts ``validate_inputs`` reads back off the barcode
workbook, which are display only: nothing here is a value the operator sets.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from sidecar_mame.handlers.analyze import (
    _barcode_axis_counts,
    _layout_is_inferred,
    _plate_capacity_finding,
    handle_analyze,
    handle_validate_inputs,
)

_F_TAIL = "cacaggaggttaaacc"
_R_TAIL = "tgcgttgcgctctag"


def _write_expected(path: Path, n_mutants: int) -> Path:
    """A minimal KURO export carrying ``n_mutants`` DESIGNED rows and no WT row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "expected_mutations"
    ws.append([
        "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
        "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
    ])
    for i in range(1, n_mutants + 1):
        ws.append([f"M{i}", i, "A", "G", "", "", "", "", "substitution", "DESIGNED"])
    wb.save(str(path))
    return path


def _write_full_barcodes(path: Path) -> Path:
    """8 reverse and 12 forward seeds, i.e. a barcode set that fills the plate."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for i in range(1, 13):
        ws.append([f"target_f_{i}", f"AATCCCACT{i:02d}" + _F_TAIL])
    for i in range(1, 9):
        ws.append([f"target_r_{i}", f"CCCTATGA{i:02d}" + _R_TAIL])
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# The three-way discriminator
# ---------------------------------------------------------------------------

def test_95_designed_fits_and_keeps_its_wt_well(tmp_path: Path) -> None:
    expected = _write_expected(tmp_path / "expected.xlsx", 95)

    error, draft = _plate_capacity_finding(expected, None, None)

    assert error is None
    assert draft is not None
    assert len(draft.layout) == 96
    assert draft.layout["H12"] == "WT"


def test_96_designed_does_not_fit_because_the_wt_control_needs_a_well(
    tmp_path: Path,
) -> None:
    """The old boundary was off by exactly the control well.

    ``len(expected) <= 96`` passed this, and ``build_draft_layout`` then asked
    ``seq_to_well`` for well 97 while placing. The refusal has to come out of the
    occupancy arithmetic, before anything is placed.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 96)

    error, draft = _plate_capacity_finding(expected, None, None)

    assert error is not None
    assert "M96" in error
    assert draft is not None
    assert draft.layout == {}


def test_the_graded_draft_is_the_layout_the_run_would_place(tmp_path: Path) -> None:
    """Handed back, not summarised: the run reuses this object as its layout."""
    expected = _write_expected(tmp_path / "expected.xlsx", 3)

    _error, draft = _plate_capacity_finding(expected, None, None)

    assert draft is not None
    # Column-major from A1, control in the last well (the default since
    # 2026-08-18). What this pins is that the graded object IS the placed one,
    # not where the control sits, so it follows the default.
    assert draft.layout == {"A1": "M1", "B1": "M2", "C1": "M3", "H12": "WT"}


def test_97_designed_is_refused_and_the_message_states_both_numbers(
    tmp_path: Path,
) -> None:
    expected = _write_expected(tmp_path / "expected.xlsx", 97)

    error, _ = _plate_capacity_finding(expected, None, None)

    assert error is not None
    assert "97" in error
    assert "96" in error
    # The reason, not just the count: the native barcode is a replicate axis.
    assert "replicates" in error
    assert "M97" in error
    # And the way out that does not involve splitting the campaign.
    assert "well layout" in error


def test_a_file_that_will_not_open_is_not_this_checks_business(
    tmp_path: Path,
) -> None:
    """A path that is not a workbook says nothing about the plate."""
    junk = tmp_path / "junk.xlsx"
    junk.write_bytes(b"not a workbook")

    assert _plate_capacity_finding(junk, None, None) == (None, None)
    assert _plate_capacity_finding(tmp_path / "absent.xlsx", None, None) == (None, None)


def test_the_readers_own_refusal_travels_up_instead_of_being_swallowed(
    tmp_path: Path,
) -> None:
    """The bug this check existed to prevent, re-created by its own except.

    A workbook that opens and says something unplaceable (here: nothing at all)
    is a refusal, not an unreadable file. Returning ``(None, None)`` for it sent
    the caller back to ``read_variant_source`` further down the handler, which
    raised the same message AFTER the multi-minute demux. Every refusal added to
    the reader (a row read and not placed, a second WT row, a duplicate variant)
    left through the same hole.
    """
    empty = tmp_path / "empty.xlsx"
    openpyxl.Workbook().save(str(empty))

    with pytest.raises(ValueError, match="is empty"):
        _plate_capacity_finding(empty, None, None)


# ---------------------------------------------------------------------------
# Who the refusal is for: only a run that drafts its own layout
# ---------------------------------------------------------------------------


def test_only_a_run_without_a_layout_of_its_own_drafts_one() -> None:
    assert _layout_is_inferred({}) is True
    assert _layout_is_inferred({"well_layout": None}) is True
    assert _layout_is_inferred({"well_layout": {"A1": "M1"}}) is False
    # An empty mapping is still a layout the caller supplied.
    assert _layout_is_inferred({"well_layout": {}}) is False
    # A well selection says which wells the campaign occupies, not what sits in
    # them. The contents still come from reading ``expected``, so the capacity
    # question is exactly as live as it is for a run with no selection.
    assert _layout_is_inferred({"selected_wells": ["A1", "B1"]}) is True


# ---------------------------------------------------------------------------
# End to end through the handler, before any demux
# ---------------------------------------------------------------------------

def _analyze_params(tmp_path: Path, expected: Path) -> dict:
    input_dir = tmp_path / "consensus"
    input_dir.mkdir(exist_ok=True)
    reference = tmp_path / "reference.fa"
    reference.write_text(">ref\nATGGGGTTT\n", encoding="utf-8")
    return {
        "input_dir": str(input_dir),
        "reference": str(reference),
        "expected": str(expected),
        "output": str(tmp_path / "result.xlsx"),
    }


def test_analyze_refuses_a_campaign_larger_than_one_plate(tmp_path: Path) -> None:
    expected = _write_expected(tmp_path / "expected.xlsx", 97)

    with pytest.raises(ValueError) as excinfo:
        handle_analyze(_analyze_params(tmp_path, expected))

    message = str(excinfo.value)
    assert "97" in message
    assert "96" in message


def test_analyze_refuses_96_designed_rows_before_placing_anything(
    tmp_path: Path,
) -> None:
    """The refusal replaces a ValueError raised from inside the placement loop.

    Before the occupancy fix this reached ``seq_to_well(97)`` and surfaced as
    ``seq must be in [1, 96]``, which names no file, no mutant, and no way out.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 96)

    with pytest.raises(ValueError) as excinfo:
        handle_analyze(_analyze_params(tmp_path, expected))

    message = str(excinfo.value)
    assert "seq must be in" not in message
    assert "M96" in message


def test_a_campaign_larger_than_one_plate_runs_when_a_layout_names_the_plate(
    tmp_path: Path,
) -> None:
    """The regression the gate must not cause.

    One KURO export per campaign and one ``well_layout`` per plate is a working
    configuration: ``run_analyze`` scopes each well through ``well_to_sample``
    and reads the designed list only as a ``mutant_id -> labels`` lookup, so the
    200 rows never reach a well. Refusing on the row count alone would block it.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 200)
    params = _analyze_params(tmp_path, expected)
    params["well_layout"] = {"A1": "M1", "B1": "M2", "H12": "WT"}

    result = handle_analyze(params)

    assert result["layout_provenance"]["source"] == "explicit_well_layout"


def test_selected_wells_says_nothing_about_a_layout_it_did_not_build(
    tmp_path: Path,
) -> None:
    """An operator layout is not a selection, and must not be reported as one.

    ``selected_wells`` is a statement about a drafted layout: which of the 96
    wells this campaign occupies. A run handed an explicit ``well_layout`` never
    drafted one, so naming its wells here would put a claim about a choice
    nobody made onto a stored result.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 95)
    params = _analyze_params(tmp_path, expected)
    params["well_layout"] = {"A1": "M1", "H12": "WT"}

    result = handle_analyze(params)

    assert result["layout_provenance"]["source"] == "explicit_well_layout"
    assert result["layout_provenance"]["selected_wells"] is None


# ---------------------------------------------------------------------------
# The button and the run agree
# ---------------------------------------------------------------------------


def _validate_params(tmp_path: Path, expected: Path) -> dict:
    input_dir = tmp_path / "consensus"
    input_dir.mkdir(exist_ok=True)
    reference = tmp_path / "reference.fa"
    reference.write_text(">ref\nATGGGGTTT\n", encoding="utf-8")
    return {
        "input_dir": str(input_dir),
        "reference": str(reference),
        "expected": str(expected),
    }


def test_validate_inputs_refuses_what_the_run_would_refuse(tmp_path: Path) -> None:
    """Otherwise the refusal arrives after a multi-minute demux instead."""
    expected = _write_expected(tmp_path / "expected.xlsx", 120)

    result = handle_validate_inputs(_validate_params(tmp_path, expected))

    assert result["valid"] is False
    assert any("120 designed" in e for e in result["errors"])

    # Same inputs, same answer from the run.
    with pytest.raises(ValueError):
        handle_analyze(_analyze_params(tmp_path, expected))


def test_validate_inputs_accepts_what_the_run_accepts(tmp_path: Path) -> None:
    expected = _write_expected(tmp_path / "expected.xlsx", 120)
    params = _validate_params(tmp_path, expected)
    params["well_layout"] = {"A1": "M1", "H12": "WT"}

    result = handle_validate_inputs(params)

    assert result["valid"] is True
    assert result["errors"] == []


def test_a_run_that_declares_no_wells_stamps_no_selection(tmp_path: Path) -> None:
    """The default has to stay indistinguishable from the pre-selection runs."""
    expected = _write_expected(tmp_path / "expected.xlsx", 95)

    result = handle_analyze(_analyze_params(tmp_path, expected))

    assert result["layout_provenance"]["source"] == "inferred_draft_layout"
    assert result["layout_provenance"]["selected_wells"] is None


def test_a_declared_selection_is_stamped_onto_the_result(tmp_path: Path) -> None:
    """A run that cannot say which wells it declared cannot be reproduced."""
    expected = _write_expected(tmp_path / "expected.xlsx", 2)
    params = _analyze_params(tmp_path, expected)
    # Deliberately out of plate order, and skipping B1: the response has to come
    # back column-major, because that is the order the assignment rule uses.
    params["selected_wells"] = ["D1", "A1", "C1"]

    result = handle_analyze(params)

    assert result["layout_provenance"]["selected_wells"] == ["A1", "C1", "D1"]
    # Two mutants draft onto A1..B1 and the control onto H12. B1 was not
    # declared, so what sits in it is off the plate; C1 and D1 are declared
    # wells the draft never reached.
    assert result["layout_provenance"]["unused_wells"] == ["C1", "D1"]
    assert list(result["layout_provenance"]["excluded_occupants"]) == ["B1", "H12"]


def test_wells_declared_beyond_the_campaign_are_named_rather_than_dropped(
    tmp_path: Path,
) -> None:
    """Selecting more wells than there are samples is not on its own a mistake.

    So it runs. What it must not do is run in silence: the placement rule uses
    up the leading wells and the rest are declarations the result would
    otherwise be unable to distinguish from never having been made.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 2)
    params = _analyze_params(tmp_path, expected)
    params["selected_wells"] = ["A1", "B1", "C1", "D1", "E1"]

    result = handle_analyze(params)

    provenance = result["layout_provenance"]
    # Two mutants take A1 and B1 and the control takes H12; C1, D1 and E1 are
    # declared and unused, and the declaration is reported whole rather than
    # trimmed to fit.
    assert provenance["selected_wells"] == ["A1", "B1", "C1", "D1", "E1"]
    assert provenance["unused_wells"] == ["C1", "D1", "E1"]


def test_a_partly_filled_plate_runs_and_names_what_it_left_out(
    tmp_path: Path,
) -> None:
    """Fewer wells than samples is a description of the bench, not a mistake.

    It used to be a refusal, because the occupants were re-seated onto the
    declared wells and a short list left one with nowhere to go. Placement is
    anchored to the plate now, so an undeclared well simply holds nothing: the
    campaign did not fill it, and what the draft put there is not sequenced.
    The variants left out get no verdict anywhere on the result, so the result
    has to name them.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 4)
    validate = _validate_params(tmp_path, expected)
    validate["selected_wells"] = ["A1", "B1"]

    result = handle_validate_inputs(validate)

    assert result["valid"] is True

    analyze = _analyze_params(tmp_path, expected)
    analyze["selected_wells"] = ["A1", "B1"]

    analyzed = handle_analyze(analyze)

    provenance = analyzed["layout_provenance"]
    assert provenance["selected_wells"] == ["A1", "B1"]
    assert provenance["unused_wells"] == []
    # Four mutants draft onto A1..D1 and the control onto H12. A1 and B1 were
    # declared, so the three occupants past them are off this plate, still under
    # the wells the draft gave them rather than shifted.
    assert list(provenance["excluded_occupants"]) == ["C1", "D1", "H12"]


def test_an_empty_selection_is_reported_as_a_selection_problem(
    tmp_path: Path,
) -> None:
    """The "Clear selection" button sends this, and it was labelled ``expected:``.

    The parameter was parsed inside the block that reports the expected
    workbook, so its message came back attached to a file that is perfectly
    fine and sent the operator to the wrong screen.
    """
    expected = _write_expected(tmp_path / "expected.xlsx", 2)
    params = _validate_params(tmp_path, expected)
    params["selected_wells"] = []

    result = handle_validate_inputs(params)

    assert result["valid"] is False
    assert any(e.startswith("selected_wells is empty") for e in result["errors"])
    assert not any(e.startswith("expected:") for e in result["errors"])


# ---------------------------------------------------------------------------
# Axis counts read back on validate_inputs
# ---------------------------------------------------------------------------

def test_axis_counts_report_the_two_axes_and_the_wells_they_name(
    tmp_path: Path,
) -> None:
    barcodes = _write_full_barcodes(tmp_path / "barcodes.xlsx")

    counts = _barcode_axis_counts(barcodes)

    assert counts == {"forward_count": 12, "reverse_count": 8, "wells": 96}
    # The three numbers state one plate, so they have to multiply out.
    assert counts["forward_count"] * counts["reverse_count"] == counts["wells"]


def test_axis_counts_say_nothing_about_a_set_that_does_not_fit(
    tmp_path: Path,
) -> None:
    """13F x 9R would render "13 x 9 seeds, 96 wells", which does not multiply.

    The counts are every index the file carries; ``wells`` counts only the
    in-range combinations. For a file the layout check already refuses, the two
    disagree, so the line is withheld and the layout error speaks alone.
    """
    path = tmp_path / "oversized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for i in range(1, 14):
        ws.append([f"target_f_{i}", f"AATCCCACT{i:02d}" + _F_TAIL])
    for i in range(1, 10):
        ws.append([f"target_r_{i}", f"CCCTATGA{i:02d}" + _R_TAIL])
    wb.save(str(path))

    assert _barcode_axis_counts(path) is None


def test_axis_counts_are_absent_when_no_barcode_workbook_was_given(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "consensus"
    input_dir.mkdir()
    reference = tmp_path / "reference.fa"
    reference.write_text(">ref\nATGGGGTTT\n", encoding="utf-8")
    expected = _write_expected(tmp_path / "expected.xlsx", 3)

    result = handle_validate_inputs(
        {
            "input_dir": str(input_dir),
            "reference": str(reference),
            "expected": str(expected),
        }
    )

    assert "barcode_axes" not in result


def test_validate_inputs_reports_the_axis_counts_of_the_workbook(
    tmp_path: Path,
) -> None:
    input_dir = tmp_path / "consensus"
    input_dir.mkdir()
    reference = tmp_path / "reference.fa"
    reference.write_text(">ref\nATGGGGTTT\n", encoding="utf-8")
    expected = _write_expected(tmp_path / "expected.xlsx", 3)
    barcodes = _write_full_barcodes(tmp_path / "barcodes.xlsx")

    result = handle_validate_inputs(
        {
            "input_dir": str(input_dir),
            "reference": str(reference),
            "expected": str(expected),
            "custom_barcodes_xlsx": str(barcodes),
        }
    )

    assert result["barcode_axes"] == {
        "forward_count": 12,
        "reverse_count": 8,
        "wells": 96,
    }
