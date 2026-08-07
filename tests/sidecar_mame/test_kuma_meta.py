"""Tests for reading __kuma_meta__ from xlsx."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from kuma_core.mame.io.kuma_meta import KumaMeta, read_kuma_meta
from kuma_core.shared.version import KUMA_VERSION, KURO_MODULE_VERSION


def _make_xlsx_with_meta(path: Path, project_id: str = "abc-123") -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["col"])
    meta = wb.create_sheet("__kuma_meta__")
    meta.sheet_state = "hidden"
    meta.append(["project_id", project_id])
    # The fixture writes what an export actually writes. A hardcoded "0.02.02"
    # round-tripped through a value no exporter has produced since, which is how
    # KUMA_VERSION sat at 0.1.0 while the release manifests reached 0.16.5
    # without a test noticing.
    meta.append(["kuma_version", KUMA_VERSION])
    meta.append(["kuro_module_version", KURO_MODULE_VERSION])
    meta.append(["exported_at", "2026-04-24T00:00:00+00:00"])
    wb.save(path)
    return path


def _make_plain_xlsx(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["only", "data"])
    wb.save(path)
    return path


def test_reads_meta_sheet_if_present(tmp_path):
    xlsx = _make_xlsx_with_meta(tmp_path / "meta.xlsx", project_id="abc-123")
    meta = read_kuma_meta(xlsx)
    assert isinstance(meta, KumaMeta)
    assert meta.project_id == "abc-123"
    assert meta.kuma_version == KUMA_VERSION


def test_returns_none_if_meta_absent(tmp_path):
    xlsx = _make_plain_xlsx(tmp_path / "plain.xlsx")
    assert read_kuma_meta(xlsx) is None
