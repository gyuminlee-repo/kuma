"""RPC-level tests for ``export_janus_mapping`` dest_layout wiring.

``tests/mame/test_janus_mapping.py`` covers the core row builder. This module
covers the handler path the app actually calls, so a param that never reaches
the core (or reaches only one of the two format branches) is caught.
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from kuma_core.mame.models import (
    BarcodeRecord,
    ReplicateResult,
    TranslatedRecord,
    VerdictClass,
    VerdictRecord,
)
from sidecar_mame.core import reset_state, set_last_analyze
from sidecar_mame.handlers.export import (
    handle_export_janus_mapping,
    handle_export_janus_mapping_dry_run,
)


def _make_replicate(mutant_id: str, nb: str, custom: str, size_kb: float) -> ReplicateResult:
    barcode = BarcodeRecord(
        native_barcode=nb,
        custom_barcode=custom,
        consensus_seq="",
        file_size_kb=size_kb,
        source_path=Path("/tmp/mock.fasta"),
    )
    translated = TranslatedRecord(
        barcode=barcode,
        aa_sequence="",
        observed_aa_changes=[],
        observed_nt_changes=[],
    )
    verdict = VerdictRecord(
        translated=translated,
        expected_mutations=[],
        verdict=VerdictClass.PASS,
        verdict_notes="",
    )
    return ReplicateResult(
        mutant_id=mutant_id,
        plate_verdicts={nb: verdict},
        selected_plate=nb,
        selection_reason="pass",
        failed=False,
    )


@pytest.fixture
def seeded_state():
    """Cache two picks at scattered source positions, highest priority first."""
    replicates = [
        _make_replicate("HIGH", "NB01", "5_7", 300.0),   # E7
        _make_replicate("LOW", "NB02", "8_12", 10.0),    # H12
    ]
    set_last_analyze([], replicates, "/tmp/out.xlsx", run_meta=None)
    yield replicates
    reset_state()


# The handler defaults to the instrument-native eight column sheet with a
# compact destination layout. Cases that assert the kuma-internal 5-column
# output at the source position pass this explicitly; the defaults themselves
# are asserted by the test_default_* cases below.
_LEGACY_SOURCE = {"output_schema": "legacy5", "dest_layout": "source"}


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def test_source_layout_mirrors_source(tmp_path: Path, seeded_state) -> None:
    out = tmp_path / "source.csv"
    handle_export_janus_mapping({"output": str(out), **_LEGACY_SOURCE})
    rows = _read_csv(out)
    assert [r["dest_well"] for r in rows] == ["E7", "H12"]


def test_default_layout_is_compact(tmp_path: Path, seeded_state) -> None:
    """A stock plate is a new plate, so the default fills it from A1."""
    out = tmp_path / "default.csv"
    handle_export_janus_mapping(
        {"output": str(out), "output_schema": "legacy5", "liquid_class": "Cell"}
    )
    rows = _read_csv(out)
    assert [r["dest_well"] for r in rows] == ["A1", "B1"]
    assert [r["source_well"] for r in rows] == ["E7", "H12"]


def test_compact_layout_reaches_core_via_csv(tmp_path: Path, seeded_state) -> None:
    out = tmp_path / "compact.csv"
    handle_export_janus_mapping(
        {"output": str(out), "dest_layout": "compact", "output_schema": "legacy5"}
    )
    rows = _read_csv(out)
    assert [r["dest_well"] for r in rows] == ["A1", "B1"]
    assert [r["source_well"] for r in rows] == ["E7", "H12"]


def test_compact_layout_reaches_core_via_xlsx(tmp_path: Path, seeded_state) -> None:
    out = tmp_path / "compact.xlsx"
    handle_export_janus_mapping(
        {
            "output": str(out),
            "format": "xlsx",
            "dest_layout": "compact",
            "output_schema": "legacy5",
        }
    )
    ws = openpyxl.load_workbook(out)["Janus Mapping"]
    header = [c.value for c in ws[1]]
    dest_col = header.index("dest_well")
    dest = [row[dest_col].value for row in ws.iter_rows(min_row=2)]
    assert dest == ["A1", "B1"]


def test_invalid_dest_layout_rejected(tmp_path: Path, seeded_state) -> None:
    out = tmp_path / "bad.csv"
    with pytest.raises(ValueError, match="Invalid dest_layout"):
        handle_export_janus_mapping({"output": str(out), "dest_layout": "grid"})
    assert not out.exists(), "invalid params must not produce an output file"


@pytest.mark.parametrize(
    "params",
    [
        {"source_racks": {"NB01": 1, "NB02": 2}},
        {"source_racks": {"NB01": 1.9}},
        {"dest_rack": 4},
        {"dest_rack": 4.7},
    ],
)
def test_a_stale_numeric_deck_is_dropped_rather_than_written(
    tmp_path: Path,
    seeded_state,
    params,
) -> None:
    """A deck saved when the columns held numbers must not reach the sheet.

    The handler is the wire edge, so it faces a client that stored its settings
    under the old format. Writing ``str(1)`` would put a bare number where the
    robot expects a labware name and every gate would stay green, so the stale
    value is dropped and the generated plate name stands. Failing the export
    instead would punish the operator for what an older build saved.

    ``JanusSettings`` does the opposite with the same value and raises, because
    in-process an integer rack is a programming error rather than old state.
    That pair is pinned in tests/mame/test_janus_policy.py.
    """
    out = tmp_path / "stale-rack.csv"
    handle_export_janus_mapping({
        "output": str(out),
        "liquid_class": "Cell 100ul",
        **params,
    })

    with out.open(encoding="utf-8") as fh:
        rows = list(csv.reader(fh))[1:]
    assert rows
    assert {r[3] for r in rows} == {"Stock plate1", "Stock plate2"}
    assert {r[5] for r in rows} == {"final culture plate"}


def test_a_project_saved_under_the_former_schema_name_still_resolves(
    tmp_path: Path, seeded_state
) -> None:
    """``device9`` named the instrument schema before it lost its ninth column.

    A saved project, and a request already in flight when the build changed
    under it, ask for the sheet by that name. It is the same sheet, so the old
    name is folded into the current one at the wire edge rather than being
    rejected by the whitelist. Without this the dialog raises on a value the
    client itself stored, so the export dies at the point the operator uses it.
    """
    out = tmp_path / "former-name.csv"
    result = handle_export_janus_mapping(
        {"output": str(out), "output_schema": "device9"}
    )

    assert Path(result["output_path"]).exists()
    with out.open(encoding="utf-8") as fh:
        rows = list(csv.reader(fh))
    assert len(rows[0]) == 8
    assert result["settings"]["output_schema"] == "device"

    # The same file the current name writes, byte for byte.
    current = tmp_path / "current-name.csv"
    handle_export_janus_mapping({"output": str(current), "output_schema": "device"})
    assert out.read_bytes() == current.read_bytes()


def test_the_former_schema_name_does_not_leak_into_the_pick_list_name(
    tmp_path: Path, seeded_state
) -> None:
    """Folding the old name must not fold ``legacy5`` with it.

    ``legacy5`` did not change, and it is the schema the automatic pick list
    pins, so a normalisation that treated any stored schema string loosely would
    silently turn the pick list into an instrument worklist.
    """
    out = tmp_path / "legacy.csv"
    result = handle_export_janus_mapping(
        {"output": str(out), "output_schema": "legacy5"}
    )
    assert result["settings"]["output_schema"] == "legacy5"
    with out.open(encoding="utf-8") as fh:
        assert next(csv.reader(fh)) == [
            "name", "source_plate", "source_well", "dest_well", "priority_score",
        ]


def test_duplicate_dest_surfaces_through_handler(tmp_path: Path) -> None:
    set_last_analyze(
        [],
        [
            _make_replicate("P1_A1", "NB01", "1_1", 200.0),
            _make_replicate("P2_A1", "NB02", "1_1", 100.0),
        ],
        "/tmp/out.xlsx",
        run_meta=None,
    )
    try:
        out = tmp_path / "dup.csv"
        with pytest.raises(ValueError, match="duplicate dest_well"):
            handle_export_janus_mapping({"output": str(out), **_LEGACY_SOURCE})
    finally:
        reset_state()


# ---------------------------------------------------------------------------
# Dry-run preview handler
# ---------------------------------------------------------------------------


def test_dry_run_requires_a_prior_analyze() -> None:
    reset_state()
    with pytest.raises(RuntimeError, match="No prior analyze result"):
        handle_export_janus_mapping_dry_run({})


def test_dry_run_returns_rows_without_writing(tmp_path: Path, seeded_state) -> None:
    result = handle_export_janus_mapping_dry_run(dict(_LEGACY_SOURCE))
    assert result["row_count"] == 2
    assert [r["dest_well"] for r in result["rows"]] == ["E7", "H12"]
    assert result["errors"] == []
    assert result["excluded"] == []
    assert result["excluded_count"] == 0
    assert list(tmp_path.iterdir()) == [], "dry run must not create files"


def test_dry_run_passes_dest_layout_to_core(seeded_state) -> None:
    result = handle_export_janus_mapping_dry_run({"dest_layout": "compact"})
    assert [r["dest_well"] for r in result["rows"]] == ["A1", "B1"]
    assert [r["source_well"] for r in result["rows"]] == ["E7", "H12"]


def test_dry_run_treats_null_dest_layout_as_compact(seeded_state) -> None:
    """An explicit JSON null must land on the default, not on a rejected value."""
    result = handle_export_janus_mapping_dry_run({"dest_layout": None})
    assert [r["dest_well"] for r in result["rows"]] == ["A1", "B1"]


def test_dry_run_rejects_invalid_dest_layout(seeded_state) -> None:
    with pytest.raises(ValueError, match="Invalid dest_layout"):
        handle_export_janus_mapping_dry_run({"dest_layout": "grid"})


def test_dry_run_reports_duplicate_instead_of_raising() -> None:
    set_last_analyze(
        [],
        [
            _make_replicate("P1_A1", "NB01", "1_1", 200.0),
            _make_replicate("P2_A1", "NB02", "1_1", 100.0),
        ],
        "/tmp/out.xlsx",
        run_meta=None,
    )
    try:
        result = handle_export_janus_mapping_dry_run(dict(_LEGACY_SOURCE))
        assert [e["code"] for e in result["errors"]] == ["duplicate_dest_well"]
        assert result["errors"][0]["mutant_ids"] == ["P1_A1", "P2_A1"]
        # Same state, compact layout: the documented way out reports clean.
        compact = handle_export_janus_mapping_dry_run(
            {"dest_layout": "compact", "output_schema": "legacy5"}
        )
        assert compact["errors"] == []
    finally:
        reset_state()
