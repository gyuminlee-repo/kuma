"""What a run does once the sample map is gone: declare wells, report strays.

Three things that used to be somebody else's job, checked at the handler:

* ``check_plate_order`` runs inside ``handle_analyze``, not only behind the
  validate button. Its only caller was ``handle_validate_inputs``, which left
  the run itself defended by nothing but the frontend's ``selectCanRun``, so a
  CLI call, a harness or a script scored every well against whichever of a
  workbook's two plate statements the reader happened to pick.
* A declared selection places the variants and is stamped onto the result.
* Reads from wells the layout does not name are counted and reported, never
  refused: the same counts appear for a well the operator declared empty and
  for barcode crosstalk, and nothing in the number separates the two.

Plus the migration path. An existing project still has a filled-in
``sample_map_template.xlsx`` on disk, and ignoring it is the one option that
cannot be defended: if it disagrees with the computed draft, one of the two
describes the tubes that were pipetted.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from sidecar_mame.handlers.analyze import handle_analyze, handle_validate_inputs

# Reference ATG GGG TTT -> M G F. Well A01 ("1_1") observes G2A.
_REFERENCE_NT = "ATGGGGTTT"
_G2A_NT = "ATGGCGTTT"
_PAD = "\n" * (52 * 1024)


def _write_fasta(path: Path, header: str, body: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(f">{header} depth=100\n{body}\n{_PAD}", encoding="utf-8")


def _reference(tmp_path: Path) -> Path:
    ref = tmp_path / "reference.fa"
    ref.write_text(f">ref\n{_REFERENCE_NT}\n", encoding="utf-8")
    return ref


def _variant_list(path: Path, variants: list[str]) -> Path:
    """A plain variant list: one column, one variant per row, in plate order."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["variant"])
    for variant in variants:
        ws.append([variant])
    wb.save(str(path))
    return path


def _plate_workbook(path: Path, plate_rows: list[tuple[str, str]], expected_ids: list[str]) -> Path:
    """A KURO export stating its plate twice: a primer sheet and expected_mutations."""
    wb = openpyxl.Workbook()
    first = wb.worksheets[0]
    first.title = "Fwd List"
    first.append(["Well", "Primer Name", "Mutation"])
    for well, mutation in plate_rows:
        first.append([well, f"{mutation}_F", mutation])
    sheet = wb.create_sheet("expected_mutations")
    sheet.append([
        "mutant_id", "position", "wt_aa", "mt_aa", "wt_codon",
        "mt_codon", "group_id", "primer_set_ref", "notation_type", "status",
    ])
    for mutant_id in expected_ids:
        sheet.append([
            mutant_id, int(mutant_id[1:-1]), mutant_id[0], mutant_id[-1],
            "", "", "", "", "substitution", "DESIGNED",
        ])
    wb.save(str(path))
    return path


def _params(tmp_path: Path, expected: Path, *, with_reads: bool = False) -> dict:
    ingest = tmp_path / "consensus"
    ingest.mkdir(exist_ok=True)
    if with_reads:
        # A01 ("1_1") and C01 ("3_1"), both carrying the G2A consensus.
        _write_fasta(ingest / "NB01" / "1_1.fasta", "1_1", _G2A_NT)
        _write_fasta(ingest / "NB01" / "3_1.fasta", "3_1", _G2A_NT)
    return {
        "input_dir": str(ingest),
        "reference": str(_reference(tmp_path)),
        "expected": str(expected),
        "output": str(tmp_path / "result.xlsx"),
        "cds_start": 0,
        "cds_end": 9,
        "min_file_size_kb": 0.0,
        "min_read_count": 0,
        "ingest_mode": "barcode",
    }


# ---------------------------------------------------------------------------
# The plate-order refusal reaches the run, not just the button
# ---------------------------------------------------------------------------

def test_analyze_refuses_a_workbook_that_states_its_plate_twice(tmp_path: Path) -> None:
    """The gap: the run had no plate-order defence of its own.

    ``handle_validate_inputs`` was the only caller of ``check_plate_order``, and
    the frontend refused to enable the button on its finding. Everything that
    does not go through that button, which is every scripted path, ran.
    """
    expected = _plate_workbook(
        tmp_path / "reordered.xlsx",
        [("A1", "S11I"), ("B1", "S22T")],
        ["S22T", "S11I"],
    )

    with pytest.raises(ValueError) as excinfo:
        handle_analyze(_params(tmp_path, expected))

    message = str(excinfo.value)
    assert "Fwd List" in message
    assert "expected_mutations" in message


def test_the_run_and_the_button_refuse_the_same_workbook(tmp_path: Path) -> None:
    """Two defences that disagree are one defence and one false reassurance."""
    expected = _plate_workbook(
        tmp_path / "reordered2.xlsx",
        [("A1", "S11I"), ("B1", "S22T")],
        ["S22T", "S11I"],
    )
    params = _params(tmp_path, expected)

    validation = handle_validate_inputs({
        key: params[key] for key in ("input_dir", "reference", "expected")
    })

    assert validation["valid"] is False
    with pytest.raises(ValueError):
        handle_analyze(params)


def test_an_agreeing_workbook_still_runs(tmp_path: Path) -> None:
    """The check has to be silent on the file it is meant to let through."""
    expected = _plate_workbook(
        tmp_path / "ordered.xlsx",
        [("A1", "S11I"), ("B1", "S22T")],
        ["S11I", "S22T"],
    )

    result = handle_analyze(_params(tmp_path, expected))

    assert result["layout_provenance"]["source"] == "inferred_draft_layout"


# ---------------------------------------------------------------------------
# Reads from wells the layout does not name
# ---------------------------------------------------------------------------

def test_reads_from_an_undeclared_well_are_counted_and_named(tmp_path: Path) -> None:
    """Reported, not refused: this is also the shape barcode crosstalk makes."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A"])
    params = _params(tmp_path, expected, with_reads=True)
    # G2A plus WT occupies two wells; declare A1 and B1, so C01 is off-layout.
    params["selected_wells"] = ["A1", "B1"]

    result = handle_analyze(params)

    off = result["off_layout_records"]
    assert off["count"] == 1
    assert off["wells"] == [{"well": "C1", "records": 1}]


def test_reads_from_wells_the_layout_names_are_not_strays(tmp_path: Path) -> None:
    """The count has to be silent on the ordinary case or it says nothing."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W", "T5A"])

    result = handle_analyze(_params(tmp_path, expected, with_reads=True))

    # Three variants plus WT fill A1..D1, so both reads land inside the layout.
    assert result["off_layout_records"] == {"count": 0, "wells": []}


def test_a_run_that_declares_nothing_still_has_a_layout_to_be_outside_of(
    tmp_path: Path,
) -> None:
    """Absent selection is not "the whole plate", it is the leading N+1 wells.

    So a read from beyond them is a stray even with nothing declared. That is
    the same read that used to land in an ``UNKNOWN_*`` group with nothing on
    the result to say a well nobody expected had produced data.
    """
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A"])

    result = handle_analyze(_params(tmp_path, expected, with_reads=True))

    assert result["layout_provenance"]["selected_wells"] is None
    assert result["off_layout_records"]["wells"] == [{"well": "C1", "records": 1}]


def test_an_empty_selection_is_refused_rather_than_treated_as_absent(
    tmp_path: Path,
) -> None:
    """A run with no wells has nothing to score, and absent already means "all"."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A"])
    params = _params(tmp_path, expected)
    params["selected_wells"] = []

    with pytest.raises(ValueError, match="omit the parameter"):
        handle_analyze(params)


def test_a_malformed_selection_is_refused_rather_than_ignored(tmp_path: Path) -> None:
    """This parameter decides where every variant lands; a silent default moves them."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A"])
    params = _params(tmp_path, expected)
    params["selected_wells"] = "A1"

    with pytest.raises(ValueError, match="list of well ids"):
        handle_analyze(params)


# ---------------------------------------------------------------------------
# An existing project's sample map, compared rather than ignored
# ---------------------------------------------------------------------------

def _legacy_sample_map(path: Path, rows: list[tuple[str, str]]) -> Path:
    """The pre-removal shape: column A sample name, column B well position."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for sample, well in rows:
        ws.append([sample, well])
    wb.save(str(path))
    return path


def _validate_params(tmp_path: Path, expected: Path) -> dict:
    ingest = tmp_path / "consensus"
    ingest.mkdir(exist_ok=True)
    return {
        "input_dir": str(ingest),
        "reference": str(_reference(tmp_path)),
        "expected": str(expected),
    }


def test_an_agreeing_legacy_sample_map_is_reported_and_lets_the_run_through(
    tmp_path: Path,
) -> None:
    """Agreement is worth saying once, so the operator can delete the file."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W"])
    params = _validate_params(tmp_path, expected)
    # Unpadded on purpose: A1 has to normalise onto the draft's A1.
    params["legacy_sample_map_xlsx"] = str(
        _legacy_sample_map(tmp_path / "map.xlsx", [("G2A", "A1"), ("F3W", "B1"), ("WT", "C1")])
    )

    result = handle_validate_inputs(params)

    assert result["valid"] is True
    assert result["legacy_sample_map"]["status"] == "matches"
    assert result["legacy_sample_map"]["differences"] == []


def test_a_disagreeing_legacy_sample_map_names_the_wells_and_blocks(
    tmp_path: Path,
) -> None:
    """One of the two describes the tubes, and the file does not say which."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W"])
    params = _validate_params(tmp_path, expected)
    # The bench sheet says the two variants are the other way round.
    params["legacy_sample_map_xlsx"] = str(
        _legacy_sample_map(tmp_path / "map.xlsx", [("F3W", "A1"), ("G2A", "B1"), ("WT", "C1")])
    )

    result = handle_validate_inputs(params)

    assert result["valid"] is False
    finding = result["legacy_sample_map"]
    assert finding["status"] == "differs"
    assert {d["well"] for d in finding["differences"]} == {"A01", "B01"}
    assert finding["differences"][0] == {"well": "A01", "file": "F3W", "draft": "G2A"}
    assert any("sample map" in message for message in result["errors"])


def test_the_legacy_comparison_reads_the_selection_the_run_would_use(
    tmp_path: Path,
) -> None:
    """Comparing against the unselected draft would flag a plate that is correct."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W"])
    params = _validate_params(tmp_path, expected)
    # The draft is A1=G2A, B1=F3W, C1=WT. Declaring all but B1 says F3W was not
    # pipetted, and the map on disk agrees: it names the two wells that were.
    # Compared against the draft instead, B1 would read as a disagreement.
    params["selected_wells"] = ["A1", "C1"]
    params["legacy_sample_map_xlsx"] = str(
        _legacy_sample_map(tmp_path / "map.xlsx", [("G2A", "A1"), ("WT", "C1")])
    )

    result = handle_validate_inputs(params)

    assert result["legacy_sample_map"]["status"] == "matches"
    assert result["valid"] is True


def test_the_run_refuses_a_disagreeing_sample_map_too(tmp_path: Path) -> None:
    """The comparison was behind the validate button and nowhere else.

    So a CLI call, a harness, a script, or an operator who pressed Run without
    validating scored the plate with a file on disk saying it was a different
    plate. Refused here before the demux, on the same comparison and with the
    same message.
    """
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W"])
    params = _params(tmp_path, expected)
    params["legacy_sample_map_xlsx"] = str(
        _legacy_sample_map(tmp_path / "map.xlsx", [("F3W", "A1"), ("G2A", "B1"), ("WT", "C1")])
    )

    with pytest.raises(ValueError, match="sample map"):
        handle_analyze(params)


def test_the_run_proceeds_when_the_sample_map_agrees(tmp_path: Path) -> None:
    """The comparison has to be silent on the file it is meant to let through."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W"])
    params = _params(tmp_path, expected)
    params["legacy_sample_map_xlsx"] = str(
        _legacy_sample_map(tmp_path / "map.xlsx", [("G2A", "A1"), ("F3W", "B1"), ("WT", "C1")])
    )

    result = handle_analyze(params)

    assert result["layout_provenance"]["source"] == "inferred_draft_layout"


def test_the_run_compares_the_map_against_the_wells_it_would_use(
    tmp_path: Path,
) -> None:
    """The selection has to be applied before the comparison, not after it."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A", "F3W"])
    params = _params(tmp_path, expected)
    params["selected_wells"] = ["A1", "C1"]
    params["legacy_sample_map_xlsx"] = str(
        _legacy_sample_map(tmp_path / "map.xlsx", [("G2A", "A1"), ("WT", "C1")])
    )

    result = handle_analyze(params)

    assert result["layout_provenance"]["selected_wells"] == ["A1", "C1"]
    assert result["layout_provenance"]["excluded_occupants"] == {"B1": "F3W"}


def test_a_project_with_no_sample_map_says_nothing_about_one(tmp_path: Path) -> None:
    """Every project created from here on, so the key must simply be absent."""
    expected = _variant_list(tmp_path / "variants.xlsx", ["G2A"])

    result = handle_validate_inputs(_validate_params(tmp_path, expected))

    assert "legacy_sample_map" not in result
