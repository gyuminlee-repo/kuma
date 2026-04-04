"""96-well plate mapping and Excel export for SDM primers."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .sdm_engine import SdmPrimerResult


@dataclass
class PlateMapping:
    """Single well assignment in a 96-well plate."""

    well: str               # e.g. "A1", "B1"
    primer_name: str        # e.g. "Q232A_F" or "Q232A_R"
    sequence: str           # Primer sequence
    primer_type: str        # "forward" or "reverse"
    mutation: str           # Original mutation notation
    tm: float | None = None          # Whole-primer Tm (Fwd or Rev)
    tm_overlap: float | None = None  # Overlap Tm
    wt_codon: str | None = None      # Wild-type codon
    mt_codon: str | None = None      # Mutant codon


def _well_name(index: int, order: str = "column") -> str:
    """Convert a 0-based index to a well name.

    Args:
        index: 0-based well index.
        order: 'column' (A1->H1->A2) or 'row' (A1->A12->B1).

    Returns:
        Well name like "A1".
    """
    rows = "ABCDEFGH"
    if order == "column":
        col = index // 8 + 1
        row = index % 8
    else:  # row
        row = index // 12
        col = index % 12 + 1
    if row >= 8 or col > 12:
        raise ValueError(f"Well index {index} exceeds 96-well plate capacity")
    return f"{rows[row]}{col}"


def _assign_well(index: int, well_order: str = "column") -> str:
    """Generate well name with overflow to additional plates."""
    plate_num = index // 96
    local_idx = index % 96
    well = _well_name(local_idx, well_order)
    if plate_num == 0:
        return well
    return f"P{plate_num + 1}-{well}"


def _parse_well_plate(well: str) -> tuple[int, str]:
    """Return (0-based plate index, base well) from an overflow well label.

    ``"A1"`` → ``(0, "A1")``,  ``"P2-A1"`` → ``(1, "A1")``.
    Inverse of ``_assign_well``.
    """
    if "-" in well:
        prefix, base = well.split("-", 1)
        return int(prefix[1:]) - 1, base
    return 0, well


def deduplicate_reverse(
    results: list[SdmPrimerResult],
) -> dict[str, list[str]]:
    """Find mutations sharing identical reverse primers.

    Returns:
        Dict mapping reverse primer sequence to list of mutation names.
    """
    rev_map: dict[str, list[str]] = {}
    for r in results:
        rev_seq = r.reverse_seq
        if rev_seq not in rev_map:
            rev_map[rev_seq] = []
        rev_map[rev_seq].append(r.mutation.raw)
    return rev_map


def generate_plate_map(
    results: list[SdmPrimerResult],
    well_order: str = "column",
    deduplicate_rev: bool = True,
) -> tuple[list[PlateMapping], list[PlateMapping]]:
    """Generate separate Fwd and Rev plate mappings.

    Forward plate: one primer per mutation, sequential well assignment.
    Reverse plate: deduplicated reverse primers, sequential well assignment.

    Returns:
        Tuple of (fwd_mappings, rev_mappings).
    """
    # Forward plate
    fwd_mappings: list[PlateMapping] = []
    for idx, r in enumerate(results):
        well = _assign_well(idx, well_order)
        fwd_mappings.append(PlateMapping(
            well=well,
            primer_name=f"{r.mutation.raw}_F",
            sequence=r.forward_seq,
            primer_type="forward",
            mutation=r.mutation.raw,
        ))

    # Reverse plate (deduplicated)
    rev_mappings: list[PlateMapping] = []
    if deduplicate_rev:
        rev_groups = deduplicate_reverse(results)
        for idx, (rev_seq, mut_names) in enumerate(rev_groups.items()):
            label = mut_names[0]
            well = _assign_well(idx, well_order)
            rev_mappings.append(PlateMapping(
                well=well,
                primer_name=f"{label}_R",
                sequence=rev_seq,
                primer_type="reverse",
                mutation=label,
            ))
    else:
        for idx, r in enumerate(results):
            well = _assign_well(idx, well_order)
            rev_mappings.append(PlateMapping(
                well=well,
                primer_name=f"{r.mutation.raw}_R",
                sequence=r.reverse_seq,
                primer_type="reverse",
                mutation=r.mutation.raw,
            ))

    return fwd_mappings, rev_mappings


def _is_shared_rev(m: PlateMapping, rev_groups: dict[str, list[str]] | None) -> bool:
    """Check if a reverse primer is shared by multiple mutations."""
    if m.primer_type != "reverse" or rev_groups is None:
        return False
    return len(rev_groups.get(m.sequence, [])) > 1


# Colors matching the UI: Fwd=green, Rev=orange, Shared=blue
_COLOR_FWD = "C6EFCE"       # light green
_COLOR_REV = "FCE4D6"       # light orange
_COLOR_SHARED = "BDD7EE"    # light blue


def _write_list_sheet(
    ws,
    mappings: list[PlateMapping],
    fill_color: str,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Write a primer list sheet with row coloring."""
    from openpyxl.styles import Font, PatternFill

    headers = ["Well", "Primer Name", "Sequence", "Length",
               "Tm", "Tm_Overlap", "WT_Codon", "MT_Codon", "Mutation"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")

    num_cols = len(headers)
    for i, m in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=m.well)
        ws.cell(row=i, column=2, value=m.primer_name)
        ws.cell(row=i, column=3, value=m.sequence)
        ws.cell(row=i, column=4, value=len(m.sequence))
        ws.cell(row=i, column=5, value=m.tm)
        ws.cell(row=i, column=6, value=m.tm_overlap)
        ws.cell(row=i, column=7, value=m.wt_codon)
        ws.cell(row=i, column=8, value=m.mt_codon)
        ws.cell(row=i, column=9, value=m.mutation)

        # Row color: shared=blue, otherwise fill_color
        row_color = _COLOR_SHARED if _is_shared_rev(m, rev_groups) else fill_color
        row_fill = PatternFill(start_color=row_color, fill_type="solid")
        for col_idx in range(1, num_cols + 1):
            ws.cell(row=i, column=col_idx).fill = row_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def _write_plate_sheet(
    ws,
    mappings: list[PlateMapping],
    fill_color: str,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Write a 96-well plate layout sheet with shared primers in blue."""
    from openpyxl.styles import Alignment, Font, PatternFill

    rows_label = "ABCDEFGH"

    for c in range(1, 13):
        cell = ws.cell(row=1, column=c + 1, value=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r_idx, r_label in enumerate(rows_label):
        ws.cell(row=r_idx + 2, column=1, value=r_label).font = Font(bold=True)

    default_fill = PatternFill(start_color=fill_color, fill_type="solid")
    shared_fill = PatternFill(start_color=_COLOR_SHARED, fill_type="solid")

    for m in mappings:
        if not m.well[0].isalpha():
            continue
        row_letter = m.well[0]
        col_num = int(m.well[1:])
        if row_letter not in rows_label or col_num > 12:
            continue
        r_idx = rows_label.index(row_letter) + 2
        c_idx = col_num + 1

        cell = ws.cell(row=r_idx, column=c_idx, value=m.primer_name)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = shared_fill if _is_shared_rev(m, rev_groups) else default_fill

    ws.column_dimensions["A"].width = 4
    for c in range(2, 14):
        ws.column_dimensions[chr(64 + c)].width = 14


def _chunk_by_plate(mappings: list[PlateMapping]) -> list[list[PlateMapping]]:
    """Split mappings into 96-well plate chunks."""
    plates: list[list[PlateMapping]] = []
    current: list[PlateMapping] = []
    for m in mappings:
        # Normalize well: strip overflow prefix like "P2-"
        well = m.well
        if "-" in well:
            well = well.split("-", 1)[1]
            m = PlateMapping(
                well=well,
                primer_name=m.primer_name,
                sequence=m.sequence,
                primer_type=m.primer_type,
                mutation=m.mutation,
                tm=m.tm,
                tm_overlap=m.tm_overlap,
                wt_codon=m.wt_codon,
                mt_codon=m.mt_codon,
            )
        current.append(m)
        if len(current) >= 96:
            plates.append(current)
            current = []
    if current:
        plates.append(current)
    return plates if plates else [[]]


def _pair_rev_per_plate(
    fwd_plates: list[list[PlateMapping]],
    rev_all: list[PlateMapping],
    rev_groups: dict[str, list[str]] | None,
) -> list[list[PlateMapping]]:
    """Pair reverse primers with each fwd plate by mutation membership.

    For each fwd plate chunk, collect the deduplicated reverse primers
    that belong to that chunk's mutations, reassigning well names.
    """
    # Build mutation → rev sequence lookup
    mut_to_rev_seq: dict[str, str] = {}
    if rev_groups:
        for seq, muts in rev_groups.items():
            for mut in muts:
                mut_to_rev_seq[mut] = seq

    rev_by_seq = {r.sequence: r for r in rev_all}

    paired: list[list[PlateMapping]] = []
    for fwd_chunk in fwd_plates:
        seen_seq: set[str] = set()
        rev_chunk: list[PlateMapping] = []
        well_idx = 0
        for fwd_m in fwd_chunk:
            rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
            if rev_seq and rev_seq not in seen_seq:
                seen_seq.add(rev_seq)
                rev_m = rev_by_seq.get(rev_seq)
                if rev_m:
                    rev_chunk.append(PlateMapping(
                        well=_well_name(well_idx),
                        primer_name=rev_m.primer_name,
                        sequence=rev_m.sequence,
                        primer_type="reverse",
                        mutation=rev_m.mutation,
                        tm=rev_m.tm,
                        tm_overlap=rev_m.tm_overlap,
                        wt_codon=rev_m.wt_codon,
                        mt_codon=rev_m.mt_codon,
                    ))
                    well_idx += 1
        paired.append(rev_chunk)
    return paired


def export_plate_excel(
    mappings: list[PlateMapping],
    output_path: Path,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Export plate mappings to Excel with per-plate sheets.

    Forward and reverse primers are paired per plate: Rev Plate N contains
    only the reverse primers for Fwd Plate N's mutations.

    For N plates, creates sheets:
    - 'Fwd List 1', 'Fwd Plate 1', 'Rev List 1', 'Rev Plate 1'
    - 'Fwd List 2', 'Fwd Plate 2', ... (if multi-plate)

    Single plate omits the number suffix.

    Args:
        rev_groups: Original reverse deduplication map (seq -> mutation names).
            Used to pair rev with fwd plates and detect shared primers.
    """
    from openpyxl import Workbook

    fwd_all = [m for m in mappings if m.primer_type == "forward"]
    rev_all = [m for m in mappings if m.primer_type == "reverse"]

    fwd_plates = _chunk_by_plate(fwd_all)
    rev_plates = _pair_rev_per_plate(fwd_plates, rev_all, rev_groups)
    plate_count = len(fwd_plates)
    suffix = plate_count > 1

    wb = Workbook()
    first_sheet = True
    chunk_rev_groups = rev_groups or {}

    for i in range(plate_count):
        tag = f" {i + 1}" if suffix else ""
        fwd_chunk = fwd_plates[i]
        rev_chunk = rev_plates[i] if i < len(rev_plates) else []

        if first_sheet:
            ws = wb.active
            ws.title = f"Fwd List{tag}"
            first_sheet = False
        else:
            ws = wb.create_sheet(f"Fwd List{tag}")
        _write_list_sheet(ws, fwd_chunk, _COLOR_FWD)

        ws = wb.create_sheet(f"Fwd Plate{tag}")
        _write_plate_sheet(ws, fwd_chunk, _COLOR_FWD)

        ws = wb.create_sheet(f"Rev List{tag}")
        _write_list_sheet(ws, rev_chunk, _COLOR_REV, rev_groups=chunk_rev_groups)

        ws = wb.create_sheet(f"Rev Plate{tag}")
        _write_plate_sheet(ws, rev_chunk, _COLOR_REV, rev_groups=chunk_rev_groups)

    wb.save(output_path)


def export_idt_csv(
    results: list[SdmPrimerResult],
    output_path: Path,
    scale: str = "25nm",
    purification: str = "STD",
) -> None:
    """Export primer order CSV in IDT OligoEntry format.

    CSV columns: Name, Sequence, Scale, Purification.
    Each mutation produces two rows (forward and reverse).

    Args:
        results: List of SdmPrimerResult from primer design.
        output_path: Path to output CSV file.
        scale: Synthesis scale (default: "25nm").
        purification: Purification type (default: "STD").
    """
    import csv

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Sequence", "Scale", "Purification"])
        for r in results:
            writer.writerow([f"{r.mutation.raw}_F", r.forward_seq, scale, purification])
            writer.writerow([f"{r.mutation.raw}_R", r.reverse_seq, scale, purification])


def export_twist_csv(
    results: list[SdmPrimerResult],
    output_path: Path,
) -> None:
    """Export primer order CSV in Twist Bioscience bulk order format.

    CSV columns: Name, Sequence, Notes.
    Each mutation produces two rows (forward and reverse).

    Args:
        results: List of SdmPrimerResult from primer design.
        output_path: Path to output CSV file.
    """
    import csv

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Sequence", "Notes"])
        for r in results:
            writer.writerow([f"{r.mutation.raw}_F", r.forward_seq, r.mutation.raw])
            writer.writerow([f"{r.mutation.raw}_R", r.reverse_seq, r.mutation.raw])


# ---------------------------------------------------------------------------
# Liquid handler mapping exports
# ---------------------------------------------------------------------------

_ROWS_384 = "ABCDEFGHIJKLMNOP"
_ROWS_96 = "ABCDEFGH"


def _to_384_well_fwd(well_96: str) -> str:
    """Map 96-well address to 384-well forward primer position (odd rows A,C,E,G,I,K,M,O)."""
    row_idx = _ROWS_96.index(well_96[0])
    return f"{_ROWS_384[row_idx * 2]}{well_96[1:]}"


def _to_384_well_rev(well_96: str) -> str:
    """Map 96-well address to 384-well reverse primer position (even rows B,D,F,H,J,L,N,P)."""
    row_idx = _ROWS_96.index(well_96[0])
    return f"{_ROWS_384[row_idx * 2 + 1]}{well_96[1:]}"


def export_echo_mapping_csv(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: int = 100,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Export Echo 525 acoustic dispenser mapping CSV.

    Source plate layout: 384-well (Eco 384PP).
      - Forward primers occupy odd rows (A, C, E, G, I, K, M, O).
      - Reverse primers occupy even rows (B, D, F, H, J, L, N, P).
      - Both use column-first well ordering matching the 96-well plate layout.

    Each row in the output corresponds to one transfer event.
    Shared reverse primers produce one row per destination well.

    Args:
        fwd_mappings: Forward primer plate mappings (96-well coordinates).
        rev_mappings: Deduplicated reverse primer plate mappings (96-well).
        output_path: Output CSV file path.
        transfer_vol: Transfer volume in nL (default 100).
        rev_groups: Reverse deduplication map {seq: [mutation_names]}.
            Used to expand shared primers to all destination wells.
    """
    import csv

    fwd_by_mut, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Source Plate Name", "Source Well Name", "Source Well",
            "Dest Plate Name", "Dest Well Name", "Dest Well", "Transfer Vol",
        ])

        # Forward: one row per mutation
        for m in fwd_mappings:
            plate_idx, base_well = _parse_well_plate(m.well)
            src_plate = f"Source [{plate_idx + 1}]"
            dest_plate = f"Destination [{plate_idx + 1}]"
            src_well = _to_384_well_fwd(base_well)
            writer.writerow([
                src_plate, m.primer_name, src_well,
                dest_plate, m.mutation, base_well, transfer_vol,
            ])

        # Reverse: one row per (primer, dest_well) pair
        for fwd_m in fwd_mappings:
            rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
            if rev_seq is None:
                continue
            rev_m = rev_by_seq.get(rev_seq)
            if rev_m is None:
                continue

            fwd_plate_idx, _ = _parse_well_plate(fwd_m.well)
            src_plate = f"Source [{fwd_plate_idx + 1}]"
            dest_plate = f"Destination [{fwd_plate_idx + 1}]"
            _, rev_base_well = _parse_well_plate(rev_m.well)
            src_well = _to_384_well_rev(rev_base_well)
            _, dest_well = _parse_well_plate(fwd_by_mut.get(fwd_m.mutation, fwd_m.well))

            writer.writerow([
                src_plate, rev_m.primer_name, src_well,
                dest_plate, fwd_m.mutation, dest_well, transfer_vol,
            ])


def export_janus_mapping_csv(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: float = 2.0,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Export JANUS liquid handler mapping CSV.

    Uses two source racks:
      - Rack 1: forward primers (96-well deep well plate).
      - Rack 2: reverse primers (96-well deep well plate).

    Shared reverse primers produce one row per destination well,
    all aspirating from the same source position.

    Args:
        fwd_mappings: Forward primer plate mappings.
        rev_mappings: Deduplicated reverse primer plate mappings.
        output_path: Output CSV file path.
        transfer_vol: Dispense volume in µL (default 2.0).
        rev_groups: Reverse deduplication map {seq: [mutation_names]}.
    """
    import csv

    fwd_by_mut, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        # Header matches JANUS format exactly (Dsp. Rack appears twice)
        writer.writerow([
            "name", "type", "Dsp. Rack", "no",
            "Asp. Rack", "Asp. Posi", "Dsp. Rack", "Dsp. Posi", "volume",
        ])

        seq_no = 1

        # Forward primers (Asp. Rack = 1)
        for m in fwd_mappings:
            writer.writerow([
                f"{m.mutation}-fw", "primer", "Oligo 5pmol/ul", seq_no,
                1, m.well, 2, m.well, transfer_vol,
            ])
            seq_no += 1

        # Reverse primers (Asp. Rack = 2), expanding shared primers
        for fwd_m in fwd_mappings:
            rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
            if rev_seq is None:
                continue
            rev_m = rev_by_seq.get(rev_seq)
            if rev_m is None:
                continue

            dest_well = fwd_by_mut.get(fwd_m.mutation, fwd_m.well)
            writer.writerow([
                f"{fwd_m.mutation}-rv", "primer", "Oligo 5pmol/ul", seq_no,
                2, rev_m.well, 2, dest_well, transfer_vol,
            ])
            seq_no += 1


# ---------------------------------------------------------------------------
# XLSX liquid handler exports (transfer sheet + plate layout)
# ---------------------------------------------------------------------------

def _build_rev_lookups(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    rev_groups: dict[str, list[str]] | None,
) -> tuple[dict[str, str], dict[str, PlateMapping], dict[str, str]]:
    """Shared lookup tables for Echo/JANUS export."""
    fwd_by_mut = {m.mutation: m.well for m in fwd_mappings}
    rev_by_seq = {m.sequence: m for m in rev_mappings}
    mut_to_rev_seq: dict[str, str] = {}
    if rev_groups:
        for seq, muts in rev_groups.items():
            for mut in muts:
                mut_to_rev_seq[mut] = seq
    else:
        for m in rev_mappings:
            mut_to_rev_seq[m.mutation] = m.sequence
    return fwd_by_mut, rev_by_seq, mut_to_rev_seq


def _write_96well_grid(
    ws, start_row: int, mappings: list[PlateMapping], label: str,
    labware: str = "96 deep well",
    value_attr: str = "primer_name",
) -> int:
    """Write a labelled 96-well plate grid and return the next free row.

    Args:
        value_attr: PlateMapping attribute to display in each cell
            ('primer_name' for source plates, 'mutation' for destination).
    """
    from openpyxl.styles import Alignment, Font

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    r = start_row

    ws.cell(row=r, column=1, value="lab ware").font = bold
    ws.cell(row=r, column=3, value=label)
    r += 1

    ws.cell(row=r, column=1, value=labware).font = bold
    for c in range(1, 13):
        cell = ws.cell(row=r, column=c + 2, value=c)
        cell.font = bold
        cell.alignment = center
    r += 1

    well_lookup: dict[str, str] = {}
    for m in mappings:
        _, base = _parse_well_plate(m.well)
        well_lookup[base] = getattr(m, value_attr)

    for row_letter in "ABCDEFGH":
        ws.cell(row=r, column=2, value=row_letter).font = bold
        for c in range(1, 13):
            name = well_lookup.get(f"{row_letter}{c}")
            if name:
                ws.cell(row=r, column=c + 2, value=name).alignment = center
        r += 1

    return r


def export_echo_mapping_xlsx(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: int = 100,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Export Echo 525 mapping as XLSX matching the lab reference format.

    Sheets:
      - layout: 384-well source plate (Fwd odd rows + Rev even rows)
                + 96-well destination plate.
      - Echo mapping file: one row per transfer event.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    fwd_by_mut, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    wb = Workbook()

    # ---- Sheet 1: layout ----
    ws = wb.active
    ws.title = "layout"

    # Title
    ws.cell(row=1, column=2, value="Primer dispensing (Echo 525)")

    # Source section header
    ws.cell(row=2, column=1, value="Source").font = bold
    ws.cell(row=3, column=1, value="labware")
    ws.cell(row=3, column=3, value="source plate")

    # "Eco 384PP" + column numbers 1-24
    ws.cell(row=4, column=1, value="Eco 384PP").font = bold
    for c in range(1, 25):
        cell = ws.cell(row=4, column=c + 2, value=c)
        cell.font = bold
        cell.alignment = center

    # Build 384-well lookup: well_384 → primer_name
    well_384: dict[str, str] = {}
    for m in fwd_mappings:
        _, base = _parse_well_plate(m.well)
        well_384[_to_384_well_fwd(base)] = m.primer_name
    for m in rev_mappings:
        _, base = _parse_well_plate(m.well)
        well_384[_to_384_well_rev(base)] = m.primer_name

    # 384-well grid: rows A-P (16 rows)
    for ri, row_letter in enumerate(_ROWS_384):
        r = 5 + ri
        ws.cell(row=r, column=2, value=row_letter).font = bold
        for c in range(1, 25):
            name = well_384.get(f"{row_letter}{c}")
            if name:
                ws.cell(row=r, column=c + 2, value=name).alignment = center

    # Destination section (96-well PCR plate)
    dest_start = 5 + len(_ROWS_384) + 1  # after 384-well grid + blank row
    ws.cell(row=dest_start, column=1, value="Destination").font = bold
    _write_96well_grid(
        ws, dest_start + 1, fwd_mappings, "PCR mixture",
        labware="96 PCR plate", value_attr="mutation",
    )

    # Auto-width
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 4
    for c in range(3, 27):
        ws.column_dimensions[chr(64 + c)].width = 12

    # ---- Sheet 2: Echo mapping file ----
    ws2 = wb.create_sheet("Echo mapping file")
    headers = [
        "Source Plate Name", "Source Well Name", "Source Well",
        "Dest Plate Name", "Dest Well Name", "Dest Well", "Transfer Vol",
    ]
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill

    row_num = 2

    for m in fwd_mappings:
        plate_idx, base_well = _parse_well_plate(m.well)
        src_plate = f"Source [{plate_idx + 1}]"
        dest_plate = f"Destination [{plate_idx + 1}]"
        src_well = _to_384_well_fwd(base_well)
        for ci, val in enumerate([
            src_plate, m.primer_name, src_well,
            dest_plate, m.mutation, base_well, transfer_vol,
        ], 1):
            ws2.cell(row=row_num, column=ci, value=val)
        row_num += 1

    for fwd_m in fwd_mappings:
        rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
        if rev_seq is None:
            continue
        rev_m = rev_by_seq.get(rev_seq)
        if rev_m is None:
            continue
        fwd_plate_idx, _ = _parse_well_plate(fwd_m.well)
        src_plate = f"Source [{fwd_plate_idx + 1}]"
        dest_plate = f"Destination [{fwd_plate_idx + 1}]"
        _, rev_base_well = _parse_well_plate(rev_m.well)
        src_well = _to_384_well_rev(rev_base_well)
        _, dest_well = _parse_well_plate(fwd_by_mut.get(fwd_m.mutation, fwd_m.well))
        for ci, val in enumerate([
            src_plate, rev_m.primer_name, src_well,
            dest_plate, fwd_m.mutation, dest_well, transfer_vol,
        ], 1):
            ws2.cell(row=row_num, column=ci, value=val)
        row_num += 1

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_path)


def export_janus_mapping_xlsx(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: float = 2.0,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Export JANUS mapping as XLSX matching the lab reference format.

    Sheets:
      - layout: Fwd 96-well plate + Rev 96-well plate
                + 96-well PCR mixture (destination) on a single sheet.
      - primer_mapping file: one row per transfer event.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    fwd_by_mut, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )
    bold = Font(bold=True)

    wb = Workbook()

    # ---- Sheet 1: layout ----
    ws = wb.active
    ws.title = "layout"

    # Title
    ws.cell(row=1, column=2, value="Primer dispensing (JANUS)")

    # Fwd plate
    r = _write_96well_grid(ws, 2, fwd_mappings, "fw plate")

    # Blank row + Rev plate
    r += 1
    r = _write_96well_grid(ws, r, rev_mappings, "rv plate")

    # Blank rows + Destination plate
    r += 2
    _write_96well_grid(
        ws, r, fwd_mappings, "PCR mixture plate",
        labware="96 PCR plate", value_attr="mutation",
    )

    # Auto-width
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 4
    for c in range(3, 15):
        ws.column_dimensions[chr(64 + c)].width = 12

    # ---- Sheet 2: primer_mapping file ----
    ws2 = wb.create_sheet("primer_mapping file")
    headers = [
        "name", "type", "Dsp. Rack", "no",
        "Asp. Rack", "Asp. Posi", "Dsp. Rack", "Dsp. Posi", "volume",
    ]
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill

    row_num = 2
    seq_no = 1

    for m in fwd_mappings:
        for ci, val in enumerate([
            f"{m.mutation}-fw", "primer", "Oligo 5pmol/ul", seq_no,
            1, m.well, 2, m.well, transfer_vol,
        ], 1):
            ws2.cell(row=row_num, column=ci, value=val)
        row_num += 1
        seq_no += 1

    for fwd_m in fwd_mappings:
        rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
        if rev_seq is None:
            continue
        rev_m = rev_by_seq.get(rev_seq)
        if rev_m is None:
            continue
        dest_well = fwd_by_mut.get(fwd_m.mutation, fwd_m.well)
        for ci, val in enumerate([
            f"{fwd_m.mutation}-rv", "primer", "Oligo 5pmol/ul", seq_no,
            2, rev_m.well, 2, dest_well, transfer_vol,
        ], 1):
            ws2.cell(row=row_num, column=ci, value=val)
        row_num += 1
        seq_no += 1

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_path)
