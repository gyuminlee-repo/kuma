"""Combinatorial barcode demux pipeline for 96-well nanopore amplicon screening.

Algorithm (minimap2 align + alignment-anchored fuzzy barcode matching):
----------------------------------------------------------------------
1. Align all raw FASTQ reads to reference using mappy (map-ont preset).
2. MAPQ >= 25 filter.
3. Coverage filter: each alignment must cover >= coverage_fraction of reference
   (default 0.98; replaces strict 100% filter to recover reads with 1-2 bp clip).
4. Chimera/concatemer splitting: ALL passing hits per read are evaluated.
   A single raw read may yield assignments to multiple wells if it contains
   two distinct amplicon copies (chimeric or concatemer read).
   Duplicate (read_id, well) pairs within the same read are deduplicated.
5. For each passing hit, extract alignment-anchored slice +/-trim_flank_bp,
   then run alignment-anchored fuzzy barcode matching on that slice.
   Strand normalisation is done inside _demux_read_anchored.
6. Barcode demux using edlib HW (infix) edit-distance search:
   Library structure (sense strand of read):
     5-[F_barcode + F_anneal]-[insert]-[RC(R_anneal) + RC(R_barcode)]-3
   - F-barcode window (5 end): [max(0, q_st - window_bp - max_f_len), q_st + window_bp]
   - R-barcode window (3 end): [max(0, q_en - window_bp), min(L, q_en + window_bp + max_r_len)]
     R barcode prefixes are reverse-complemented before searching (RC form in read).
   - For each barcode, best infix edit distance is computed; only accept if
     edit_distance <= int(len(bc) * edit_dist_ratio)  (floor, conservative).
   - Ambiguity: if best == second-best edit distance -> drop (ambiguous).
   - Exactly 1 R + 1 F unambiguous match required; otherwise dropped.
7. Per-well consensus: majority-vote per position (N if depth < min_depth).

Barcode loading:
- Annealing tails stripped; only prefix portion used for fuzzy matching.
- F tail: 'cacaggaggttaaacc' (16 bp), R tail: 'tgcgttgcgctctag' (15 bp).
- Fallback prefix length if tail absent: 11 bp (F) / 10 bp (R).

Assumptions:
- Reference FASTA has exactly one sequence record.
- Barcodes xlsx rows: isps_f_1..12 then isps_r_1..8.
- mappy and edlib available (pyproject.toml restricts mappy to Linux).
- Edit-distance threshold uses floor(len * ratio), not ceil, to stay
  conservative on 10 bp barcodes (floor gives max 2 edits at ratio=0.20).
"""

from __future__ import annotations

import contextlib
import gzip
import logging
import queue
import threading
import multiprocessing
import tempfile
import time
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
import os
import re
import sys
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from typing import Any, Callable, Iterator, TypeVar

from kuma_core.mame.ingest.align import (
    align_reads_grouped,
    align_reads_multi_with_gate_counts,
    align_reads_with_gate_counts,
    build_minimap2_index,
    _get_reference_length,
    Alignment,
)
from kuma_core.mame.ingest.consensus import call_consensus_with_metrics
from kuma_core.mame.ingest.consensus_metadata import (
    BASIS_COVERED,
    ConsensusMetadata,
    format_consensus_fasta_record,
)
from kuma_core.mame.ingest.stage_marker import (
    is_unit_complete,
    marker_inputs_match,
    read_stage_marker,
    reference_fingerprint,
    write_stage_marker,
)
from kuma_core.mame.ingest.well_consensus import _read_reference_seq
from kuma_core.mame.perf import TIMER, timed_iter
from kuma_core.shared.atomic_write import atomic_write_text, fsync_directory

T = TypeVar("T")

log = logging.getLogger(__name__)

# NOTE (2026-07-31): a former `_is_frozen_win()` guard disabled both the per-NB
# and the per-read ProcessPool on frozen Windows builds, blaming "PyInstaller
# onefile re-extracts the archive per spawned worker and deadlocks". That
# diagnosis was wrong and is deliberately not restored:
#   * The observed symptom (Windows frozen smoke appearing to hang) was root-
#     caused later the same day to stdin read-ahead buffering in the sidecar
#     dispatcher, and fixed there, see the readline() loop in
#     python-core/sidecar_mame/dispatcher.py. The demux itself ran in ~10 s once
#     the request actually reached it; it was never a compute cold-start.
#   * The guard also claimed "Linux/macOS is fine (fork)", which is false: this
#     module never used the plain "fork" start method, and the Linux frozen
#     onefile build exercises the same spawn ProcessPool path in CI without
#     deadlocking.
# Re-entrancy of spawned children is handled by multiprocessing.freeze_support()
# in python-core/sidecar_main_mame.py, not by disabling parallelism. Frozen
# builds still take exactly that spawn path; see _mp_start_method() below, which
# only upgrades non-frozen POSIX to forkserver. Escape hatches remain
# env-driven: KUMA_MAME_NB_PARALLEL=0 disables the per-NB pool, a large
# KUMA_MAME_PERREAD_THRESHOLD disables the per-read pool, and
# KUMA_MAME_MP_START=spawn pins the old start method.

_F_TAIL = "cacaggaggttaaacc"
_R_TAIL = "tgcgttgcgctctag"

_F_FALLBACK_LEN = 11  # prefix length if F tail absent
_R_FALLBACK_LEN = 10  # prefix length if R tail absent

# Gene-agnostic barcode row-name patterns (mirror sort_barcode.py).
# Match any "<prefix>_f_<int>" / "<prefix>_r_<int>" — not limited to "isps".
_FWD_ROW_RE = re.compile(r"^(?P<prefix>.+?)_f_(?P<n>\d+)$")
_REV_ROW_RE = re.compile(r"^(?P<prefix>.+?)_r_(?P<n>\d+)$")

_COMP = str.maketrans("ACGTacgtNn", "TGCAtgcaNn")

# Default worker count for per-well consensus ThreadPool.
# KUMA_MAME_CONSENSUS_WORKERS env var takes priority.
_CONSENSUS_WORKERS: int = int(
    os.environ.get("KUMA_MAME_CONSENSUS_WORKERS", "")
    or str(max(1, (os.cpu_count() or 4) - 1))
)

# Filename for the combined single-file consensus FASTA (all wells in one
# multi-record file), written in output_dir alongside the per-well consensus/
# directory. Mirrors the Aporva pipeline's final/<...>_consensus_dna.fasta so
# downstream tools that expect one multi-record FASTA keep working.
_COMBINED_CONSENSUS_FILENAME = "consensus_all_dna.fasta"

# Modules the forkserver helper imports once, so every forked demux worker
# starts with them already resident. Only this module is listed: importing it
# pulls the whole worker-side chain (align, consensus -> numpy, edlib, ...).
_MP_PRELOAD = ("kuma_core.mame.ingest.combinatorial_demux",)


def _mp_start_method() -> str:
    """Start method for the demux ProcessPools.

    "forkserver" launches one helper process, imports ``_MP_PRELOAD`` in it
    once, then forks already-warm children. "spawn" instead pays a full
    interpreter start plus the whole import chain in every worker, every run.

    Restricted to POSIX, non-frozen builds:

    * forkserver is POSIX only, so Windows always reports "spawn".
    * ``multiprocessing.forkserver.ensure_running`` launches the helper as
      ``[sys.executable] + interpreter_flags + ["-c", <code>]`` (CPython 3.11
      forkserver.py:148-151). Unlike ``spawn.get_command_line``, it carries no
      ``sys.frozen`` branch, so inside a PyInstaller bundle it would re-exec
      the frozen sidecar binary, which ignores ``-c`` and would restart the
      JSON-RPC loop instead of becoming a forkserver. Frozen therefore stays on
      spawn, which multiprocessing.freeze_support() in
      python-core/sidecar_main_mame.py already handles.

    Forked children are safe here despite the parent holding threads (the
    progress drain thread below, plus the sidecar heartbeat/stdin threads):
    the forkserver helper is created by fork+exec (util.spawnv_passfds ->
    _posixsubprocess.fork_exec), so it never inherits parent threads or their
    lock state, and workers are forked from that single-threaded helper rather
    than from the parent.

    ``KUMA_MAME_MP_START`` forces a method by name (escape hatch).
    """
    available = multiprocessing.get_all_start_methods()
    forced = os.environ.get("KUMA_MAME_MP_START", "").strip().lower()
    if forced:
        if forced in available:
            return forced
        log.warning(
            "KUMA_MAME_MP_START=%r is not available (have %s); using spawn",
            forced, available,
        )
        return "spawn"
    if getattr(sys, "frozen", False):
        return "spawn"
    if "forkserver" in available:
        return "forkserver"
    return "spawn"


def _demux_mp_context():
    """Multiprocessing context for the demux pools, preloaded when forkserver."""
    method = _mp_start_method()
    ctx = multiprocessing.get_context(method)
    if method == "forkserver":
        # A preload entry that fails to import is skipped by the forkserver
        # (forkserver.main swallows ImportError), so this cannot make workers
        # unusable, they would just import lazily as under spawn.
        ctx.set_forkserver_preload(list(_MP_PRELOAD))
    return ctx


def _warm_mp_context(ctx) -> threading.Thread | None:
    """Start the forkserver helper off the critical path. No-op for spawn.

    The helper is otherwise created lazily by the first ``Process.start()``,
    which puts its preload import (~0.5 s) in front of the first worker. Kicking
    it here lets it run concurrently with the parent-side barcode parse below,
    so even the first pool in a process starts warm.

    ``forkserver.ensure_running`` is idempotent and holds its own lock, so the
    pool calling it again later is harmless.
    """
    if ctx.get_start_method() != "forkserver":
        return None
    from multiprocessing import forkserver as _forkserver

    def _warm() -> None:
        try:
            _forkserver.ensure_running()
        except Exception:  # warm-up is an optimization, never fatal
            log.debug("forkserver warm-up failed", exc_info=True)

    thread = threading.Thread(
        target=_warm, name="mame-forkserver-warmup", daemon=True
    )
    thread.start()
    return thread


def _reverse_complement(seq: str) -> str:
    return seq.translate(_COMP)[::-1]


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class DemuxStats:
    """Summary counters from a single run_combinatorial_demux call."""

    total_reads: int = 0
    passed_mapq: int = 0
    passed_coverage: int = 0
    assigned_reads: int = 0
    ambiguous_dropped: int = 0
    chimera_splits: int = 0   # extra well assignments from multi-hit reads
    wells_with_reads: int = 0
    wells_with_min_reads: int = 0


@dataclass
class DemuxResult:
    """Return value of run_combinatorial_demux.

    ``per_well_reads`` carries the assigned read slices per well, but ONLY when
    the run fitted inside the in-memory well buffer (see ``_WellReadBuffer``).
    A run large enough to spill leaves it empty rather than pulling gigabytes
    of reads back off disk to hand to a caller that, in every production path,
    only counts them.  ``per_well_read_counts`` is always populated and is what
    those callers should read.
    """

    stats: DemuxStats
    per_well_reads: dict[str, list[tuple[str, str]]] = field(default_factory=dict)
    per_well_consensus: dict[str, str] = field(default_factory=dict)
    per_well_read_counts: dict[str, int] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Barcode utilities
# ---------------------------------------------------------------------------


def _extract_barcode_prefix(seq: str, tail: str) -> str:
    """Return the prefix before the annealing tail.

    Fallback: first 11 bp if tail absent.  Public for external callers and
    backward compatibility.
    """
    idx = seq.lower().find(tail.lower())
    if idx >= 0:
        return seq[:idx]
    return seq[:11]


def load_barcode_prefixes(
    barcodes_xlsx: Path,
) -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    """Load F and R barcode *prefix* sequences from xlsx (annealing tail stripped).

    Returns
    -------
    (r_barcodes, f_barcodes)
        r_barcodes: 8-element list of (name, prefix_seq) tuples (index 0 = R1).
        f_barcodes: 12-element list of (name, prefix_seq) tuples (index 0 = F1).

    The prefix is the barcode-unique region only (tail excluded).  Prefix lengths
    are typically 10 bp (R and most F) or 11 bp (F1-F3 in the standard plate).
    """
    try:
        import openpyxl  # type: ignore[import]
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for barcode loading. "
            "Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(barcodes_xlsx, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Empty workbook: no active sheet in " + str(barcodes_xlsx))

    f_entries: list[tuple[int, str, str]] = []  # (idx, name, prefix)
    r_entries: list[tuple[int, str, str]] = []

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip().lower()
        seq_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not seq_val:
            continue

        m_f = _FWD_ROW_RE.match(name)
        m_r = _REV_ROW_RE.match(name)
        if m_f is not None:
            idx = int(m_f.group("n"))
            prefix = _extract_f_prefix(seq_val)
            f_entries.append((idx, name, prefix.upper()))

        elif m_r is not None:
            idx = int(m_r.group("n"))
            prefix = _extract_r_prefix(seq_val)
            r_entries.append((idx, name, prefix.upper()))

    wb.close()

    f_entries.sort(key=lambda x: x[0])
    r_entries.sort(key=lambda x: x[0])

    f_barcodes = [(name, prefix) for _, name, prefix in f_entries]
    r_barcodes = [(name, prefix) for _, name, prefix in r_entries]

    if len(f_barcodes) != 12:
        log.warning("Expected 12 F barcodes, got %d", len(f_barcodes))
    if len(r_barcodes) != 8:
        log.warning("Expected 8 R barcodes, got %d", len(r_barcodes))

    return r_barcodes, f_barcodes


def _extract_f_prefix(seq: str) -> str:
    """Strip F annealing tail; fallback to first _F_FALLBACK_LEN bases."""
    idx = seq.lower().find(_F_TAIL.lower())
    return seq[:idx] if idx >= 0 else seq[:_F_FALLBACK_LEN]


def _extract_r_prefix(seq: str) -> str:
    """Strip R annealing tail; fallback to first _R_FALLBACK_LEN bases."""
    idx = seq.lower().find(_R_TAIL.lower())
    return seq[:idx] if idx >= 0 else seq[:_R_FALLBACK_LEN]


def load_barcodes(barcodes_xlsx: Path) -> tuple[list[str], list[str]]:
    """Load F and R barcode full sequences from xlsx (legacy, full seq).

    Returns
    -------
    (f_barcodes, r_barcodes)
        f_barcodes: 12-element list, uppercase full sequences (index 0 = F1).
        r_barcodes: 8-element list, uppercase full sequences (index 0 = R1).

    .. note::
        This function is kept for backward compatibility.  The main pipeline
        now uses :func:`load_barcode_prefixes` which strips annealing tails.
    """
    try:
        import openpyxl  # type: ignore[import]
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for barcode loading. "
            "Install with: pip install openpyxl"
        ) from exc

    wb = openpyxl.load_workbook(barcodes_xlsx, read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Empty workbook: no active sheet in " + str(barcodes_xlsx))

    f_entries: list[tuple[int, str]] = []
    r_entries: list[tuple[int, str]] = []

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip().lower()
        seq_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        if not seq_val:
            continue

        m_f = _FWD_ROW_RE.match(name)
        m_r = _REV_ROW_RE.match(name)
        if m_f is not None:
            idx = int(m_f.group("n"))
            f_entries.append((idx, seq_val.upper()))

        elif m_r is not None:
            idx = int(m_r.group("n"))
            r_entries.append((idx, seq_val.upper()))

    wb.close()

    f_entries.sort(key=lambda x: x[0])
    r_entries.sort(key=lambda x: x[0])

    f_barcodes = [s for _, s in f_entries]
    r_barcodes = [s for _, s in r_entries]

    if len(f_barcodes) != 12:
        log.warning("Expected 12 F barcodes, got %d", len(f_barcodes))
    if len(r_barcodes) != 8:
        log.warning("Expected 8 R barcodes, got %d", len(r_barcodes))

    return f_barcodes, r_barcodes


# ---------------------------------------------------------------------------
# FASTQ parsing
# ---------------------------------------------------------------------------


def _iter_fastq(paths: list[Path]) -> Iterator[tuple[str, str]]:
    """Yield (read_id, sequence) from one or more FASTQ(.gz) files."""
    for path in paths:
        opener = gzip.open if str(path).endswith(".gz") else open
        with opener(path, "rt") as fh:
            while True:
                header = fh.readline()
                if not header:
                    break
                seq = fh.readline().rstrip("\n")
                fh.readline()   # '+'
                fh.readline()   # quality
                if seq:
                    read_id = header[1:].split()[0].rstrip("\n")
                    yield read_id, seq


#: Queue depth (in chunks) of the FASTQ prefetch thread. 0 disables prefetch
#: and restores the fully serial read-then-align path.
_FASTQ_PREFETCH_DEFAULT = 1


def _prefetch(it: Iterator[T], depth: int) -> Iterator[T]:
    """Yield from *it* while a background thread runs it ``depth`` items ahead.

    Ordering: a single producer thread drains *it* sequentially into a FIFO
    queue, so consumed order is the producing order verbatim. This matters for
    the demux, whose consensus tie-break depends on within-well read order
    (``kuma_core/mame/consensus.py`` ``first_touch``).

    Why a thread: measured on the reference workload, 89% of the FASTQ read
    cost is zlib decompression, which releases the GIL, and the consumer spends
    most of its time waiting on the minimap2 subprocess. Both leave the GIL
    free for the producer.

    Memory: the queue holds at most *depth* ready items plus the one in flight.

    Errors: an exception raised by *it* is captured and re-raised in the
    consumer at the point it would have surfaced serially, so a truncated or
    corrupt gzip still aborts the run instead of looking like a short input.
    """
    q: "queue.Queue[tuple[bool, object]]" = queue.Queue(maxsize=max(1, depth))
    stop = threading.Event()
    _DONE = object()

    def _produce() -> None:
        try:
            for item in it:
                if stop.is_set():
                    return
                q.put((True, item))
        except BaseException as exc:  # noqa: BLE001 - re-raised in the consumer
            q.put((False, exc))
            return
        q.put((True, _DONE))

    thread = threading.Thread(target=_produce, name="fastq-prefetch", daemon=True)
    thread.start()
    try:
        while True:
            ok, payload = q.get()
            if not ok:
                raise payload  # type: ignore[misc]
            if payload is _DONE:
                return
            yield payload  # type: ignore[misc]
    finally:
        # Unblock the producer if the consumer abandoned the iterator early
        # (exception upstream, generator close) so the thread cannot leak.
        stop.set()
        with contextlib.suppress(queue.Empty):
            while True:
                q.get_nowait()


# ---------------------------------------------------------------------------
# Alignment-anchored fuzzy barcode matching
# ---------------------------------------------------------------------------


def _best_infix_match(
    query: str,
    window: str,
    max_edit: int,
) -> int | None:
    """Return the best infix edit distance of *query* inside *window*.

    Uses edlib HW mode (infix / semi-global on query).  Returns the edit
    distance if <= max_edit, otherwise None.
    """
    try:
        import edlib  # type: ignore[import]
    except ImportError as exc:
        raise ImportError(
            "edlib is required for fuzzy barcode matching. "
            "Install with: pip install edlib"
        ) from exc

    if not window or not query:
        return None

    result = edlib.align(query, window, mode="HW", task="distance", k=max_edit)
    dist = result["editDistance"]
    if dist < 0 or dist > max_edit:
        return None
    return dist


def _find_best_barcode(
    barcodes: list[tuple[str, str]],
    window: str,
    edit_dist_ratio: float,
    max_edits: list[int] | None = None,
) -> tuple[int, int] | None:
    """Find the unambiguous best-matching barcode in *window*.

    Parameters
    ----------
    barcodes:
        List of (name, prefix_seq) tuples (1-indexed position = list index + 1).
    window:
        Sequence window extracted from the read.
    edit_dist_ratio:
        Max allowed edit distance = int(len(bc) * edit_dist_ratio).
    max_edits:
        Optional precomputed per-barcode thresholds, positionally aligned with
        *barcodes*.  When omitted they are derived from *edit_dist_ratio* exactly
        as before; passing them in only hoists a read-invariant computation out
        of the per-read loop (identical values, no behaviour change).

    Returns
    -------
    (1-based index, edit_distance) if exactly one barcode is unambiguously
    best, None otherwise (no match or ambiguous).
    """
    best_dist: int = 10**6
    best_idx: int = -1
    second_best_dist: int = 10**6

    for i, (_, prefix) in enumerate(barcodes):
        if max_edits is None:
            max_edit = int(len(prefix) * edit_dist_ratio)
        else:
            max_edit = max_edits[i]
        dist = _best_infix_match(prefix, window, max_edit)
        if dist is None:
            continue
        if dist < best_dist:
            second_best_dist = best_dist
            best_dist = dist
            best_idx = i
        elif dist < second_best_dist:
            second_best_dist = dist

    if best_idx < 0:
        return None  # no match within threshold

    # Ambiguity guard: if best == second_best, result is ambiguous -> drop
    if best_dist == second_best_dist:
        return None

    return best_idx + 1, best_dist  # 1-based index


@dataclass(frozen=True)
class _BarcodePlan:
    """Read-invariant barcode preprocessing, computed once per run.

    Everything here depends only on (r_barcodes, f_barcodes, edit_dist_ratio),
    never on the read, so hoisting it out of the per-read loop cannot change
    results.  Module-level dataclass => picklable for the ``spawn`` ProcessPool.
    """

    r_barcodes_rc: list[tuple[str, str]]
    f_barcodes: list[tuple[str, str]]
    max_r_len: int
    max_f_len: int
    r_max_edits: list[int]
    f_max_edits: list[int]
    edit_dist_ratio: float


def _build_barcode_plan(
    r_barcodes: list[tuple[str, str]],
    f_barcodes: list[tuple[str, str]],
    edit_dist_ratio: float,
) -> _BarcodePlan:
    """Precompute the read-invariant parts of :func:`_demux_read_anchored`.

    Mirrors the former per-read expressions verbatim:
    ``max(len(p))`` defaults, ``_reverse_complement`` of each R prefix, and
    ``int(len(prefix) * edit_dist_ratio)``.  RC preserves length, so the R
    thresholds computed from the RC'd prefixes equal the original ones.
    """
    r_rc = [(name, _reverse_complement(prefix)) for name, prefix in r_barcodes]
    return _BarcodePlan(
        r_barcodes_rc=r_rc,
        f_barcodes=f_barcodes,
        max_r_len=max((len(p) for _, p in r_barcodes), default=10),
        max_f_len=max((len(p) for _, p in f_barcodes), default=11),
        r_max_edits=[int(len(p) * edit_dist_ratio) for _, p in r_rc],
        f_max_edits=[int(len(p) * edit_dist_ratio) for _, p in f_barcodes],
        edit_dist_ratio=edit_dist_ratio,
    )


def _extract_barcode_windows(
    read_seq: str,
    q_st: int,
    q_en: int,
    strand: int,
    window_bp: int,
    max_f_len: int,
    max_r_len: int,
) -> tuple[str, str]:
    """Return the (f_window, r_window) barcode search slices for one read.

    Window extraction works on two short slices instead of upper()/RC'ing the
    whole (up to multi-kb) read. This is exact, not an approximation:
    ``upper()`` and ``_reverse_complement()`` (str.translate + reversal) are
    both per-character on the ASCII bases a FASTQ read carries, so
        upper(S)[a:b]                == upper(S[a:b])
        RC(upper(S))[a:b]            == RC(upper(S[L-b:L-a]))
    The -1 branch below is the second identity with the window bounds folded
    in; see the comment there for the coordinate derivation.

    Window rationale (unchanged):
    F barcode is strictly 5' of alignment start (F_barcode + F_anneal tail).
    The inner edge stops at the anchor and must NOT extend into the aligned
    insert: the reference 5'-start can sit within the edit threshold of a
    forward barcode (e.g. ispS starts "TGGCTTGCTC", edit distance 2 from the
    F9 prefix "TGCCTTGATC"). If the insert is inside the search window, every
    read whose real F barcode is degraded matches that barcode against the gene
    start and funnels into a single well, contaminating its whole column. The
    barcode + anneal lie wholly 5' of the insert, so excluding the insert loses
    no real barcode signal. The R barcode window is the mirror image: strictly
    3' of the alignment end, where the barcode appears as RC(R_barcode) (hence
    the R prefixes are searched in RC form, precomputed in the plan).
    """
    L = len(read_seq)

    if strand == -1:
        # Original code normalised via norm_q_st = L - q_en, norm_q_en = L - q_st
        # on rc = RC(upper(read_seq)), then took
        #   f_window = rc[max(0, norm_q_st - window_bp - max_f_len) : min(L, norm_q_st)]
        #   r_window = rc[max(0, norm_q_en) : min(L, norm_q_en + window_bp + max_r_len)]
        # Mapping rc[a:b] back to the read via RC(upper(read_seq[L-b:L-a])) and
        # substituting L - norm_q_st = q_en, L - norm_q_en = q_st gives:
        #   f source = read_seq[max(0, q_en) : min(L, q_en + window_bp + max_f_len)]
        #   r source = read_seq[max(0, q_st - window_bp - max_r_len) : min(L, q_st)]
        # (min/max survive the mapping because L - min(L, x) == max(0, L - x).)
        f_window = _reverse_complement(
            read_seq[max(0, q_en):min(L, q_en + window_bp + max_f_len)].upper()
        )
        r_window = _reverse_complement(
            read_seq[max(0, q_st - window_bp - max_r_len):min(L, q_st)].upper()
        )
    else:
        f_window = read_seq[
            max(0, q_st - window_bp - max_f_len):min(L, q_st)
        ].upper()
        r_window = read_seq[
            max(0, q_en):min(L, q_en + window_bp + max_r_len)
        ].upper()

    return f_window, r_window


def _demux_read_anchored(
    read_seq: str,
    q_st: int,
    q_en: int,
    strand: int,
    r_barcodes: list[tuple[str, str]],
    f_barcodes: list[tuple[str, str]],
    window_bp: int = 30,
    edit_dist_ratio: float = 0.20,
    plan: _BarcodePlan | None = None,
) -> tuple[int, int] | None:
    """Demux one read using alignment anchors and edlib fuzzy matching.

    Parameters
    ----------
    read_seq:
        Full read sequence as returned by the FASTQ parser (not RC'd).
    q_st, q_en:
        Alignment start/end on the read (from mappy, 0-based half-open).
    strand:
        +1 or -1 from mappy.
    r_barcodes, f_barcodes:
        (name, prefix) tuples as returned by load_barcode_prefixes.
    window_bp:
        Window radius around anchor points (default 30 bp).
    edit_dist_ratio:
        Max allowed edit distance fraction of barcode length (default 0.20).
        Threshold = floor(len(bc) * ratio).  At ratio=0.20: 10 bp -> 2 edits,
        11 bp -> 2 edits, 15 bp -> 3 edits.
    plan:
        Optional :class:`_BarcodePlan` from :func:`_build_barcode_plan`, letting
        callers hoist the read-invariant barcode preprocessing out of their
        per-read loop.  Omitted (or built for a different *edit_dist_ratio*) it
        is rebuilt here, so existing call sites keep working unchanged.

    Returns
    -------
    (r_idx_1based, f_idx_1based) or None if demux fails.

    Notes
    -----
    Strand normalisation:
      strand +1: use read_seq as-is.
      strand -1: work on RC(read_seq) and remap anchor coords via
                 norm_q_st = L - q_en, norm_q_en = L - q_st.
    Biological window layout (after normalisation to +1 orientation):
      F barcode is 5' of the amplicon -> search before norm_q_st.
      R barcode is 3' of the amplicon in RC form -> search RC(r_prefix) after norm_q_en.

    Library structure on sense strand of read (strand +1):
      5'-[F_barcode + F_anneal]-[insert]-[RC(R_anneal) + RC(R_barcode)]-3'
    So the 5' window contains F_barcode (as-is) and the 3' window contains
    RC(R_barcode). R barcode prefixes are reverse-complemented before searching.
    """
    if plan is None or plan.edit_dist_ratio != edit_dist_ratio:
        plan = _build_barcode_plan(r_barcodes, f_barcodes, edit_dist_ratio)

    # Window extraction (slice-then-normalise) lives in
    # :func:`_extract_barcode_windows`; see its docstring for the equivalence
    # proof against the original whole-read upper()/RC() formulation and for
    # the biological rationale of the window bounds.
    f_window, r_window = _extract_barcode_windows(
        read_seq,
        q_st,
        q_en,
        strand,
        window_bp,
        plan.max_f_len,
        plan.max_r_len,
    )

    f_result = _find_best_barcode(
        plan.f_barcodes, f_window, edit_dist_ratio, plan.f_max_edits
    )
    r_result = _find_best_barcode(
        plan.r_barcodes_rc, r_window, edit_dist_ratio, plan.r_max_edits
    )

    if r_result is None or f_result is None:
        return None

    r_idx, _ = r_result
    f_idx, _ = f_result
    return r_idx, f_idx


def _demux_read(
    trimmed_seq: str,
    f_barcodes: list[str],
    r_barcodes: list[str],
) -> tuple[int, int] | None:
    """Exact substring demux (legacy, no alignment context).

    Kept for backward compatibility with existing tests.  The main pipeline
    uses :func:`_demux_read_anchored`.

    Returns (r_idx_1based, f_idx_1based) or None.
    """
    seq_upper = trimmed_seq.upper()
    seq_rc = _reverse_complement(seq_upper)

    matched_r = [
        i + 1
        for i, bc in enumerate(r_barcodes)
        if bc in seq_upper or bc in seq_rc
    ]
    matched_f = [
        i + 1
        for i, bc in enumerate(f_barcodes)
        if bc in seq_upper or bc in seq_rc
    ]

    if len(matched_r) == 1 and len(matched_f) == 1:
        return matched_r[0], matched_f[0]
    return None


# ---------------------------------------------------------------------------
# Per-read chimera-path matching (extracted for optional ProcessPool fan-out)
# ---------------------------------------------------------------------------

# Default read-count threshold above which the chimera-path per-read matching
# loop is fanned out to a ProcessPool (only when this run owns the cores, i.e.
# n_nb == 1). Read at call time via os.environ so tests can lower it; a
# module-level constant bound at import could not be overridden by monkeypatch.
_PERREAD_THRESHOLD_DEFAULT = 10000

# Default read-chunk size for the alignment stage. Reads are loaded and aligned
# in chunks of this size instead of materialising the whole FASTQ in memory, so
# the per-chunk minimap2 input/SAM and the per-chunk Alignment lists are dropped
# between chunks (lowers alignment-stage peak RAM only; per_well accumulates to
# consensus as before). Read at call time via os.environ (KUMA_MAME_READ_CHUNK)
# so tests can lower it; a module-level constant bound at import could not be
# overridden by monkeypatch.
#
# CORRECTION (2026-08-01): an earlier version of this comment blamed minimap2,
# claiming it "is NOT split-invariant" and that the chunk size therefore could
# not preserve output. That was wrong. minimap2 aligns each query independently
# and IS split-invariant for a fixed set of query NAMES; the earlier measurement
# was confounded by our own aligner adapter, which renamed reads to their
# position within the call (align._write_reads_fasta wrote ">0", ">1", ... from
# scratch per call). minimap2 seeds its per-read RNG from a hash of the query
# name, so restarting the numbering in every chunk changed the effective seed of
# every read and moved a handful of alignments. Proof: aligning the SAME 3000
# reads in the SAME order, numbered 0.. versus 100000.., produced different SAM
# at -t 1 and -t 7 alike (MAPQ 60 -> 1 on read index 235, primary/supplementary
# swap on 124, a different chain on 1830).
#
# The loop below now passes a running ``name_offset`` so each read keeps the
# QNAME it would have had in a single whole-set call, which makes the output
# invariant to this value. It is a performance knob again. Any future change
# that re-derives query names from a chunk-local counter re-introduces the bug.
#
# Sizing (measured 2026-08-01, 10-core WSL2 box, reference workload of 3 native
# barcodes / 34.3k reads). The value only matters through the number of chunks,
# because that is what the FASTQ prefetch above has to overlap with: at 50000
# the reference workload is ONE chunk per barcode and there is nothing to hide
# the gzip behind. Interleaved 6-round A/B, prefetch on, e2e wall / demux
# medians and the residual fastq_read wait:
#
#   chunk 10000  ( 2 chunks/NB)  7.951 s / 6.206 s   fastq_read 1.765 s
#   chunk  5000  ( 3 chunks/NB)  7.791 s / 6.021 s   fastq_read 0.853 s
#   chunk  2500  ( 6 chunks/NB)  7.699 s / 5.947 s   fastq_read 0.436 s
#
# against 8.568 s / 6.686 s for the old 50000 default with prefetch off. Going
# below 2500 loses: the per-chunk fixed cost (tempdir, reads FASTA write,
# minimap2 spawn, reference index build) is ~0.016 s, which is +1.4% of the
# alignment phase at 2500 but +9% at 1000 (single-barcode min-of-4, -t 3).
#
# Production scale (~1e6 reads per barcode) was checked as an argument, not
# measured. Chunk COUNT is not a resource: the prefetch queue is bounded
# (maxsize=depth), so a slow consumer applies backpressure to the reader and at
# most depth+1 chunks are ever resident. 2500 reads x ~1.5 kb is ~4 MB per
# resident chunk against ~75 MB for the old 50000, i.e. this lowers the
# alignment-stage peak RAM that the chunk loop exists to bound. The per-chunk
# fixed cost stays a constant ~1.4% of alignment because chunk count and total
# alignment work grow together. The one input that would change this answer is a
# large multi-record reference, where the per-chunk index build stops being
# negligible; that case should pass a prebuilt ``reference_index`` rather than
# raise the chunk size.
#
# PORTABILITY (measured 2026-08-02). Asked whether this should adapt to core
# count or filesystem the way the memory bounds now adapt to RAM. Measured, and
# the answer is NO: keep it fixed. Recorded here so the next person does not
# repeat the sweep.
#
# Both sides of the trade-off turn out to be environment-insensitive in SHAPE.
# The term the chunk size buys is the residual `fastq_read` wait, i.e. the gzip
# the prefetch failed to hide, and it moves monotonically and almost identically
# in two environments that differ in both axes at once (medians of 3, chunk ->
# fastq_read seconds):
#
#   9p share, 10 cores   1000: 0.233  2500: 0.861  5000: 1.254  50000: 2.929
#   ext4,      4 cores   1000: 0.219  2500: 0.685  5000: 1.033  50000: 2.682
#
# Same curve to within a few percent across a filesystem change that the ingest
# fan-out DOES have to probe for, and across a 2.5x core change. The opposing
# term, the ~0.016 s per-chunk fixed cost, is tempdir + FASTA write + minimap2
# spawn + index build, none of which scales with either axis either.
#
# The resulting wall is flat across a 50x span of chunk sizes in both
# environments (e2e min of 3: 8.51 / 9.18 / 8.69 / 9.00 / 8.85 s on the share,
# 9.43 / 9.55 / 10.21 / 9.60 / 9.58 s on 4 cores, for 1000 / 2500 / 5000 /
# 10000 / 50000). No optimum resolves above run-to-run noise, so there is
# nothing for an adaptive rule to track: it would add a code path and a failure
# mode to chase differences smaller than the measurement error.
# ``KUMA_MAME_READ_CHUNK`` remains the escape hatch if some future environment
# does show a gradient.
_READ_CHUNK_DEFAULT = 2500


# Memory bound for the assigned read slices held between the read loop and
# consensus (MB of sequence text; the Python object overhead on top is roughly
# another 60%). Measured 2026-08-01: this buffer is 0.71 MB of resident set per
# MB of input FASTQ and, before this bound existed, it was the term that made
# the real 5.9 GB run need ~14 GB of RSS. Above the budget every well is
# appended to its own spill file and the RAM lists are dropped; the reads come
# back one consensus batch at a time.
#
# 512 MB is chosen so the reference 54 MB fixture (4 MB of slices) and any run
# up to ~700 MB of FASTQ per barcode never spill at all, i.e. the default costs
# nothing on every workload measured so far, while the real run (2975 MB in the
# largest barcode) spills and stays bounded. 0 disables the bound.
#
# This is now the FALLBACK only, used when the box will not tell us how much
# memory it has (see _memory_limit_bytes). The live default is derived from the
# limit; _WELL_BUFFER_FRACTION reproduces 512 on the 15 GiB box it was tuned on.
_WELL_BUFFER_MB_DEFAULT = 512

# Memory bound for the consensus stage (MB of sequence text per batch of
# wells). Wells are aligned and consensus-called one batch at a time instead of
# all at once, which bounds BOTH the Alignment objects held (~1.6 kB per read,
# measured) and the per-well pileup arrays inside call_consensus_with_metrics
# (~45 B per aligned base, measured, and allocated for every read of a well in
# one vectorised pass).
#
# 32 MB keeps the batch pileup near the 1.4 GB mark for a worker whose threads
# all land on batch-sized wells at once, and leaves the reference fixture and
# every barcode up to ~65 MB of slices in a single batch (i.e. unchanged from
# the previous all-wells-at-once path). A single well larger than the budget
# still forms its own batch: bounding one well below its own depth is not
# possible from here, it needs the accumulation inside consensus.py to become
# incremental. 0 disables batching.
#
# Fallback only, as with _WELL_BUFFER_MB_DEFAULT above.
_CONSENSUS_BATCH_MB_DEFAULT = 32


# --------------------------------------------------------------------------
# Adaptive sizing of the two memory bounds above
# --------------------------------------------------------------------------
#
# Both defaults were measured on ONE box (10 cores, 15 GiB, 3 native barcodes)
# and neither number is a property of the workload: they are a property of how
# much RAM that box had. Shipping them fixed is wrong in both directions. An
# 8 GiB laptop running the same three workers gets 3 x 512 MB of slice text
# (~2.5 GB of RSS after the ~60% object overhead) plus 3 x the batch pileup on
# top of the aligner, which is the OOM the bound exists to prevent; a 64 GiB
# workstation spills and re-reads for no reason and batches finer than it needs
# to, paying the ~0.016 s per-batch fixed cost more often than necessary.
#
# So derive both from the memory limit, the same philosophy the ingest fan-out
# uses for the filesystem (``fasta_parser._PROBE_LATENCY_S``): measure or read
# the environment, never guess it from a path or a hardcoded profile. Here the
# environment can simply be ASKED rather than probed, so there is no probe.
#
# Denominator choice, stated because it is the load-bearing assumption:
#
# * The limit, not the free memory. ``MemAvailable`` moves with whatever else
#   the user has open, so deriving from it would make two runs of the same data
#   on the same box pick different budgets, and a run started next to a browser
#   would silently size itself for a machine it is not on. Output is identical
#   either way (see the identity note below), but a value that jitters is not
#   one you can reason about from a bug report. The limit is a stable property
#   of the box, and the clamp floor is what protects the genuinely small box.
# * cgroup before /proc/meminfo. Inside a container ``MemTotal`` is the HOST's
#   RAM, which is exactly the case where over-sizing is fatal: the kernel OOM
#   killer enforces ``memory.max``, and the host figure can be an order of
#   magnitude larger. cgroup v2 nests, so the effective limit is the tightest
#   value on the path from this process's cgroup up to the root, not just the
#   leaf's.
#
# IDENTITY: neither budget can change demux output. The well buffer only
# decides whether a slice list is spilled to disk and re-read (``_WellReadBuffer``
# preserves append order across a spill by construction), and the consensus
# batch only decides how many wells share one minimap2 call, with ``name_offset``
# keeping every QNAME identical to the all-wells case. Both invariants predate
# this change; the acceptance test below drives the budgets to their extremes
# and checks the tree hash is unmoved.

#: cgroup v2 leaf: this many bytes is "no limit".
_CGROUP_UNLIMITED = "max"
#: cgroup v1 writes a huge sentinel rather than a word, anything at or above
#: this is "no limit" (the exact value varies with PAGE_SIZE).
_CGROUP_V1_UNLIMITED = 1 << 62

#: Per-worker share of the limit given to the assigned-slice buffer. 0.10 of
#: 16.04 GB / 3 workers is 535 MB, i.e. it reproduces the measured 512 on the
#: box the constant was tuned on. Slice text costs ~1.6x its size in RSS, so
#: this is ~16% of the share in resident terms and leaves the majority to the
#: aligner, the pileup transients and the interpreter.
_WELL_BUFFER_FRACTION = 0.10
#: Floor/ceiling in MB. The floor keeps a tiny container from spilling on every
#: append; the ceiling stops a big box from turning the bound into a no-op
#: (past a few GB of buffered text the spill path is the cheaper behaviour
#: anyway, and an unbounded buffer is what needed 14 GB on the real run).
_WELL_BUFFER_MB_MIN = 64
_WELL_BUFFER_MB_MAX = 4096

#: Per-worker share of the limit given to one consensus batch of sequence text.
#: 0.006 of 16.04 GB / 3 is 32.1 MB, reproducing the measured 32. The pileup
#: transients this bounds are ~45 B per aligned base and are multiplied by the
#: consensus ThreadPool width, which is why the fraction is so much smaller
#: than the buffer's.
_CONSENSUS_BATCH_FRACTION = 0.006
#: Floor/ceiling in MB. The floor is what keeps batch COUNT from exploding:
#: each batch costs ~0.016 s of fixed work (tempdir, FASTA write, minimap2
#: spawn, index build) and at s3 the 25 batches already came to 8.6 s, 4.8% of
#: the run, so a budget small enough to double that is worse than the spill it
#: avoids. The ceiling is set past the largest value ever measured as useful
#: (524288 query bases of batch was neutral against 262144, consensus-depth.md
#: section 4) so a huge box widens but does not run unbounded.
_CONSENSUS_BATCH_MB_MIN = 8
_CONSENSUS_BATCH_MB_MAX = 256


def _read_cgroup_v2_limit(
    root: Path = Path("/sys/fs/cgroup"),
    proc_cgroup: Path = Path("/proc/self/cgroup"),
) -> int | None:
    """Tightest ``memory.max`` from this process's cgroup up to the root.

    cgroup v2 limits nest: a pod can sit inside a slice that is itself capped,
    and the kernel enforces the minimum. Reading only the leaf would over-size
    in exactly that layout. Returns ``None`` when unified cgroups are absent or
    every level on the path says ``max``.

    The two paths are parameters purely so the nesting behaviour can be tested
    against a temp-dir replica; nothing in production passes them.
    """
    try:
        # Unified hierarchy always presents "0::<relative path>".
        rel = ""
        for line in proc_cgroup.read_text().splitlines():
            parts = line.split(":", 2)
            if len(parts) == 3 and parts[0] == "0":
                rel = parts[2].strip()
                break
        if not rel:
            return None
    except OSError:
        return None

    node = root / rel.lstrip("/")
    best: int | None = None
    # Walk leaf -> root. Bounded by the path depth; `root` itself has no
    # memory.max, so the loop simply stops finding files.
    while True:
        try:
            raw = (node / "memory.max").read_text().strip()
        except OSError:
            raw = ""
        if raw and raw != _CGROUP_UNLIMITED:
            try:
                val = int(raw)
            except ValueError:
                val = 0
            if val > 0:
                best = val if best is None else min(best, val)
        if node == root or root not in node.parents:
            break
        node = node.parent
    return best


def _read_cgroup_v1_limit() -> int | None:
    """``memory.limit_in_bytes`` from a legacy cgroup v1 memory controller."""
    try:
        val = int(
            Path("/sys/fs/cgroup/memory/memory.limit_in_bytes").read_text().strip()
        )
    except (OSError, ValueError):
        return None
    return val if 0 < val < _CGROUP_V1_UNLIMITED else None


def _read_phys_mem() -> int | None:
    """Physical RAM in bytes, or ``None`` if the platform will not say.

    ``os.sysconf`` covers Linux and macOS. Windows has neither it nor
    ``/proc``, so frozen Windows builds fall through to the fixed defaults;
    that is the pre-existing behaviour, not a regression.
    """
    try:
        pages = os.sysconf("SC_PHYS_PAGES")
        page_size = os.sysconf("SC_PAGE_SIZE")
    except (AttributeError, ValueError, OSError):
        return None
    if pages > 0 and page_size > 0:
        return pages * page_size
    return None


def _memory_limit_bytes() -> tuple[int | None, str]:
    """Bytes this process tree may use, plus a one-word provenance tag.

    The tag is carried into the timing record so a support question can be
    answered without reproducing the environment: it says whether the number
    came from a container limit or from the physical box.
    """
    phys = _read_phys_mem()
    for reader, tag in (
        (_read_cgroup_v2_limit, "cgroup_v2"),
        (_read_cgroup_v1_limit, "cgroup_v1"),
    ):
        limit = reader()
        if limit is None:
            continue
        # A cgroup may be capped ABOVE the physical RAM (common on unconstrained
        # container runtimes). The binding constraint is whichever is smaller.
        if phys is not None and phys < limit:
            return phys, "meminfo"
        return limit, tag
    if phys is not None:
        return phys, "meminfo"
    return None, "unknown"


def _derive_mb(
    limit_bytes: int | None,
    workers: int,
    fraction: float,
    lo_mb: int,
    hi_mb: int,
    fallback_mb: int,
) -> int:
    """One budget in MB: a clamped fraction of this worker's share of *limit*.

    *workers* is the number of native-barcode processes that will hold such a
    budget CONCURRENTLY. Dividing by it is the whole point: the constants were
    read off a single-worker RSS figure but three of them run at once, so an
    undivided budget is a 3x under-count of what the box is asked for.
    """
    if limit_bytes is None:
        return fallback_mb
    per_worker_mb = limit_bytes / max(1, workers) / 1_000_000
    return int(max(lo_mb, min(hi_mb, per_worker_mb * fraction)))


def _resolve_memory_budgets(workers: int) -> dict[str, Any]:
    """Well-buffer and consensus-batch budgets in MB, with their provenance.

    An explicit environment variable always wins, including the documented
    ``0`` that disables a bound outright; adaptation only fills in the value
    nobody chose.
    """
    limit, source = _memory_limit_bytes()
    info: dict[str, Any] = {
        "mem_limit_bytes": limit,
        "mem_limit_source": source,
        "mem_workers": workers,
    }

    for key, env_name, fraction, lo, hi, fallback in (
        (
            "well_buffer_mb",
            "KUMA_MAME_WELL_BUFFER_MB",
            _WELL_BUFFER_FRACTION,
            _WELL_BUFFER_MB_MIN,
            _WELL_BUFFER_MB_MAX,
            _WELL_BUFFER_MB_DEFAULT,
        ),
        (
            "consensus_batch_mb",
            "KUMA_MAME_CONSENSUS_BATCH_MB",
            _CONSENSUS_BATCH_FRACTION,
            _CONSENSUS_BATCH_MB_MIN,
            _CONSENSUS_BATCH_MB_MAX,
            _CONSENSUS_BATCH_MB_DEFAULT,
        ),
    ):
        raw = os.environ.get(env_name, "").strip()
        if raw:
            try:
                info[key] = max(0, int(raw))
                info[key + "_source"] = "env"
                continue
            except ValueError:
                log.warning(
                    "%s=%r is not an integer; deriving the budget instead.",
                    env_name,
                    raw,
                )
        info[key] = _derive_mb(limit, workers, fraction, lo, hi, fallback)
        info[key + "_source"] = "derived" if limit is not None else "fallback"
    return info


# Dynamic core budget for the per-native-barcode ProcessPool. The work unit is
# one native barcode, and the three barcodes of the real plate measure
# 2837 : 1945 : 849 MB, a 3.34 : 1 spread that the worker walls track almost
# exactly. Under a static cpu//P share the two small workers therefore finish
# around a third and two thirds of the way through and their cores then idle
# (scale-profile section 5, unit-balance section 1). Worth 8.6 percent of the
# s3 wall, measured. ``KUMA_MAME_CORE_BUDGET=0`` restores the static share.
_CORE_BUDGET_DEFAULT = 1


class _CoreBudget:
    """This worker's current share of the box, as siblings finish.

    Holds a shared counter of live workers (a ``Manager().Value`` proxy, the
    same transport the progress queue already uses) plus the static floor. The
    parent decrements the counter as each unit completes, so ``threads()``
    only ever grows and never drops below the static share.

    Only minimap2 ``-t`` is widened. The consensus ThreadPool is deliberately
    left at its static width: its concurrency multiplies the per-batch pileup
    arrays, which is exactly the term ``KUMA_MAME_CONSENSUS_BATCH_MB`` exists
    to bound, so widening it would trade the memory bound for wall.

    Thread count does not change demux output: the P=3x3, P=2x5 and P=1x10
    arrangements in scale-profile section 5 all produced tree digest
    9d106bae4d32. minimap2 seeds its per-read RNG from the query NAME, not from
    the thread that happens to pick the read up.
    """

    __slots__ = ("_live", "_floor", "_cpu")

    def __init__(self, live, floor: int, cpu: int) -> None:
        self._live = live
        self._floor = max(1, floor)
        self._cpu = max(1, cpu)

    def threads(self) -> int:
        try:
            live = int(self._live.value)
        except Exception:  # noqa: BLE001 - proxy gone: fall back to the static share
            return self._floor
        return max(self._floor, self._cpu // max(1, live))


class _WellReadBuffer:
    """Assigned read slices per well, with a bounded in-memory footprint.

    Appends land in RAM. Once the buffered sequence text passes ``budget``
    every well is flushed to a spill file under a private temp dir and the RAM
    lists are cleared, so the resident set stops tracking the input size.
    :meth:`load` concatenates a well's spill file with its RAM tail, which
    reproduces the append order exactly whether or not a spill happened.

    That order is load-bearing twice over: it fixes each read's synthetic QNAME
    in the consensus alignment, and consensus resolves per-position ties by
    first touch (``first_touch`` in ``ingest/consensus.py``). Nothing here may
    reorder reads within a well.

    The spill format is one ``read_id<TAB>sequence`` line per read.
    ``_iter_fastq`` takes the read id as ``header[1:].split()[0]``, so it can
    contain neither a tab nor a newline and needs no escaping.
    """

    __slots__ = ("_budget", "_mem", "_counts", "_sizes", "_bytes", "_tmp", "_spilled")

    def __init__(self, budget_bytes: int) -> None:
        self._budget = budget_bytes
        self._mem: dict[tuple[int, int], list[tuple[str, str]]] = defaultdict(list)
        self._counts: dict[tuple[int, int], int] = {}
        self._sizes: dict[tuple[int, int], int] = {}
        self._bytes = 0
        self._tmp: tempfile.TemporaryDirectory | None = None
        self._spilled: set[tuple[int, int]] = set()

    @property
    def spilled(self) -> bool:
        """True once anything has been written to disk (final after the read loop)."""
        return bool(self._spilled)

    def append(self, well: tuple[int, int], read_id: str, seq: str) -> None:
        self._mem[well].append((read_id, seq))
        self._counts[well] = self._counts.get(well, 0) + 1
        self._sizes[well] = self._sizes.get(well, 0) + len(seq)
        self._bytes += len(seq) + len(read_id) + 1
        if self._budget and self._bytes > self._budget:
            self.flush()

    def wells(self) -> list[tuple[int, int]]:
        """Well keys in first-append order (the old ``per_well`` dict order)."""
        return list(self._counts)

    def counts(self) -> dict[tuple[int, int], int]:
        return dict(self._counts)

    def sizes(self) -> dict[tuple[int, int], int]:
        """Sequence bytes per well, whether resident or spilled."""
        return dict(self._sizes)

    def _path(self, well: tuple[int, int]) -> Path:
        assert self._tmp is not None
        return Path(self._tmp.name) / f"{well[0]}_{well[1]}.tsv"

    def flush(self) -> None:
        """Append every buffered well to its spill file and drop the RAM lists."""
        if self._tmp is None:
            self._tmp = tempfile.TemporaryDirectory(prefix="kuma_mame_wells_")
        with TIMER.phase("well_buffer_spill"):
            for well, reads in self._mem.items():
                if not reads:
                    continue
                with self._path(well).open("a", encoding="utf-8") as fh:
                    fh.writelines(f"{rid}\t{seq}\n" for rid, seq in reads)
                self._spilled.add(well)
        self._mem.clear()
        self._bytes = 0

    def load(self, well: tuple[int, int]) -> list[tuple[str, str]]:
        """Return this well's reads in append order (spilled part first)."""
        out: list[tuple[str, str]] = []
        if well in self._spilled:
            with TIMER.phase("well_buffer_reload"):
                with self._path(well).open(encoding="utf-8") as fh:
                    for line in fh:
                        rid, _tab, seq = line.rstrip("\n").partition("\t")
                        out.append((rid, seq))
        out.extend(self._mem.get(well, ()))
        return out

    def release(self, well: tuple[int, int]) -> None:
        """Drop a well's reads once its consensus is written."""
        reads = self._mem.pop(well, None)
        if reads:
            self._bytes -= sum(len(s) + len(r) + 1 for r, s in reads)
        if well in self._spilled:
            self._spilled.discard(well)
            self._path(well).unlink(missing_ok=True)

    def close(self) -> None:
        if self._tmp is not None:
            self._tmp.cleanup()
            self._tmp = None


def _iter_chunks(
    it: Iterator[tuple[str, str]], size: int
) -> Iterator[list[tuple[str, str]]]:
    """Yield successive ``size``-length lists from *it* (last may be shorter)."""
    chunk: list[tuple[str, str]] = []
    for item in it:
        chunk.append(item)
        if len(chunk) >= size:
            yield chunk
            chunk = []
    if chunk:
        yield chunk


def _match_reads_chunk(
    chunk: list[tuple[int, str, str, list[Alignment]]],
    r_barcodes: list[tuple[str, str]],
    f_barcodes: list[tuple[str, str]],
    window_bp: int,
    edit_dist_ratio: float,
    trim_flank_bp: int,
) -> list[tuple[int, list[tuple[int, int, str]], int, int, int]]:
    """Pure per-read barcode matching for the chimera (multi-hit) path.

    Module-level (no closure) so it is picklable for a ``spawn`` ProcessPool.
    Mirrors the serial loop body in :func:`run_combinatorial_demux` exactly
    (slice extraction, per-hit ``_demux_read_anchored``, read-local dedup,
    first-hit vs chimera-split classification). The matching logic itself is
    unchanged; only the accumulation is returned to the caller instead of
    mutating shared ``per_well``/``stats``.

    Parameters
    ----------
    chunk:
        ``(read_index, read_id, read_seq, hits)`` tuples. ``read_index`` is the
        position in the original ``multi_results`` list, used by the caller to
        re-sort results into input order before appending.

    Returns
    -------
    One tuple per input read: ``(read_index, appends, assigned_delta,
    chimera_delta, ambiguous_delta)`` where ``appends`` is the ordered list of
    ``(r_idx, f_idx, slice_seq)`` to push onto ``per_well[(r_idx, f_idx)]`` and
    the three deltas are this read's contribution to the matching stats.
    """
    # Read-invariant barcode preprocessing, hoisted out of the per-read loop.
    plan = _build_barcode_plan(r_barcodes, f_barcodes, edit_dist_ratio)

    out: list[tuple[int, list[tuple[int, int, str]], int, int, int]] = []
    for read_index, _read_id, read_seq, hits in chunk:
        assigned_wells_this_read: set[tuple[int, int]] = set()
        is_first_hit = True
        appends: list[tuple[int, int, str]] = []
        assigned_delta = 0
        chimera_delta = 0
        ambiguous_delta = 0

        for hit in hits:
            slice_start = max(0, hit.q_st - trim_flank_bp)
            slice_end = min(len(read_seq), hit.q_en + trim_flank_bp)
            slice_seq = read_seq[slice_start:slice_end]

            q_st_in_slice = hit.q_st - slice_start
            q_en_in_slice = hit.q_en - slice_start

            result = _demux_read_anchored(
                read_seq=slice_seq,
                q_st=q_st_in_slice,
                q_en=q_en_in_slice,
                strand=hit.strand,
                r_barcodes=r_barcodes,
                f_barcodes=f_barcodes,
                window_bp=window_bp,
                edit_dist_ratio=edit_dist_ratio,
                plan=plan,
            )
            if result is None:
                # A hit that resolved to no well is not an assignment, so it
                # must not consume the read's "first assignment" slot; doing so
                # mis-filed the next successful hit as a chimera split.
                ambiguous_delta += 1
                continue

            r_idx, f_idx = result
            well = (r_idx, f_idx)

            if well in assigned_wells_this_read:
                # Duplicate of a well already assigned for this read; the slot
                # was consumed by that earlier hit, not by this one.
                continue

            assigned_wells_this_read.add(well)
            appends.append((r_idx, f_idx, slice_seq))

            if is_first_hit:
                assigned_delta += 1
            else:
                chimera_delta += 1
            is_first_hit = False

        out.append(
            (read_index, appends, assigned_delta, chimera_delta, ambiguous_delta)
        )
    return out


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------


def run_combinatorial_demux(
    raw_fastq_paths: list[Path],
    reference_fasta: Path,
    barcodes_xlsx: Path,
    output_dir: Path,
    mapq_threshold: int = 25,
    coverage_fraction: float = 0.98,
    trim_flank_bp: int = 30,
    min_depth: int = 3,
    window_bp: int = 30,
    edit_dist_ratio: float = 0.25,
    chimera_split: bool = True,
    well_consensus_at_root: bool = False,
    minimap2_threads: int | None = None,
    consensus_workers: int | None = None,
    per_read_parallel: bool = False,
    progress_callback: Callable[[int, int, str], None] | None = None,
    barcode_prefixes: tuple[list[tuple[str, str]], list[tuple[str, str]]] | None = None,
    core_budget: "_CoreBudget | None" = None,
    mem_workers: int = 1,
) -> DemuxResult:
    """MAPQ-filtered alignment-anchored fuzzy per-well demux with chimera splitting.

    Aligns pooled reads to a single reference, applies coverage filter,
    assigns each read (or each hit within a chimeric/concatemer read) to an
    R x F well by alignment-anchored edlib fuzzy barcode matching, and calls
    majority-vote consensus per well.

    Parameters
    ----------
    raw_fastq_paths:
        FASTQ(.gz) input files (all reads pooled before alignment).
    reference_fasta:
        Single-record DNA FASTA used as alignment reference.
    barcodes_xlsx:
        xlsx with isps_f_1..12 and isps_r_1..8 barcode sequences.
    output_dir:
        Directory for output files.
        Per-well FASTA: ``{output_dir}/{R_idx}_{F_idx}.fasta``
        Consensus FASTA: ``{output_dir}/consensus/{R_idx}_{F_idx}.fasta``
    mapq_threshold:
        Minimum MAPQ per alignment hit (default 25).
    coverage_fraction:
        Minimum fraction of reference covered by each hit (default 0.98).
        Replaces the former strict 100%-span filter; recovers reads with
        1-2 bp end-clip while keeping spurious partial alignments out.
    trim_flank_bp:
        Bases to include on each side of each hit's aligned region when
        extracting the slice written to FASTA (default 30).
    min_depth:
        Minimum read depth per position for consensus call (default 3).
    window_bp:
        Search window radius around alignment anchors for barcode matching
        (default 30 bp).
    edit_dist_ratio:
        Max allowed edit distance fraction of barcode length (default 0.25).
        Threshold = floor(len(bc) * ratio).
    chimera_split:
        When True (default), iterate ALL passing alignment hits per read and
        attempt demux for each hit independently.  A chimeric read carrying
        two different amplicon copies may contribute to two wells.
        When False, only the first passing hit is used (legacy behaviour).
    well_consensus_at_root:
        When True, write single-record per-well consensus FASTA files at the
        top level of ``output_dir`` (so a non-recursive top-level ``*.fasta``
        glob sees only consensus files), the multi-record per-well reads under
        ``output_dir/reads/``, and the combined consensus FASTA under
        ``output_dir/final/``.  When False (default), keep the legacy layout
        (reads at root, consensus under ``output_dir/consensus/``, combined at
        root).
    minimap2_threads:
        Thread count passed through to the alignment minimap2 invocation.
        ``None`` (default) keeps the module-level auto-detected default.
    consensus_workers:
        Worker count for the per-well consensus ThreadPool.  ``None`` (default)
        keeps the module-level ``_CONSENSUS_WORKERS`` default.
    barcode_prefixes:
        Pre-parsed ``(r_barcodes, f_barcodes)`` as returned by
        :func:`load_barcode_prefixes`.  ``None`` (default) parses
        *barcodes_xlsx* here, preserving the original behaviour.  Supplying it
        lets a caller that already parsed the workbook (e.g. the per-native-
        barcode ProcessPool parent) skip both the parse and the ~1.4 s
        ``openpyxl`` import inside every worker process.
    core_budget:
        Optional :class:`_CoreBudget` shared with the per-native-barcode parent.
        When supplied, every minimap2 invocation raises its ``-t`` to this
        process's current share of the box as sibling workers finish, instead of
        holding the static share for the whole run.  ``None`` (default) keeps
        *minimap2_threads* fixed.  Not part of the RPC surface: it is a
        parent-to-worker scheduling handle, not a demux parameter.
    mem_workers:
        How many native-barcode processes hold a memory budget concurrently,
        i.e. the divisor for this process's share of the box (see
        :func:`_resolve_memory_budgets`).  ``1`` (default) is correct for the
        serial path and for any direct caller.  Like *core_budget* this is a
        parent-to-worker scheduling handle and not part of the RPC surface: it
        cannot change output, only how often the run spills and re-reads.

    Returns
    -------
    DemuxResult with stats, per_well_reads, per_well_consensus.
    """
    # Resolve the two memory bounds once, here, so they can be both LOGGED and
    # folded into this worker's timing record. A user hitting an out-of-memory
    # or an unexpected spill can be asked for one line instead of for their
    # container spec: the record carries the limit, where it was read from, the
    # divisor, each budget and whether it was derived or forced by env.
    budgets = _resolve_memory_budgets(mem_workers)
    log.info(
        "Memory budgets: well_buffer=%d MB (%s), consensus_batch=%d MB (%s); "
        "limit=%s MB via %s / %d worker(s)",
        budgets["well_buffer_mb"],
        budgets["well_buffer_mb_source"],
        budgets["consensus_batch_mb"],
        budgets["consensus_batch_mb_source"],
        round(budgets["mem_limit_bytes"] / 1_000_000)
        if budgets["mem_limit_bytes"]
        else "unknown",
        budgets["mem_limit_source"],
        budgets["mem_workers"],
    )
    with TIMER.session("demux", output_dir=str(output_dir), **budgets):
        return _run_combinatorial_demux_body(
            budgets=budgets,
            raw_fastq_paths=raw_fastq_paths,
            reference_fasta=reference_fasta,
            barcodes_xlsx=barcodes_xlsx,
            output_dir=output_dir,
            mapq_threshold=mapq_threshold,
            coverage_fraction=coverage_fraction,
            trim_flank_bp=trim_flank_bp,
            min_depth=min_depth,
            window_bp=window_bp,
            edit_dist_ratio=edit_dist_ratio,
            chimera_split=chimera_split,
            well_consensus_at_root=well_consensus_at_root,
            minimap2_threads=minimap2_threads,
            consensus_workers=consensus_workers,
            per_read_parallel=per_read_parallel,
            progress_callback=progress_callback,
            barcode_prefixes=barcode_prefixes,
            core_budget=core_budget,
        )


def _run_combinatorial_demux_body(
    raw_fastq_paths: list[Path],
    reference_fasta: Path,
    barcodes_xlsx: Path,
    output_dir: Path,
    mapq_threshold: int,
    coverage_fraction: float,
    trim_flank_bp: int,
    min_depth: int,
    window_bp: int,
    edit_dist_ratio: float,
    chimera_split: bool,
    well_consensus_at_root: bool,
    minimap2_threads: int | None,
    consensus_workers: int | None,
    per_read_parallel: bool,
    progress_callback: Callable[[int, int, str], None] | None,
    barcode_prefixes: tuple[list[tuple[str, str]], list[tuple[str, str]]] | None = None,
    core_budget: "_CoreBudget | None" = None,
    budgets: dict[str, Any] | None = None,
) -> DemuxResult:
    """Body of :func:`run_combinatorial_demux` (see there for semantics).

    Split out only so the public entry point can wrap the whole run in one
    :meth:`PhaseTimer.session`; behaviour is unchanged.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    if well_consensus_at_root:
        reads_dir = output_dir / "reads"
        reads_dir.mkdir(exist_ok=True)
        final_dir = output_dir / "final"
        final_dir.mkdir(exist_ok=True)
        consensus_dir = output_dir
        combined_path = final_dir / _COMBINED_CONSENSUS_FILENAME
    else:
        consensus_dir = output_dir / "consensus"
        consensus_dir.mkdir(exist_ok=True)
        reads_dir = output_dir
        combined_path = output_dir / _COMBINED_CONSENSUS_FILENAME

    stats = DemuxStats()

    with TIMER.phase("load_barcodes"):
        # A caller that already parsed the workbook passes the result in; that
        # skips this process's first `import openpyxl` too (the import lives
        # inside load_barcode_prefixes and dominates the phase: ~1.4 s import
        # vs ~0.01 s parse), which matters once per worker process.
        if barcode_prefixes is not None:
            r_barcodes, f_barcodes = barcode_prefixes
        else:
            r_barcodes, f_barcodes = load_barcode_prefixes(barcodes_xlsx)
    log.info(
        "Loaded %d R barcodes, %d F barcodes (prefix-only, annealing tail stripped)",
        len(r_barcodes),
        len(f_barcodes),
    )

    # Read-invariant barcode preprocessing (RC'd R prefixes, max prefix lengths,
    # per-barcode edit thresholds) computed once for the whole run instead of
    # per read. The ProcessPool path rebuilds it inside each worker chunk
    # (_match_reads_chunk) rather than pickling it, keeping the payload as-is.
    barcode_plan = _build_barcode_plan(r_barcodes, f_barcodes, edit_dist_ratio)

    ref_len = _get_reference_length(reference_fasta)
    log.info("Reference length: %d bp", ref_len)

    def _mm_threads() -> int | None:
        """minimap2 ``-t`` for the next invocation.

        Re-read per call rather than once per run: with a core budget the share
        grows as sibling native-barcode workers finish, and every minimap2 call
        is a fresh subprocess, so the newer share applies from the next chunk on
        with no restart. Without a budget this is the static value and the call
        is a plain attribute read.
        """
        if core_budget is None:
            return minimap2_threads
        return core_budget.threads()

    # Assigned read slices, held between the read loop and consensus. Bounded:
    # see _WELL_BUFFER_MB_DEFAULT and _resolve_memory_budgets. Chunking bounds
    # the alignment stage only, this buffer is what grows with the whole barcode.
    # A direct caller of the body (tests) gets a single-worker derivation.
    if budgets is None:
        budgets = _resolve_memory_budgets(1)
    per_well = _WellReadBuffer(max(0, int(budgets["well_buffer_mb"])) * 1_000_000)

    # Chunk-stream read loading: load + align in N-read chunks instead of
    # materialising the whole FASTQ. Each chunk's minimap2 input/SAM and its
    # Alignment lists are dropped between iterations, lowering alignment-stage
    # peak RAM only (per_well still accumulates across chunks to consensus).
    # Chunks run in input order and the per-read pool re-sorts each chunk by
    # read_index, so the per_well append order (and thus the consensus tie-break)
    # is stable; stats are accumulated across chunks (total_reads/passed_*/
    # per-read deltas all use +=).
    #
    # Changing the chunk size preserves the output, but only because the running
    # name_offset below keeps every read's synthetic QNAME identical to the
    # single-chunk case. See the _READ_CHUNK_DEFAULT comment for why the QNAME
    # is load-bearing.
    _chunk_size = int(
        os.environ.get("KUMA_MAME_READ_CHUNK", str(_READ_CHUNK_DEFAULT))
    )
    _chunk_size = max(1, _chunk_size)
    _read_chunks = _iter_chunks(_iter_fastq(raw_fastq_paths), _chunk_size)

    # Read/align pipelining: run the reader one chunk ahead on a background
    # thread so gzip decompression of chunk N+1 overlaps the alignment of chunk
    # N instead of sitting on the critical path. Order is preserved (single
    # producer, FIFO queue), so consensus tie-break is unaffected. With a single
    # chunk there is nothing to overlap and this is a no-op.
    _prefetch_depth = int(
        os.environ.get("KUMA_MAME_FASTQ_PREFETCH", str(_FASTQ_PREFETCH_DEFAULT))
    )
    if _prefetch_depth > 0:
        _read_chunks = _prefetch(_read_chunks, _prefetch_depth)

    # Running count of reads already handed to the aligner, passed as
    # name_offset so a read keeps the same synthetic QNAME whatever the chunk
    # size. Without it every chunk restarts the numbering at 0 and minimap2
    # (per-read RNG seeded from the QNAME hash) returns different hits for a
    # handful of reads, which is what made the chunk size change the output.
    # It advances by the number of NON-EMPTY reads because _write_reads_fasta
    # skips empty sequences when it assigns indices.
    #
    # The prefetch thread only reorders WHEN a chunk is produced, never in which
    # order chunks are consumed (single producer + FIFO), so the offset the loop
    # below assigns to each chunk is unchanged by it.
    _name_offset = 0

    # Phase timing: charging only the generator's `next()` to fastq_read keeps
    # gzip decompression + parsing separate from the per-chunk work below, with
    # one timer pair per chunk (never per read). With prefetch on, this measures
    # residual *waiting* for the reader, not the reader's total cost.
    for chunk_reads in timed_iter(_read_chunks, "fastq_read"):
        stats.total_reads += len(chunk_reads)
        _chunk_offset = _name_offset
        _name_offset += sum(1 for _rid, _seq in chunk_reads if _seq)

        if chimera_split:
            # --- multi-hit path: chimera / concatemer splitting ------------
            with TIMER.phase("align_minimap2"):
                multi_results, gate_counts = align_reads_multi_with_gate_counts(
                    reads=chunk_reads,
                    reference_fasta=reference_fasta,
                    preset="map-ont",
                    min_mapq=mapq_threshold,
                    coverage_fraction=coverage_fraction,
                    threads=_mm_threads(),
                    name_offset=_chunk_offset,
                )

            reads_with_hits = len(multi_results)
            total_hit_count = sum(len(hits) for _, _, hits in multi_results)
            # The two counters are per-gate and NOT interchangeable:
            # passed_mapq counts reads clearing MAPQ alone, passed_coverage the
            # subset that also cleared the coverage gate. Assigning both the
            # post-filter number (as this did before) made a coverage wipeout --
            # e.g. a whole-plasmid reference against amplicon reads -- read as a
            # MAPQ wipeout in the marker stats.
            stats.passed_mapq += gate_counts.reads_passed_mapq
            stats.passed_coverage += gate_counts.reads_passed_coverage
            log.info(
                "Passed MAPQ filter (chunk): %d reads; also passed coverage: "
                "%d reads / %d total hits",
                gate_counts.reads_passed_mapq,
                reads_with_hits,
                total_hit_count,
            )

            _demux_total = len(multi_results)
            _threshold = int(
                os.environ.get(
                    "KUMA_MAME_PERREAD_THRESHOLD", str(_PERREAD_THRESHOLD_DEFAULT)
                )
            )
            # Per-read ProcessPool fan-out is only safe/beneficial when this run
            # owns the cores (n_nb == 1, signalled by per_read_parallel) and the
            # dataset clears the spawn/pickle overhead break-even threshold. Below
            # the threshold, or when n_nb > 1 (the per-NB pool already saturates
            # cores), stay on the serial path. Both paths produce byte-identical
            # per_well/stats: parallel results are re-sorted to input (read_index)
            # order before append, and the matching logic is shared verbatim via
            # _match_reads_chunk.
            _use_perread_pool = (
                per_read_parallel and _demux_total >= _threshold
                and _demux_total > 0
            )

            if _use_perread_pool:
                # oversubscription guard: the per-NB pool is 1 here (n_nb == 1)
                # and the consensus ThreadPool runs after this loop, so the
                # per-read pool may use all cores.
                cpu = os.cpu_count() or 4
                _env_w = os.environ.get("KUMA_MAME_PERREAD_WORKERS", "").strip()
                if _env_w.isdigit() and int(_env_w) > 0:
                    pool_workers = min(int(_env_w), cpu)
                else:
                    pool_workers = cpu
                pool_workers = max(1, min(pool_workers, _demux_total))

                # Parent-side wall clock only. The matching work happens in
                # spawned workers, so this is NOT decomposable into their
                # internal phases and is not CPU time either (it overlaps
                # pool_workers processes).
                _t_pool = time.perf_counter()

                indexed = [
                    (i, rid, rseq, hits)
                    for i, (rid, rseq, hits) in enumerate(multi_results)
                ]
                chunk_size = max(1, (_demux_total + pool_workers - 1) // pool_workers)
                chunks = [
                    indexed[i : i + chunk_size]
                    for i in range(0, _demux_total, chunk_size)
                ]

                collected: list[
                    tuple[int, list[tuple[int, int, str]], int, int, int]
                ] = []
                ctx = _demux_mp_context()
                with ProcessPoolExecutor(
                    max_workers=pool_workers, mp_context=ctx
                ) as ex:
                    futs = [
                        ex.submit(
                            _match_reads_chunk,
                            chunk,
                            r_barcodes,
                            f_barcodes,
                            window_bp,
                            edit_dist_ratio,
                            trim_flank_bp,
                        )
                        for chunk in chunks
                    ]
                    _done = 0
                    for fut in as_completed(futs):
                        collected.extend(fut.result())
                        _done += 1
                        if progress_callback is not None:
                            progress_callback(_done, len(futs), "demux")

                # Re-sort to input order so per_well append order (and thus
                # consensus tie-break) matches the serial path exactly.
                collected.sort(key=lambda r: r[0])
                id_by_index = {i: rid for i, rid, _, _ in indexed}
                for (
                    read_index,
                    appends,
                    assigned_d,
                    chimera_d,
                    ambiguous_d,
                ) in collected:
                    read_id = id_by_index[read_index]
                    for r_idx, f_idx, slice_seq in appends:
                        per_well.append((r_idx, f_idx), read_id, slice_seq)
                    stats.assigned_reads += assigned_d
                    stats.chimera_splits += chimera_d
                    stats.ambiguous_dropped += ambiguous_d

                TIMER.add(
                    "barcode_match_parallel_wall", time.perf_counter() - _t_pool
                )
            else:
                _t_match = time.perf_counter()
                _demux_step = max(1, _demux_total // 100)  # ~1% interval throttle
                for _demux_i, (read_id, read_seq, hits) in enumerate(multi_results):
                    if progress_callback is not None and _demux_i % _demux_step == 0:
                        progress_callback(_demux_i, _demux_total, "demux")
                    # Track which wells this read has already been assigned to
                    # to prevent double-counting the same read in the same well.
                    assigned_wells_this_read: set[tuple[int, int]] = set()
                    is_first_hit = True

                    for hit in hits:
                        # Extract aligned slice + flanks from the raw read.
                        # Coordinates are in read (query) space; no strand flip.
                        slice_start = max(0, hit.q_st - trim_flank_bp)
                        slice_end = min(len(read_seq), hit.q_en + trim_flank_bp)
                        slice_seq = read_seq[slice_start:slice_end]

                        # Alignment anchors within the slice coordinate space.
                        # q_st/q_en are absolute positions in read_seq.
                        q_st_in_slice = hit.q_st - slice_start
                        q_en_in_slice = hit.q_en - slice_start

                        result = _demux_read_anchored(
                            read_seq=slice_seq,
                            q_st=q_st_in_slice,
                            q_en=q_en_in_slice,
                            strand=hit.strand,
                            r_barcodes=r_barcodes,
                            f_barcodes=f_barcodes,
                            window_bp=window_bp,
                            edit_dist_ratio=edit_dist_ratio,
                            plan=barcode_plan,
                        )
                        if result is None:
                            # A hit that resolved to no well is not an
                            # assignment, so it must not consume the read's
                            # "first assignment" slot; doing so mis-filed the
                            # next successful hit as a chimera split.
                            stats.ambiguous_dropped += 1
                            continue

                        r_idx, f_idx = result
                        well = (r_idx, f_idx)

                        if well in assigned_wells_this_read:
                            # Already assigned to this well from an earlier hit;
                            # that hit consumed the slot, not this one.
                            continue

                        assigned_wells_this_read.add(well)
                        per_well.append(well, read_id, slice_seq)

                        if is_first_hit:
                            stats.assigned_reads += 1
                        else:
                            stats.chimera_splits += 1
                        is_first_hit = False

                TIMER.add("barcode_match", time.perf_counter() - _t_match)

        else:
            # --- legacy single-hit path ------------------------------------
            with TIMER.phase("align_minimap2"):
                alignments, gate_counts = align_reads_with_gate_counts(
                    reads=chunk_reads,
                    reference_fasta=reference_fasta,
                    preset="map-ont",
                    min_mapq=mapq_threshold,
                    # Apply the SAME graded coverage filter as the multi-hit path
                    # (align_reads_multi). Collapsing it to require_full_span=
                    # (coverage_fraction >= 1.0) dropped the span filter entirely for
                    # any coverage_fraction < 1.0 (e.g. the 0.98 default), admitting
                    # partial-coverage reads into wells on the chimera_split=False path.
                    require_full_span=False,
                    coverage_fraction=coverage_fraction,
                    threads=_mm_threads(),
                    name_offset=_chunk_offset,
                )
            # Per-gate counters; see the multi-hit path above for why these two
            # must not both be set to the post-filter number.
            stats.passed_mapq += gate_counts.reads_passed_mapq
            stats.passed_coverage += gate_counts.reads_passed_coverage
            log.info(
                "Passed MAPQ filter (chunk): %d / %d; also passed coverage: %d",
                gate_counts.reads_passed_mapq,
                len(chunk_reads),
                len(alignments),
            )

            _t_match = time.perf_counter()
            for aln in alignments:
                trimmed = _trim_read(aln, aln.read_seq, trim_flank_bp)
                result = _demux_read_anchored(
                    read_seq=aln.read_seq,
                    q_st=aln.q_st,
                    q_en=aln.q_en,
                    strand=aln.strand,
                    r_barcodes=r_barcodes,
                    f_barcodes=f_barcodes,
                    window_bp=window_bp,
                    edit_dist_ratio=edit_dist_ratio,
                    plan=barcode_plan,
                )
                if result is None:
                    stats.ambiguous_dropped += 1
                    continue
                r_idx, f_idx = result
                per_well.append((r_idx, f_idx), aln.read_id, trimmed)
                stats.assigned_reads += 1
            TIMER.add("barcode_match", time.perf_counter() - _t_match)

    log.info("Total reads: %d", stats.total_reads)
    log.info(
        "Barcode-assigned reads: %d  chimera splits: %d  (ambiguous/no-match dropped: %d)",
        stats.assigned_reads,
        stats.chimera_splits,
        stats.ambiguous_dropped,
    )

    # Well inventory. The read slices themselves stay in the bounded buffer and
    # are pulled back one consensus batch at a time below, so nothing here
    # materialises the whole barcode.
    well_counts = per_well.counts()
    well_sizes = per_well.sizes()
    well_keys = per_well.wells()

    # The on-disk per-well reads FASTA is not read by any production code path;
    # it is off by default and only written when KUMA_MAME_KEEP_WELL_READS=1 is
    # set for post-hoc forensics. Writing one small file per well dominates
    # wall time on network/9p-backed output dirs. It is written inside the
    # consensus batch loop, where the reads are resident anyway.
    keep_well_reads = os.environ.get("KUMA_MAME_KEEP_WELL_READS", "").strip() == "1"

    stats.wells_with_reads = sum(1 for n in well_counts.values() if n >= 1)
    stats.wells_with_min_reads = sum(
        1 for n in well_counts.values() if n >= min_depth
    )
    log.info(
        "Wells with >=1 read: %d/96, wells with >=%d reads: %d/96",
        stats.wells_with_reads,
        min_depth,
        stats.wells_with_min_reads,
    )

    # Per-well consensus — parallel across wells (each well is independent)
    ref_seq = _read_reference_seq(reference_fasta)
    per_well_consensus: dict[str, str] = {}

    _consensus_total = len(well_keys)

    # Reads handed back to the caller. Populated only when the buffer never
    # spilled, i.e. when holding them costs nothing beyond what the run already
    # held; a spilled run reports read COUNTS instead (per_well_read_counts,
    # which is what every production consumer of this field actually reads).
    per_well_reads: dict[str, list[tuple[str, str]]] = {}
    _materialise_reads = not per_well.spilled

    # Build the reference minimap2 index once (map-ont preset, identical to the
    # consensus alignment preset) so the batched alignment below skips the
    # on-the-fly index build. The .mmi lives in a tempdir dropped right after
    # the last batch.
    _index_tmp = tempfile.TemporaryDirectory(prefix="kuma_mame_idx_")
    well_index: Path | None
    try:
        with TIMER.phase("build_index"):
            well_index = build_minimap2_index(
                reference_fasta, Path(_index_tmp.name) / "reference.mmi"
            )
    except Exception as exc:  # noqa: BLE001
        # Index prebuild is a pure performance optimisation. On any failure,
        # fall back to per-well on-the-fly indexing (reference_index=None) so
        # alignment output stays identical.
        log.warning(
            "minimap2 index prebuild failed (%s); per-well alignment will "
            "index the reference FASTA on the fly", exc
        )
        well_index = None

    # One minimap2 call per BATCH of wells instead of one per well.
    # Per-read results were verified identical over 92 wells and 4936 reads.
    # Note the independence argument is necessary but not sufficient: minimap2
    # maps each query independently and the seed-occurrence cutoffs come from
    # the prebuilt index rather than the query set, but its per-read RNG is
    # seeded from the query NAME hash, so regrouping that renumbers reads can
    # move a few alignments (see the CORRECTION above _READ_CHUNK_DEFAULT).
    # align_reads_grouped keeps each well's reads in input order, which
    # the consensus tie-break depends on. Threads: the per-well calls had to
    # stay at 1 because up to n_workers wells aligned concurrently; a batched
    # call can use this worker's whole allotted share.
    #
    # Batching is a memory bound, not a speed change: it caps the Alignment
    # objects and the consensus pileup arrays that are live at once (see
    # _CONSENSUS_BATCH_MB_DEFAULT). The running ``name_offset`` gives every
    # read the QNAME it would have had in a single all-wells call, so the batch
    # size does not move a single alignment; it is the same invariant the read
    # loop above maintains for the chunk size.
    n_workers = consensus_workers if consensus_workers is not None else _CONSENSUS_WORKERS
    _batch_threads_static = (
        minimap2_threads if minimap2_threads is not None else n_workers
    )

    def _batch_threads() -> int:
        """minimap2 ``-t`` for the next consensus batch alignment.

        Same widening as the read loop. ``n_workers`` (the consensus ThreadPool)
        stays static on purpose: see :class:`_CoreBudget`.
        """
        if core_budget is None:
            return max(1, _batch_threads_static)
        return max(1, core_budget.threads())

    _cons_batch_bytes = max(0, int(budgets["consensus_batch_mb"])) * 1_000_000

    def _well_batches() -> Iterator[list[tuple[int, int]]]:
        """Group wells, in append order, into batches of bounded sequence bytes.

        A well larger than the budget forms its own batch; splitting one well
        is not possible here, consensus needs all of its reads at once.
        """
        cur: list[tuple[int, int]] = []
        cur_bytes = 0
        for key in well_keys:
            size = well_sizes.get(key, 0)
            if cur and _cons_batch_bytes and cur_bytes + size > _cons_batch_bytes:
                yield cur
                cur, cur_bytes = [], 0
            cur.append(key)
            cur_bytes += size
        if cur:
            yield cur

    def _run_well(
        well_name: str,
        reads: list[tuple[str, str]],
        alignments: list[Alignment],
    ) -> tuple[
        str, str, int, int, float, int, float, int, int, int, int, int, int, float,
        int, int, int,
    ]:
        """Worker: returns consensus sequence, depth, and mix metrics."""
        (
            seq,
            depth,
            mixed_positions,
            max_minor_fraction,
            low_depth_positions,
            n_fraction,
            low_quality_bases,
            input_reads,
            aligned_reads,
            mapq_failed,
            span_failed,
            n_indel_event_positions,
            max_indel_event_fraction,
            max_del_run_length,
            consensus_net_indel,
            read_net_indel,
            min_variant_support,
            variant_positions,
            min_variant_support_depth,
        ) = _compute_well_consensus(
            well_name, reads, alignments, ref_seq, ref_len, min_depth,
        )
        return (
            well_name,
            seq,
            depth,
            mixed_positions,
            max_minor_fraction,
            low_depth_positions,
            n_fraction,
            low_quality_bases,
            input_reads,
            aligned_reads,
            mapq_failed,
            span_failed,
            n_indel_event_positions,
            max_indel_event_fraction,
            max_del_run_length,
            consensus_net_indel,
            read_net_indel,
            min_variant_support,
            variant_positions,
            min_variant_support_depth,
        )

    _consensus_done = 0
    # Wall time of the whole per-well consensus stage in this process. The
    # ``*_sum`` keys added inside _compute_well_consensus are summed over the
    # ThreadPool workers and can exceed this wall.
    _t_cons = time.perf_counter()
    # Running count of non-empty reads already numbered by align_reads_grouped,
    # so each batch continues the single-call numbering (see above).
    _cons_name_offset = 0
    try:
        with ThreadPoolExecutor(max_workers=n_workers) as pool:
            for _batch in _well_batches():
                groups = [
                    (f"{r_idx}_{f_idx}", per_well.load((r_idx, f_idx)))
                    for r_idx, f_idx in _batch
                ]
                with TIMER.phase("well_consensus.align_minimap2_batch"):
                    well_alignments_map = align_reads_grouped(
                        groups=groups,
                        reference_fasta=reference_fasta,
                        preset="map-ont",
                        min_mapq=0,           # trimmed reads; already filtered upstream
                        require_full_span=False,
                        threads=_batch_threads(),
                        reference_index=well_index,
                        name_offset=_cons_name_offset,
                    )
                # align_reads_grouped skips empty sequences when it assigns indices, so
                # the offset must advance by the same count.
                _cons_name_offset += sum(
                    1 for _wn, rds in groups for _rid, _seq in rds if _seq
                )

                if keep_well_reads:
                    # fsync=False: these per-well reads FASTA are an intermediate
                    # artifact fully reconstructible by re-running the unit (whose
                    # completion marker is written last, and IS fsync'd). Final
                    # consensus FASTA, the combined FASTA, and stage markers keep the
                    # default fsync=True.
                    with TIMER.phase("write_well_fasta"):
                        for well_name, reads in groups:
                            atomic_write_text(
                                reads_dir / f"{well_name}.fasta",
                                "".join(f">{rid}\n{seq}\n" for rid, seq in reads),
                                fsync=False,
                            )
                if _materialise_reads:
                    per_well_reads.update(groups)

                futures = {
                    pool.submit(_run_well, wn, rds, well_alignments_map.get(wn, [])): wn
                    for wn, rds in groups
                }
                for fut in as_completed(futures):
                    (
                        wn,
                        seq,
                        depth,
                        mixed_positions,
                        max_minor_fraction,
                        low_depth_positions,
                        n_fraction,
                        low_quality_bases,
                        input_reads,
                        aligned_reads,
                        mapq_failed,
                        span_failed,
                        n_indel_event_positions,
                        max_indel_event_fraction,
                        max_del_run_length,
                        consensus_net_indel,
                        read_net_indel,
                        min_variant_support,
                        variant_positions,
                        min_variant_support_depth,
                    ) = fut.result()
                    per_well_consensus[wn] = seq
                    atomic_write_text(
                        consensus_dir / f"{wn}.fasta",
                        format_consensus_fasta_record(
                            wn,
                            seq,
                            ConsensusMetadata(
                                depth=depth,
                                input_reads=input_reads,
                                aligned_reads=aligned_reads,
                                mapq_failed=mapq_failed,
                                span_failed=span_failed,
                                mixed_positions=mixed_positions,
                                max_minor_allele_fraction=max_minor_fraction,
                                low_depth_positions=low_depth_positions,
                                consensus_n_fraction=n_fraction,
                                low_quality_bases=low_quality_bases,
                                n_indel_event_positions=n_indel_event_positions,
                                max_indel_event_fraction=max_indel_event_fraction,
                                max_del_run_length=max_del_run_length,
                                consensus_net_indel=consensus_net_indel,
                                read_net_indel=read_net_indel,
                                consensus_n_fraction_basis=BASIS_COVERED,
                                min_variant_support=min_variant_support,
                                variant_positions=variant_positions,
                                min_variant_support_depth=min_variant_support_depth,
                            ),
                        ),
                        # fsync=False here, one fsync_directory below instead. Per-file
                        # fsync costs a filesystem round trip per well (~280 of them) and
                        # bought no end-to-end guarantee anyway: atomic_write_text never
                        # fsync'd the parent directory, so the os.replace that publishes
                        # the final name was not durable. Batching the durability point
                        # into a single directory fsync commits all those renames at
                        # once, and on ext4 (data=ordered) that metadata commit forces
                        # the newly allocated data blocks out first, so the batch stays
                        # "absent or complete". The authoritative completion signal is
                        # still the stage marker, written afterwards with fsync=True,
                        # and validate_marker rejects a missing or zero-length well.
                        fsync=False,
                    )
                    _consensus_done += 1
                    if progress_callback is not None:
                        progress_callback(_consensus_done, _consensus_total, "consensus")

                # This batch is written; drop its reads (RAM and spill file
                # alike) before the next batch pulls the next ones in. Skipped
                # when the reads were handed to the caller, which then owns
                # them and expects them to stay alive.
                if not _materialise_reads:
                    for key in _batch:
                        per_well.release(key)
                del groups, well_alignments_map
    finally:
        # Every batch aligned; the .mmi and any spill files are done with.
        _index_tmp.cleanup()
        per_well.close()

    fsync_directory(consensus_dir)

    TIMER.add("well_consensus_wall", time.perf_counter() - _t_cons)

    # Combined single-file consensus FASTA (all wells, sorted by R then F),
    # mirroring the Aporva pipeline's final/<...>_consensus_dna.fasta output.
    # The per-well consensus/ files above are still written.
    _combined_order = sorted(
        per_well_consensus,
        key=lambda w: tuple(int(part) for part in w.split("_")),
    )
    atomic_write_text(
        combined_path,
        "".join(f">{wn}\n{per_well_consensus[wn]}\n" for wn in _combined_order),
    )

    return DemuxResult(
        stats=stats,
        per_well_reads=per_well_reads,
        per_well_consensus=per_well_consensus,
        per_well_read_counts={
            f"{r_idx}_{f_idx}": n for (r_idx, f_idx), n in well_counts.items()
        },
    )


# ---------------------------------------------------------------------------
# Helpers for per-alignment processing
# ---------------------------------------------------------------------------


def _trim_read(aln: Alignment, original_seq: str, flank_bp: int) -> str:
    """Return the aligned region of a read with +/-flank_bp flanks."""
    start = max(0, aln.q_st - flank_bp)
    end = min(len(original_seq), aln.q_en + flank_bp)
    return original_seq[start:end]


def _compute_well_consensus(
    well_name: str,
    reads: list[tuple[str, str]],
    well_alignments: list[Alignment],
    ref_seq: str,
    ref_len: int,
    min_depth: int,
) -> tuple[
    str, int, int, float, int, float, int, int, int, int, int, int, float, int,
    int, int, float | None, int, int,
]:
    """Call consensus for one well from its (pre-computed) alignments.

    ``well_alignments`` comes from the single batched :func:`align_reads_grouped`
    call for the whole unit and is in this well's original read order, which the
    consensus tie-break depends on.
    """
    if not reads:
        return (
            "N" * ref_len,
            0,
            0,
            0.0,
            ref_len,
            1.0 if ref_len > 0 else 0.0,
            0,
            0,
            0,
            0,
            0,
            0,
            0.0,
            0,
            0,
            0,
            None,
            0,
            0,
        )

    if not well_alignments:
        log.debug(
            "Well %s: 0 alignments from %d trimmed reads", well_name, len(reads)
        )
        return (
            "N" * ref_len,
            0,
            0,
            0.0,
            ref_len,
            1.0 if ref_len > 0 else 0.0,
            0,
            len(reads),
            0,
            0,
            0,
            0,
            0.0,
            0,
            0,
            0,
            None,
            0,
            0,
        )

    with TIMER.phase("well_consensus.compute_sum"):
        consensus_call = call_consensus_with_metrics(
            well_alignments,
            ref_seq,
            min_depth=min_depth,
        )
    return (
        consensus_call.consensus_seq,
        len(well_alignments),
        consensus_call.n_mixed_positions,
        consensus_call.max_minor_allele_fraction,
        consensus_call.n_low_depth_positions,
        consensus_call.consensus_n_fraction,
        consensus_call.n_low_quality_bases,
        len(reads),
        len(well_alignments),
        0,
        0,
        consensus_call.n_indel_event_positions,
        consensus_call.max_indel_event_fraction,
        consensus_call.max_del_run_length,
        consensus_call.consensus_net_indel_bp,
        consensus_call.median_read_net_indel_bp,
        consensus_call.min_variant_support,
        consensus_call.n_variant_positions,
        consensus_call.min_variant_support_depth,
    )


# ---------------------------------------------------------------------------
# Per-native-barcode parallel orchestration
# ---------------------------------------------------------------------------

# The 8 DemuxStats counters carried in each per-NB summary and summed into
# merged_stats. Single source of truth so the worker summary, the resume-seed
# from a marker, and the merge step all agree on the key set.
_DEMUX_NB_STAT_KEYS: tuple[str, ...] = (
    "total_reads", "passed_mapq", "passed_coverage", "assigned_reads",
    "ambiguous_dropped", "chimera_splits", "wells_with_reads",
    "wells_with_min_reads",
)


class _DirectProgressSink:
    """In-process stand-in for a multiprocessing progress queue (serial demux).

    Mirrors the ``put_nowait((nb_name, fraction))`` contract so the worker uses a
    single code path; forwards straight to the aggregate callback.
    """

    __slots__ = ("_cb",)

    def __init__(self, cb: Callable[[str, float], None]) -> None:
        self._cb = cb

    def put_nowait(self, item: tuple[str, float]) -> None:
        try:
            self._cb(item[0], item[1])
        except Exception:
            pass


def _demux_one_nb(payload: dict) -> dict:
    """ProcessPool worker: run one native barcode, return a picklable summary."""
    fastq = [Path(s) for s in payload["fastq_paths"]]
    q = payload.get("progress_queue")
    nb_name = payload["nb_name"]
    inner_cb: Callable[[int, int, str], None] | None = None
    if q is not None:
        _last = [0.0]

        def _inner_cb(done: int, total: int, stage: str) -> None:
            if total <= 0:
                return
            frac = done / total
            # Fold the two inner sub-phases into one 0..1 fraction for this NB:
            # read demux fills 0..0.85, per-well consensus 0.85..1.0.
            nb_f = 0.85 * frac if stage == "demux" else 0.85 + 0.15 * frac
            nb_f = max(0.0, min(1.0, nb_f))
            if nb_f - _last[0] >= 0.01 or nb_f >= 1.0:
                _last[0] = nb_f
                try:
                    q.put_nowait((nb_name, nb_f))
                except Exception:
                    pass
        inner_cb = _inner_cb

    # Reclaim the cores of native barcodes that already finished. Absent (or a
    # Manager that failed to start) => the static minimap2_threads share, as
    # before.
    _live = payload.get("live_workers")
    budget = (
        _CoreBudget(_live, payload["minimap2_threads"], payload.get("cpu_total", 1))
        if _live is not None
        else None
    )

    result = run_combinatorial_demux(
        raw_fastq_paths=fastq, reference_fasta=Path(payload["reference_fasta"]),
        barcodes_xlsx=Path(payload["barcodes_xlsx"]), output_dir=Path(payload["output_dir"]),
        mapq_threshold=payload["mapq_threshold"], coverage_fraction=payload["coverage_fraction"],
        trim_flank_bp=payload["trim_flank_bp"], edit_dist_ratio=payload["edit_dist_ratio"],
        chimera_split=payload["chimera_split"], well_consensus_at_root=True,
        minimap2_threads=payload["minimap2_threads"], consensus_workers=payload["consensus_workers"],
        per_read_parallel=payload.get("per_read_parallel", False),
        progress_callback=inner_cb,
        # Parsed once in the parent and shipped in the payload (picklable list
        # of (name, prefix) tuples). Absent => this worker parses the xlsx
        # itself, as before.
        barcode_prefixes=payload.get("barcode_prefixes"),
        core_budget=budget,
        # How many of these workers run at once, so each one asks for its own
        # share of the box rather than all of it. Absent (a caller that did not
        # set it) => 1, the pre-existing single-worker sizing.
        mem_workers=payload.get("mem_workers", 1))
    s = result.stats
    return {"nb_name": payload["nb_name"], "sort_barcode_name": payload["sort_barcode_name"],
            "output_dir": str(Path(payload["output_dir"]).resolve()),
            "stats": {k: getattr(s, k) for k in _DEMUX_NB_STAT_KEYS},
            "per_well_read_counts": dict(result.per_well_read_counts)}


def _summary_from_marker(sort_barcode_name: str, nb_out: Path, marker: dict) -> dict:
    """Reconstruct a per-NB summary dict from a completed unit's stage marker.

    Mirrors the picklable dict returned by :func:`_demux_one_nb` so a skipped
    (already-complete) native barcode contributes identically to ``per_nb`` and
    ``merged_stats`` as a freshly-processed one.  ``per_well_read_counts`` is the
    marker's recorded ``per_well_counts``; the 8 DemuxStats counters come from
    the marker's optional ``stats`` block (absent in older/foreign markers, in
    which case those counters seed 0, never a crash).

    ``nb_name`` is left as the sort_barcode name here; the caller overwrites it
    with the real input nb_name before ordering so resume ordering matches.
    """
    marker_stats = marker.get("stats") or {}
    stats = {k: int(marker_stats.get(k, 0)) for k in _DEMUX_NB_STAT_KEYS}
    per_well = {
        str(w): int(c) for w, c in (marker.get("per_well_counts") or {}).items()
    }
    return {
        "nb_name": sort_barcode_name,
        "sort_barcode_name": sort_barcode_name,
        "output_dir": str(nb_out.resolve()),
        "stats": stats,
        "per_well_read_counts": per_well,
    }


def _marker_has_usable_alignment(marker: dict) -> bool:
    """True when a completed unit's marker records reads that survived alignment.

    Reads ``passed_coverage``, the LAST alignment gate, because that is the
    count of reads that actually reached barcode matching; a unit whose reads
    were all dropped is re-run rather than resumed as an empty result.

    This used to read ``passed_mapq`` back when both counters were assigned the
    same post-filter number, so the two were interchangeable. Now that the gates
    are counted separately, ``passed_mapq`` alone would call a unit usable even
    when every read died at the coverage gate. Markers written before the split
    still carry ``passed_mapq == passed_coverage``, so the decision for old
    markers is unchanged.
    """
    stats = marker.get("stats")
    if not isinstance(stats, dict):
        return True
    total_reads = int(stats.get("total_reads", 0))
    passed_coverage = int(stats.get("passed_coverage", 0))
    return total_reads == 0 or passed_coverage > 0


def run_combinatorial_demux_per_nb(
    nb_to_fastq: dict[str, list[Path]],
    reference_fasta: Path,
    barcodes_xlsx: Path,
    output_dir: Path,
    *,
    mapq_threshold: int = 25,
    coverage_fraction: float = 0.98,
    trim_flank_bp: int = 30,
    edit_dist_ratio: float = 0.25,
    chimera_split: bool = True,
    parallel: bool = True,
    max_workers: int | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> dict:
    """Run combinatorial demux per native barcode, in parallel across barcodes.

    Each native barcode in *nb_to_fastq* is demuxed into its own
    ``output_dir/sort_barcode{NN}/`` subdir (well_consensus_at_root layout),
    optionally across worker processes.  Iterates *nb_to_fastq* in dict order
    (insertion order = caller order).

    Returns a dict with ``merged_stats`` (8 stat keys summed across barcodes),
    ``per_nb`` (per-barcode summaries in input order), ``parallel`` (whether the
    parallel path was used), and ``workers`` (process count).
    """
    from kuma_core.mame.ingest.sort_barcode import _nb_to_sort_barcode_name

    # Parent-side measurement window. On the serial path the child
    # run_combinatorial_demux phases land in this same process and therefore
    # show up decomposed; on the ProcessPool path only the parent wall is
    # measured here and each worker emits its own "demux" record.
    _perf_base = TIMER.begin()

    cpu = os.cpu_count() or 4
    n = len(nb_to_fastq)
    env_off = os.environ.get("KUMA_MAME_NB_PARALLEL", "1") == "0"
    use_parallel = parallel and n > 1 and not env_off
    if use_parallel:
        _env_workers = os.environ.get("KUMA_MAME_NB_WORKERS", "").strip()
        if max_workers:
            P = max_workers
        elif _env_workers.isdigit() and int(_env_workers) > 0:
            P = int(_env_workers)
        else:
            P = min(n, cpu)
        P = max(1, min(P, n, cpu))
    else:
        P = 1
    # Work unit is one native barcode, so P is capped by the barcode count and
    # each worker gets cpu // P minimap2 threads. On the reference workload that
    # is 3 processes x 3 threads on a 10-core box.
    #
    # Flattening the unit to (barcode, read_chunk) on a single cpu-wide pool was
    # measured on 2026-08-01 and does NOT pay, so it is deliberately not done.
    # The premise for it was that minimap2 scales sublinearly in -t, making
    # "more processes x fewer threads" win. That premise is false for this
    # workload: on one barcode (13190 reads, ispS 1683 bp, min of 4) minimap2
    # measures 5.973 s at -t 1, 3.120 s at -t 2, 2.250 s at -t 3 and 1.824 s at
    # -t 4, i.e. 3.28x on 4 threads. With no sublinearity to exploit, any split
    # that keeps the same core count is a wash, and a real ProcessPool fan-out
    # over (barcode, chunk) units confirmed it (medians / mins, 4 interleaved
    # rounds): 3px3t 3.379/2.839, 6px2t 2.936/2.833, 9px1t 3.028/2.709,
    # 10px1t 3.005/2.732. Nor is there an idle core to claim: raising the
    # per-worker budget to -t 4 or -t 5 (10 to 15 threads on 10 cores) measured
    # 3.279/2.899 against 3.157 for -t 3, inside the noise. The alignment stage
    # already saturates the box, so the flattened version would only add
    # cross-process pickling of read chunks and a coarser resume unit.
    #
    # What DID pay on the same axis is overlapping the stages inside a worker
    # rather than adding workers: see _READ_CHUNK_DEFAULT and _prefetch.
    #
    # threads_per is the FLOOR, not the whole story: on the parallel path a
    # shared live-worker count lets each survivor raise its minimap2 -t to
    # cpu // live once a sibling unit finishes (see _CoreBudget below). The
    # units are 3.34 : 1 on the real plate, so the static share alone leaves
    # cores idle for most of the run.
    threads_per = max(1, cpu // P)
    # n_nb == 1: this single NB runs in the main process (no per-NB pool), so
    # the per-read matching loop may fan out to its own ProcessPool. With n>1
    # the per-NB pool already owns the cores, so per-read stays serial.
    #
    # Nesting is not the reason. Nesting a ProcessPoolExecutor inside one of its
    # own workers is legal (its workers are created non-daemonic, unlike
    # multiprocessing.Pool workers, which refuse children); an earlier note here
    # claimed otherwise. The reason is throughput: allowing the nested fan-out
    # with n_nb=3 on a 10-core box, each NB worker capped at its cpu//n_nb share,
    # measured 19% slower end to end and was slower in all 5 paired rounds. The
    # nested spawn warm-up plus pickling the per-chunk read payload costs more
    # than the ~0.3 s of per-NB matching work it parallelises. Output stayed
    # byte-identical, so this is purely a cost decision, not a safety one.
    per_read_parallel = n == 1

    payloads: list[dict] = []
    for nb_name, paths in nb_to_fastq.items():
        sort_barcode_name = _nb_to_sort_barcode_name(nb_name)
        payloads.append({
            "nb_name": nb_name,
            "sort_barcode_name": sort_barcode_name,
            "output_dir": str(output_dir / sort_barcode_name),
            "fastq_paths": [str(p) for p in paths],
            "reference_fasta": str(reference_fasta),
            "barcodes_xlsx": str(barcodes_xlsx),
            "mapq_threshold": mapq_threshold,
            "coverage_fraction": coverage_fraction,
            "trim_flank_bp": trim_flank_bp,
            "edit_dist_ratio": edit_dist_ratio,
            "chimera_split": chimera_split,
            "minimap2_threads": threads_per,
            "consensus_workers": threads_per,
            "per_read_parallel": per_read_parallel,
            # Memory-budget divisor. P is the count of these processes that are
            # resident at once, which is what the box actually has to hold; the
            # serial path leaves this at 1 because there is only ever one.
            "mem_workers": P,
        })

    # Fail fast if two entries map to the same sort_barcode output dir
    # (e.g. both "barcode06" and "NB06"), which would silently overwrite.
    _sort_names = [pl["sort_barcode_name"] for pl in payloads]
    if len(set(_sort_names)) != len(_sort_names):
        raise ValueError(
            f"Native barcodes map to colliding sort_barcode output dirs: {_sort_names}"
        )

    # Identity of the inputs that decide this run's per-well output. Recorded in
    # every marker written below and compared before any unit is reused, so a
    # rerun that swaps the reference (or a gate) recomputes instead of reusing
    # consensus called against something else.
    _marker_reference = reference_fingerprint(reference_fasta)
    _marker_params = {
        "mapq_threshold": int(mapq_threshold),
        "coverage_fraction": float(coverage_fraction),
        "trim_flank_bp": int(trim_flank_bp),
        "edit_dist_ratio": float(edit_dist_ratio),
        "chimera_split": bool(chimera_split),
    }

    # ── Resume: which per-NB units are already complete? ─────────────────
    # A unit (one output_dir/sort_barcode{NN}/ dir) is "done" ONLY when it
    # carries a valid completion marker whose recorded inventory matches the
    # consensus FASTA on disk.  Directory existence alone never counts.  Each
    # completed unit's summary is reconstructed from its marker (mirroring the
    # _demux_one_nb dict) so it is NOT re-demuxed/aligned, yet still seeds
    # merged_stats and per_nb identically to a freshly-processed one.
    #
    # Inventory alone is not enough: it says the files are all there, never what
    # they were made from. A unit whose recorded reference/parameters differ from
    # this run is reprocessed, which is the same remedy this loop already applies
    # to a unit with no usable alignment, and never a silent reuse.
    completed_summaries: dict[str, dict] = {}  # nb_name -> summary dict
    for pl in payloads:
        nb_out = output_dir / pl["sort_barcode_name"]
        if is_unit_complete(nb_out):
            marker = read_stage_marker(nb_out)
            if marker is None:
                continue
            inputs_ok, inputs_reason = marker_inputs_match(
                marker, _marker_reference, _marker_params
            )
            if not inputs_ok:
                log.info(
                    "Reprocessing %s: %s", pl["sort_barcode_name"], inputs_reason
                )
                continue
            if _marker_has_usable_alignment(marker):
                summ = _summary_from_marker(pl["sort_barcode_name"], nb_out, marker)
                summ["nb_name"] = pl["nb_name"]  # real input nb_name for ordering
                completed_summaries[pl["nb_name"]] = summ

    # Only dispatch payloads for units that are NOT already complete.
    pending = [pl for pl in payloads if pl["nb_name"] not in completed_summaries]

    def _commit_marker(summ: dict) -> None:
        """Write the unit's completion marker LAST (atomic commit point).

        Called only after the unit's consensus FASTA are all on disk (worker
        returned).  Records per_well_counts (= read counts), the 8 DemuxStats
        counters under ``stats`` so a full resume can reseed merged_stats, and
        consensus=True since the per-NB path always runs consensus.  A failure
        to write the marker must not lose the completed work, but here a write
        failure is unexpected (atomic temp+replace) and should surface, so it is
        not swallowed.
        """
        nb_out = output_dir / summ["sort_barcode_name"]
        # The worker (run_combinatorial_demux) already created nb_out; mkdir here
        # is an idempotent guard so the marker write never fails on a missing
        # parent.  An empty/interrupted unit whose consensus FASTA never landed
        # stays "not complete" (validate_marker inventory mismatch), so this does
        # not falsely mark an empty dir done.
        nb_out.mkdir(parents=True, exist_ok=True)
        write_stage_marker(
            nb_out,
            per_well_counts={
                str(w): int(c) for w, c in summ["per_well_read_counts"].items()
            },
            consensus=True,
            stats={k: int(summ["stats"][k]) for k in _DEMUX_NB_STAT_KEYS},
            reference=_marker_reference,
            params=_marker_params,
        )

    summaries: list[dict] = list(completed_summaries.values())

    # ── Smooth aggregate progress across barcodes ─────────────────────────
    # Each pending NB contributes a 0..1 fraction (streamed from the worker via
    # progress_q); resume-skipped units count as 1.0. The bar reported to the
    # caller = (completed + sum(in-flight fractions)) / n as parts-per-1000, so
    # the demux phase advances continuously instead of only at NB boundaries.
    nb_frac: dict[str, float] = {pl["nb_name"]: 0.0 for pl in pending}
    n_seed_done = len(completed_summaries)
    _agg_lock = threading.Lock()
    _agg_last = [-1.0]

    def _emit_agg(force: bool = False) -> None:
        if progress_callback is None:
            return
        with _agg_lock:
            agg = (n_seed_done + sum(nb_frac.values())) / n if n else 1.0
            agg = max(0.0, min(1.0, agg))
            if not force and abs(agg - _agg_last[0]) < 0.003:
                return
            _agg_last[0] = agg
            done_ct = n_seed_done + sum(1 for f in nb_frac.values() if f >= 1.0)
        progress_callback(int(round(agg * 1000)), 1000, f"{done_ct}/{n} barcodes")

    def _note_frac(nb_name: str, frac: float) -> None:
        with _agg_lock:
            if nb_name in nb_frac and frac > nb_frac[nb_name]:
                nb_frac[nb_name] = frac
        _emit_agg()

    _emit_agg(force=True)  # tick past resume-skipped units immediately

    if P > 1 and pending:
        # Pick the start method first and, on forkserver, let its helper boot in
        # the background so its preload import overlaps the barcode parse below
        # instead of delaying the first worker. No-op under spawn.
        mp_ctx = _demux_mp_context()
        _warm_mp_context(mp_ctx)
        # Parse the barcode workbook once here and ship the result in each
        # payload, instead of every spawned worker re-parsing it. The dominant
        # cost is not the parse (~0.01 s) but the first `import openpyxl`
        # (~1.4 s) inside load_barcode_prefixes, which a worker now skips
        # entirely. Done only on this branch: the serial path below runs the
        # unit in this same process, where openpyxl is already imported, so
        # there is nothing to save and the workbook stays unread when a caller
        # stubs the worker out.
        with TIMER.phase("load_barcodes_parent"):
            _barcode_prefixes = load_barcode_prefixes(barcodes_xlsx)
        for pl in pending:
            pl["barcode_prefixes"] = _barcode_prefixes
        manager = None
        progress_q = None
        try:
            manager = mp_ctx.Manager()
            progress_q = manager.Queue()
        except Exception:  # Manager unavailable — degrade to per-NB completion only
            manager = None
            progress_q = None
        if progress_q is not None:
            for pl in pending:
                pl["progress_queue"] = progress_q
        # Dynamic core budget: a shared count of workers still running, so a
        # worker whose siblings have finished raises its minimap2 -t to the
        # freed cores instead of holding cpu//P for the whole run. Sizes on the
        # real plate are 3.34 : 1, so under the static share the two small
        # workers finish at 80 s and 179 s of a 251 s run.
        live_val = None
        if os.environ.get(
            "KUMA_MAME_CORE_BUDGET", str(_CORE_BUDGET_DEFAULT)
        ).strip() != "0" and manager is not None:
            try:
                live_val = manager.Value("i", len(pending))
            except Exception:  # noqa: BLE001 - proxy unavailable: static share
                live_val = None
        if live_val is not None:
            for pl in pending:
                pl["live_workers"] = live_val
                pl["cpu_total"] = cpu
        _drain_stop = threading.Event()

        def _drainer(progress_queue) -> None:
            while not _drain_stop.is_set():
                try:
                    nb_name, frac = progress_queue.get(timeout=0.4)
                except Exception:
                    continue
                _note_frac(nb_name, frac)

        _drain_thread = None
        if progress_q is not None:
            _drain_thread = threading.Thread(
                target=_drainer, args=(progress_q,), daemon=True
            )
            _drain_thread.start()
        try:
            with ProcessPoolExecutor(max_workers=P, mp_context=mp_ctx) as ex:
                futs = {ex.submit(_demux_one_nb, pl): pl["nb_name"] for pl in pending}
                for fut in as_completed(futs):
                    summ = fut.result()  # propagate worker exceptions (fail-fast)
                    if live_val is not None:
                        # Publish the freed cores before anything else: the
                        # survivors pick them up on their next minimap2 call.
                        # Only the parent writes, so read-modify-write here
                        # needs no lock beyond as_completed's single thread.
                        try:
                            live_val.value = max(1, int(live_val.value) - 1)
                        except Exception:  # noqa: BLE001 - proxy gone, static share
                            pass
                    _commit_marker(summ)  # commit point: unit files all on disk now
                    with _agg_lock:
                        nb_frac[summ["nb_name"]] = 1.0
                    _emit_agg(force=True)
                    summaries.append(summ)
        finally:
            _drain_stop.set()
            if _drain_thread is not None:
                _drain_thread.join(timeout=1.5)
            if manager is not None:
                try:
                    manager.shutdown()
                except Exception:
                    pass
    else:
        for pl in pending:
            # Serial path: forward inner progress in-process via a queue-shaped
            # shim so _demux_one_nb uses one code path.
            pl["progress_queue"] = _DirectProgressSink(_note_frac)
            summ = _demux_one_nb(pl)
            _commit_marker(summ)  # commit point: unit files all on disk now
            summaries.append(summ)
            with _agg_lock:
                nb_frac[pl["nb_name"]] = 1.0
            _emit_agg(force=True)

    # Order summaries by input nb order.
    by_name = {s["nb_name"]: s for s in summaries}
    ordered_summaries = [by_name[pl["nb_name"]] for pl in payloads]

    # Merge stats: sum each of the 8 stat keys across summaries (processed +
    # resume-seeded units alike).
    merged = {
        k: sum(s["stats"][k] for s in ordered_summaries) for k in _DEMUX_NB_STAT_KEYS
    }

    TIMER.end(
        "demux_per_nb", _perf_base, workers=P, barcodes=n, parallel=P > 1,
    )

    return {"merged_stats": merged, "per_nb": ordered_summaries,
            "parallel": P > 1, "workers": P}


__all__ = [
    # Public types
    "DemuxResult",
    "DemuxStats",
    # Public entry points
    "load_barcodes",
    "load_barcode_prefixes",
    "run_combinatorial_demux",
    "run_combinatorial_demux_per_nb",
    "_demux_one_nb",
    # Semi-private helpers exported for tests and diagnostic scripts
    "_extract_barcode_prefix",
    "_extract_f_prefix",
    "_extract_r_prefix",
    "_find_best_barcode",
    "_demux_read_anchored",
    "_demux_read",
]
