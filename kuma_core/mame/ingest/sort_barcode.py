"""sort_barcode -- Combinatorial barcode xlsx parsers for 96-well plates.

This module provides **xlsx parsing utilities** used by the barcode setup
pipeline.  The combinatorial read-sorting algorithm (sliding-window + edlib
fuzzy matching) has been removed in PR-B; the canonical pipeline is now
``kuma_core.mame.ingest.combinatorial_demux.run_combinatorial_demux``.

Remaining public API
--------------------
parse_combinatorial_barcodes
    Read isps_f_* / isps_r_* barcodes from xlsx, expand to 96-well map.
parse_sample_map
    Read well_id -> sample_name mapping from a legacy sample-map xlsx. Kept for
    migration only; nothing places wells from it any more.
_nb_to_sort_barcode_name
    Convert native-barcode dir basename to sort_barcode dir name.

These utilities are also used by ``kuma_core.mame.ingest.barcode_package``
to validate generated xlsx files in tests.
"""

from __future__ import annotations

import re
from pathlib import Path

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_ROW_LETTERS = list("ABCDEFGH")   # rows 1-8 -> A-H
_N_COLS = 12
_N_ROWS = 8

_NB_DIR_PATTERN = re.compile(r"^(?:barcode|NB)(\d{1,3})$", re.IGNORECASE)

# Barcode row names follow ``<gene_prefix>_f_<int>`` / ``<gene_prefix>_r_<int>``
# where ``<gene_prefix>`` is derived from the user-provided gene_name
# (see ``barcode_package._sanitize_gene_prefix``). The reader is gene-agnostic
# and only requires the trailing ``_f_<n>`` / ``_r_<n>`` marker. Backward
# compatible with legacy ``isps_f_*`` / ``isps_r_*`` xlsx files.
_FWD_ROW_RE = re.compile(r"^(?P<prefix>.+?)_f_(?P<n>\d+)$")
_REV_ROW_RE = re.compile(r"^(?P<prefix>.+?)_r_(?P<n>\d+)$")


# ---------------------------------------------------------------------------
# NB dir -> sort_barcode name
# ---------------------------------------------------------------------------


def _nb_to_sort_barcode_name(nb_basename: str) -> str:
    """Map a native-barcode dir name to a sort_barcode output dir name.

    Examples
    --------
    ``barcode06``  ->  ``sort_barcode06``
    ``NB06``       ->  ``sort_barcode06``
    ``barcode100`` ->  ``sort_barcode100``

    Raises
    ------
    ValueError
        If ``nb_basename`` does not match the expected pattern.
    """
    m = _NB_DIR_PATTERN.match(nb_basename)
    if m is None:
        raise ValueError(
            f"Cannot convert {nb_basename!r} to sort_barcode name: "
            "expected 'barcodeN' or 'NBN' (1-3 digit suffix)"
        )
    n = int(m.group(1))
    # 2-digit zero-pad for n < 100; 3-digit (no pad needed) for n >= 100.
    padded = f"{n:02d}" if n < 100 else str(n)
    return f"sort_barcode{padded}"


# ---------------------------------------------------------------------------
# Sample map parser
# ---------------------------------------------------------------------------


def parse_sample_map(path: Path) -> dict[str, str]:
    """Parse a legacy sample-map xlsx into a well_id -> sample dict.

    MIGRATION READER. Nothing places wells from this file any more: the run
    layout is computed from the variant list, and the sample map was the second
    statement of the same plate that nobody kept in step with the first. It is
    read only so ``validate_inputs`` can compare an old project's file against
    the computed draft and name the wells where the two disagree, which is a
    better answer than deleting the file or ignoring it. Expected to be removed
    a release after the projects that carry one have been opened once.

    File format (Sheet1)
    --------------------
    Column A: sample name  (e.g. ``V5F``, ``K53R``, ``WT``)
    Column B: well position in plate notation (e.g. ``A1``, ``H12``)

    Well positions are normalised to zero-padded format (``A1`` -> ``A01``).
    Rows with missing or malformed well positions are silently skipped, which is
    safe here for the reason it was not safe in the variant list: this file no
    longer decides where anything goes.

    Returns
    -------
    ``dict[str, str]``: {well_id: sample_name}  e.g. ``{"A01": "V5F", "H12": "WT"}``

    Raises
    ------
    FileNotFoundError
        If ``path`` does not exist.
    """
    if not path.exists():
        raise FileNotFoundError(f"sample_map_path not found: {path}")

    import openpyxl  # local import: keeps cold-start fast

    from kuma_core.mame.plate_geometry import PLATE_COLS, PLATE_ROWS, norm_well

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        result: dict[str, str] = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            sample_raw = str(row[0]).strip()
            if not sample_raw:
                continue
            well_raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            if not well_raw:
                continue
            well_id = norm_well(well_raw)
            if len(well_id) != 3 or not well_id[1:].isdigit():
                continue
            if not ("A" <= well_id[0] < chr(ord("A") + PLATE_ROWS)):
                continue
            if not 1 <= int(well_id[1:]) <= PLATE_COLS:
                continue
            if well_id not in result:
                result[well_id] = sample_raw
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Barcode xlsx parser
# ---------------------------------------------------------------------------


def parse_combinatorial_barcodes(path: Path) -> dict[str, tuple[str, str]]:
    """Parse ``<gene>_f_*`` / ``<gene>_r_*`` barcodes from xlsx and expand to 96 wells.

    File format (Sheet1)
    --------------------
    Column A: barcode name  (e.g. ``isps_f_1``, ``mygene_r_3``)
    Column B: sequence (ACGT, any length >= 5)

    The ``<gene>`` prefix is gene-agnostic -- any string ending in
    ``_f_<int>`` / ``_r_<int>`` is accepted. Legacy ``isps_f_*`` / ``isps_r_*``
    files remain compatible.

    Expansion rule
    --------------
    - ``<gene>_f_N`` (N=1..12): forward barcode, used for column index.
    - ``<gene>_r_N`` (N=1..8):  reverse barcode, used for row index.
    - ``well_id = f"{ROW_LETTER[r-1]}{c:02d}"`` where r = rev index, c = fwd index.
    - Example: fwd=1, rev=1 -> A01; fwd=12, rev=8 -> H12.

    Returns
    -------
    ``dict[str, tuple[str, str]]``: {well_id: (fwd_seq, rev_seq)} for all 96 wells.

    Raises
    ------
    FileNotFoundError
        If ``path`` does not exist.
    ValueError
        If not all 12 fwd or not all 8 rev barcodes are present, or if any
        barcode name prefix is unrecognised.
    """
    if not path.exists():
        raise FileNotFoundError(f"custom_barcode_xlsx not found: {path}")

    import openpyxl  # local import: keeps cold-start fast

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        fwd_map: dict[int, str] = {}   # {N: sequence}
        rev_map: dict[int, str] = {}   # {N: sequence}

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            name_raw = str(row[0]).strip()
            seq_raw = str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ""
            if not seq_raw or len(seq_raw) < 5:
                continue

            fwd_match = _FWD_ROW_RE.match(name_raw)
            if fwd_match is not None:
                n = int(fwd_match.group("n"))
                fwd_map[n] = seq_raw
                continue

            rev_match = _REV_ROW_RE.match(name_raw)
            if rev_match is not None:
                n = int(rev_match.group("n"))
                rev_map[n] = seq_raw
                continue

            # Rows that do not match either pattern are silently skipped
            # (header rows, comments, blank names).
    finally:
        wb.close()

    # Validate completeness.
    missing_fwd = [i for i in range(1, _N_COLS + 1) if i not in fwd_map]
    if missing_fwd:
        raise ValueError(
            f"Missing forward barcodes (expected <gene>_f_1 ... <gene>_f_{_N_COLS}): "
            f"indices {missing_fwd}"
        )
    missing_rev = [i for i in range(1, _N_ROWS + 1) if i not in rev_map]
    if missing_rev:
        raise ValueError(
            f"Missing reverse barcodes (expected <gene>_r_1 ... <gene>_r_{_N_ROWS}): "
            f"indices {missing_rev}"
        )

    # Build 96-well map.
    well_map: dict[str, tuple[str, str]] = {}
    for r in range(1, _N_ROWS + 1):       # rev idx -> row letter
        for c in range(1, _N_COLS + 1):   # fwd idx -> column number
            well_id = f"{_ROW_LETTERS[r - 1]}{c:02d}"
            well_map[well_id] = (fwd_map[c], rev_map[r])

    return well_map


__all__ = [
    "parse_combinatorial_barcodes",
    # Migration reader only. The sample map no longer places wells; this stays
    # one release so an existing project's file can be compared against the
    # computed draft and its disagreements named, then goes.
    "parse_sample_map",
    "_nb_to_sort_barcode_name",
]
