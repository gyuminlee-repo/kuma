"""barcode_package -- Generate MAME input package from barcode seeds and a CDS FASTA.

This module produces three output files and a context JSON:
- ``barcodes_sequence.xlsx``    : 20-row combinatorial barcode table (12 fwd + 8 rev)
- ``{sanitized_gene_name}_amplicon.fa`` : single-entry FASTA for the target gene region
- ``sample_map_template.xlsx``  : well-map template for the MAME operator (blank
  headers, or pre-filled with a draft placement when ``expected_mutations_path``
  is supplied)
- ``mame_context.json``         : machine-readable pointer file (schema 1)

Typical call site::

    result = generate_mame_package(
        fasta_path=Path("seq/target.fa"),
        gene_start=400,
        gene_end=700,
        barcode_seeds_path=Path("design/barcode_seeds.xlsx"),
        output_dir=Path("project/design"),
        project_root=Path("project"),
        gene_name="target_gene",
        polymerase="Q5",
    )
"""

from __future__ import annotations

import datetime
import json
import warnings
from dataclasses import dataclass, field
from pathlib import Path

import primer3

from kuma_core.mame.ingest.polymerase import PolymeraseProfile, get_profile

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_N_FWD = 12
_N_REV = 8
_MIN_SEED_LEN = 5
_MAX_SEED_LEN = 30
_VALID_BASES = frozenset("ATGC")

# Prefix convention for barcode_seeds.xlsx rows
_FWD_PREFIX = "fwd_"
_REV_PREFIX = "rev_"

# Suffix convention for the output barcodes_sequence.xlsx (MAME side).
# Row names are derived as ``{sanitized_gene_name}{_FWD_SUFFIX}{i}`` /
# ``{sanitized_gene_name}{_REV_SUFFIX}{i}``. The matching reader in
# ``sort_barcode.parse_combinatorial_barcodes`` is gene-agnostic — it parses
# any prefix that ends with ``_f_<int>`` / ``_r_<int>``.
_FWD_SUFFIX = "_f_"
_REV_SUFFIX = "_r_"


def _sanitize_gene_prefix(gene_name: str) -> str:
    """Return a filesystem- and excel-safe prefix derived from ``gene_name``.

    - Lower-cases.
    - Replaces any run of non-alphanumeric chars with a single ``_``.
    - Strips leading/trailing ``_``.
    - Strips embedded ``_f_`` / ``_r_`` substrings so the suffix on row names
      stays unambiguous for the reader's regex parser.

    Raises
    ------
    ValueError
        If the sanitized result is empty (pathological input such as ``""``
        or all-punctuation).
    """
    import re as _re
    s = _re.sub(r"[^a-z0-9]+", "_", gene_name.lower()).strip("_")
    # Strip any embedded _f_ / _r_ runs so the suffix on row names remains
    # the only such marker in the final string.
    s = s.replace("_f_", "_").replace("_r_", "_").strip("_")
    if not s:
        raise ValueError(
            f"gene_name {gene_name!r} sanitizes to an empty prefix; "
            "provide a gene name containing at least one alphanumeric character."
        )
    return s


# ---------------------------------------------------------------------------
# Reverse complement (local to avoid cross-layer coupling with kuro.overlap)
# ---------------------------------------------------------------------------

def _reverse_complement(seq: str) -> str:
    """Return the reverse complement of a DNA sequence (case-preserving)."""
    table = str.maketrans("ACGTacgt", "TGCAtgca")
    return seq.translate(table)[::-1]


# ---------------------------------------------------------------------------
# Tm calculation
# ---------------------------------------------------------------------------

def _calc_tm(seq: str, profile: PolymeraseProfile) -> float:
    """Calculate Tm using the given polymerase salt profile.

    Uses primer3.calc_tm with SantaLucia 1998 nearest-neighbour parameters.
    ``seq`` is converted to uppercase before passing to primer3.
    """
    return primer3.calc_tm(
        seq.upper(),
        mv_conc=profile.mv_conc,
        dv_conc=profile.dv_conc,
        dntp_conc=profile.dntp_conc,
        dna_conc=profile.dna_conc,
        tm_method=profile.tm_method,
        salt_corrections_method=profile.salt_corrections_method,
    )


# ---------------------------------------------------------------------------
# Sequence parser (FASTA / GenBank / SnapGene)
# ---------------------------------------------------------------------------

_GENBANK_SUFFIXES = {".gb", ".gbk", ".gbff"}
_SNAPGENE_SUFFIXES = {".dna"}
_FASTA_SUFFIXES = {".fa", ".fasta", ".fna"}


def _parse_first_cds_sequence(seq_path: Path) -> tuple[str, str]:
    """Return (sequence, topology) for the first record in a FASTA, GenBank, or
    SnapGene file.

    GenBank/SnapGene are routed to Biopython via kuro's ``load_sequence`` for the
    sequence, and via kuro's ``detect_topology`` for the topology annotation
    ("circular" or "linear"). FASTA uses a lightweight inline parser to keep
    the dependency surface small and has no topology annotation, so it is
    always reported as "linear".

    Raises
    ------
    FileNotFoundError
        If ``seq_path`` does not exist.
    ValueError
        If no record is found or the resulting sequence is empty.
    """
    if not seq_path.exists():
        raise FileNotFoundError(f"Sequence file not found: {seq_path}")

    suffix = seq_path.suffix.lower()

    if suffix in _GENBANK_SUFFIXES or suffix in _SNAPGENE_SUFFIXES:
        from kuma_core.kuro.sdm_engine import detect_topology, load_sequence

        _header, sequence, _genes = load_sequence(seq_path)
        if not sequence:
            raise ValueError(f"Empty sequence in: {seq_path}")
        topology = detect_topology(seq_path)
        return sequence.upper(), topology

    if suffix not in _FASTA_SUFFIXES:
        raise ValueError(
            f"Unsupported sequence file extension {suffix!r}; "
            "use .fa/.fasta/.fna, .gb/.gbk/.gbff, or .dna."
        )

    return _parse_first_fasta_sequence(seq_path), "linear"


def _parse_first_fasta_sequence(fasta_path: Path) -> str:
    """Return the sequence of the first record in a FASTA file.

    Issues a UserWarning when more than one record is present (first is used).

    Raises
    ------
    FileNotFoundError
        If ``fasta_path`` does not exist.
    ValueError
        If the file contains no valid FASTA record or the sequence is empty.
    """
    if not fasta_path.exists():
        raise FileNotFoundError(f"FASTA file not found: {fasta_path}")

    seq_parts: list[str] = []
    header_count = 0
    with fasta_path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.rstrip("\r\n")
            if line.startswith(">"):
                header_count += 1
                if header_count > 1:
                    break   # stop reading after second header
            elif header_count == 1 and line:
                seq_parts.append(line.strip())

    if header_count == 0:
        raise ValueError(f"No FASTA record found in: {fasta_path}")
    if header_count > 1:
        warnings.warn(
            f"{fasta_path} contains {header_count} sequences; using the first one.",
            UserWarning,
            stacklevel=3,
        )

    seq = "".join(seq_parts)
    if not seq:
        raise ValueError(f"FASTA record in {fasta_path} has an empty sequence.")
    return seq


# ---------------------------------------------------------------------------
# Core functions
# ---------------------------------------------------------------------------

def parse_barcode_seeds(path: Path) -> dict[str, str]:
    """Parse fwd/rev seed sequences from an xlsx file.

    File format (Sheet1):
        Column A: name  (``fwd_1`` .. ``fwd_12``, ``rev_1`` .. ``rev_8``)
        Column B: sequence (A/T/G/C, case-insensitive, 5-30 bp)

    Returns
    -------
    dict mapping ``"fwd_1"`` .. ``"fwd_12"`` and ``"rev_1"`` .. ``"rev_8"``
    to their uppercase seed sequences.

    Raises
    ------
    FileNotFoundError
        If ``path`` does not exist.
    ValueError
        If fwd count != 12, rev count != 8, any sequence contains non-ATGC
        characters, any sequence is outside the 5-30 bp range, or duplicate
        sequences are found among all 20 seeds.
    """
    if not path.exists():
        raise FileNotFoundError(f"barcode_seeds file not found: {path}")

    import openpyxl  # local import: keeps cold-start fast

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        fwd_map: dict[int, str] = {}
        rev_map: dict[int, str] = {}

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            name_raw = str(row[0]).strip().lower()
            seq_raw = str(row[1]).strip().upper() if len(row) > 1 and row[1] is not None else ""
            if not seq_raw:
                continue

            if name_raw.startswith(_FWD_PREFIX):
                try:
                    n = int(name_raw[len(_FWD_PREFIX):])
                except ValueError:
                    raise ValueError(
                        f"Cannot parse fwd barcode index from {name_raw!r}. "
                        f"Expected format: {_FWD_PREFIX}<integer>"
                    )
                fwd_map[n] = seq_raw
            elif name_raw.startswith(_REV_PREFIX):
                try:
                    n = int(name_raw[len(_REV_PREFIX):])
                except ValueError:
                    raise ValueError(
                        f"Cannot parse rev barcode index from {name_raw!r}. "
                        f"Expected format: {_REV_PREFIX}<integer>"
                    )
                rev_map[n] = seq_raw
            # rows that match neither prefix (header, comment) are silently skipped
    finally:
        wb.close()

    # Completeness check
    missing_fwd = [i for i in range(1, _N_FWD + 1) if i not in fwd_map]
    if missing_fwd:
        raise ValueError(
            f"Missing forward barcode seeds (expected fwd_1 .. fwd_{_N_FWD}): "
            f"indices {missing_fwd}"
        )
    missing_rev = [i for i in range(1, _N_REV + 1) if i not in rev_map]
    if missing_rev:
        raise ValueError(
            f"Missing reverse barcode seeds (expected rev_1 .. rev_{_N_REV}): "
            f"indices {missing_rev}"
        )

    # Per-seed validation (bases and length)
    all_seeds: dict[str, str] = {}
    for i in range(1, _N_FWD + 1):
        key = f"fwd_{i}"
        seq = fwd_map[i]
        _validate_seed_sequence(key, seq)
        all_seeds[key] = seq
    for i in range(1, _N_REV + 1):
        key = f"rev_{i}"
        seq = rev_map[i]
        _validate_seed_sequence(key, seq)
        all_seeds[key] = seq

    # Duplicate check across all 20 seeds
    seen_seqs: dict[str, str] = {}  # seq -> first key
    for key, seq in all_seeds.items():
        if seq in seen_seqs:
            raise ValueError(
                f"Duplicate seed sequence detected: {key!r} and "
                f"{seen_seqs[seq]!r} share the same sequence ({seq!r}). "
                "All 20 barcode seeds must be unique."
            )
        seen_seqs[seq] = key

    return all_seeds


def _validate_seed_sequence(key: str, seq: str) -> None:
    """Raise ValueError if seq contains non-ATGC chars or is out of length range."""
    if not (_MIN_SEED_LEN <= len(seq) <= _MAX_SEED_LEN):
        raise ValueError(
            f"Barcode seed {key!r} has length {len(seq)}, "
            f"expected {_MIN_SEED_LEN}-{_MAX_SEED_LEN} bp."
        )
    invalid = set(seq) - _VALID_BASES
    if invalid:
        raise ValueError(
            f"Barcode seed {key!r} contains non-ATGC characters: "
            f"{sorted(invalid)!r} (only A, T, G, C allowed)."
        )


def _circular_slice(seq: str, start: int, length: int, seq_len: int) -> str:
    """Return a length-character substring of seq starting at start, wrapping
    around the origin (position 0) if start is negative or start + length
    exceeds seq_len.

    Intended only for circular topology, where a primer binding site may
    physically span the origin of the template. start may be any integer
    (negative offsets count backwards from the end via Python's modulo).
    """
    return "".join(seq[(start + i) % seq_len] for i in range(length))


def design_flanking_primers(
    cds_sequence: str,
    gene_start: int,
    gene_end: int,
    profile: PolymeraseProfile,
    flank_min: int = 100,
    flank_max: int = 400,
    binding_min_len: int = 18,
    binding_max_len: int = 35,
    tm_min: float = 55.0,
    tm_max: float = 68.0,
    require_gc_clamp: bool = True,
    topology: str = "linear",
) -> tuple[str, str, list[str]]:
    """Design Tm-guided flanking primers flanking a gene region.

    Search strategy
    ---------------
    For the forward primer, the binding site is sought in the region
    ``[gene_start - flank_max, gene_start - flank_min)``.
    Outer loop: start position ascending from ``gene_start - flank_max``.
    Inner loop: binding length ascending from ``binding_min_len`` to ``binding_max_len``.
    First candidate satisfying both the Tm window and the GC-clamp (if requested)
    is returned immediately.
    If no candidate meets the criteria, the candidate whose Tm is closest to
    ``(tm_min + tm_max) / 2`` (Tm midpoint) is returned instead, and a warning
    is appended.

    For the reverse primer, the binding site is sought in
    ``[gene_end + flank_min, gene_end + flank_max)``.
    The candidate sequence is ``reverse_complement(cds_sequence[end - length : end])``.
    Same selection logic applies.

    Parameters
    ----------
    cds_sequence:
        Full CDS nucleotide string (any case).
    gene_start:
        0-based inclusive start position of the gene within ``cds_sequence``.
    gene_end:
        0-based exclusive end position of the gene within ``cds_sequence``.
    profile:
        PolymeraseProfile supplying salt concentrations for Tm calculation.
    flank_min:
        Minimum distance (bp) upstream/downstream of the gene boundary where the
        primer binding site must end/start.
    flank_max:
        Maximum distance (bp) upstream/downstream of the gene boundary that is
        searched for a binding site.
    binding_min_len:
        Minimum primer binding length to try.
    binding_max_len:
        Maximum primer binding length to try.
    tm_min:
        Lower bound of the acceptable Tm window (deg C).
    tm_max:
        Upper bound of the acceptable Tm window (deg C).
    require_gc_clamp:
        If True, the 3' terminal base of every candidate must be G or C.
    topology:
        Either "linear" (default) or "circular". When "linear", a search
        window that falls outside ``cds_sequence`` boundaries raises
        ValueError (unchanged behaviour). When "circular", the forward and
        reverse search windows are allowed to wrap around the sequence
        origin, since the corresponding template region physically exists on
        a circular molecule.

    Returns
    -------
    (fwd_flanking, rev_flanking, warnings) where both sequences are lowercase
    strings and warnings is a (possibly empty) list of human-readable messages.

    Raises
    ------
    ValueError
        If topology is not "linear" or "circular", if the flank search window
        falls outside ``cds_sequence`` boundaries under linear topology, if
        wrapping under circular topology would require reading past a full
        revolution of the sequence, or if ``gene_start >= gene_end``, or if
        parameter ranges are invalid.
    """
    seq_len = len(cds_sequence)

    if topology not in ("linear", "circular"):
        raise ValueError(
            f"topology must be \"linear\" or \"circular\", got {topology!r}."
        )

    if gene_start < 0:
        raise ValueError(f"gene_start must be >= 0, got {gene_start}.")
    if gene_end > seq_len:
        raise ValueError(
            f"gene_end ({gene_end}) exceeds sequence length ({seq_len})."
        )
    if gene_start >= gene_end:
        raise ValueError(
            f"gene_start ({gene_start}) must be < gene_end ({gene_end})."
        )
    if flank_min < 0 or flank_max <= flank_min:
        raise ValueError(
            f"flank_min ({flank_min}) must be >= 0 and < flank_max ({flank_max})."
        )
    if binding_min_len < 1 or binding_max_len < binding_min_len:
        raise ValueError(
            f"binding_min_len ({binding_min_len}) must be >= 1 and "
            f"<= binding_max_len ({binding_max_len})."
        )

    if topology == "circular" and (
        (flank_max - flank_min) > seq_len or binding_max_len > seq_len
    ):
        raise ValueError(
            f"Circular wrap search window (flank_max - flank_min = "
            f"{flank_max - flank_min}) or binding_max_len ({binding_max_len}) "
            f"exceeds the sequence length (seq_len={seq_len}); wrapping would "
            "read the same base more than once. Reduce flank_max/binding_max_len "
            "or use a longer template."
        )

    # Forward primer search window: positions [fwd_window_start, fwd_window_end)
    # The primer starts at `pos` and extends binding_len bases to the right.
    # The primer must end no later than gene_start - flank_min,
    # so pos + length <= gene_start - flank_min  =>  pos <= gene_start - flank_min - length.
    # The primer starts no earlier than gene_start - flank_max.
    fwd_region_start = gene_start - flank_max
    fwd_region_end = gene_start - flank_min  # exclusive upper bound for pos

    if topology == "linear" and fwd_region_start < 0:
        raise ValueError(
            f"Forward primer search window starts at {fwd_region_start} "
            f"(gene_start={gene_start}, flank_max={flank_max}); "
            "sequence is too short upstream of the gene."
        )
    if fwd_region_end <= fwd_region_start:
        raise ValueError(
            f"Forward primer search window [{fwd_region_start}, {fwd_region_end}) "
            "is empty. Increase the gap between gene_start and flank_min/flank_max."
        )

    # Reverse primer search window: binding ends at `end`, starts at `end - length`.
    # end must satisfy gene_end + flank_min <= end <= gene_end + flank_max.
    rev_region_start = gene_end + flank_min  # inclusive lower bound for `end`
    rev_region_end = gene_end + flank_max    # inclusive upper bound for `end`

    if topology == "linear" and rev_region_end > seq_len:
        raise ValueError(
            f"Reverse primer search window ends at {rev_region_end} "
            f"(gene_end={gene_end}, flank_max={flank_max}); "
            "sequence is too short downstream of the gene."
        )
    if rev_region_start > rev_region_end:
        raise ValueError(
            f"Reverse primer search window [{rev_region_start}, {rev_region_end}] "
            "is empty. Increase the gap between gene_end and flank_min/flank_max."
        )

    collected_warnings: list[str] = []
    tm_target = (tm_min + tm_max) / 2.0

    # --- Forward primer -------------------------------------------------------
    fwd_candidates: list[tuple[float, str]] = []  # (abs(Tm - target), seq)
    fwd_chosen: str | None = None

    for pos in range(fwd_region_start, fwd_region_end):
        for length in range(binding_min_len, binding_max_len + 1):
            if topology == "circular":
                candidate = _circular_slice(cds_sequence, pos, length, seq_len)
            else:
                candidate = cds_sequence[pos: pos + length]
                if len(candidate) < length:
                    break  # hit end of sequence
            tm = _calc_tm(candidate, profile)
            gc_ok = (not require_gc_clamp) or (candidate[-1].upper() in "GC")
            fwd_candidates.append((abs(tm - tm_target), candidate))
            if tm_min <= tm <= tm_max and gc_ok:
                fwd_chosen = candidate
                break
        if fwd_chosen is not None:
            break

    if fwd_chosen is None:
        if fwd_candidates:
            fwd_candidates.sort(key=lambda x: x[0])
            fwd_chosen = fwd_candidates[0][1]
            best_tm = _calc_tm(fwd_chosen, profile)
            collected_warnings.append(
                f"No forward primer candidate met Tm [{tm_min}, {tm_max}] "
                f"(require_gc_clamp={require_gc_clamp}). "
                f"Using closest candidate (Tm={best_tm:.1f} C): {fwd_chosen.upper()}"
            )
        else:
            raise ValueError(
                "Forward primer search produced no candidates. "
                f"Check flank_min={flank_min}, flank_max={flank_max}, "
                f"binding_min_len={binding_min_len}, binding_max_len={binding_max_len}."
            )

    # --- Reverse primer -------------------------------------------------------
    rev_candidates: list[tuple[float, str]] = []
    rev_chosen: str | None = None

    for end in range(rev_region_start, rev_region_end + 1):
        for length in range(binding_min_len, binding_max_len + 1):
            start = end - length
            if topology == "circular":
                candidate_raw = _circular_slice(cds_sequence, start, length, seq_len)
            else:
                if start < 0:
                    break
                candidate_raw = cds_sequence[start:end]
                if len(candidate_raw) < length:
                    break
            candidate = _reverse_complement(candidate_raw)
            tm = _calc_tm(candidate, profile)
            gc_ok = (not require_gc_clamp) or (candidate[-1].upper() in "GC")
            rev_candidates.append((abs(tm - tm_target), candidate))
            if tm_min <= tm <= tm_max and gc_ok:
                rev_chosen = candidate
                break
        if rev_chosen is not None:
            break

    if rev_chosen is None:
        if rev_candidates:
            rev_candidates.sort(key=lambda x: x[0])
            rev_chosen = rev_candidates[0][1]
            best_tm = _calc_tm(rev_chosen, profile)
            collected_warnings.append(
                f"No reverse primer candidate met Tm [{tm_min}, {tm_max}] "
                f"(require_gc_clamp={require_gc_clamp}). "
                f"Using closest candidate (Tm={best_tm:.1f} C): {rev_chosen.upper()}"
            )
        else:
            raise ValueError(
                "Reverse primer search produced no candidates. "
                f"Check flank_min={flank_min}, flank_max={flank_max}, "
                f"binding_min_len={binding_min_len}, binding_max_len={binding_max_len}."
            )

    return fwd_chosen.lower(), rev_chosen.lower(), collected_warnings


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------

@dataclass
class MamePackageResult:
    """Paths produced by :func:`generate_mame_package`."""
    barcodes_xlsx: Path
    amplicon_fa: Path
    sample_map_template: Path
    context_json: Path
    warnings: list[str] = field(default_factory=list)
    amplicon_length: int | None = None
    #: Number of pre-filled data rows in ``sample_map_template`` (0 = header
    #: only, i.e. no ``expected_mutations_path`` was supplied). The UI reports
    #: this so an operator can tell a drafted template from a blank one.
    sample_map_prefilled_rows: int = 0
    #: True when an existing ``sample_map_template.xlsx`` already carried well
    #: assignments and was therefore left untouched instead of regenerated.
    sample_map_preserved: bool = False


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def generate_mame_package(
    fasta_path: Path,
    gene_start: int,
    gene_end: int,
    barcode_seeds_path: Path,
    output_dir: Path,
    project_root: Path,
    gene_name: str,
    polymerase: str = "Q5",
    flank_min: int = 100,
    flank_max: int = 400,
    binding_min_len: int = 18,
    binding_max_len: int = 35,
    tm_min: float = 55.0,
    tm_max: float = 68.0,
    require_gc_clamp: bool = True,
    topology: str | None = None,
    expected_mutations_path: Path | None = None,
    variant_sheet: str | None = None,
    variant_column: str | None = None,
) -> MamePackageResult:
    """Generate the complete MAME input package for a sequencing run.

    Steps
    -----
    1. Read the first sequence from ``fasta_path`` (warn if multi-record).
    2. Call :func:`design_flanking_primers` to obtain fwd/rev flanking sequences.
    3. Call :func:`parse_barcode_seeds` to obtain 12 fwd + 8 rev seed sequences.
    4. Write ``barcodes_sequence.xlsx`` (20 data rows, 1 header row).
       Row format: name ``{sanitized_gene}_f_N`` / ``{sanitized_gene}_r_N``,
       sequence = SEED + flanking (all upper). The ``sanitized_gene`` prefix
       is derived from the ``gene_name`` parameter via
       :func:`_sanitize_gene_prefix`.
    5. Write ``{sanitized_gene_name}_amplicon.fa`` containing the target gene region.
    6. Write ``sample_map_template.xlsx`` (column A "sample_name", column B
       "well"). Header-only by default; pre-filled with a draft placement when
       ``expected_mutations_path`` is supplied. An existing template that
       already holds data rows is left untouched (operator well assignments
       cannot be regenerated) and reported via ``sample_map_preserved``.
    7. Write ``mame_context.json`` at ``project_root`` with schema 1.

    Parameters
    ----------
    fasta_path:
        Path to a FASTA file containing the full CDS sequence.
    gene_start:
        0-based inclusive start of the gene region within the CDS.
    gene_end:
        0-based exclusive end of the gene region within the CDS.
    barcode_seeds_path:
        Path to the barcode seeds xlsx (fwd_1..12, rev_1..8).
    output_dir:
        Destination for barcodes_sequence.xlsx, amplicon .fa, and template.
        Created automatically (``parents=True``) if it does not exist.
    project_root:
        Root of the KUMA project. ``mame_context.json`` is written here.
        All paths in the JSON are relative to this directory.
    gene_name:
        Required gene label derived from input annotation or explicit user entry.
        It is sanitized for barcode row names and the amplicon filename.
    polymerase:
        Name of the polymerase profile to use for Tm calculation.
        Must be one of the keys in ``POLYMERASE_PROFILES`` (default "Q5").
    flank_min:
        Minimum distance (bp) from the gene boundary to the primer binding site.
    flank_max:
        Maximum distance (bp) from the gene boundary searched for a binding site.
    binding_min_len:
        Minimum primer binding length to try.
    binding_max_len:
        Maximum primer binding length to try.
    tm_min:
        Lower bound of the acceptable Tm window (deg C).
    tm_max:
        Upper bound of the acceptable Tm window (deg C).
    require_gc_clamp:
        If True, the 3' terminal base of every primer must be G or C.
    topology:
        Either "linear", "circular", or None (default). None means
        auto-detect from the sequence file: GenBank/SnapGene records carry an
        explicit topology annotation (falls back to "linear" if absent or
        unrecognised); plain FASTA has no topology annotation and is always
        treated as "linear". Pass "linear" or "circular" explicitly to
        override auto-detection.
    variant_sheet:
        Sheet holding the variant list, when the file is not a KURO export and
        the sheet cannot be inferred. Ignored for KURO exports.
    variant_column:
        Column holding the variant labels, for the same case. Ignored for KURO
        exports and unnecessary when the header is a recognised name.
    expected_mutations_path:
        Optional path to a variant list. A KURO results xlsx carrying an
        ``expected_mutations``
        sheet. When given, ``sample_map_template.xlsx`` is pre-filled with one
        row per designed mutant in column-major well order plus a trailing
        ``WT`` control row, so the operator verifies a draft instead of
        authoring the map from scratch. The placement is the same draft
        assumption used by ``mame.build_well_layout`` and still requires
        reconciliation against the physical plate. Omit for a header-only
        template.

    Returns
    -------
    :class:`MamePackageResult` with absolute paths of all four output files
    and a (possibly empty) ``warnings`` list.

    Raises
    ------
    FileNotFoundError
        If ``fasta_path`` or ``barcode_seeds_path`` does not exist.
    ValueError
        If FASTA parsing fails, gene range is invalid, barcode seeds fail
        validation, or the flank search window falls outside the CDS.
    """
    # Ensure output_dir exists
    output_dir = Path(output_dir)
    project_root = Path(project_root)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Resolve polymerase profile
    profile = get_profile(polymerase)

    # Step 1: parse CDS
    cds_seq, detected_topology = _parse_first_cds_sequence(Path(fasta_path))
    effective_topology = topology if topology is not None else detected_topology

    # Step 2: flanking primers (Tm-guided)
    fwd_flanking, rev_flanking, pkg_warnings = design_flanking_primers(
        cds_seq,
        gene_start=gene_start,
        gene_end=gene_end,
        profile=profile,
        flank_min=flank_min,
        flank_max=flank_max,
        binding_min_len=binding_min_len,
        binding_max_len=binding_max_len,
        tm_min=tm_min,
        tm_max=tm_max,
        require_gc_clamp=require_gc_clamp,
        topology=effective_topology,
    )

    # Step 3: barcode seeds
    seeds = parse_barcode_seeds(Path(barcode_seeds_path))

    # Step 4: barcodes_sequence.xlsx
    # Derive sanitized prefix from gene_name (raises ValueError on empty).
    gene_prefix = _sanitize_gene_prefix(gene_name)
    barcodes_xlsx_path = output_dir / "barcodes_sequence.xlsx"
    _write_barcodes_xlsx(
        path=barcodes_xlsx_path,
        seeds=seeds,
        fwd_flanking=fwd_flanking,
        rev_flanking=rev_flanking,
        gene_prefix=gene_prefix,
    )

    # Step 5: amplicon FASTA (gene region only)
    amplicon_fa_path = output_dir / f"{gene_prefix}_amplicon.fa"
    _write_amplicon_fasta(
        path=amplicon_fa_path,
        cds_seq=cds_seq,
        gene_start=gene_start,
        gene_end=gene_end,
        gene_name=gene_prefix,
    )

    # Step 6: sample map template. Pre-filled with a draft well placement when
    # a KURO expected_mutations xlsx is supplied, header-only otherwise.
    #
    # An existing template that already carries data rows is left untouched: it
    # holds hand-entered well assignments, and regeneration is routinely re-run
    # to adjust the gene range or polymerase. The caller confirms overwriting
    # output_dir as a whole, which is not informed consent for discarding the
    # one file in it that only the operator can reproduce.
    template_path = output_dir / "sample_map_template.xlsx"
    sample_map_rows: list[tuple[str, str]] = []
    sample_map_preserved = _sample_map_has_data(template_path)
    if expected_mutations_path is not None and not sample_map_preserved:
        sample_map_rows = _build_sample_map_rows(
            Path(expected_mutations_path),
            variant_sheet=variant_sheet,
            variant_column=variant_column,
        )
    if sample_map_preserved:
        pkg_warnings.append(
            f"sample_map_template.xlsx already contains well assignments and was "
            f"left unchanged ({template_path}). Delete or rename it to regenerate "
            "a draft template."
        )
    else:
        if sample_map_rows:
            pkg_warnings.append(
                f"sample_map_template.xlsx pre-filled with {len(sample_map_rows)} draft "
                "rows from expected_mutations (column-major placement, WT last). "
                "Verify against the physical plate before running analyze."
            )
        _write_sample_map_template(template_path, rows=sample_map_rows or None)

    # Step 7: mame_context.json
    context_json_path = project_root / "mame_context.json"
    _write_mame_context_json(
        path=context_json_path,
        project_root=project_root,
        barcodes_xlsx=barcodes_xlsx_path,
        amplicon_fa=amplicon_fa_path,
        sample_map_template=template_path,
    )

    amplicon_length = _compute_amplicon_length(
        cds_seq=cds_seq,
        fwd_flanking=fwd_flanking,
        rev_flanking=rev_flanking,
    )

    return MamePackageResult(
        barcodes_xlsx=barcodes_xlsx_path,
        amplicon_fa=amplicon_fa_path,
        sample_map_template=template_path,
        context_json=context_json_path,
        warnings=pkg_warnings,
        amplicon_length=amplicon_length,
        sample_map_prefilled_rows=len(sample_map_rows),
        sample_map_preserved=sample_map_preserved,
    )


def _compute_amplicon_length(
    cds_seq: str,
    fwd_flanking: str,
    rev_flanking: str,
) -> int | None:
    """Locate primer binding sites on the template and return PCR amplicon length.

    Returns None if either primer is not found on the template (defensive — this
    should not happen since design_flanking_primers picks the sequences directly
    from cds_seq, but search may fail if rev primer is reverse-complemented).
    """
    seq_upper = cds_seq.upper()
    fwd_upper = fwd_flanking.upper()
    fwd_pos = seq_upper.find(fwd_upper)
    if fwd_pos < 0:
        return None

    complement = str.maketrans("ACGTNacgtn", "TGCANtgcan")
    rev_binding = rev_flanking.translate(complement)[::-1].upper()
    rev_pos = seq_upper.rfind(rev_binding)
    if rev_pos < 0:
        return None

    return (rev_pos + len(rev_binding)) - fwd_pos


# ---------------------------------------------------------------------------
# Private write helpers
# ---------------------------------------------------------------------------

def _write_barcodes_xlsx(
    path: Path,
    seeds: dict[str, str],
    fwd_flanking: str,
    rev_flanking: str,
    gene_prefix: str,
) -> None:
    """Write the 20-row barcodes_sequence.xlsx consumed by MAME.

    Output sequence = SEED (uppercase) + flanking (lowercase), concatenated.
    The MAME sort_barcode module trims the flanking part during matching, so
    including it here satisfies the full-primer column B requirement.

    Row names are ``{gene_prefix}{_FWD_SUFFIX}{i}`` /
    ``{gene_prefix}{_REV_SUFFIX}{i}``. The reader (``parse_combinatorial_barcodes``)
    is gene-agnostic and matches any prefix that ends with ``_f_<int>`` /
    ``_r_<int>``.
    """
    import openpyxl  # local import

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Barcodes"
    ws.append(["name", "sequence"])

    for i in range(1, _N_FWD + 1):
        name = f"{gene_prefix}{_FWD_SUFFIX}{i}"
        seq = seeds[f"fwd_{i}"].upper() + fwd_flanking
        ws.append([name, seq])

    for i in range(1, _N_REV + 1):
        name = f"{gene_prefix}{_REV_SUFFIX}{i}"
        seq = seeds[f"rev_{i}"].upper() + rev_flanking
        ws.append([name, seq])

    wb.save(str(path))


def _write_amplicon_fasta(
    path: Path,
    cds_seq: str,
    gene_start: int,
    gene_end: int,
    gene_name: str,
) -> None:
    """Write a single-record FASTA of the gene region."""
    amplicon = cds_seq[gene_start:gene_end].upper()
    with path.open("w", encoding="utf-8") as fh:
        fh.write(f">{gene_name}_amplicon start={gene_start} end={gene_end}\n")
        # 60-character line wrap (standard FASTA)
        for i in range(0, len(amplicon), 60):
            fh.write(amplicon[i:i + 60] + "\n")


def _write_sample_map_template(
    path: Path,
    rows: list[tuple[str, str]] | None = None,
) -> None:
    """Write the sample-map xlsx consumed by MAME ``analyze``.

    With ``rows`` omitted the sheet carries the header only, leaving the
    operator to author every well by hand. With ``rows`` supplied (see
    :func:`_build_sample_map_rows`) the sheet is pre-filled with one
    ``(sample_name, well)`` pair per row, turning the operator task into
    "verify and correct" instead of "author from scratch".

    The pre-filled placement is a *draft*: it assumes KURO row order maps to
    column-major wells with a single WT control after the last mutant. The
    operator remains responsible for reconciling it with the physical plate.
    """
    import openpyxl  # local import

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "SampleMap"
    ws.append(["sample_name", "well"])
    for sample_name, well in rows or []:
        ws.append([sample_name, well])
    wb.save(str(path))


def _sample_map_has_data(path: Path) -> bool:
    """Return True if ``path`` is an xlsx carrying at least one data row.

    Used to decide whether re-running package generation may overwrite an
    existing ``sample_map_template.xlsx``. A header-only sheet holds no operator
    work and is safe to replace; a sheet with data rows holds hand-entered well
    assignments and must survive. An unreadable or corrupt file is treated as
    "no data" so a broken template never blocks regeneration.
    """
    if not path.exists():
        return False

    import openpyxl  # local import

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - unreadable file: fall through to rewrite
        return False
    try:
        ws = wb.worksheets[0]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:  # header
                continue
            if any(c is not None and str(c).strip() != "" for c in row):
                return True
        return False
    finally:
        wb.close()


def _build_sample_map_rows(
    expected_mutations_path: Path,
    variant_sheet: str | None = None,
    variant_column: str | None = None,
) -> list[tuple[str, str]]:
    """Return draft ``(sample_name, well)`` rows from a KURO results xlsx.

    Delegates placement to :func:`kuma_core.mame.layout.build_draft_layout` so
    the template and the ``mame.build_well_layout`` RPC cannot drift apart:
    column-major wells for the designed mutants, then a single ``WT`` control.

    Raises
    ------
    FileNotFoundError
        If ``expected_mutations_path`` does not exist.
    ValueError
        If the xlsx carries no ``expected_mutations`` sheet or no DESIGNED rows,
        or if the mutation set does not fit one 96-well plate. Callers pass this
        path deliberately, so an unusable file is an error rather than a reason
        to emit a misleading template.
    """
    # Local imports: keep module import cost off the barcode-only code path.
    from kuma_core.mame.io.variant_list import read_variant_source
    from kuma_core.mame.layout import build_draft_layout

    # Accepts a KURO export or a plain variant list. The KURO shape is detected
    # and routed to the strict reader, so its behaviour is unchanged.
    read = read_variant_source(
        Path(expected_mutations_path),
        sheet=variant_sheet,
        variant_column=variant_column,
    )
    expected = read.expected
    if not expected:
        raise ValueError(
            f"No DESIGNED expected mutations found in {expected_mutations_path}; "
            "cannot pre-fill the sample map template. Omit "
            "expected_mutations_path to emit a header-only template."
        )
    # A source that lists its own WT row must not also get an appended one.
    draft = build_draft_layout(expected, include_wt=not read.has_explicit_wt)
    # A clamped draft is refused rather than written. The template is a file the
    # operator edits with no confirmation step in between, and a truncated sheet
    # reads as a correct full plate, so writing one would hide the truncation
    # until wells past the cut are mis-scored.
    if not draft.is_complete:
        detail = (
            f"{len(draft.dropped_mutant_ids)} mutants past well 96 "
            f"({', '.join(draft.dropped_mutant_ids[:5])}"
            f"{', ...' if len(draft.dropped_mutant_ids) > 5 else ''})"
            if draft.dropped_mutant_ids
            else "no well left for the WT control"
        )
        raise ValueError(
            f"{len(expected)} designed mutations do not fit one 96-well plate: "
            f"{detail}. The combinatorial barcode space is 12 fwd x 8 rev, so "
            "wells past the 96th cannot be told apart in the reads. Split the "
            "campaign across plates (MAME separates plates by native barcode) "
            "and supply one sample map per plate, or omit "
            "expected_mutations_path for a header-only template."
        )
    # build_draft_layout returns an insertion-ordered dict[well, sample] in
    # column-major order (WT last when it fits); the sheet stores the columns
    # in sample_name/well order, matching parse_sample_map.
    return [(sample, well) for well, sample in draft.layout.items()]


def _ctx_path(p: Path, root: Path) -> str:
    """Return relative posix path if inside root, else absolute posix path."""
    rp = p.resolve()
    try:
        return rp.relative_to(root).as_posix()
    except ValueError:
        return rp.as_posix()


def _write_mame_context_json(
    path: Path,
    project_root: Path,
    barcodes_xlsx: Path,
    amplicon_fa: Path,
    sample_map_template: Path,
) -> None:
    """Write mame_context.json with relative paths for files inside project_root,
    or absolute paths for files outside it.
    """
    root = project_root.resolve()
    context = {
        "schema": 1,
        "published_at": datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "custom_barcodes_path": _ctx_path(barcodes_xlsx, root),
        "reference_path": _ctx_path(amplicon_fa, root),
        "sample_map_template_path": _ctx_path(sample_map_template, root),
    }
    path.write_text(
        json.dumps(context, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
