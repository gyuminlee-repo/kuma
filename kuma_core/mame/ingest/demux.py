"""A1 — Custom barcode demultiplexer.

Separates reads from a MinKNOW native-barcode FASTQ folder into per-well FASTA
files using 96 custom barcodes.

Strategy
--------
1. If cutadapt is available and ``use_cutadapt=True``, dispatch to cutadapt
   via subprocess with ``-g file:<adapters_fasta>`` (one adapter per well).
2. Otherwise fall back to pure-Python prefix-scan with Hamming distance ≤
   ``ceil(len(barcode) * error_tolerance)``.

Cutadapt details
----------------
- Adapters written to a temporary FASTA to avoid huge command lines.
- ``-e <error_tolerance>`` passed as a single float.
- ``--discard-untrimmed`` routes unmatched reads to a separate sink.
- Invoked with ``shell=False`` and ``list`` args (security: code-security §SQL).
- When ``linked_trim=True`` and ``rev_primer_universal`` is provided, cutadapt
  receives both ``-g file:<fwd_adapters>`` and ``-a <rc(rev)>`` in one pass,
  trimming both ends simultaneously (cutadapt linked-adapter mode).

Pure-Python fallback
--------------------
- Iterates over every read; for each barcode computes Hamming distance to the
  read prefix (up to 3 × barcode length to allow for indels in the prefix).
- Best-match wins; ties (equal distance) → unassigned.
- When ``linked_trim=True`` and ``rev_primer_universal`` is provided, the
  reverse complement of ``rev_primer_universal`` is searched near the read
  3′-end (last 60 bp) using Hamming distance ≤ ``ceil(len * error_tolerance)``;
  if found the read is trimmed to that position.
- Reads are written as FASTA (one record per read, ID taken from FASTQ header).

Header normalisation
--------------------
When ``normalize_headers=True``, every FASTA record in the output file for
well ``{well}`` receives the header ``>{well}`` (replacing the original
ONT UUID). This matches the sort_barcode convention observed in external
pipeline output. All records in a per-well file share the same header —
``fasta_parser.parse_fasta_file`` counts ``>`` lines as ``read_count``, so
this is downstream-compatible.

Output
------
``{output_dir}/{well_name}.fasta`` for each well with ≥ 1 assigned read.
Unassigned reads are discarded (counted only).
"""

from __future__ import annotations

import gzip
import math
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

_BARCODE_RE = re.compile(r"^[ACGTacgt]{5,60}$")
_WELL_NAME_RE = re.compile(r"^[A-Za-z0-9_\-]{1,32}$")


@dataclass
class DemuxResult:
    output_dir: Path
    n_input_reads: int
    n_assigned: int
    n_unassigned: int
    per_well_counts: dict[str, int] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Validation helpers
# ---------------------------------------------------------------------------


def _validate_custom_barcodes(barcodes: dict[str, str]) -> None:
    """Validate barcode name → sequence mapping. Raises ValueError on bad data."""
    if not barcodes:
        raise ValueError("custom_barcodes must not be empty")
    for name, seq in barcodes.items():
        if not _WELL_NAME_RE.match(name):
            raise ValueError(
                f"Barcode name {name!r} contains invalid characters "
                "(allowed: A-Z a-z 0-9 _ -)"
            )
        if not _BARCODE_RE.match(seq):
            raise ValueError(
                f"Barcode sequence for {name!r} is invalid: {seq!r}. "
                "Must be 5–60 ACGT characters."
            )


def _validate_error_tolerance(tol: float) -> float:
    """Clamp and validate error_tolerance to [0.0, 0.5]."""
    if not (0.0 <= tol <= 0.5):
        raise ValueError(
            f"error_tolerance must be in [0.0, 0.5], got {tol!r}"
        )
    return tol


# ---------------------------------------------------------------------------
# FASTQ reading (handles gzip-compressed files transparently)
# ---------------------------------------------------------------------------


def _open_fastq(path: Path):
    """Open a FASTQ file, decompressing on-the-fly if .gz."""
    if path.suffix.lower() == ".gz":
        return gzip.open(path, "rt", encoding="utf-8")
    return open(path, "r", encoding="utf-8")  # noqa: WPS515


def _iter_fastq_records(
    path: Path,
) -> Iterator[tuple[str, str]]:
    """Yield (read_id, sequence) pairs from a FASTQ file."""
    with _open_fastq(path) as fh:
        while True:
            header = fh.readline()
            if not header:
                break
            seq = fh.readline().rstrip("\r\n")
            fh.readline()  # '+'
            fh.readline()  # quality
            read_id = header.lstrip("@").split()[0].rstrip("\r\n")
            yield read_id, seq.upper()


# ---------------------------------------------------------------------------
# Reference xlsx parsing (Sheet1, column L)
# ---------------------------------------------------------------------------

_XLSX_SUFFIXES = {".xlsx"}
_CSV_SUFFIXES = {".csv"}


def parse_custom_barcodes(
    path: Path,
    sheet: str = "Sheet1",
    col: str = "L",
) -> dict[str, str]:
    """Parse custom barcodes from an xlsx or csv file.

    xlsx
    ----
    Reads ``sheet`` and looks for:
    - Column ``col`` (1-based letter, e.g. "L" = 12th column) for the barcode
      sequence.
    - First column (A) for the well name; if blank the row index is used.
    - Rows without a valid ACGT sequence in column ``col`` are skipped.

    csv
    ---
    Expected columns: ``name`` and ``sequence`` (header row required).

    Returns
    -------
    ``dict[str, str]``: well_name → barcode_sequence (upper-cased).
    Empty dict if no valid rows found.
    """
    suffix = path.suffix.lower()
    if suffix in _XLSX_SUFFIXES:
        return _parse_barcodes_xlsx(path, sheet=sheet, col=col)
    if suffix in _CSV_SUFFIXES:
        return _parse_barcodes_csv(path)
    raise ValueError(
        f"Unsupported barcode file format: {suffix!r}. "
        "Supported: .xlsx, .csv"
    )


def _col_letter_to_index(col: str) -> int:
    """Convert Excel-style column letter to 0-based index ('A'→0, 'L'→11)."""
    col = col.strip().upper()
    result = 0
    for ch in col:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _parse_barcodes_xlsx(
    path: Path, sheet: str, col: str
) -> dict[str, str]:
    import openpyxl  # local import: keeps cold-start fast

    col_idx = _col_letter_to_index(col)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            # Graceful: try first sheet.
            ws = wb.worksheets[0]
        else:
            ws = wb[sheet]

        result: dict[str, str] = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Name from column A (index 0); fallback to row index.
            name_raw = row[0] if row else None
            name = str(name_raw).strip() if name_raw is not None else f"well_{i}"

            # Sequence from target column.
            if col_idx >= len(row):
                continue
            seq_raw = row[col_idx]
            if seq_raw is None:
                continue
            seq = str(seq_raw).strip().upper()
            if not _BARCODE_RE.match(seq):
                continue
            # Skip duplicates silently (first wins).
            if name not in result:
                result[name] = seq
        return result
    finally:
        wb.close()


def _parse_barcodes_csv(path: Path) -> dict[str, str]:
    import csv

    result: dict[str, str] = {}
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            name = (row.get("name") or "").strip()
            seq = (row.get("sequence") or "").strip().upper()
            if not name or not _BARCODE_RE.match(seq):
                continue
            if name not in result:
                result[name] = seq
    return result


# ---------------------------------------------------------------------------
# Reverse-complement helper (no external deps)
# ---------------------------------------------------------------------------

_COMPLEMENT: dict[str, str] = {
    "A": "T", "T": "A", "C": "G", "G": "C",
    "a": "t", "t": "a", "c": "g", "g": "c",
}


def _rc(seq: str) -> str:
    """Return the reverse complement of a DNA sequence."""
    return "".join(_COMPLEMENT.get(b, "N") for b in reversed(seq))


# ---------------------------------------------------------------------------
# Hamming distance utility
# ---------------------------------------------------------------------------


def _hamming_prefix(read_seq: str, barcode: str) -> int:
    """Compute Hamming distance between barcode and the read prefix.

    Only the first ``len(barcode)`` bases of the read are compared.
    If the read is shorter than the barcode the distance is treated as infinite
    (returns ``len(barcode) + 1``).
    """
    if len(read_seq) < len(barcode):
        return len(barcode) + 1
    mismatches = 0
    for a, b in zip(barcode, read_seq[: len(barcode)]):
        if a != b:
            mismatches += 1
    return mismatches


# ---------------------------------------------------------------------------
# Pure-Python demux
# ---------------------------------------------------------------------------


def _trim_rev_primer(
    seq: str,
    rev_rc: str,
    max_mismatches: int,
    search_tail_bp: int = 60,
) -> str:
    """Trim the 3′ end of a read at the rightmost occurrence of rev_rc within tolerance.

    Searches only the last ``search_tail_bp`` bases to avoid false hits in the
    amplicon body. Scans right-to-left so that the primer closest to the actual
    3′ end is preferred over an earlier false positive caused by 1-mismatch hits
    in the amplicon body.

    Returns the original sequence unchanged if no hit is found.
    """
    rev_len = len(rev_rc)
    if len(seq) < rev_len:
        return seq
    tail_start = max(0, len(seq) - search_tail_bp)
    tail = seq[tail_start:]
    # Scan right-to-left: prefer the rightmost (3′-most) hit.
    for i in range(len(tail) - rev_len, -1, -1):
        mismatches = 0
        for a, b in zip(rev_rc, tail[i:i + rev_len]):
            if a != b:
                mismatches += 1
            if mismatches > max_mismatches:
                break
        if mismatches <= max_mismatches:
            cut = tail_start + i
            return seq[:cut]
    return seq


def _demux_python(
    fastq_files: list[Path],
    custom_barcodes: dict[str, str],
    output_dir: Path,
    error_tolerance: float,
    linked_trim: bool = False,
    rev_primer_universal: str | None = None,
    normalize_headers: bool = True,
) -> DemuxResult:
    """Demux using pure-Python Hamming-distance prefix matching.

    Parameters
    ----------
    linked_trim:
        When True and ``rev_primer_universal`` is provided, trim the 3′ end of
        each read at the reverse complement of the primer.
    rev_primer_universal:
        Sequence of the universal reverse primer (5′→3′).  Its RC is searched
        near the 3′ end of each read and clipped off.
    normalize_headers:
        When True, write ``>{well_name}`` as the FASTA header for every read
        in the well file instead of the original ONT UUID.
    """
    # Precompute max allowed mismatches per barcode.
    max_mm: dict[str, int] = {
        name: math.ceil(len(seq) * error_tolerance)
        for name, seq in custom_barcodes.items()
    }

    # Precompute rev-primer RC and its mismatch threshold (if applicable).
    rev_rc: str | None = None
    rev_max_mm: int = 0
    if linked_trim and rev_primer_universal:
        rev_rc = _rc(rev_primer_universal)
        rev_max_mm = math.ceil(len(rev_rc) * error_tolerance)

    per_well_counts: dict[str, int] = {}
    writers: dict[str, object] = {}
    output_dir.mkdir(parents=True, exist_ok=True)

    n_input = 0
    n_assigned = 0
    n_unassigned = 0

    try:
        for fastq_path in fastq_files:
            for read_id, seq in _iter_fastq_records(fastq_path):
                n_input += 1

                best_name: str | None = None
                best_dist = 999

                for name, barcode in custom_barcodes.items():
                    dist = _hamming_prefix(seq, barcode)
                    if dist < best_dist:
                        best_dist = dist
                        best_name = name
                    elif dist == best_dist:
                        # Tie → unresolved; keep best_name=None only if same dist
                        best_name = None  # ambiguous tie

                if best_name is None:
                    n_unassigned += 1
                    continue

                # Check threshold.
                threshold = max_mm[best_name]
                if best_dist > threshold:
                    n_unassigned += 1
                    continue

                # Trim forward barcode from 5′ end (only in linked_trim mode;
                # preserves R6 full-sequence behaviour when linked_trim=False).
                bc_len = len(custom_barcodes[best_name])
                trimmed_seq = seq[bc_len:] if linked_trim else seq

                # Optionally trim reverse primer from 3′ end.
                if rev_rc is not None:
                    trimmed_seq = _trim_rev_primer(trimmed_seq, rev_rc, rev_max_mm)

                # Write FASTA record.
                if best_name not in writers:
                    fasta_path = output_dir / f"{best_name}.fasta"
                    writers[best_name] = open(  # noqa: WPS515
                        fasta_path, "w", encoding="utf-8"
                    )
                fh = writers[best_name]
                header = best_name if normalize_headers else read_id
                fh.write(f">{header}\n{trimmed_seq}\n")  # type: ignore[union-attr]

                per_well_counts[best_name] = per_well_counts.get(best_name, 0) + 1
                n_assigned += 1
    finally:
        for fh in writers.values():
            fh.close()  # type: ignore[union-attr]

    return DemuxResult(
        output_dir=output_dir,
        n_input_reads=n_input,
        n_assigned=n_assigned,
        n_unassigned=n_unassigned,
        per_well_counts=per_well_counts,
    )


# ---------------------------------------------------------------------------
# cutadapt-backed demux
# ---------------------------------------------------------------------------


def _build_adapters_fasta(barcodes: dict[str, str], tmp_dir: str) -> str:
    """Write barcode sequences to a temporary FASTA file and return its path."""
    fasta_path = Path(tmp_dir) / "adapters.fasta"
    with open(fasta_path, "w", encoding="utf-8") as fh:
        for name, seq in barcodes.items():
            fh.write(f">{name}\n{seq}\n")
    return str(fasta_path)


def _collect_cutadapt_outputs(
    output_dir: Path,
    custom_barcodes: dict[str, str],
) -> tuple[dict[str, int], int, int]:
    """Count reads in per-well FASTA files written by cutadapt.

    Returns (per_well_counts, n_assigned, n_unassigned).
    Unassigned: reads in ``_unassigned.fasta`` (if present).
    """
    per_well_counts: dict[str, int] = {}
    n_assigned = 0
    for name in custom_barcodes:
        fp = output_dir / f"{name}.fasta"
        if not fp.exists():
            continue
        count = sum(1 for ln in fp.read_text(encoding="utf-8").splitlines() if ln.startswith(">"))
        if count:
            per_well_counts[name] = count
            n_assigned += count

    unassigned_file = output_dir / "_unassigned.fasta"
    n_unassigned = 0
    if unassigned_file.exists():
        n_unassigned = sum(
            1 for ln in unassigned_file.read_text(encoding="utf-8").splitlines()
            if ln.startswith(">")
        )
    return per_well_counts, n_assigned, n_unassigned


def _demux_cutadapt(
    fastq_files: list[Path],
    custom_barcodes: dict[str, str],
    output_dir: Path,
    error_tolerance: float,
    linked_trim: bool = False,
    rev_primer_universal: str | None = None,
    normalize_headers: bool = True,
) -> DemuxResult:
    """Demux via cutadapt subprocess.

    Calls cutadapt with:
    - ``-g file:<adapters.fasta>`` for all 96 barcodes in one shot.
    - ``-e <error_tolerance>`` as per-adapter error rate.
    - ``--discard-untrimmed`` to route unmatched reads to ``_unassigned.fasta``.
    - FASTQ → FASTA conversion via ``--fasta``.
    - When ``linked_trim=True`` and ``rev_primer_universal`` is provided,
      ``-a <rc(rev)>`` is added to trim the 3′ end of every matched read.

    When ``normalize_headers=True``, a post-processing pass rewrites every FASTA
    header in the per-well output files to ``>{well_name}``.
    """
    import logging

    logger = logging.getLogger(__name__)

    output_dir.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp_dir:
        adapters_fasta = _build_adapters_fasta(custom_barcodes, tmp_dir)

        # Concatenate all input FASTQ paths (cutadapt accepts multiple inputs).
        input_paths = [str(p) for p in fastq_files]

        cmd = [
            "cutadapt",
            "-g", f"file:{adapters_fasta}",
            "-e", str(error_tolerance),
            "--discard-untrimmed",
            "--fasta",
            "-o", str(output_dir / "{name}.fasta"),
            "--untrimmed-output", str(output_dir / "_unassigned.fasta"),
        ]

        # Linked adapter: trim rev-primer RC from 3′ end.
        if linked_trim and rev_primer_universal:
            rev_rc_seq = _rc(rev_primer_universal)
            cmd += ["-a", rev_rc_seq]

        cmd += input_paths

        proc = subprocess.run(
            cmd,
            shell=False,
            capture_output=True,
            text=True,
            timeout=600,
            check=False,
        )

        if proc.returncode != 0:
            logger.warning(
                "cutadapt exited %d; stderr: %s",
                proc.returncode,
                proc.stderr[:500],
            )
            raise RuntimeError(
                f"cutadapt failed (exit {proc.returncode}): {proc.stderr[:200]}"
            )

    per_well_counts, n_assigned, n_unassigned = _collect_cutadapt_outputs(
        output_dir, custom_barcodes
    )

    # Header normalization: rewrite every per-well FASTA header to >{well_name}.
    if normalize_headers:
        for well_name in list(per_well_counts.keys()):
            fp = output_dir / f"{well_name}.fasta"
            if not fp.exists():
                continue
            lines = fp.read_text(encoding="utf-8").splitlines(keepends=True)
            normalized: list[str] = []
            for ln in lines:
                if ln.startswith(">"):
                    normalized.append(f">{well_name}\n")
                else:
                    normalized.append(ln)
            fp.write_text("".join(normalized), encoding="utf-8")

    n_input = n_assigned + n_unassigned

    return DemuxResult(
        output_dir=output_dir,
        n_input_reads=n_input,
        n_assigned=n_assigned,
        n_unassigned=n_unassigned,
        per_well_counts=per_well_counts,
    )


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def demux_native_barcode(
    fastq_dir: Path,
    custom_barcodes: dict[str, str],
    output_dir: Path,
    error_tolerance: float = 0.1,
    use_cutadapt: bool = True,
    linked_trim: bool = False,
    rev_primer_universal: str | None = None,
    normalize_headers: bool = True,
) -> DemuxResult:
    """Demux a native barcode folder into per-well FASTA files.

    Parameters
    ----------
    fastq_dir:
        Path to a MinKNOW native-barcode directory containing ``*.fastq`` or
        ``*.fastq.gz`` files (e.g. ``fastq_pass/barcode06/``).
    custom_barcodes:
        Mapping of well name → barcode sequence (ACGT, 5–60 bp).
    output_dir:
        Destination directory; created if it does not exist.
    error_tolerance:
        Per-base mismatch rate [0.0, 0.5]. Pure-Python: mapped to
        ``ceil(len * tol)`` max mismatches. cutadapt: passed as ``-e``.
    use_cutadapt:
        If True *and* cutadapt is on PATH, use the cutadapt backend.
        Otherwise falls back to pure-Python Hamming matching.
    linked_trim:
        When True *and* ``rev_primer_universal`` is provided, trim the
        reverse complement of ``rev_primer_universal`` from the 3′ end of
        each read after forward-barcode demux.  Defaults to False — safe
        to enable only once the reverse primer sequence is known.
    rev_primer_universal:
        Sequence of the universal reverse primer (5′→3′). Required when
        ``linked_trim=True``; ignored otherwise.
    normalize_headers:
        When True, write ``>{well_name}`` as the FASTA record header for
        every read in the well file.  When False, the original FASTQ read ID
        is preserved.  Defaults to True (matches sort_barcode convention).

    Returns
    -------
    DemuxResult with n_input_reads, n_assigned, n_unassigned, per_well_counts.

    Raises
    ------
    ValueError
        If ``linked_trim=True`` but ``rev_primer_universal`` is None or empty.
    """
    # Input validation (security: code-security §Input Validation).
    if not fastq_dir.is_dir():
        raise FileNotFoundError(f"fastq_dir is not a directory: {fastq_dir}")
    _validate_custom_barcodes(custom_barcodes)
    error_tolerance = _validate_error_tolerance(error_tolerance)

    if linked_trim and not rev_primer_universal:
        raise ValueError(
            "linked_trim=True requires rev_primer_universal to be a non-empty string. "
            "Provide the universal reverse primer sequence or set linked_trim=False."
        )

    # Collect FASTQ files.
    fastq_files: list[Path] = sorted(
        [p for p in fastq_dir.rglob("*.fastq")]
        + [p for p in fastq_dir.rglob("*.fastq.gz")]
    )
    if not fastq_files:
        raise FileNotFoundError(
            f"No FASTQ files found in {fastq_dir}. "
            "Expected *.fastq or *.fastq.gz files."
        )

    # Backend selection.
    if use_cutadapt and shutil.which("cutadapt") is not None:
        return _demux_cutadapt(
            fastq_files,
            custom_barcodes,
            output_dir,
            error_tolerance,
            linked_trim=linked_trim,
            rev_primer_universal=rev_primer_universal,
            normalize_headers=normalize_headers,
        )

    return _demux_python(
        fastq_files,
        custom_barcodes,
        output_dir,
        error_tolerance,
        linked_trim=linked_trim,
        rev_primer_universal=rev_primer_universal,
        normalize_headers=normalize_headers,
    )


__all__ = [
    "DemuxResult",
    "demux_native_barcode",
    "parse_custom_barcodes",
    "_rc",
    "_trim_rev_primer",
]
