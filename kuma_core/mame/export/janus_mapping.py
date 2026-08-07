"""Janus mapping export for final cell-stock pick (mame K4 spec).

Header design follows the 260428 meeting §2.5 decision:
  name | source_plate | source_well | dest_well | priority_score

- ``source_plate``: the plate label ``nb_label`` produces from the barcode
  directory the replicate was selected from ("sort_barcode07" -> "NB07",
  "NB01" -> "NB01", a name without digits unchanged). That helper is the single
  source of truth every MAME export uses, the result workbook included, so the
  two files a run produces name the same plate the same way. A fixed
  NB01->P1 / NB02->P2 / NB03->P3 dictionary lived here until v0.15.7 and never
  matched anything: the selected plate is a barcode directory name, so the
  lookup missed and the raw name was written.
- ``source_well``:  well label in the NB plate (e.g. "A1")
- ``dest_well``:    destination well in the final 96-well plate.
                    Auto-filled from the custom_barcode position.
                    Users can overwrite in the saved CSV/XLSX.
- ``priority_score``: ``read_count`` when available (G6/A6+); otherwise
                      ``file_size_kb`` as a volume proxy (Phase 1 fallback).

Row order: the plate map, not sequencing depth
-----------------------------------------------
Rows are laid down in ``source_well`` order, column-major (A1, B1 ... H1, then
A2), which is the one traversal
:data:`~kuma_core.mame.plate_geometry.DEFAULT_ADDRESSING` defines and the one
the result table already reads. The operator fills the final plate by hand with
the step 2.2 plate map in front of them, so this file has to run in the same
direction as the plate they are reading; screen, file and bench then agree on
one axis.

It sorted by ``priority_score`` DESC before (the §2.5 note of the 260428
meeting recommended placing the highest-volume clones first). That put the
deepest-sequenced clone in A1 regardless of where it sat on the source plate,
which turned every next well into a search back through the plate map. The
score itself is unchanged and still written: it ranks depth, which is worth
reading, it just no longer decides where anything goes.

Ties are only reachable across plates, since one plate holds at most one pick
per position, and they break by plate order (``nb_order_key`` then the label),
the same order ``JanusSettings.resolve_deck`` numbers the stock plates in. A
pick whose ``custom_barcode`` names no well has no position to be placed by and
sorts before every placed pick, as
:meth:`~kuma_core.mame.plate_geometry.PlateAddressing.sort_key` does for the
same reason: at the top it is seen. It is never only sorted, either, since
``_find_unresolved_wells`` names each one as an error.

read_count policy (G6/A6)
--------------------------
``BarcodeRecord.read_count`` is populated by the consensus parser from
``depth=N`` header metadata when available, falling back to single-record
counts for legacy consensus files. ``priority_score`` uses read_count when
non-None; falls back to file_size_kb. Column name ``priority_score`` is kept
for downstream consumers regardless of which underlying metric is used.

G3 run-meta embedding
---------------------
``export_mame_janus_csv`` and ``export_mame_janus_xlsx`` accept an optional
``ngs_run_meta`` argument (``NgsRunMeta | None``).

- CSV: when *ngs_run_meta* is not ``None``, a single comment line is prepended
  before the header row::

      # kuma_run_meta: flow_cell=PAX12345, kit=SQK-LSK109, started=2024-01-01T00:00:00Z

  When ``None`` no comment line is written, preserving backward compatibility
  with existing tests that use ``csv.DictReader`` directly.

- XLSX: a ``__kuma_meta__`` sheet is appended with key/value rows.  The sheet
  is always present; content is optional (placeholder when meta is ``None``).
"""

from __future__ import annotations

import csv
import datetime
from collections.abc import Iterable
from dataclasses import dataclass, replace
from pathlib import Path
from typing import TYPE_CHECKING

from kuma_core.mame.export.nb_label import nb_label, nb_order_key
from kuma_core.mame.models import ReplicateResult, VerdictClass
from kuma_core.mame.plate_geometry import DEFAULT_ADDRESSING, PlateAddressing
from kuma_core.shared.janus_deck import JANUS_DEVICE_HEADER

if TYPE_CHECKING:
    from kuma_core.mame.ingest.run_meta import NgsRunMeta

_JANUS_HEADER = [
    "name",
    "source_plate",
    "source_well",
    "dest_well",
    "priority_score",
]

# Instrument-native worksheet header. The literal lives in
# ``kuma_core/shared/janus_deck.py`` so KURO and MAME write the same columns
# from one definition; the provenance comment (which workbook, which fixture
# pins it) lives alongside it there.
_JANUS_DEVICE_HEADER = JANUS_DEVICE_HEADER

# Finding severity. ``error`` withholds the file (the rows themselves cannot be
# written correctly); ``warning`` names a value the export generated rather than
# was given, which the operator has to see but which is never a reason to
# produce no file.
SEVERITY_ERROR = "error"
SEVERITY_WARNING = "warning"

SCHEMA_DEVICE = "device"
SCHEMA_LEGACY5 = "legacy5"
_SCHEMAS = (SCHEMA_DEVICE, SCHEMA_LEGACY5)

# The name this schema carried while the sheet had nine columns. The sheet has
# eight, so the old name states a fact that is no longer true, but a project
# saved before the rename still asks for it by that name and a request already
# in flight still carries it. ``_janus_settings_from_params`` folds it into
# ``SCHEMA_DEVICE`` on the way in; the constant lives here so the schema
# vocabulary stays defined in one place rather than as a literal in the handler.
SCHEMA_DEVICE_FORMER_NAME = "device9"

DEST_LAYOUT_COMPACT = "compact"
DEST_LAYOUT_SOURCE = "source"
_DEST_LAYOUTS = (DEST_LAYOUT_SOURCE, DEST_LAYOUT_COMPACT)

# Cell-stock picking keeps only fully verified clones by default:
# AMBIGUOUS carries the designed change plus a side indel, so its activity
# measurement would be mislabelled; LOWDEPTH is simply unverified.
DEFAULT_INCLUDE_VERDICTS: tuple[str, ...] = (VerdictClass.PASS.value,)

# The cell-stock transfer volume this lab uses for this run, given by the
# operator who runs the instrument. It is not derived from anything else in this
# repository: the KURO default (2.0 µL) is primer dispensing, a different
# operation with a different volume, so the two numbers are not comparable and
# neither one can be read off the other. Still editable in the export dialog for
# a run that transfers something else; the preview shows the value that will be
# written.
DEFAULT_VOLUME_UL = 70.0

# Assumption: ``type`` labels the transferred material on the instrument sheet.
# The lab workbook writes "cell stock" on every seeding row (KURO writes
# "primer" for primer dispensing, a different operation). Editable.
DEFAULT_SAMPLE_TYPE = "cell stock"

# The ``Asp. Rack`` / ``Dsp. Rack`` cells carry plate NAMES, because the JANUS
# software matches labware by the name printed on it rather than by deck slot.
# The names are generated from the run, not asked for: the plates a run picked
# from take "Stock plate1", "Stock plate2" ... in plate order, and everything is
# dispensed into one "final culture plate", which is the shape of the seeding
# workbook the lab imports.
#
# They are derived from the plates a run actually used rather than from a fixed
# dictionary, because which NB plates exist is a property of the run: a fixed
# ``NB01/NB02/NB03`` map (here until v0.15.7 as P1/P2/P3, then as NB labels)
# left every other run unnamed.
#
# Empty tuple / ``None`` on ``JanusSettings`` mean "derive"; see ``resolve_deck``.
SOURCE_PLATE_PREFIX = "Stock plate"
DEST_PLATE_NAME = "final culture plate"
DEFAULT_SOURCE_RACKS: dict[str, str] = {}
DEFAULT_DEST_RACK: str | None = None

# No default value: the liquid class drives the pipetting behaviour of the
# robot, so a guessed value would silently change how cells are handled. The
# eight column sheet has no liquid class cell, so nothing is written for it
# either way; the value is recorded with the run because the operator sets it
# and it describes how the run was pipetted. Blank is not reported: warning
# about a value that reaches no file is noise.
DEFAULT_LIQUID_CLASS = ""


def _require_plate_name(value: object, label: str) -> str:
    """Guard an operator override: a plate name the instrument can match.

    Replaces a positive-integer guard. The cells used to carry deck numbers, so
    the check that kept an unusable deck value off the sheet is now the check
    that keeps an unusable plate name off it: a blank cell is an instruction the
    robot cannot follow either way.
    """
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"Invalid {label} {value!r}. Expected a plate name.")
    return value


def normalize_include_verdicts(raw: object | None) -> tuple[str, ...]:
    """Validate and normalise the requested verdict classes.

    ``None`` selects the default (PASS only). An empty selection is a
    ``ValueError`` rather than "include nothing": a silently empty stock plate
    is the failure mode this export exists to prevent.
    """
    if raw is None:
        return DEFAULT_INCLUDE_VERDICTS
    if isinstance(raw, str):
        raise ValueError(
            "include_verdicts must be a list of verdict names, not a single string."
        )
    try:
        names = [str(v).strip().upper() for v in raw]  # type: ignore[union-attr]
    except TypeError as exc:  # not iterable
        raise ValueError("include_verdicts must be a list of verdict names.") from exc
    if not names:
        raise ValueError(
            "include_verdicts is empty: the export would contain no clones. "
            f"Choose at least one of {[v.value for v in VerdictClass]}."
        )
    valid = {v.value for v in VerdictClass}
    unknown = [n for n in names if n not in valid]
    if unknown:
        raise ValueError(
            f"Unknown verdict class(es) {unknown}. Expected any of {sorted(valid)}."
        )
    # Preserve order, drop duplicates.
    return tuple(dict.fromkeys(names))


def _exclusion_reason(
    rr: ReplicateResult,
    include_verdicts: tuple[str, ...],
    include_fallback: bool,
) -> tuple[str, str] | None:
    """Return ``(reason, verdict_value)`` when *rr* must be left out, else ``None``.

    Reason precedence is fixed and pinned by test: ``failed`` > ``no_selection``
    > ``missing_verdict`` > ``verdict_class`` > ``fallback``. ``fallback`` is
    evaluated independently of the verdict class, so a fallback pick whose plate
    happens to read PASS is still excluded by default.
    """
    if rr.failed:
        return "failed", ""
    if rr.selected_plate is None:
        return "no_selection", ""
    vr = rr.plate_verdicts.get(rr.selected_plate)
    if vr is None:
        return "missing_verdict", ""
    verdict_value = vr.verdict.value
    if verdict_value not in include_verdicts:
        return "verdict_class", verdict_value
    if rr.is_fallback and not include_fallback:
        return "fallback", verdict_value
    return None


def _find_unresolved_wells(
    bad_barcodes: list[tuple[str, str]],
    addressing: PlateAddressing = DEFAULT_ADDRESSING,
) -> dict[str, object] | None:
    """Report rows whose ``custom_barcode`` could not be mapped to a well.

    A blank well silently shipped to JANUS is an unusable instruction, so the
    export refuses it (no silent fallback) and the preview surfaces it.

    The plate comes in the same way it does for the row builder that produced
    ``bad_barcodes``, so the range this message quotes is the range that
    rejected them rather than a second copy of the grid.

    The message also states where these rows sit, because rows are now ordered
    by source position and these have none. Landing at the top of the list is a
    consequence of that, not a ranking, and saying so is what keeps it from
    reading as one.
    """
    if not bad_barcodes:
        return None
    detail = ", ".join(f"{mid} (custom_barcode={cb!r})" for mid, cb in bad_barcodes)
    return {
        "code": "unresolved_well",
        "severity": SEVERITY_ERROR,
        "message": (
            "Janus mapping: unparseable custom_barcode, well position unknown for "
            f"{len(bad_barcodes)} mutant(s): {detail}. "
            f"Expected '<row>_<col>' with row 1-{addressing.rows} "
            f"and col 1-{addressing.cols}. "
            "Rows are ordered by source well, so these are listed first, "
            "having no position to be ordered by."
        ),
        "mutant_ids": [mid for mid, _ in bad_barcodes],
    }


def _find_plate_overflow(
    rows: list[dict[str, object]],
    addressing: PlateAddressing = DEFAULT_ADDRESSING,
) -> dict[str, object] | None:
    """Report selections larger than one destination plate.

    Evaluated before compact reassignment so the message names the real problem
    rather than surfacing ``seq_to_well``'s internal range error.
    """
    capacity = addressing.capacity
    if len(rows) <= capacity:
        return None
    return {
        "code": "plate_capacity",
        "severity": SEVERITY_ERROR,
        "message": (
            f"Janus mapping: {len(rows)} picks exceed the {capacity}-well "
            f"destination plate capacity. Reduce the selection to at most "
            f"{capacity} clones."
        ),
        "mutant_ids": [str(row["name"]) for row in rows[capacity:]],
    }


def _find_duplicate_dests(rows: list[dict[str, object]]) -> dict[str, object] | None:
    """Report duplicate ``dest_well`` (JANUS would double-dispense one well).

    Duplicates arise in source layout when two plates carry a pick at the same
    position, which is normal in multi-plate runs, so the message names the
    compact layout as the way out.

    Blank destinations are skipped: they mean the well never resolved, which
    ``_find_unresolved_wells`` already reports. In the export path that check
    raises first, so skipping blanks here leaves export behaviour unchanged.
    """
    seen: dict[str, list[str]] = {}
    for row in rows:
        well = str(row["dest_well"])
        if not well:
            continue
        seen.setdefault(well, []).append(str(row["name"]))
    dups = {well: names for well, names in seen.items() if len(names) > 1}
    if not dups:
        return None
    detail = "; ".join(
        f"{well} <- {', '.join(names)}" for well, names in sorted(dups.items())
    )
    mutant_ids: list[str] = []
    for _, names in sorted(dups.items()):
        mutant_ids.extend(names)
    return {
        "code": "duplicate_dest_well",
        "severity": SEVERITY_ERROR,
        "message": (
            "Janus mapping: duplicate dest_well would dispense multiple clones "
            f"into the same well: {detail}. "
            "Use dest_layout='compact' to assign destinations sequentially."
        ),
        "mutant_ids": mutant_ids,
    }


def _raise_if(finding: dict[str, object] | None) -> None:
    """Fail-fast wrapper: turn a structured finding into the export ``ValueError``."""
    if finding is not None:
        raise ValueError(str(finding["message"]))


def _assemble_janus_rows(
    replicates: list[ReplicateResult],
    dest_layout: str,
    include_verdicts: tuple[str, ...] = DEFAULT_INCLUDE_VERDICTS,
    include_fallback: bool = False,
    addressing: PlateAddressing = DEFAULT_ADDRESSING,
) -> tuple[list[dict[str, object]], list[tuple[str, str]], list[dict[str, object]]]:
    """Build rows in plate-map order without validating them.

    Returns ``(rows, bad_barcodes, excluded)``. Destinations still mirror the
    source position; compact reassignment is applied by the caller so that both
    the export and the preview can decide what to do about capacity first.

    Rows come back in ``source_well`` order, column-major, with the reasoning at
    the sort call below.

    Every replicate that does not make the cut lands in *excluded* with the
    reason, its verdict class, and the plate it was selected from, so a retry
    plan can be built from the same call that produces the picks.
    """
    if dest_layout not in _DEST_LAYOUTS:
        raise ValueError(
            f"Invalid dest_layout {dest_layout!r}. Expected 'source' or 'compact'."
        )

    # ``(order key, row)`` pairs rather than a key on the row: the row dict is
    # written to the file verbatim by ``csv.DictWriter``, so it carries the five
    # columns of the schema and nothing else. The ordering value rides beside it.
    keyed: list[tuple[tuple[int, int, str], dict[str, object]]] = []
    bad_barcodes: list[tuple[str, str]] = []
    excluded: list[dict[str, object]] = []

    for rr in replicates:
        dropped = _exclusion_reason(rr, include_verdicts, include_fallback)
        if dropped is not None:
            reason, verdict_value = dropped
            excluded.append(
                {
                    "mutant_id": rr.mutant_id,
                    "reason": reason,
                    "verdict": verdict_value,
                    "selected_plate": (
                        nb_label(rr.selected_plate) if rr.selected_plate else ""
                    ),
                    "is_fallback": bool(rr.is_fallback),
                }
            )
            continue

        # _exclusion_reason already proved both are present.
        assert rr.selected_plate is not None
        vr = rr.plate_verdicts[rr.selected_plate]

        bc = vr.translated.barcode
        custom_barcode = bc.custom_barcode
        seq = addressing.token_to_seq(custom_barcode)
        if seq is None:
            well_label = ""
            bad_barcodes.append((rr.mutant_id, custom_barcode))
        else:
            well_label = addressing.seq_to_well(seq)

        # nb_label is what every other MAME export writes (the result workbook
        # included), so the two files a run produces name one plate one way.
        source_plate = nb_label(rr.selected_plate)
        # G6/A6: read_count preferred; fall back to file_size_kb proxy.
        rc = bc.read_count
        priority_score: float = float(rc) if rc is not None else round(bc.file_size_kb, 3)

        # 0 for a pick with no readable position, so it sorts before every
        # placed pick. ``PlateAddressing.sort_key`` makes the same choice for
        # the same reason: an unreadable token is not dropped and not hidden at
        # the bottom of a long list, it is put where the operator opens the
        # preview. It is never only sorted, either: ``_find_unresolved_wells``
        # names each one as an error, which withholds the exported file
        # entirely and lists the clone in the preview.
        order_seq = seq if seq is not None else 0

        keyed.append(
            (
                (order_seq, nb_order_key(source_plate), source_plate),
                {
                    "name": rr.mutant_id,
                    "source_plate": source_plate,
                    "source_well": well_label,
                    "dest_well": well_label,  # default = same position; user may override
                    "priority_score": priority_score,
                },
            )
        )

    # Plate-map order, column-major (A1, B1 ... H1, A2), not sequencing depth:
    # the operator fills the final plate by hand while reading the step 2.2
    # plate map, so this file has to run in the same direction as the plate in
    # front of them. Sorting by priority_score DESC (the §2.5 recommendation
    # this used to follow) put the deepest-sequenced clone in A1 wherever it
    # sat on the source plate, and every next well had to be looked up. The
    # score keeps its value and its column; it just no longer places anything.
    #
    # The traversal is not re-derived here: ``order_seq`` came from
    # ``addressing.token_to_seq`` above, so this order is plate_geometry's one
    # rule and cannot drift from the result table or from ``seq_to_well``.
    #
    # Ties are reachable only across plates, since one plate holds at most one
    # pick per position. They break by plate, in the ``(nb_order_key, label)``
    # order ``JanusSettings.resolve_deck`` numbers the stock plates in, so a
    # position held on two plates is poured Stock plate1 first. A same-plate
    # tie would need one barcode map to give two clones one well; if that ever
    # arrives, the stable sort leaves them in the order *replicates* came in.
    keyed.sort(key=lambda pair: pair[0])
    rows: list[dict[str, object]] = [row for _, row in keyed]
    return rows, bad_barcodes, excluded


def _apply_compact_layout(
    rows: list[dict[str, object]],
    addressing: PlateAddressing = DEFAULT_ADDRESSING,
) -> None:
    """Reassign ``dest_well`` sequentially from A1, in place.

    Rows arrive in source-plate order, so pouring them out in list order is what
    makes the destination plate read the way the source plate map does: the pick
    from the lowest source position takes A1, the next B1, and holes close.

    Only the first ``addressing.capacity`` rows get a destination;
    ``seq_to_well`` rejects anything past the plate. Overflow rows keep a blank
    destination and are reported by ``_find_plate_overflow``.

    The plate is passed in rather than read from anywhere: ``kuma_core`` does
    not import the sidecar, and a destination plate that depended on sidecar
    state would make the same picks land differently depending on which process
    wrote them.
    """
    capacity = addressing.capacity
    for idx, row in enumerate(rows):
        row["dest_well"] = addressing.seq_to_well(idx + 1) if idx < capacity else ""


@dataclass(frozen=True)
class JanusSettings:
    """One policy object shared by the export and the preview.

    Both paths resolve their behaviour from a single instance so the plate the
    operator approves in the preview is the plate the file describes. The
    handler builds one instance and passes it to both calls.

    Instrument fields (``volume``, ``sample_type``, ``source_racks``,
    ``dest_rack``) only affect the ``device`` schema. ``liquid_class`` affects
    no file at all: the sheet has no column for it and it is carried as a record
    of the run.

    Only ``volume`` needs an operator decision: how much of a cell stock is
    transferred is an experimental condition that cannot be derived from the
    run. The plate names are derived from the plates of the run
    (``resolve_deck``) and overridden by anything the operator enters.
    """

    dest_layout: str = DEST_LAYOUT_COMPACT
    include_verdicts: tuple[str, ...] = DEFAULT_INCLUDE_VERDICTS
    include_fallback: bool = False
    output_schema: str = SCHEMA_DEVICE
    volume: float = DEFAULT_VOLUME_UL
    sample_type: str = DEFAULT_SAMPLE_TYPE
    liquid_class: str = DEFAULT_LIQUID_CLASS
    #: Operator overrides only, plate label -> ``Asp. Rack`` plate name. Empty
    #: means "derive from the run". Step 3 collects none of these since the
    #: names are generated, so in practice this arrives empty.
    source_racks: tuple[tuple[str, str], ...] = ()
    #: ``None`` means the generated destination plate name.
    dest_rack: str | None = None

    def __post_init__(self) -> None:
        if self.dest_layout not in _DEST_LAYOUTS:
            raise ValueError(
                f"Invalid dest_layout {self.dest_layout!r}. "
                "Expected 'source' or 'compact'."
            )
        if self.output_schema not in _SCHEMAS:
            raise ValueError(
                f"Invalid output_schema {self.output_schema!r}. "
                f"Expected one of {list(_SCHEMAS)}."
            )
        if self.output_schema == SCHEMA_DEVICE and not self.volume > 0:
            raise ValueError(
                f"Invalid volume {self.volume!r}. Expected a positive number of µL."
            )
        if self.output_schema == SCHEMA_DEVICE:
            for label, plate_name in self.source_racks:
                _require_plate_name(
                    plate_name, f"source plate name for plate {label!r}"
                )
            if self.dest_rack is not None:
                _require_plate_name(self.dest_rack, "dest_rack")
        object.__setattr__(
            self, "include_verdicts", normalize_include_verdicts(self.include_verdicts)
        )

    @property
    def rack_map(self) -> dict[str, str]:
        """The operator's overrides alone, echoed back to the dialog unchanged."""
        return dict(self.source_racks)

    def resolve_deck(self, plate_labels: "Iterable[str]") -> tuple[dict[str, str], str]:
        """Plate names for a run whose picks come from *plate_labels*.

        The ordering rule is the one this method has always used: plates sort by
        natural plate order (``nb_order_key``, so NB07 < NB08 < NB10) and take
        the positions 1, 2, 3 ... in that order. Only what the position is
        called has changed. The first plate of the run is "Stock plate1", the
        second "Stock plate2", and everything is dispensed into one
        "final culture plate", which is what the seeding workbook writes and
        what the instrument matches labware by.

        The position number is a rank within this run, not a plate number: a run
        of NB07 and NB10 gives "Stock plate1" and "Stock plate2".

        Anything the operator entered wins: ``source_racks`` overrides the
        generated name for the plates it names, and an explicit ``dest_rack``
        replaces the generated destination name.
        """
        ordered = sorted({p for p in plate_labels if p}, key=lambda p: (nb_order_key(p), p))
        names = {label: f"{SOURCE_PLATE_PREFIX}{idx}" for idx, label in enumerate(ordered, start=1)}
        names.update(self.rack_map)
        dest = self.dest_rack if self.dest_rack is not None else DEST_PLATE_NAME
        return names, dest

    @property
    def header(self) -> list[str]:
        """Column names of the file this policy writes."""
        if self.output_schema == SCHEMA_DEVICE:
            return list(_JANUS_DEVICE_HEADER)
        return list(_JANUS_HEADER)

    def to_payload(self) -> dict[str, object]:
        """JSON-safe view for the preview payload, so the dialog can show it."""
        return {
            "dest_layout": self.dest_layout,
            "include_verdicts": list(self.include_verdicts),
            "include_fallback": self.include_fallback,
            "output_schema": self.output_schema,
            "volume": self.volume,
            "sample_type": self.sample_type,
            "liquid_class": self.liquid_class,
            "source_racks": self.rack_map,
            "dest_rack": self.dest_rack,
            "columns": self.header,
        }


def _resolve_settings(
    settings: JanusSettings | None,
    dest_layout: str | None,
) -> JanusSettings:
    """Single resolution point for the policy.

    *dest_layout* is a convenience override kept for the many callers that only
    care about the layout; when given it replaces the field on *settings*.
    """
    resolved = settings if settings is not None else JanusSettings()
    if dest_layout is not None and dest_layout != resolved.dest_layout:
        resolved = replace(resolved, dest_layout=dest_layout)
    return resolved


def _find_derived_source_racks(
    rows: list[dict[str, object]],
    settings: JanusSettings,
) -> dict[str, object] | None:
    """Report which plates got a plate name generated rather than entered.

    Not an error: the names follow the seeding workbook (see
    ``JanusSettings.resolve_deck``). It is reported so the operator can see
    which names came from the run rather than from the plates in the room, and
    can check that the labware on the deck is labelled to match.
    """
    if settings.output_schema != SCHEMA_DEVICE:
        return None
    rack_map, dest_rack = settings.resolve_deck(
        str(row["source_plate"]) for row in rows
    )
    entered = settings.rack_map
    derived = {
        plate: name for plate, name in sorted(rack_map.items()) if plate not in entered
    }
    if not derived and settings.dest_rack is not None:
        return None
    detail = ", ".join(f"{plate} -> {name}" for plate, name in derived.items())
    if settings.dest_rack is None:
        detail = f"{detail}, destination -> {dest_rack}" if detail else (
            f"destination -> {dest_rack}"
        )
    return {
        "code": "derived_source_rack",
        "severity": SEVERITY_WARNING,
        "message": (
            "Janus mapping: plate names generated from the plates of this run "
            f"({detail}). Source plates are named in plate order and everything "
            "is dispensed into one culture plate, as in the seeding workbook. "
            "Label the labware on the deck to match."
        ),
        "mutant_ids": [],
    }


def project_device_rows(
    rows: list[dict[str, object]],
    settings: JanusSettings,
) -> list[list[object]]:
    """Project canonical rows onto the instrument-native column layout.

    Positional lists, not dicts: the canonical row supplies three of these cells
    (name and the two wells) and the policy and the deck supply the rest, so
    there is no one mapping to key by. ``no`` is the 1-based position in the
    already sorted row list, so the sheet counts off the transfers in the order
    the plate is filled.

    Plate names come from ``resolve_deck``, so every plate of the run carries
    one. A plate that somehow still has none writes an empty cell rather than
    raising: a blank the operator can see beats a file that never arrives.

    Nothing is written for the liquid class. The sheet has no column for it, and
    the file format is followed exactly.
    """
    rack_map, dest_rack = settings.resolve_deck(
        str(row["source_plate"]) for row in rows
    )
    projected: list[list[object]] = []
    for idx, row in enumerate(rows, start=1):
        plate = str(row["source_plate"])
        projected.append(
            [
                row["name"],
                settings.sample_type,
                idx,
                rack_map.get(plate, ""),
                row["source_well"],
                dest_rack,
                row["dest_well"],
                settings.volume,
            ]
        )
    return projected


def _collect_janus_rows(
    replicates: list[ReplicateResult],
    settings: JanusSettings,
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    """Build rows, exclusions, and findings for *settings*.

    The single collector behind both ``_build_janus_rows`` (raises) and
    ``build_janus_preview_rows`` (reports), so the two cannot drift.

    Returns ``(rows, excluded, findings)`` with destinations already assigned
    for the chosen layout.
    """
    rows, bad_barcodes, excluded = _assemble_janus_rows(
        replicates,
        settings.dest_layout,
        include_verdicts=settings.include_verdicts,
        include_fallback=settings.include_fallback,
    )

    findings: list[dict[str, object]] = []
    # Empty-well and capacity checks run before compact reassignment so the
    # capacity error names the picks instead of tripping seq_to_well's range.
    for finding in (
        _find_unresolved_wells(bad_barcodes),
        _find_plate_overflow(rows),
    ):
        if finding is not None:
            findings.append(finding)

    if settings.dest_layout == DEST_LAYOUT_COMPACT:
        _apply_compact_layout(rows)

    # Evaluated after compaction: compact layout is the documented way out of a
    # duplicate, so reporting a pre-compaction duplicate would be misleading.
    for finding in (
        _find_duplicate_dests(rows),
        _find_derived_source_racks(rows, settings),
    ):
        if finding is not None:
            findings.append(finding)

    return rows, excluded, findings


def _build_janus_rows(
    replicates: list[ReplicateResult],
    dest_layout: str | None = None,
    settings: JanusSettings | None = None,
) -> list[dict[str, object]]:
    """Build sorted Janus mapping rows from replicate results.

    Only clones whose selected plate carries an included verdict class survive;
    the default is PASS alone, and fallback picks are dropped unless
    ``include_fallback`` is set. Rows are sorted by ``source_well``,
    column-major (see the module docstring for why, and for the tiebreak).

    ``settings.dest_layout`` controls ``dest_well`` assignment:

    - ``"compact"`` (default): destinations are assigned sequentially from A1 in
      that same source order, following the column-major ``seq_to_well``
      convention (A1, B1, ... H1, A2, ...), so the destination plate reads the
      way the source plate map does with the holes closed. A stock plate is a
      new plate, so filling it from the front is the normal case.
    - ``"source"``: ``dest_well`` mirrors ``source_well``.

    Raises ``ValueError`` on empty wells, >96 rows, or duplicate destinations.
    A generated plate name is a warning, not an error: the file ships and names
    what it generated.
    """
    resolved = _resolve_settings(settings, dest_layout)
    rows, _, findings = _collect_janus_rows(replicates, resolved)
    for finding in findings:
        if finding.get("severity") == SEVERITY_ERROR:
            _raise_if(finding)
    return rows


def build_janus_preview_rows(
    replicates: list[ReplicateResult],
    dest_layout: str | None = None,
    settings: JanusSettings | None = None,
) -> dict[str, object]:
    """Build Janus mapping rows in-memory, collecting problems instead of raising.

    Shares ``_collect_janus_rows`` with ``_build_janus_rows`` so the preview
    shows exactly what the export would write for the same settings, but the
    guards come back as structured entries. Showing every problem at once is the
    point: the export path keeps its fail-fast behaviour.

    Returns ``{"rows", "errors", "warnings", "row_count", "excluded",
    "excluded_count", "settings"}``. Each entry of ``errors`` / ``warnings`` is
    ``{"code", "severity", "message", "mutant_ids"}``; each excluded entry is
    ``{"mutant_id", "reason", "verdict", "selected_plate", "is_fallback"}``.

    ``errors`` alone blocks the export. ``warnings`` names what was generated
    (the plate names), which the operator has to see but which never withholds a
    file.

    Rows with an unresolved well are kept with blank ``source_well`` and
    ``dest_well`` so the broken clone stays visible in the preview.
    """
    resolved = _resolve_settings(settings, dest_layout)
    rows, excluded, findings = _collect_janus_rows(replicates, resolved)
    rack_map, dest_rack = resolved.resolve_deck(
        str(row["source_plate"]) for row in rows
    )

    payload = resolved.to_payload()
    # The deck the file will carry, next to the operator's own entries: the
    # dialog shows these, and comparing the echoed entries with what it holds is
    # how it knows the preview is current.
    payload["resolved_source_racks"] = rack_map
    payload["resolved_dest_rack"] = dest_rack

    return {
        "rows": rows,
        "errors": [f for f in findings if f.get("severity") != SEVERITY_WARNING],
        "warnings": [f for f in findings if f.get("severity") == SEVERITY_WARNING],
        "row_count": len(rows),
        "excluded": excluded,
        "excluded_count": len(excluded),
        "settings": payload,
    }


def _meta_comment_line(meta: "NgsRunMeta") -> str:
    """Build a single-line CSV comment from *meta* (G3 spec).

    Format: ``# kuma_run_meta: flow_cell=X, kit=Y, started=Z``
    Fields that are ``None`` are omitted from the comment.
    """
    parts: list[str] = []
    if meta.flow_cell_id:
        parts.append(f"flow_cell={meta.flow_cell_id}")
    if meta.kit:
        parts.append(f"kit={meta.kit}")
    if meta.started:
        parts.append(f"started={meta.started}")
    if meta.instrument:
        parts.append(f"instrument={meta.instrument}")
    if meta.position:
        parts.append(f"position={meta.position}")
    return "# kuma_run_meta: " + ", ".join(parts)


def _write_janus_kuma_meta_sheet(
    wb: "object",
    meta: "NgsRunMeta | None",
    kuma_version: str,
) -> None:
    """Append ``__kuma_meta__`` sheet to an openpyxl Workbook.

    Mirrors the logic in excel_writer._write_kuma_meta_sheet but is a
    standalone helper to avoid a circular import between the two modules.
    """
    from openpyxl.styles import Font as _Font  # local import keeps cold-start fast

    ws = wb.create_sheet("__kuma_meta__")  # type: ignore[union-attr]
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40

    ws.append(["key", "value"])
    for cell in ws[1]:
        cell.font = _Font(bold=True)

    ws.append(["kuma_version", kuma_version])
    ws.append([
        "generated_at",
        datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    ])

    if meta is None:
        ws.append(["ngs_run_meta", "(not found — no MinKNOW run folder detected)"])
        return

    fields: list[tuple[str, object]] = [
        ("instrument", meta.instrument),
        ("position", meta.position),
        ("flow_cell_id", meta.flow_cell_id),
        ("sample_id", meta.sample_id),
        ("kit", meta.kit),
        ("started", meta.started),
        ("basecalling_enabled", (
            None if meta.basecalling_enabled is None
            else ("true" if meta.basecalling_enabled else "false")
        )),
        ("raw_run_dir", meta.raw_run_dir),
    ]
    for key, value in fields:
        ws.append([key, "" if value is None else value])


def export_mame_janus_csv(
    replicates: list[ReplicateResult],
    output_path: Path,
    ngs_run_meta: "NgsRunMeta | None" = None,
    dest_layout: str | None = None,
    settings: JanusSettings | None = None,
) -> Path:
    """Export final cell-stock Janus mapping as CSV.

    Two headers are available through ``settings.output_schema``:

    - ``"device"`` (default) writes the instrument-native worksheet columns
      ``name | type | no | Asp. Rack | Asp. Posi | Dsp. Rack | Dsp. Posi |
      volume``. The two rack cells carry plate names, and there is no liquid
      class column.
    - ``"legacy5"`` writes ``name | source_plate | source_well | dest_well |
      priority_score``, the kuma-internal column set.

    Sorted by ``source_well``, column-major, so the file reads in the direction
    the operator fills the plate. Only clones with an included verdict class
    (PASS by default) are written.

    The return value stays ``Path``: exclusions are reported by
    ``build_janus_preview_rows`` for the same ``settings``, which the RPC
    handler calls alongside this function.

    G3: when *ngs_run_meta* is not ``None``, a ``# kuma_run_meta: ...`` comment
    line is prepended before the header row.  When *ngs_run_meta* is ``None``
    no comment is written (backward-compatible with existing consumers).

    Phase 1: priority_score = file_size_kb proxy.
    G6/A6 round: replace with BarcodeRecord.read_count when available.

    *dest_layout* overrides ``settings.dest_layout``: ``"compact"``
    (destinations assigned sequentially from A1 in sorted order, default) or
    ``"source"`` (dest mirrors source position).
    Raises ``ValueError`` on unresolved wells, >96 picks, or duplicate dests.
    Generated plate names do not withhold the file; they come back as warnings
    from ``build_janus_preview_rows``.
    """
    resolved = _resolve_settings(settings, dest_layout)
    rows = _build_janus_rows(replicates, settings=resolved)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as fh:
        if ngs_run_meta is not None:
            fh.write(_meta_comment_line(ngs_run_meta) + "\n")
        if resolved.output_schema == SCHEMA_DEVICE:
            # Positional writer: the canonical row dict does not carry the
            # instrument cells, so ``project_device_rows`` builds them in header
            # order and there is no fieldname mapping to hand a DictWriter.
            writer = csv.writer(fh)
            writer.writerow(_JANUS_DEVICE_HEADER)
            writer.writerows(project_device_rows(rows, resolved))
        else:
            dict_writer = csv.DictWriter(fh, fieldnames=_JANUS_HEADER)
            dict_writer.writeheader()
            dict_writer.writerows(rows)

    return output_path


def export_mame_janus_xlsx(
    replicates: list[ReplicateResult],
    output_path: Path,
    ngs_run_meta: "NgsRunMeta | None" = None,
    kuma_version: str = "",
    dest_layout: str | None = None,
    settings: JanusSettings | None = None,
) -> Path:
    """Export final cell-stock Janus mapping as XLSX.

    Same data and same ``settings.output_schema`` choice as the CSV variant.
    Provides header bold-styling and column freeze for readability.

    The worksheet keeps the kuma sheet name ``Janus Mapping``; only the header
    row follows the instrument workbook.

    G3: when *ngs_run_meta* is provided, a ``__kuma_meta__`` sheet is appended.
    The sheet is always written (placeholder when meta is ``None``).

    Phase 1: priority_score = file_size_kb proxy.

    *dest_layout* behaves exactly as in ``export_mame_janus_csv``.
    """
    import openpyxl
    from openpyxl.styles import Font

    resolved = _resolve_settings(settings, dest_layout)
    rows = _build_janus_rows(replicates, settings=resolved)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Janus Mapping")
    else:
        ws.title = "Janus Mapping"

    ws.append(resolved.header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    if resolved.output_schema == SCHEMA_DEVICE:
        for device_row in project_device_rows(rows, resolved):
            ws.append(device_row)
    else:
        for row in rows:
            ws.append(
                [
                    row["name"],
                    row["source_plate"],
                    row["source_well"],
                    row["dest_well"],
                    row["priority_score"],
                ]
            )

    # G3: always append __kuma_meta__ sheet.
    _write_janus_kuma_meta_sheet(wb, ngs_run_meta, kuma_version)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


__all__ = [
    "JanusSettings",
    "SEVERITY_ERROR",
    "SEVERITY_WARNING",
    "build_janus_preview_rows",
    "export_mame_janus_csv",
    "export_mame_janus_xlsx",
    "normalize_include_verdicts",
    "project_device_rows",
]
