"""Excel export: per-native-barcode sheets + Final 96-well matrix.

G6/A6 read_count policy
------------------------
``BarcodeRecord.read_count`` is populated by the consensus parser from
``depth=N`` header metadata when available, falling back to single-record
counts for legacy consensus files.  The "NGS Results" sheet ``<NB>_reads``
columns (header names built dynamically from ``nb_label``) carry read_count
verbatim; a missing read_count (``None``) is written as a blank cell — the
legacy file_size_kb volume proxy has been removed so a blank never masks a
real depth shortage.
The per-plate ``NB`` sheets retain ``file_size_kb`` in their header (Sheet1
format) for backward compatibility; only the unified "NGS Results" sheet uses
the ``reads`` naming.

A11 / G3 __kuma_meta__ sheet
------------------------------
``write_excel`` appends a ``__kuma_meta__`` sheet with MinKNOW run metadata
when ``ngs_run_meta`` is supplied.  The sheet contains key/value rows for
instrument, flow_cell_id, sample_id, kit, started, position,
basecalling_enabled, and raw_run_dir.  A ``kuma_version`` row is always
written so consumers can identify the generating software version.
When ``ngs_run_meta`` is ``None`` a single placeholder row is written to
keep the sheet structure consistent.
"""

from __future__ import annotations

import datetime
from collections.abc import Iterable
from pathlib import Path
from typing import TYPE_CHECKING, Literal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook

from kuma_core.mame.export.well_mapper import WellMapper, seq_to_well
from kuma_core.mame.models import ReplicateResult, VerdictClass, VerdictRecord
from kuma_core.mame.detected import compute_recovery, replicate_is_recovered
from kuma_core.mame.export.nb_label import nb_label, nb_order_key, well_sort_key

if TYPE_CHECKING:
    from kuma_core.mame.ingest.run_meta import NgsRunMeta

# Confirmed color map (040 AC-09).
VERDICT_FILL: dict[VerdictClass, str] = {
    VerdictClass.PASS: "C6EFCE",       # soft green
    VerdictClass.AMBIGUOUS: "FFF2CC",  # soft amber
    VerdictClass.MIXED: "FCE4D6",      # soft peach
    VerdictClass.FRAMESHIFT: "F8CBCB", # soft red
    VerdictClass.MANY: "F8CBCB",
    VerdictClass.WRONG_AA: "F8CBCB",
    VerdictClass.LOWDEPTH: "ECEEF1",   # light gray
    VerdictClass.NO_CALL: "DFE3E8",    # gray
}

FAILED_FILL = "F4B6B6"            # soft red (clear, not garish)
WELL_HIGHLIGHT_PURPLE = "44506A"  # Final sheet coord column — refined slate
SELECTED_PLATE_YELLOW = "FFF3B0"  # Final sheet chosen plate — soft highlight

_SHEET1_HEADER = [
    "well_id",
    "file_size_kb",
    "read_count",
    "input_reads",
    "aligned_reads",
    "mapq_failed",
    "span_failed",
    "low_depth_positions",
    "consensus_n_fraction",
    "low_quality_bases",
    "mixed_positions",
    "max_minor_allele_fraction",
    "custom_barcode",
    "observed_mutations",
    "observed_aa",
    "verdict",
    "verdict_notes",
    "selected",
    "is_fallback",
    "fallback_reason",
]

_FINAL_HEADER = [
    "well_id",
    "selected_plate",
    "custom_barcode",
    "mutant_id",
    "verdict",
    "is_fallback",
    "fallback_reason",
    "notes",
]

# Reference-format "NGS Results" / "Final (matrix)" sheets adapt to the actual
# set of native barcodes present in a run (any count / naming). Reads columns
# carry BarcodeRecord.read_count verbatim (blank when None); the legacy
# file_size_kb proxy substitution has been removed.


def _run_native_barcodes(
    verdict_records: list[VerdictRecord],
    replicate_results: list[ReplicateResult],
) -> list[str]:
    """Distinct native barcodes present in this run, naturally sorted.

    Sources: every verdict record's ``native_barcode`` plus every plate key
    appearing in a replicate result's ``plate_verdicts``. Replaces the legacy
    fixed ``_KNOWN_PLATES`` triple so exports adapt to any NB count/naming.
    """
    nbs: set[str] = {vr.translated.barcode.native_barcode for vr in verdict_records}
    for rr in replicate_results:
        nbs.update(rr.plate_verdicts.keys())
    return sorted(nbs, key=nb_order_key)


def _ngs_header(nbs: list[str]) -> list[str]:
    header = ["index", "mutant", "well", "selected_NB", "custom_barcode"]
    for nb in nbs:
        label = nb_label(nb)
        header.extend([f"{label}_detected", f"{label}_reads", f"{label}_quality"])
    header.append("recovered")
    return header


def _matrix_header(nbs: list[str]) -> list[str]:
    return ["index", "mutant", "well", "selected_NB", *[nb_label(nb) for nb in nbs]]


def _fill(color_hex: str) -> PatternFill:
    return PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

# ── Clean, professional sheet styling ───────────────────────────────────────
_HEADER_FILL = "1F2937"  # slate-800 header band
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ZEBRA_FILL = "F5F7FA"   # very light row stripe
_BORDER_SIDE = Side(style="thin", color="D9DEE6")
_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE
)


def _style_header(ws) -> None:
    """Apply the slate header band (white bold, centered) to row 1."""
    for cell in ws[1]:
        cell.fill = _fill(_HEADER_FILL)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 22


def _finalize(
    ws,
    *,
    freeze: str | None = "A2",
    autofit: bool = True,
    zebra: bool = False,
) -> None:
    """Apply thin borders, optional zebra striping, autofit, and a freeze pane."""
    from openpyxl.utils import get_column_letter

    max_row = ws.max_row
    max_col = ws.max_column
    if zebra:
        stripe = _fill(_ZEBRA_FILL)
        for r in range(3, max_row + 1, 2):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill is None or cell.fill.fill_type is None:
                    cell.fill = stripe
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = _BORDER
    if autofit:
        for c in range(1, max_col + 1):
            length = 0
            for r in range(1, max_row + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    length = max(length, len(str(v)))
            ws.column_dimensions[get_column_letter(c)].width = min(max(length + 2, 9), 46)
    if freeze:
        ws.freeze_panes = freeze


def _custom_barcode_to_seq(custom: str) -> int | None:
    """`{R}_{F}` -> 1-based column-major sequence index.

    Uses F as the well column (1..12) and R as the row contribution within the
    column. Returns None if the label cannot be parsed.
    """

    parts = custom.split("_")
    if len(parts) != 2:
        return None
    try:
        r = int(parts[0])
        f = int(parts[1])
    except ValueError:
        return None
    if not (1 <= r <= 8 and 1 <= f <= 12):
        return None
    return (f - 1) * 8 + r


def _write_sheet1(
    wb: Workbook,
    native_barcode: str,
    records: Iterable[VerdictRecord],
    replicate_results: Iterable[ReplicateResult] = (),
) -> None:
    ws = wb.create_sheet(nb_label(native_barcode))
    ws.append(_SHEET1_HEADER)
    _style_header(ws)

    # Map each final selection (selected plate + chosen well) to its replicate so
    # the per-NB sheet can flag which well was the final pick and whether that
    # pick was a non-PASS fallback. Keyed by (native_barcode, custom_barcode).
    selected_lookup: dict[tuple[str, str], ReplicateResult] = {}
    for rr in replicate_results:
        if rr.selected_plate is None or rr.failed:
            continue
        sel_vr = rr.plate_verdicts.get(rr.selected_plate)
        if sel_vr is None:
            continue
        selected_lookup[
            (rr.selected_plate, sel_vr.translated.barcode.custom_barcode)
        ] = rr

    for vr in sorted(
        records,
        key=lambda r: well_sort_key(r.translated.barcode.custom_barcode),
    ):
        br = vr.translated.barcode
        seq = _custom_barcode_to_seq(br.custom_barcode)
        well_id = seq_to_well(seq) if seq else ""
        sel_rr = selected_lookup.get((native_barcode, br.custom_barcode))
        selected_marker = "Y" if sel_rr is not None else ""
        is_fallback = "Y" if sel_rr is not None and sel_rr.is_fallback else ""
        fallback_reason = (
            sel_rr.fallback_reason or ""
            if sel_rr is not None and sel_rr.is_fallback
            else ""
        )
        row = [
            well_id,
            round(br.file_size_kb, 3),
            br.read_count,
            br.n_input_reads,
            br.n_aligned_reads,
            br.n_mapq_failed,
            br.n_span_failed,
            br.n_low_depth_positions,
            round(br.consensus_n_fraction, 4),
            br.n_low_quality_bases,
            br.n_mixed_positions,
            round(br.max_minor_allele_fraction, 4),
            br.custom_barcode,
            ", ".join(vr.translated.observed_nt_changes),
            ", ".join(vr.translated.observed_aa_changes),
            vr.verdict.value,
            vr.verdict_notes,
            selected_marker,
            is_fallback,
            fallback_reason,
        ]
        ws.append(row)
        fill = _fill(VERDICT_FILL[vr.verdict])
        for cell in ws[ws.max_row]:
            cell.fill = fill
    _finalize(ws)


def _write_final(
    wb: Workbook,
    replicate_results: Iterable[ReplicateResult],
    mapper: WellMapper,
) -> None:
    ws = wb.create_sheet("Final")
    ws.append(_FINAL_HEADER)
    _style_header(ws)

    # Index replicate results by the chosen custom_barcode's well when selectable.
    replicate_by_seq: dict[int, ReplicateResult] = {}
    failed_wells: list[str] = []
    unassigned: list[ReplicateResult] = []

    for rr in replicate_results:
        seq: int | None = None
        if rr.selected_plate is not None:
            vr = rr.plate_verdicts.get(rr.selected_plate)
            if vr is not None:
                seq = _custom_barcode_to_seq(vr.translated.barcode.custom_barcode)
        else:
            # Failed — try to borrow a well from any available verdict to place
            # a FAILED marker.
            for vr in rr.plate_verdicts.values():
                s = _custom_barcode_to_seq(vr.translated.barcode.custom_barcode)
                if s is not None:
                    seq = s
                    break
        if seq is not None and seq not in replicate_by_seq:
            replicate_by_seq[seq] = rr
        else:
            unassigned.append(rr)

    confirmed = 0
    failed_count = 0
    redo_targets: list[str] = []

    for seq in range(1, 97):
        well = mapper.seq_to_well(seq)
        rr = replicate_by_seq.get(seq)
        if rr is None:
            ws.append([well, "", "", "", "", "", "", ""])
            _highlight_well_cell(ws, ws.max_row)
            continue

        if rr.failed or rr.selected_plate is None:
            failed_wells.append(well)
            failed_count += 1
            redo_targets.append(rr.mutant_id)
            ws.append([well, "FAILED", "-", rr.mutant_id, "-", "", "", ""])
            for cell in ws[ws.max_row]:
                cell.fill = _fill(FAILED_FILL)
        else:
            vr = rr.plate_verdicts[rr.selected_plate]
            ws.append(
                [
                    well,
                    nb_label(rr.selected_plate),
                    vr.translated.barcode.custom_barcode,
                    rr.mutant_id,
                    vr.verdict.value,
                    "Y" if rr.is_fallback else "",
                    (rr.fallback_reason or "") if rr.is_fallback else "",
                    vr.verdict_notes,
                ]
            )
            row_idx = ws.max_row
            # Selected-plate yellow highlight per 050 spec.
            ws.cell(row=row_idx, column=2).fill = _fill(SELECTED_PLATE_YELLOW)
            # Verdict-class fill on the verdict column for visual continuity.
            ws.cell(row=row_idx, column=5).fill = _fill(VERDICT_FILL[vr.verdict])
            confirmed += 1
        _highlight_well_cell(ws, ws.max_row)

    # Summary footer.
    ws.append([])
    summary = (
        f"confirmed: {confirmed}/96 | FAILED: {failed_count} | "
        f"REDO targets: {', '.join(redo_targets) if redo_targets else '(none)'}"
    )
    ws.append([summary])
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="left")

    # Any replicate results that could not be assigned (extra wells) go in a
    # secondary block so they are not silently dropped.
    if unassigned:
        ws.append([])
        ws.append(["unassigned replicate results:"])
        for rr in unassigned:
            ws.append(
                [
                    "",
                    nb_label(rr.selected_plate) if rr.selected_plate else "FAILED",
                    "",
                    rr.mutant_id,
                    rr.selection_reason,
                    "",
                    "",
                    "",
                ]
            )
    _finalize(ws)


def _highlight_well_cell(ws, row_idx: int) -> None:
    ws.cell(row=row_idx, column=1).fill = _fill(WELL_HIGHLIGHT_PURPLE)
    ws.cell(row=row_idx, column=1).font = Font(color="FFFFFF", bold=True)


# ---------------------------------------------------------------------------
# Reference-format sheets (G7 + G2 implementation).
# ---------------------------------------------------------------------------


def _representative_custom_barcode(
    rr: ReplicateResult,
    nbs: list[str],
) -> str:
    """Custom barcode of the row's representative well.

    Prefers the selected plate's chosen well; otherwise falls back to the first
    available plate verdict (in natural NB order). Empty when no verdict carries
    identity. Used to keep "NGS Results" / "Final (matrix)" rows in natural well
    order (AC-2.1).
    """
    ref_vr: VerdictRecord | None = None
    if rr.selected_plate is not None:
        ref_vr = rr.plate_verdicts.get(rr.selected_plate)
    if ref_vr is None:
        for plate in nbs:
            ref_vr = rr.plate_verdicts.get(plate)
            if ref_vr is not None:
                break
    return ref_vr.translated.barcode.custom_barcode if ref_vr is not None else ""


def _build_unified_ngs_data(
    replicate_results: list[ReplicateResult],
    nbs: list[str],
) -> list[dict]:
    """Build per-mutant rows for the "NGS Results" sheet.

    Each row holds a mutant_id key, the selected NB label, and per-plate
    detected/reads/quality values keyed by NB label. Rows are sorted by each
    mutant's representative well (selected plate's chosen well, else first
    available plate verdict) in natural well order (AC-2.1); ``index`` is
    re-assigned 1..N after sorting.

    Reads carry ``BarcodeRecord.read_count`` verbatim (blank when ``None``); the
    legacy ``file_size_kb`` proxy substitution has been removed.
    """
    rows: list[dict] = []
    ordered = sorted(
        replicate_results,
        key=lambda rr: well_sort_key(_representative_custom_barcode(rr, nbs)),
    )
    for idx, rr in enumerate(ordered, start=1):
        # Determine the well from the selected plate if available, otherwise
        # from any available plate verdict.
        ref_vr: VerdictRecord | None = None
        if rr.selected_plate is not None:
            ref_vr = rr.plate_verdicts.get(rr.selected_plate)
        if ref_vr is None:
            for plate in nbs:
                ref_vr = rr.plate_verdicts.get(plate)
                if ref_vr is not None:
                    break

        seq: int | None = None
        cb: str = ""
        if ref_vr is not None:
            cb = ref_vr.translated.barcode.custom_barcode
            seq = _custom_barcode_to_seq(cb)

        well_id = seq_to_well(seq) if seq is not None and 1 <= seq <= 96 else ""

        selected_nb = (
            nb_label(rr.selected_plate)
            if rr.selected_plate and not rr.failed
            else ""
        )

        row: dict = {
            "index": idx,
            "mutant": rr.mutant_id,
            "well": well_id,
            "selected_NB": selected_nb,
            "custom_barcode": cb,
        }

        for plate in nbs:
            label = nb_label(plate)
            vr = rr.plate_verdicts.get(plate)
            if vr is not None:
                detected = ", ".join(vr.translated.observed_aa_changes) or vr.verdict.value
                bc = vr.translated.barcode
                # read_count verbatim; blank when absent (no file_size_kb proxy).
                reads_val: int | str = bc.read_count if bc.read_count is not None else ""
                quality = (
                    f"N={bc.consensus_n_fraction:.3f}; "
                    f"low_depth={bc.n_low_depth_positions}; "
                    f"low_q={bc.n_low_quality_bases}; "
                    f"mix={bc.n_mixed_positions}/"
                    f"{bc.max_minor_allele_fraction:.3f}; "
                    f"drop_mapq={bc.n_mapq_failed}; "
                    f"drop_span={bc.n_span_failed}"
                )
            else:
                detected = ""
                reads_val = ""
                quality = ""
            row[f"{label}_detected"] = detected
            row[f"{label}_reads"] = reads_val
            row[f"{label}_quality"] = quality

        # 재현(recovered): OR across this mutant's plate verdicts (PASS/AMBIGUOUS).
        row["recovered"] = "Y" if replicate_is_recovered(rr) else "N"
        rows.append(row)

    return rows


def _write_unified_ngs_sheet(
    wb: Workbook,
    replicate_results: list[ReplicateResult],
    nbs: list[str],
    designed_mutant_ids: frozenset[str] | None = None,
) -> None:
    """Write the reference-format "NGS Results" sheet (G7 spec)."""
    ws = wb.create_sheet("NGS Results")
    header = _ngs_header(nbs)
    ws.append(header)
    _style_header(ws)

    rows = _build_unified_ngs_data(replicate_results, nbs)
    for row in rows:
        ws.append([row.get(col, "") for col in header])

    # Recovery (재현율) summary area below the per-mutant rows.
    recovery = compute_recovery(replicate_results, designed_mutant_ids)
    ws.append([])
    if recovery is None:
        ws.append(["Recovery (재현율)", "n/a"])
    else:
        ws.append([
            "Recovery (재현율)",
            f"{recovery.recovery_rate * 100:.1f}%",
        ])
        ws.append(["recovered_mutants", recovery.recovered_mutants])
        ws.append(["total_mutants", recovery.total_mutants])

    # Freeze top header row for readability.
    _finalize(ws, zebra=True)


def _write_final_matrix_sheet(
    wb: Workbook,
    replicate_results: list[ReplicateResult],
    nbs: list[str],
) -> None:
    """Write the reference-format binary selection matrix (G2 spec).

    Sheet name is "Final (matrix)" to avoid collision with the legacy "Final"
    sheet (which is preserved for backward compatibility).

    Cell values: "O" if that plate's verdict is PASS for the mutant, blank
    otherwise. The single final selection (selected plate with a PASS verdict,
    not failed) is bolded — at most one bold "O" per mutant. A non-PASS
    fallback selection produces no matrix "O" (surfaced only via selected_NB).
    """
    ws = wb.create_sheet("Final (matrix)")
    ws.append(_matrix_header(nbs))
    _style_header(ws)

    ordered = sorted(
        replicate_results,
        key=lambda rr: well_sort_key(_representative_custom_barcode(rr, nbs)),
    )
    for idx, rr in enumerate(ordered, start=1):
        # Determine well from selected plate or any available verdict.
        ref_vr: VerdictRecord | None = None
        if rr.selected_plate is not None:
            ref_vr = rr.plate_verdicts.get(rr.selected_plate)
        if ref_vr is None:
            for plate in nbs:
                ref_vr = rr.plate_verdicts.get(plate)
                if ref_vr is not None:
                    break

        seq: int | None = None
        if ref_vr is not None:
            seq = _custom_barcode_to_seq(ref_vr.translated.barcode.custom_barcode)

        well_id = seq_to_well(seq) if seq is not None and 1 <= seq <= 96 else ""

        selected_nb = (
            nb_label(rr.selected_plate)
            if rr.selected_plate and not rr.failed
            else ""
        )

        # The single final selection is the bold "O" — PASS verdict only.
        bold_plate: str | None = None
        if rr.selected_plate is not None and not rr.failed:
            sel_vr = rr.plate_verdicts.get(rr.selected_plate)
            if sel_vr is not None and sel_vr.verdict == VerdictClass.PASS:
                bold_plate = rr.selected_plate

        cells: list[str] = []
        for plate in nbs:
            vr = rr.plate_verdicts.get(plate)
            cells.append("O" if vr is not None and vr.verdict == VerdictClass.PASS else "")

        ws.append([idx, rr.mutant_id, well_id, selected_nb, *cells])

        # Bold + highlight the single final-selection cell.
        if bold_plate is not None and bold_plate in nbs:
            col_idx = nbs.index(bold_plate)
            # +5: 1-based, 4 leading cols (index, mutant, well, selected_NB) + 1
            cell = ws.cell(row=ws.max_row, column=5 + col_idx)
            cell.font = Font(bold=True)
            cell.fill = _fill(SELECTED_PLATE_YELLOW)

    _finalize(ws, zebra=True)


# ---------------------------------------------------------------------------
# __kuma_meta__ sheet (A11 / G3).
# ---------------------------------------------------------------------------


def _write_kuma_meta_sheet(
    wb: Workbook,
    meta: "NgsRunMeta | None",
    kuma_version: str,
) -> None:
    """Append a ``__kuma_meta__`` sheet to *wb*.

    Row format: col-A = key, col-B = value.
    When *meta* is ``None``, only the ``kuma_version`` and a ``ngs_run_meta``
    placeholder row are written so the sheet is always present.
    """
    ws = wb.create_sheet("__kuma_meta__")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40

    # Header row.
    ws.append(["key", "value"])
    _style_header(ws)

    # Always include software version and generation timestamp.
    ws.append(["kuma_version", kuma_version])
    ws.append([
        "generated_at",
        datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    ])

    if meta is None:
        ws.append(["ngs_run_meta", "(not found — no MinKNOW run folder detected)"])
        _finalize(ws, freeze=None, autofit=False)
        return

    # MinKNOW run fields.
    fields: list[tuple[str, object]] = [
        ("instrument", meta.instrument),
        ("position", meta.position),
        ("flow_cell_id", meta.flow_cell_id),
        ("sample_id", meta.sample_id),
        ("kit", meta.kit),
        ("started", meta.started),
        ("basecalling_enabled", (
            None if meta.basecalling_enabled is None
            else ("true" if meta.basecalling_enabled else "false")
        )),
        ("raw_run_dir", meta.raw_run_dir),
    ]
    for key, value in fields:
        ws.append([key, "" if value is None else value])
    _finalize(ws, freeze=None, autofit=False)


def write_excel(
    verdict_records: list[VerdictRecord],
    replicate_results: list[ReplicateResult],
    output_path: Path,
    mapper: WellMapper | None = None,
    mode: Literal["amplicon", "plasmid"] = "amplicon",  # reserved for Phase 2
    ngs_run_meta: "NgsRunMeta | None" = None,
    kuma_version: str = "",
    designed_mutant_ids: frozenset[str] | None = None,
) -> Path:
    """Write the combined Excel report to ``output_path``. Returns the path.

    Parameters
    ----------
    ngs_run_meta:
        MinKNOW run metadata discovered by ``discover_run_meta``.  When
        ``None`` the ``__kuma_meta__`` sheet is still written with a
        placeholder row so the sheet always exists (A11).
    kuma_version:
        Software version string injected into the ``__kuma_meta__`` sheet.
    designed_mutant_ids:
        Distinct designed ``mutant_id`` set used to compute the recovery
        (재현율) summary in the "NGS Results" sheet.  ``None`` → recovery is
        rendered as ``n/a`` (designed set unavailable).
    """

    del mode  # Phase 1: mode does not alter Excel layout.
    mapper = mapper or WellMapper()

    wb = openpyxl.Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    by_nb: dict[str, list[VerdictRecord]] = {}
    for vr in verdict_records:
        by_nb.setdefault(vr.translated.barcode.native_barcode, []).append(vr)

    for nb in sorted(by_nb, key=nb_order_key):
        _write_sheet1(wb, nb, by_nb[nb], replicate_results)

    _write_final(wb, replicate_results, mapper)

    # Reference-format sheets (G7 + G2): appended after legacy sheets so that
    # existing consumers of the per-NB sheets and "Final" are not disturbed.
    nbs = _run_native_barcodes(verdict_records, replicate_results)
    _write_unified_ngs_sheet(wb, replicate_results, nbs, designed_mutant_ids)
    _write_final_matrix_sheet(wb, replicate_results, nbs)

    # A11 / G3: MinKNOW run metadata sheet — always present, content optional.
    _write_kuma_meta_sheet(wb, ngs_run_meta, kuma_version)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
