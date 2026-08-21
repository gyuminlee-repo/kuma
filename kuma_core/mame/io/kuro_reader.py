"""KURO xlsx `expected_mutations` sheet adapter (Blocker B resolved)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from kuma_core.mame.models import WT_LABELS, ExpectedMutation

_EXPECTED_SHEET = "expected_mutations"
_EXPECTED_HEADER = [
    "mutant_id",
    "position",
    "wt_aa",
    "mt_aa",
    "wt_codon",
    "mt_codon",
    "group_id",
    "primer_set_ref",
    "notation_type",
    "status",
]
_DESIGNED_STATUSES = {
    "DESIGNED",
    "SAME_POSITION",
    "DIFF_POSITION",
    "AUTO_SUGGESTION",
    "AUTO_SUGGESTION_L1",
    "AUTO_SUGGESTION_L2",
    "AUTO_SUGGESTION_L3",
    "AUTO_SUGGESTION_L4",
    "POOL_CASCADE",
    "AUTO_RELAX",
}


@dataclass(frozen=True)
class KuroReadResult:
    """The designed rows of a KURO export, plus the rows the filter removed.

    Row numbers travel with both lists because a dropped row is not a neutral
    omission: MAME reads row order as plate order, so every row removed between
    two kept rows moves each later mutant one well up. Nothing downstream can
    see that shift once the rows are gone, which is why the caller is handed the
    row numbers rather than a count.
    """

    #: Designed rows, in sheet order.
    expected: list[ExpectedMutation]
    #: 1-based sheet row number of each element of :attr:`expected`, same order.
    row_numbers: list[int] = field(default_factory=list)
    #: ``(row number, status)`` for rows the status filter removed.
    dropped_rows: list[tuple[int, str]] = field(default_factory=list)


def read_expected_mutations(path: Path) -> list[ExpectedMutation]:
    """Read the designed rows of the `expected_mutations` sheet.

    Designed rows and the wild-type control row are returned; FAILED rows are
    Phase 2. Interim KURO exports that stored rescue stage names in status are
    accepted as designed. The control row is kept regardless of its status
    (KURO writes it as ``control``) because it occupies a well, and a plate
    occupant removed mid-list moves every later mutant one well up.
    Raises ValueError if the expected sheet is missing (old KURO version).

    Thin wrapper over :func:`read_expected_mutations_with_rows` for callers that
    only need the mutations. A caller that places wells wants the other reader:
    the rows this one drops shift every later well.
    """
    return read_expected_mutations_with_rows(path).expected


def read_expected_mutations_with_rows(path: Path) -> KuroReadResult:
    """Read the `expected_mutations` sheet, reporting what the filter removed.

    The filter reads status alone: a row whose status is outside
    ``_DESIGNED_STATUSES`` is not returned, the sole exception being the
    wild-type control row, which is a plate occupant rather than a design and is
    identified by its ``mutant_id``. The filter is also no longer silent about
    what it removed: a dropped row moves every later mutant one well up, and the
    result still reads like a full plate, so the row numbers travel with the
    result and let a caller refuse the file instead of scoring the shifted plate.
    """

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if _EXPECTED_SHEET not in wb.sheetnames:
            raise ValueError(
                f"'{path}' cannot be read as an expected-variant list. MAME accepts "
                "either a KURO export (a workbook carrying an "
                f"'{_EXPECTED_SHEET}' sheet) or a plain variant list: one variant "
                "per row under a single variant column, in plate order. This file "
                "is neither."
            )
        ws = wb[_EXPECTED_SHEET]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError(f"'{_EXPECTED_SHEET}' sheet in '{path}' is empty.")
        header_list = [str(c) if c is not None else "" for c in header]
        expected = [h.lower() for h in _EXPECTED_HEADER]
        got = [h.strip().lower() for h in header_list]
        if got[: len(expected)] != expected:
            raise ValueError(
                f"'{_EXPECTED_SHEET}' header mismatch. Expected {expected}, got {got}."
            )

        results: list[ExpectedMutation] = []
        row_numbers: list[int] = []
        dropped_rows: list[tuple[int, str]] = []
        # Row 1 is the header, so data rows are numbered from 2 and the number
        # matches what the operator sees in Excel.
        for row_number, raw in enumerate(rows_iter, start=2):
            if raw is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
                continue
            cells = list(raw) + [None] * (len(_EXPECTED_HEADER) - len(raw))
            status = _s(cells[9])
            mutant_id = _s(cells[0])
            # A wild-type control row is kept whatever its status says. Both
            # shipped workbooks carry it as `status=control` (`templates/` and
            # `src-tauri/samples/mame/`), which is not a designed status, so
            # the filter used to remove it -- and a removed row is not a neutral
            # omission here: it moves every later mutant one well up, and the
            # caller refuses the whole file over it. The row is a plate
            # occupant, not a design, so membership is decided by the label.
            # Status stays the only test for every other row, which is why this
            # is keyed on `mutant_id` rather than by widening the status set.
            is_wt = mutant_id.lower() in WT_LABELS
            if not is_wt and status.upper() not in _DESIGNED_STATUSES:
                dropped_rows.append((row_number, status))
                continue
            row_numbers.append(row_number)
            results.append(
                ExpectedMutation(
                    mutant_id=mutant_id,
                    position=_position(
                        cells[1],
                        row_number=row_number,
                        mutant_id=mutant_id,
                        is_wt=is_wt,
                        path=path,
                    ),
                    wt_aa=_s(cells[2]),
                    mt_aa=_s(cells[3]),
                    wt_codon=_s(cells[4]),
                    mt_codon=_s(cells[5]),
                    group_id=_s(cells[6]),
                    primer_set_ref=_s(cells[7]),
                    notation_type=_s(cells[8]),
                    status=status,
                )
            )
        return KuroReadResult(
            expected=results,
            row_numbers=row_numbers,
            dropped_rows=dropped_rows,
        )
    finally:
        wb.close()


def expected_to_labels(expected: list[ExpectedMutation]) -> list[str]:
    """Produce the human-readable mutation label list consumed by compare.verdict."""

    return [f"{m.wt_aa}{m.position}{m.mt_aa}" for m in expected]


def _s(cell: object) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


#: A position cell written as text: a whole number, optionally with a zero
#: fraction, which is how a spreadsheet sometimes spells an integer ("232.0").
_POSITION_TEXT = re.compile(r"^[+-]?\d+(?:\.0*)?$")


def _position(
    cell: object,
    *,
    row_number: int,
    mutant_id: str,
    is_wt: bool,
    path: Path,
) -> int:
    """The `position` column of one row, or a refusal naming the cell.

    A position that could not be read used to come back as 0, and 0 is silent:
    the row still became an ``ExpectedMutation``, ``expected_to_labels`` wrote it
    as ``V0F``, and no observed mutation ever carries position 0, so the row
    scored as a mismatch that reads like a failed clone rather than like a
    workbook this reader could not parse. The operator saw a wrong answer where
    there was an unreadable cell.

    Two cases keep the lenient 0 rather than raising, and both are real files
    rather than tolerances:

    * The wild-type control row. Both shipped workbooks
      (``templates/03_mame_expected_mutations.xlsx`` and the copy under
      ``src-tauri/samples/mame/``) end with
      ``('WT', 0, '-', '-', '-', '-', 'G0', '-', 'wt', 'control')``. Position 0
      is what the control *is*, so ``position <= 0`` is not the test. The test is
      whether the cell parses at all.
    * A blank cell. Blank is the shape a column the exporter did not fill takes,
      so refusing it would refuse exports this reader accepts today. Only a cell
      holding something unreadable is refused.

    Excel hands numeric cells over as ``int`` or ``float``, so ``232.0`` is the
    same position as ``232`` and is read as one. A non-integral number is not a
    residue index and is refused with everything else.
    """
    if cell is None:
        return 0
    if isinstance(cell, bool):
        # openpyxl gives a TRUE/FALSE cell as bool, and bool is an int subclass,
        # so this is checked before the numeric branch would silently read 1.
        return _refuse_position(cell, row_number, mutant_id, is_wt, path)
    if isinstance(cell, (int, float)):
        if float(cell).is_integer():
            return int(cell)
        return _refuse_position(cell, row_number, mutant_id, is_wt, path)
    text = str(cell).strip()
    if not text:
        return 0
    if _POSITION_TEXT.match(text):
        return int(float(text))
    return _refuse_position(cell, row_number, mutant_id, is_wt, path)


def _refuse_position(
    cell: object,
    row_number: int,
    mutant_id: str,
    is_wt: bool,
    path: Path,
) -> int:
    """Raise for a designed row; keep 0 for the control row.

    The control carries no residue and its label is never compared, so an
    unreadable position on it costs nothing downstream. A designed row is the
    opposite: its position is the whole of what is being scored.
    """
    if is_wt:
        return 0
    raise ValueError(
        f"'{_EXPECTED_SHEET}' row {row_number} of '{path}' has an unreadable "
        f"position: {cell!r}"
        + (f" (mutant_id '{mutant_id}')" if mutant_id else "")
        + ". The position column holds the residue index as a whole number; a "
        "blank cell is accepted, anything else is not. Fix the cell and export "
        "again."
    )
