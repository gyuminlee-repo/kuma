"""Read __kuma_meta__ sheet from Kuro-exported xlsx."""
from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional


@dataclass
class KumaMeta:
    project_id: str
    kuma_version: str
    kuro_module_version: str
    exported_at: str

    def to_dict(self) -> dict:
        return asdict(self)


def read_kuma_meta(xlsx_path: Path | str) -> Optional[KumaMeta]:
    """Return KumaMeta if __kuma_meta__ sheet present, else None."""
    import openpyxl

    path = Path(xlsx_path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "__kuma_meta__" not in wb.sheetnames:
            return None
        sheet = wb["__kuma_meta__"]
        kv: dict[str, str] = {}
        for row in sheet.iter_rows(max_col=2, values_only=True):
            if row and row[0]:
                kv[str(row[0])] = "" if row[1] is None else str(row[1])
        if "project_id" not in kv:
            return None
        return KumaMeta(
            project_id=kv["project_id"],
            kuma_version=kv.get("kuma_version", ""),
            kuro_module_version=kv.get("kuro_module_version", ""),
            exported_at=kv.get("exported_at", ""),
        )
    finally:
        wb.close()
