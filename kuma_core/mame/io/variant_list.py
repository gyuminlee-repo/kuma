"""Generic variant-list adapter for the MAME expected-variant input.

MAME needs one thing from this input: which variant is meant to be in each well,
in plate order. Until now the only accepted shape was a KURO results xlsx with an
``expected_mutations`` sheet whose first ten columns matched exactly, which left
anyone running MAME without KURO unable to supply a list they already had.

This module makes the file shape a detail instead of a requirement, the same way
the KURO input step stopped distinguishing an "EVOLVEpro CSV" from an "Others"
file and simply reads whichever sheet and column the user points at. A KURO
export keeps working untouched: it is recognised and routed to the strict reader,
so its ``status`` filter and rescue-stage handling are unchanged.

Recognition is a default, not a verdict. Naming a sheet overrides it, because one
workbook can carry the strict sheet next to the sheet that describes the plate on
the bench, and the two can disagree on both membership and order. Preferring the
strict sheet in that case places every well from a list nobody chose, and the
result reads like a finished plate rather than a mis-set one.

A file can state where its samples sit, and when it does that statement wins.
A ``Well`` column makes the address data, so the reader computes nothing: the
occupant of ``D7`` is whatever the row for ``D7`` says, the control is wherever
the file put it, and a plate with no control row has no control well. This is
the shape that lets a partly filled plate be described honestly, because a row
naming a well and no variant is a well this campaign did not use.

Without a ``Well`` column, row order is plate order. ``build_draft_layout``
assigns well *i* from occupant *i*, so a plain list needs no extra ordering
rule, and the two consequences of that rule hold: a wild-type row occupies a
well (its ordinal travels with the read instead of being discarded), and a row
that is read and not placed is refused by row number rather than absorbed,
because absorbing it moves every later mutant one well up and leaves the result
looking like a correct plate. An ordinal is not an address, though, so where the
control lands is a decision rather than a reading: ``layout.WtPlacement`` names
the choices and defaults to the last well of the plate.

The five fields MAME never reads (``wt_codon``, ``mt_codon``, ``group_id``,
``primer_set_ref``, ``notation_type``) are left empty rather than invented.
"""

from __future__ import annotations

import csv
from collections.abc import Sequence
from dataclasses import dataclass, field
from pathlib import Path

from kuma_core.kuro.mutation import parse_mutation_notation
from kuma_core.mame.io.kuro_reader import read_expected_mutations_with_rows
from kuma_core.mame.models import WT_LABELS, ExpectedMutation
from kuma_core.mame.plate_geometry import canonical_well

#: Sheet that marks a file as a KURO export. Routed to the strict reader.
KURO_SHEET = "expected_mutations"

#: Header names accepted as the well column. Matched case-insensitively after
#: stripping, like the variant candidates.
#:
#: Deliberately short. ``position`` is NOT a candidate and must not become one:
#: in a KURO sheet that header is the residue number of a substitution, so
#: accepting it would read a column of integers as plate coordinates and place
#: the whole plate somewhere else.
_WELL_HEADER_CANDIDATES = (
    "well",
    "wells",
    "well_id",
)

#: Header names accepted as the variant column when the caller does not name one.
#: Matched case-insensitively after stripping.
_VARIANT_HEADER_CANDIDATES = (
    "variant",
    "variants",
    "mutant",
    "mutants",
    "mutation",
    "mutations",
    "mutant_id",
    "variant_id",
)

#: Labels that denote the wild-type control rather than a mutant. Defined in
#: ``mame.models`` so ``io/kuro_reader`` can share it without importing this
#: module (the dependency already runs the other way). Kept as a module name
#: because existing references read it here.
_WT_LABELS = WT_LABELS

#: Reported as ``variant_column`` when the file carries no header row. Stating
#: the absence is the point: the alternative is naming the first variant, which
#: is exactly the value a headerless file would otherwise have eaten.
_HEADERLESS_COLUMN = "(no header)"


@dataclass(frozen=True)
class DroppedRow:
    """A row the reader did not place, and why.

    Not a diagnostic afterthought. Row order is plate order, so a row removed
    ahead of the last variant moves every later mutant one well up, and nothing
    downstream can see that it happened. The row number is what lets a caller
    point at the file instead of scoring the shifted plate.
    """

    #: 1-based sheet row number, matching what the operator sees in Excel.
    row: int
    #: ``"blank"`` (empty variant cell), ``"short"`` (row ends before the
    #: variant column), or ``"status"`` (KURO status outside the designed set).
    reason: str
    #: Extra context for the message. The status text, for ``"status"`` rows.
    detail: str = ""


@dataclass(frozen=True)
class VariantListReadResult:
    """Parsed variant list plus what the caller needs to know about it."""

    #: Plate-ordered mutants. Excludes any explicit WT row.
    expected: list[ExpectedMutation]
    #: 1-based placement ordinal of the wild-type well among plate occupants,
    #: or ``None`` when the source listed no WT row of its own. Counting the WT
    #: row rather than discarding it is what keeps the wells after it in place:
    #: dropping it shifted every later mutant one well up and reported nothing.
    #: ``build_draft_layout`` appends WT after the last mutant for ``None``.
    wt_ordinal: int | None
    #: Sheet the values were read from. ``None`` for CSV.
    sheet: str | None
    #: Header of the column the values were read from, for reporting back.
    variant_column: str
    #: Rows that were read but not placed. Empty on every result
    #: :func:`read_variant_source` returns: a non-empty list is what it refuses
    #: on, and the refusal message is built from these entries.
    dropped_rows: list[DroppedRow] = field(default_factory=list)
    #: Well the source named for each entry of :attr:`expected`, same order and
    #: same length, or ``None`` when the source carried no ``Well`` column and
    #: said nothing about wells.
    #:
    #: The distinction is the point: ``None`` means the placement has to be
    #: computed from row order, and a list means it must not be. A caller that
    #: re-orders or filters :attr:`expected` without carrying this alongside
    #: re-seats the plate, which is why ``build_draft_layout`` refuses a pair
    #: whose lengths disagree.
    wells: list[str] | None = None
    #: Well the source named for the wild-type control, or ``None`` when it
    #: named none. Only a ``Well``-column source fills this.
    #:
    #: ``None`` here is a statement, not a gap: on this path a plate without a
    #: control row simply has no control well, and nothing invents one. That is
    #: what a ``Well`` column buys, and it is why :attr:`wt_ordinal` stays
    #: ``None`` on this path even for a file that names a control.
    wt_well: str | None = None

    @property
    def has_explicit_wt(self) -> bool:
        """True when the source listed a wild-type row itself.

        Derived rather than stored so it cannot disagree with the two fields it
        is a statement about. Kept because it is a cross-layer symbol.

        Both are consulted, because the two paths record the same fact in
        different places: a row-order source states an ordinal, a
        ``Well``-column source states a well. Reading only the ordinal made this
        answer ``False`` for a file that named its control in ``D7``, which is
        the opposite of what the file says.
        """
        return self.wt_ordinal is not None or self.wt_well is not None


@dataclass(frozen=True)
class VariantSourceInfo:
    """What a file offers, so a caller can present sheet and column choices."""

    #: True when this is a KURO export and needs no column mapping.
    is_kuro_export: bool
    #: Sheet names in workbook order. Empty for CSV.
    sheets: list[str]
    #: ``{sheet: [header, ...]}``. CSV uses the single key ``""``.
    headers: dict[str, list[str]]
    #: Column this module would pick on its own, when it can pick one.
    suggested_column: str | None


def _is_csv(path: Path) -> bool:
    return path.suffix.lower() in {".csv", ".tsv", ".txt"}


def _csv_rows(path: Path) -> list[list[str]]:
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with open(path, newline="", encoding="utf-8-sig") as handle:
        return [list(row) for row in csv.reader(handle, delimiter=delimiter)]


def _looks_like_a_variant(text: str) -> bool:
    """True when *text* reads as a mutation label or the wild-type control.

    Case is ignored on both tests. ``parse_mutation_notation`` itself accepts
    only upper case, but this function decides whether a row is data rather
    than whether it is valid, and ``s65t`` is unmistakably an attempt at a
    variant. Reading it as a column name instead loses it silently, which is
    the failure this check exists to prevent; recognising it hands the row to
    the parser, which refuses it by number and says why.
    """
    if not text:
        return False
    if text.lower() in _WT_LABELS:
        return True
    try:
        parse_mutation_notation(text.upper())
    except ValueError:
        return False
    return True


def _is_headerless(headers: list[str]) -> bool:
    """True when the first row holds a variant rather than a column name.

    A one-column list written without a header puts its first variant where a
    header belongs. Read as a header, that variant is consumed and every later
    one moves a well up, which is the same silent shift this module refuses
    everywhere else, so it is detected instead of accepted.
    """
    non_empty = [h for h in headers if h]
    return len(non_empty) == 1 and _looks_like_a_variant(non_empty[0])


def _suggest_column(headers: list[str]) -> str | None:
    """Pick the variant column from headers, or None when it is ambiguous."""
    normalised = [(h or "").strip() for h in headers]
    for candidate in _VARIANT_HEADER_CANDIDATES:
        for header in normalised:
            if header.lower() == candidate:
                return header
    non_empty = [h for h in normalised if h]
    if len(non_empty) == 1:
        # A lone cell that is itself a variant is data, not a column name.
        # Offering it as a column would invite the caller to name it and lose
        # it; the headerless branch of the reader handles this file instead.
        return None if _looks_like_a_variant(non_empty[0]) else non_empty[0]
    return None


def _resolve_well_index(headers: list[str]) -> int | None:
    """Index of the ``Well`` column, or ``None`` when the file has none.

    Never guessed. A column is the well column because its header says so, and
    a file without one is read by row order exactly as before. Guessing here
    would put a plate somewhere the file never named.
    """
    normalised = [(h or "").strip().lower() for h in headers]
    for candidate in _WELL_HEADER_CANDIDATES:
        for index, header in enumerate(normalised):
            if header == candidate:
                return index
    return None


def inspect_variant_source(path: Path) -> VariantSourceInfo:
    """Report the sheets and headers a file offers, without reading its rows.

    Lets a caller show the same sheet/column pickers the KURO input step uses
    instead of failing on a file whose layout is merely unfamiliar.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"variant list not found: {path}")

    if _is_csv(path):
        rows = _csv_rows(path)
        csv_headers = [str(c).strip() for c in rows[0]] if rows else []
        return VariantSourceInfo(
            is_kuro_export=False,
            sheets=[],
            headers={"": csv_headers},
            suggested_column=_suggest_column(csv_headers),
        )

    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = list(workbook.sheetnames)
        is_kuro = KURO_SHEET in sheets
        sheet_headers: dict[str, list[str]] = {}
        for name in sheets:
            first = next(workbook[name].iter_rows(values_only=True), None)
            sheet_headers[name] = [
                "" if cell is None else str(cell).strip() for cell in (first or ())
            ]
        # Headers are reported for every sheet even on a KURO export. A workbook can
        # carry the strict sheet next to the sheet that actually describes the plate
        # in front of the operator, and without headers there is nothing to pick from.
        first_sheet = KURO_SHEET if is_kuro else (sheets[0] if sheets else "")
        return VariantSourceInfo(
            is_kuro_export=is_kuro,
            sheets=sheets,
            headers=sheet_headers,
            suggested_column=(
                None if is_kuro else _suggest_column(sheet_headers.get(first_sheet, []))
            ),
        )
    finally:
        workbook.close()


def _rows_from_xlsx(path: Path, sheet: str | None) -> tuple[str, list[list[object]]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is not None and sheet not in workbook.sheetnames:
            raise ValueError(
                f"sheet '{sheet}' not in {path.name}. "
                f"Available: {', '.join(workbook.sheetnames)}"
            )
        name = sheet if sheet is not None else workbook.sheetnames[0]
        rows: list[list[object]] = [
            [cell for cell in row]
            for row in workbook[name].iter_rows(values_only=True)
        ]
        return name, rows
    finally:
        workbook.close()


def _resolve_column_index(headers: list[str], variant_column: str | None) -> int:
    normalised = [(h or "").strip() for h in headers]
    if variant_column is not None:
        wanted = variant_column.strip().lower()
        for index, header in enumerate(normalised):
            if header.lower() == wanted:
                return index
        raise ValueError(
            f"column '{variant_column}' not found. "
            f"Available: {', '.join(h for h in normalised if h) or '(none)'}"
        )
    suggested = _suggest_column(normalised)
    if suggested is None:
        raise ValueError(
            "cannot tell which column holds the variants; name one explicitly. "
            f"Available: {', '.join(h for h in normalised if h) or '(none)'}"
        )
    return normalised.index(suggested)


def _variant_shaped_header_error(path: Path, header: str) -> ValueError:
    """A named column whose header is itself a variant, in a multi-column file.

    The single-column case is decided rather than refused: one non-empty cell
    is the whole first row, so a variant there is unambiguously data. A file
    with more columns says nothing of the kind. The other cells of that row may
    be headers (``plate1``) or values, and reading the row either way is a
    guess: as a header it eats the first variant and moves every later one a
    well up, as data it invents a column name for the rest. So the file is
    refused and the operator settles it.
    """
    return ValueError(
        f"the column named '{header}' in {path.name} is itself a variant, so "
        "the first row of this file holds data rather than column names. The "
        "file has more than one column and nothing in it says whether the "
        "other cells of that row are headers or values, and MAME reads row "
        "order as plate order, so guessing wrong loses a well silently. Add a "
        "header row that names each column and read again."
    )


def _unreadable_well_error(path: Path, row: int, raw: str) -> ValueError:
    """A Well cell that is not a well on this plate.

    Not repaired and not skipped. The whole reason to read a ``Well`` column is
    that the file, rather than the reader, decides where an occupant sits, so a
    coordinate the plate does not have leaves that occupant with no place and
    no way to invent one. ``I13`` and ``A0`` are off an 8x12 plate; a plate of
    another size is a different instrument, not a wider grid.
    """
    return ValueError(
        f"{path.name} row {row}: '{raw}' is not a well on a 96-well plate. "
        "Wells run A1 to H12 (a row letter A-H and a column number 1-12, "
        "zero-padding optional). Correct the cell and read again."
    )


def _duplicate_well_error(
    path: Path, well: str, first_row: int, second_row: int
) -> ValueError:
    """One well named twice is two occupants for one position.

    Which of the two is actually in the well is not something the file says, and
    either answer scores one variant against the other one reads. ``A1`` and
    ``A01`` are the same well, so they collide here too.
    """
    return ValueError(
        f"{path.name} names well {well} twice (rows {first_row} and "
        f"{second_row}). One well holds one sample, and the file does not say "
        "which of the two is in it. Give one of them a different well and read "
        "again."
    )


def _missing_well_error(path: Path, row: int, label: str) -> ValueError:
    """A variant with no well, in a file whose wells are the placement.

    Filling one in is the guess this column exists to remove: a file that
    states wells states all of them, and the only address left over is one the
    operator has not decided yet.
    """
    return ValueError(
        f"{path.name} row {row}: '{label}' has no well. This file has a Well "
        "column, so the well is where each variant sits and nothing else in "
        "the file says where this one goes. Fill the well in, or delete the "
        "row, and read again."
    )


def _duplicate_wt_error(path: Path, first_row: int, second_row: int) -> ValueError:
    """Two wild-type rows in one list is a plate nobody can place.

    A plate carries exactly one WT well, so a second row is either a duplicate
    or a variant labelled like the control. Both readings change which mutant
    every later well holds, and the file does not say which was meant.
    """
    return ValueError(
        f"{path.name} lists a wild-type row twice (rows {first_row} and "
        f"{second_row}). One plate carries exactly one WT well, so the file "
        "has to say which of the two it is. Delete the other row and read "
        "again."
    )


def _dropped_rows_error(path: Path, dropped: list[DroppedRow]) -> ValueError:
    """The refusal a set of unplaced rows produces.

    Row order is plate order, so a row that is read and not placed moves every
    mutant after it one well up. That shift is invisible in the result: each
    well still carries a plausible variant, and the workbook reads like a
    correct plate. So the rows are named and the read is refused.
    """
    wording = {
        "blank": "empty variant cell",
        "short": "row ends before the variant column",
        "status": "status outside the designed set",
    }
    listed = ", ".join(
        f"row {d.row} ({wording.get(d.reason, d.reason)}"
        + (f": {d.detail}" if d.detail else "")
        + ")"
        for d in dropped[:5]
    )
    if len(dropped) > 5:
        listed += f", and {len(dropped) - 5} more"
    return ValueError(
        f"{path.name} has {len(dropped)} row(s) that cannot be placed on the "
        f"plate: {listed}. MAME reads row order as plate order, so a row read "
        "and not placed moves every mutant after it one well up, and the "
        "result still looks like a full plate. Remove the rows or give them a "
        "variant, and read again."
    )


def _cell(row: Sequence[object], index: int) -> str:
    """The stripped text of *index* in *row*, or ``""`` when the row is shorter."""
    if index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def _read_stated_wells(
    path: Path,
    data_rows: Sequence[tuple[int, Sequence[object]]],
    variant_index: int,
    well_index: int,
    sheet_name: str | None,
    column_label: str,
) -> VariantListReadResult:
    """Read a file whose ``Well`` column is the placement.

    The column changes what a row means, so it changes which rows are faults.

    An empty variant cell is refused on the row-order path because row order is
    plate order there: the row is read, nothing is placed, and every mutant
    after it moves one well up while the result still reads as a full plate.
    Here nothing moves. A row that names a well and no variant says that well is
    not part of this campaign, which is a plate a partly filled campaign
    actually has and had no way to express before. So it is skipped, the well is
    still claimed (naming it twice is still a contradiction), and
    ``dropped_rows`` stays empty.

    The refusals are the ones a stated address can produce and a row ordinal
    cannot: a coordinate off the plate, one well named twice, and a variant with
    no well at all. Each carries the row number and the offending value, because
    the file is where the fix is.

    A wild-type row puts the control in the well it names, wherever that is. The
    control is not appended anywhere: a file with a ``Well`` column states which
    wells this plate has, and a plate without a control row simply has none.
    """
    expected: list[ExpectedMutation] = []
    wells: list[str] = []
    wt_well: str | None = None
    wt_row: int | None = None
    seen_wells: dict[str, int] = {}
    seen_variants: dict[str, int] = {}

    for offset, row in data_rows:
        raw_well = _cell(row, well_index)
        label = _cell(row, variant_index)
        if not raw_well and not label:
            # Neither column says anything: a trailing blank row, or an
            # openpyxl phantom from a sheet formatted wider than it is filled.
            continue
        if not raw_well:
            raise _missing_well_error(path, offset, label)
        try:
            well = canonical_well(raw_well)
        except (ValueError, IndexError) as exc:
            raise _unreadable_well_error(path, offset, raw_well) from exc
        if well in seen_wells:
            raise _duplicate_well_error(path, well, seen_wells[well], offset)
        seen_wells[well] = offset
        if not label:
            continue
        if label.lower() in _WT_LABELS:
            if wt_row is not None:
                raise _duplicate_wt_error(path, wt_row, offset)
            wt_well = well
            wt_row = offset
            continue
        if label in seen_variants:
            raise ValueError(
                f"duplicate variant '{label}' in {path.name} "
                f"(rows {seen_variants[label]} and {offset}). "
                "Each well needs a distinct variant to be scored."
            )
        seen_variants[label] = offset
        try:
            wt_aa, position, mt_aa = parse_mutation_notation(label)
        except ValueError as exc:
            raise ValueError(f"{path.name} row {offset}: {exc}") from exc
        expected.append(
            ExpectedMutation(
                mutant_id=label,
                position=position,
                wt_aa=wt_aa,
                mt_aa=mt_aa,
                wt_codon="",
                mt_codon="",
                group_id="",
                primer_set_ref="",
                notation_type="",
                status="DESIGNED",
            )
        )
        wells.append(well)

    if not expected:
        raise ValueError(
            f"no variants found in {path.name} "
            f"(column '{column_label}'). "
            "The file needs one variant per row below the header."
        )

    return VariantListReadResult(
        expected=expected,
        # The ordinal is the row-order path's answer to "where is the control",
        # and this path has a better one. Filling both in would leave two
        # statements about one well with nothing keeping them in step.
        wt_ordinal=None,
        sheet=sheet_name,
        variant_column=column_label,
        wells=wells,
        wt_well=wt_well,
    )


def _read_variant_source(
    path: Path,
    sheet: str | None,
    variant_column: str | None,
) -> VariantListReadResult:
    """Read *path* and report everything found, placing no judgement on it.

    Separated from :func:`read_variant_source` so the refusal is built from the
    very rows that were read, rather than from a second pass that could see a
    different file.
    """
    if not _is_csv(path):
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            is_kuro = KURO_SHEET in workbook.sheetnames
        finally:
            workbook.close()
        # An explicit sheet wins. A workbook can hold the strict sheet alongside the
        # sheet that describes the plate on the bench, and those two can disagree on
        # both membership and order; silently preferring the strict one then places
        # every well from a list the operator did not choose. Naming no sheet, or
        # naming the strict one, keeps the previous behaviour exactly.
        if is_kuro and sheet in (None, KURO_SHEET):
            return _read_kuro_export(path)

    if _is_csv(path):
        rows = _csv_rows(path)
        sheet_name = None
    else:
        sheet_name, rows = _rows_from_xlsx(path, sheet)

    if not rows:
        raise ValueError(f"{path.name} is empty")

    headers = ["" if c is None else str(c).strip() for c in rows[0]]
    # Whether the file has a header is decided before any column is resolved,
    # and without consulting *variant_column*. The UI hands back whatever
    # `inspect_variant_source` suggested, so honouring a named column first
    # would let a headerless file arrive as ``variant_column="S65T"`` and eat
    # its own first variant, which is the exact shift this module refuses.
    well_index: int | None = None
    if _is_headerless(headers):
        index = next(i for i, h in enumerate(headers) if h)
        # Data starts at the first row, so row numbers start at 1 and match
        # what the operator sees in Excel.
        data_rows = list(enumerate(rows, start=1))
        column_label = _HEADERLESS_COLUMN
    else:
        well_index = _resolve_well_index(headers)
        # The well column is taken out of the running for the variant column
        # before auto-detection, so a two-column ``Well, Sample`` file resolves
        # ``Sample`` instead of reporting two candidates and refusing. A caller
        # that named the column explicitly is left alone, and told below if what
        # it named is the well column.
        candidate_headers = list(headers)
        if well_index is not None and variant_column is None:
            candidate_headers[well_index] = ""
        index = _resolve_column_index(candidate_headers, variant_column)
        if well_index is not None and index == well_index:
            raise ValueError(
                f"column '{headers[index]}' in {path.name} is the well column, "
                "which says where each sample sits rather than what it is. "
                "Name the column holding the variants."
            )
        # Past the single-column case, a header that is itself a variant is not
        # something to resolve either way: the row is refused rather than
        # guessed at. See `_variant_shaped_header_error`.
        if _looks_like_a_variant(headers[index]):
            raise _variant_shaped_header_error(path, headers[index])
        data_rows = list(enumerate(rows[1:], start=2))
        column_label = headers[index] or str(index)

    if well_index is not None:
        return _read_stated_wells(
            path, data_rows, index, well_index, sheet_name, column_label
        )

    expected: list[ExpectedMutation] = []
    dropped: list[DroppedRow] = []
    wt_ordinal: int | None = None
    wt_row: int | None = None
    # Occupants placed so far, WT included. This is the plate ordinal, which is
    # why the WT row increments it instead of being skipped.
    placed = 0
    last_value_row = 0
    seen: dict[str, int] = {}
    for offset, row in data_rows:
        if index >= len(row):
            dropped.append(DroppedRow(row=offset, reason="short"))
            continue
        cell = row[index]
        label = "" if cell is None else str(cell).strip()
        if not label:
            dropped.append(DroppedRow(row=offset, reason="blank"))
            continue
        last_value_row = offset
        placed += 1
        if label.lower() in _WT_LABELS:
            if wt_row is not None:
                raise _duplicate_wt_error(path, wt_row, offset)
            wt_ordinal = placed
            wt_row = offset
            continue
        if label in seen:
            raise ValueError(
                f"duplicate variant '{label}' in {path.name} "
                f"(rows {seen[label]} and {offset}). "
                "Each well needs a distinct variant to be scored."
            )
        seen[label] = offset
        try:
            wt_aa, position, mt_aa = parse_mutation_notation(label)
        except ValueError as exc:
            raise ValueError(f"{path.name} row {offset}: {exc}") from exc
        expected.append(
            ExpectedMutation(
                mutant_id=label,
                position=position,
                wt_aa=wt_aa,
                mt_aa=mt_aa,
                wt_codon="",
                mt_codon="",
                group_id="",
                primer_set_ref="",
                notation_type="",
                status="DESIGNED",
            )
        )

    if not expected:
        raise ValueError(
            f"no variants found in {path.name} "
            f"(column '{column_label}'). "
            "The file needs one variant per row below the header."
        )

    # Rows after the last value are openpyxl phantoms (a sheet remembers the
    # extent it was once formatted to) and a CSV trailing newline. They shift
    # nothing, so they are not reported.
    return VariantListReadResult(
        expected=expected,
        wt_ordinal=wt_ordinal,
        sheet=sheet_name,
        variant_column=column_label,
        dropped_rows=[d for d in dropped if d.row < last_value_row],
    )


def _read_kuro_export(path: Path) -> VariantListReadResult:
    """Read a KURO export under the same WT and dropped-row rules as a plain list.

    The two branches used to differ: this one hard-coded ``has_explicit_wt`` to
    false, so a KURO sheet carrying its own WT row got a second control well
    appended, and the status filter removed rows without saying so, so the
    wells after each removed row were scored one place off.
    """
    read = read_expected_mutations_with_rows(path)
    expected: list[ExpectedMutation] = []
    wt_ordinal: int | None = None
    wt_row: int | None = None
    placed = 0
    for row_number, mutation in zip(read.row_numbers, read.expected):
        placed += 1
        if mutation.mutant_id.strip().lower() in _WT_LABELS:
            if wt_row is not None:
                raise _duplicate_wt_error(path, wt_row, row_number)
            wt_ordinal = placed
            wt_row = row_number
            continue
        expected.append(mutation)
    return VariantListReadResult(
        expected=expected,
        wt_ordinal=wt_ordinal,
        sheet=KURO_SHEET,
        variant_column="mutant_id",
        dropped_rows=[
            DroppedRow(row=row_number, reason="status", detail=status or "(blank)")
            for row_number, status in read.dropped_rows
        ],
    )


def read_variant_source(
    path: Path,
    sheet: str | None = None,
    variant_column: str | None = None,
) -> VariantListReadResult:
    """Read a variant list from *path*, whatever shape it comes in.

    A KURO export (a workbook carrying an ``expected_mutations`` sheet) is routed
    to :func:`read_expected_mutations_with_rows`, and its status filter is
    unchanged. What changed is that a filtered row is no longer silent: the
    plate-order check reads the same sheet without looking at status, so the two
    readers saw different row sets and named different wells for the same mutant.

    Anything else is read as a plain list: one variant per row, in plate order,
    from the named sheet and column. A wild-type row is recognised and its plate
    ordinal recorded rather than parsed as a mutation.

    Raises:
        ValueError: on an unreadable notation, a duplicate variant, a second
            wild-type row, an empty list, a column that cannot be identified, a
            named column of a multi-column file whose header is itself a
            variant, or a row that was read and could not be placed. Each is a
            mis-scored plate if it were let through, so none of them is a
            warning.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"variant list not found: {path}")

    result = _read_variant_source(path, sheet, variant_column)
    if result.dropped_rows:
        raise _dropped_rows_error(path, result.dropped_rows)
    return result


__all__ = [
    "KURO_SHEET",
    "DroppedRow",
    "VariantListReadResult",
    "VariantSourceInfo",
    "inspect_variant_source",
    "read_variant_source",
]
