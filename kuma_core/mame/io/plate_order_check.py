"""Does an exported workbook describe one plate, or two?

A KURO export carries the primer plate on ``Fwd List``/``Fwd Plate`` and the expected
variants on ``expected_mutations``. MAME turns the expected sheet into a plate, so the
two are the same statement written twice and they have to agree.

Exports written before v0.14.3 did not agree. The plate sheets came from the plate
mapping while the expected sheet followed the design ranking, so the same mutants sat
in different wells depending on which sheet was read. Nothing failed: every well got a
variant, the counts looked right, and the verdicts were scored against a plate nobody
had built. On the 260722 R2-1 export `K53I` sits at A2 by the primer list and `I92D`
sits there by the expected sheet.

A silent wrong answer is worth more noise than a loud one, so this reports the
disagreement rather than repairing it. Repair needs the operator to say which sheet
describes the tubes they actually pipetted, and that is not a guess to make for them.

**The comparison axis is the well, not the row number.** Both sides already know
their own wells and this module used to throw that away: the plate sheets were
collapsed into a dense list and the expected sheet was counted off one row at a
time, then each side was re-labelled from its position in that list. Three
things went wrong at once and all three were shifts of exactly the kind this
module exists to catch. A plate with a gap in it renumbered every later well. A
wild-type row, which occupies a well and has no primer, made the two lists
differ in length and put ``WT`` in ``absent_from_plate``. A row the expected
reader drops on status was counted here as though it were placed.

Keying on the well removes all three without a case for any of them, because it
stops re-deriving what each side already states: the plate sheets carry a Well
column and a labelled grid, and the expected side is placed by
``io.variant_list.read_variant_source`` together with ``layout.build_draft_layout``,
which are what the run itself uses. Reading the expected sheet a second way here
was the whole defect: two readers of one sheet disagreeing about which well a
mutant sits in is precisely what this check was written to detect.

The wild-type well is left out of the comparison. A plate sheet lists primers,
the control has none, and its absence is not a disagreement.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from kuma_core.mame.io.variant_list import read_variant_source
from kuma_core.mame.layout import (
    DEFAULT_WT_PLACEMENT,
    WtPlacement,
    build_draft_layout,
)
from kuma_core.mame.plate_geometry import canonical_well, well_to_seq

#: The sheet MAME reads as the plate order.
EXPECTED_SHEET = "expected_mutations"
#: Sheets that carry the primer plate, in preference order. The first one present wins.
PLATE_SHEETS = ("Fwd List", "Fwd Plate")

#: The occupant ``build_draft_layout`` writes for the control well.
_WT_SAMPLE = "WT"

_MUTATION = re.compile(r"^([A-Z])(\d+)([A-Z])$")


@dataclass(frozen=True)
class PlateOrderReport:
    """What the two sheets disagree about, if anything."""

    #: True when the file could be compared at all (both kinds of sheet present).
    comparable: bool
    #: True when a comparison ran and the orders differ.
    mismatched: bool = False
    #: ``(well, from_plate_sheet, from_expected_sheet)`` for the first few wells that
    #: disagree. Wells are labelled column-major, the order MAME assigns. A side that
    #: places nothing in a disagreeing well contributes an empty string.
    examples: list[tuple[str, str, str]] = field(default_factory=list)
    #: Mutants on the plate with no row in the expected sheet. Each one shifts every
    #: later well by one, so this is a mismatch even when the shared rows line up.
    missing_from_expected: list[str] = field(default_factory=list)
    #: Rows in the expected sheet naming a mutant the plate does not carry. Never
    #: contains the wild-type control: it occupies a well and has no primer, so a
    #: plate sheet without it is not disagreeing about anything.
    absent_from_plate: list[str] = field(default_factory=list)
    #: Which sheet supplied the plate order.
    plate_sheet: str | None = None
    #: Wells the plate sheet declares more than once, with two different
    #: occupants. The sheet states two plates in one, and which one was pipetted
    #: is not in the file, so this is reported rather than resolved. A well
    #: repeated with the *same* occupant is not listed: both readings name the
    #: same plate, so nothing downstream can differ.
    duplicate_wells: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        """True when nothing was found to report."""
        return (
            not self.mismatched
            and not self.missing_from_expected
            and not self.duplicate_wells
        )


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _canonical_well(raw: str) -> str:
    """Well label in the one spelling both sides use, or ``""`` if it is not one.

    The spelling rule itself lives in
    :func:`kuma_core.mame.plate_geometry.canonical_well`, which is where every
    other reader of a Well column gets it. Only the verdict differs here: a
    sheet row that is not a well is a row to ignore rather than a fault, so the
    refusal is turned into an empty string instead of travelling up.
    """
    try:
        return canonical_well(raw)
    except (ValueError, IndexError):
        return ""


def _mutation_from_primer(name: str) -> str:
    """Strip the ``_F``/``_R`` suffix a primer name carries."""
    stem = name[:-2] if name.upper().endswith(("_F", "_R")) else name
    return stem if _MUTATION.match(stem) else ""


def _in_plate_order(cells: dict[str, str]) -> dict[str, str]:
    """Re-key a well map into column-major order, the order MAME assigns."""
    return {well: cells[well] for well in sorted(cells, key=well_to_seq)}


def _place(
    cells: dict[str, str],
    duplicates: list[str],
    well: str,
    mutation: str,
) -> None:
    """Put one occupant in a well, recording a well that was already taken.

    Assigning straight into the dict is last-write-wins, and last-write-wins is
    the wrong answer twice over here: the second declaration overwrites the first
    with no trace, and a sheet that names one well twice necessarily leaves some
    other well unnamed, so the plate this module compares against is not the
    plate the sheet describes. Neither reading is more correct than the other,
    which is why the well is reported rather than resolved. The first occupant is
    kept so the comparison still has something to run on.
    """
    if well in cells:
        if cells[well] != mutation and well not in duplicates:
            duplicates.append(well)
        return
    cells[well] = mutation


def _plate_layout_from_list_sheet(worksheet) -> tuple[dict[str, str], list[str]]:
    """Read a ``Fwd List`` sheet into ``{well: mutation}`` plus repeated wells.

    The Well column is the coordinate, so the row order of the sheet says
    nothing and no sorting rule is needed to recover one. A well named by two
    rows that disagree comes back in the second element; see :func:`_place`.
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}, []
    headers = [_cell_text(c).lower() for c in rows[0]]

    def column(*names: str) -> int | None:
        for name in names:
            if name in headers:
                return headers.index(name)
        return None

    well_at = column("well")
    label_at = column("mutation", "mutant_id", "primer name", "primer_name")
    if well_at is None or label_at is None:
        return {}, []
    cells: dict[str, str] = {}
    duplicates: list[str] = []
    for row in rows[1:]:
        if max(well_at, label_at) >= len(row):
            continue
        well = _canonical_well(_cell_text(row[well_at]))
        label = _cell_text(row[label_at])
        mutation = label if _MUTATION.match(label) else _mutation_from_primer(label)
        if well and mutation:
            _place(cells, duplicates, well, mutation)
    return _in_plate_order(cells), duplicates


def _plate_layout_from_grid_sheet(worksheet) -> tuple[dict[str, str], list[str]]:
    """Read a ``Fwd Plate`` grid into ``{well: mutation}`` plus repeated wells.

    Row labels down the side and column numbers on top are the well, so an empty
    cell leaves a gap in the plate instead of pulling the rest of the grid up. A
    grid can name one well twice as well: a repeated row label, or a repeated
    column number in the header. Those come back in the second element; see
    :func:`_place`.
    """
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}, []
    header = [_cell_text(c) for c in rows[0]]
    cells: dict[str, str] = {}
    duplicates: list[str] = []
    for row in rows[1:]:
        if not row:
            continue
        row_label = _cell_text(row[0]).upper()
        if len(row_label) != 1 or not ("A" <= row_label <= "P"):
            continue
        for index, value in enumerate(row[1:], start=1):
            if index >= len(header) or not header[index].isdigit():
                continue
            label = _cell_text(value)
            mutation = label if _MUTATION.match(label) else _mutation_from_primer(label)
            well = _canonical_well(f"{row_label}{int(header[index])}")
            if well and mutation:
                _place(cells, duplicates, well, mutation)
    return _in_plate_order(cells), duplicates


def _expected_layout(
    path: Path,
    wt_placement: WtPlacement = DEFAULT_WT_PLACEMENT,
) -> dict[str, str] | None:
    """Where the expected sheet puts each mutant, or ``None`` if it cannot say.

    This is the run's own placement, not a second reading of it:
    :func:`~kuma_core.mame.io.variant_list.read_variant_source` decides which
    rows are occupants and at which ordinal the control sits, and
    :func:`~kuma_core.mame.layout.build_draft_layout` turns that into wells.
    Deriving it here again is what let this check and the run name different
    wells for one mutant.

    ``read_variant_source`` refuses a file for several reasons, all of them
    reported to the operator on the path that actually loads the workbook. This
    check is not that path, so a refusal here is lowered to "cannot compare"
    rather than raised: a check that threw would turn every one of those
    refusals into a second, differently worded failure, and would let a
    diagnostic block the gate that is already blocking.
    """
    try:
        read = read_variant_source(path)
    except (ValueError, FileNotFoundError, KeyError):
        return None
    # Every argument the run passes, and the same control-well default. This
    # check is only worth anything while it names the wells the run names, so it
    # follows ``build_draft_layout`` rather than pinning a policy of its own.
    draft = build_draft_layout(
        read.expected,
        wt_ordinal=read.wt_ordinal,
        wells=read.wells,
        wt_well=read.wt_well,
        wt_placement=wt_placement,
    )
    # Over capacity the draft places nothing, so there is no placement to compare.
    return draft.layout or None


def check_plate_order(
    path: Path,
    max_examples: int = 5,
    wt_placement: WtPlacement = DEFAULT_WT_PLACEMENT,
) -> PlateOrderReport:
    """Compare the plate sheets against ``expected_mutations`` in *path*.

    A file missing either kind of sheet is reported as not comparable rather than as
    consistent, so a caller cannot read silence as agreement. So is a file the
    expected-variant reader refuses; see :func:`_expected_layout` for why that is
    lowered rather than raised. This function does not raise.

    ``wt_placement`` is the control-well policy the run would use, forwarded so
    this check names the wells the run names. It has to travel: the check is a
    comparison against a hand-written primer plate, and answering it under a
    different placement than the run uses turns every agreement into a
    disagreement and back again.
    """
    path = Path(path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"} or not path.exists():
        return PlateOrderReport(comparable=False)

    import openpyxl

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return PlateOrderReport(comparable=False)
    try:
        if EXPECTED_SHEET not in workbook.sheetnames:
            return PlateOrderReport(comparable=False)
        plate: dict[str, str] = {}
        duplicate_wells: list[str] = []
        plate_sheet: str | None = None
        for name in PLATE_SHEETS:
            if name not in workbook.sheetnames:
                continue
            reader = (
                _plate_layout_from_list_sheet
                if name.endswith("List")
                else _plate_layout_from_grid_sheet
            )
            plate, duplicate_wells = reader(workbook[name])
            if plate:
                plate_sheet = name
                break
    finally:
        workbook.close()

    if not plate:
        return PlateOrderReport(comparable=False)

    layout = _expected_layout(path, wt_placement)
    if layout is None:
        return PlateOrderReport(comparable=False, plate_sheet=plate_sheet)

    # The control occupies a well and has no primer, so its well is not a place
    # the two sheets can disagree. Leaving it in is what put `WT` in
    # `absent_from_plate` and made every export carrying a WT row mismatch.
    expected = {
        well: sample for well, sample in layout.items() if sample != _WT_SAMPLE
    }
    wt_wells = {well for well, sample in layout.items() if sample == _WT_SAMPLE}

    examples: list[tuple[str, str, str]] = []
    for well in sorted(set(plate) | set(expected), key=well_to_seq):
        if well in wt_wells:
            continue
        on_plate = plate.get(well, "")
        in_expected = expected.get(well, "")
        if on_plate != in_expected and len(examples) < max_examples:
            examples.append((well, on_plate, in_expected))

    expected_set = set(expected.values())
    plate_set = set(plate.values())
    missing_from_expected = [m for m in plate.values() if m not in expected_set]
    absent_from_plate = [m for m in expected.values() if m not in plate_set]

    return PlateOrderReport(
        comparable=True,
        # A repeated well is counted here as well as in its own field: the
        # sheet describes two plates, which is the same fault this flag has
        # always stood for, and a caller that reads only the flag (the wire
        # payload in ``sidecar_mame.handlers.barcode_package`` is one) would
        # otherwise be told the file is clean.
        mismatched=bool(
            examples or missing_from_expected or absent_from_plate or duplicate_wells
        ),
        examples=examples,
        missing_from_expected=missing_from_expected,
        absent_from_plate=absent_from_plate,
        plate_sheet=plate_sheet,
        duplicate_wells=duplicate_wells,
    )


__all__ = ["EXPECTED_SHEET", "PLATE_SHEETS", "PlateOrderReport", "check_plate_order"]
