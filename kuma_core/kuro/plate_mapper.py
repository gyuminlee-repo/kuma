"""96-well plate mapping and Excel export for SDM primers."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from kuma_core.shared.janus_deck import (
    JANUS_DEVICE_HEADER,
    KURO_PRIMER_DECK,
    JanusDeck,
)

from .sdm_engine import SdmPrimerResult


@dataclass
class PlateMapping:
    """Single well assignment in a 96-well plate."""

    well: str               # e.g. "A1", "B1"
    primer_name: str        # e.g. "Q232A_F" or "Q232A_R"
    sequence: str           # Primer sequence
    primer_type: str        # "forward" or "reverse"
    mutation: str           # Original mutation notation
    tm: float | None = None          # Whole-primer Tm (Fwd or Rev)
    tm_overlap: float | None = None  # Overlap Tm
    wt_codon: str | None = None      # Wild-type codon
    mt_codon: str | None = None      # Mutant codon


def _well_name(index: int, order: str = "column") -> str:
    """Convert a 0-based index to a well name.

    Args:
        index: 0-based well index.
        order: 'column' (A1->H1->A2) or 'row' (A1->A12->B1).

    Returns:
        Well name like "A1".
    """
    rows = "ABCDEFGH"
    if order == "column":
        col = index // 8 + 1
        row = index % 8
    else:  # row
        row = index // 12
        col = index % 12 + 1
    if row >= 8 or col > 12:
        raise ValueError(f"Well index {index} exceeds 96-well plate capacity")
    return f"{rows[row]}{col}"


def _assign_well(index: int, well_order: str = "column") -> str:
    """Generate well name with overflow to additional plates."""
    plate_num = index // 96
    local_idx = index % 96
    well = _well_name(local_idx, well_order)
    if plate_num == 0:
        return well
    return f"P{plate_num + 1}-{well}"


def _parse_well_plate(well: str) -> tuple[int, str]:
    """Return (0-based plate index, base well) from an overflow well label.

    ``"A1"`` → ``(0, "A1")``,  ``"P2-A1"`` → ``(1, "A1")``.
    Inverse of ``_assign_well``.
    """
    if "-" in well:
        prefix, base = well.split("-", 1)
        return int(prefix[1:]) - 1, base
    return 0, well


def deduplicate_reverse(
    results: list[SdmPrimerResult],
) -> dict[str, list[str]]:
    """Find mutations sharing identical reverse primers.

    Returns:
        Dict mapping reverse primer sequence to list of mutation names.
    """
    rev_map: dict[str, list[str]] = {}
    for r in results:
        rev_seq = r.reverse_seq
        if rev_seq not in rev_map:
            rev_map[rev_seq] = []
        rev_map[rev_seq].append(r.mutation.raw)
    return rev_map


def _capacity_for_range(mapping_range: tuple[str, str] | None) -> int:
    """Return per-96-plate well capacity given an optional 384-row range.

    Default capacity is 96. With ``mapping_range=(row_start, row_end)`` (inclusive
    over 384 rows A-P), capacity is ``(rows_in_range // 2) * 12`` because each
    96-row consumes one fwd row + one rev row in the 384 destination.
    """
    rng = _validate_mapping_range(mapping_range)
    if rng is None:
        return 96
    s, e = rng
    pair_count = (e - s + 1) // 2
    return pair_count * 12


def _assign_well_with_range(
    index: int,
    well_order: str,
    mapping_range: tuple[str, str] | None,
) -> str:
    """Assign well with overflow respecting an optional mapping_range capacity."""
    capacity = _capacity_for_range(mapping_range)
    plate_num = index // capacity
    local_idx = index % capacity
    # within a plate, use the same row/col scheme but constrained row count
    rng = _validate_mapping_range(mapping_range)
    if rng is None:
        well = _well_name(local_idx, well_order)
    else:
        s, e = rng
        rows_96 = (e - s + 1) // 2
        rows = "ABCDEFGH"[:rows_96]
        if well_order == "column":
            col = local_idx // rows_96 + 1
            row = local_idx % rows_96
        else:
            row = local_idx // 12
            col = local_idx % 12 + 1
        if row >= rows_96 or col > 12:
            raise ValueError(
                f"Well index {local_idx} exceeds reduced plate capacity "
                f"({rows_96} rows × 12)"
            )
        well = f"{rows[row]}{col}"
    if plate_num == 0:
        return well
    return f"P{plate_num + 1}-{well}"


def generate_plate_map(
    results: list[SdmPrimerResult],
    well_order: str = "column",
    deduplicate_rev: bool = True,
    mapping_range: tuple[str, str] | None = None,
) -> tuple[list[PlateMapping], list[PlateMapping]]:
    """Generate separate Fwd and Rev plate mappings.

    Forward plate: one primer per mutation, sequential well assignment.
    Reverse plate: deduplicated reverse primers, sequential well assignment.

    Args:
        mapping_range: Optional (row_start, row_end) inclusive over 384-well rows
            A-P. When set, 96-well per-plate capacity shrinks to
            ``(rows_in_range // 2) * 12`` and overflow rolls to P2 etc.

    Returns:
        Tuple of (fwd_mappings, rev_mappings).
    """
    # Forward plate
    fwd_mappings: list[PlateMapping] = []
    for idx, r in enumerate(results):
        well = _assign_well_with_range(idx, well_order, mapping_range)
        fwd_mappings.append(PlateMapping(
            well=well,
            primer_name=f"{r.mutation.raw}_F",
            sequence=r.forward_seq,
            primer_type="forward",
            mutation=r.mutation.raw,
        ))

    # Reverse plate (deduplicated)
    rev_mappings: list[PlateMapping] = []
    if deduplicate_rev:
        rev_groups = deduplicate_reverse(results)
        for idx, (rev_seq, mut_names) in enumerate(rev_groups.items()):
            label = mut_names[0]
            well = _assign_well_with_range(idx, well_order, mapping_range)
            rev_mappings.append(PlateMapping(
                well=well,
                primer_name=f"{label}_R",
                sequence=rev_seq,
                primer_type="reverse",
                mutation=label,
            ))
    else:
        for idx, r in enumerate(results):
            well = _assign_well_with_range(idx, well_order, mapping_range)
            rev_mappings.append(PlateMapping(
                well=well,
                primer_name=f"{r.mutation.raw}_R",
                sequence=r.reverse_seq,
                primer_type="reverse",
                mutation=r.mutation.raw,
            ))

    return fwd_mappings, rev_mappings


def _is_shared_rev(m: PlateMapping, rev_groups: dict[str, list[str]] | None) -> bool:
    """Check if a reverse primer is shared by multiple mutations."""
    if m.primer_type != "reverse" or rev_groups is None:
        return False
    return len(rev_groups.get(m.sequence, [])) > 1


# Colors matching the UI: Fwd=green, Rev=orange, Shared=blue
_COLOR_FWD = "C6EFCE"       # light green
_COLOR_REV = "FCE4D6"       # light orange
_COLOR_SHARED = "BDD7EE"    # light blue


def _write_list_sheet(
    ws,
    mappings: list[PlateMapping],
    fill_color: str,
    rev_groups: dict[str, list[str]] | None = None,
    overlap_mode: str = "partial",
) -> None:
    """Write a primer list sheet with row coloring."""
    from openpyxl.styles import Font, PatternFill

    # In full-overlap mode the Tm_Overlap column is renamed Tm_Primer (same value as Tm),
    # keeping column count identical for downstream parser compatibility.
    tm_overlap_header = "Tm_Primer" if overlap_mode == "full" else "Tm_Overlap"
    headers = ["Well", "Primer Name", "Sequence", "Length",
               "Tm", tm_overlap_header, "WT_Codon", "MT_Codon", "Mutation"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")

    num_cols = len(headers)
    for i, m in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=m.well)
        ws.cell(row=i, column=2, value=m.primer_name)
        ws.cell(row=i, column=3, value=m.sequence)
        ws.cell(row=i, column=4, value=len(m.sequence))
        ws.cell(row=i, column=5, value=m.tm)
        ws.cell(row=i, column=6, value=m.tm_overlap)
        ws.cell(row=i, column=7, value=m.wt_codon)
        ws.cell(row=i, column=8, value=m.mt_codon)
        ws.cell(row=i, column=9, value=m.mutation)

        # Row color: shared=blue, otherwise fill_color
        row_color = _COLOR_SHARED if _is_shared_rev(m, rev_groups) else fill_color
        row_fill = PatternFill(start_color=row_color, fill_type="solid")
        for col_idx in range(1, num_cols + 1):
            ws.cell(row=i, column=col_idx).fill = row_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def _write_plate_sheet(
    ws,
    mappings: list[PlateMapping],
    fill_color: str,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Write a 96-well plate layout sheet with shared primers in blue."""
    from openpyxl.styles import Alignment, Font, PatternFill

    rows_label = "ABCDEFGH"

    for c in range(1, 13):
        cell = ws.cell(row=1, column=c + 1, value=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r_idx, r_label in enumerate(rows_label):
        ws.cell(row=r_idx + 2, column=1, value=r_label).font = Font(bold=True)

    default_fill = PatternFill(start_color=fill_color, fill_type="solid")
    shared_fill = PatternFill(start_color=_COLOR_SHARED, fill_type="solid")

    for m in mappings:
        if not m.well[0].isalpha():
            continue
        row_letter = m.well[0]
        col_num = int(m.well[1:])
        if row_letter not in rows_label or col_num > 12:
            continue
        r_idx = rows_label.index(row_letter) + 2
        c_idx = col_num + 1

        cell = ws.cell(row=r_idx, column=c_idx, value=m.primer_name)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = shared_fill if _is_shared_rev(m, rev_groups) else default_fill

    ws.column_dimensions["A"].width = 4
    for c in range(2, 14):
        ws.column_dimensions[chr(64 + c)].width = 14


def _chunk_by_plate(mappings: list[PlateMapping]) -> list[list[PlateMapping]]:
    """Split mappings into 96-well plate chunks."""
    plates: list[list[PlateMapping]] = []
    current: list[PlateMapping] = []
    for m in mappings:
        # Normalize well: strip overflow prefix like "P2-"
        well = m.well
        if "-" in well:
            well = well.split("-", 1)[1]
            m = PlateMapping(
                well=well,
                primer_name=m.primer_name,
                sequence=m.sequence,
                primer_type=m.primer_type,
                mutation=m.mutation,
                tm=m.tm,
                tm_overlap=m.tm_overlap,
                wt_codon=m.wt_codon,
                mt_codon=m.mt_codon,
            )
        current.append(m)
        if len(current) >= 96:
            plates.append(current)
            current = []
    if current:
        plates.append(current)
    return plates if plates else [[]]


def _pair_rev_per_plate(
    fwd_plates: list[list[PlateMapping]],
    rev_all: list[PlateMapping],
    rev_groups: dict[str, list[str]] | None,
) -> list[list[PlateMapping]]:
    """Pair reverse primers with each fwd plate by mutation membership.

    For each fwd plate chunk, collect the deduplicated reverse primers
    that belong to that chunk's mutations, reassigning well names.
    """
    # Build mutation → rev sequence lookup
    mut_to_rev_seq: dict[str, str] = {}
    if rev_groups:
        for seq, muts in rev_groups.items():
            for mut in muts:
                mut_to_rev_seq[mut] = seq

    rev_by_seq = {r.sequence: r for r in rev_all}

    paired: list[list[PlateMapping]] = []
    for fwd_chunk in fwd_plates:
        seen_seq: set[str] = set()
        rev_chunk: list[PlateMapping] = []
        well_idx = 0
        for fwd_m in fwd_chunk:
            rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
            if rev_seq and rev_seq not in seen_seq:
                seen_seq.add(rev_seq)
                rev_m = rev_by_seq.get(rev_seq)
                if rev_m:
                    rev_chunk.append(PlateMapping(
                        well=_well_name(well_idx),
                        primer_name=rev_m.primer_name,
                        sequence=rev_m.sequence,
                        primer_type="reverse",
                        mutation=rev_m.mutation,
                        tm=rev_m.tm,
                        tm_overlap=rev_m.tm_overlap,
                        wt_codon=rev_m.wt_codon,
                        mt_codon=rev_m.mt_codon,
                    ))
                    well_idx += 1
        paired.append(rev_chunk)
    return paired



def _order_results_by_plate(results: list, mappings: list[PlateMapping]) -> list:
    """Return *results* in the well order the plate sheets use.

    MAME assigns well *i* from row *i* of this sheet (``build_draft_layout``), so the
    row order here is a plate coordinate, not a presentation choice. The plate sheets
    are written from ``mappings`` while this sheet used to iterate ``results``, which
    arrives in design order (the ranking a diversity or EVOLVEpro run produced). One
    workbook then described two different plates, and reading the wrong one places
    every well from a list nobody built. Observed on the 260722 R2-1 export: the same
    94 mutants in both sheets, `K53I` at well A2 by the primer list and `I92D` at A2
    by this sheet, which scored 94 of 95 wells against the wrong design.

    Forward mappings carry the well order. Results with no forward mapping keep their
    original relative order at the end rather than being dropped, since a missing
    mapping is not a reason to lose a designed mutation from the sheet.

    Ordering alone is not enough, which the same export also shows. `mappings` arrives
    from the UI and carries entries added while relaxing conditions on a failed
    mutation, while `results` is the design state, so a filled well can have a primer
    and no result. `V263I` at C7 is one: a primer, a reverse partner, and no row here.
    Dropping it shifts every later well by one, so :func:`_plate_ordered_row_sources`
    emits a row per well instead and this function only settles the order.
    """
    rank: dict[str, int] = {}
    for index, mapping in enumerate(mappings):
        if mapping.primer_type != "forward":
            continue
        rank.setdefault(mapping.mutation, index)
    fallback = len(mappings)
    ordered = sorted(
        enumerate(results),
        key=lambda pair: (rank.get(pair[1].mutation.raw, fallback), pair[0]),
    )
    return [result for _, result in ordered]


def _plate_ordered_row_sources(
    results: list, mappings: list[PlateMapping]
) -> list[tuple[object, PlateMapping | None]]:
    """Pair every plate well with the result describing it, in well order.

    Returns ``(result_or_None, mapping_or_None)``. One entry per forward mapping, in
    mapping order, followed by any result the plate does not carry. A well whose
    mutation has no result yields ``(None, mapping)``: the caller synthesises that row
    from the mapping, because MAME counts rows to find wells and a gap here silently
    renames every later well.
    """
    by_mutation: dict[str, object] = {}
    for result in results:
        by_mutation.setdefault(result.mutation.raw, result)

    sources: list[tuple[object, PlateMapping | None]] = []
    seen: set[str] = set()
    for mapping in mappings:
        if mapping.primer_type != "forward" or mapping.mutation in seen:
            continue
        seen.add(mapping.mutation)
        sources.append((by_mutation.get(mapping.mutation), mapping))
    for result in results:
        if result.mutation.raw not in seen:
            sources.append((result, None))
    return sources


def _synthesised_row(mapping: PlateMapping, headers: list[str]) -> list:
    """Build an expected-mutations row for a well whose mutation has no result.

    Only what the mapping actually states is written. The residue fields come from the
    notation, the codons from the mapping when it carries them, and everything else is
    left empty rather than guessed. ``status`` is ``DESIGNED`` because a primer for this
    well exists, which is the question that field answers for MAME.
    """
    match = re.match(r"^([A-Za-z])(\d+)([A-Za-z*])$", mapping.mutation.strip())
    values: dict[str, object] = {
        "mutant_id": mapping.mutation,
        "position": int(match.group(2)) if match else None,
        "wt_aa": match.group(1).upper() if match else "",
        "mt_aa": match.group(3).upper() if match else "",
        "wt_codon": mapping.wt_codon or "",
        "mt_codon": mapping.mt_codon or "",
        "group_id": "",
        "primer_set_ref": mapping.mutation,
        "notation_type": "substitution" if match else "",
        "status": "DESIGNED",
    }
    return [values.get(header, "") for header in headers]


def _write_expected_mutations_sheet(
    wb,                              # openpyxl.Workbook
    row_sources: list,               # list[tuple[SdmPrimerResult | None, PlateMapping | None]]
    rescued_info: list[dict] | None = None,  # list of RescuedMutation dicts from frontend
) -> None:
    """Append 'expected_mutations' sheet to workbook.

    One row per plate well, in well order, because MAME reads row *i* of this sheet as
    well *i*. Callers build the pairs with :func:`_plate_ordered_row_sources`; a well
    whose mutation has no design result is written from its mapping rather than
    skipped, since a skipped row renames every later well.
    Multi-notation (A40P/E61Y) produces one row per sub-mutation;
    they are linked via group_id.
    FAILED mutations excluded in Phase 1.

    Sheet columns (in order):
        mutant_id, position, wt_aa, mt_aa, wt_codon, mt_codon,
        group_id, primer_set_ref, notation_type, status,
        rescue_type, rescue_stage, rescued_from

    status values:
        "DESIGNED" — primer exists and should be included by downstream readers.

    rescue_type values:
        same_position, diff_position, auto_suggestion_l1~l4,
        pool_cascade, auto_relax, auto_suggestion.
    """
    # Build lookup: rescued_by (mutation_raw in results) -> rescue metadata.
    rescued_lookup: dict[str, dict] = {}
    if rescued_info:
        for item in rescued_info:
            if hasattr(item, "model_dump"):
                item = item.model_dump(exclude_none=True)
            if not isinstance(item, dict):
                continue
            key = item.get("rescued_by")
            stage_type = item.get("type")
            if key and stage_type:
                rescued_lookup[str(key)] = item

    HEADERS = [
        "mutant_id", "position", "wt_aa", "mt_aa",
        "wt_codon", "mt_codon", "group_id", "primer_set_ref",
        "notation_type", "status", "rescue_type", "rescue_stage", "rescued_from",
    ]
    ws = wb.create_sheet("expected_mutations")
    ws.append(HEADERS)
    for r, mapping in row_sources:
        if r is None:
            if mapping is None:
                continue
            ws.append(_synthesised_row(mapping, HEADERS))
            continue
        m = r.mutation
        rescue_meta = rescued_lookup.get(m.raw, {})
        ws.append([
            m.raw,
            m.position,
            m.wt_aa,
            m.mt_aa,
            m.wt_codon,
            m.mt_codon,
            m.group_id or "",
            m.raw,           # primer_set_ref == Mutation.raw
            "substitution",  # notation_type: Phase 1 constant
            "DESIGNED",
            rescue_meta.get("type", ""),
            rescue_meta.get("stage", ""),
            rescue_meta.get("original", ""),
        ])


def export_plate_excel(
    mappings: list[PlateMapping],
    output_path: Path,
    rev_groups: dict[str, list[str]] | None = None,
    results: list | None = None,          # list[SdmPrimerResult] — added in Phase 1
    overlap_mode: str = "partial",
    rescued_info: list[dict] | None = None,  # rescue stage info from frontend
) -> None:
    """Export plate mappings to Excel with per-plate sheets.

    Forward and reverse primers are paired per plate: Rev Plate N contains
    only the reverse primers for Fwd Plate N's mutations.

    For N plates, creates sheets:
    - 'Fwd List 1', 'Fwd Plate 1', 'Rev List 1', 'Rev Plate 1'
    - 'Fwd List 2', 'Fwd Plate 2', ... (if multi-plate)

    Single plate omits the number suffix.

    Args:
        rev_groups: Original reverse deduplication map (seq -> mutation names).
            Used to pair rev with fwd plates and detect shared primers.
        overlap_mode: "partial" (Gibson-style) or "full" (NEB Q5 SDM style).
            Controls Tm_Overlap vs Tm_Primer header in list sheets.
        rescued_info: List of RescuedMutation dicts (keys: rescued_by, type, original).
            When provided, rescue_type/rescue_stage columns record the cascade
            stage while status remains "DESIGNED" for downstream readers.
    """
    from openpyxl import Workbook

    fwd_all = [m for m in mappings if m.primer_type == "forward"]
    rev_all = [m for m in mappings if m.primer_type == "reverse"]

    fwd_plates = _chunk_by_plate(fwd_all)
    rev_plates = _pair_rev_per_plate(fwd_plates, rev_all, rev_groups)
    plate_count = len(fwd_plates)
    suffix = plate_count > 1

    wb = Workbook()
    first_sheet = True
    chunk_rev_groups = rev_groups or {}

    for i in range(plate_count):
        tag = f" {i + 1}" if suffix else ""
        fwd_chunk = fwd_plates[i]
        rev_chunk = rev_plates[i] if i < len(rev_plates) else []

        if first_sheet:
            ws = wb.active
            ws.title = f"Fwd List{tag}"
            first_sheet = False
        else:
            ws = wb.create_sheet(f"Fwd List{tag}")
        _write_list_sheet(ws, fwd_chunk, _COLOR_FWD, overlap_mode=overlap_mode)

        ws = wb.create_sheet(f"Fwd Plate{tag}")
        _write_plate_sheet(ws, fwd_chunk, _COLOR_FWD)

        ws = wb.create_sheet(f"Rev List{tag}")
        _write_list_sheet(ws, rev_chunk, _COLOR_REV, rev_groups=chunk_rev_groups,
                          overlap_mode=overlap_mode)

        ws = wb.create_sheet(f"Rev Plate{tag}")
        _write_plate_sheet(ws, rev_chunk, _COLOR_REV, rev_groups=chunk_rev_groups)

    # Phase 1: append expected_mutations sheet when results are provided, in the same
    # well order the plate sheets above use. MAME reads row i of that sheet as well i.
    if results:
        _write_expected_mutations_sheet(
            wb,
            _plate_ordered_row_sources(results, mappings),
            rescued_info=rescued_info,
        )

    wb.save(output_path)


def export_idt_csv(
    results: list[SdmPrimerResult],
    output_path: Path,
    scale: str = "25nm",
    purification: str = "STD",
    encoding: str = "utf-8",
) -> None:
    """Export primer order CSV in IDT OligoEntry format.

    CSV columns: Name, Sequence, Scale, Purification.
    Each mutation produces two rows (forward and reverse).

    Args:
        results: List of SdmPrimerResult from primer design.
        output_path: Path to output CSV file.
        scale: Synthesis scale (default: "25nm").
        purification: Purification type (default: "STD").
        encoding: File encoding (default "utf-8"; use "utf-8-sig" for BOM).
    """
    import csv

    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Sequence", "Scale", "Purification"])
        for r in results:
            writer.writerow([f"{r.mutation.raw}_F", r.forward_seq, scale, purification])
            writer.writerow([f"{r.mutation.raw}_R", r.reverse_seq, scale, purification])


def export_twist_csv(
    results: list[SdmPrimerResult],
    output_path: Path,
    encoding: str = "utf-8",
) -> None:
    """Export primer order CSV in Twist Bioscience bulk order format.

    CSV columns: Name, Sequence, Notes.
    Each mutation produces two rows (forward and reverse).

    Args:
        results: List of SdmPrimerResult from primer design.
        output_path: Path to output CSV file.
        encoding: File encoding (default "utf-8"; use "utf-8-sig" for BOM).
    """
    import csv

    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Sequence", "Notes"])
        for r in results:
            writer.writerow([f"{r.mutation.raw}_F", r.forward_seq, r.mutation.raw])
            writer.writerow([f"{r.mutation.raw}_R", r.reverse_seq, r.mutation.raw])


# ---------------------------------------------------------------------------
# Liquid handler mapping exports
# ---------------------------------------------------------------------------

_ROWS_384 = "ABCDEFGHIJKLMNOP"
_ROWS_96 = "ABCDEFGH"
_ECHO_MAX_TRANSFER_NL = 500  # Echo 525 single-transfer limit


def _split_echo_volume(total_vol: int) -> list[int]:
    """Split a total Echo transfer volume into ≤500 nL steps.

    Echo 525 allows a maximum of 500 nL per single transfer event.
    Volumes above this threshold are split into multiple rows in the
    mapping file (low-repeat transfers to the same dest well).

    Example: 1000 nL → [500, 500]
             600 nL  → [500, 100]
             300 nL  → [300]
    """
    if total_vol <= 0:
        return []
    steps: list[int] = []
    remaining = total_vol
    while remaining > _ECHO_MAX_TRANSFER_NL:
        steps.append(_ECHO_MAX_TRANSFER_NL)
        remaining -= _ECHO_MAX_TRANSFER_NL
    if remaining > 0:
        steps.append(remaining)
    return steps


def _validate_mapping_range(
    mapping_range: tuple[str, str] | None,
) -> tuple[int, int] | None:
    """Validate (row_start, row_end) inclusive over 384-well rows A-P.

    Returns (start_idx, end_idx) 0-based, or None if mapping_range is None.
    The interval length must be even (fwd/rev row pairing) and >= 2.
    """
    if mapping_range is None:
        return None
    row_start, row_end = mapping_range
    if row_start not in _ROWS_384 or row_end not in _ROWS_384:
        raise ValueError(
            f"mapping_range rows must be within A-P, got {mapping_range!r}"
        )
    s = _ROWS_384.index(row_start)
    e = _ROWS_384.index(row_end)
    if e < s:
        raise ValueError(
            f"mapping_range row_end ({row_end}) precedes row_start ({row_start})"
        )
    span = e - s + 1
    if span % 2 != 0:
        raise ValueError(
            f"mapping_range span must be even (fwd/rev pairs); got {span} rows"
        )
    return s, e


def _to_384_well_fwd(
    well_96: str,
    mapping_range: tuple[str, str] | None = None,
    quadrant: str | None = None,
) -> str:
    """Map 96-well address to 384-well forward primer position.

    Default: forward primers occupy even-indexed 384 rows (A,C,E,G,I,K,M,O).
    With ``mapping_range=(row_start, row_end)`` (inclusive, A-P), forward rows
    are restricted to the even sub-offsets within that range.

    ``quadrant`` selects one of the four interleaved sets a 96-head Zephyr can
    stamp (A1/A2/B1/B2) and takes precedence over ``mapping_range``, which
    describes row bands of the older row-doubled layout and cannot express a
    column offset. See ``plate_quadrant`` for the geometry.
    """
    if quadrant is not None:
        from kuma_core.kuro.plate_quadrant import to_384_well

        return to_384_well(well_96, quadrant)
    rng = _validate_mapping_range(mapping_range)
    row_idx = _ROWS_96.index(well_96[0])
    if rng is None:
        return f"{_ROWS_384[row_idx * 2]}{well_96[1:]}"
    s, e = rng
    pair_count = (e - s + 1) // 2
    sub = row_idx % pair_count
    return f"{_ROWS_384[s + sub * 2]}{well_96[1:]}"


def _to_384_well_rev(
    well_96: str,
    mapping_range: tuple[str, str] | None = None,
    quadrant: str | None = None,
) -> str:
    """Map 96-well address to 384-well reverse primer position.

    Default: reverse primers occupy odd-indexed 384 rows (B,D,F,H,J,L,N,P).
    With ``mapping_range=(row_start, row_end)`` (inclusive, A-P), reverse rows
    are restricted to the odd sub-offsets within that range.

    ``quadrant`` here is the *forward* quadrant; the reverse set lands in its
    row-paired partner (A1 -> B1, A2 -> B2) so forward and reverse stay on
    adjacent rows of the same columns, as they did before.
    """
    if quadrant is not None:
        from kuma_core.kuro.plate_quadrant import paired_quadrant, to_384_well

        return to_384_well(well_96, paired_quadrant(quadrant))
    rng = _validate_mapping_range(mapping_range)
    row_idx = _ROWS_96.index(well_96[0])
    if rng is None:
        return f"{_ROWS_384[row_idx * 2 + 1]}{well_96[1:]}"
    s, e = rng
    pair_count = (e - s + 1) // 2
    sub = row_idx % pair_count
    return f"{_ROWS_384[s + sub * 2 + 1]}{well_96[1:]}"


#: Echo worklist columns, in the order the instrument reads them.
ECHO_DEVICE_HEADER: list[str] = [
    "Source Plate Name", "Source Well Name", "Source Well",
    "Dest Plate Name", "Dest Well Name", "Dest Well", "Transfer Vol",
]

# Row dict key per instrument column, in ``ECHO_DEVICE_HEADER`` order. The
# header carries the instrument spellings ("Source Well Name") while the row
# dicts are keyed by the sidecar preview field names, so neither can be derived
# from the other. This list is the single place the two orders are tied
# together, which is what lets the CSV, the XLSX worklist sheet and the preview
# emit the same seven columns without any of them restating the order.
_ECHO_ROW_KEYS: list[str] = [
    "source_plate",
    "source_well_name",
    "source_well",
    "dest_plate",
    "dest_well_name",
    "dest_well",
    "transfer_vol",
]


def build_echo_rows(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    rev_groups: dict[str, list[str]] | None = None,
    transfer_vol: int = 100,
    mapping_range: tuple[str, str] | None = None,
    quadrant: str | None = None,
    used_quadrants: list[str] | None = None,
) -> list[dict]:
    """Build Echo 525 transfer rows, one dict per transfer event.

    Single source of the row content behind the CSV export, the XLSX
    "Echo mapping file" sheet and the sidecar preview, so the three cannot drift
    apart. They did drift: the CSV honoured both ``quadrant`` and
    ``mapping_range``, the XLSX worklist sheet honoured neither and the preview
    only ``mapping_range``, so one ``export_all`` left two files naming
    different source wells for the same primer while the preview the operator
    checked agreed with neither.

    Source plate layout: 384-well (Eco 384PP).
      - Default: forward primers occupy odd rows (A, C, E, ...), reverse
        primers the even rows, both keeping the 96-well column ordering.
      - ``mapping_range`` restricts those rows to an inclusive A-P band.
      - ``quadrant`` places forward primers in one interleaved 96-head set and
        reverse primers in its row-paired partner. Takes precedence over
        ``mapping_range``, which cannot express a column offset.

    Row order is forward primers first (in ``fwd_mappings`` order), then reverse
    primers expanded so every forward mutation gets its own transfer row,
    aspirating from the shared source well when the reverse primer is shared.
    A volume above the Echo per-transfer ceiling is split into several rows.

    Keys are the sidecar preview field names, plus ``mutation`` (the mutation
    the row belongs to, not an instrument column).

    Args:
        fwd_mappings: Forward primer plate mappings (96-well coordinates).
        rev_mappings: Deduplicated reverse primer plate mappings (96-well).
        rev_groups: Reverse deduplication map {seq: [mutation_names]}.
            Used to expand shared primers to all destination wells.
        transfer_vol: Transfer volume in nL (default 100).
        mapping_range: Inclusive 384 row band (row_start, row_end).
        quadrant: Forward-primer quadrant (A1/A2/B1/B2).
        used_quadrants: Quadrants already spent on a part-used plate.
    """
    if quadrant is not None:
        # 이미 쓴 quadrant 위에 덮어쓰면 그 안의 프라이머가 사라진다. 경고가 아니라
        # 거부다. 판단 근거는 작업자가 입력한 현재 plate 상태뿐이다. 여기에 두어야
        # preview 도 같은 거부를 내고, 작업자가 못 쓸 배치를 검산하지 않는다.
        from kuma_core.kuro.plate_quadrant import check_quadrants_available

        check_quadrants_available(quadrant, used_quadrants)

    fwd_by_mut, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )

    rows: list[dict] = []

    # Forward: one row per mutation (split if over the per-transfer ceiling)
    for m in fwd_mappings:
        plate_idx, base_well = _parse_well_plate(m.well)
        src_well = _to_384_well_fwd(
            base_well, mapping_range=mapping_range, quadrant=quadrant,
        )
        for vol in _split_echo_volume(transfer_vol):
            rows.append({
                "source_plate": f"Source [{plate_idx + 1}]",
                "source_well_name": m.primer_name,
                "source_well": src_well,
                "dest_plate": f"Destination [{plate_idx + 1}]",
                "dest_well_name": m.mutation,
                "dest_well": base_well,
                "transfer_vol": vol,
                "mutation": m.mutation,
            })

    # Reverse: one row per (primer, dest_well) pair
    for fwd_m in fwd_mappings:
        rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
        if rev_seq is None:
            continue
        rev_m = rev_by_seq.get(rev_seq)
        if rev_m is None:
            continue

        fwd_plate_idx, _ = _parse_well_plate(fwd_m.well)
        _, rev_base_well = _parse_well_plate(rev_m.well)
        src_well = _to_384_well_rev(
            rev_base_well, mapping_range=mapping_range, quadrant=quadrant,
        )
        _, dest_well = _parse_well_plate(fwd_by_mut.get(fwd_m.mutation, fwd_m.well))

        for vol in _split_echo_volume(transfer_vol):
            rows.append({
                "source_plate": f"Source [{fwd_plate_idx + 1}]",
                "source_well_name": rev_m.primer_name,
                "source_well": src_well,
                "dest_plate": f"Destination [{fwd_plate_idx + 1}]",
                "dest_well_name": fwd_m.mutation,
                "dest_well": dest_well,
                "transfer_vol": vol,
                "mutation": fwd_m.mutation,
            })

    return rows


def echo_row_values(row: dict) -> list:
    """Flatten one ``build_echo_rows`` dict into ``ECHO_DEVICE_HEADER`` order."""
    return [row[key] for key in _ECHO_ROW_KEYS]


def export_echo_mapping_csv(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: int = 100,
    rev_groups: dict[str, list[str]] | None = None,
    encoding: str = "utf-8",
    mapping_range: tuple[str, str] | None = None,
    quadrant: str | None = None,
    used_quadrants: list[str] | None = None,
) -> None:
    """Export Echo 525 acoustic dispenser mapping CSV.

    Rows come from :func:`build_echo_rows`, so this file, the XLSX
    "Echo mapping file" sheet and the sidecar preview always describe the same
    transfers. Where the source wells land (default row-doubled,
    ``mapping_range`` band, or ``quadrant`` set) is decided there, not here.

    Args:
        fwd_mappings: Forward primer plate mappings (96-well coordinates).
        rev_mappings: Deduplicated reverse primer plate mappings (96-well).
        output_path: Output CSV file path.
        transfer_vol: Transfer volume in nL (default 100).
        rev_groups: Reverse deduplication map {seq: [mutation_names]}.
            Used to expand shared primers to all destination wells.
        encoding: File encoding (default "utf-8"; use "utf-8-sig" for BOM).
        mapping_range: Inclusive 384 row band (row_start, row_end).
        quadrant: Forward-primer quadrant (A1/A2/B1/B2).
        used_quadrants: Quadrants already spent on a part-used plate. A clash is
            refused before the file is opened, so nothing is written.
    """
    import csv

    rows = build_echo_rows(
        fwd_mappings, rev_mappings, rev_groups, transfer_vol,
        mapping_range=mapping_range,
        quadrant=quadrant,
        used_quadrants=used_quadrants,
    )

    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        writer.writerow(ECHO_DEVICE_HEADER)
        for row in rows:
            writer.writerow(echo_row_values(row))


# Row dict key per instrument column, in ``JANUS_DEVICE_HEADER`` order. The
# header carries the instrument spellings ("Asp. Rack") while the row dicts are
# keyed by the sidecar preview field names, so neither can be derived from the
# other. This list is the single place the two orders are tied together, which
# is what lets the CSV, the XLSX sheet and the preview emit the same eight
# columns without any of them restating the order.
_JANUS_ROW_KEYS: list[str] = [
    "name",
    "type",
    "no",
    "asp_rack",
    "asp_posi",
    "dsp_rack",
    "dsp_posi",
    "volume",
]


def build_janus_rows(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    rev_groups: dict[str, list[str]] | None = None,
    transfer_vol: float = 2.0,
    deck: JanusDeck = KURO_PRIMER_DECK,
) -> list[dict]:
    """Build JANUS transfer rows, one dict per transfer event.

    Single source of the row content behind the CSV export, the XLSX
    "primer_mapping file" sheet and the sidecar preview, so the three cannot
    drift apart. Row order is forward primers first (in ``fwd_mappings`` order),
    then reverse primers expanded so that every forward mutation gets its own
    reverse transfer row, aspirating from the shared source position when the
    reverse primer is shared. ``no`` runs 1..N across both blocks.

    Keys are the sidecar preview field names, plus:
      - ``mutation``: the mutation the row belongs to (not an instrument column).
      - ``role``: ``"fwd"`` or ``"rev"``. Stated so consumers do not have to
        infer direction from the source plate name, which is deck policy and not
        a stable direction marker.

    Args:
        fwd_mappings: Forward primer plate mappings.
        rev_mappings: Deduplicated reverse primer plate mappings.
        rev_groups: Reverse deduplication map {seq: [mutation_names]}.
        transfer_vol: Dispense volume in µL.
        deck: Plate-name and sample-type policy. Defaults to the KURO primer
            deck. ``deck.liquid_class`` is not written: the sheet has no column
            for it.
    """
    fwd_by_mut, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )

    rows: list[dict] = []
    seq_no = 1

    for m in fwd_mappings:
        rows.append({
            "name": f"{m.mutation}-F",
            "type": deck.sample_type,
            "no": seq_no,
            "asp_rack": deck.fwd_rack,
            "asp_posi": m.well,
            "dsp_rack": deck.dest_rack,
            "dsp_posi": m.well,
            "volume": transfer_vol,
            "mutation": m.mutation,
            "role": "fwd",
        })
        seq_no += 1

    for fwd_m in fwd_mappings:
        rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
        if rev_seq is None:
            continue
        rev_m = rev_by_seq.get(rev_seq)
        if rev_m is None:
            continue

        dest_well = fwd_by_mut.get(fwd_m.mutation, fwd_m.well)
        rows.append({
            "name": f"{fwd_m.mutation}-R",
            "type": deck.sample_type,
            "no": seq_no,
            "asp_rack": deck.rev_rack,
            "asp_posi": rev_m.well,
            "dsp_rack": deck.dest_rack,
            "dsp_posi": dest_well,
            "volume": transfer_vol,
            "mutation": fwd_m.mutation,
            "role": "rev",
        })
        seq_no += 1

    return rows


def janus_row_values(row: dict) -> list:
    """Flatten one ``build_janus_rows`` dict into ``JANUS_DEVICE_HEADER`` order."""
    return [row[key] for key in _JANUS_ROW_KEYS]


def export_janus_mapping_csv(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: float = 2.0,
    rev_groups: dict[str, list[str]] | None = None,
    encoding: str = "utf-8",
    deck: JanusDeck = KURO_PRIMER_DECK,
) -> None:
    """Export JANUS liquid handler mapping CSV.

    Rows come from :func:`build_janus_rows`, so this file, the XLSX
    "primer_mapping file" sheet and the sidecar preview always describe the same
    transfers. Which plate holds what is the *deck* argument, not a literal here.

    Shared reverse primers produce one row per destination well,
    all aspirating from the same source position.

    Args:
        fwd_mappings: Forward primer plate mappings.
        rev_mappings: Deduplicated reverse primer plate mappings.
        output_path: Output CSV file path.
        transfer_vol: Dispense volume in µL (default 2.0).
        rev_groups: Reverse deduplication map {seq: [mutation_names]}.
        encoding: File encoding (default "utf-8"; use "utf-8-sig" for BOM).
        deck: Plate-name and sample-type policy. Defaults to the KURO primer
            deck.
    """
    import csv

    rows = build_janus_rows(
        fwd_mappings, rev_mappings, rev_groups, transfer_vol, deck,
    )

    with open(output_path, "w", newline="", encoding=encoding) as f:
        writer = csv.writer(f)
        writer.writerow(JANUS_DEVICE_HEADER)
        for row in rows:
            writer.writerow(janus_row_values(row))


# ---------------------------------------------------------------------------
# XLSX liquid handler exports (transfer sheet + plate layout)
# ---------------------------------------------------------------------------

def _build_rev_lookups(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    rev_groups: dict[str, list[str]] | None,
) -> tuple[dict[str, str], dict[str, PlateMapping], dict[str, str]]:
    """Shared lookup tables for Echo/JANUS export."""
    fwd_by_mut = {m.mutation: m.well for m in fwd_mappings}
    rev_by_seq = {m.sequence: m for m in rev_mappings}
    mut_to_rev_seq: dict[str, str] = {}
    if rev_groups:
        for seq, muts in rev_groups.items():
            for mut in muts:
                mut_to_rev_seq[mut] = seq
    else:
        for m in rev_mappings:
            mut_to_rev_seq[m.mutation] = m.sequence
        # ``rev_mappings`` is already deduplicated, so it only carries the
        # representative mutation of each shared group. Mutations that are not
        # representatives have no entry, and every downstream writer would drop
        # their reverse transfer row without a trace (a reaction with no reverse
        # primer). Reconstruction is impossible here: forward mappings carry the
        # forward sequence only, so nothing links a non-representative mutation
        # to its reverse primer. Fail loudly instead of exporting a silently
        # incomplete instrument file.
        missing = [
            m.mutation for m in fwd_mappings
            if m.mutation not in mut_to_rev_seq
            or mut_to_rev_seq[m.mutation] not in rev_by_seq
        ]
        if missing:
            preview = ", ".join(missing[:10])
            if len(missing) > 10:
                preview += f", ... (+{len(missing) - 10} more)"
            raise ValueError(
                "Reverse primer grouping information (dedup_info) is missing, "
                f"so {len(missing)} mutation(s) cannot be matched to a reverse "
                f"primer: {preview}. Re-run the primer design so the reverse "
                "deduplication map is regenerated, then export again."
            )
    return fwd_by_mut, rev_by_seq, mut_to_rev_seq


def _write_96well_grid(
    ws, start_row: int, mappings: list[PlateMapping], label: str,
    labware: str = "96 deep well",
    value_attr: str = "primer_name",
) -> int:
    """Write a labelled 96-well plate grid and return the next free row.

    Args:
        value_attr: PlateMapping attribute to display in each cell
            ('primer_name' for source plates, 'mutation' for destination).
    """
    from openpyxl.styles import Alignment, Font

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    r = start_row

    ws.cell(row=r, column=1, value="lab ware").font = bold
    ws.cell(row=r, column=3, value=label)
    r += 1

    ws.cell(row=r, column=1, value=labware).font = bold
    for c in range(1, 13):
        cell = ws.cell(row=r, column=c + 2, value=c)
        cell.font = bold
        cell.alignment = center
    r += 1

    well_lookup: dict[str, str] = {}
    for m in mappings:
        _, base = _parse_well_plate(m.well)
        well_lookup[base] = getattr(m, value_attr)

    for row_letter in "ABCDEFGH":
        ws.cell(row=r, column=2, value=row_letter).font = bold
        for c in range(1, 13):
            name = well_lookup.get(f"{row_letter}{c}")
            if name:
                ws.cell(row=r, column=c + 2, value=name).alignment = center
        r += 1

    return r


def _rev_usage_counts(
    fwd_mappings: list[PlateMapping],
    rev_by_seq: dict[str, PlateMapping],
    mut_to_rev_seq: dict[str, str],
) -> dict[str, int]:
    """Count how many destination reactions each reverse sequence feeds.

    Counted from ``fwd_mappings`` (one destination per forward mutation) so the
    number always matches the transfer rows actually written.
    """
    counts: dict[str, int] = {seq: 0 for seq in rev_by_seq}
    for fwd_m in fwd_mappings:
        rev_seq = mut_to_rev_seq.get(fwd_m.mutation)
        if rev_seq is None or rev_seq not in counts:
            continue
        counts[rev_seq] += 1
    return counts


def _write_rev_usage_table(
    ws,
    start_row: int,
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    rev_by_seq: dict[str, PlateMapping],
    mut_to_rev_seq: dict[str, str],
    transfer_vol: float,
    unit: str,
) -> int:
    """Write the reverse source well usage table and return the next free row.

    Shared reverse primers occupy a single source well that the instrument
    aspirates once per destination reaction, so the volume leaving that well is
    ``shared_count x transfer_vol``. Dead volume is a labware property that KUMA
    does not know, so it is called out in a note instead of being added.
    """
    from openpyxl.styles import Font

    bold = Font(bold=True)
    counts = _rev_usage_counts(fwd_mappings, rev_by_seq, mut_to_rev_seq)

    r = start_row
    ws.cell(row=r, column=1, value="Reverse primer usage").font = bold
    r += 1

    headers = [
        "Source Well",
        "Primer Name",
        "Shared By (reactions)",
        f"Volume Per Reaction ({unit})",
        f"Total Transfer Volume ({unit})",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci, value=h).font = bold
    r += 1

    for m in rev_mappings:
        count = counts.get(m.sequence, 0)
        for ci, val in enumerate([
            m.well, m.primer_name, count, transfer_vol, count * transfer_vol,
        ], 1):
            ws.cell(row=r, column=ci, value=val)
        r += 1

    ws.cell(
        row=r, column=1,
        value=(
            "Total Transfer Volume excludes instrument dead volume. "
            "Fill each source well with this total plus the dead volume of the "
            "labware in use."
        ),
    )
    r += 1
    return r


def export_echo_mapping_xlsx(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: int = 100,
    rev_groups: dict[str, list[str]] | None = None,
    mapping_range: tuple[str, str] | None = None,
    quadrant: str | None = None,
    used_quadrants: list[str] | None = None,
) -> None:
    """Export Echo 525 mapping as XLSX matching the lab reference format.

    Sheets:
      - layout: 384-well source plate (Fwd odd rows + Rev even rows)
                + 96-well destination plate.
      - Echo mapping file: one row per transfer event, from
        :func:`build_echo_rows` (the same rows the CSV export and the sidecar
        preview show), so a single ``export_all`` cannot leave a csv and an xlsx
        that name different source wells for the same primer.

    ``mapping_range`` / ``quadrant`` / ``used_quadrants`` reach the worklist
    sheet only. The layout sheet keeps the row-doubled 384 view it always drew,
    which is a picture of the default plate and not of this transfer list.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    # Only the reverse-usage table below reads these; the transfer rows come
    # from ``build_echo_rows``, which builds its own.
    _, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )
    # Built before the workbook so a spent-quadrant refusal happens with no file
    # written, the way the CSV export already behaved.
    echo_rows = build_echo_rows(
        fwd_mappings, rev_mappings, rev_groups, transfer_vol,
        mapping_range=mapping_range,
        quadrant=quadrant,
        used_quadrants=used_quadrants,
    )
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    wb = Workbook()

    # ---- Sheet 1: layout ----
    ws = wb.active
    ws.title = "layout"

    # Title
    ws.cell(row=1, column=2, value="Primer dispensing (Echo 525)")

    # Source section header
    ws.cell(row=2, column=1, value="Source").font = bold
    ws.cell(row=3, column=1, value="labware")
    ws.cell(row=3, column=3, value="source plate")

    # "Eco 384PP" + column numbers 1-24
    ws.cell(row=4, column=1, value="Eco 384PP").font = bold
    for c in range(1, 25):
        cell = ws.cell(row=4, column=c + 2, value=c)
        cell.font = bold
        cell.alignment = center

    # Build 384-well lookup: well_384 → primer_name
    well_384: dict[str, str] = {}
    for m in fwd_mappings:
        _, base = _parse_well_plate(m.well)
        well_384[_to_384_well_fwd(base)] = m.primer_name
    for m in rev_mappings:
        _, base = _parse_well_plate(m.well)
        well_384[_to_384_well_rev(base)] = m.primer_name

    # 384-well grid: rows A-P (16 rows)
    for ri, row_letter in enumerate(_ROWS_384):
        r = 5 + ri
        ws.cell(row=r, column=2, value=row_letter).font = bold
        for c in range(1, 25):
            name = well_384.get(f"{row_letter}{c}")
            if name:
                ws.cell(row=r, column=c + 2, value=name).alignment = center

    # Destination section (96-well PCR plate)
    dest_start = 5 + len(_ROWS_384) + 1  # after 384-well grid + blank row
    ws.cell(row=dest_start, column=1, value="Destination").font = bold
    next_row = _write_96well_grid(
        ws, dest_start + 1, fwd_mappings, "PCR mixture",
        labware="96 PCR plate", value_attr="mutation",
    )

    # Reverse source well usage (shared count + total transfer volume)
    _write_rev_usage_table(
        ws, next_row + 1, fwd_mappings, rev_mappings,
        rev_by_seq, mut_to_rev_seq, transfer_vol, "nL",
    )

    # Auto-width
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 4
    for c in range(3, 27):
        ws.column_dimensions[chr(64 + c)].width = 12

    # ---- Sheet 2: Echo mapping file ----
    ws2 = wb.create_sheet("Echo mapping file")
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    for col_idx, h in enumerate(ECHO_DEVICE_HEADER, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill

    for row_num, row in enumerate(echo_rows, 2):
        for ci, val in enumerate(echo_row_values(row), 1):
            ws2.cell(row=row_num, column=ci, value=val)

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_path)


def export_janus_mapping_xlsx(
    fwd_mappings: list[PlateMapping],
    rev_mappings: list[PlateMapping],
    output_path: Path,
    transfer_vol: float = 2.0,
    rev_groups: dict[str, list[str]] | None = None,
    deck: JanusDeck = KURO_PRIMER_DECK,
) -> None:
    """Export JANUS mapping as XLSX matching the lab reference format.

    Sheets:
      - layout: Fwd 96-well plate + Rev 96-well plate
                + 96-well PCR mixture (destination) on a single sheet.
      - primer_mapping file: one row per transfer event, from
        :func:`build_janus_rows` (same rows the CSV export and the sidecar
        preview use).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    # The mapping rows come from ``build_janus_rows``; these lookups feed the
    # layout sheet reverse-usage table only.
    _, rev_by_seq, mut_to_rev_seq = _build_rev_lookups(
        fwd_mappings, rev_mappings, rev_groups,
    )
    bold = Font(bold=True)

    wb = Workbook()

    # ---- Sheet 1: layout ----
    ws = wb.active
    ws.title = "layout"

    # Title
    ws.cell(row=1, column=2, value="Primer dispensing (JANUS)")

    # Fwd plate
    r = _write_96well_grid(ws, 2, fwd_mappings, "fw plate")

    # Blank row + Rev plate
    r += 1
    r = _write_96well_grid(ws, r, rev_mappings, "rv plate")

    # Blank rows + Destination plate
    r += 2
    next_row = _write_96well_grid(
        ws, r, fwd_mappings, "PCR mixture plate",
        labware="96 PCR plate", value_attr="mutation",
    )

    # Reverse source well usage (shared count + total transfer volume)
    _write_rev_usage_table(
        ws, next_row + 1, fwd_mappings, rev_mappings,
        rev_by_seq, mut_to_rev_seq, transfer_vol, "µL",
    )

    # Auto-width
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 4
    for c in range(3, 15):
        ws.column_dimensions[chr(64 + c)].width = 12

    # ---- Sheet 2: primer_mapping file ----
    ws2 = wb.create_sheet("primer_mapping file")
    header_fill = PatternFill(start_color="D9E1F2", fill_type="solid")
    for col_idx, h in enumerate(JANUS_DEVICE_HEADER, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill

    for row_num, row in enumerate(
        build_janus_rows(fwd_mappings, rev_mappings, rev_groups, transfer_vol, deck),
        2,
    ):
        for ci, val in enumerate(janus_row_values(row), 1):
            ws2.cell(row=row_num, column=ci, value=val)

    for col in ws2.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Macrogen Plate Oligo .xls exporter
# ---------------------------------------------------------------------------

import re as _re

_MACROGEN_NAME_RE = _re.compile(r"^[A-Za-z0-9_-]{1,20}$")
_MACROGEN_HEADERS = [
    "No.",
    "Plate Name",
    "Well",
    "Oligo Name",
    "5' - Oligo Seq - 3'",
    "Amount",
    "Purification",
]


def _macrogen_column_major_wells() -> list[str]:
    """Return 96-well names in column-major order (A1..H1, A2..H2, ...)."""
    return [f"{r}{c}" for c in range(1, 13) for r in "ABCDEFGH"]


def _validate_macrogen_plate_name(name: str, label: str) -> None:
    if not _MACROGEN_NAME_RE.fullmatch(name):
        raise ValueError(
            f"{label} plate name '{name}' violates ^[A-Za-z0-9_-]{{1,20}}$"
        )


def _validate_macrogen_oligo_names(primers, label: str) -> None:
    for p in primers:
        if not _MACROGEN_NAME_RE.fullmatch(p.primer_name):
            raise ValueError(
                f"{label} oligo name '{p.primer_name}' violates ^[A-Za-z0-9_-]{{1,20}}$"
            )


def export_macrogen_xls(
    fwd_primers,
    rev_primers,
    fwd_plate_name: str,
    rev_plate_name: str,
    amount: str,
    purification: str,
    output_path: str,
) -> None:
    """Export forward/reverse primer plates to a Macrogen Plate Oligo .xls file.

    Layout: 1 header row, then 96 rows per plate (column-major wells).
    Empty wells receive only No./Plate/Well; oligo columns blank.
    """
    import xlwt

    if len(fwd_primers) > 96:
        raise ValueError(
            f"fwd primer count {len(fwd_primers)} exceeds 96 well limit"
        )
    if len(rev_primers) > 96:
        raise ValueError(
            f"rev primer count {len(rev_primers)} exceeds 96 well limit"
        )

    plates = []
    if fwd_primers:
        _validate_macrogen_plate_name(fwd_plate_name, "fwd")
        _validate_macrogen_oligo_names(fwd_primers, "fwd")
        plates.append((fwd_plate_name, fwd_primers))
    if rev_primers:
        _validate_macrogen_plate_name(rev_plate_name, "rev")
        _validate_macrogen_oligo_names(rev_primers, "rev")
        plates.append((rev_plate_name, rev_primers))

    wb = xlwt.Workbook(encoding="utf-8")
    sheet = wb.add_sheet("Sheet")
    for c, h in enumerate(_MACROGEN_HEADERS):
        sheet.write(0, c, h)

    wells = _macrogen_column_major_wells()
    row = 1
    no = 1
    for plate_name, primers in plates:
        for i, well in enumerate(wells):
            sheet.write(row, 0, no)
            sheet.write(row, 1, plate_name)
            sheet.write(row, 2, well)
            if i < len(primers):
                p = primers[i]
                sheet.write(row, 3, p.primer_name)
                sheet.write(row, 4, p.sequence)
                sheet.write(row, 5, amount)
                sheet.write(row, 6, purification)
            row += 1
            no += 1
    wb.save(output_path)
