"""96-well plate mapping and Excel export for SDM primers."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .sdm_engine import SdmPrimerResult


@dataclass
class PlateMapping:
    """Single well assignment in a 96-well plate."""

    well: str               # e.g. "A1", "B1"
    primer_name: str        # e.g. "Q232A_F" or "Q232A_R"
    sequence: str           # Primer sequence
    primer_type: str        # "forward" or "reverse"
    mutation: str           # Original mutation notation


def _well_name(index: int, order: str = "column") -> str:
    """Convert a 0-based index to a well name.

    Args:
        index: 0-based well index.
        order: 'column' (A1->H1->A2) or 'row' (A1->A12->B1).

    Returns:
        Well name like "A1".
    """
    rows = "ABCDEFGH"
    if order == "column":
        col = index // 8 + 1
        row = index % 8
    else:  # row
        row = index // 12
        col = index % 12 + 1
    if row >= 8 or col > 12:
        raise ValueError(f"Well index {index} exceeds 96-well plate capacity")
    return f"{rows[row]}{col}"


def _assign_well(index: int, well_order: str = "column") -> str:
    """Generate well name with overflow to second plate."""
    if index < 96:
        return _well_name(index, well_order)
    return f"P2-{_well_name(index - 96, well_order)}"


def deduplicate_reverse(
    results: list[SdmPrimerResult],
) -> dict[str, list[str]]:
    """Find mutations sharing identical reverse primers.

    Returns:
        Dict mapping reverse primer sequence to list of mutation names.
    """
    rev_map: dict[str, list[str]] = {}
    for r in results:
        rev_seq = r.reverse_seq
        if rev_seq not in rev_map:
            rev_map[rev_seq] = []
        rev_map[rev_seq].append(r.mutation.raw)
    return rev_map


def generate_plate_map(
    results: list[SdmPrimerResult],
    well_order: str = "column",
    deduplicate_rev: bool = True,
) -> tuple[list[PlateMapping], list[PlateMapping]]:
    """Generate separate Fwd and Rev plate mappings.

    Forward plate: one primer per mutation, sequential well assignment.
    Reverse plate: deduplicated reverse primers, sequential well assignment.

    Returns:
        Tuple of (fwd_mappings, rev_mappings).
    """
    # Forward plate
    fwd_mappings: list[PlateMapping] = []
    for idx, r in enumerate(results):
        well = _assign_well(idx, well_order)
        fwd_mappings.append(PlateMapping(
            well=well,
            primer_name=f"{r.mutation.raw}_F",
            sequence=r.forward_seq,
            primer_type="forward",
            mutation=r.mutation.raw,
        ))

    # Reverse plate (deduplicated)
    rev_mappings: list[PlateMapping] = []
    if deduplicate_rev:
        rev_groups = deduplicate_reverse(results)
        for idx, (rev_seq, mut_names) in enumerate(rev_groups.items()):
            label = mut_names[0]
            well = _assign_well(idx, well_order)
            rev_mappings.append(PlateMapping(
                well=well,
                primer_name=f"{label}_R",
                sequence=rev_seq,
                primer_type="reverse",
                mutation=label,
            ))
    else:
        for idx, r in enumerate(results):
            well = _assign_well(idx, well_order)
            rev_mappings.append(PlateMapping(
                well=well,
                primer_name=f"{r.mutation.raw}_R",
                sequence=r.reverse_seq,
                primer_type="reverse",
                mutation=r.mutation.raw,
            ))

    return fwd_mappings, rev_mappings


def _is_shared_rev(m: PlateMapping, rev_groups: dict[str, list[str]] | None) -> bool:
    """Check if a reverse primer is shared by multiple mutations."""
    if m.primer_type != "reverse" or rev_groups is None:
        return False
    return len(rev_groups.get(m.sequence, [])) > 1


# Colors matching the UI: Fwd=green, Rev=orange, Shared=blue
_COLOR_FWD = "C6EFCE"       # light green
_COLOR_REV = "FCE4D6"       # light orange
_COLOR_SHARED = "BDD7EE"    # light blue


def _write_list_sheet(
    ws,
    mappings: list[PlateMapping],
    fill_color: str,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Write a primer list sheet with row coloring."""
    from openpyxl.styles import Font, PatternFill

    headers = ["Well", "Primer Name", "Sequence", "Length", "Mutation"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")

    for i, m in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=m.well)
        ws.cell(row=i, column=2, value=m.primer_name)
        ws.cell(row=i, column=3, value=m.sequence)
        ws.cell(row=i, column=4, value=len(m.sequence))
        ws.cell(row=i, column=5, value=m.mutation)

        # Row color: shared=blue, otherwise fill_color
        row_color = _COLOR_SHARED if _is_shared_rev(m, rev_groups) else fill_color
        row_fill = PatternFill(start_color=row_color, fill_type="solid")
        for col_idx in range(1, 6):
            ws.cell(row=i, column=col_idx).fill = row_fill

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def _write_plate_sheet(
    ws,
    mappings: list[PlateMapping],
    fill_color: str,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Write a 96-well plate layout sheet with shared primers in blue."""
    from openpyxl.styles import Alignment, Font, PatternFill

    rows_label = "ABCDEFGH"

    for c in range(1, 13):
        cell = ws.cell(row=1, column=c + 1, value=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r_idx, r_label in enumerate(rows_label):
        ws.cell(row=r_idx + 2, column=1, value=r_label).font = Font(bold=True)

    default_fill = PatternFill(start_color=fill_color, fill_type="solid")
    shared_fill = PatternFill(start_color=_COLOR_SHARED, fill_type="solid")

    for m in mappings:
        if not m.well[0].isalpha():
            continue
        row_letter = m.well[0]
        col_num = int(m.well[1:])
        if row_letter not in rows_label or col_num > 12:
            continue
        r_idx = rows_label.index(row_letter) + 2
        c_idx = col_num + 1

        # Display mutation name (without _F/_R suffix) like the UI
        display_name = m.primer_name.rsplit("_", 1)[0] if "_" in m.primer_name else m.primer_name

        cell = ws.cell(row=r_idx, column=c_idx, value=display_name)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = shared_fill if _is_shared_rev(m, rev_groups) else default_fill

    ws.column_dimensions["A"].width = 4
    for c in range(2, 14):
        ws.column_dimensions[chr(64 + c)].width = 14


def _chunk_by_plate(mappings: list[PlateMapping]) -> list[list[PlateMapping]]:
    """Split mappings into 96-well plate chunks."""
    plates: list[list[PlateMapping]] = []
    current: list[PlateMapping] = []
    for m in mappings:
        # Normalize well: strip overflow prefix like "P2-"
        well = m.well
        if "-" in well:
            well = well.split("-", 1)[1]
            m = PlateMapping(
                well=well,
                primer_name=m.primer_name,
                sequence=m.sequence,
                primer_type=m.primer_type,
                mutation=m.mutation,
            )
        current.append(m)
        if len(current) >= 96:
            plates.append(current)
            current = []
    if current:
        plates.append(current)
    return plates if plates else [[]]


def export_plate_excel(
    mappings: list[PlateMapping],
    output_path: Path,
    rev_groups: dict[str, list[str]] | None = None,
) -> None:
    """Export plate mappings to Excel with per-plate sheets.

    For N plates, creates sheets:
    - 'Fwd List 1', 'Fwd Plate 1', 'Rev List 1', 'Rev Plate 1'
    - 'Fwd List 2', 'Fwd Plate 2', ... (if multi-plate)

    Single plate omits the number suffix.

    Args:
        rev_groups: Original reverse deduplication map (seq -> mutation names).
            Used to detect shared reverse primers for blue coloring.
    """
    from openpyxl import Workbook

    fwd_all = [m for m in mappings if m.primer_type == "forward"]
    rev_all = [m for m in mappings if m.primer_type == "reverse"]

    fwd_plates = _chunk_by_plate(fwd_all)
    rev_plates = _chunk_by_plate(rev_all)
    plate_count = max(len(fwd_plates), len(rev_plates))
    suffix = plate_count > 1

    wb = Workbook()
    first_sheet = True

    for i in range(plate_count):
        tag = f" {i + 1}" if suffix else ""
        fwd_chunk = fwd_plates[i] if i < len(fwd_plates) else []
        rev_chunk = rev_plates[i] if i < len(rev_plates) else []

        # Use original rev_groups for shared detection (covers deduplicated entries)
        chunk_rev_groups = rev_groups or {}

        if first_sheet:
            ws = wb.active
            ws.title = f"Fwd List{tag}"
            first_sheet = False
        else:
            ws = wb.create_sheet(f"Fwd List{tag}")
        _write_list_sheet(ws, fwd_chunk, _COLOR_FWD)

        ws = wb.create_sheet(f"Fwd Plate{tag}")
        _write_plate_sheet(ws, fwd_chunk, _COLOR_FWD)

        ws = wb.create_sheet(f"Rev List{tag}")
        _write_list_sheet(ws, rev_chunk, _COLOR_REV, rev_groups=chunk_rev_groups)

        ws = wb.create_sheet(f"Rev Plate{tag}")
        _write_plate_sheet(ws, rev_chunk, _COLOR_REV, rev_groups=chunk_rev_groups)

    wb.save(output_path)
