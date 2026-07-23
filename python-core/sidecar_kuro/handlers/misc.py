"""Handlers: polymerase list, organism list, EVOLVEpro CSV, benchmark."""

import csv
from dataclasses import asdict
from pathlib import Path

from kuma_core.kuro.evolvepro import load_evolvepro_csv
from kuma_core.kuro.polymerase import _dict_to_profile

import sidecar_kuro.core as _core
from sidecar_kuro.core import (
    _validate_filepath,
    _poly_registry,
    _codon_registry,
    _CUSTOM_POLYMERASE_PATH,
    _get_cached_ca_coords,
    _get_cached_ca_seq,
)
from sidecar_kuro.models import (
    LoadEvolveproParams,
    PolymeraseProfileModel,
    PreviewEvolveproSourceParams,
    RunBenchmarkParams,
    SaveCustomPolymeraseResultModel,
)

# EVOLVEpro-specific table extensions: CSV, TSV, and XLSX.
# .txt is excluded intentionally (ambiguous delimiter, not supported for column mapping).
_ALLOWED_TABLE_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls"}


_POLYMERASE_META = {
    "Benchling": {"manufacturer": "SantaLucia 1998", "fidelity": "standard"},
    "Q5": {"manufacturer": "NEB", "fidelity": "high"},
    "Phusion": {"manufacturer": "Thermo", "fidelity": "high"},
    "Taq": {"manufacturer": "Various", "fidelity": "low"},
    "DreamTaq": {"manufacturer": "Thermo", "fidelity": "low"},
    "KOD": {"manufacturer": "Toyobo", "fidelity": "high"},
    "TAKARA_GXL": {"manufacturer": "Takara", "fidelity": "high"},
    "Q5 SDM": {"manufacturer": "NEB", "fidelity": "high"},
}


def handle_list_polymerases(_params: dict) -> list[dict]:
    """Return available polymerase profiles."""
    names = _poly_registry.list_names()
    result = []
    for name in names:
        meta = _POLYMERASE_META.get(name, {"manufacturer": "", "fidelity": ""})
        result.append(
            {
                "name": name,
                "manufacturer": meta["manufacturer"],
                "fidelity": meta["fidelity"],
            }
        )
    return result


def handle_get_polymerase_details(params: dict) -> dict:
    """Return full polymerase profile for the selected name."""
    name = params.get("name", "")
    profile = PolymeraseProfileModel.model_validate(asdict(_poly_registry.get(name)))
    return profile.to_rpc_dict()


def handle_save_custom_polymerase(params: dict) -> dict:
    """Persist a custom polymerase profile and keep it available after restart."""
    profile = _dict_to_profile(params)
    _CUSTOM_POLYMERASE_PATH.parent.mkdir(parents=True, exist_ok=True)
    _poly_registry.save_custom(profile, _CUSTOM_POLYMERASE_PATH)
    return SaveCustomPolymeraseResultModel(name=profile.name).to_rpc_dict()


def handle_list_organisms(_params: dict) -> list[dict]:
    """Return available organism codon tables for the UI dropdown."""
    return _codon_registry.list_organisms_detailed()


def _preview_csv(filepath: str, max_rows: int) -> dict:
    """Read headers and first max_rows data rows from a CSV or TSV file."""
    delimiter = "\t" if Path(filepath).suffix.lower() == ".tsv" else ","
    with open(filepath, encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        headers = next(reader, [])
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append([str(c) for c in row])
    return {"sheets": [], "headers": headers, "rows": rows}


def _preview_xlsx(filepath: str, sheet_name: str | None, max_rows: int) -> dict:
    """Read sheet names, headers, and first max_rows rows from an XLSX file."""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    target = sheet_name if sheet_name is not None else sheet_names[0]
    if target not in sheet_names:
        wb.close()
        raise ValueError(
            f"sheet '{target}' not found in {filepath}. "
            f"Available sheets: {sheet_names}"
        )

    ws = wb[target]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return {"sheets": sheet_names, "headers": [], "rows": []}

    headers = [str(c) if c is not None else "" for c in all_rows[0]]
    data_rows = [
        [str(c) if c is not None else "" for c in row]
        for row in all_rows[1 : max_rows + 1]
    ]
    return {"sheets": sheet_names, "headers": headers, "rows": data_rows}


def _preview_xls(filepath: str, sheet_name: str | None, max_rows: int) -> dict:
    """Read sheet names, headers, and first max_rows rows from a legacy XLS file."""
    import xlrd

    wb = xlrd.open_workbook(filepath)
    sheet_names = wb.sheet_names()
    target = sheet_name if sheet_name is not None else sheet_names[0]
    if target not in sheet_names:
        raise ValueError(
            f"sheet '{target}' not found in {filepath}. "
            f"Available sheets: {sheet_names}"
        )

    ws = wb.sheet_by_name(target)
    if ws.nrows == 0:
        return {"sheets": sheet_names, "headers": [], "rows": []}

    headers = [str(c.value) if c.value is not None else "" for c in ws.row(0)]
    rows = []
    for ridx in range(1, min(ws.nrows, max_rows + 1)):
        rows.append([str(c.value) if c.value is not None else "" for c in ws.row(ridx)])
    return {"sheets": sheet_names, "headers": headers, "rows": rows}


def handle_preview_evolvepro_source(params: dict) -> dict:
    """Preview first rows and column headers of a CSV or XLSX EVOLVEpro source."""
    p = PreviewEvolveproSourceParams(**params)
    resolved = _validate_filepath(p.filepath, allowed_extensions=_ALLOWED_TABLE_EXTENSIONS)
    ext = Path(str(resolved)).suffix.lower()
    if ext == ".xlsx":
        return _preview_xlsx(str(resolved), p.sheet_name, p.max_rows)
    if ext == ".xls":
        return _preview_xls(str(resolved), p.sheet_name, p.max_rows)
    return _preview_csv(str(resolved), p.max_rows)


def _frame_checked_ca_coords(
    ca_coords: list | None,
    structure_accession: str | None,
    ref_seq: str,
) -> tuple[list | None, bool]:
    """Drop cached coordinates that do not cover the reference frame.

    3D coordinates are indexed by reference-sequence position, so a structure
    that is not identical to (or a clean substring of) the reference would place
    Cα coordinates on the wrong residues, silently corrupting any 3D-aware
    selection. When it does not match, return None so callers fall back to 1-D
    distance, plus a flag the UI can surface. Both load_evolvepro_csv and
    run_benchmark share this so their handling cannot diverge.

    A user-supplied structure caches its own sequence, so use that and skip the
    network. Only fall back to fetching when the coordinates came from an
    accession, which has no cached sequence.
    """
    if ca_coords is None or not structure_accession or not ref_seq.strip():
        return ca_coords, False
    from kuma_core.kuro.interface import structure_matches_reference

    structure_seq = _get_cached_ca_seq(structure_accession)
    if structure_seq is None:
        from kuma_core.kuro.alphafold import fetch_ca_seq

        structure_seq = fetch_ca_seq(structure_accession)
    if structure_seq and not structure_matches_reference(structure_seq, ref_seq):
        return None, True
    return ca_coords, False


def handle_load_evolvepro_csv(params: dict) -> dict:
    """Load EVOLVEpro df_test.csv, sort by y_pred descending, return top-N variants."""
    p = LoadEvolveproParams(**params)
    if not p.filepath:
        raise ValueError("filepath is required")
    resolved = _validate_filepath(p.filepath, allowed_extensions=_ALLOWED_TABLE_EXTENSIONS)

    ca_coords = _get_cached_ca_coords(p.structure_accession)

    # Structure-accuracy guard: 3D coordinates are only valid when the loaded
    # structure exactly covers the reference frame (identity or clean substring).
    # A near-but-not-exact structure would place Cα coordinates on the wrong
    # residues, silently corrupting dispersion / structural-diversity selection.
    # When it does not match, drop to None so those paths fall back to 1-D
    # distance, and record why for the UI.
    ca_coords, structure_frame_mismatch = _frame_checked_ca_coords(
        ca_coords, p.structure_accession, p.ref_seq
    )
    result = load_evolvepro_csv(
        filepath=str(resolved),
        top_n=p.top_n,
        max_per_position=p.max_per_position,
        domains=p.domains,
        excluded_ranges=[{"start": r.start, "end": r.end} for r in p.excluded_ranges],
        domain_diversity=p.domain_diversity,
        domain_strategy=p.domain_strategy,
        domain_overlap_policy=p.domain_overlap_policy,
        linker_handling=p.linker_handling,
        domain_quota_min=p.domain_quota_min,
        pareto_diversity=p.pareto_diversity,
        entropy_weight=p.entropy_weight,
        pool_multiplier=p.pool_multiplier,
        distance_mode=p.distance_mode,
        ca_coords=ca_coords,
        evolvepro_round=p.evolvepro_round,
        round_size=p.round_size,
        ref_seq=p.ref_seq,
        variant_column=p.variant_column,
        score_column=p.score_column,
        score_order=p.score_order,
        sheet_name=p.sheet_name,
        domain_pool_autoexpand=p.domain_pool_autoexpand,
        domain_pool_max_multiplier=p.domain_pool_max_multiplier,
        structural_diversity=p.structural_diversity,
        structural_kappa=p.structural_kappa,
        anchor_variants=p.anchor_variants,
    )
    result["structure_frame_mismatch"] = structure_frame_mismatch
    return result


def handle_run_benchmark(params: dict) -> dict:
    """Run benchmark simulation on provided fitness landscape."""
    from kuma_core.kuro.benchmark import run_benchmark

    p = RunBenchmarkParams(**params)

    if not p.landscape:
        raise ValueError("landscape data is required")

    landscape = [(v.variant, v.fitness) for v in p.landscape]

    ca_coords = _get_cached_ca_coords(p.structure_accession)
    # Same frame guard load_evolvepro_csv uses: a structure that does not cover
    # the reference frame must not feed 3D coordinates into the benchmark.
    ca_coords, structure_frame_mismatch = _frame_checked_ca_coords(
        ca_coords, p.structure_accession, p.ref_seq
    )

    bench_results = run_benchmark(
        landscape,
        p.ground_truth,
        p.n_select,
        n_random_trials=p.n_random_trials,
        top_percentile=p.top_percentile,
        strategies=p.strategies,
        domains=p.domains,
        domain_strategy=p.domain_strategy,
        max_per_position=p.max_per_position,
        entropy_weight=p.entropy_weight,
        pool_multiplier=p.pool_multiplier,
        distance_mode=p.distance_mode,
        random_seed=p.random_seed,
        ca_coords=ca_coords,
    )
    return {"results": bench_results, "structure_frame_mismatch": structure_frame_mismatch}
