"""Handlers: Excel/CSV export, plate map, workspace save/load."""

import csv
import json
from dataclasses import fields as dc_fields
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace

import openpyxl

from kuma_core.kuro.plate_mapper import (
    PlateMapping,
    deduplicate_reverse,
    export_idt_csv,
    export_echo_mapping_csv,
    export_echo_mapping_xlsx,
    export_janus_mapping_csv,
    export_janus_mapping_xlsx,
    export_plate_excel,
    export_twist_csv,
    generate_plate_map,
)
from kuma_core.shared.version import KUMA_VERSION, KURO_MODULE_VERSION
from kuma_core.shared.run_manifest import (
    build_run_manifest,
    write_run_manifest,
)
from kuma_core.shared.output_hash import write_output_checksum

import sidecar_kuro.core as _core
from sidecar_kuro.core import (
    logger,
    _validate_filepath,
    _validate_output_path,
    _ALLOWED_EXCEL_EXTENSIONS,
    _ALLOWED_CSV_EXTENSIONS,
)
from sidecar_kuro.models import (
    ExportExcelParams,
    ExportMappingResultModel,
    ExportMappingParams,
    ExportMappingDryRunParams,
    ExportOrderParams,
    ExportOrderResultModel,
    ExportBenchmarkCsvParams,
    FileExportResultModel,
    SaveWorkspaceParams,
    SaveJsonParams,
    LoadWorkspaceParams,
    validate_workspace_data,
)

_ALLOWED_MAPPING_EXTENSIONS = {".xlsx", ".csv"}

_PLATE_MAPPING_KEYS = {f.name for f in dc_fields(PlateMapping)}


def _resolve_rev_groups(
    param_dedup: dict | None,
    state_dedup: dict | None,
    results,
    *,
    context: str,
) -> dict:
    """Resolve the reverse deduplication map used by every mapping exporter.

    Fallback ladder:
      1. ``dedup_info`` shipped by the frontend,
      2. ``dedup_info`` held in sidecar state,
      3. recomputation from the design ``results``,
      4. empty dict, which lets ``plate_mapper`` raise its explicit error.

    Step 3 covers workspaces saved before ``dedupInfo`` was persisted: the
    frontend migration fills the field with ``{}``, which would otherwise
    block the export outright. Recomputation is logged, never silent.

    This helper must not touch ``_core._state_lock``: callers already read
    state under that non-reentrant lock and pass the values in.

    Args:
        param_dedup: dedup map supplied with the request, if any.
        state_dedup: dedup map currently held in sidecar state, if any.
        results: design results (needs ``mutation.raw`` and ``reverse_seq``).
        context: handler name used in the recomputation log line.
    """
    if param_dedup:
        return dict(param_dedup)
    if state_dedup:
        return dict(state_dedup)
    if not results:
        return {}
    recomputed = deduplicate_reverse(results)
    if not recomputed:
        return {}
    logger.warning(
        "%s: dedup_info missing, reverse grouping recomputed from %d design "
        "result(s) into %d group(s). This happens with workspaces saved "
        "before dedup_info was persisted.",
        context, len(results), len(recomputed),
    )
    return recomputed


def _write_report_sheet(wb: openpyxl.Workbook, report_data: dict) -> None:
    if "Report" in wb.sheetnames:
        del wb["Report"]
    ws = wb.create_sheet("Report")
    ws.append(["Section", "Label", "Value", "Warn"])

    for section in report_data.get("sections", []):
        title = section.get("title", "")
        for item in section.get("items", []):
            ws.append([
                title,
                item.get("label", ""),
                str(item.get("value", "")),
                "Y" if item.get("warn") else "",
            ])


def _write_benchmark_raw_sheet(wb: openpyxl.Workbook, benchmark_raw: dict) -> None:
    if "Benchmark Raw" in wb.sheetnames:
        del wb["Benchmark Raw"]
    ws = wb.create_sheet("Benchmark Raw")

    ws.append(["Benchmark Raw Export"])
    ws.append(["exported_at", benchmark_raw.get("exported_at", "")])
    ws.append([])

    ws.append(["Settings"])
    ws.append(["key", "value"])
    for key, value in (benchmark_raw.get("settings") or {}).items():
        ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value])
    ws.append([])

    ws.append(["Domains"])
    ws.append(["bucket", "name", "id", "start", "end", "db"])
    domains = benchmark_raw.get("domains") or {}
    for bucket in ("active", "excluded"):
        for domain in domains.get(bucket, []):
            ws.append([
                bucket,
                domain.get("name", ""),
                domain.get("id", ""),
                domain.get("start", ""),
                domain.get("end", ""),
                domain.get("db", ""),
            ])
    ws.append([])

    ws.append(["Results"])
    ws.append([
        "strategy", "n_selected", "hit_rate", "mean_fitness", "unique_positions",
        "position_coverage", "domain_coverage", "structural_spread", "hits",
        "threshold", "n_trials",
    ])
    for strategy, metrics in (benchmark_raw.get("results") or {}).items():
        ws.append([
            strategy,
            metrics.get("n_selected", ""),
            metrics.get("hit_rate", ""),
            metrics.get("mean_fitness", ""),
            metrics.get("unique_positions", ""),
            metrics.get("position_coverage", ""),
            metrics.get("domain_coverage", ""),
            metrics.get("structural_spread", ""),
            metrics.get("hits", ""),
            metrics.get("threshold", ""),
            metrics.get("n_trials", ""),
        ])
    ws.append([])

    ws.append(["Landscape"])
    ws.append(["variant", "fitness"])
    for row in benchmark_raw.get("landscape", []):
        ws.append([row.get("variant", ""), row.get("fitness", "")])


def _pydantic_to_plate_mappings(items) -> list[PlateMapping]:
    """Convert a list of Pydantic PlateMappingItem objects to PlateMapping dataclasses."""
    return [
        PlateMapping(**{k: v for k, v in m.model_dump().items() if k in _PLATE_MAPPING_KEYS})
        for m in items
    ]


def _resolve_mapping_transfer_volume(fmt: str, transfer_vol: float | None) -> int | float:
    """Validate and normalize mapping transfer volume by instrument."""
    if fmt == "echo":
        if transfer_vol is None:
            return 100
        if transfer_vol <= 0:
            raise ValueError("Echo transfer volume must be greater than 0 nL.")
        if not float(transfer_vol).is_integer():
            raise ValueError("Echo transfer volume must be a whole number of nL.")
        return int(transfer_vol)

    if transfer_vol is None:
        return 2.0
    if transfer_vol <= 0:
        raise ValueError("JANUS transfer volume must be greater than 0 uL.")
    return float(transfer_vol)


def _manifest_path_for(output_path: Path) -> Path:
    """Return the sibling ``.run.json`` path for *output_path*.

    Examples:
        /out/primers.xlsx -> /out/primers.run.json
        /out/order.csv   -> /out/order.run.json
    """
    return output_path.parent / (output_path.stem + ".run.json")


def _design_provenance_for_manifest(
    results_source: str,
) -> tuple[dict[str, Path], dict[str, object]]:
    """Manifest ``inputs`` and ``extra`` describing what produced these rows.

    *results_source* says where the rows in this export came from: ``"state"``
    when they were read out of the session's last design, ``"payload"`` when the
    caller handed them in.

    On the payload path the session provenance is deliberately NOT attached.
    ``handle_load_workspace`` restores a saved workspace without repopulating
    ``_core._state.results``, so after a load the last design this process ran
    can easily be a different one, and stamping its fasta digest and its
    intervention log onto an unrelated export would put a confident falsehood in
    the one artifact whose whole job is to be trusted. An empty ``inputs`` reads
    as "not recorded"; a wrong ``inputs`` reads as "recorded", which is worse.

    Everything here was fixed at design time (see
    ``handlers.design._build_design_provenance``). This function only serialises
    it; the digests it returns paths for are re-taken by ``build_run_manifest``
    at export time on purpose, so a fasta edited between designing and exporting
    shows up as two digests that disagree instead of passing unnoticed.
    """
    if results_source != "state":
        return {}, {
            "results_source": results_source,
            "design": None,
            "interventions": None,
            "provenance_omitted": (
                "rows supplied by the caller, so the session design provenance "
                "may describe a different run"
            ),
        }

    provenance, interventions = _core._provenance_snapshot()
    inputs: dict[str, Path] = {}
    if provenance:
        fasta_path = provenance.get("fasta_path")
        if fasta_path:
            inputs["design_fasta"] = Path(str(fasta_path))
        mutations = provenance.get("mutations") or {}
        # Only a real file. Mutations typed into the UI went through a temporary
        # CSV this process already deleted, and naming a path that no longer
        # exists would be dropped from the manifest and become indistinguishable
        # from no mutations at all. Those are recorded inline under extra.design.
        if mutations.get("source") == "file" and mutations.get("path"):
            inputs["design_mutations"] = Path(str(mutations["path"]))

    return inputs, {
        "results_source": "state",
        "design": provenance,
        "interventions": interventions,
    }


def handle_get_plate_map(_params: dict) -> dict:  # noqa: ARG001
    """Return the plate map from last design."""
    with _core._state_lock:
        if not _core._state.results:
            raise ValueError("No design available. Run design_sdm_primers first.")

        return {
            "mappings": [
                {
                    "well": m.well,
                    "primer_name": m.primer_name,
                    "sequence": m.sequence,
                    "primer_type": m.primer_type,
                    "mutation": m.mutation,
                }
                for m in _core._state.plate_mappings
            ],
            "dedup_info": _core._state.dedup_info,
        }


def handle_export_excel(params: dict) -> dict:
    """Export plate map to Excel.

    Accepts optional 'mappings' and 'dedup_info' from the frontend to reflect
    the current UI state (sorted order, custom additions from failed mutations).
    Falls back to backend state only when 'mappings' is absent (CLI usage).

    Absence and emptiness are different statements. ``None`` means the caller
    said nothing about the layout, so the stored one is used. ``[]`` means the
    caller states there is no layout, and that is honoured: every exporter this
    handler reaches accepts an empty list and writes a header-only file, so the
    empty payload is exported rather than silently swapped for whatever state
    happens to hold. ``mappings_source`` in the manifest names the branch taken.
    """
    started_at = datetime.now(timezone.utc)

    p = ExportExcelParams(**params)
    resolved = _validate_output_path(
        p.filepath, allowed_extensions=_ALLOWED_EXCEL_EXTENSIONS
    )

    if p.mappings is not None:
        mappings = _pydantic_to_plate_mappings(p.mappings)
        with _core._state_lock:
            state_dedup = _core._state.dedup_info
            results_for_export = list(_core._state.results)
        rev_groups = _resolve_rev_groups(
            p.dedup_info, state_dedup, results_for_export, context="export_excel",
        )
    else:
        with _core._state_lock:
            if not _core._state.results:
                raise ValueError("No design available")
            mappings = _core._state.plate_mappings
            results_for_export = list(_core._state.results)
            rev_groups = _resolve_rev_groups(
                None, _core._state.dedup_info, results_for_export,
                context="export_excel",
            )

    # Derive overlap_mode from the first result (all results in a run share the same mode).
    run_overlap_mode = results_for_export[0].overlap_mode if results_for_export else "partial"

    rescued_info = (
        [item.model_dump(exclude_none=True) for item in p.rescued_info]
        if p.rescued_info
        else None
    )

    export_plate_excel(
        mappings, resolved,
        rev_groups=rev_groups,
        results=results_for_export,
        overlap_mode=run_overlap_mode,
        rescued_info=rescued_info,
    )
    if p.project_id or p.report_data or p.benchmark_raw:
        wb = openpyxl.load_workbook(resolved)
        if p.report_data and isinstance(p.report_data, dict):
            _write_report_sheet(wb, p.report_data)
        if p.benchmark_raw and isinstance(p.benchmark_raw, dict):
            _write_benchmark_raw_sheet(wb, p.benchmark_raw)
        if "__kuma_meta__" in wb.sheetnames:
            del wb["__kuma_meta__"]
        meta = wb.create_sheet("__kuma_meta__")
        meta.sheet_state = "hidden"
        meta.append(["project_id", p.project_id or ""])
        meta.append(["kuma_version", p.kuma_version or KUMA_VERSION])
        meta.append(["kuro_module_version", KURO_MODULE_VERSION])
        meta.append(["exported_at", datetime.now(timezone.utc).isoformat()])
        meta.append(["overlap_mode", run_overlap_mode])
        wb.save(resolved)

    finished_at = datetime.now(timezone.utc)

    # Sanitise params for manifest: exclude large non-serialisable items.
    manifest_params = {
        k: v for k, v in params.items()
        if k not in ("mappings", "rescued_info", "report_data", "benchmark_raw")
    }
    # The exported rows come out of _core._state.results on both branches above;
    # only the well layout can arrive as a payload, which is recorded separately
    # rather than folded into results_source.
    manifest_inputs, manifest_extra = _design_provenance_for_manifest("state")
    manifest_extra["mappings_source"] = (
        "payload" if p.mappings is not None else "state"
    )
    manifest = build_run_manifest(
        method="export_excel",
        inputs=manifest_inputs,
        params=manifest_params,
        started_at=started_at,
        finished_at=finished_at,
        extra=manifest_extra,
    )
    mpath = _manifest_path_for(resolved)
    write_run_manifest(mpath, manifest)
    cpath = write_output_checksum(resolved)

    result = FileExportResultModel(filepath=str(resolved)).to_rpc_dict()
    result["manifest_path"] = str(mpath)
    result["checksum_path"] = str(cpath)
    return result


def _order_payload_to_results(items):
    """Build the minimal result shape needed by order CSV exporters."""
    return [
        SimpleNamespace(
            mutation=SimpleNamespace(raw=item.mutation),
            forward_seq=item.forward_seq,
            reverse_seq=item.reverse_seq,
        )
        for item in items
    ]


def handle_export_order(params: dict) -> dict:
    """Export primer order CSV for IDT or Twist."""
    started_at = datetime.now(timezone.utc)

    p = ExportOrderParams(**params)
    resolved = _validate_output_path(p.filepath, allowed_extensions=_ALLOWED_CSV_EXTENSIONS)

    if p.results is not None:
        results = _order_payload_to_results(p.results)
    else:
        with _core._state_lock:
            if not _core._state.results:
                raise ValueError("No design available. Run design_sdm_primers first.")
            results = list(_core._state.results)

    encoding = "utf-8-sig" if p.bom else "utf-8"
    if p.format == "idt":
        export_idt_csv(results, resolved, encoding=encoding)  # pyright: ignore[reportArgumentType]
    else:
        export_twist_csv(results, resolved, encoding=encoding)  # pyright: ignore[reportArgumentType]

    finished_at = datetime.now(timezone.utc)

    manifest_params = {"filepath": params.get("filepath"), "format": p.format}
    manifest_inputs, manifest_extra = _design_provenance_for_manifest(
        "payload" if p.results is not None else "state"
    )
    manifest = build_run_manifest(
        method="export_order",
        inputs=manifest_inputs,
        params=manifest_params,
        started_at=started_at,
        finished_at=finished_at,
        extra=manifest_extra,
    )
    mpath = _manifest_path_for(resolved)
    write_run_manifest(mpath, manifest)
    cpath = write_output_checksum(resolved)

    result = ExportOrderResultModel(
        filepath=str(resolved),
        format=p.format,
        primer_count=len(results) * 2,
    ).to_rpc_dict()
    result["manifest_path"] = str(mpath)
    result["checksum_path"] = str(cpath)
    return result


def handle_export_mapping(params: dict) -> dict:
    """Export liquid handler mapping file (Echo 525 or JANUS, CSV or XLSX)."""
    started_at = datetime.now(timezone.utc)

    p = ExportMappingParams(**params)
    resolved = _validate_output_path(p.filepath, allowed_extensions=_ALLOWED_MAPPING_EXTENSIONS)

    mapping_range = (
        (p.mapping_range.row_start, p.mapping_range.row_end)
        if p.mapping_range is not None
        else None
    )

    if p.mappings is not None:
        mappings = _pydantic_to_plate_mappings(p.mappings)
        fwd_mappings = [m for m in mappings if m.primer_type == "forward"]
        rev_mappings = [m for m in mappings if m.primer_type == "reverse"]
        with _core._state_lock:
            state_dedup = _core._state.dedup_info
            state_results = list(_core._state.results)
        rev_groups = _resolve_rev_groups(
            p.dedup_info, state_dedup, state_results, context="export_mapping",
        )
    else:
        with _core._state_lock:
            if not _core._state.results:
                raise ValueError("No design available. Run design_sdm_primers first.")
            results = _core._state.results
            rev_groups = _resolve_rev_groups(
                None, _core._state.dedup_info, results, context="export_mapping",
            )

        fwd_mappings, rev_mappings = generate_plate_map(
            results,
            deduplicate_rev=True,
            mapping_range=mapping_range,
        )

    use_xlsx = resolved.suffix.lower() == ".xlsx"

    encoding = "utf-8-sig" if p.bom else "utf-8"
    if p.format == "echo":
        vol = int(_resolve_mapping_transfer_volume(p.format, p.transfer_vol))
        if use_xlsx:
            # Same placement parameters as the csv branch: one export_mapping
            # call must not answer differently for .xlsx than for .csv.
            export_echo_mapping_xlsx(fwd_mappings, rev_mappings, resolved,
                                     transfer_vol=vol, rev_groups=rev_groups,
                                     mapping_range=mapping_range,
                                     quadrant=p.quadrant,
                                     used_quadrants=list(p.used_quadrants or []))
        else:
            export_echo_mapping_csv(fwd_mappings, rev_mappings, resolved,
                                    transfer_vol=vol, rev_groups=rev_groups,
                                    encoding=encoding,
                                    mapping_range=mapping_range,
                                    quadrant=p.quadrant,
                                    used_quadrants=list(p.used_quadrants or []))
    else:
        vol = _resolve_mapping_transfer_volume(p.format, p.transfer_vol)
        if use_xlsx:
            export_janus_mapping_xlsx(fwd_mappings, rev_mappings, resolved,
                                      transfer_vol=vol, rev_groups=rev_groups)
        else:
            export_janus_mapping_csv(fwd_mappings, rev_mappings, resolved,
                                     transfer_vol=vol, rev_groups=rev_groups,
                                     encoding=encoding)

    finished_at = datetime.now(timezone.utc)

    primer_count = len(fwd_mappings) + len(rev_mappings)

    manifest_params = {
        "filepath": params.get("filepath"),
        "format": p.format,
        "transfer_vol": p.transfer_vol,
    }
    # Unlike export_excel and export_all, the rows this handler writes really do
    # come from the payload when one is supplied: the state branch derives them
    # with generate_plate_map, the payload branch does not consult
    # _core._state.results at all. So results_source tracks p.mappings here and
    # the comment in handle_export_all warns only about copying this reading
    # into handlers whose rows always come from state.
    manifest_inputs, manifest_extra = _design_provenance_for_manifest(
        "payload" if p.mappings is not None else "state"
    )
    manifest = build_run_manifest(
        method="export_mapping",
        inputs=manifest_inputs,
        params=manifest_params,
        started_at=started_at,
        finished_at=finished_at,
        extra=manifest_extra,
    )
    mpath = _manifest_path_for(resolved)
    write_run_manifest(mpath, manifest)
    cpath = write_output_checksum(resolved)

    result = ExportMappingResultModel(
        filepath=str(resolved),
        format=p.format,
        primer_count=primer_count,
    ).to_rpc_dict()
    result["manifest_path"] = str(mpath)
    result["checksum_path"] = str(cpath)
    return result


def handle_save_workspace(params: dict) -> dict:
    """Save workspace JSON to file."""
    p = SaveWorkspaceParams(**params)
    if not p.filepath or p.data is None:
        raise ValueError("filepath and data are required")
    resolved = _validate_output_path(p.filepath, allowed_extensions={".json"})
    with open(resolved, "w", encoding="utf-8") as f:
        json.dump(p.data, f, ensure_ascii=False, indent=2)
    return FileExportResultModel(filepath=str(resolved)).to_rpc_dict()


def handle_save_json(params: dict) -> dict:
    """Save generic JSON payload to file."""
    p = SaveJsonParams(**params)
    if not p.filepath or p.data is None:
        raise ValueError("filepath and data are required")
    resolved = _validate_output_path(p.filepath, allowed_extensions={".json"})
    with open(resolved, "w", encoding="utf-8") as f:
        json.dump(p.data, f, ensure_ascii=False, indent=2)
    return FileExportResultModel(filepath=str(resolved)).to_rpc_dict()


def handle_load_workspace(params: dict) -> dict:
    """Load workspace JSON from file."""
    p = LoadWorkspaceParams(**params)
    resolved = _validate_filepath(p.filepath, allowed_extensions={".json"})
    if not resolved.exists():
        raise FileNotFoundError(f"File not found: {p.filepath}")
    file_size = resolved.stat().st_size
    if file_size > 50 * 1024 * 1024:
        raise ValueError(f"Workspace file too large: {file_size / 1024 / 1024:.1f} MB (max 50 MB)")
    with open(resolved, encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError("Workspace file must contain a JSON object")
    result_list = None
    if isinstance(data.get("results"), list):
        result_list = data["results"]
    elif isinstance(data.get("results"), dict) and isinstance(data["results"].get("designResults"), list):
        result_list = data["results"]["designResults"]

    if result_list is not None and len(result_list) > 10_000:
        raise ValueError(f"Workspace contains {len(result_list)} results, exceeding 10,000 limit")

    validated = validate_workspace_data(data)
    return validated.to_rpc_dict(exclude_unset=True, round_trip=True)


def handle_export_benchmark_csv(params: dict) -> dict:
    """Export benchmark result table to CSV."""
    p = ExportBenchmarkCsvParams(**params)
    if not p.results:
        raise ValueError("Benchmark results are required")

    resolved = _validate_output_path(p.filepath, allowed_extensions=_ALLOWED_CSV_EXTENSIONS)
    fieldnames = [
        "strategy",
        "n_selected",
        "hit_rate",
        "mean_fitness",
        "unique_positions",
        "position_coverage",
        "domain_coverage",
        "structural_spread",
        "hits",
        "threshold",
        "n_trials",
    ]
    encoding = "utf-8-sig" if p.bom else "utf-8"
    with open(resolved, "w", encoding=encoding, newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for strategy, metrics in p.results.items():
            writer.writerow({"strategy": strategy, **metrics})
    return FileExportResultModel(filepath=str(resolved)).to_rpc_dict()


# ---------------------------------------------------------------------------
# Export All + Macrogen (spec 2026-05-13)
# ---------------------------------------------------------------------------

from datetime import datetime as _dt

from kuma_core.kuro.plate_mapper import export_macrogen_xls
from sidecar_kuro.models import ExportAllParams, ExportMacrogenParams


def _split_fwd_rev(mappings: list[PlateMapping]) -> tuple[list[PlateMapping], list[PlateMapping]]:
    fwd = [m for m in mappings if m.primer_type == "forward"]
    rev = [m for m in mappings if m.primer_type == "reverse"]
    return fwd, rev


def handle_export_macrogen(params: dict) -> dict:
    """Export forward/reverse plate primers to Macrogen Plate Oligo .xls.

    Writes a sibling ``.run.json`` the way the three single-file export handlers
    do. This handler also produces exactly one file at a path the operator
    picked, so the sibling convention fits; the batch pipeline cannot use it
    because ``handle_export_all`` already names one ``{prefix}_run.json`` for
    the whole folder.

    The returned dict is deliberately left at ``{ok, path}``. Not because
    anything rejects a third key (``src/types/validators.ts`` only checks that
    ``ok`` and ``path`` are present), but because no caller would read it:
    ``handleExportMacrogen`` types the result as those two keys and this change
    stops at the Python layer, so a ``manifest_path`` nobody declared would just
    be dead weight.

    Known gap, and it needs a TypeScript-only follow-up: ``src/lib/reRun.ts``
    treats any ``.run.json`` as a manifest, and ``export_macrogen`` is in
    neither its runnable nor its export-only set, so dropping this file on the
    app reports "unsupported method" instead of "export not runnable". Adding
    the method to ``EXPORT_ONLY_METHODS`` fixes the wording. Naming a different
    method here to dodge it would be a lie in the one file meant to be trusted.
    """
    started_at = datetime.now(timezone.utc)

    p = ExportMacrogenParams(**params)
    with _core._state_lock:
        mappings = list(_core._state.plate_mappings)

    fwd, rev = _split_fwd_rev(mappings)

    if fwd and not p.fwd_plate_name:
        raise ValueError("fwd_plate_name is required when forward primers exist")
    if rev and not p.rev_plate_name:
        raise ValueError("rev_plate_name is required when reverse primers exist")

    resolved = _validate_output_path(p.output_path, allowed_extensions={".xls"})
    export_macrogen_xls(
        fwd_primers=fwd,
        rev_primers=rev,
        fwd_plate_name=p.fwd_plate_name,
        rev_plate_name=p.rev_plate_name,
        amount=p.amount,
        purification=p.purification,
        output_path=str(resolved),
    )

    finished_at = datetime.now(timezone.utc)

    # The plate primers came out of state above and there is no payload branch
    # to take them from, so "state" is the only honest value here.
    manifest_inputs, manifest_extra = _design_provenance_for_manifest("state")
    write_run_manifest(
        _manifest_path_for(resolved),
        build_run_manifest(
            method="export_macrogen",
            inputs=manifest_inputs,
            params={
                "output_path": params.get("output_path"),
                "fwd_plate_name": p.fwd_plate_name,
                "rev_plate_name": p.rev_plate_name,
                "amount": p.amount,
                "purification": p.purification,
            },
            started_at=started_at,
            finished_at=finished_at,
            extra=manifest_extra,
        ),
    )

    return {"ok": True, "path": str(resolved)}


def _export_primers_fasta(mappings: list[PlateMapping], output_path: Path) -> None:
    lines = []
    for m in mappings:
        lines.append(f">{m.primer_name}\n{m.sequence}\n")
    output_path.write_text("".join(lines), encoding="utf-8")


def _export_echo_for_all(
    fwd: list[PlateMapping],
    rev: list[PlateMapping],
    output_path: Path,
    transfer_vol: int,
    rev_groups: dict,
    bom: bool,
    quadrant: str | None = None,
    used_quadrants: list[str] | None = None,
) -> None:
    export_echo_mapping_csv(
        fwd, rev, output_path,
        transfer_vol=transfer_vol,
        rev_groups=rev_groups,
        encoding="utf-8-sig" if bom else "utf-8",
        quadrant=quadrant,
        used_quadrants=used_quadrants,
    )


def _export_janus_for_all(
    fwd: list[PlateMapping],
    rev: list[PlateMapping],
    output_path: Path,
    transfer_vol: float,
    rev_groups: dict,
    bom: bool,
) -> None:
    export_janus_mapping_csv(
        fwd, rev, output_path,
        transfer_vol=transfer_vol,
        rev_groups=rev_groups,
        encoding="utf-8-sig" if bom else "utf-8",
    )


def _export_platemap_for_all(
    mappings: list[PlateMapping],
    results,
    rev_groups: dict,
    output_path: Path,
) -> None:
    export_plate_excel(
        mappings, output_path,
        rev_groups=rev_groups,
        results=results,
    )


def _export_run_json(
    mappings: list[PlateMapping],
    results,
    rev_groups: dict,
    output_path: Path,
    *,
    manifest: dict,
) -> None:
    """Write the batch ``{prefix}_run.json``: a run manifest plus its plate map.

    *manifest* comes from ``build_run_manifest`` and its keys go in first; the
    plate dump earlier versions wrote follows. One merged file rather than a
    second one next to it, because ``src/components/layout/export-handlers.ts``
    maps the ``_run.json`` suffix to the ``kuro_run_json`` artifact type, so a
    rename or an extra sibling would turn this into a cross-layer change. The
    ``mappings``, ``dedup_info`` and ``result_count`` keys keep their names and
    shapes so anything already reading this file keeps working.

    *manifest* is keyword-only and has no default on purpose: the bug being
    fixed here was that this file carried no manifest at all, and a default
    would let a future caller reintroduce it silently.
    """
    payload = {
        **manifest,
        "exported_at": _dt.now().isoformat(),
        "mappings": [
            {
                "well": m.well,
                "primer_name": m.primer_name,
                "sequence": m.sequence,
                "primer_type": m.primer_type,
                "mutation": m.mutation,
            }
            for m in mappings
        ],
        "dedup_info": rev_groups,
        "result_count": len(results) if results else 0,
    }
    # write_run_manifest, not write_text: the plate map is now sitting in a
    # manifest, and a half-written manifest is worse than none.
    write_run_manifest(output_path, payload)


def _build_echo_preview_rows(
    fwd: list[PlateMapping],
    rev: list[PlateMapping],
    transfer_vol: int,
    rev_groups: dict,
    mapping_range: tuple[str, str] | None = None,
    quadrant: str | None = None,
    used_quadrants: list[str] | None = None,
) -> list[dict]:
    """Build Echo mapping preview rows in-memory (no file I/O).

    Delegates to ``build_echo_rows``, the same builder the CSV and XLSX exports
    use, so the preview cannot describe a different plate than the file the
    operator ends up loading onto the instrument. It used to re-derive the rows
    and honour ``mapping_range`` alone, which put the preview above the quadrant
    selector in ExportStepView showing wells the exported csv would not use.
    """
    from kuma_core.kuro.plate_mapper import build_echo_rows

    return build_echo_rows(
        fwd, rev, rev_groups, transfer_vol,
        mapping_range=mapping_range,
        quadrant=quadrant,
        used_quadrants=used_quadrants,
    )


def _build_janus_preview_rows(
    fwd: list[PlateMapping],
    rev: list[PlateMapping],
    transfer_vol: float,
    rev_groups: dict,
    mapping_range: tuple[str, str] | None = None,
) -> list[dict]:
    """Build JANUS mapping preview rows in-memory (no file I/O).

    Delegates to ``build_janus_rows``, the same builder the CSV and XLSX
    exports use, so the preview cannot describe a different run than the file
    the operator ends up loading onto the instrument. The plate names written
    into ``Asp. Rack`` and ``Dsp. Rack`` come from the deck policy, not from
    literals here.

    ``mapping_range`` is accepted for parity with the Echo dry-run preview but
    JANUS layout uses 96-well source/dest so the range only flows through
    if/when the plate_mapper Janus paths consume it (currently unused, kept
    future-proof).
    """
    from kuma_core.kuro.plate_mapper import build_janus_rows

    return build_janus_rows(fwd, rev, rev_groups, transfer_vol)


def handle_export_echo_mapping_dry_run(params: dict) -> dict:
    """Return Echo 525 mapping rows for preview without writing a file.

    Params:
        transfer_vol: optional override (nL, integer). Default 100.
        mappings: optional pre-reordered plate layout (matches frontend sort).
        dedup_info: companion to ``mappings``.

    Returns:
        ``{"rows": [...], "total": N, "transfer_vol": int}``.
    """
    p = ExportMappingDryRunParams(**params)
    vol = _resolve_mapping_transfer_volume("echo", p.transfer_vol)
    mapping_range = (
        (p.mapping_range.row_start, p.mapping_range.row_end)
        if p.mapping_range is not None
        else None
    )
    if p.mappings is not None:
        mappings = _pydantic_to_plate_mappings(p.mappings)
        if not mappings:
            return {"rows": [], "total": 0, "transfer_vol": int(vol)}
        with _core._state_lock:
            state_dedup = _core._state.dedup_info
            state_results = list(_core._state.results)
        rev_groups = _resolve_rev_groups(
            p.dedup_info, state_dedup, state_results,
            context="export_echo_mapping_dry_run",
        )
    else:
        with _core._state_lock:
            if not _core._state.results:
                return {"rows": [], "total": 0, "transfer_vol": int(vol)}
            mappings = list(_core._state.plate_mappings)
            rev_groups = _resolve_rev_groups(
                None, _core._state.dedup_info, _core._state.results,
                context="export_echo_mapping_dry_run",
            )
    fwd, rev = _split_fwd_rev(mappings)
    rows = _build_echo_preview_rows(
        fwd, rev, int(vol), rev_groups,
        mapping_range=mapping_range,
        quadrant=p.quadrant,
        used_quadrants=list(p.used_quadrants or []),
    )
    return {"rows": rows, "total": len(rows), "transfer_vol": int(vol)}


def handle_export_janus_mapping_dry_run(params: dict) -> dict:
    """Return JANUS mapping rows for preview without writing a file.

    Params:
        transfer_vol: optional override (µL, float). Default 2.0.
        mappings: optional pre-reordered plate layout (matches frontend sort).
        dedup_info: companion to ``mappings``.

    Returns:
        ``{"rows": [...], "total": N, "transfer_vol": float}``.
    """
    p = ExportMappingDryRunParams(**params)
    vol = _resolve_mapping_transfer_volume("janus", p.transfer_vol)
    mapping_range = (
        (p.mapping_range.row_start, p.mapping_range.row_end)
        if p.mapping_range is not None
        else None
    )
    if p.mappings is not None:
        mappings = _pydantic_to_plate_mappings(p.mappings)
        if not mappings:
            return {"rows": [], "total": 0, "transfer_vol": float(vol)}
        with _core._state_lock:
            state_dedup = _core._state.dedup_info
            state_results = list(_core._state.results)
        rev_groups = _resolve_rev_groups(
            p.dedup_info, state_dedup, state_results,
            context="export_janus_mapping_dry_run",
        )
    else:
        with _core._state_lock:
            if not _core._state.results:
                return {"rows": [], "total": 0, "transfer_vol": float(vol)}
            mappings = list(_core._state.plate_mappings)
            rev_groups = _resolve_rev_groups(
                None, _core._state.dedup_info, _core._state.results,
                context="export_janus_mapping_dry_run",
            )
    fwd, rev = _split_fwd_rev(mappings)
    rows = _build_janus_preview_rows(
        fwd, rev, float(vol), rev_groups, mapping_range=mapping_range,
    )
    return {"rows": rows, "total": len(rows), "transfer_vol": float(vol)}


def handle_export_all(params: dict) -> dict:
    """Run the 6-file batch export pipeline.

    Returns ``{"success": [filename, ...], "failed": [{path, reason}, ...], "output_dir": str}``.
    Individual exporter failures are recorded but do not raise.
    """
    started_at = datetime.now(timezone.utc)

    p = ExportAllParams(**params)
    out_dir = Path(p.output_dir).expanduser().resolve()
    if not out_dir.is_absolute():
        raise ValueError(f"output_dir must be absolute: {p.output_dir}")
    if out_dir.exists() and not out_dir.is_dir():
        raise ValueError(f"output_dir exists but is not a directory: {out_dir}")
    out_dir.mkdir(parents=True, exist_ok=True)

    if p.mappings is not None:
        mappings = _pydantic_to_plate_mappings(p.mappings)
        # Filter backend results to only those present in frontend mappings
        # so capped designs (e.g. maxPrimers=95) export the capped set.
        mut_keys = {m.mutation for m in mappings}
        with _core._state_lock:
            state_dedup = _core._state.dedup_info
            results = [r for r in _core._state.results if r.mutation.raw in mut_keys]
        rev_groups = _resolve_rev_groups(
            p.dedup_info, state_dedup, results, context="export_all",
        )
    else:
        with _core._state_lock:
            mappings = list(_core._state.plate_mappings)
            results = list(_core._state.results)
            rev_groups = _resolve_rev_groups(
                None, _core._state.dedup_info, results, context="export_all",
            )

    fwd, rev = _split_fwd_rev(mappings)

    now = _dt.now()
    if p.project_name:
        base_folder_name = f"{p.project_name}_{now.strftime('%Y%m%d')}"
    else:
        base_folder_name = f"kuro_{now.strftime('%y%m%d_%H%M')}"
    target_dir = out_dir / base_folder_name
    suffix = 1
    while target_dir.exists():
        suffix += 1
        target_dir = out_dir / f"{base_folder_name}_{suffix}"
    target_dir.mkdir(parents=True)

    file_prefix = target_dir.name
    ECHO_CSV = f"{file_prefix}_echo.csv"
    ECHO_XLSX = f"{file_prefix}_echo.xlsx"
    JANUS_CSV = f"{file_prefix}_janus.csv"
    JANUS_XLSX = f"{file_prefix}_janus.xlsx"
    MACROGEN = f"{file_prefix}_macrogen.xls"
    PRIMERS_FASTA = f"{file_prefix}_primers.fasta"
    PLATEMAP_XLSX = f"{file_prefix}_platemap.xlsx"
    RUN_JSON = f"{file_prefix}_run.json"

    success: list[str] = []
    failed: list[dict] = []

    def _try(name: str, fn) -> None:
        try:
            fn()
            success.append(name)
        except Exception as exc:  # noqa: BLE001 -- intentionally aggregating per-file
            failed.append({"path": name, "reason": str(exc)})

    _try(MACROGEN, lambda: export_macrogen_xls(
        fwd_primers=fwd,
        rev_primers=rev,
        fwd_plate_name=p.fwd_plate_name,
        rev_plate_name=p.rev_plate_name,
        amount=p.amount,
        purification=p.purification,
        output_path=str(target_dir / MACROGEN),
    ))

    _try(PRIMERS_FASTA, lambda: _export_primers_fasta(mappings, target_dir / PRIMERS_FASTA))

    _try(ECHO_CSV, lambda: _export_echo_for_all(
        fwd, rev, target_dir / ECHO_CSV,
        transfer_vol=int(p.echo_transfer_vol),
        rev_groups=rev_groups,
        bom=p.bom,
        quadrant=p.quadrant,
        used_quadrants=list(p.used_quadrants or []),
    ))

    # The same quadrant the csv above is written with. One export_all used to
    # leave a csv and an xlsx naming different source wells for the same primer,
    # and nothing in either file says which one the plate was stamped from.
    _try(ECHO_XLSX, lambda: export_echo_mapping_xlsx(
        fwd, rev, target_dir / ECHO_XLSX,
        transfer_vol=int(p.echo_transfer_vol),
        rev_groups=rev_groups,
        quadrant=p.quadrant,
        used_quadrants=list(p.used_quadrants or []),
    ))

    _try(JANUS_CSV, lambda: _export_janus_for_all(
        fwd, rev, target_dir / JANUS_CSV,
        transfer_vol=float(p.janus_transfer_vol),
        rev_groups=rev_groups,
        bom=p.bom,
    ))

    _try(JANUS_XLSX, lambda: export_janus_mapping_xlsx(
        fwd, rev, target_dir / JANUS_XLSX,
        transfer_vol=float(p.janus_transfer_vol),
        rev_groups=rev_groups,
    ))

    _try(PLATEMAP_XLSX, lambda: _export_platemap_for_all(
        mappings, results, rev_groups, target_dir / PLATEMAP_XLSX,
    ))

    finished_at = datetime.now(timezone.utc)

    # Same reading as handle_export_excel, and for the same reason: the design
    # results come out of _core._state.results on both branches above (the
    # payload branch only filters them by the mutations the caller kept), so the
    # session provenance describes this export either way, and where the well
    # layout came from is recorded separately instead of being folded into
    # results_source. Keying results_source off p.mappings the way
    # handle_export_mapping does would drop the provenance on the only path
    # operators actually use: src/components/layout/export-handlers.ts always
    # sends mappings for export_all, so every real batch export would come out
    # stamped provenance_omitted.
    manifest_inputs, manifest_extra = _design_provenance_for_manifest("state")
    manifest_extra["mappings_source"] = (
        "payload" if p.mappings is not None else "state"
    )
    manifest = build_run_manifest(
        method="export_all",
        inputs=manifest_inputs,
        # mappings and dedup_info are the plate map, which the same file already
        # carries in full below. Copying them into params would double the file.
        params={
            k: v for k, v in params.items()
            if k not in ("mappings", "dedup_info")
        },
        started_at=started_at,
        finished_at=finished_at,
        extra=manifest_extra,
    )

    _try(RUN_JSON, lambda: _export_run_json(
        mappings, results, rev_groups, target_dir / RUN_JSON, manifest=manifest,
    ))

    return {
        "success": success,
        "failed": failed,
        "output_dir": str(target_dir),
    }
