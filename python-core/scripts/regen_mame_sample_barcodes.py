#!/usr/bin/env python3
"""Regenerate the shipped MAME custom-barcode workbooks from their own inputs.

Why this script exists
----------------------
``src-tauri/samples/mame/04_mame_custom_barcodes.xlsx`` and its ``templates/``
twin used to hold 24 bp sequences with a 0 bp common suffix on both axes: no
shared annealing region at all. They loaded only because the demux fell back to
a fixed 11 bp / 10 bp seed cut. That fallback is gone (a seed length nothing in
the workbook states is a guess, and a wrong guess names the wrong plate row), so
those files would now be refused by the very reader the sample is meant to
demonstrate.

The fix is to give the samples the shape a real package has, ``[seed][shared
flank]``, and the only honest way to do that is to run kuma's own generator over
kuma's own sample inputs rather than to hand-author a binary nobody can
reproduce. Rerun this after changing either input:

    python python-core/scripts/regen_mame_sample_barcodes.py

Inputs (both already shipped as samples)
----------------------------------------
* ``src-tauri/samples/mame/egfp_with_flanks.fa`` - 1620 bp, EGFP CDS at 450..1170,
  with synthetic flanks the flanking primers are designed in.
* ``src-tauri/samples/mame/02_mame_barcode_seeds.xlsx`` - the 12 fwd + 8 rev
  11 bp seeds, which is the file an operator ordered primers from and therefore
  the ground truth the regenerated workbook has to hand back.

Note on ``src-tauri/samples/mame/reference.fasta``: it is the bare 720 bp EGFP
CDS and is deliberately NOT touched. The flanking primers designed here bind
outside the CDS, so ``resolve_amplicon_reference`` reports NOT_FOUND against it.
That is the documented normal path for a bare-CDS reference (see the
``_SpanReason`` comments in ``kuma_core/mame/ingest/amplicon_reference.py``), not
a defect introduced by this script.

It is, however, the wrong file to hand the demo as its analyze reference, which
is why ``loadSampleData`` (``src/store/mame/slices/analysisSlice.ts``) points
that at ``egfp_with_flanks.fa`` instead: the workbook this script writes and the
reference the demo analyzes have to be the same construct, or the demo can only
ever show the NOT_FOUND path. ``reference.fasta`` remains bundled as the
bare-CDS sample and as that prefill's fallback.
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

from kuma_core.mame.ingest.barcode_package import (  # noqa: E402
    generate_mame_package,
    parse_barcode_seeds,
)
from kuma_core.mame.ingest.combinatorial_demux import (  # noqa: E402
    load_barcode_prefixes_with_provenance,
)

_FASTA = _REPO_ROOT / "src-tauri" / "samples" / "mame" / "egfp_with_flanks.fa"
_SEEDS = _REPO_ROOT / "src-tauri" / "samples" / "mame" / "02_mame_barcode_seeds.xlsx"
_GENE_START = 450
_GENE_END = 1170

#: Destination, and the row-name prefix that destination already uses. The two
#: differ (the sample says ``egfp``, the template says ``isps``) and both are
#: kept as they are: ``templates/README.md`` documents the template prefix, and
#: the demux is gene-agnostic, so the prefix carries no behaviour either way.
_DESTINATIONS = (
    (
        _REPO_ROOT / "src-tauri" / "samples" / "mame" / "04_mame_custom_barcodes.xlsx",
        "egfp",
    ),
    (_REPO_ROOT / "templates" / "04_mame_custom_barcodes.xlsx", "isps"),
)

#: Sheet name and header row the shipped files already have; the reader ignores
#: both (it matches on the row-name pattern), but a sample is also something a
#: user opens, and the columns should still say what they are.
_SHEET_TITLE = "barcodes"
_HEADER = ("barcode_name", "sequence")


def _design_rows() -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    """Run the real generator once and read its 12 fwd + 8 rev rows back.

    Returns ``(forward, reverse)`` as ``(index_suffix, sequence)`` pairs, prefix
    stripped, so each destination can re-attach its own.
    """
    import openpyxl

    with tempfile.TemporaryDirectory(prefix="mame-sample-barcodes-") as tmp:
        tmp_path = Path(tmp)
        result = generate_mame_package(
            fasta_path=_FASTA,
            gene_start=_GENE_START,
            gene_end=_GENE_END,
            barcode_seeds_path=_SEEDS,
            output_dir=tmp_path / "out",
            project_root=tmp_path,
            gene_name="egfp",
        )
        for warning in result.warnings:
            print(f"  generator warning: {warning}")
        workbook = openpyxl.load_workbook(result.barcodes_xlsx)
        sheet = workbook.active
        if sheet is None:
            workbook.close()
            raise RuntimeError(
                f"generated barcode workbook has no sheet: {result.barcodes_xlsx}"
            )
        forward: list[tuple[str, str]] = []
        reverse: list[tuple[str, str]] = []
        for name, sequence in sheet.iter_rows(min_row=2, values_only=True):
            text = str(name)
            if "_f_" in text:
                forward.append((text.split("_f_")[1], str(sequence).upper()))
            elif "_r_" in text:
                reverse.append((text.split("_r_")[1], str(sequence).upper()))
        workbook.close()
    return forward, reverse


def _write(
    dest: Path,
    prefix: str,
    forward: list[tuple[str, str]],
    reverse: list[tuple[str, str]],
) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # A freshly created Workbook always carries one active worksheet;
    # None here would mean openpyxl handed back an empty book.
    assert sheet is not None
    sheet.title = _SHEET_TITLE
    sheet.append(list(_HEADER))
    for index, sequence in forward:
        sheet.append([f"{prefix}_f_{index}", sequence])
    for index, sequence in reverse:
        sheet.append([f"{prefix}_r_{index}", sequence])
    workbook.save(dest)


def _verify(dest: Path) -> None:
    """Read the file back the way MAME reads it, and check it against the seeds.

    Two claims, both checked rather than assumed: the reader derives a tail on
    both axes (it raises otherwise, so reaching the assertions is already most
    of it), and every seed it hands back is the seed the operator ordered.
    """
    expected = parse_barcode_seeds(_SEEDS)
    resolution = load_barcode_prefixes_with_provenance(dest)

    matched = 0
    for axis, key_prefix in ((resolution.forward, "fwd"), (resolution.reverse, "rev")):
        for position, (name, seed) in enumerate(axis.barcodes, start=1):
            want = expected[f"{key_prefix}_{position}"].upper()
            if seed != want:
                raise SystemExit(
                    f"{dest.name}: {name} came back as {seed!r}, "
                    f"but {_SEEDS.name} states {want!r}"
                )
            matched += 1

    print(f"  {dest.relative_to(_REPO_ROOT)}")
    print(
        f"    forward tail {resolution.forward.tail} "
        f"({len(resolution.forward.tail)} bp), "
        f"reverse tail {resolution.reverse.tail} "
        f"({len(resolution.reverse.tail)} bp)"
    )
    print(f"    seeds matching {_SEEDS.name}: {matched}/{len(expected)}")


def main() -> int:
    forward, reverse = _design_rows()
    print(f"designed {len(forward)} forward and {len(reverse)} reverse rows")
    for dest, prefix in _DESTINATIONS:
        _write(dest, prefix, forward, reverse)
    for dest, _ in _DESTINATIONS:
        _verify(dest)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
