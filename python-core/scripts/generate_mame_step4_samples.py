#!/usr/bin/env python3
"""Regenerate the bundled MAME sample set as one coherent campaign.

Every file this writes describes the same plate: the EGFP reference in
``src-tauri/samples/mame/reference.fasta``, ten designed variants, and the
placement kuma computes for them. Nothing here is hand-authored, so the sample
set cannot drift into the state it shipped in, where the expected-mutation
positions were numbered against a different protein than the reference and no
combination of step 4 inputs could finish a build.

What it writes, all under ``src-tauri/samples/mame/``:

* ``03_mame_expected_mutations.xlsx`` - the designed variant list. Positions are
  numbered against the shipped reference and every ``wt_aa`` is read off it, so
  the list cannot claim a residue the sequence does not carry.
* ``sample_analysis_result.json`` - the Analyze screen fixture, produced by a
  real in-process pipeline run over synthetic consensus FASTA. No FASTQ and no
  MinKNOW run folder is bundled: a raw nanopore run is large and the app only
  needs the result to show one.
* ``13_mame_verdict.xlsx`` - the Analyze workbook from that same run. Step 4
  requires a verdict workbook and had none, which is why the sample data could
  not reach a finished build.
* ``06_mame_plate_layout.xlsx`` - the placement, one well per variant.
* ``07_mame_activity_long.csv`` / ``.xlsx`` - well-labeled long format already
  relative to wild-type.
* ``14_mame_activity_long_raw.csv`` - the same measurements unnormalised, with
  the ``WT_1``..``WT_3`` rows that ``activity_scale="raw"`` divides by.
* ``15_mame_activity_variant.csv`` - variant-labeled long format, the branch
  that needs no plate layout at all.
* ``10_mame_gc_prenormalised.xlsx`` - GC sheet already relative to wild-type.
* ``11_mame_gc_fid_round1_raw.xlsx`` - raw Agilent FID report.
* ``09_mame_agilent_rep_batch.xlsx`` - variant-labeled confirmation report.
* ``12_mame_agilent_numeric_index.xlsx`` - numeric-ID confirmation report, whose
  identifier count now matches the plate order it is decoded against.

Usage::

    PYTHONPATH=.:python-core python python-core/scripts/generate_mame_step4_samples.py

The script is re-runnable and overwrites each file. It verifies what it wrote by
running every step 4 input branch through ``build_evolvepro_input`` and refuses
to leave a sample behind that its own parser rejects.
"""

from __future__ import annotations

import json
import logging
import shutil
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

_SCRIPT_DIR = Path(__file__).resolve().parent
_PYTHON_CORE = _SCRIPT_DIR.parent
_REPO_ROOT = _PYTHON_CORE.parent

for _p in [str(_REPO_ROOT), str(_PYTHON_CORE)]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

SAMPLES = _REPO_ROOT / "src-tauri" / "samples" / "mame"
REFERENCE = SAMPLES / "reference.fasta"

#: The KURO half of the same campaign. ``Help > Load Sample Data`` opens these
#: two in KURO: the round-0 EVOLVEpro prediction the operator picks from, and
#: the plasmid the primers are designed against. The MAME plate is built from
#: them rather than from a list of its own, because a demo whose two halves name
#: different variants is two demos. The plasmid CDS and ``reference.fasta`` are
#: the same 239-residue protein, which :func:`select_variants` re-checks.
EVOLVEPRO_CSV = _REPO_ROOT / "src-tauri" / "samples" / "sample_evolvepro.csv"
PLASMID = _REPO_ROOT / "src-tauri" / "samples" / "sample_plasmid.gb"

#: How many of the predicted candidates reach the plate. 10 resolve to a clean
#: PASS at FINAL (selected-replicate) verdict -- one of them, G190A, also
#: demonstrates why a plate is triplicate-sequenced (see
#: ``_LOWDEPTH_REPLICATE_VARIANT`` below), and another, K239A, demonstrates it
#: from the opposite direction (see ``_AMBIGUOUS_REPLICATE_VARIANT`` below).
#: The other 6 each demonstrate one of the remaining non-PASS, non-AMBIGUOUS
#: ``VerdictClass`` values at FINAL, so 7 of the 8 classes in
#: ``kuma_core.mame.models.VerdictClass`` are reachable at FINAL from the
#: bundled sample (all but AMBIGUOUS: see ``_AMBIGUOUS_REPLICATE_VARIANT`` for
#: why FINAL never selects it here). AMBIGUOUS is still reachable, only at the
#: replicate-comparison layer rather than FINAL. See ``_TARGET_VERDICT`` below
#: for which variant carries which class and how its consensus is built to
#: earn it honestly through ``kuma_core.mame.compare.verdict.classify_verdict``
#: rather than being asserted.
VARIANT_COUNT = 16

#: Activity relative to wild-type, per variant, as the demo reports it. Three
#: replicate measurements each, spread by a per-variant amount rather than a
#: single constant so a reader cannot mistake the spread for a fixed artefact.
#:
#: The ranking is deliberately not the prediction ranking. ``A88V`` is the best
#: measured variant while ``F28A`` was predicted highest, and three candidates
#: land below wild-type. A demo where the model ordered the assay perfectly
#: would be teaching the wrong thing about what a round of screening is for.
#:
#: Keys must be exactly the selected variants. :func:`select_variants` raises
#: when they drift apart rather than letting a renamed variant fall out of the
#: measurements without a word.
ACTIVITY: dict[str, tuple[float, float, float]] = {
    "F28A": (2.41, 2.38, 2.44),
    "H78A": (1.86, 1.83, 1.90),
    "A88V": (3.12, 3.07, 3.15),
    "R97A": (0.62, 0.59, 0.66),
    "I124A": (1.54, 1.51, 1.57),
    "N150A": (1.19, 1.16, 1.22),
    "K163A": (0.88, 0.85, 0.91),
    "G175A": (1.33, 1.30, 1.36),
    "G190A": (0.47, 0.44, 0.50),
    # The 7 below are measured like every other well (a wet-lab operator plates
    # and reads activity before NGS comes back) but do not survive NGS
    # confirmation; see ``_TARGET_VERDICT``. Their values are ordinary points on
    # the same scale, not flagged in any way, because the assay has no way to
    # know a well's NGS class in advance.
    "H200A": (1.05, 1.02, 1.08),
    "K215A": (0.95, 0.92, 0.98),
    "Q205A": (1.42, 1.38, 1.46),
    "K210A": (0.71, 0.68, 0.74),
    "A227V": (1.68, 1.64, 1.72),
    "D235A": (0.55, 0.52, 0.58),
    "K239A": (1.20, 1.16, 1.24),
}

#: Wild-type replicates on the raw scale. Every raw value below is this mean
#: times the relative activity above, so the two long-format files state the
#: same measurements on two scales and can be compared against each other.
WT_RAW: tuple[float, float, float] = (1011.0, 962.0, 982.0)

#: Confirmation re-measurement, a subset re-run on the instrument. Values sit
#: close to but not on the primary numbers, which is what a repeat measurement
#: looks like and what the merge step exists to reconcile.
CONFIRMATION: dict[str, tuple[float, float]] = {
    "A88V": (3.18, 3.21),
    "F28A": (2.36, 2.39),
    "H78A": (1.89, 1.91),
}


@dataclass(frozen=True)
class Variant:
    """One designed substitution, resolved against the reference."""

    mutant_id: str      # internal notation, e.g. "Y67H"
    position: int
    wt_aa: str
    mt_aa: str
    wt_codon: str
    mt_codon: str
    note: str

    @property
    def short(self) -> str:
        """EVOLVEpro notation, e.g. ``67H``."""
        return f"{self.position}{self.mt_aa}"


# ---------------------------------------------------------------------------
# reference
# ---------------------------------------------------------------------------

def read_reference_cds() -> str:
    lines = REFERENCE.read_text(encoding="utf-8").splitlines()
    return "".join(line.strip() for line in lines if line and not line.startswith(">"))


def translate(cds: str) -> str:
    from Bio.Seq import Seq

    return str(Seq(cds).translate())


def plasmid_cds_start() -> int:
    """Where the CDS begins in the KURO sample plasmid."""
    from Bio import SeqIO

    record = next(SeqIO.parse(str(PLASMID), "genbank"))
    coding = [f for f in record.features if f.type == "CDS"]
    if not coding:
        raise ValueError(f"{PLASMID.name} has no CDS feature")
    return int(coding[0].location.start)


def read_predictions() -> list[str]:
    """The round-0 candidates, best prediction first."""
    import csv

    with EVOLVEPRO_CSV.open(encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    ranked = sorted(rows, key=lambda row: float(row["y_pred"]), reverse=True)
    return [row["variant"].strip() for row in ranked]


def select_variants(cds: str) -> list[Variant]:
    """The variants this campaign actually put on a plate.

    The list is derived rather than written down. The predictions name the
    candidates, KURO says which of them it can build primers for, and the best
    predicted survivors go to the bench. A candidate whose primers do not design
    never reaches a well, so including one would put a variant on the plate that
    the campaign could not have made.

    Codons come from the KURO design rather than from a table here, so the
    expected-mutation sheet states the codon the primers actually carry.
    """
    from kuma_core.kuro.sdm_engine import design_sdm_primers

    protein = translate(cds)
    predictions = read_predictions()

    with tempfile.TemporaryDirectory(prefix="kuro_design_") as _tmp:
        csv_path = Path(_tmp) / "mutations.csv"
        csv_path.write_text(
            "mutation\n" + "".join(f"{v}\n" for v in predictions), encoding="utf-8"
        )
        results, _candidates, failed = design_sdm_primers(
            fasta_path=PLASMID,
            target_start=plasmid_cds_start(),
            mutations_csv=csv_path,
            polymerase="KOD",
        )

    designed = {str(result.mutation.raw): result.mutation for result in results}
    print(
        f"KURO designed {len(designed)} of {len(predictions)} predicted candidates; "
        f"{len(failed)} could not be built"
    )
    for variant, reason in failed.items():
        print(f"  no primers for {variant}: {str(reason).split(' - ')[0]}")

    picked = [v for v in predictions if v in designed][:VARIANT_COUNT]
    if len(picked) < VARIANT_COUNT:
        raise ValueError(
            f"only {len(picked)} of the predicted candidates designed, "
            f"{VARIANT_COUNT} are needed for the plate"
        )

    missing = set(picked) - set(ACTIVITY)
    extra = set(ACTIVITY) - set(picked)
    if missing or extra:
        raise ValueError(
            "ACTIVITY does not describe the selected variants: "
            f"missing {sorted(missing)}, unused {sorted(extra)}"
        )

    resolved: list[Variant] = []
    for name in picked:
        mutation = designed[name]
        position = int(mutation.position)
        if position > len(protein) or protein[position - 1] != mutation.wt_aa:
            raise ValueError(
                f"{name} names {mutation.wt_aa}{position}, but the MAME "
                f"reference holds {protein[position - 1: position] or '(past the end)'}"
            )
        resolved.append(
            Variant(
                mutant_id=name,
                position=position,
                wt_aa=str(mutation.wt_aa),
                mt_aa=str(mutation.mt_aa),
                wt_codon=str(mutation.wt_codon),
                mt_codon=str(mutation.mt_codon),
                note=f"EVOLVEpro rank {predictions.index(name) + 1}",
            )
        )
    # Ascending position, which is not cosmetic. The plate is filled in file
    # order by ``build_draft_layout`` while the numeric-ID decoder derives its
    # plate order from ``expected_variant_order``, which sorts by position. The
    # two agree only when the file is already sorted, and a set that disagreed
    # with itself would place a variant in one well and score it in another.
    return sorted(resolved, key=lambda variant: variant.position)


def apply_substitution(cds: str, variant: Variant) -> str:
    start = (variant.position - 1) * 3
    return cds[:start] + variant.mt_codon + cds[start + 3 :]


# ---------------------------------------------------------------------------
# 03 expected mutations
# ---------------------------------------------------------------------------

_EXPECTED_HEADER = [
    "mutant_id",
    "position",
    "wt_aa",
    "mt_aa",
    "wt_codon",
    "mt_codon",
    "group_id",
    "primer_set_ref",
    "notation_type",
    "status",
]


def write_expected_mutations(variants: list[Variant], path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "expected_mutations"
    sheet.append(_EXPECTED_HEADER)
    for index, variant in enumerate(variants, 1):
        sheet.append(
            [
                variant.mutant_id,
                variant.position,
                variant.wt_aa,
                variant.mt_aa,
                variant.wt_codon,
                variant.mt_codon,
                f"G{index}",
                f"PS_{index:03d}",
                "single",
                "designed",
            ]
        )
    # The control row, last. It is an occupant rather than an annotation: the
    # plate reserves a well for it, and the row number it sits at is the well it
    # gets. A file without it places the control by default somewhere at the end
    # of the plate, which is a different plate than the one the operator wrote
    # down. It carries no residue, so its position reads as 0 and nothing
    # compares it.
    sheet.append(["WT", "n/a", "-", "-", "-", "-", "G0", "-", "wt", "control"])
    workbook.save(path)


# ---------------------------------------------------------------------------
# consensus FASTA, one well per occupant
# ---------------------------------------------------------------------------

def seq_to_token(seq: int) -> str:
    """``{reverse}_{forward}`` barcode token for a 1-based occupant index."""
    from kuma_core.mame.plate_geometry import DEFAULT_ADDRESSING

    row, col = DEFAULT_ADDRESSING.seq_to_rc(seq)
    return f"{row}_{col}"


#: Three native barcodes: the same plate, sequenced on three separate nanopore
#: loads. MAME's plate model allows exactly one well per variant (see
#: :func:`write_plate_layout`), so the NGS replicate axis this campaign needs
#: lives here, in which native-barcode consensus directory a well's FASTA
#: sits in, not in extra wells.
NATIVE_BARCODES: tuple[str, ...] = ("NB01", "NB02", "NB03")

#: Per-run depth and per-run background purity, one nanopore load each. No two
#: loads off the same library pool return the same read count or the same
#: background error rate, so the three are given different baselines rather
#: than three copies of one number: NB02 ran deepest but noisiest, NB03 ran
#: shallow but purest, NB01 sits in between. Numbers are illustrative
#: nanopore-range values, not measurements.
_NB_PROFILE: dict[str, dict[str, float]] = {
    "NB01": {"depth": 230.0, "purity": 0.960},
    "NB02": {"depth": 280.0, "purity": 0.945},
    "NB03": {"depth": 195.0, "purity": 0.970},
}

#: Per-well multiplicative/additive jitter around a load's baseline, so a well
#: is not simply its native barcode's constant. ``random.Random`` is seeded on
#: ``f"{mutant_id}:{nb}"`` so the jitter is reproducible across re-runs of this
#: script (see module docstring: "The script is re-runnable") while still
#: differing well to well and barcode to barcode. With this spread,
#: :func:`kuma_core.mame.select.best_pick.pick_best_replicate`'s Wilson-bound
#: tiebreak (kuma_core/mame/select/best_pick.py) picks NB01 for 5 of the 10
#: variants that resolve to PASS at FINAL, NB03 for 5, and NB02 for 1 in the
#: current fixture (counts are read off the regenerated fixture, not
#: hand-kept, because they move whenever a variant or its jitter changes):
#: sequencing order (NB ascending) does not decide the winner, purity does.
_DEPTH_JITTER = (0.80, 1.20)
_PURITY_JITTER = (-0.025, 0.025)


def _nb_stats(mutant_id: str, nb: str) -> tuple[int, float]:
    """Return ``(depth, purity)`` for one (variant, native-barcode) well.

    ``purity`` is the fraction of reads at the designed substitution's codon
    that agree with the called allele, i.e. the ``min_variant_support`` this
    campaign's consensus caller would have reported. Only meaningful for a
    variant well: a WT well calls no substitution, so callers must not read
    this ``purity`` for the control.
    """
    import random

    profile = _NB_PROFILE[nb]
    rng = random.Random(f"{mutant_id}:{nb}")
    depth = max(30, round(profile["depth"] * rng.uniform(*_DEPTH_JITTER)))
    purity = min(0.995, max(0.80, profile["purity"] + rng.uniform(*_PURITY_JITTER)))
    return depth, purity


#: Standard codon for each amino acid, used only to manufacture an AA change
#: at a chosen position. Any codon coding the target amino acid works; these
#: are arbitrary picks, not a usage table.
_AA_CODON: dict[str, str] = {
    "A": "GCC", "R": "CGC", "N": "AAC", "D": "GAC", "C": "TGC", "Q": "CAG",
    "E": "GAG", "G": "GGC", "H": "CAC", "I": "ATC", "L": "CTC", "K": "AAG",
    "M": "ATG", "F": "TTC", "P": "CCC", "S": "AGC", "T": "ACC", "W": "TGG",
    "Y": "TAC", "V": "GTC",
}


def _codon_aa(codon: str) -> str:
    from Bio.Seq import Seq

    return str(Seq(codon).translate())


def _codon_for_other_aa(seq: str, position: int, *avoid: str) -> str:
    """A codon for 1-based amino-acid ``position`` that differs from what
    ``seq`` currently holds there and from every amino acid in ``avoid``."""
    start = (position - 1) * 3
    current_aa = _codon_aa(seq[start : start + 3])
    banned = {current_aa, *avoid}
    for aa, codon in _AA_CODON.items():
        if aa not in banned:
            return codon
    raise AssertionError("no amino acid left to pick")


def _apply_other_aa(seq: str, position: int, *avoid: str) -> str:
    start = (position - 1) * 3
    codon = _codon_for_other_aa(seq, position, *avoid)
    return seq[:start] + codon + seq[start + 3 :]


#: Each of these variants demonstrates one non-PASS ``VerdictClass`` at FINAL
#: (selected-replicate) verdict, applied identically across all three native
#: barcodes so the class is not an artefact of the triplicate tiebreak. Gate
#: order and thresholds are read from ``kuma_core/mame/compare/verdict.py``
#: (module docstring: "LOWDEPTH -> FRAMESHIFT -> INDEL_EVENT(-> AMBIGUOUS) ->
#: NO_CALL -> MANY -> MIXED -> WRONG_AA -> AMBIGUOUS -> PASS") and
#: ``kuma_core/mame/models.py`` (``CompareParams`` defaults):
#:
#: * WRONG_AA (``H200A``): the consensus carries a different amino acid than
#:   designed at the expected position (``classify_verdict`` step "4) WRONG_AA").
#: * MANY (``K215A``): the designed substitution is present, plus 5 more AA
#:   changes elsewhere -- over ``CompareParams.many_mutation_cutoff`` (5) and
#:   over the well's own 1 expected mutation (step "3) MANY").
#: * FRAMESHIFT (``Q205A``): the consensus header's ``consensus_net_indel`` is
#:   not a multiple of 3 (the net-indel check, second gate, right after
#:   LOWDEPTH: ``net_indel is not None and net_indel % 3 != 0``).
#: * MIXED (``K210A``): header ``mixed_positions`` > 0 at a depth at/above the
#:   confident-mixed floor (``min_read_count`` (30) x ``_MIXED_CONFIDENT_DEPTH_FACTOR``
#:   (3) = 90), so it reads as contamination rather than being downgraded to
#:   LOWDEPTH by the floor check inside the MIXED gate.
#: * LOWDEPTH (``A227V``): header ``depth``/``aligned_reads`` sits under
#:   ``CompareParams.min_read_count`` (30), the very first gate.
#: * NO_CALL (``D235A``): the consensus sequence itself carries N bases well
#:   away from the designed codon, so the covered-scoped ``consensus_n_fraction``
#:   recovered from them (``fasta_parser._recover_covered_n_fraction``) exceeds
#:   ``CompareParams.max_consensus_n_fraction`` (0.0).
#: * AMBIGUOUS (``K239A``): the designed substitution is present, plus one
#:   extra AA change within ``CompareParams.indel_window_codon`` (5) codons of
#:   it (step "5) AMBIGUOUS", the window check). Unlike the other six classes
#:   above, this departure is applied to only one of the three native barcodes
#:   (``_AMBIGUOUS_REPLICATE_NB``, see :func:`_special_sequence` for why); the
#:   FINAL (selected-replicate) verdict for ``K239A`` is PASS, by way of the
#:   other two barcodes, and AMBIGUOUS itself is reachable only in the
#:   replicate-comparison view. FINAL therefore surfaces 7 of the 8
#:   ``VerdictClass`` values (all but AMBIGUOUS); the 8th is intentionally not
#:   pushed through the picker (see ``_AMBIGUOUS_REPLICATE_VARIANT``).
_TARGET_VERDICT: dict[str, str] = {
    "H200A": "WRONG_AA",
    "K215A": "MANY",
    "Q205A": "FRAMESHIFT",
    "K210A": "MIXED",
    "A227V": "LOWDEPTH",
    "D235A": "NO_CALL",
    "K239A": "AMBIGUOUS",
}

#: The one PASS variant whose three native barcodes do not all read the same
#: class. Its NB01 consensus is deliberately shallow (LOWDEPTH, header only);
#: NB02 and NB03 read the clean designed substitution at full depth.
#: :func:`kuma_core.mame.select.best_pick.pick_best_replicate` ranks PASS
#: above LOWDEPTH (``PRIORITY_ORDER``), so the FINAL verdict is PASS by way of
#: NB02/NB03 -- the sample demonstrates why a plate is sequenced three times:
#: one bad load does not sink a well the other two loads confirm.
_LOWDEPTH_REPLICATE_VARIANT = "G190A"
_LOWDEPTH_REPLICATE_NB = "NB01"

#: The one AMBIGUOUS variant whose three native barcodes do not all read the
#: same class, for the opposite reason ``_LOWDEPTH_REPLICATE_VARIANT`` exists.
#: ``kuma_core.mame.select.best_pick.PRIORITY_ORDER`` is
#: ``[PASS, AMBIGUOUS, LOWDEPTH]``, so a variant whose three native barcodes
#: are *all* AMBIGUOUS has no PASS candidate for the picker to prefer and
#: AMBIGUOUS becomes its FINAL (selected-replicate) verdict -- on the Analyze
#: screen that reads as "the app chose an ambiguous well", which the
#: selection layer (``pick_best_replicate``) exists specifically to avoid for
#: a variant that has a clean well available. Only ``_AMBIGUOUS_REPLICATE_NB``
#: carries the AMBIGUOUS departure (see :func:`_special_sequence`); the other
#: two native barcodes read the plain, correct substitution and score PASS,
#: so the picker resolves this variant to PASS at FINAL and AMBIGUOUS is
#: reachable only by comparing its three replicates against each other. Do
#: not apply this departure to all three barcodes to "recover" an AMBIGUOUS
#: FINAL: that is the exact state the user reported as wrong (2026-08-31,
#: bundled K239A sample had FINAL=AMBIGUOUS chosen over two other AMBIGUOUS
#: wells) and is not a state a triplicate-sequenced plate should model as a
#: normal outcome for a variant that also has good data.
_AMBIGUOUS_REPLICATE_VARIANT = "K239A"
_AMBIGUOUS_REPLICATE_NB = "NB01"

#: Header-only LOWDEPTH depth: under ``CompareParams.min_read_count`` (30).
_SHALLOW_DEPTH = 14

#: Header-only MIXED depth: at/above the confident-mixed floor (90) so the
#: mixed signal is not downgraded to LOWDEPTH by the floor inside that gate.
_MIXED_DEPTH = 200
_MIXED_MINOR_ALLELE_FRACTION = 0.30


def _special_sequence(
    sample: str, sequence: str, variant: Variant, protein_len: int, nb: str
) -> str:
    """Depart from the correct single substitution for a ``_TARGET_VERDICT`` well.

    Only WRONG_AA, MANY, AMBIGUOUS and NO_CALL act on the sequence; FRAMESHIFT,
    MIXED and LOWDEPTH act on header metadata only (see :func:`_consensus_header`),
    so this returns ``sequence`` unchanged for them.

    AMBIGUOUS is the one class this campaign's selection layer, not just its
    scoring layer, treats as non-clean: ``kuma_core.mame.select.best_pick``'s
    ``PRIORITY_ORDER`` ranks AMBIGUOUS above only LOWDEPTH, so a variant whose
    three native barcodes are *all* AMBIGUOUS gets AMBIGUOUS picked as its
    FINAL (selected-replicate) verdict -- there is no PASS well for the picker
    to prefer. That reads on the Analyze screen as "the app picked an
    ambiguous well", which is not what triplicate sequencing is supposed to
    let happen. So unlike WRONG_AA/MANY/FRAMESHIFT/MIXED/NO_CALL, which stay
    applied to every native barcode because nothing above them in
    ``PRIORITY_ORDER`` (or ``_FALLBACK_ELIGIBLE``) can rescue a clean well from
    the fallback path anyway, AMBIGUOUS is applied to exactly one native
    barcode (``_AMBIGUOUS_REPLICATE_NB``). The other two read the plain,
    correct substitution and score PASS, so the picker still resolves
    ``_AMBIGUOUS_REPLICATE_VARIANT`` to PASS at FINAL -- the same shape as
    ``_LOWDEPTH_REPLICATE_VARIANT`` below, one bad load among three good ones.
    AMBIGUOUS itself remains fully reachable, just at the replicate-comparison
    layer rather than at FINAL: see ``_AMBIGUOUS_REPLICATE_VARIANT``.
    """
    target = _TARGET_VERDICT.get(sample)
    if target == "WRONG_AA":
        return _apply_other_aa(sequence, variant.position, variant.wt_aa)
    if target == "MANY":
        for offset in (3, 6, 9, -3, -6):
            pos = variant.position + offset
            if 2 <= pos <= protein_len - 1:
                sequence = _apply_other_aa(sequence, pos)
        return sequence
    if target == "AMBIGUOUS":
        if nb != _AMBIGUOUS_REPLICATE_NB:
            return sequence
        pos = variant.position + 3
        if pos > protein_len - 1:
            pos = variant.position - 3
        return _apply_other_aa(sequence, pos)
    if target == "NO_CALL":
        # First two codons: far from every designed position in this campaign
        # (all sit past residue 28), so the N run never touches the codon
        # NO_CALL is meant to be scored against, and NO_CALL returns before
        # any AA diff is inspected anyway.
        start = 3
        return sequence[:start] + "N" * 5 + sequence[start + 5 :]
    return sequence


def _consensus_header(nb: str, mutant_id: str, is_control: bool) -> str:
    """One consensus FASTA header's metadata, depth/purity resolved per well.

    ``max_minor_allele_fraction``/``mixed_positions`` and depth stay fixed at
    their clean defaults across every well and native barcode except where
    ``_TARGET_VERDICT`` or ``_LOWDEPTH_REPLICATE_VARIANT`` names a departure
    (see the constants above for exactly which class each departure earns and
    why).

    ``variant_positions``/``min_variant_support``/``min_variant_support_depth``
    are only written for a variant well. A WT well calls no substitution, and
    :func:`kuma_core.mame.select.purity.support_lower_bound` already treats a
    header missing them as "not evaluable" rather than zero support, so
    omitting them for WT states exactly that.
    """
    depth, purity = _nb_stats(mutant_id, nb)
    mixed_positions = 0
    max_minor_allele_fraction = 0.04
    net_indel: int | None = None

    target = _TARGET_VERDICT.get(mutant_id)
    if target == "LOWDEPTH":
        depth = _SHALLOW_DEPTH
    elif target == "MIXED":
        depth = _MIXED_DEPTH
        mixed_positions = 1
        max_minor_allele_fraction = _MIXED_MINOR_ALLELE_FRACTION
    elif target == "FRAMESHIFT":
        net_indel = 1  # not a multiple of 3
    elif mutant_id == _LOWDEPTH_REPLICATE_VARIANT and nb == _LOWDEPTH_REPLICATE_NB:
        depth = _SHALLOW_DEPTH

    aligned_reads = depth
    parts = [
        f"depth={depth}",
        f"input_reads={depth + 8}",
        f"aligned_reads={aligned_reads}",
        "mapq_failed=5",
        "span_failed=3",
        "low_depth_positions=0",
        "consensus_n_fraction=0.000",
        "low_quality_bases=0",
        f"max_minor_allele_fraction={max_minor_allele_fraction:.2f}",
        f"mixed_positions={mixed_positions}",
    ]
    if net_indel is not None:
        parts.append(f"consensus_net_indel={net_indel}")
    if not is_control:
        parts.append("variant_positions=1")
        parts.append(f"min_variant_support={purity:.4f}")
        parts.append(f"min_variant_support_depth={depth}")
    return " ".join(parts)


def build_consensus_dir(
    workdir: Path,
    cds: str,
    layout: dict[str, str],
    by_id: dict[str, Variant],
    native_barcodes: tuple[str, ...] = NATIVE_BARCODES,
) -> Path:
    """One consensus FASTA per occupied well, once per native barcode.

    Writing the same well into ``NATIVE_BARCODES`` directories is the
    triplicate: the plate is not re-arranged, it is re-sequenced.
    :func:`kuma_core.mame.pipeline.run_analyze` groups verdicts by
    ``native_barcode`` (the directory name), so this is what turns three
    directories into three replicates per mutant on the Analyze screen.
    """
    from kuma_core.mame.plate_geometry import well_to_seq

    protein_len = len(cds) // 3
    consensus = workdir / "consensus"
    for nb in native_barcodes:
        native = consensus / nb
        native.mkdir(parents=True)
        for well, sample in layout.items():
            token = seq_to_token(well_to_seq(well))
            if sample == "WT":
                sequence = cds
            else:
                sequence = apply_substitution(cds, by_id[sample])
                sequence = _special_sequence(sample, sequence, by_id[sample], protein_len, nb)
            header = _consensus_header(nb, sample, is_control=sample == "WT")
            (native / f"{token}.fasta").write_text(
                f">{token} {header}\n{sequence}\n", encoding="utf-8"
            )
    return consensus


# ---------------------------------------------------------------------------
# activity samples, all derived from the placement
# ---------------------------------------------------------------------------

def _variant_wells(layout: dict[str, str]) -> list[tuple[str, str]]:
    """``(well, mutant_id)`` for the occupied non-control wells, in plate order."""
    from kuma_core.mame.plate_geometry import well_to_seq

    wells = [(well, sample) for well, sample in layout.items() if sample != "WT"]
    return sorted(wells, key=lambda item: well_to_seq(item[0]))


def _wt_wells(layout: dict[str, str]) -> list[str]:
    from kuma_core.mame.plate_geometry import well_to_seq

    return sorted(
        (well for well, sample in layout.items() if sample == "WT"),
        key=well_to_seq,
    )


def write_plate_layout(layout: dict[str, str], path: Path) -> None:
    """The placement as the operator-facing layout sheet.

    One well per variant. A repeated well would name one variant twice, and both
    the layout reader and the verdict reader refuse that: a variant that sits in
    two wells has no single well for the NGS evidence to be attached to.
    """
    import openpyxl
    from kuma_core.mame.plate_geometry import well_to_seq

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Plate Layout"
    sheet.append(["Mutant", "Well Pos."])
    for well in sorted(layout, key=well_to_seq):
        sheet.append([layout[well], well])
    workbook.save(path)


def write_long_relative(layout: dict[str, str], csv_path: Path, xlsx_path: Path) -> None:
    """Well-labeled long format, already relative to wild-type."""
    import openpyxl

    rows: list[tuple[str, str, float, int]] = []
    for well, mutant_id in _variant_wells(layout):
        for replicate, value in enumerate(ACTIVITY[mutant_id], 1):
            rows.append(("plate01", well, value, replicate))
    # The wild-type rows carry a ``WT_n`` label rather than their well, the same
    # spelling the raw file uses. On this scale they are already 1.0 and nothing
    # divides by them, but labelling them as an occupant would send them into
    # the well<->variant mapping, where the control well is deliberately absent.
    for replicate in (1, 2, 3):
        rows.append(("plate01", f"WT_{replicate}", 1.0, replicate))

    header = "plate_id,well_id,value,replicate_idx\n"
    body = "".join(f"{p},{w},{v},{r}\n" for p, w, v, r in rows)
    csv_path.write_text(header + body, encoding="utf-8")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "activity"
    sheet.append(["plate_id", "well_id", "value", "replicate_idx"])
    for row in rows:
        sheet.append(list(row))
    workbook.save(xlsx_path)


def write_long_raw(layout: dict[str, str], path: Path) -> None:
    """The same measurements before normalisation.

    ``activity_scale="raw"`` divides each cohort by the mean of its own ``WT_n``
    rows, so those rows carry a wild-type label rather than a well: they are the
    denominator, not an occupant to be scored.
    """
    wt_mean = sum(WT_RAW) / len(WT_RAW)
    lines = ["plate_id,well_id,value,replicate_idx"]
    for index, value in enumerate(WT_RAW, 1):
        lines.append(f"plate01,WT_{index},{value:.1f},{index}")
    for well, mutant_id in _variant_wells(layout):
        for replicate, relative in enumerate(ACTIVITY[mutant_id], 1):
            lines.append(f"plate01,{well},{relative * wt_mean:.1f},{replicate}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_long_variant(path: Path) -> None:
    """Variant-labeled long format.

    The label is the variant itself, so this branch needs no plate layout and no
    well<->variant mapping: it states what was measured rather than where.
    """
    lines = ["variant,activity,replicate_idx"]
    for mutant_id, values in ACTIVITY.items():
        short = f"{mutant_id[1:-1]}{mutant_id[-1]}"
        for replicate, value in enumerate(values, 1):
            lines.append(f"{short},{value},{replicate}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_gc_prenormalised(layout: dict[str, str], path: Path) -> None:
    """GC sheet whose areas are already wild-type relative."""
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "GC_normalised"
    sheet.append(["Sample Name", "Area"])
    for well in _wt_wells(layout):
        sheet.append([well, 1.0])
    for well, mutant_id in _variant_wells(layout):
        sheet.append([well, round(sum(ACTIVITY[mutant_id]) / 3, 3)])
    workbook.save(path)


def _agilent_blocks(rows: list[tuple[str, float]], title: str, path: Path) -> None:
    """An Agilent FID report: one two-line block per injection.

    The shape is the instrument export, not a table: each injection restates the
    signal header, then its ``Area``/``Sample Name`` pair, then a ``Sum`` line.
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = title
    for name, area in rows:
        sheet.append(["Signal:", "FID1B"])
        sheet.append(["Area", "Sample Name"])
        sheet.append([area, name])
        sheet.append(["Sum", area])
        sheet.append(["", ""])
    workbook.save(path)


def write_round1_raw(layout: dict[str, str], path: Path) -> None:
    """Raw FID areas, normalised downstream against this report's own WT block."""
    wt_mean = sum(WT_RAW) / len(WT_RAW)
    rows: list[tuple[str, float]] = [
        (f"WT{index}", value) for index, value in enumerate(WT_RAW, 1)
    ]
    for well, mutant_id in _variant_wells(layout):
        for relative in ACTIVITY[mutant_id]:
            rows.append((well, round(relative * wt_mean, 1)))
    _agilent_blocks(rows, "GC-FID round1 raw", path)


def write_confirmation_report(path: Path) -> None:
    """Variant-labeled re-measurement of a handful of wells."""
    wt_mean = sum(WT_RAW) / len(WT_RAW)
    rows: list[tuple[str, float]] = [
        (f"WT{index}", value) for index, value in enumerate(WT_RAW, 1)
    ]
    for mutant_id, values in CONFIRMATION.items():
        short = f"{mutant_id[1:-1]}{mutant_id[-1]}"
        for relative in values:
            rows.append((short, round(relative * wt_mean, 1)))
    _agilent_blocks(rows, "Agilent", path)


def _above_wt(layout: dict[str, str]) -> list[tuple[str, str]]:
    """``(well, mutant_id)`` for the wells the primary screen put above WT.

    A numeric-ID confirmation indexes this subset rather than the whole plate,
    because an instrument re-runs the hits and numbers what it was given.
    """
    return [
        (well, mutant_id)
        for well, mutant_id in _variant_wells(layout)
        if sum(ACTIVITY[mutant_id]) / 3 > 1.0
    ]


def write_numeric_confirmation(layout: dict[str, str], path: Path) -> None:
    """Numeric-ID re-measurement of the wells the primary put above wild-type.

    Its identifiers count the above-WT subset, not the plate: an identifier that
    indexes a different set than the decoder derives names the wrong variant,
    which is a silent relabelling rather than an error.
    """
    wt_mean = sum(WT_RAW) / len(WT_RAW)
    rows: list[tuple[str, float]] = [
        (f"WT{index}", value) for index, value in enumerate(WT_RAW, 1)
    ]
    for index, (_, mutant_id) in enumerate(_above_wt(layout), 1):
        for relative in ACTIVITY[mutant_id][:2]:
            rows.append((str(index), round(relative * wt_mean, 1)))
    _agilent_blocks(rows, "Agilent numeric confirmation", path)


def write_numeric_report(layout: dict[str, str], path: Path) -> None:
    """The same report with the instrument numbering the samples.

    The identifiers index the plate order rather than naming variants, so the
    count has to match the order it is decoded against. Shipping six identifiers
    for a ten-variant plate is what made the previous file undecodable under
    every order source.
    """
    wt_mean = sum(WT_RAW) / len(WT_RAW)
    rows: list[tuple[str, float]] = [
        (f"WT{index}", value) for index, value in enumerate(WT_RAW, 1)
    ]
    for index, (_, mutant_id) in enumerate(_variant_wells(layout), 1):
        for relative in ACTIVITY[mutant_id]:
            rows.append((str(index), round(relative * wt_mean, 1)))
    _agilent_blocks(rows, "Agilent numeric index", path)


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    from kuma_core.mame.io.variant_list import read_variant_source
    from kuma_core.mame.layout import build_draft_layout

    cds = read_reference_cds()
    variants = select_variants(cds)
    by_id = {variant.mutant_id: variant for variant in variants}
    print(f"reference: {REFERENCE.name}, {len(cds)} bp, {len(cds) // 3} residues")
    for variant in variants:
        print(f"  {variant.mutant_id:>6}  {variant.wt_codon}->{variant.mt_codon}  {variant.note}")

    expected_path = SAMPLES / "03_mame_expected_mutations.xlsx"
    write_expected_mutations(variants, expected_path)
    print(f"wrote {expected_path.relative_to(_REPO_ROOT)}")

    read = read_variant_source(expected_path)
    draft = build_draft_layout(read.expected, wt_ordinal=read.wt_ordinal)
    if draft.dropped_mutant_ids:
        raise SystemExit(f"placement dropped {draft.dropped_mutant_ids}")
    layout = dict(draft.layout)
    print(f"placement: {len(layout)} wells -> {layout}")

    with tempfile.TemporaryDirectory(prefix="mame_step4_samples_") as _tmp:
        workdir = Path(_tmp)
        consensus = build_consensus_dir(workdir, cds, layout, by_id)
        run_workbook = workdir / "analysis.xlsx"

        from kuma_core.mame.ingest import IngestMode
        from kuma_core.mame.pipeline import run_analyze

        # min_read_count=30 and max_consensus_n_fraction=0.0 are the shipped
        # CompareParams defaults (kuma_core/mame/models.py); a previous version
        # of this script passed None for both, disabling the LOWDEPTH read-count
        # gate and the NO_CALL N-fraction gate outright. Real thresholds are
        # needed for the plate to demonstrate every VerdictClass (see
        # _TARGET_VERDICT above) with LOWDEPTH and NO_CALL among them.
        verdicts, replicates = run_analyze(
            input_dir=consensus,
            reference_path=REFERENCE,
            expected_path=expected_path,
            output_path=run_workbook,
            cds_start=0,
            cds_end=len(cds),
            mode="amplicon",
            min_file_size_kb=0.0,
            min_read_count=30,
            max_consensus_n_fraction=0.0,
            many_cutoff=5,
            ingest_mode=IngestMode.BARCODE,
            well_layout=layout,
            scored_wells=set(layout),
        )
        print(f"analyze: {len(verdicts)} verdicts, {len(replicates)} replicates")

        verdict_path = SAMPLES / "13_mame_verdict.xlsx"
        shutil.copyfile(run_workbook, verdict_path)
        print(f"wrote {verdict_path.relative_to(_REPO_ROOT)}")

        write_result_fixture(verdicts, replicates, run_workbook)

    write_plate_layout(layout, SAMPLES / "06_mame_plate_layout.xlsx")
    write_long_relative(
        layout,
        SAMPLES / "07_mame_activity_long.csv",
        SAMPLES / "07_mame_activity_long.xlsx",
    )
    write_long_raw(layout, SAMPLES / "14_mame_activity_long_raw.csv")
    write_long_variant(SAMPLES / "15_mame_activity_variant.csv")
    write_gc_prenormalised(layout, SAMPLES / "10_mame_gc_prenormalised.xlsx")
    write_round1_raw(layout, SAMPLES / "11_mame_gc_fid_round1_raw.xlsx")
    write_confirmation_report(SAMPLES / "09_mame_agilent_rep_batch.xlsx")
    write_numeric_report(layout, SAMPLES / "12_mame_agilent_numeric_index.xlsx")
    write_numeric_confirmation(
        layout, SAMPLES / "16_mame_agilent_numeric_confirmation.xlsx"
    )
    print("wrote the activity sample set")


def write_result_fixture(verdicts, replicates, run_workbook: Path) -> None:
    """Serialise the Analyze screen fixture through the sidecar handlers.

    The handlers rather than a hand-built dict: a copy of their output goes
    stale silently, and a fixture that has lost a field shows the user a panel
    that cannot state what it is missing.
    """
    from sidecar_mame import core as sidecar_core
    from sidecar_mame.core import SidecarState, set_last_analyze
    from sidecar_mame.handlers.analyze import (
        _serialize_replicate,
        _serialize_verdict,
        _summarize,
    )
    from sidecar_mame.handlers.export import handle_get_plate_data
    from sidecar_mame.handlers.health import handle_get_run_health

    with sidecar_core._state_lock:
        sidecar_core._state = SidecarState()
    set_last_analyze(verdicts, replicates, str(run_workbook), run_meta=None)

    fixture = {
        "schema": 1,
        "verdicts": [_serialize_verdict(v) for v in verdicts],
        "replicates": [_serialize_replicate(r) for r in replicates],
        "summary": _summarize(verdicts),
        "wells": handle_get_plate_data({})["wells"],
        "runHealth": handle_get_run_health({}),
    }
    path = SAMPLES / "sample_analysis_result.json"
    path.write_text(json.dumps(fixture, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {path.relative_to(_REPO_ROOT)}: {fixture['summary']}")


def verify() -> None:
    """Run every step 4 input branch over what was just written.

    Existence is not the property that matters here. A sample file is only a
    sample if the parser it was written for accepts it, and the previous set
    shipped files that no combination of inputs could carry to a finished
    build. Each case below is one branch of the primary/confirmation contract,
    and a branch that raises stops the script rather than leaving the sample
    set in the state this function exists to rule out.
    """
    from kuma_core.mame.activity.build_evolvepro_input import build_evolvepro_input

    verdict = SAMPLES / "13_mame_verdict.xlsx"
    layout = SAMPLES / "06_mame_plate_layout.xlsx"
    expected = SAMPLES / "03_mame_expected_mutations.xlsx"

    cases: list[tuple[str, dict]] = [
        (
            "long format, well labels, already relative, with the layout sheet",
            {"activity_path": SAMPLES / "07_mame_activity_long.csv",
             "activity_scale": "relative_to_wt", "layout_xlsx": layout},
        ),
        (
            "long format, well labels, mapping taken from the verdict workbook",
            {"activity_path": SAMPLES / "07_mame_activity_long.csv",
             "activity_scale": "relative_to_wt"},
        ),
        (
            "long format, xlsx twin of the same measurements",
            {"activity_path": SAMPLES / "07_mame_activity_long.xlsx",
             "activity_scale": "relative_to_wt", "layout_xlsx": layout},
        ),
        (
            "long format, raw values divided by the WT rows",
            {"activity_path": SAMPLES / "14_mame_activity_long_raw.csv",
             "activity_scale": "raw", "layout_xlsx": layout},
        ),
        (
            "long format, variant labels, no plate layout involved",
            {"activity_path": SAMPLES / "15_mame_activity_variant.csv",
             "activity_scale": "relative_to_wt"},
        ),
        (
            "GC sheet, already relative",
            {"gc_data_xlsx": SAMPLES / "10_mame_gc_prenormalised.xlsx",
             "layout_xlsx": layout},
        ),
        (
            "raw Agilent report, normalised against its own WT block",
            {"round1_report_xlsx": SAMPLES / "11_mame_gc_fid_round1_raw.xlsx",
             "layout_xlsx": layout},
        ),
        (
            "numeric-ID report as the primary screen",
            {"numeric_report_xlsx": SAMPLES / "12_mame_agilent_numeric_index.xlsx",
             "layout_xlsx": layout, "expected_xlsx": expected},
        ),
        (
            "variant-labeled confirmation on top of the long format",
            {"activity_path": SAMPLES / "07_mame_activity_long.csv",
             "activity_scale": "relative_to_wt", "layout_xlsx": layout,
             "remeasure_report_xlsx": SAMPLES / "09_mame_agilent_rep_batch.xlsx"},
        ),
        # The design list is the order source here and the plate file is left
        # out, because a numeric confirmation accepts exactly one of the two.
        # The primary numeric case above takes both, so the two numeric paths
        # disagree about how many order sources they will hold; the samples pick
        # the spelling each one accepts rather than papering over that.
        (
            "numeric-ID confirmation on top of the long format",
            {"activity_path": SAMPLES / "07_mame_activity_long.csv",
             "activity_scale": "relative_to_wt",
             "expected_xlsx": expected,
             "remeasure_numeric_xlsx":
                 SAMPLES / "16_mame_agilent_numeric_confirmation.xlsx"},
        ),
    ]

    #: The first case is also the shipped example of what step 4 produces, so
    #: the example is the build rather than a picture of one.
    exported_example = SAMPLES / "08_mame_evolvepro_raw.xlsx"

    failures: list[str] = []
    with tempfile.TemporaryDirectory(prefix="mame_step4_verify_") as _tmp:
        out_dir = Path(_tmp)
        for index, (label, kwargs) in enumerate(cases, 1):
            destination = out_dir / f"case{index:02d}.xlsx"
            try:
                result = build_evolvepro_input(
                    destination, verdict_xlsx=verdict, **kwargs
                )
            except Exception as exc:  # noqa: BLE001 - the report is the point
                failures.append(f"  FAIL {label}\n        {type(exc).__name__}: {exc}")
                continue
            print(
                f"  ok   {label}: {result.n_variants} variants"
                f"{', ' + str(len(result.warnings)) + ' warning(s)' if result.warnings else ''}"
            )
            for warning in result.warnings:
                print(f"         warning: {warning}")
            if index == 1:
                shutil.copyfile(destination, exported_example)
                print(f"         kept as {exported_example.name}")

    if failures:
        print("\n".join(failures))
        raise SystemExit(
            f"{len(failures)} of {len(cases)} step 4 branches refused the samples"
        )
    print(f"all {len(cases)} step 4 branches built from the bundled samples")


if __name__ == "__main__":
    main()
    verify()
