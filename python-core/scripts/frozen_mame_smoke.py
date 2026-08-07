#!/usr/bin/env python3
"""Frozen MAME sidecar smoke test.

Verifies that the frozen (PyInstaller) MAME sidecar binary correctly handles:
  - ProcessPoolExecutor(mp_context=spawn) + multiprocessing.freeze_support()
  - per-NB parallel combinatorial demux over JSON-RPC 2.0
  - per-read ProcessPool demux (single native barcode, lowered threshold)
  - a barcode workbook whose annealing tail is NOT the ispS constant, i.e.
    the shape every package `barcode_package` generates per gene

Usage:
    python frozen_mame_smoke.py <path-to-frozen-mame-sidecar>

Exit codes:
    0 = PASS
    1 = FAIL (with diagnostic output)
"""

from __future__ import annotations

import gzip
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path

# Sibling import. Normally sys.path[0] is already this directory, but an
# explicit insert keeps the import working under -P / PYTHONSAFEPATH and from
# any working directory.
sys.path.insert(0, str(Path(__file__).resolve().parent))

from smoke_sidecar_io import SidecarIO, rpc_request as _rpc  # noqa: E402

# ---------------------------------------------------------------------------
# Synthetic constants (mirrored from tests/mame/test_combinatorial_demux.py)
# Self-contained: do NOT import the pytest test module.
# ---------------------------------------------------------------------------

_REF_SEQ = "ATGGCTTGCTCTGTATCCACTGAGAACGTATCTTTCACTGAGACTGAAACTGAGACCCGT"

_F_BARCODES = [
    "AATCCCACTAC",  # F1 (11 bp)
    "TGAACTGAGCG",  # F2 (11 bp)
    "TATCTGACCTT",  # F3 (11 bp)
    "ATATGAGACG",   # F4 (10 bp)
    "CGCTCATTAG",   # F5 (10 bp)
    "TAATCTCGTC",   # F6 (10 bp)
    "GCGCGATTTT",   # F7 (10 bp)
    "AGAGCACTAG",   # F8 (10 bp)
    "TGCCTTGATC",   # F9 (10 bp)
    "CTACTCAGTC",   # F10 (10 bp)
    "TCGTCTGACT",   # F11 (10 bp)
    "GAACATACGG",   # F12 (10 bp)
]

_R_BARCODES = [
    "CCCTATGACA",  # R1 (10 bp)
    "TAATGGCAAG",  # R2 (10 bp)
    "AACAAGGCGT",  # R3 (10 bp)
    "GTATGTAGAA",  # R4 (10 bp)
    "TTCTATGGGG",  # R5 (10 bp)
    "CCTCGCAACC",  # R6 (10 bp)
    "TGGATGCTTA",  # R7 (10 bp)
    "AGAGTGCGGC",  # R8 (10 bp)
]

_F_TAIL = "cacaggaggttaaacc"
_R_TAIL = "tgcgttgcgctctag"

# ---------------------------------------------------------------------------
# Non-ispS fixture (mirrored from tests/mame/test_barcode_prefix_derivation.py)
#
# Every barcode above is the ispS shape, which is why this smoke passed for as
# long as the demux held the ispS annealing tail as a constant: the frozen
# binary never once read a workbook that constant could not explain. Every
# package `barcode_package` generates is such a workbook, because it designs a
# fresh flanking primer per gene.
#
# So a second, non-ispS plate is exercised alongside the first rather than
# instead of it. Seeds are 9 bp forward and 13 bp reverse, neither of which was
# the fallback length (11 / 10), and R1 and R2 share their first 10 bases: under
# the fallback both cut to the same string and every read carrying either was
# dropped as ambiguous, leaving 4 wells of the 12 that went in. Measured on this
# exact fixture: 12 wells / 24 reads derived, 4 wells / 8 reads under the
# fallback. The fallback is deleted now (a workbook stating no shared tail is
# refused), so this step no longer distinguishes two behaviours; it pins that
# the frozen binary derives a per-gene tail at all, which is the case the ispS
# steps above cannot reach.
# ---------------------------------------------------------------------------

_GENE_F_TAIL = "GGTTCAGACGTATCCTGA"  # 18 bp, shares nothing with _F_TAIL
_GENE_R_TAIL = "AACCTGGTATCGAGCTTA"  # 18 bp, shares nothing with _R_TAIL

_GENE_F_SEEDS = [
    "AACGTTCAG", "TTGCAACGT", "CGATTGCAA", "GCTAACGTT", "TACGGTTCA",
    "ATCCGTTAG", "CCATGGATC", "GGTACCTAG", "TGACCAGTT", "AGTCCTGAA",
    "CTGAAGGTC", "GACTTCCAG",
]
_GENE_R_SEEDS = [
    "AACCGGTTACGAT", "AACCGGTTACTCA", "CGGATCCATTAGC", "GCCTAGGTAACGT",
    "TACCGGATCCAAG", "ATGCCATGGTTCA", "CCGGTTAACCGAT", "GGCCAATTGGCTA",
]

#: Rows and columns the non-ispS fixture puts on the plate, and the reads each
#: well gets. 3 x 4 x 2 = 24 reads over 12 wells.
_GENE_ROWS = (1, 2, 3)
_GENE_COLS = (1, 2, 3, 4)
_GENE_READS_PER_WELL = 2


def _reverse_complement(seq: str) -> str:
    complement = str.maketrans("ACGTacgt", "TGCAtgca")
    return seq.translate(complement)[::-1]


def _build_read(r_idx: int, f_idx: int, amplicon: str) -> str:
    """Build a synthetic read matching the real library layout (1-indexed).

    Real library structure (sense strand):
      5'-[F_barcode + F_anneal]-[insert]-[RC(R_anneal) + RC(R_barcode)]-3'
    """
    return (
        _F_BARCODES[f_idx - 1] + _F_TAIL
        + amplicon
        + _reverse_complement(_R_TAIL.upper()) + _reverse_complement(_R_BARCODES[r_idx - 1])
    )


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------

def _build_reference_fasta(workdir: Path) -> Path:
    ref = workdir / "reference.fasta"
    ref.write_text(f">sispS_test\n{_REF_SEQ}\n", encoding="utf-8")
    return ref


def _build_barcodes_xlsx(workdir: Path) -> Path:
    try:
        import openpyxl
    except ImportError as exc:
        print(f"FAIL: openpyxl not available — install it with: pip install openpyxl")
        print(f"  ImportError: {exc}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    for i, bc in enumerate(_F_BARCODES, start=1):
        ws.append([f"isps_f_{i}", bc.lower() + _F_TAIL])
    for i, bc in enumerate(_R_BARCODES, start=1):
        ws.append([f"isps_r_{i}", bc.lower() + _R_TAIL])

    path = workdir / "barcodes.xlsx"
    wb.save(path)
    return path


def _build_fastq_gz(barcode_dir: Path, n_reads: int = 18) -> None:
    """Write ~18 gzipped synthetic reads into barcode_dir/reads.fastq.gz."""
    barcode_dir.mkdir(parents=True, exist_ok=True)
    fastq_path = barcode_dir / "reads.fastq.gz"
    amplicon = _REF_SEQ  # full 60 bp; proven to align in unit tests

    reads: list[tuple[str, str]] = []
    for i in range(n_reads):
        # Rotate through r_idx in {1..3} and f_idx in {1..4} for variety
        r_idx = (i % 3) + 1
        f_idx = (i % 4) + 1
        seq = _build_read(r_idx, f_idx, amplicon)
        reads.append((f"read_{i}", seq))

    with gzip.open(fastq_path, "wt", encoding="utf-8") as fh:
        for read_id, seq in reads:
            qual = "I" * len(seq)
            fh.write(f"@{read_id}\n{seq}\n+\n{qual}\n")


def _build_run_dir(workdir: Path) -> Path:
    """Create MinKNOW run dir with barcode06 and barcode20 fastq_pass dirs."""
    run_dir = workdir / "RUN"
    for barcode in ("barcode06", "barcode20"):
        _build_fastq_gz(run_dir / "fastq_pass" / barcode)
    return run_dir


def _build_gene_read(r_idx: int, f_idx: int, amplicon: str) -> str:
    """The same library layout as ``_build_read``, with the non-ispS flanks."""
    return (
        _GENE_F_SEEDS[f_idx - 1] + _GENE_F_TAIL
        + amplicon
        + _reverse_complement(_GENE_R_TAIL) + _reverse_complement(_GENE_R_SEEDS[r_idx - 1])
    )


def _build_gene_barcodes_xlsx(workdir: Path) -> Path:
    """A barcode workbook of the shape ``barcode_package`` writes per gene."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None

    for i, seed in enumerate(_GENE_F_SEEDS, start=1):
        ws.append([f"mygene_f_{i}", seed + _GENE_F_TAIL.lower()])
    for i, seed in enumerate(_GENE_R_SEEDS, start=1):
        ws.append([f"mygene_r_{i}", seed + _GENE_R_TAIL.lower()])

    path = workdir / "barcodes_gene.xlsx"
    wb.save(path)
    return path


def _build_gene_run_dir(workdir: Path) -> Path:
    """A one-native-barcode run whose reads carry the non-ispS plate."""
    run_dir = workdir / "RUN_GENE"
    barcode_dir = run_dir / "fastq_pass" / "barcode06"
    barcode_dir.mkdir(parents=True, exist_ok=True)
    fastq_path = barcode_dir / "reads.fastq.gz"
    with gzip.open(fastq_path, "wt", encoding="utf-8") as fh:
        for r_idx in _GENE_ROWS:
            for f_idx in _GENE_COLS:
                for k in range(_GENE_READS_PER_WELL):
                    seq = _build_gene_read(r_idx, f_idx, _REF_SEQ)
                    fh.write(f"@gene_{r_idx}_{f_idx}_{k}\n{seq}\n+\n{'I' * len(seq)}\n")
    return run_dir


def _build_single_nb_run_dir(workdir: Path) -> Path:
    """Create a MinKNOW run dir with exactly ONE native barcode.

    ``run_combinatorial_demux_multi`` only sets ``per_read_parallel=True`` when
    n_nb == 1 (with n_nb > 1 the per-NB pool already owns the cores). Combined
    with a lowered ``KUMA_MAME_PERREAD_THRESHOLD`` this is what drives the
    per-read spawn ProcessPool, a path the two-barcode fixture never reaches.
    """
    run_dir = workdir / "RUN_SINGLE"
    _build_fastq_gz(run_dir / "fastq_pass" / "barcode06")
    return run_dir


def _build_expected_mutations_xlsx(workdir: Path) -> Path:
    """Write a minimal KURO results xlsx with an ``expected_mutations`` sheet.

    The sheet carries the exact 10-column header that
    ``kuma_core.mame.io.kuro_reader.read_expected_mutations`` requires and no
    designed data rows. A header-only sheet parses cleanly into an empty
    expected-mutation list (WT-only), so ``analyze`` runs end-to-end without
    requiring any specific mutants to be present in the synthetic reads.
    """
    try:
        import openpyxl
    except ImportError as exc:
        print(f"FAIL: openpyxl not available — install it with: pip install openpyxl")
        print(f"  ImportError: {exc}")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "expected_mutations"
    # Header must match kuro_reader._EXPECTED_HEADER exactly (order + names).
    ws.append([
        "mutant_id",
        "position",
        "wt_aa",
        "mt_aa",
        "wt_codon",
        "mt_codon",
        "group_id",
        "primer_set_ref",
        "notation_type",
        "status",
    ])

    path = workdir / "expected_mutations.xlsx"
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main smoke test
# ---------------------------------------------------------------------------

def run_smoke(binary: Path) -> None:
    workdir = Path(tempfile.mkdtemp(prefix="mame_smoke_"))
    stderr_path = workdir / "sidecar_stderr.txt"
    sio: SidecarIO | None = None
    failures: list[str] = []

    try:
        # Build fixtures
        ref_fasta = _build_reference_fasta(workdir)
        xlsx = _build_barcodes_xlsx(workdir)
        run_dir = _build_run_dir(workdir)
        single_run_dir = _build_single_nb_run_dir(workdir)
        gene_xlsx = _build_gene_barcodes_xlsx(workdir)
        gene_run_dir = _build_gene_run_dir(workdir)
        expected_xlsx = _build_expected_mutations_xlsx(workdir)
        out_dir = workdir / "output"
        out_dir.mkdir()

        # KUMA_MAME_PERREAD_THRESHOLD=1 lets step [5/7] (single native barcode)
        # reach the per-read spawn ProcessPool with the tiny synthetic fixture.
        # Steps [3/7] and [4/7] use two native barcodes, where per_read_parallel
        # is False regardless of the threshold, so they stay on the serial
        # per-read path and are unaffected.
        # KUMA_MAME_TIMING_JSON gives a frozen-safe positive signal: the
        # per-read pool branch records a "barcode_match_parallel_wall" phase.
        timing_json = workdir / "timing.jsonl"
        child_env = dict(os.environ)
        child_env["KUMA_MAME_PERREAD_THRESHOLD"] = "1"
        child_env["KUMA_MAME_TIMING_JSON"] = str(timing_json)

        sio = SidecarIO(binary, stderr_path, env=child_env)

        # --- ping ---
        print("[1/7] ping ...")
        sio.send(_rpc(1, "ping", {}))
        try:
            ping_resp = sio.recv(1, timeout=30.0)
            if ping_resp.get("result", {}).get("ok") is not True:
                failures.append(f"ping: ok is not True — got {ping_resp!r}")
            else:
                print("      ping OK")
        except (TimeoutError, RuntimeError) as exc:
            failures.append(f"ping timed out or process died: {exc}")

        # --- detect ---
        print("[2/7] mame.detect_native_barcodes ...")
        sio.send(_rpc(2, "mame.detect_native_barcodes", {
            "minknow_run_dir": str(run_dir),
        }))
        try:
            detect_resp = sio.recv(2, timeout=60.0)
            if "error" in detect_resp:
                failures.append(f"detect RPC error: {detect_resp['error']}")
            else:
                detect_result = detect_resp.get("result", {})
                total = detect_result.get("total_count", 0)
                if total < 1:
                    failures.append(f"detect: total_count={total}, expected >= 1")
                else:
                    print(f"      detect OK — total_count={total}, "
                          f"native_barcodes={detect_result.get('native_barcodes')}")
        except (TimeoutError, RuntimeError) as exc:
            failures.append(f"detect timed out or process died: {exc}")

        # --- per-NB parallel demux ---
        print("[3/7] mame.run_combinatorial_demux (per-NB parallel) ...")
        sio.send(_rpc(3, "mame.run_combinatorial_demux", {
            "minknow_run_dir": str(run_dir),
            "custom_barcodes_xlsx": str(xlsx),
            "reference_fasta": str(ref_fasta),
            "output_dir": str(out_dir),
            "native_barcodes": ["barcode06", "barcode20"],
        }))
        try:
            demux_resp = sio.recv(3, timeout=300.0)
            if "error" in demux_resp:
                failures.append(f"demux RPC error: {demux_resp['error']}")
            else:
                demux_result = demux_resp.get("result", {})
                per_nb = demux_result.get("native_barcodes")
                if not isinstance(per_nb, list):
                    failures.append(
                        f"demux: native_barcodes is not a list — got {type(per_nb).__name__!r}"
                    )
                elif len(per_nb) != 2:
                    failures.append(
                        f"demux: native_barcodes length={len(per_nb)}, expected 2"
                    )
                else:
                    print(f"      demux OK — native_barcodes list length={len(per_nb)}")
        except (TimeoutError, RuntimeError) as exc:
            failures.append(f"demux timed out or process died: {exc}")

        # --- analyze (raw MinKNOW run folder: folds demux + analyze) ---
        analyze_out = workdir / "analyze_out.xlsx"
        print("[4/7] mame.analyze (raw MinKNOW run folder) ...")
        sio.send(_rpc(4, "analyze", {
            "input_dir": str(run_dir),
            "reference": str(ref_fasta),
            "expected": str(expected_xlsx),
            "output": str(analyze_out),
            "custom_barcodes_xlsx": str(xlsx),
            "native_barcodes": ["barcode06", "barcode20"],
        }))
        try:
            analyze_resp = sio.recv(4, timeout=300.0)
            if "error" in analyze_resp:
                failures.append(f"analyze RPC error: {analyze_resp['error']}")
            else:
                analyze_result = analyze_resp.get("result", {})
                verdicts = analyze_result.get("verdicts")
                if not isinstance(verdicts, list):
                    failures.append(
                        f"analyze: verdicts is not a list — got {type(verdicts).__name__!r}"
                    )
                elif "assigned_reads" not in analyze_result:
                    failures.append(
                        "analyze: raw-run extra 'assigned_reads' missing from result"
                    )
                elif "wells_with_reads" not in analyze_result:
                    failures.append(
                        "analyze: raw-run extra 'wells_with_reads' missing from result"
                    )
                else:
                    print(
                        f"      analyze OK — verdicts={len(verdicts)}, "
                        f"assigned_reads={analyze_result.get('assigned_reads')}, "
                        f"wells_with_reads={analyze_result.get('wells_with_reads')}"
                    )
        except (TimeoutError, RuntimeError) as exc:
            failures.append(f"analyze timed out or process died: {exc}")

        # --- per-read ProcessPool demux (single native barcode) ---
        # The steps above never reach the per-read pool: with 2 native barcodes
        # per_read_parallel is False. One native barcode + a threshold of 1
        # forces the per-read spawn ProcessPool, which is the newly enabled
        # path on frozen Windows builds. The stderr check at the end must still
        # see exactly one "MAME sidecar started" line: if a spawned per-read
        # worker re-entered the RPC loop it would print a second one.
        perread_out = workdir / "output_perread"
        perread_out.mkdir()
        print("[5/7] mame.run_combinatorial_demux (per-read ProcessPool, 1 NB) ...")
        sio.send(_rpc(5, "mame.run_combinatorial_demux", {
            "minknow_run_dir": str(single_run_dir),
            "custom_barcodes_xlsx": str(xlsx),
            "reference_fasta": str(ref_fasta),
            "output_dir": str(perread_out),
            "native_barcodes": ["barcode06"],
        }))
        try:
            pr_resp = sio.recv(5, timeout=300.0)
            if "error" in pr_resp:
                failures.append(f"per-read demux RPC error: {pr_resp['error']}")
            else:
                pr_result = pr_resp.get("result", {})
                pr_nb = pr_result.get("native_barcodes")
                if not isinstance(pr_nb, list) or len(pr_nb) != 1:
                    failures.append(
                        f"per-read demux: native_barcodes expected list of 1, got {pr_nb!r}"
                    )
                elif not pr_result.get("assigned_reads"):
                    failures.append(
                        f"per-read demux: assigned_reads={pr_result.get('assigned_reads')!r}, "
                        "expected a positive count"
                    )
                else:
                    print(
                        f"      per-read demux OK, assigned_reads="
                        f"{pr_result.get('assigned_reads')}, "
                        f"wells_with_reads={pr_result.get('wells_with_reads')}"
                    )
        except (TimeoutError, RuntimeError) as exc:
            failures.append(f"per-read demux timed out or process died: {exc}")

        # Positive proof the pool actually ran: only the ProcessPool branch
        # records a "barcode_match_parallel_wall" phase into the timing JSONL.
        perread_phase_seen = False
        if timing_json.exists():
            for raw in timing_json.read_text(
                encoding="utf-8", errors="replace"
            ).splitlines():
                if not raw.strip():
                    continue
                try:
                    rec = json.loads(raw)
                except json.JSONDecodeError:
                    continue
                if "barcode_match_parallel_wall" in rec.get("phases_s", {}):
                    perread_phase_seen = True
                    break
        if perread_phase_seen:
            print("      per-read ProcessPool path confirmed "
                  "(barcode_match_parallel_wall recorded)")
        else:
            failures.append(
                "per-read demux: 'barcode_match_parallel_wall' phase absent from "
                f"{timing_json}, the per-read ProcessPool branch never ran, so this "
                "step did not exercise the path it is meant to cover"
            )

        # --- non-ispS barcode plate ---
        # Every step above feeds the frozen binary an ispS-shaped workbook, so
        # none of them ever exercised the barcode reader on a file the ispS tail
        # constants cannot explain. That is the whole population of packages
        # `barcode_package` generates. The counts asserted here are the ones the
        # derived tail produces; the deleted fixed-length fallback collapsed R1
        # and R2 into one string and returned 4 wells / 8 reads.
        gene_out = workdir / "output_gene"
        gene_out.mkdir()
        print("[6/7] mame.run_combinatorial_demux (non-ispS barcode plate) ...")
        sio.send(_rpc(6, "mame.run_combinatorial_demux", {
            "minknow_run_dir": str(gene_run_dir),
            "custom_barcodes_xlsx": str(gene_xlsx),
            "reference_fasta": str(ref_fasta),
            "output_dir": str(gene_out),
            "native_barcodes": ["barcode06"],
        }))
        expected_wells = len(_GENE_ROWS) * len(_GENE_COLS)
        expected_reads = expected_wells * _GENE_READS_PER_WELL
        try:
            gene_resp = sio.recv(6, timeout=300.0)
            if "error" in gene_resp:
                failures.append(f"non-ispS demux RPC error: {gene_resp['error']}")
            else:
                gene_result = gene_resp.get("result", {})
                gene_wells = gene_result.get("wells_with_reads")
                gene_reads = gene_result.get("assigned_reads")
                if gene_wells != expected_wells or gene_reads != expected_reads:
                    failures.append(
                        "non-ispS demux: expected "
                        f"{expected_wells} wells / {expected_reads} assigned reads, "
                        f"got {gene_wells!r} wells / {gene_reads!r} reads. The "
                        "barcode seeds were not cut at the tail derived from the "
                        "workbook, so this plate was read with the wrong seeds."
                    )
                else:
                    print(
                        f"      non-ispS demux OK, wells_with_reads={gene_wells}, "
                        f"assigned_reads={gene_reads}"
                    )
        except (TimeoutError, RuntimeError) as exc:
            failures.append(f"non-ispS demux timed out or process died: {exc}")

        # --- shutdown ---
        print("[7/7] shutdown ...")
        sio.send(_rpc(7, "shutdown", {}))
        try:
            sio.recv(7, timeout=15.0)
            print("      shutdown ack received")
        except (TimeoutError, RuntimeError) as exc:
            # Shutdown ack may not arrive before EOF; process exit is the criterion
            print(f"      shutdown ack not received ({exc}); waiting for process exit")

        rc = sio.close(timeout=15.0)
        print(f"      sidecar exit code: {rc}")
        sio = None

    except OSError as exc:
        failures.append(f"OS error launching sidecar: {exc}")
    finally:
        if sio is not None:
            sio.close(timeout=5.0)
            sio = None

    # --- stderr analysis (after process is dead so buffer is fully flushed) ---
    print("\n[stderr analysis]")
    stderr_lines: list[str] = []
    if stderr_path.exists():
        stderr_lines = stderr_path.read_text(encoding="utf-8", errors="replace").splitlines()

    started_count = sum(
        1 for line in stderr_lines if "MAME sidecar started" in line
    )
    print(f"      'MAME sidecar started' occurrences in stderr: {started_count}")

    if started_count == 0:
        failures.append(
            "freeze_support check: 'MAME sidecar started' not found in stderr — "
            "sidecar may not have launched correctly"
        )
    elif started_count > 1:
        failures.append(
            f"freeze_support BROKEN: 'MAME sidecar started' appeared {started_count}x "
            f"(expected exactly 1) — spawned children are re-running the server loop"
        )
    else:
        print("      freeze_support OK: exactly 1 'MAME sidecar started' line")

    # --- Final report ---
    print()
    if failures:
        print("=" * 60)
        print("FROZEN MAME SMOKE: FAIL")
        print("=" * 60)
        for i, msg in enumerate(failures, 1):
            print(f"  [{i}] {msg}")
        print()
        print("--- last 20 stderr lines ---")
        for line in stderr_lines[-20:]:
            print(f"  {line}")
        shutil.rmtree(workdir, ignore_errors=True)
        sys.exit(1)
    else:
        print("=" * 60)
        print("FROZEN MAME SMOKE: PASS")
        print("=" * 60)
        shutil.rmtree(workdir, ignore_errors=True)
        sys.exit(0)


def main() -> None:
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <path-to-frozen-mame-sidecar>")
        sys.exit(1)

    binary = Path(sys.argv[1])
    if not binary.exists():
        print(f"FAIL: sidecar binary not found: {binary}")
        sys.exit(1)
    if not os.access(binary, os.X_OK):
        print(f"FAIL: sidecar binary is not executable: {binary}")
        sys.exit(1)

    print(f"Frozen MAME sidecar smoke test")
    print(f"Binary: {binary}")
    print(f"Platform: {sys.platform}")
    print()

    run_smoke(binary)


if __name__ == "__main__":
    main()
