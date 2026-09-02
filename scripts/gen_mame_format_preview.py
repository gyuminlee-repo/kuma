"""Generate the MAME step 4.1 file-shape previews from the bundled templates.

Run via:
    python3 scripts/gen_mame_format_preview.py

Outputs:
    src/data/mameFormatPreviews.generated.json

Every cell comes out of `templates/`, none of it is written by hand: a copied
table goes stale the moment a template changes and a stale copy shows the
operator a spreadsheet that no longer exists.
`tests/scripts/test_mame_format_preview_drift.py` re-runs `build_previews()`
in memory and fails when the checked-in JSON no longer matches.

Three of the sources (09, 11, 12) are byte-identical for their first fifteen
rows: three wild-type blocks with the same signal header, the same area and the
same name. They diverge in the sample names further down, which is why the
preview shows two windows (one WT block, one sample block) rather than the top
of the file, and why the sample name cell is reported as the highlight.

Which sample block a preview lands on is chosen per preview, as an index into
the non-wild-type blocks rather than as a row number: template 12 backs both
the primary numeric screen and the numeric confirmation, and the two would show
the identical table if both took the first one. The confirmation takes the
second, whose name (`1-2`) is a repeat measurement of the first, so the table
carries the one thing the two formats really differ in.
"""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path
from typing import Any, NamedTuple

import openpyxl

_REPO_ROOT = Path(__file__).resolve().parent.parent
_OUTPUT = _REPO_ROOT / "src" / "data" / "mameFormatPreviews.generated.json"

sys.path.insert(0, str(_REPO_ROOT))

from kuma_core.mame.activity.constants import WT_PATTERN  # noqa: E402

#: Rows shown for a flat sheet: the column header plus enough data rows to make
#: the repeating shape obvious without turning the popover into a scroll box.
_FLAT_DATA_ROWS = 3

#: First column marker that opens an Agilent-style block.
_BLOCK_MARKER = "Signal:"

#: Row offset of the sample name inside a block, counted from the marker row.
_BLOCK_NAME_OFFSET = 2
_BLOCK_NAME_COL = 1


def _cell(value: Any) -> str:
    """Render one spreadsheet cell the way the operator sees it.

    Strings so the JSON is display-ready and so a float never re-formats
    differently in JavaScript than it did here.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _rectangular(rows: list[list[str]]) -> list[list[str]]:
    width = max((len(row) for row in rows), default=0)
    return [row + [""] * (width - len(row)) for row in rows]


def _sheet_rows(path: Path) -> list[list[str]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = [[_cell(value) for value in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows


def _csv_rows(path: Path) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return [list(row) for row in csv.reader(handle)]


def _read_rows(path: Path) -> list[list[str]]:
    return _csv_rows(path) if path.suffix.lower() == ".csv" else _sheet_rows(path)


def _is_blank(row: list[str]) -> bool:
    return all(cell == "" for cell in row)


def _blocks(rows: list[list[str]]) -> list[tuple[int, list[list[str]]]]:
    """Split an Agilent-style sheet into its blocks.

    Returns (zero-based start row, block rows). Boundaries come from the
    `Signal:` marker and the blank row that closes a block, never from a row
    number: a real run carries a different number of blocks than a template.
    """
    found: list[tuple[int, list[list[str]]]] = []
    index = 0
    while index < len(rows):
        row = rows[index]
        if row and row[0] == _BLOCK_MARKER:
            end = index + 1
            while (
                end < len(rows)
                and not _is_blank(rows[end])
                and rows[end][0] != _BLOCK_MARKER
            ):
                end += 1
            found.append((index, rows[index:end]))
            index = end
        else:
            index += 1
    return found


def _block_sample_name(block: list[list[str]]) -> str:
    if len(block) <= _BLOCK_NAME_OFFSET:
        return ""
    name_row = block[_BLOCK_NAME_OFFSET]
    return name_row[_BLOCK_NAME_COL] if len(name_row) > _BLOCK_NAME_COL else ""


def _is_wt(name: str) -> bool:
    return bool(WT_PATTERN.match(name)) or name.upper() == "WT"


def _flat_preview(source: str) -> dict[str, Any]:
    rows = _read_rows(_REPO_ROOT / source)
    window = _rectangular(rows[: _FLAT_DATA_ROWS + 1])
    return {
        "source": source,
        "headerRow": True,
        "windows": [{"startRow": 1, "rows": window}],
        "ellipsisBetweenWindows": False,
        "truncatedAfter": len(rows) > len(window),
        "highlight": None,
    }


def _block_preview(source: str, sample_block: int = 0) -> dict[str, Any]:
    """Two windows out of an Agilent-style sheet: a WT block and a sample block.

    *sample_block* indexes the non-wild-type blocks, found by their marker and
    their name. A row number would break the moment a template gains a block.
    """
    rows = _read_rows(_REPO_ROOT / source)
    blocks = _blocks(rows)
    wt = next((b for b in blocks if _is_wt(_block_sample_name(b[1]))), None)
    samples = [b for b in blocks if not _is_wt(_block_sample_name(b[1]))]
    if wt is None or len(samples) <= sample_block:
        raise SystemExit(
            f"{source}: expected a WT block and at least "
            f"{sample_block + 1} sample block(s)"
        )
    sample = samples[sample_block]
    windows = [
        {"startRow": wt[0] + 1, "rows": _rectangular(wt[1])},
        {"startRow": sample[0] + 1, "rows": _rectangular(sample[1])},
    ]
    return {
        "source": source,
        "headerRow": False,
        "windows": windows,
        "ellipsisBetweenWindows": True,
        "truncatedAfter": sample[0] + len(sample[1]) < len(rows),
        "highlight": {
            "window": 1,
            "row": _BLOCK_NAME_OFFSET,
            "col": _BLOCK_NAME_COL,
        },
    }


class _Source(NamedTuple):
    """Where one preview reads from and how the file is shaped."""

    path: str
    kind: str
    #: Which non-wild-type block to show. Block sources only.
    sample_block: int = 0


#: preview id -> source. Ids match `MeasurementSource` plus the supporting
#: files every other file field in the app asks for. Folder pickers, sequence
#: and structure files, app-written manifests and every output path are absent
#: on purpose: none of them is a table, and the app fills the outputs itself.
_PREVIEWS: dict[str, _Source] = {
    "longFormat": _Source("templates/07_mame_activity_long.csv", "flat"),
    "gcSheet": _Source("templates/10_mame_gc_prenormalised.xlsx", "flat"),
    "rawReport": _Source("templates/11_mame_gc_fid_round1_raw.xlsx", "block"),
    "numericReport": _Source("templates/12_mame_agilent_numeric_index.xlsx", "block"),
    "confirmationVariantLabels": _Source(
        "templates/09_mame_agilent_rep_batch.xlsx", "block"
    ),
    # The second sample block, not the first: see the module docstring.
    "confirmationNumericIds": _Source(
        "templates/12_mame_agilent_numeric_index.xlsx", "block", sample_block=1
    ),
    "plateLayout": _Source("templates/06_mame_plate_layout.xlsx", "flat"),
    "expectedMutations": _Source("templates/03_mame_expected_mutations.xlsx", "flat"),
    "customBarcodes": _Source("templates/04_mame_custom_barcodes.xlsx", "flat"),
    # The only one of these that ships beside the sample data rather than in
    # `templates/`; it is still the file the sample loader hands the operator.
    "barcodeSeeds": _Source("src-tauri/samples/mame/02_mame_barcode_seeds.xlsx", "flat"),
    "evolveproPrediction": _Source("templates/01_kuro_evolvepro_pred.csv", "flat"),
}


def build_previews() -> dict[str, Any]:
    """Read every template and return the preview payload."""
    built: dict[str, Any] = {}
    for preview_id, source in _PREVIEWS.items():
        built[preview_id] = (
            _flat_preview(source.path)
            if source.kind == "flat"
            else _block_preview(source.path, source.sample_block)
        )
    return built


def render(previews: dict[str, Any]) -> str:
    payload = {
        "_generated": "scripts/gen_mame_format_preview.py -- do not edit by hand",
        "previews": previews,
    }
    return json.dumps(payload, indent=2, ensure_ascii=False) + "\n"


def highlighted_cell(preview: dict[str, Any]) -> str:
    """The one cell a block format is told apart by, or an empty string."""
    highlight = preview["highlight"]
    if highlight is None:
        return ""
    window = preview["windows"][highlight["window"]]
    return window["rows"][highlight["row"]][highlight["col"]]


def main() -> None:
    previews = build_previews()
    _OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    _OUTPUT.write_text(render(previews), encoding="utf-8")
    print(f"wrote {len(previews)} previews to {_OUTPUT.relative_to(_REPO_ROOT)}")
    for preview_id, preview in previews.items():
        shape = ", ".join(
            f"{len(w['rows'])}x{len(w['rows'][0])}@row{w['startRow']}"
            for w in preview["windows"]
        )
        print(
            f"  {preview_id}: {shape} "
            f"highlight={highlighted_cell(preview)!r} <- {preview['source']}"
        )


if __name__ == "__main__":
    main()
